#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-

"""
----------------------------------------------------------------------------------------------
  SISTEMA ..: 
  MODULO ...: 
  SCRIPT ...: rel_comparativo_sped_gia_conv115_por_serie.py
  CRIACAO ..: 05/03/2021
  AUTOR ....: AIRTON BORGES DA SILVA FILHO / KYROS CONSULTORIA
  DESCRICAO : Relatório em excel para validação das informações do SPED, Convênio 115 e GIA. utilizando como insumos os relatórios
      “Análise_SPED”, “Conciliação Série” e “Comando SO”.
      O relatório deve conter as abas “Resumo SPED” e “Resumo GIA”
      Nomenclatura:  Analise_SPED_Convenio115_GIA_<MESANO>_<UF>_<IE>_ V<N>.xlsx
        Exemplo:
            Analise_SPED_Convenio115_GIA_042020_SP_108383949112_V001.xlsx
      
      O arquivo gerado deve ser disponibilizado na pasta:
            /arquivos/RELATORIOS/MENSAIS/YYYY/MM
        Exemplo
            /arquivos/RELATORIOS/MENSAIS/2018/01

           

  ANEXO ....: Demais dados e documentação na pasta documentação, arquivos :
                D:\CLONE1\arquivos\TESHUVA\melhorias\09 - Fase 2 - Relatórios JP 20

----------------------------------------------------------------------------------------------
  HISTORICO : 
https://www.letscode.com.br/blog/aprenda-a-integrar-python-e-excel
  * 19/08/2021 - Marcelo Gremonesi            / Kyros Consultoria - Adquacoes para o novo painel 
----------------------------------------------------------------------------------------------
"""

import sys
import os
dir_base = os.path.join( os.path.realpath('.').split('/PROD/')[0], 'PROD') if os.path.realpath('.').__contains__('/PROD/') else os.path.join( os.path.realpath('.').split('/DEV/')[0], 'DEV')
sys.path.append(dir_base)
import configuracoes

import datetime
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.reader.excel import load_workbook 

import comum


SD = ('/' if os.name == 'posix' else '\\')
# DEBUG = True
DEBUG = False

def validauf(uf):
    return(True if (uf.upper() in ('AC','AL','AM','AP','BA','CE','DF','ES','GO','MA','MG','MS','MT','PA','PB','PE','PI','PR','RJ','RN','RO','RR','RS','SC','SE','SP','TO')) else False)
           
def dtf():
    return (datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
 
def ultimodia(ano,mes):
   return(31 if mes == 12 else datetime.date.fromordinal((datetime.date(ano,mes+1,1)).toordinal()-1).day)

def proximo_arquivo(mascara,diretorio):
    qdade = 0
    nomearq = "" 
    directory = Path(diretorio)
    files = directory.glob(mascara)
    sorted_files = sorted(files, reverse=False)    
   
    nomearq = mascara.replace("*", "V000")
    proximo = "001"
    
    if sorted_files:
        for f in sorted_files:
            qdade = qdade + 1
            nomearq = f
        proximo = '{:03d}'.format(int((str(nomearq).split(".")[0]).split("_")[-1][1:]) + 1)
    nomearq = mascara.replace("*", "V"+proximo)
    return(nomearq)

def nome_arquivo(mascara,diretorio):
    qdade = 0
    nomearq = "" 
    directory = Path(diretorio)
    files = directory.glob(mascara)
    sorted_files = sorted(files, reverse=False)
    if sorted_files:
        for f in sorted_files:
            qdade = qdade + 1
            nomearq = f
    else: 
        log('#### Erro encontrado : Arquivo %s não está na pasta %s'%(mascara,diretorio))
    return(nomearq)

def semespacos(frase):
    retorno = "" 
    for l in frase:
        if (l != " "):
            retorno = retorno + l
    return(retorno)

def formata_SPED(planilha):

    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
    fontMasterAzul     = Font(color='000000FF', bold=True, size=12)
    fontMasterVermelha = Font(color='00FF0000', bold=True, size=12)
    
    planilha.column_dimensions['A'].width = 10  
    planilha.column_dimensions['B'].width = 20   
    planilha.column_dimensions['C'].width = 20  
    planilha.column_dimensions['D'].width = 20  
    planilha.column_dimensions['E'].width = 20 
    planilha.column_dimensions['F'].width = 20  
    planilha.column_dimensions['G'].width = 20  
    planilha.column_dimensions['H'].width = 20  
    planilha.column_dimensions['I'].width = 20  
    planilha.column_dimensions['J'].width = 20  
    planilha.column_dimensions['K'].width = 20  
    planilha.column_dimensions['L'].width = 20  
    planilha.column_dimensions['M'].width = 20  
    
    planilha.merge_cells('C1:G1')
    planilha.merge_cells('H1:J1')
    planilha.merge_cells('K1:M1')



    planilha.cell(1,  2).font = fontMasterPreta
    planilha.cell(1,  3).font = fontMasterPreta
    planilha.cell(1,  8).font = fontMasterPreta
    planilha.cell(1, 11).font = fontMasterPreta
    
    planilha.cell(2,  1).font = fontMasterPreta
    planilha.cell(2,  2).font = fontMasterPreta
    planilha.cell(2,  3).font = fontMasterPreta
    planilha.cell(2,  4).font = fontMasterPreta
    planilha.cell(2,  5).font = fontMasterPreta
    planilha.cell(2,  6).font = fontMasterPreta
    planilha.cell(2,  7).font = fontMasterPreta
    planilha.cell(2,  8).font = fontMasterPreta
    planilha.cell(2,  9).font = fontMasterPreta
    planilha.cell(2, 10).font = fontMasterPreta
    planilha.cell(2, 11).font = fontMasterPreta
    planilha.cell(2, 12).font = fontMasterPreta
    planilha.cell(2, 13).font = fontMasterPreta

   
    planilha.cell(1,  2).alignment = Alignment(horizontal='center')
    planilha.cell(1,  3).alignment = Alignment(horizontal='center')
    planilha.cell(1,  8).alignment = Alignment(horizontal='center')
    planilha.cell(1, 11).alignment = Alignment(horizontal='center')

    planilha.cell(2,  1).alignment = Alignment(horizontal='center')
    planilha.cell(2,  2).alignment = Alignment(horizontal='center')
    planilha.cell(2,  3).alignment = Alignment(horizontal='center')
    planilha.cell(2,  4).alignment = Alignment(horizontal='center')
    planilha.cell(2,  5).alignment = Alignment(horizontal='center')
    planilha.cell(2,  6).alignment = Alignment(horizontal='center')
    planilha.cell(2,  7).alignment = Alignment(horizontal='center')
    planilha.cell(2,  8).alignment = Alignment(horizontal='center')
    planilha.cell(2,  9).alignment = Alignment(horizontal='center')
    planilha.cell(2, 10).alignment = Alignment(horizontal='center')
    planilha.cell(2, 11).alignment = Alignment(horizontal='center')
    planilha.cell(2, 12).alignment = Alignment(horizontal='center')
    planilha.cell(2, 13).alignment = Alignment(horizontal='center')

  
    linha = 1
    for row in planilha.rows:
      
        if (linha > 2):
            planilha.cell(linha,  2).number_format = "#,##0.00"
            planilha.cell(linha,  3).number_format = "#,##0.00"
            planilha.cell(linha,  4).number_format = "#,##0.00"
            planilha.cell(linha,  5).number_format = "#,##0.00"
            planilha.cell(linha,  6).number_format = "#,##0.00"
            planilha.cell(linha,  7).number_format = "#,##0.00"
            planilha.cell(linha,  8).number_format = "#,##0.00"
            planilha.cell(linha,  9).number_format = "#,##0.00"
            planilha.cell(linha, 10).number_format = "#,##0.00"
            planilha.cell(linha, 11).number_format = "#,##0.00"
            planilha.cell(linha, 12).number_format = "#,##0.00"
            planilha.cell(linha, 13).number_format = "#,##0.00"
            

        if (planilha.cell(linha,  1).value == 'TOTAIS'):
            planilha.cell(linha,  1).alignment = Alignment(horizontal='center')
            planilha.cell(linha,  1).font = fontMasterAzul
            planilha.cell(linha,  2).font = fontMasterAzul
            planilha.cell(linha,  3).font = fontMasterAzul
            planilha.cell(linha,  4).font = fontMasterAzul
            planilha.cell(linha,  5).font = fontMasterAzul
            planilha.cell(linha,  6).font = fontMasterAzul
            planilha.cell(linha,  7).font = fontMasterAzul
            planilha.cell(linha,  8).font = fontMasterAzul
            planilha.cell(linha,  9).font = fontMasterAzul
            planilha.cell(linha, 10).font = fontMasterAzul
            planilha.cell(linha, 11).font = fontMasterAzul
            planilha.cell(linha, 12).font = fontMasterAzul
            planilha.cell(linha, 13).font = fontMasterAzul

        linha = linha + 1  
  
    s=Side(border_style=BORDER_THIN, color='00000000')
    n=Side(border_style=None, color='00000000') 
    
  
    for a in range(1,linha):
         planilha.cell(row=a, column=1).border = Border(s,s,n,n)
         planilha.cell(row=a, column=2).border = Border(s,s,n,n)
         planilha.cell(row=a, column=8).border = Border(s,n,n,n)
         planilha.cell(row=a, column=11).border = Border(s,n,n,n)
         planilha.cell(row=a, column=14).border = Border(s,n,n,n)
 
    planilha.cell(row=linha-1, column=1).border = Border(s,s,s,s)
    planilha.cell(row=linha-1, column=2).border = Border(s,s,s,s)
    planilha.cell(row=linha-1, column=2).border = Border(s,s,s,s)
    planilha.cell(row=linha-1, column=2).border = Border(s,s,s,s)

    for b in range(1,14):
        planilha.cell(row=linha-1, column=b).border = Border(n,n,s,s)

    planilha.cell(row=linha-1, column=2).border  = Border(s,s,s,s)
    planilha.cell(row=linha-1, column=8).border  = Border(s,n,s,s)
    planilha.cell(row=linha-1, column=11).border = Border(s,n,s,s)
    planilha.cell(row=linha-1, column=13).border = Border(n,s,s,s) 
    
       
def formata_GIA(planilha):

    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
    fontMasterAzul     = Font(color='000000FF', bold=True, size=12)
    fontMasterVermelha = Font(color='00FF0000', bold=True, size=12)
    fontMasterVerde    = Font(color='00008000', bold=True, size=12)
    
    planilha.column_dimensions['A'].width = 10  
    planilha.column_dimensions['B'].width = 17   
    planilha.column_dimensions['C'].width = 17  
    planilha.column_dimensions['D'].width = 17  
    planilha.column_dimensions['E'].width = 17 
    planilha.column_dimensions['F'].width = 17  
    planilha.column_dimensions['G'].width = 17  
    planilha.column_dimensions['H'].width = 17  
    planilha.column_dimensions['I'].width = 17  
    planilha.column_dimensions['J'].width = 17  
    planilha.column_dimensions['K'].width = 17  
    planilha.column_dimensions['L'].width = 17  
    planilha.column_dimensions['M'].width = 17  
    planilha.column_dimensions['N'].width = 17  
    planilha.column_dimensions['O'].width = 17  
    planilha.column_dimensions['P'].width = 17  

    planilha.merge_cells('B1:F1')
    planilha.merge_cells('G1:K1')
    planilha.merge_cells('L1:P1')
    
    planilha.cell(1,  2).font = fontMasterPreta
    planilha.cell(1,  7).font = fontMasterPreta
    planilha.cell(1, 12).font = fontMasterVermelha
    
    planilha.cell(1,  2).alignment = Alignment(horizontal='center')
    planilha.cell(1,  7).alignment = Alignment(horizontal='center')
    planilha.cell(1, 12).alignment = Alignment(horizontal='center')
 
    
    planilha.cell(2,  1).font = fontMasterPreta
    planilha.cell(2,  2).font = fontMasterPreta
    planilha.cell(2,  3).font = fontMasterPreta
    planilha.cell(2,  4).font = fontMasterPreta
    planilha.cell(2,  5).font = fontMasterPreta
    planilha.cell(2,  6).font = fontMasterPreta
    planilha.cell(2,  7).font = fontMasterPreta
    planilha.cell(2,  8).font = fontMasterPreta
    planilha.cell(2,  9).font = fontMasterPreta
    planilha.cell(2, 10).font = fontMasterPreta
    planilha.cell(2, 11).font = fontMasterPreta
    planilha.cell(2, 12).font = fontMasterPreta
    planilha.cell(2, 13).font = fontMasterPreta
    planilha.cell(2, 14).font = fontMasterPreta
    planilha.cell(2, 15).font = fontMasterPreta
    planilha.cell(2, 16).font = fontMasterPreta
   


    planilha.cell(2,  1).alignment = Alignment(horizontal='center')
    planilha.cell(2,  2).alignment = Alignment(horizontal='center')
    planilha.cell(2,  3).alignment = Alignment(horizontal='center')
    planilha.cell(2,  4).alignment = Alignment(horizontal='center')
    planilha.cell(2,  5).alignment = Alignment(horizontal='center')
    planilha.cell(2,  6).alignment = Alignment(horizontal='center')
    planilha.cell(2,  7).alignment = Alignment(horizontal='center')
    planilha.cell(2,  8).alignment = Alignment(horizontal='center')
    planilha.cell(2,  9).alignment = Alignment(horizontal='center')
    planilha.cell(2, 10).alignment = Alignment(horizontal='center')
    planilha.cell(2, 11).alignment = Alignment(horizontal='center')
    planilha.cell(2, 12).alignment = Alignment(horizontal='center')
    planilha.cell(2, 13).alignment = Alignment(horizontal='center')
    planilha.cell(2, 14).alignment = Alignment(horizontal='center')
    planilha.cell(2, 15).alignment = Alignment(horizontal='center')
    planilha.cell(2, 16).alignment = Alignment(horizontal='center')
  
    linha = 1
    for row in planilha.rows:
      
        if (linha > 2):
            planilha.cell(linha,  2).number_format = "#,##0.00"
            planilha.cell(linha,  3).number_format = "#,##0.00"
            planilha.cell(linha,  4).number_format = "#,##0.00"
            planilha.cell(linha,  5).number_format = "#,##0.00"
            planilha.cell(linha,  6).number_format = "#,##0.00"
            planilha.cell(linha,  7).number_format = "#,##0.00"
            planilha.cell(linha,  8).number_format = "#,##0.00"
            planilha.cell(linha,  9).number_format = "#,##0.00"
            planilha.cell(linha, 10).number_format = "#,##0.00"
            planilha.cell(linha, 11).number_format = "#,##0.00"
            planilha.cell(linha, 12).number_format = "#,##0.00"
            planilha.cell(linha, 13).number_format = "#,##0.00"
            planilha.cell(linha, 14).number_format = "#,##0.00"
            planilha.cell(linha, 15).number_format = "#,##0.00"
            planilha.cell(linha, 16).number_format = "#,##0.00"
            

        if (planilha.cell(linha,  1).value == 'TOTAIS'):
            planilha.cell(linha,  1).alignment = Alignment(horizontal='center')
            planilha.cell(linha,  1).font = fontMasterAzul
            planilha.cell(linha,  2).font = fontMasterAzul
            planilha.cell(linha,  3).font = fontMasterAzul
            planilha.cell(linha,  4).font = fontMasterAzul
            planilha.cell(linha,  5).font = fontMasterAzul
            planilha.cell(linha,  6).font = fontMasterAzul
            planilha.cell(linha,  7).font = fontMasterAzul
            planilha.cell(linha,  8).font = fontMasterAzul
            planilha.cell(linha,  9).font = fontMasterAzul
            planilha.cell(linha, 10).font = fontMasterAzul
            planilha.cell(linha, 11).font = fontMasterAzul
            planilha.cell(linha, 12).font = fontMasterVermelha
            planilha.cell(linha, 13).font = fontMasterVermelha
            planilha.cell(linha, 14).font = fontMasterVermelha
            planilha.cell(linha, 15).font = fontMasterVermelha
            planilha.cell(linha, 16).font = fontMasterVermelha
            
            
        if (planilha.cell(linha,  6).value == 'TOTAIS SPED'):
            planilha.cell(linha,  6).alignment = Alignment(horizontal='center')
            planilha.cell(linha,  6).font = fontMasterVerde
            planilha.cell(linha,  7).font = fontMasterVerde
            planilha.cell(linha,  8).font = fontMasterVerde
            planilha.cell(linha,  9).font = fontMasterVerde
 
        if (planilha.cell(linha,  6).value == 'DIFERENCA'):
            planilha.cell(linha,  6).alignment = Alignment(horizontal='center')
            planilha.cell(linha,  6).font = fontMasterVermelha
            planilha.cell(linha,  7).font = fontMasterVermelha
            planilha.cell(linha,  8).font = fontMasterVermelha
            planilha.cell(linha,  9).font = fontMasterVermelha

        linha = linha + 1  
        
    s=Side(border_style=BORDER_THIN, color='00000000')
    n=Side(border_style=None, color='00000000') 
    
  
    for a in range(1,linha-2):
         planilha.cell(row=a, column=1).border = Border(s,s,n,n)
         planilha.cell(row=a, column=7).border = Border(s,n,n,n)
         planilha.cell(row=a, column=12).border = Border(s,n,n,n)
         planilha.cell(row=a, column=16).border = Border(n,s,n,n)
 
    for b in range(1,16):
        planilha.cell(row=linha-3, column=b).border = Border(n,n,s,s)

    planilha.cell(row=linha-3, column=2).border  = Border(s,n,s,s)
    planilha.cell(row=linha-3, column=7).border  = Border(s,n,s,s)
    planilha.cell(row=linha-3, column=12).border = Border(s,n,s,s)
    planilha.cell(row=linha-3, column=16).border = Border(n,s,s,s) 

    planilha.cell(row=linha-2, column=6).border = Border(s,s,s,s)
    planilha.cell(row=linha-2, column=7).border = Border(s,n,s,s)
    planilha.cell(row=linha-2, column=8).border = Border(n,n,s,s)
    planilha.cell(row=linha-2, column=9).border = Border(n,s,s,s)
    
    planilha.cell(row=linha-1, column=6).border = Border(s,s,s,s) 
    planilha.cell(row=linha-1, column=7).border = Border(s,n,s,s)
    planilha.cell(row=linha-1, column=8).border = Border(n,n,s,s)
    planilha.cell(row=linha-1, column=9).border = Border(n,s,s,s)
    




def processar():
    ret = 0
    ufi = ""
    mesanoi = ""
    mesi = ""
    anoi = "" 
    try :
        if (len(sys.argv) == 4 and len(sys.argv[1])==2 and len(sys.argv[2])==6 ) :
            iei = "*"
            ufi = sys.argv[1].upper()
            mesanoi = sys.argv[2].upper()
            mesi = sys.argv[2][:2].upper()
            anoi = sys.argv[2][2:].upper()
            iei = sys.argv[3].upper()
            iei = re.sub('[^0-9]','',iei)
        else :
            ret = 99
    except Exception as e :
        log('#### Erro encontrado :', e)
        ret = 99
    if ret == 99:
        log("-" * 100)
        log("#### ")
        log('#### ERRO - Erro nos parametros do script.')
        log("#### ")
        log('#### Exemplo de como deve ser :')
        log('####      %s <UF> <MMAAAA> <IE>'%(sys.argv[0]))
        log("#### ")
        log('#### Onde')
        log('####      <UF> = estado. Ex: SP')
        log('####      <MMAAAA> = mês e ano. Ex: Para junho de 2020 informe 062020')
        log('####      <IE> = inscrição desejada. Ex: 108383949112')
        log("#### ")
        log('#### Portanto, se o estado = SP, o mes = 06 e o ano = 2020, e IE = 108383949112, o comando correto deve ser :')  
        log('####      %s SP 062020 108383949112'%(sys.argv[0]))  
        log("#### ")
        log('#### Neste exemplo, a planilha resultado será:') 
        log('####      %sarquivos%sRELATORIOS%sMENSAIS%s2020%s06%sAnalise_SPED_Convenio115_GIA_062020_SP_108383949112.xlsx'%(SD,SD,SD,SD,SD,SD)) 
        log('#### ')
        log("-" * 100)
        print ("")
        print ("Retorno = 99")
        return(ret)

  # dir_rel_VALIDA  = os.path.join(SD,'arquivos','RELATORIOS','MENSAIS',anoi, mesi)
  # dir_rel_GIA     = os.path.join(SD,'arquivos','GIA'+ufi,'RELATORIOS',anoi,mesi)
  # dir_rel_SPED    = os.path.join(SD,'arquivos','SPED_FISCAL','RELATORIOS',ufi,anoi, mesi)
  # dir_rel_CONV115 = os.path.join(SD,'arquivos','RELATORIOS','CONV115',ufi,anoi, mesi)
   

    dir_rel_VALIDA  = os.path.join(SD,configuracoes.diretorio_rel_VALIDA, ufi, anoi, mesi)
    dir_rel_GIA     = os.path.join(SD,configuracoes.diretorio_rel_GIA.replace('<<uf>>',ufi),anoi,mesi)
    dir_rel_SPED    = os.path.join(SD,configuracoes.diretorio_rel_SPED,ufi,anoi, mesi)
    dir_rel_CONV115 = os.path.join(SD,configuracoes.diretorio_rel_CONV115,ufi,anoi, mesi)

    log('Diretorio',dir_rel_VALIDA)
    log('Diretorio',dir_rel_GIA)
    log('Diretorio',dir_rel_SPED)
    log('Diretorio',dir_rel_CONV115)
    
   
    mask_VALIDA = "Analise_SPED_Convenio115_GIA_"+mesanoi+"_"+ufi+"_"+iei+"_*.xlsx"    
    mask_CONV115 = "Relatorio_Conciliacao_"+ ufi + "_" + mesanoi + ".xlsx"
    mask_GIA = "relatorio_gia_totais_por_cfop_"+mesanoi+"_"+ufi+"_"+iei+"_*.xlsx"    
    mask_SPED = "Analise_SPED_"+mesanoi+"_"+ufi+"_"+iei+"_*.xlsx"   
    
    



    rel_VALIDA = proximo_arquivo(mask_VALIDA,dir_rel_VALIDA)
    R_VALIDA = os.path.join(dir_rel_VALIDA,rel_VALIDA)

    rel_GIA = nome_arquivo(mask_GIA,dir_rel_GIA)
    R_GIA = os.path.join(dir_rel_GIA,rel_GIA)

    rel_SPED = nome_arquivo(mask_SPED,dir_rel_SPED)
    R_SPED = os.path.join(dir_rel_SPED,rel_SPED)

    rel_CONV115 = nome_arquivo(mask_CONV115,dir_rel_CONV115)
    R_CONV115 = os.path.join(dir_rel_CONV115,rel_CONV115)

    if(DEBUG):  

        log("  = ", )
        log("mask_GIA     => ",mask_GIA )
        log("dir_rel_GIA  => ",dir_rel_GIA )   
        log("rel_GIA      => ",rel_GIA )
        log("R_GIA        => ",R_GIA )
        input("GIA OK?")
        log("  = ", )
    
        log("  = ", )    
        log("mask_SPED     => ",mask_SPED )
        log("dir_rel_SPED  => ",dir_rel_SPED )  
        log("rel_SPED      => ",rel_SPED )
        log("R_SPED        => ",R_SPED )
        input("SPED OK?")
        log("  = ", )
        
        log("  = ", )    
        log("mask_CONV115     => ",mask_CONV115 )
        log("dir_rel_CONV115  => ",dir_rel_CONV115 )  
        log("rel_CONV115      => ",rel_CONV115 )
        log("R_CONV115        => ",R_CONV115 )
        input("CONV115 OK?")
        log("  = ", )
        
        log("  = ", ) 
        log("mask_VALIDA     => ",mask_VALIDA )
        log("R_VALIDA        => ",R_VALIDA )   
        log("R_GIA           => ",R_GIA )   
        log("R_SPED          => ",R_SPED )  
        log("R_CONV115       => ",R_VALIDA )  
        log("  = ", ) 
         
        log("dir_rel_VALIDA   = ",dir_rel_VALIDA )
        log("dir_rel_SPED     = ",dir_rel_SPED )
        log("dir_rel_GIA      = ",dir_rel_GIA )
        log("dir_rel_CONV115  = ",dir_rel_CONV115 )
        log("  = ", )


    log("-"*100)
    log('#### Relatórios de entrada para análise:')
    log('####') 
    log('#### '+ R_GIA) 
    log('#### '+ R_SPED) 
    log('#### '+ R_CONV115) 
    log('####') 
    log("-"*100)
    log('#### Relatório de saida:')    
    log('####') 
    log('#### '+ R_VALIDA)           
    log('####')           
    log("-"*100)   

    
# =============================================================================
#     log(R_GIA)
#     log(dir_rel_GIA)
#     log(R_SPED)
#     log(dir_rel_SPED)
#     log(R_CONV115)
#     log(dir_rel_CONV115)
# 
# =============================================================================


    if (R_GIA == dir_rel_GIA+SD): 
        log('#### ERRO - Relatório de entrada GIA não encontrado na pasta: '+ R_GIA)  
        ret = 99
        
    if (R_SPED == dir_rel_SPED+SD): 
        log('#### ERRO - Relatório de entrada SPED não encontrado na pasta: '+ R_SPED)
        ret = 99
        
    if (R_CONV115 == dir_rel_CONV115+SD): 
        log('#### ERRO - Relatório de entrada CONV115 não encontrado na pasta: '+ R_CONV115)
        ret = 99
        
    if (ret == 99):
        print ("Retorno = 99")
        return(ret)
      
#    log("ANALISE_SPED")  
    asp = load_workbook(os.path.join(dir_rel_SPED, R_SPED))

#    log("COMANDO_SO")
    cso = load_workbook(os.path.join(dir_rel_GIA, R_GIA))
    
#    log("RELATORIO CONCILIACAO")    
    rc = load_workbook(os.path.join(dir_rel_CONV115, R_CONV115))
        

#### Se a pasta do relatório não existir, cria
    if not os.path.isdir(dir_rel_VALIDA) :
        os.makedirs(dir_rel_VALIDA)
   
#### Cria a planilha excel em memória....
    arquivo_excel = Workbook()
    RSPED = arquivo_excel.active
    RSPED.title = "Resumo SPED"
    RGIA = arquivo_excel.create_sheet("Resumo GIA", 1)

   
    


##########  ABA 1-   Resumo SPED
##########  ABA 1-   Resumo SPED
##########  ABA 1-   Resumo SPED
##########  ABA 1-   Resumo SPED
##########  ABA 1-   Resumo SPED
##########  ABA 1-   Resumo SPED



    rcp = rc['REG X PROT']
    rasp = asp['SPED_Regerado']
    
    serie        = {}
    vl_completo  = {}
    vl_contabil  = {}
    bc_icms      = {}
    icms         = {}
    isentos      = {}
    outros       = {}
    svl_contabil = {}
    sbc_icms     = {}
    sicms        = {}
    
    
    for line in rcp:

        if (line[12].value == 'REGERADO'):
            lserie = semespacos(line[2].value)

            if (serie.get(lserie)== None):
                vl_completo[lserie]  = 0.00
                vl_contabil[lserie]  = 0.00
                bc_icms[lserie]      = 0.00
                icms[lserie]         = 0.00
                isentos[lserie]      = 0.00
                outros[lserie]       = 0.00
                svl_contabil[lserie] = 0.00
                sbc_icms[lserie]     = 0.00
                sicms[lserie]        = 0.00


            serie.update({lserie:lserie})
            vl_completo.update({lserie:(vl_completo[lserie] + float(line[4].value))})
            
            if (line[3].value != '0000' ):
                vl_contabil.update      ({lserie: vl_contabil[lserie] + float(line[4].value if line[4].value else 0)})
                bc_icms.update          ({lserie: bc_icms[lserie]     + float(line[6].value if line[6].value else 0)})  
                icms.update             ({lserie: icms[lserie]        + float(line[7].value if line[7].value else 0)})
                isentos.update          ({lserie: isentos[lserie]     + float(line[8].value if line[8].value else 0)})
                outros.update           ({lserie: outros[lserie]      + float(line[9].value if line[9].value else 0)})
            
    for line in rasp:

        if (line[12].value != 'ICMS'):
            lserie = semespacos(line[1].value)

            if (serie.get(lserie)== None):
                vl_completo[lserie]  = 0.00
                vl_contabil[lserie]  = 0.00
                bc_icms[lserie]      = 0.00
                icms[lserie]         = 0.00
                isentos[lserie]      = 0.00
                outros[lserie]       = 0.00
                svl_contabil[lserie] = 0.00
                sbc_icms[lserie]     = 0.00
                sicms[lserie]        = 0.00

            serie.update({lserie:lserie})
            svl_contabil.update      ({lserie: svl_contabil[lserie] + float(line[8].value if line[8].value else 0)})
            sbc_icms.update          ({lserie: sbc_icms[lserie]     + float(line[11].value if line[11].value else 0)})  
            sicms.update             ({lserie: sicms[lserie]        + float(line[12].value if line[12].value else 0)})
    
    l=1

    
    RSPED.cell(row=l, column=1,  value="")
    RSPED.cell(row=l, column=2,  value="Regerado")
    RSPED.cell(row=l, column=3,  value="1 - Arquivos Regerados (sem CFOP000)")
    RSPED.cell(row=l, column=4,  value="")
    RSPED.cell(row=l, column=5,  value="")
    RSPED.cell(row=l, column=6,  value="")
    RSPED.cell(row=l, column=7,  value="")
    RSPED.cell(row=l, column=8,  value="2 - Bloco D Final")
    RSPED.cell(row=l, column=9,  value="")
    RSPED.cell(row=l, column=10, value="")
    RSPED.cell(row=l, column=11, value="Diferencas (2 - 1)")
    RSPED.cell(row=l, column=12, value="")
    RSPED.cell(row=l, column=13, value="")
    
    l=l+1

    RSPED.cell(row=l, column=1,  value="Serie")
    RSPED.cell(row=l, column=2,  value="Vl_Completo")
    RSPED.cell(row=l, column=3,  value="Vl_Liquido")
    RSPED.cell(row=l, column=4,  value="BC_ICMS")
    RSPED.cell(row=l, column=5,  value="Vl_ICMS")
    RSPED.cell(row=l, column=6,  value="Isentos")
    RSPED.cell(row=l, column=7,  value="Outros")
    RSPED.cell(row=l, column=8,  value="Vl_Liquido")
    RSPED.cell(row=l, column=9,  value="BC_ICMS")
    RSPED.cell(row=l, column=10, value="Vl_ICMS")
    RSPED.cell(row=l, column=11, value="Vl_Liquido")
    RSPED.cell(row=l, column=12, value="BC_ICMS")
    RSPED.cell(row=l, column=13, value="Vl_ICMS")
    
    l=l+1

    lseriet=[]
    for seriet in serie:
        lseriet.append(seriet)

    lseriet.sort()

    for lserie in lseriet:

        RSPED.cell(row=l, column=1,  value=serie[lserie])
        RSPED.cell(row=l, column=2,  value=vl_completo[lserie])
        RSPED.cell(row=l, column=3,  value=vl_contabil[lserie])
        RSPED.cell(row=l, column=4,  value=bc_icms[lserie])
        RSPED.cell(row=l, column=5,  value=icms[lserie])
        RSPED.cell(row=l, column=6,  value=isentos[lserie])
        RSPED.cell(row=l, column=7,  value=outros[lserie])
        RSPED.cell(row=l, column=8,  value=svl_contabil[lserie])
        RSPED.cell(row=l, column=9,  value=sbc_icms[lserie])
        RSPED.cell(row=l, column=10, value=sicms[lserie])
        RSPED.cell(row=l, column=11, value=svl_contabil[lserie]-vl_contabil[lserie])
        RSPED.cell(row=l, column=12, value=sbc_icms[lserie]-bc_icms[lserie])
        RSPED.cell(row=l, column=13, value=sicms[lserie]-icms[lserie])
        
        l=l+1
         
    somab="=SUM(B3:B" + str(l-1) + ")"
    somac="=SUM(C3:C" + str(l-1) + ")"
    somad="=SUM(D3:D" + str(l-1) + ")"
    somae="=SUM(E3:E" + str(l-1) + ")"
    somaf="=SUM(F3:F" + str(l-1) + ")"
    somag="=SUM(G3:G" + str(l-1) + ")"
    somah="=SUM(H3:H" + str(l-1) + ")"
    somai="=SUM(I3:I" + str(l-1) + ")"
    somaj="=SUM(J3:J" + str(l-1) + ")"
    somak="=SUM(K3:K" + str(l-1) + ")"
    somal="=SUM(L3:L" + str(l-1) + ")"
    somam="=SUM(M3:M" + str(l-1) + ")"

    RSPED.cell(row=l, column=1,  value = "TOTAIS")
    RSPED.cell(row=l, column=2,  value = somab)
    RSPED.cell(row=l, column=3,  value = somac)
    RSPED.cell(row=l, column=4,  value = somad)
    RSPED.cell(row=l, column=5,  value = somae)
    RSPED.cell(row=l, column=6,  value = somaf)
    RSPED.cell(row=l, column=7,  value = somag)
    RSPED.cell(row=l, column=8,  value = somah)
    RSPED.cell(row=l, column=9,  value = somai)
    RSPED.cell(row=l, column=10, value = somaj)
    RSPED.cell(row=l, column=11, value = somak)
    RSPED.cell(row=l, column=12, value = somal)
    RSPED.cell(row=l, column=13, value = somam)    

    formata_SPED(RSPED)
    
    
 
    
    lt = l


    
    
      
    arquivo_excel.save(R_VALIDA)

   
##########  ABA 2 -   Resumo GIA
##########  ABA 2 -   Resumo GIA
##########  ABA 2 -   Resumo GIA
##########  ABA 2 -   Resumo GIA
##########  ABA 2 -   Resumo GIA


    rcp = rc['REG X PROT']
    rcsp = cso['Resumo']
    rasp = asp['SPED_Regerado']
    
    cfop         = {}
    serie        = {}
    vl_completo  = {}
    vl_contabil  = {}
    bc_icms      = {}
    icms         = {}
    isentos      = {}
    outros       = {}
    svl_contabil = {}
    sbc_icms     = {}
    sicms        = {}
    sisentos     = {}
    soutros      = {}
    
    
    for line in rcp:
        lcfop = line[3].value
        if (line[12].value == 'REGERADO' and lcfop != '0000'):
            if (cfop.get(lcfop) == None):
                vl_contabil[lcfop]  = 0.00
                bc_icms[lcfop]      = 0.00
                icms[lcfop]         = 0.00
                isentos[lcfop]      = 0.00
                outros[lcfop]       = 0.00
                svl_contabil[lcfop] = 0.00
                sbc_icms[lcfop]     = 0.00
                sicms[lcfop]        = 0.00
                sisentos[lcfop]     = 0.00
                soutros[lcfop]      = 0.00
            cfop.update             ({lcfop:lcfop})
            vl_contabil.update      ({lcfop: vl_contabil[lcfop] + float(line[4].value)})
            bc_icms.update          ({lcfop: bc_icms[lcfop]     + float(line[6].value)})  
            icms.update             ({lcfop: icms[lcfop]        + float(line[7].value)})
            isentos.update          ({lcfop: isentos[lcfop]     + float(line[8].value)})
            outros.update           ({lcfop: outros[lcfop]      + float(line[9].value)})
            
    for line in rcsp:
       
        if (line[0].value != 'CFOP' and line[0].value != None):
            lcfop = line[0].value
            if (cfop.get(lcfop) == None):
                vl_contabil[lcfop]  = 0.00
                bc_icms[lcfop]      = 0.00
                icms[lcfop]         = 0.00
                isentos[lcfop]      = 0.00
                outros[lcfop]       = 0.00
                svl_contabil[lcfop] = 0.00
                sbc_icms[lcfop]     = 0.00
                sicms[lcfop]        = 0.00
                sisentos[lcfop]     = 0.00
                soutros[lcfop]      = 0.00
            cfop.update({lcfop:lcfop})
            svl_contabil.update({lcfop: svl_contabil[lcfop] + float(line[1].value)})
            sbc_icms.update({lcfop: sbc_icms[lcfop] + float(line[2].value)})  
            sicms.update({lcfop: sicms[lcfop] + float(line[3].value)})
            sisentos.update({lcfop: sisentos[lcfop] + float(line[4].value)})
            soutros.update({lcfop: soutros[lcfop] + float(line[5].value)})
    
    l=1
    # CABECALHO
    RGIA.cell(row=l, column=2,  value="REGERADO")
    RGIA.cell(row=l, column=7,  value="GIA")
    RGIA.cell(row=l, column=12, value="DIFERENCAS")

    
    l=l+1
    #DESCRICAO DAS COLUNAS 
    RGIA.cell(row=l, column=1,  value="CFOP")
    RGIA.cell(row=l, column=2,  value="VL_LIQUIDO")
    RGIA.cell(row=l, column=3,  value="BC_ICMS")
    RGIA.cell(row=l, column=4,  value="VL_ICMS")
    RGIA.cell(row=l, column=5,  value="ISENTOS")
    RGIA.cell(row=l, column=6,  value="OUTROS")
    RGIA.cell(row=l, column=7,  value="VL_LIQUIDO")
    RGIA.cell(row=l, column=8,  value="BC_ICMS")
    RGIA.cell(row=l, column=9,  value="VL_ICMS")
    RGIA.cell(row=l, column=10, value="ISENTOS")
    RGIA.cell(row=l, column=11, value="OUTROS")
    RGIA.cell(row=l, column=12, value="VL_LIQUIDO")
    RGIA.cell(row=l, column=13, value="BC_ICMS")
    RGIA.cell(row=l, column=14,  value="VL_ICMS")
    RGIA.cell(row=l, column=15, value="ISENTOS")
    RGIA.cell(row=l, column=16, value="OUTROS")    
    l=l+1


    lcfopt=[]
    for cfopt in cfop:
        lcfopt.append(cfopt)
   
    lcfopt.sort()

    for lcfop in lcfopt:
        #VALORES DAS COLUNAS
        RGIA.cell(row=l, column=1,  value=cfop[lcfop])
        RGIA.cell(row=l, column=2,  value=vl_contabil[lcfop])
        RGIA.cell(row=l, column=3,  value=bc_icms[lcfop])
        RGIA.cell(row=l, column=4,  value=icms[lcfop])
        RGIA.cell(row=l, column=5,  value=isentos[lcfop])
        RGIA.cell(row=l, column=6,  value=outros[lcfop])
        RGIA.cell(row=l, column=7,  value=svl_contabil[lcfop])
        RGIA.cell(row=l, column=8,  value=sbc_icms[lcfop])
        RGIA.cell(row=l, column=9, value=sicms[lcfop])
        RGIA.cell(row=l, column=10,  value=sisentos[lcfop])
        RGIA.cell(row=l, column=11,  value=soutros[lcfop])
        RGIA.cell(row=l, column=12, value=svl_contabil[lcfop]-vl_contabil[lcfop])
        RGIA.cell(row=l, column=13, value=sbc_icms[lcfop]-bc_icms[lcfop])
        RGIA.cell(row=l, column=14, value=sicms[lcfop]-icms[lcfop])
        RGIA.cell(row=l, column=15,  value=sisentos[lcfop]-isentos[lcfop])
        RGIA.cell(row=l, column=16,  value=soutros[lcfop]-outros[lcfop])
   
        
        l=l+1
    #SOMA DAS COLUNAS     
    somab="=SUM(B3:B" + str(l-1) + ")"
    somac="=SUM(C3:C" + str(l-1) + ")"
    somad="=SUM(D3:D" + str(l-1) + ")"
    somae="=SUM(E3:E" + str(l-1) + ")"
    somaf="=SUM(F3:F" + str(l-1) + ")"
    somag="=SUM(G3:G" + str(l-1) + ")"
    somah="=SUM(H3:H" + str(l-1) + ")"
    somai="=SUM(I3:I" + str(l-1) + ")"
    somaj="=SUM(J3:J" + str(l-1) + ")"
    somak="=SUM(K3:K" + str(l-1) + ")"
    somal="=SUM(L3:L" + str(l-1) + ")"
    somam="=SUM(M3:M" + str(l-1) + ")"
    soman="=SUM(N3:N" + str(l-1) + ")"
    somao="=SUM(O3:O" + str(l-1) + ")"
    somap="=SUM(P3:P" + str(l-1) + ")"

    RGIA.cell(row=l, column=1,  value = 'TOTAIS')
    RGIA.cell(row=l, column=2,  value = somab)
    RGIA.cell(row=l, column=3,  value = somac)
    RGIA.cell(row=l, column=4,  value = somad)
    RGIA.cell(row=l, column=5,  value = somae)
    RGIA.cell(row=l, column=6,  value = somaf)
    RGIA.cell(row=l, column=7,  value = somag)
    RGIA.cell(row=l, column=8,  value = somah)
    RGIA.cell(row=l, column=9,  value = somai)
    RGIA.cell(row=l, column=10, value = somaj)
    RGIA.cell(row=l, column=11, value = somak)
    RGIA.cell(row=l, column=12, value = somal)
    RGIA.cell(row=l, column=13, value = somam)    
    RGIA.cell(row=l, column=14, value = soman)    
    RGIA.cell(row=l, column=15, value = somao)    
    RGIA.cell(row=l, column=16, value = somap)    
 
    l=l+1

    somavc = "='Resumo SPED'!H" + str(lt) 
    somaci = "='Resumo SPED'!I" + str(lt) 
    somavi = "='Resumo SPED'!J" + str(lt) 

    RGIA.cell(row=l, column=6,  value = 'TOTAIS SPED')
    RGIA.cell(row=l, column=7, value = somavc )
    RGIA.cell(row=l, column=8, value = somaci )
    RGIA.cell(row=l, column=9, value = somavi )
    

    
    difvc = "=G" + str(l-1) + "-G" + str(l)
    difci = "=H" + str(l-1) + "-H" + str(l)
    difvi = "=I" + str(l-1) + "-I" + str(l)
    
    l=l+1
    RGIA.cell(row=l, column=6,  value = 'DIFERENCA')
    RGIA.cell(row=l, column=7, value = difvc )
    RGIA.cell(row=l, column=8, value = difci )
    RGIA.cell(row=l, column=9, value = difvi )
  
    formata_GIA(RGIA)
    
    
    arquivo_excel.save(R_VALIDA)

    return(ret)


if __name__ == "__main__":
    global arquivo_destino
    arquivo_destino= ""
    log("-"*100)
    log("#### ",dtf(), " INICIO DO VALIDA RELATORIOS SPED_CONVENIO115_GIA #### " ,sys.argv[0])
    variaveis = comum.carregaConfiguracoes(configuracoes)
    ret = processar()
    if (ret > 0) :
        if(arquivo_destino):
            if os.path.isfile(arquivo_destino):
                os.remove(arquivo_destino)
    log("#### Código de execução = ", ret)
    log("#### ",dtf(), " FIM DO VALIDA RELATORIOS SPED_CONVENIO115_GIA #### ",sys.argv[0])
    sys.exit(ret)
