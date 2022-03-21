#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-

"""
-----------------------------------------------------------------------------------------------------------------------
  SISTEMA ..: SPARTA
  MODULO ...: TESHUVA
  SCRIPT ...: relatorio_sped_telcom_resumo_serie_cfop.py
  CRIACAO ..: 06/12/2021
  AUTOR ....: Eduardo da Silva Ferreira (eduardof@kyros.com.br) 
              / KYROS Consultoria
  DESCRICAO : Este relatório tem por finalidade gerar um comparativo do SPED entre o Arquivo de PROTOCOLADO e o ENXERTADO . 
-----------------------------------------------------------------------------------------------------------------------
  HISTORICO :
    * 06/12/2021 - Eduardo da Silva Ferreira (eduardof@kyros.com.br) 
                   / KYROS Consultoria - Criacao do script.
------------------------------------------------------------------------------------------------------------------------
"""
#### PATRONIZACAO PARA O PAINEL DE EXECUCOES....
import sys
import os
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)
import configuracoes
import comum
import sql
import layout
import util
import vie 

#
import datetime
import traceback
import re
import string

#
import operator
import glob
from pathlib import Path
import openpyxl
import copy
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# Demais configuração
comum.carregaConfiguracoes(configuracoes)
#layout.carregaLayout()
log.gerar_log_em_arquivo = True

class cls_sped(object):
    """
        Classe para guardar as informações de SPED 
    """
    nr_count = 0
    def __init__(self):
        cls_sped.nr_count += 1
        self.REG               = ''
        self.SEQ               = int(0)
        self.TIPO              = ''
        self.VOLUME            = int(0)
        self.CFOP              = ''
        self.SERIE             = ''
        self.MODELO            = ''
        self.NF_INI            = ''
        self.DT_INI            = ''
        self.NF_FIM            = ''
        self.DT_FIM            = ''
        self.NM_ARQ            = ''
        self.HASHCODE          = ''
        self.CST               = ''
        self.VL_LIQUIDO        = float(0)
        self.VL_BASE           = float(0)
        self.VL_ICMS           = float(0)

def as_text(value):
    if value is None:
        return ""
    return str(value)      
    
def fnc_column_dimensions(p_ob_ws):
    for column_cells in p_ob_ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        p_ob_ws.column_dimensions[column_cells[0].column_letter].width = length+5        
    
def ob_global():
    """
        Funcao falsa apenas para sergir de apoio para armazenar as variaveis globais 
    """
    pass


def fnc_ultimo_arquivo_diretorio(p_ds_mascara,p_ds_diretorio):
    """
        Funcao para retornar arquivos existentes
    """
    v_nr_qt = 0
    v_nm_arquivo = ""
    v_ds_dir = Path(p_ds_diretorio)
    v_nm_arq = v_ds_dir.glob(p_ds_mascara)
    v_cc_procura_arquivos = sorted(v_nm_arq, reverse=False)
    if v_cc_procura_arquivos:        
        for arquivo in v_cc_procura_arquivos:
            v_nr_qt = v_nr_qt + 1
            v_nm_arquivo = str(arquivo)

    return(v_nm_arquivo)

def fnc_conectar_banco_dados():
    """
        Funcao para conectar na base de dados 
    """
    try:
        
        ob_global.gv_ob_conexao = sql.geraCnxBD(configuracoes)
        v_ds_sql="""
        SELECT 'PAINELEXECUCAO_'||TO_CHAR(SYSDATE,'YYYYMMDD_HH24MISS') NM_JOB FROM DUAL
        """
        ob_global.gv_ob_conexao.executa(v_ds_sql)
        
        v_ob_cursor = ob_global.gv_ob_conexao.fetchone()
        if (v_ob_cursor):    
            for campo in v_ob_cursor:
                log(str(campo) + " SUCESSO CONEXAO BANCO DE DADOS") 
                ob_global.gv_nm_job = str(campo)
                break
        
        else:
            log("ERRO CONEXAO BANCO DE DADOS ") 
            return 91
            
        return 0
                
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO CONEXAO BANCO DE DADOS .: " + str(e) + " - TRACE - " + v_ds_trace)
        return 91

def fnc_processa_sped(p_nm_arquivo,p_tp_arquivo="PROTOCOLADO"):
    """
        Funcao processa arquivo do tipo SPED
    """
    
    # inicializa a variaveis de controle
    v_cc_resumo={}
    v_cc_reg_rel={} 
    v_nr_contador_nf = 0
    v_fl_nf = 0
    v_ds_linha = ""
    v_cc_chave_01 = ""
    v_cc_chave_02 = ""
    try:
    
        if os.path.isfile(p_nm_arquivo):
            
            log(p_tp_arquivo + " >> Processando leitura do arquivo : " + p_nm_arquivo)
            
            v_ob_encoding = comum.encodingDoArquivo( p_nm_arquivo )
            v_ob_fd = open(p_nm_arquivo, 'r', encoding=v_ob_encoding)
                        
            for v_cc_linha in v_ob_fd:
                # Quebra em vetor        
                v_ds_linha = v_cc_linha
                v_cc_dados = v_ds_linha.split("|")
    
                if (len(v_cc_dados) < 7):
                    continue
                
                # Valida se os tipos        
                if v_cc_dados[1].upper().strip() \
                in ('D695','D696'):
                    # inicio if
                    
                    if v_cc_dados[1].upper().strip() \
                    in ('D695'):
                        ob_sped = cls_sped()                    
                        v_nr_contador_nf += 1
                        v_fl_nf = 0 
                        ob_sped.SEQ               = v_fl_nf
                        ob_sped.REG               = v_cc_dados[1].upper().strip()
                        ob_sped.TIPO              = p_tp_arquivo
                        ob_sped.MODELO            = str(v_cc_dados[2].upper().strip())
                        ob_sped.SERIE             = str(v_cc_dados[3].upper().strip())
                        ob_sped.NF_INI            = str(v_cc_dados[4].upper().strip())
                        ob_sped.NF_FIM            = str(v_cc_dados[5].upper().strip())
                        ob_sped.DT_INI            = str(v_cc_dados[6].upper().strip())
                        ob_sped.DT_FIM            = str(v_cc_dados[7].upper().strip())
                        ob_sped.NM_ARQ            = str(v_cc_dados[8].upper().strip())
                        ob_sped.HASHCODE          = str(v_cc_dados[9].strip())
                        ob_sped.CFOP              = ''
                        ob_sped.CST               = ''
                        ob_sped.VL_LIQUIDO        = float(0)
                        ob_sped.VL_BASE           = float(0)
                        ob_sped.VL_ICMS           = float(0)
                        try:
                            ob_sped.VOLUME = int(str(ob_sped.NM_ARQ).split(".")[1])
                        except:
                            ob_sped.VOLUME = int(0)
                    if v_cc_dados[1].upper().strip() \
                    in ('D696'):
                        ob_sped.REG    = ob_sped.REG \
                                         + "|" + v_cc_dados[1].upper().strip()
                        
                        v_cc_chave_01     = ob_sped.TIPO \
                                         + "|" + ob_sped.MODELO \
                                         + "|" + ob_sped.SERIE \
                                         + "|" + ob_sped.NF_INI \
                                         + "|" + ob_sped.NF_FIM \
                                         + "|" + ob_sped.DT_INI \
                                         + "|" + ob_sped.DT_FIM \
                                         + "|" + ob_sped.NM_ARQ \
                                         + "|" + ob_sped.HASHCODE \
                                         + "|" + str(v_cc_dados[2].upper().strip()) \
                                         + "|" + str(v_cc_dados[3].upper().strip())
                        
                        v_cc_chave_02     = str(v_cc_dados[3].upper().strip()) \
                                         + "|" + ob_sped.SERIE                                         
                        
                        if v_cc_chave_02 in v_cc_resumo:
                            try:
                                v_cc_resumo[v_cc_chave_02]['VL_LIQUIDO'] += float(str(v_cc_dados[5].upper().strip()).replace(',', '.'))
                            except:
                                v_cc_resumo[v_cc_chave_02]['VL_LIQUIDO'] += float(0)                                
                            try:
                                v_cc_resumo[v_cc_chave_02]['VL_BASE'] += float(str(v_cc_dados[6].upper().strip()).replace(',', '.'))
                            except:
                                v_cc_resumo[v_cc_chave_02]['VL_BASE'] += float(0)                                
                            try:
                                v_cc_resumo[v_cc_chave_02]['VL_ICMS'] += float(str(v_cc_dados[7].upper().strip()).replace(',', '.'))
                            except:
                                v_cc_resumo[v_cc_chave_02]['VL_ICMS'] += float(0)   
                        else:
                            v_vl = {}
                            v_vl_aux   = float(0)        
                            v_vl['EXISTE'] = 0    
                            try:
                                v_vl_aux = float(str(v_cc_dados[5].upper().strip()).replace(',', '.'))
                            except:
                                v_vl_aux = float(0)   
                            v_vl['VL_LIQUIDO'] = v_vl_aux 
                            try:
                                v_vl_aux = float(str(v_cc_dados[6].upper().strip()).replace(',', '.'))
                            except:
                                v_vl_aux = float(0)   
                            v_vl['VL_BASE'] = v_vl_aux    
                            try:
                                v_vl_aux = float(str(v_cc_dados[7].upper().strip()).replace(',', '.'))
                            except:
                                v_vl_aux = float(0)   
                            v_vl['VL_ICMS'] = v_vl_aux
                            v_cc_resumo[v_cc_chave_02] = v_vl
                            
                        if v_fl_nf: 
                            if v_cc_chave_01 in v_cc_reg_rel:
                                try:
                                    v_cc_reg_rel[v_cc_chave_01].VL_LIQUIDO += float(str(v_cc_dados[5].upper().strip()).replace(',', '.'))
                                except:
                                    v_cc_reg_rel[v_cc_chave_01].VL_LIQUIDO += float(0)                                
                                try:
                                    v_cc_reg_rel[v_cc_chave_01].VL_BASE += float(str(v_cc_dados[6].upper().strip()).replace(',', '.'))
                                except:
                                    v_cc_reg_rel[v_cc_chave_01].VL_BASE += float(0)                                
                                try:
                                    v_cc_reg_rel[v_cc_chave_01].VL_ICMS += float(str(v_cc_dados[7].upper().strip()).replace(',', '.'))
                                except:
                                    v_cc_reg_rel[v_cc_chave_01].VL_ICMS += float(0)   
                                continue            
                        
                        v_fl_nf += 1
                        ob_sped_resumo = copy.deepcopy(ob_sped)
                        ob_sped_resumo.SEQ   = v_fl_nf
                        ob_sped_resumo.CST   = str(v_cc_dados[2].upper().strip())
                        ob_sped_resumo.CFOP  = str(v_cc_dados[3].upper().strip())
                        try:
                            ob_sped_resumo.VL_LIQUIDO = float(str(v_cc_dados[5].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_resumo.VL_LIQUIDO = float(0)                                
                        try:
                            ob_sped_resumo.VL_BASE = float(str(v_cc_dados[6].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_resumo.VL_BASE = float(0)                                
                        try:
                            ob_sped_resumo.VL_ICMS = float(str(v_cc_dados[7].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_resumo.VL_ICMS = float(0)                                
                        v_cc_reg_rel[v_cc_chave_01] = ob_sped_resumo                        
                        
                #if len(v_cc_reg_rel) > 15:
                #    break
   
            # Verifica se o arquivo teve algum processamentp                
            if not v_cc_reg_rel:
                log("# " + str(len(v_cc_reg_rel)) + " >> " + " -> Não processou nenhum dados do arquivo : " + p_nm_arquivo)
                return None, None 
            else:
                log("# " + str(len(v_cc_reg_rel)) + " >> " + " -> Dados processados com sucesso do arquivo : " + p_nm_arquivo)            
                return [value for value in v_cc_reg_rel.values()], v_cc_resumo
        
        else:
            return None, None    

    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO LEITURA DO ARQUIVO .: " + str(e) + " - TRACE - " + v_ds_trace + " >> ARQUIVO: " + p_nm_arquivo + " >> TIPO: " + p_tp_arquivo + " >> LINHA: " + v_ds_linha)
        return None, None

       
def fnc_validar_entrada():
    """
        Retorna a validação de entrada e dos arquivos de configuração
    """
    try:
        v_nr_retorno = 0        
        
        comum.addParametro( 'MESANO',  None, "MESANO (MMYYYY) dos arquivos", True, '122015' )
        comum.addParametro( 'UF', None, 'Unidade Federativa (UF) dos arquivo  ', True, 'SP')
        comum.addParametro( 'IE', None, 'Inscricao Estadual (IE) dos arquivo  ', True, '108383949112')
        
        # Validacao dos parametros de entrada
        if not comum.validarParametros() :
            v_nr_retorno = 91
            return v_nr_retorno     
            
        else:
            ob_global.gv_mes_ano = comum.getParametro('MESANO').upper().strip()
            ob_global.gv_uf = comum.getParametro('UF').upper().strip()
            ob_global.gv_ie = comum.getParametro('IE').upper().strip()
        
            if (len(ob_global.gv_mes_ano) != 6):
                log("PARAMETRO MES ANO: Invalido!") 
                v_nr_retorno = 91

            if not v_nr_retorno:
                ob_global.gv_mes = ob_global.gv_mes_ano[0:2]
                ob_global.gv_ano = ob_global.gv_mes_ano[2:6]

            if not v_nr_retorno:
                try:
                    if (int(ob_global.gv_mes) < 1
                    or int(ob_global.gv_mes) > 12 
                    ):
                        log("PARAMETRO MES : Invalido!") 
                        v_nr_retorno = 91
                except:
                    log("PARAMETRO MES : Invalido!") 
                    v_nr_retorno = 91

            if not v_nr_retorno :
                try:
                    if (
                       int(ob_global.gv_ano) > datetime.datetime.now().year
                    or int(ob_global.gv_ano) < (datetime.datetime.now().year)-100
                    ):
                        log("PARAMETRO ANO : Invalido!") 
                        v_nr_retorno = 91
                except:
                    log("PARAMETRO ANO : Invalido!") 
                    v_nr_retorno = 91
            
            if not v_nr_retorno :
                if len(ob_global.gv_uf) != 2:
                    log("PARAMETRO UF: Invalido!") 
                    v_nr_retorno = 91

            if not v_nr_retorno :
                try:
                    v_nr_iei = re.sub('[^0-9]','',ob_global.gv_ie)
                    if ( (v_nr_iei == "") or (v_nr_iei == "''") or (v_nr_iei == '""') or (int("0"+v_nr_iei) == 0)):
                        log("PARAMETRO IE : Invalido!") 
                        v_nr_retorno = 91        
                except:
                    log("PARAMETRO IE : Invalido!") 
                    v_nr_retorno = 91    
        
        return v_nr_retorno            
        
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO VALIDAÇÃO DOS PARAMETROS DE ENTRADA: " + str(e)+ " >> " + v_ds_trace)
        return 93        

def fnc_validar_configuracao():
    """
        Retorna a validação dos arquivos de configuração
    """
    try:
        v_nr_retorno = 0        

        ob_global.gv_diretorio_sped_fiscal_enxertado = configuracoes.dir_sped_fiscal_enxertado.replace("<<MM>>",ob_global.gv_mes).replace("<<AAAA>>",ob_global.gv_ano).replace("<<UF>>",ob_global.gv_uf).replace("<<IE>>",ob_global.gv_ie).replace("<<MESANO>>",ob_global.gv_mes_ano).strip()
        ob_global.gv_diretorio_sped_fiscal_protocolado = configuracoes.dir_sped_fiscal_protocolado.replace("<<MM>>",ob_global.gv_mes).replace("<<AAAA>>",ob_global.gv_ano).replace("<<UF>>",ob_global.gv_uf).replace("<<IE>>",ob_global.gv_ie).replace("<<MESANO>>",ob_global.gv_mes_ano).strip()
        ob_global.gv_diretorio_insumo_sped = configuracoes.dir_insumo_sped.replace("<<MM>>",ob_global.gv_mes).replace("<<AAAA>>",ob_global.gv_ano).replace("<<UF>>",ob_global.gv_uf).replace("<<IE>>",ob_global.gv_ie).replace("<<MESANO>>",ob_global.gv_mes_ano).strip()
        
        ob_global.gv_arq_sped_fiscal_enxertado = configuracoes.arq_sped_fiscal_enxertado.replace("<<MM>>",ob_global.gv_mes).replace("<<AAAA>>",ob_global.gv_ano).replace("<<UF>>",ob_global.gv_uf).replace("<<IE>>",ob_global.gv_ie).replace("<<MESANO>>",ob_global.gv_mes_ano).replace("<<NNN>>","*").strip()
        ob_global.gv_arq_sped_fiscal_protocolado = configuracoes.arq_sped_fiscal_protocolado.replace("<<MM>>",ob_global.gv_mes).replace("<<AAAA>>",ob_global.gv_ano).replace("<<UF>>",ob_global.gv_uf).replace("<<IE>>",ob_global.gv_ie).replace("<<MESANO>>",ob_global.gv_mes_ano).replace("<<NNN>>","*").strip()
        ob_global.gv_arq_insumo_sped = configuracoes.arq_insumo_saida.replace("<<MM>>",ob_global.gv_mes).replace("<<AAAA>>",ob_global.gv_ano).replace("<<UF>>",ob_global.gv_uf).replace("<<IE>>",ob_global.gv_ie).replace("<<MESANO>>",ob_global.gv_mes_ano).replace("<<NNN>>","*").strip()

        if not os.path.isdir(ob_global.gv_diretorio_sped_fiscal_enxertado):
            log("Diretório não localizado sped ENXERTADO : " + ob_global.gv_diretorio_sped_fiscal_enxertado)        
            return 1
        else:    
            ob_global.gv_lst_arq_sped_fiscal_enxertado = fnc_ultimo_arquivo_diretorio(ob_global.gv_arq_sped_fiscal_enxertado,ob_global.gv_diretorio_sped_fiscal_enxertado).strip()
            if not ob_global.gv_lst_arq_sped_fiscal_enxertado:
                log(ob_global.gv_lst_arq_sped_fiscal_enxertado + " Arquivo não localizado sped ENXERTADO : " + ob_global.gv_diretorio_sped_fiscal_enxertado)        
                return 1
        
        if not os.path.isdir(ob_global.gv_diretorio_sped_fiscal_protocolado):
            log("Diretório não localizado sped protocolado : " + ob_global.gv_diretorio_sped_fiscal_protocolado)        
            return 1
        else:    
            ob_global.gv_lst_arq_sped_fiscal_protocolado = fnc_ultimo_arquivo_diretorio(ob_global.gv_arq_sped_fiscal_protocolado,ob_global.gv_diretorio_sped_fiscal_protocolado).strip()
            if not ob_global.gv_lst_arq_sped_fiscal_protocolado:
                log(ob_global.gv_lst_arq_sped_fiscal_protocolado + " Arquivo não localizado sped protocolado : " + ob_global.gv_diretorio_sped_fiscal_protocolado)        
                return 1
                
        if not os.path.isdir(ob_global.gv_diretorio_insumo_sped):
            log("Diretório não localizado insumo, porem está sendo criado : " + ob_global.gv_diretorio_insumo_sped)     
            try:    
                os.mkdir(ob_global.gv_diretorio_insumo_sped)
            except Exception as e:
                v_ds_trace = traceback.format_exc()
                log("ERRO CRIAÇÃO DO DIRETÓRIO: " + str(e)+ " >> " + v_ds_trace)
                return 93     
        
        if not ob_global.gv_arq_insumo_sped:
            log("Arquivo invalido insumo : " + ob_global.gv_arq_insumo_sped)        
            return 1        
        else:    
            ob_global.gv_lst_arq_insumo_sped = fnc_ultimo_arquivo_diretorio(ob_global.gv_arq_insumo_sped,ob_global.gv_diretorio_insumo_sped).strip()
            if len(ob_global.gv_lst_arq_insumo_sped) < 4:
                ob_global.gv_lst_arq_insumo_sped = os.path.join(ob_global.gv_diretorio_insumo_sped, ob_global.gv_arq_insumo_sped.replace("*","001"))
            else:
                if (ob_global.gv_lst_arq_insumo_sped.strip().upper().endswith(".TXT")
                or ob_global.gv_lst_arq_insumo_sped.strip().upper().find(".")):
                    ob_global.gv_lst_arq_insumo_sped = ob_global.gv_lst_arq_insumo_sped.split(".")[0][:-3]+ str(int(ob_global.gv_lst_arq_insumo_sped.split(".")[0][-3:])+1).rjust(3,'0') +  "." +ob_global.gv_lst_arq_insumo_sped.split(".")[1] 
                else:
                    ob_global.gv_lst_arq_insumo_sped = os.path.join(ob_global.gv_diretorio_insumo_sped, ob_global.gv_arq_insumo_sped.replace("*","001"))


        return v_nr_retorno
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO VALIDAÇÃO DOS PARAMETROS DE CONFIGURACAO: " + str(e)+ " >> " + v_ds_trace)
        return 93     
 
def fnc_imprimi_variaveis():
    """
        Imprimi variaveis
    """
    try:
    
        if ob_global.gv_ob_conexao:
            log("conexao  : ATIVO") 
        else:
            log("conexao  : DESATIVO")
            
        log("job      : " + str(ob_global.gv_nm_job)) 
        log("mes_ano  : " + str(ob_global.gv_mes_ano)) 
        log("uf       : " + str(ob_global.gv_uf)) 
        log("ie       : " + str(ob_global.gv_ie)) 
        log("mes      : " + str(ob_global.gv_mes)) 
        log("ano      : " + str(ob_global.gv_ano)) 
        
        log("diretorio_sped_fiscal_enxertado       : " + str(ob_global.gv_diretorio_sped_fiscal_enxertado)) 
        log("diretorio_sped_fiscal_protocolado : " + str(ob_global.gv_diretorio_sped_fiscal_protocolado))
        log("diretorio_insumo_sped             : " + str(ob_global.gv_diretorio_insumo_sped))
        
        log("arq_sped_fiscal_enxertado             : " + str(ob_global.gv_arq_sped_fiscal_enxertado))
        log("arq_sped_fiscal_protocolado       : " + str(ob_global.gv_arq_sped_fiscal_protocolado))
        log("arq_insumo_sped                   : " + str(ob_global.gv_arq_insumo_sped))
        
        log("lst_arq_sped_fiscal_enxertado         : " + str(ob_global.gv_lst_arq_sped_fiscal_enxertado))
        log("lst_arq_sped_fiscal_protocolado   : " + str(ob_global.gv_lst_arq_sped_fiscal_protocolado))
        log("lst_arq_insumo_sped               : " + str(ob_global.gv_lst_arq_insumo_sped))
        
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO IMPRESSAO DE VARIAVEIS: " + str(e)+ " >> " + v_ds_trace)
 
def fnc_validar_leitura_carregamento_arquivo():
    """
        Retorna a validação e leitura do carregamento do arquivo
    """
    v_nr_retorno = 0     
    try:
        ob_global.gv_ob_lista_sped_enxertado, ob_global.gv_ob_dict_sped_enxertado_resumo = fnc_processa_sped(ob_global.gv_lst_arq_sped_fiscal_enxertado,"ENXERTADO")
        if not ob_global.gv_ob_lista_sped_enxertado:
            log("FALHA NA LEITURA DO ARQUIVO DE SPED ENXERTADO: " + ob_global.gv_lst_arq_sped_fiscal_enxertado)
            return 93     

        ob_global.gv_ob_lista_sped_protocolado, ob_global.gv_ob_dict_sped_protocolado_resumo = fnc_processa_sped(ob_global.gv_lst_arq_sped_fiscal_protocolado,"PROTOCOLADO")
        if not ob_global.gv_ob_lista_sped_protocolado:
            log("FALHA NA LEITURA DO ARQUIVO DE SPED PROTOCOLADO: " + ob_global.gv_lst_arq_sped_fiscal_protocolado)
            return 93    
               
        ob_global.gv_ob_lista_sped_enxertado.sort(key=operator.attrgetter('VOLUME','CST','CFOP','SERIE','MODELO','DT_INI','NF_INI','DT_FIM','NF_FIM'))
        ob_global.gv_ob_lista_sped_protocolado.sort(key=operator.attrgetter('VOLUME','CST','CFOP','SERIE','MODELO','DT_INI','NF_INI','DT_FIM','NF_FIM'))
       
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO LEITURA DO CARREGAMENTO DO ARQUIVO: " + str(e)+ " >> " + v_ds_trace)
        v_nr_retorno = 93
    
    return v_nr_retorno         
    
def fnc_processar():
    """
        Funcao principal para processar as informacoes
    """
    try:
        v_nr_retorno = 0
        ob_global.gv_ob_existe_lista_sped_enxertado = []
        
        log("Criando arquivo : " + ob_global.gv_lst_arq_insumo_sped)
        #### Cria a planilha em memória....
        v_obj_arq_excel = Workbook()

        v_obj_wks_excel_nr_0 = v_obj_arq_excel.active
        v_obj_wks_excel_nr_0.title = "SPED - ENXERTADO x PROTOCOLADO"
        log("Criando wks : " + v_obj_wks_excel_nr_0.title)
        v_nr_linha = 0        
        v_nr_linha += 1
        v_obj_wks_excel_nr_0.cell(v_nr_linha,1 ,'VOLUME')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,2 ,'CST')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,3 ,'CFOP')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,4 ,'SERIE')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,5 ,'MODELO \nPROT')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,6 ,'MODELO \nENXE')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,7 ,'MODELO \nCOMPARA')                
        v_obj_wks_excel_nr_0.cell(v_nr_linha,8 ,'NF_INI \nPROT')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,9 ,'NF_INI \nENXE')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,10 ,'NF_INI \nCOMPARA')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,11 ,'NF_FIM \nPROT')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,12 ,'NF_FIM \nENXE')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,13 ,'NF_FIM \nCOMPARA')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,14 ,'DT_INI \nPROT')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,15 ,'DT_INI \nENXE')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,16 ,'DT_INI \nCOMPARA')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,17 ,'DT_FIM \nPROT')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,18 ,'DT_FIM \nENXE')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,19 ,'DT_FIM \nCOMPARA')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,20 ,'NM_ARQ \nPROT')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,21 ,'NM_ARQ \nENXE')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,22 ,'NM_ARQ \nCOMPARA')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,23 ,'HASHCODE \nPROT')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,24 ,'HASHCODE \nENXE')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,25 ,'HASHCODE \nCOMPARA')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,26 ,'VL_LIQUIDO \nPROT')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,27 ,'VL_LIQUIDO \nENXE')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,28 ,'VL_LIQUIDO \nCOMPARA')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,29 ,'VL_BASE \nPROT')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,30 ,'VL_BASE \nENXE')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,31 ,'VL_BASE \nCOMPARA')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,32 ,'VL_ICMS \nPROT')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,33 ,'VL_ICMS \nENXE')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,34 ,'VL_ICMS \nCOMPARA')
        v_obj_wks_excel_nr_0.cell(v_nr_linha,35 ,'VALIDAÇÃO')
        for nr_col in range(1,36):
            v_obj_wks_excel_nr_0.cell(v_nr_linha,nr_col).font=Font(bold=True)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,nr_col).alignment = Alignment(horizontal='center')
            if (nr_col > 4 and nr_col < 35) and (((nr_col-1) % 3) == 0):
                v_obj_wks_excel_nr_0.cell(v_nr_linha,nr_col).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")        
        for idx, reg in enumerate(ob_global.gv_ob_lista_sped_protocolado):    
            ob_sped    = None
            v_fl_achou = 0 
            v_ds_grupo = "0-EXISTE PROTOCOLADO e NÃO EXISTE NO ENXERTADO"  
            for idxa, rega in enumerate(ob_global.gv_ob_lista_sped_enxertado): 
                if rega.VOLUME > reg.VOLUME:
                    break                
                if rega.VOLUME != reg.VOLUME:
                    continue
                if rega.CST > reg.CST:
                    break
                if rega.CST != reg.CST:
                    continue    
                if rega.CFOP > reg.CFOP:
                    break
                if rega.CFOP != reg.CFOP:
                    continue
                if rega.SERIE > reg.SERIE:
                    break
                if rega.SERIE != reg.SERIE:
                    continue
                if rega not in ob_global.gv_ob_existe_lista_sped_enxertado:
                    if v_fl_achou < 1:
                        v_fl_achou = 1
                        v_ds_grupo = "1-VOLUME+CST+CFOP+SERIE"
                        ob_sped = rega
                    
                    if reg.MODELO == rega.MODELO:
                        if v_fl_achou < 2:
                            v_fl_achou = 2
                            v_ds_grupo = "2-VOLUME+CST+CFOP+SERIE+MODELO"
                            ob_sped = rega
                            break                        
                                
            if v_fl_achou:
                ob_global.gv_ob_existe_lista_sped_enxertado.append(ob_sped)
                
            v_nr_linha += 1
            v_obj_wks_excel_nr_0.cell(v_nr_linha,1  ,reg.VOLUME)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,2  ,reg.CST)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,3  ,reg.CFOP)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,4  ,reg.SERIE)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,5  ,reg.MODELO)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,6  ,ob_sped.MODELO if v_fl_achou else "")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,7  ,"VERDADEIRO" if v_fl_achou and reg.MODELO == ob_sped.MODELO else "FALSO")                
            v_obj_wks_excel_nr_0.cell(v_nr_linha,8  ,reg.NF_INI)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,9  ,ob_sped.NF_INI if v_fl_achou else "")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,10  ,"VERDADEIRO" if v_fl_achou and int(reg.NF_INI) == int(ob_sped.NF_INI) else "FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,11 ,reg.NF_FIM)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,12 ,ob_sped.NF_FIM if v_fl_achou else "")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,13 ,"VERDADEIRO" if v_fl_achou and int(reg.NF_FIM) == int(ob_sped.NF_FIM) else "FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,14 ,reg.DT_INI)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,15 ,ob_sped.DT_INI if v_fl_achou else "")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,16 ,"VERDADEIRO" if v_fl_achou and reg.DT_INI == ob_sped.DT_INI else "FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,17 ,reg.DT_FIM)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,18 ,ob_sped.DT_FIM if v_fl_achou else "")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,19 ,"VERDADEIRO" if v_fl_achou and reg.DT_FIM == ob_sped.DT_FIM else "FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,20 ,reg.NM_ARQ)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,21 ,ob_sped.NM_ARQ if v_fl_achou else "")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,22 ,"VERDADEIRO" if v_fl_achou and reg.NM_ARQ == ob_sped.NM_ARQ else "FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,23 ,reg.HASHCODE)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,24 ,ob_sped.HASHCODE if v_fl_achou else "")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,25 ,"VERDADEIRO" if v_fl_achou and reg.HASHCODE == ob_sped.HASHCODE else "FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,26 ,reg.VL_LIQUIDO)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,27 ,ob_sped.VL_LIQUIDO if v_fl_achou else "")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,28 ,"VERDADEIRO" if v_fl_achou and reg.VL_LIQUIDO == ob_sped.VL_LIQUIDO else "FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,29 ,reg.VL_BASE)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,30 ,ob_sped.VL_BASE if v_fl_achou else "")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,31 ,"VERDADEIRO" if v_fl_achou and reg.VL_BASE == ob_sped.VL_BASE else "FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,32 ,reg.VL_ICMS)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,33 ,ob_sped.VL_ICMS if v_fl_achou else "")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,34 ,"VERDADEIRO" if v_fl_achou and reg.VL_ICMS == ob_sped.VL_ICMS else "FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,35 ,v_ds_grupo)
            for idx,nr_col in enumerate(range(26,34)):
                if not (((nr_col-1) % 3) == 0):        
                    v_obj_wks_excel_nr_0.cell(v_nr_linha,nr_col).number_format = "#,##0.00"             
        
        v_ds_grupo = "100-EXISTE ENXERTADO e NÃO EXISTE NO PROTOCOLADO"
        ob_lista_sped = list(set(ob_global.gv_ob_lista_sped_enxertado) - set(ob_global.gv_ob_existe_lista_sped_enxertado))    
        for idxa, ob_sped in enumerate(ob_lista_sped): 
            v_nr_linha += 1
            v_obj_wks_excel_nr_0.cell(v_nr_linha,1  ,ob_sped.VOLUME)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,2  ,ob_sped.CST)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,3  ,ob_sped.CFOP)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,4  ,ob_sped.SERIE)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,5  ,"")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,6  ,ob_sped.MODELO)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,7  ,"FALSO")                
            v_obj_wks_excel_nr_0.cell(v_nr_linha,8  ,"")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,9  ,ob_sped.NF_INI)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,10 ,"FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,11 ,"")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,12 ,ob_sped.NF_FIM)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,13 ,"FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,14 ,"")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,15 ,ob_sped.DT_INI)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,16 ,"FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,17 ,"")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,18 ,ob_sped.DT_FIM)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,19 ,"FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,20 ,"")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,21 ,ob_sped.NM_ARQ)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,22 ,"FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,23 ,"")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,24 ,ob_sped.HASHCODE)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,25 ,"FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,26 ,"")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,27 ,ob_sped.VL_LIQUIDO)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,28 ,"FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,29 ,"")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,30 ,ob_sped.VL_BASE)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,31 ,"FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,32 ,"")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,33 ,ob_sped.VL_ICMS)
            v_obj_wks_excel_nr_0.cell(v_nr_linha,34 ,"FALSO")
            v_obj_wks_excel_nr_0.cell(v_nr_linha,35 ,v_ds_grupo)
            for idx,nr_col in enumerate(range(26,34)):
                if not (((nr_col-1) % 3) == 0):           
                    v_obj_wks_excel_nr_0.cell(v_nr_linha,nr_col).number_format = "#,##0.00"             
                        
        v_ob_filter_wks_0 = "A1:" + get_column_letter(v_obj_wks_excel_nr_0.max_column) \
                            + str(v_obj_wks_excel_nr_0.max_row)
        v_obj_wks_excel_nr_0.auto_filter.ref = v_ob_filter_wks_0        


        v_obj_wks_excel_nr_1 = v_obj_arq_excel.create_sheet("RESUMO POR CFOP E SERIE", 1)
        log("Planilha : " + v_obj_wks_excel_nr_1.title)
        v_nr_linha = 0

        v_nr_linha += 1
        v_obj_wks_excel_nr_1.cell(v_nr_linha,1  ,'CFOP')
        v_obj_wks_excel_nr_1.cell(v_nr_linha,2  ,'SERIE')
        v_obj_wks_excel_nr_1.cell(v_nr_linha,3  ,'VL_LIQUIDO \nPROT')
        v_obj_wks_excel_nr_1.cell(v_nr_linha,4  ,'VL_LIQUIDO \nENXE')
        v_obj_wks_excel_nr_1.cell(v_nr_linha,5  ,'VL_LIQUIDO \nCOMPARA')
        v_obj_wks_excel_nr_1.cell(v_nr_linha,6  ,'VL_BASE \nPROT')
        v_obj_wks_excel_nr_1.cell(v_nr_linha,7  ,'VL_BASE \nENXE')
        v_obj_wks_excel_nr_1.cell(v_nr_linha,8  ,'VL_BASE \nCOMPARA')
        v_obj_wks_excel_nr_1.cell(v_nr_linha,9  ,'VL_ICMS \nPROT')
        v_obj_wks_excel_nr_1.cell(v_nr_linha,10 ,'VL_ICMS \nENXE')
        v_obj_wks_excel_nr_1.cell(v_nr_linha,11 ,'VL_ICMS \nCOMPARA')
        v_obj_wks_excel_nr_1.cell(v_nr_linha,12 ,'VALIDAÇÃO')
        for nr_col in range(1,13):
            v_obj_wks_excel_nr_1.cell(v_nr_linha,nr_col).font=Font(bold=True)
            v_obj_wks_excel_nr_1.cell(v_nr_linha,nr_col).alignment = Alignment(horizontal='center')
            if (nr_col > 5 and nr_col < 12) and (((nr_col+1) % 3) == 0):
                v_obj_wks_excel_nr_1.cell(v_nr_linha,nr_col).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")        

        for key in sorted(ob_global.gv_ob_dict_sped_protocolado_resumo.keys()) :        
            v_nr_linha += 1
            v_cc_chave = key.split("|")
            v_cc_cfop = v_cc_chave[0]
            v_cc_serie = v_cc_chave[1]    
            v_obj_wks_excel_nr_1.cell(v_nr_linha,1  ,v_cc_cfop)
            v_obj_wks_excel_nr_1.cell(v_nr_linha,2  ,v_cc_serie)
            v_obj_wks_excel_nr_1.cell(v_nr_linha,3  ,ob_global.gv_ob_dict_sped_protocolado_resumo[key]['VL_LIQUIDO'])
            v_obj_wks_excel_nr_1.cell(v_nr_linha,4  ,'')
            v_obj_wks_excel_nr_1.cell(v_nr_linha,5  ,'FALSO')
            v_obj_wks_excel_nr_1.cell(v_nr_linha,6  ,ob_global.gv_ob_dict_sped_protocolado_resumo[key]['VL_BASE'])
            v_obj_wks_excel_nr_1.cell(v_nr_linha,7  ,'')
            v_obj_wks_excel_nr_1.cell(v_nr_linha,8  ,'FALSO')
            v_obj_wks_excel_nr_1.cell(v_nr_linha,9  ,ob_global.gv_ob_dict_sped_protocolado_resumo[key]['VL_ICMS'])
            v_obj_wks_excel_nr_1.cell(v_nr_linha,10 ,'')
            v_obj_wks_excel_nr_1.cell(v_nr_linha,11 ,'FALSO')
            v_obj_wks_excel_nr_1.cell(v_nr_linha,12 ,'0-EXISTE PROTOCOLADO e NÃO EXISTE NO ENXERTADO')
            if key in ob_global.gv_ob_dict_sped_enxertado_resumo:
                ob_global.gv_ob_dict_sped_enxertado_resumo[key]['EXISTE'] += 1
                v_obj_wks_excel_nr_1.cell(v_nr_linha,4  ,ob_global.gv_ob_dict_sped_enxertado_resumo[key]['VL_LIQUIDO'])
                if ob_global.gv_ob_dict_sped_enxertado_resumo[key]['VL_LIQUIDO'] == ob_global.gv_ob_dict_sped_enxertado_resumo[key]['VL_LIQUIDO']:
                    v_obj_wks_excel_nr_1.cell(v_nr_linha,5  ,'VERDADEIRO')    
                v_obj_wks_excel_nr_1.cell(v_nr_linha,7  ,ob_global.gv_ob_dict_sped_enxertado_resumo[key]['VL_BASE'])
                if ob_global.gv_ob_dict_sped_enxertado_resumo[key]['VL_BASE'] == ob_global.gv_ob_dict_sped_enxertado_resumo[key]['VL_BASE']:
                    v_obj_wks_excel_nr_1.cell(v_nr_linha,8  ,'VERDADEIRO')    
                v_obj_wks_excel_nr_1.cell(v_nr_linha,10 ,ob_global.gv_ob_dict_sped_enxertado_resumo[key]['VL_ICMS'])
                if ob_global.gv_ob_dict_sped_enxertado_resumo[key]['VL_ICMS'] == ob_global.gv_ob_dict_sped_enxertado_resumo[key]['VL_ICMS']:
                    v_obj_wks_excel_nr_1.cell(v_nr_linha,11  ,'VERDADEIRO')    
                v_obj_wks_excel_nr_1.cell(v_nr_linha,12 ,'1-CFOP+SERIE')
        
        for key in sorted(ob_global.gv_ob_dict_sped_enxertado_resumo.keys()) :        
            if not ob_global.gv_ob_dict_sped_enxertado_resumo[key]['EXISTE']:
                v_nr_linha += 1
                v_cc_chave = key.split("|")
                v_cc_cfop = v_cc_chave[0]
                v_cc_serie = v_cc_chave[1]    
                v_obj_wks_excel_nr_1.cell(v_nr_linha,1  ,v_cc_cfop)
                v_obj_wks_excel_nr_1.cell(v_nr_linha,2  ,v_cc_serie)
                v_obj_wks_excel_nr_1.cell(v_nr_linha,3  ,'')
                v_obj_wks_excel_nr_1.cell(v_nr_linha,4  ,ob_global.gv_ob_dict_sped_enxertado_resumo[key]['VL_LIQUIDO'])
                v_obj_wks_excel_nr_1.cell(v_nr_linha,5  ,'FALSO')
                v_obj_wks_excel_nr_1.cell(v_nr_linha,6  ,'')
                v_obj_wks_excel_nr_1.cell(v_nr_linha,7  ,ob_global.gv_ob_dict_sped_enxertado_resumo[key]['VL_BASE'])
                v_obj_wks_excel_nr_1.cell(v_nr_linha,8  ,'FALSO')
                v_obj_wks_excel_nr_1.cell(v_nr_linha,9  ,'')
                v_obj_wks_excel_nr_1.cell(v_nr_linha,10 ,ob_global.gv_ob_dict_sped_enxertado_resumo[key]['VL_ICMS'])
                v_obj_wks_excel_nr_1.cell(v_nr_linha,11 ,'FALSO')
                v_obj_wks_excel_nr_1.cell(v_nr_linha,12 ,'100-NÃO EXISTE PROTOCOLADO e EXISTE NO ENXERTADO')
        
        v_ob_filter_wks_1 = "A1:" + get_column_letter(v_obj_wks_excel_nr_1.max_column) \
                            + str(v_obj_wks_excel_nr_1.max_row)
        v_obj_wks_excel_nr_1.auto_filter.ref = v_ob_filter_wks_1        


        v_obj_wks_excel_nr_2 = v_obj_arq_excel.create_sheet("ENXERTADO", 2)
        log("Planilha : " + v_obj_wks_excel_nr_2.title)
        v_nr_linha = 0
        for idx, reg in enumerate(ob_global.gv_ob_lista_sped_enxertado):    
            if idx == 0:
                v_nr_linha += 1
                v_obj_wks_excel_nr_2.cell(v_nr_linha,1 ,'TIPO')
                v_obj_wks_excel_nr_2.cell(v_nr_linha,2 ,'CST')
                v_obj_wks_excel_nr_2.cell(v_nr_linha,3 ,'CFOP')
                v_obj_wks_excel_nr_2.cell(v_nr_linha,4 ,'SERIE')
                v_obj_wks_excel_nr_2.cell(v_nr_linha,5 ,'MODELO')
                v_obj_wks_excel_nr_2.cell(v_nr_linha,6 ,'NF_INI')
                v_obj_wks_excel_nr_2.cell(v_nr_linha,7 ,'NF_FIM')
                v_obj_wks_excel_nr_2.cell(v_nr_linha,8 ,'DT_INI')
                v_obj_wks_excel_nr_2.cell(v_nr_linha,9 ,'DT_FIM')
                v_obj_wks_excel_nr_2.cell(v_nr_linha,10,'NM_ARQ')
                v_obj_wks_excel_nr_2.cell(v_nr_linha,11,'VOLUME')                    
                v_obj_wks_excel_nr_2.cell(v_nr_linha,12,'HASHCODE')
                v_obj_wks_excel_nr_2.cell(v_nr_linha,13,'VL_LIQUIDO')
                v_obj_wks_excel_nr_2.cell(v_nr_linha,14,'VL_BASE')
                v_obj_wks_excel_nr_2.cell(v_nr_linha,15,'VL_ICMS')
                for nr_col in range(1,16):
                    v_obj_wks_excel_nr_2.cell(v_nr_linha,nr_col).font=Font(bold=True)
                    v_obj_wks_excel_nr_2.cell(v_nr_linha,nr_col).alignment = Alignment(horizontal='center')
                    v_obj_wks_excel_nr_2.cell(v_nr_linha,nr_col).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")                    
                
            v_nr_linha += 1
            v_obj_wks_excel_nr_2.cell(v_nr_linha,1 ,reg.TIPO)
            v_obj_wks_excel_nr_2.cell(v_nr_linha,2 ,reg.CST)
            v_obj_wks_excel_nr_2.cell(v_nr_linha,3 ,reg.CFOP)
            v_obj_wks_excel_nr_2.cell(v_nr_linha,4 ,reg.SERIE)
            v_obj_wks_excel_nr_2.cell(v_nr_linha,5 ,reg.MODELO)
            v_obj_wks_excel_nr_2.cell(v_nr_linha,6 ,reg.NF_INI)
            v_obj_wks_excel_nr_2.cell(v_nr_linha,7 ,reg.NF_FIM)
            v_obj_wks_excel_nr_2.cell(v_nr_linha,8 ,reg.DT_INI)
            v_obj_wks_excel_nr_2.cell(v_nr_linha,9 ,reg.DT_FIM)
            v_obj_wks_excel_nr_2.cell(v_nr_linha,10,reg.NM_ARQ)
            v_obj_wks_excel_nr_2.cell(v_nr_linha,11,reg.VOLUME)
            v_obj_wks_excel_nr_2.cell(v_nr_linha,12,reg.HASHCODE)
            v_obj_wks_excel_nr_2.cell(v_nr_linha,13,reg.VL_LIQUIDO)
            v_obj_wks_excel_nr_2.cell(v_nr_linha,14,reg.VL_BASE)
            v_obj_wks_excel_nr_2.cell(v_nr_linha,15,reg.VL_ICMS)
            for idx,nr_col in enumerate(range(13,16)):
                v_obj_wks_excel_nr_2.cell(v_nr_linha,nr_col).number_format = "#,##0.00"             

        v_ob_filter_wks_2 = "A1:" + get_column_letter(v_obj_wks_excel_nr_2.max_column) \
                            + str(v_obj_wks_excel_nr_2.max_row)
        v_obj_wks_excel_nr_2.auto_filter.ref = v_ob_filter_wks_2        

        v_obj_wks_excel_nr_3 = v_obj_arq_excel.create_sheet("PROTOCOLADO", 3)  
        log("Planilha : " + v_obj_wks_excel_nr_3.title)    
        v_nr_linha = 0
        for idx, reg in enumerate(ob_global.gv_ob_lista_sped_protocolado):    
            if idx == 0:
                v_nr_linha += 1
                v_obj_wks_excel_nr_3.cell(v_nr_linha,1 ,'TIPO')
                v_obj_wks_excel_nr_3.cell(v_nr_linha,2 ,'CST')
                v_obj_wks_excel_nr_3.cell(v_nr_linha,3 ,'CFOP')
                v_obj_wks_excel_nr_3.cell(v_nr_linha,4 ,'SERIE')
                v_obj_wks_excel_nr_3.cell(v_nr_linha,5 ,'MODELO')
                v_obj_wks_excel_nr_3.cell(v_nr_linha,6 ,'NF_INI')
                v_obj_wks_excel_nr_3.cell(v_nr_linha,7 ,'NF_FIM')
                v_obj_wks_excel_nr_3.cell(v_nr_linha,8 ,'DT_INI')
                v_obj_wks_excel_nr_3.cell(v_nr_linha,9 ,'DT_FIM')
                v_obj_wks_excel_nr_3.cell(v_nr_linha,10,'NM_ARQ')
                v_obj_wks_excel_nr_3.cell(v_nr_linha,11,'VOLUME')                    
                v_obj_wks_excel_nr_3.cell(v_nr_linha,12,'HASHCODE')
                v_obj_wks_excel_nr_3.cell(v_nr_linha,13,'VL_LIQUIDO')
                v_obj_wks_excel_nr_3.cell(v_nr_linha,14,'VL_BASE')
                v_obj_wks_excel_nr_3.cell(v_nr_linha,15,'VL_ICMS')
                for nr_col in range(1,16):
                    v_obj_wks_excel_nr_3.cell(v_nr_linha,nr_col).font=Font(bold=True)
                    v_obj_wks_excel_nr_3.cell(v_nr_linha,nr_col).alignment = Alignment(horizontal='center')
                    v_obj_wks_excel_nr_3.cell(v_nr_linha,nr_col).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")                   
                
            v_nr_linha += 1
            v_obj_wks_excel_nr_3.cell(v_nr_linha,1 ,reg.TIPO)
            v_obj_wks_excel_nr_3.cell(v_nr_linha,2 ,reg.CST)
            v_obj_wks_excel_nr_3.cell(v_nr_linha,3 ,reg.CFOP)
            v_obj_wks_excel_nr_3.cell(v_nr_linha,4 ,reg.SERIE)
            v_obj_wks_excel_nr_3.cell(v_nr_linha,5 ,reg.MODELO)
            v_obj_wks_excel_nr_3.cell(v_nr_linha,6 ,reg.NF_INI)
            v_obj_wks_excel_nr_3.cell(v_nr_linha,7 ,reg.NF_FIM)
            v_obj_wks_excel_nr_3.cell(v_nr_linha,8 ,reg.DT_INI)
            v_obj_wks_excel_nr_3.cell(v_nr_linha,9 ,reg.DT_FIM)
            v_obj_wks_excel_nr_3.cell(v_nr_linha,10,reg.NM_ARQ)
            v_obj_wks_excel_nr_3.cell(v_nr_linha,11,reg.VOLUME)
            v_obj_wks_excel_nr_3.cell(v_nr_linha,12,reg.HASHCODE)
            v_obj_wks_excel_nr_3.cell(v_nr_linha,13,reg.VL_LIQUIDO)
            v_obj_wks_excel_nr_3.cell(v_nr_linha,14,reg.VL_BASE)
            v_obj_wks_excel_nr_3.cell(v_nr_linha,15,reg.VL_ICMS)
            for idx,nr_col in enumerate(range(13,16)):
                v_obj_wks_excel_nr_3.cell(v_nr_linha,nr_col).number_format = "#,##0.00"             

        v_ob_filter_wks_3 = "A1:" + get_column_letter(v_obj_wks_excel_nr_3.max_column) \
                            + str(v_obj_wks_excel_nr_3.max_row)
        v_obj_wks_excel_nr_3.auto_filter.ref = v_ob_filter_wks_3             
        
        log("Redimensiona a planilha Excel : " + v_obj_wks_excel_nr_0.title)
        fnc_column_dimensions(v_obj_wks_excel_nr_0) 
        log("Redimensiona a planilha Excel : " + v_obj_wks_excel_nr_1.title)
        fnc_column_dimensions(v_obj_wks_excel_nr_1) 
        log("Redimensiona a planilha Excel : " + v_obj_wks_excel_nr_2.title)
        fnc_column_dimensions(v_obj_wks_excel_nr_2)   
        log("Redimensiona a planilha Excel : " + v_obj_wks_excel_nr_3.title)    
        fnc_column_dimensions(v_obj_wks_excel_nr_3)        
        try:
            log("freeze_panes a planilha Excel : " + v_obj_wks_excel_nr_0.title)
            v_obj_wks_excel_nr_0.freeze_panes = 'A2'
            log("freeze_panes a planilha Excel : " + v_obj_wks_excel_nr_1.title)
            v_obj_wks_excel_nr_1.freeze_panes = 'A2'
            log("freeze_panes a planilha Excel : " + v_obj_wks_excel_nr_2.title)
            v_obj_wks_excel_nr_2.freeze_panes = 'A2'
            log("freeze_panes a planilha Excel : " + v_obj_wks_excel_nr_3.title)
            v_obj_wks_excel_nr_3.freeze_panes = 'A2'
        except:
            pass
            
        # Grava a planilha Excel
        log("Grava a planilha Excel : " + ob_global.gv_lst_arq_insumo_sped)
        v_obj_arq_excel.save(ob_global.gv_lst_arq_insumo_sped)
        log(str(v_nr_retorno) + " Salvo a planilha Excel : " + ob_global.gv_lst_arq_insumo_sped)
        
        return v_nr_retorno
        
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO PROCESSAMENTO: " + str(e)+ " >> " + v_ds_trace)
        v_nr_retorno = 93
        return v_nr_retorno        


if __name__ == "__main__" :
    """
        Ponto de partida
    """    
    # Codigo de Retorno
    v_nr_ret = 0

    # Tratamento de excessao
    v_ds_trace = ''

    # Tratamento de variaveis globais
    ob_global.gv_ob_conexao = None
    # Parametros do arquivo de configuração
    ob_global.gv_nm_job = "" 

    # Lista de String
    ob_global.gv_ob_lista_string = list(string.ascii_lowercase)

    try:

        log("-"*100)
        log("INICIO DA EXECUÇÃO".center(120,'#'))
       
        # Validacao dos parametros de entrada
        if not v_nr_ret:
            v_nr_ret = fnc_validar_entrada()
        
        # Verificar conexao com o banco
        if not v_nr_ret:
            v_nr_ret = fnc_conectar_banco_dados()   

        # Validacao dos parametros de configuracoes
        if not v_nr_ret:
            v_nr_ret = fnc_validar_configuracao()

        # Validacao da leitura e carregamento do arquivo
        if not v_nr_ret:
            fnc_imprimi_variaveis()
            v_nr_ret = fnc_validar_leitura_carregamento_arquivo()
            
        # Processar         
        if not v_nr_ret:            
            v_nr_ret = fnc_processar()                    
        
        log("-"*100)
        # Finalizacao
        if not v_nr_ret:
            log("SUCESSO NA EXECUÇÃO")
        else:
            log("ERRO NA EXECUÇÃO")
    
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO .: " + str(e) + " >> " + v_ds_trace)
        v_nr_ret = 93
    
    sys.exit(v_nr_ret if v_nr_ret >= log.ret else log.ret )