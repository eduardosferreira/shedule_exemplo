#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-
"""
----------------------------------------------------------------------------------------------
  MODULO ...: TESHUVA
  SCRIPT ...: loader_DUB.py
  CRIACAO ..: 16/07/2021
  AUTOR ....: VICTOR SANTOS / KYROS TECNOLOGIA
  DESCRICAO : 

----------------------------------------------------------------------------------------------
  HISTORICO : 
----------------------------------------------------------------------------------------------
"""

import os
import sys

global SD, dir_base
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)
import configuracoes

import datetime
import cx_Oracle
import string
import shutil
from pathlib import Path
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

import comum
import sql
import layout
import util

log.gerar_log_em_arquivo = True
comum.carregaConfiguracoes(configuracoes)

def dtf():
    return (datetime.datetime.now().strftime('%Y%m%d%H%M%S'))

def lerXlsx(diretorio, arquivo) :
    print('-'*100)
    ret = True
    path_arq = os.path.join( diretorio, arquivo )
    print('Processando o arquivo:', path_arq)
    
    xls = load_workbook(path_arq)
    p1 = xls.sheetnames[0]
    print("Planilhas do arquivo: ",p1)
    sh = xls[p1]
    ncolunas = sh.max_column
    
    planilha = xls.active
    dados = []

    if (ncolunas != 2):
        log("ERRO - Formatação da planilha, arquivo não está no padrão correto.")
        return False

    r = 0
    for row in planilha.rows :
        dados.append([])
        for col in row :
            dados[r].append(col.value)
        r += 1 
    reg = 0
    
    cmd_sql1="""DELETE from gfcadastro.tb_clientes_consulados"""

    try : 
        con.executa(cmd_sql1)
    except Exception as e :
        print(cmd_sql1)
        print(str(e))      
  
#### Insere os novos registros    
    
    for row in dados:

        for c in string.punctuation:
            dados[reg][1] = dados[reg][1].replace(c, "")

        cmd_sql2=""" 
            insert into gfcadastro.tb_clientes_consulados(
                            RAZAO_SOCIAL,
                            CNPJ)
            values ('%s','%s') 
            """%(dados[reg][0],dados[reg][1]) 
    
        reg = reg + 1
        
        try : 
            con.executa(cmd_sql2)
        except Exception as e :
            print('Erro ao executar o Sql ...')
            print(str(e))
            print(cmd_sql2)
            ret = False
            break
        con.commit()
    print("Foram incluidos", reg, "novos registros")        
    con.commit()
    
    return ret

def processar() :

    print ('')
    print("Verificando arquivos para processar no diretorio de entrada ...")
    print ('')
    print('- Diretorio : %s'%(configuracoes.dir_arquivos))
    
    if not os.path.isdir(configuracoes.dir_arquivos) :
        os.makedirs(configuracoes.dir_arquivos)
    
    if not os.path.isdir(configuracoes.diretorio_processados) :
        os.makedirs(configuracoes.diretorio_processados)
    
    qtd_arqs = 0
    ret = 99
    for arq in os.listdir(configuracoes.dir_arquivos) :
        ret = 0
        if os.path.isfile( os.path.join(configuracoes.dir_arquivos, arq)) :
            if arq.endswith('.xlsx') :
                print('-'*100)
                qtd_arqs += 1
                print('Verificando o arquivo :',arq)
                
                if not lerXlsx(configuracoes.dir_arquivos, arq) :
                    ret = 99
                else :
                    vdataArq= dtf()
                    os.rename(os.path.join(configuracoes.dir_arquivos, arq),os.path.join(configuracoes.dir_arquivos, arq)+vdataArq)
                    shutil.move(os.path.join(configuracoes.dir_arquivos, arq)+vdataArq, configuracoes.diretorio_processados)
 
    if (qtd_arqs == 0) :
        print("ERRO - Nao foi encontrado nenhum arquivo xlxs a ser processado.")
        ret = 99

    print('-'*100)

    return ret

if __name__ == "__main__" :

    con=sql.geraCnxBD(configuracoes)
    
    ret = 0
    txt = ''
    ret = processar() 
    print("Codigo de retorno =", ret)
    if (ret != 0 ):
        print('ERRO no processamento.')
        ret = 92

