#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-

"""
----------------------------------------------------------------------------------------------
MODULO ...: TESHUVA
SCRIPT ...: relatorio_conciliacao.py
CRIACAO ..: 09/09/2021
AUTOR ....: Victor Santos / KYROS TECNOLOGIA
DESCRICAO.: Insumos - Relatorio de Conciliacao
----------------------------------------------------------------------------------------------
PARAMETROS: 

----------------------------------------------------------------------------------------------
    HISTORICO : 
        * 09/09/2021 - Victor Santos / KYROS TECNOLOGIA
        - Criacao do script.    
        * 24/09/2021 - EDUARDO DA SILVA FERREIRA / KYROS TECNOLOGIA (eduardof@kyros.com.br)
        PTITES-866 - ET - Relatório Conciliação Série - Inclusão Totalizadores - Documento Técnico e cenários de teste de produto
        - Colocado a serie no nome do arquivo;
        - Adição de dados no item "MESTRE DE NOTA FISCAL" : BASE_ICMS	VALOR_ICMS	ISENTAS_ICMS	OUTROS_VALORES	DESCONTO	VALOR_CONTABIL
     * 13/01/2022 - Eduardo da Silva Ferreira - Kyros Tecnologia
        PTITES-1367 : Aonde que faz referencia na leitura das informações do diretório do protocolado, 
        verificar se a ultima entrega de acordo com a informação 
        "ORIGEM_PROTOCOLADO" de acordo com a tabela "TSH_SERIE_LEVANTAMENTO"
        https://jira.telefonica.com.br/browse/PTITES-1367
        https://wikicorp.telefonica.com.br/x/JKMPDQ
    
    * 20/01/2022 - Victor Santos - Kyros Consultoria
        ALT006
            Documentação:
                1) Atualizar aba "Batimento"
                    Onde encontra-se "Regerado" no excel deve ser ajustado para "ATUAL_TI" tabelas 1,2 e 3.
                2) Atualizar aba "Checklist_Cadastro"
                    2.1) Alterar cabeçalho "Regerado" para "Atual TI" 
                    2.2) Alterar "Regerado" para "Atual TI" tabela 2.
    
----------------------------------------------------------------------------------------------
"""

import sys
import os
SD = '/' if os.name == 'posix' else '\\'
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV')[0], 'DEV')
sys.path.append(dir_base)
import configuracoes
import comum
import sql
import layout
import util
import vie 

comum.carregaConfiguracoes(configuracoes)
layout.carregaLayout()

import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
import operator
import glob

global strglobal

strglobal = "_OK.xlsx"

log.gerar_log_em_arquivo = True

dic_registros = {}
dic_layouts = {}
dic_campos = {}
dic_enderecos_uf = {}
variaveis = {}
dic_fd = {}
ano = 0 
serie = ""
ano_mes = ""
uf = ""
abaRS = []
abaBT = []
abaBQ = []
abaCC = []
config = "" 
id_serie = ".xlsx"
seriesRelatorio = {}

global dadosSerie #PTITES-1367
dadosSerie = {} #

def processar():
    global dadosSerie #PTITES-1367
    global config   
    global id_serie 
    global ano 
    global serie 
    global ano_mes
    global uf
    global strglobal
    global ERRO

    if len(sys.argv) < 2:
        log('#### ERRO ') 
        log('-'* 100)
        log('QUANTIDADE DE PARAMETROS INVALIDA')
        log('-'* 100)
        log('EXEMPLO')
        log('-'* 100)
        log( '%s <ID SERIE LEVANTAMENTO>'%( sys.argv[0] ) )
        log('-'* 100)
        return(99)
    else:
        id_serie = sys.argv[1]
    
        dadosSerie = comum.buscaDadosSerie(id_serie)   
        uf               = dadosSerie['uf']
        mes              = dadosSerie['mes']
        ano              = dadosSerie['ano'][2:]
        ano_mes          = ano + mes
        ano_aaaa         = dadosSerie['ano']
        serie            = dadosSerie['serie']
        diretorio        = dadosSerie['dir_serie']
        # -- Inicio PTITES-1367
        configuracoes.dir_leitura_default = "/portaloptrib/LEVCV115" 
        dir_leitura     = os.path.join( getattr(configuracoes, 'dir_leitura', configuracoes.dir_leitura_default), dadosSerie['sub_dir_serie'] )
        # -- Fim PTITES-1367
        ano              = int(ano)
        
        diretorioATUAL_TI    =  dir_leitura + "/OBRIGACAO"
        diretorioProtocolado =  dir_leitura + "/PROTOCOLADO"    
        diretorioInsumo      =  dir_leitura + "/INSUMOS" #PTITES-1367
        # -- Inicio PTITES-1367
        configuracoes.origem_protocolado  = "PROTOCOLADO"
        if dadosSerie.get("origem_protocolado"):
            if str(dadosSerie.get("origem_protocolado")).strip().upper()=="ULTIMO_ENTREGUE":    
                diretorioProtocolado =  dir_leitura + "/ULTIMA_ENTREGA" 
                configuracoes.origem_protocolado  = "ULTIMO_PROTOCOLADO"
        # -- Fim PTITES-1367
        

        log("-------------------------------------------------------------------------")    
        log("- Iniciando geracao do Relatorio de conciliacao...")    
        log(" ")    
        log("- diretorioATUAL_TI    = ", diretorioATUAL_TI)    
        log("- diretorioProtocolado = ", diretorioProtocolado)        
        log("-------------------------------------------------------------------------")    

        log("-------------------------------------------------------------------------")
        log("- Processando ATUAL_TI... ")

        if  not os.path.exists(diretorioATUAL_TI):
            log("! Diretorio ATUAL_TI Inexistente !")
        else:
            sumarizaValores(diretorioATUAL_TI, 'ATUAL_TI')

        log("- Concluido processamento do ATUAL_TI... ")
        log("-------------------------------------------------------------------------")

        log("-------------------------------------------------------------------------")
        log("- Processando Protocolado... ")
            
        if  not os.path.exists(diretorioProtocolado):
            log("! Diretorio Protocolado Inexistente !")
        else:
            sumarizaValores(diretorioProtocolado, configuracoes.origem_protocolado) # PTITES-1367

        log("- Concluido processamento do Protocolado... ")
        log("-------------------------------------------------------------------------")
        
        
        abaBT.sort(key=operator.attrgetter('AREA', 'EMPS_COD', 'FILI_COD', 'MES_ANO', 'SERIE', 'CFOP'))
        abaBQ.sort(key=operator.attrgetter('EMPS_COD', 'FILI_COD', 'MES_ANO', 'SERIE', 'AREA'))
        abaCC.sort(key=operator.attrgetter('EMPS_COD', 'FILI_COD', 'MES_ANO', 'SERIE', 'AREA'))
        
        log("-------------------------------------------------------------------------")
        log("- Processando Gravacao da Planilha... ")

        wb = openpyxl.Workbook()
        
        log("- Preenchendo a planilha...")
        ERRO = carregaDadosPlanliha(wb)
        
        nomePlanilha = 'Conciliacao_Serie'+ "_" + serie + strglobal # PTITES-866
        caminhoPlanilha = os.path.join( diretorioInsumo, nomePlanilha )
        log("- Gravando a planilha...")
        log("- Caminho da planilha = ", caminhoPlanilha )
        
        if not os.path.isdir(diretorioInsumo):
            os.makedirs(diretorioInsumo)
        
        wb.save(caminhoPlanilha)
            
        log("- Concluido Gravacao da Planilha... ")
        log("-------------------------------------------------------------------------")
        log("-")
        log("-------------------------------------------------------------------------")
        log("- Processamento Finalizado !")
        log("-------------------------------------------------------------------------")
        
        return caminhoPlanilha,ERRO

class resumo(object):
    def __init__(self):
        self.EMPS_COD           = ''
        self.FILI_COD           = ''
        self.MES_ANO            = ''
        self.SERIE              = ''
        self.CFOP               = ''
        self.RESUMO             = ''

class LinhaRelatorio:
    valorTotal = baseICMS = valorICMS = isentasICMS = outrosValores = valorContabil = None
    desconto = None
    def __init__(self, valorTotal, baseICMS, valorICMS, isentasICMS, outrosValores, desconto, valorContabil):
        self.valorTotal = valorTotal
        self.baseICMS = baseICMS
        self.valorICMS = valorICMS
        self.isentasICMS = isentasICMS
        self.outrosValores = outrosValores
        self.desconto = desconto
        self.valorContabil = valorContabil

class Serie:
    emps_cod = fili_cod = mes_ano = serie = cfop = None
    areas = {}

    def __init__(self, emps_cod, fili_cod, mes_ano, serie, cfop):
        self.emps_cod = emps_cod
        self.fili_cod = fili_cod
        self.mes_ano  = mes_ano
        self.serie    = serie
        self.cfop     = cfop
        self.areas = {}
        
class batimento(object):
    def __init__(self):
        self.EMPS_COD           = ''
        self.FILI_COD           = ''
        self.MES_ANO            = ''
        self.SERIE              = ''
        self.CFOP               = ''
        self.VALOR_TOTAL        = 0  
        self.BASE_ICMS          = 0  
        self.VALOR_ICMS         = 0  
        self.ISENTAS_ICMS       = 0 
        self.OUTROS_VALORES     = 0  
        self.DESCONTO           = 0  
        self.VALOR_CONTABIL     = 0
        self.AREA               = ''
        self.VALIDACAO          = ''

class batimentoQuantidade(object):
    def __init__(self):
        self.EMPS_COD           = ''
        self.FILI_COD           = ''
        self.MES_ANO            = ''
        self.SERIE              = ''
        self.AREA               = ''
        self.QTD_NFS            = 0
        self.VALOR_TOTAL_MESTRE = 0
        self.BASE_ICMS_MESTRE          = 0  # PTITES-866
        self.VALOR_ICMS_MESTRE         = 0  # PTITES-866
        self.ISENTAS_ICMS_MESTRE       = 0  # PTITES-866
        self.OUTROS_VALORES_MESTRE     = 0  # PTITES-866

class checkListCadastro(object):
    def __init__(self):
        self.EMPS_COD                   = ''
        self.FILI_COD                   = ''
        self.MES_ANO                    = ''
        self.SERIE                      = ''
        self.AREA                       = ''
        self.QTD_NFS                    = 0
        self.QTD_NFS_TOTAL              = 0

        self.TOTAL_NOME_CLIENTE         = 0
        self.TOTAL_NOME_CLIENTE_INV     = 0
        self.TOTAL_NOME_FANTASIA        = 0
        self.TOTAL_CPF_CNPJ             = 0
        self.TOTAL_CPF_CNPJ_INV         = 0
        self.TOTAL_TERMINAL             = 0
        self.TOTAL_TERMINAL_INV         = 0
        self.TOTAL_TEL_CONTATO          = 0
        self.TOTAL_TEL_CONTATO_INV      = 0
        self.TOTAL_ENDERECO_CLIENTE     = 0
        self.TOTAL_ENDERECO_CLIENTE_INV = 0
        self.TOTAL_INSCRICAO_INV        = 0
        
        ## Coluna : VALOR_TOTAL_NF
        self.TOTAL_NOME_CLIENTE_VALOR_TOTAL_NF         = 0
        self.TOTAL_NOME_CLIENTE_INV_VALOR_TOTAL_NF     = 0
        self.TOTAL_NOME_FANTASIA_VALOR_TOTAL_NF        = 0
        self.TOTAL_CPF_CNPJ_VALOR_TOTAL_NF             = 0
        self.TOTAL_CPF_CNPJ_INV_VALOR_TOTAL_NF         = 0
        self.TOTAL_TERMINAL_VALOR_TOTAL_NF             = 0
        self.TOTAL_TERMINAL_INV_VALOR_TOTAL_NF         = 0
        self.TOTAL_TEL_CONTATO_VALOR_TOTAL_NF          = 0
        self.TOTAL_TEL_CONTATO_INV_VALOR_TOTAL_NF      = 0
        self.TOTAL_ENDERECO_CLIENTE_VALOR_TOTAL_NF     = 0
        self.TOTAL_ENDERECO_CLIENTE_INV_VALOR_TOTAL_NF = 0
        self.TOTAL_INSCRICAO_INV_VALOR_TOTAL_NF        = 0
        
        ## Coluna : BASE_ICMS
        self.TOTAL_NOME_CLIENTE_BASE_ICMS         = 0
        self.TOTAL_NOME_CLIENTE_INV_BASE_ICMS     = 0
        self.TOTAL_NOME_FANTASIA_BASE_ICMS        = 0
        self.TOTAL_CPF_CNPJ_BASE_ICMS             = 0
        self.TOTAL_CPF_CNPJ_INV_BASE_ICMS         = 0
        self.TOTAL_TERMINAL_BASE_ICMS             = 0
        self.TOTAL_TERMINAL_INV_BASE_ICMS         = 0
        self.TOTAL_TEL_CONTATO_BASE_ICMS          = 0
        self.TOTAL_TEL_CONTATO_INV_BASE_ICMS      = 0
        self.TOTAL_ENDERECO_CLIENTE_BASE_ICMS     = 0
        self.TOTAL_ENDERECO_CLIENTE_INV_BASE_ICMS = 0
        self.TOTAL_INSCRICAO_INV_BASE_ICMS        = 0
        
        ## Coluna : VALOR_ICMS
        self.TOTAL_NOME_CLIENTE_VALOR_ICMS         = 0
        self.TOTAL_NOME_CLIENTE_INV_VALOR_ICMS     = 0
        self.TOTAL_NOME_FANTASIA_VALOR_ICMS        = 0
        self.TOTAL_CPF_CNPJ_VALOR_ICMS             = 0
        self.TOTAL_CPF_CNPJ_INV_VALOR_ICMS         = 0
        self.TOTAL_TERMINAL_VALOR_ICMS             = 0
        self.TOTAL_TERMINAL_INV_VALOR_ICMS         = 0
        self.TOTAL_TEL_CONTATO_VALOR_ICMS          = 0
        self.TOTAL_TEL_CONTATO_INV_VALOR_ICMS      = 0
        self.TOTAL_ENDERECO_CLIENTE_VALOR_ICMS     = 0
        self.TOTAL_ENDERECO_CLIENTE_INV_VALOR_ICMS = 0
        self.TOTAL_INSCRICAO_INV_VALOR_ICMS        = 0

def buscaFilial(Empresa_Emitente\
                , Cnpj_Emitente\
                , Arquivo_uf\
                , nm_arquivo=""): #PTITES-1367
    
    cursor = sql.geraCnxBD(configuracoes)

    SQL = """ SELECT F.FILI_COD       FIILAL   
                FROM OPENRISOW.FILIAL F        
               WHERE F.EMPS_COD     = '%s' 
                 AND F.FILI_COD_CGC = '%s'    
                 AND F.UNFE_SIG     = '%s'      
    """%(Empresa_Emitente, Cnpj_Emitente, Arquivo_uf)
    cursor.executa(SQL)
    linha = cursor.fetchone()

    # PTITES-1367
    try: 
        v_cd_erro = 0
        if not linha:
            v_cd_erro = 1
        else:
            if not linha[0]:
                v_cd_erro = 1                    
        if v_cd_erro:
            configuracoes.numero_erro += 1 # PTITES-1367
            configuracoes.descricao_erro += 'ERRO >> Nao encontrado cadastro para a serie na tabela de filiais ( OPENRISOW.FILIAL) '\
             + " - Empresa >>" + Empresa_Emitente\
             + " - Cnpj_Emitente >> " + Cnpj_Emitente\
             + " - Arquivo_uf >> " + Arquivo_uf\
             + " - Arquivo >> " + nm_arquivo + " >> " # PTITES-1367        
            log('ERRO >> Nao encontrado cadastro para a serie na tabela de filiais ( OPENRISOW.FILIAL) '\
             + " - Empresa >>" + Empresa_Emitente\
             + " - Cnpj_Emitente >> " + Cnpj_Emitente\
             + " - Arquivo_uf >> " + Arquivo_uf\
             + " - Arquivo >> " + nm_arquivo + " >> ") # PTITES-1367         # raise Exception('Nao encontrado cadastro para a serie na tabela de filiais ( OPENRISOW.FILIAL) ')  
            # raise Exception('Nao encontrado cadastro para a serie na tabela de filiais ( OPENRISOW.FILIAL) ')
            return None            
        return(linha[0])
    except:
        configuracoes.numero_erro += 1 # PTITES-1367
        configuracoes.descricao_erro += 'ERRO >> Nao encontrado cadastro para a serie na tabela de filiais ( OPENRISOW.FILIAL) '\
             + " - Empresa >>" + Empresa_Emitente\
             + " - Cnpj_Emitente >> " + Cnpj_Emitente\
             + " - Arquivo_uf >> " + Arquivo_uf\
             + " - Arquivo >> " + nm_arquivo + " >> " # PTITES-1367        
        log('ERRO >> Nao encontrado cadastro para a serie na tabela de filiais ( OPENRISOW.FILIAL) '\
             + " - Empresa >>" + Empresa_Emitente\
             + " - Cnpj_Emitente >> " + Cnpj_Emitente\
             + " - Arquivo_uf >> " + Arquivo_uf\
             + " - Arquivo >> " + nm_arquivo + " >> ") # PTITES-1367         # raise Exception('Nao encontrado cadastro para a serie na tabela de filiais ( OPENRISOW.FILIAL) ')  
        return None
    # PTITES-1367

def retornaEnderecoVivo( uf ) :
    if not dic_enderecos_uf.get(uf, False) :
        cursor = sql.geraCnxBD(configuracoes)
        SQL = """ SELECT DISTINCT END.ENDERECO, END.NUMERO
                    FROM GFCADASTRO.GF_CAD_RED_SANEADA1_END_VIVO END
                   WHERE TRIM(END.UF) = REPLACE(TRIM('%s'),'0','O')
        """%(uf)      
        cursor.executa(SQL)
        linha = cursor.fetchone()
        dic_enderecos_uf[uf] = [ x for x in linha ]
    endereco = dic_enderecos_uf.get(uf, '')
    return endereco

def sumarizaValores(diretorio, area):
    global abaRS
    global abaBT
    global abaBQ
    global abaCC
    global strglobal
    global dadosSerie #PTITES-1367
    separadorDiretorio = ('/' if os.name == 'posix' else '\\')
    
    listaArquivos = [f for f in glob.glob(diretorio + "/*.*", recursive=False)]    
    listaArquivos.sort(reverse=True)
    total_notas = {}
    dic_campos = {}
    for arq in listaArquivos:
        nome_arquivo = arq.split(separadorDiretorio)[-1]
        vol = nome_arquivo[-3:]

        if  os.path.isfile(arq):
            # INICIO #PTITES-1367
            nome_arquivo = str(nome_arquivo).strip()
            if not nome_arquivo\
            or not nome_arquivo.upper().startswith(dadosSerie.get('uf'))\
            or not str(nome_arquivo[-3:]).isnumeric()\
            or nome_arquivo[-4] != ".":
                log('Ignorando arquivo com nome fora do padrao', nome_arquivo + " ...")
                continue
            # FIM #PTITES-1367

            if  ano <= 16:
                Arquivo_serie        = nome_arquivo[2:5].strip()
                Arquivo_ano_mes      = nome_arquivo[5:9]
                Arquivo_uf           = nome_arquivo[0:2]
                Arquivo_tipo         = nome_arquivo[10]
                LayoutMestre         = 'LayoutMestre_Antigo'
                LayoutItem           = 'LayoutItem_Antigo'
                LayoutCadastro       = 'LayoutCadastro_Antigo'
                Empresa_Emitente     = 'TBRA'
                Cnpj_Emitente        = '0'
                Filial_Emitente      = '0'
            else:
                if len(nome_arquivo) < 28 :
                    log('Ignorando arquivo com nome fora do padrao', nome_arquivo + " ...")
                    continue

                Arquivo_serie        = nome_arquivo[18:21].strip()
                Arquivo_ano_mes      = nome_arquivo[21:25]
                Arquivo_uf           = nome_arquivo[0:2]
                Arquivo_tipo         = nome_arquivo[28]
                LayoutMestre         = 'LayoutMestre'
                LayoutItem           = 'LayoutItem'
                LayoutCadastro       = 'LayoutCadastro'
                Empresa_Emitente     = 'TBRA'
                Cnpj_Emitente        = nome_arquivo[2:16]
                Filial_Emitente      = buscaFilial(Empresa_Emitente, Cnpj_Emitente, Arquivo_uf,arq)  
                if not Filial_Emitente:
                    log("arquivo: " + nome_arquivo)
                    continue
            
            dic_campos[LayoutMestre]   = dic_campos.get(LayoutMestre, layout.carregaLayout.dic_layouts[LayoutMestre]['dic_campos'])
            dic_campos[LayoutItem]     = dic_campos.get(LayoutItem, layout.carregaLayout.dic_layouts[LayoutItem]['dic_campos'])
            dic_campos[LayoutCadastro] = dic_campos.get(LayoutCadastro, layout.carregaLayout.dic_layouts[LayoutCadastro]['dic_campos'])
            if (Arquivo_serie == serie and Arquivo_ano_mes == ano_mes and Arquivo_uf == uf):
                if  Arquivo_tipo == 'I':
                    log('Arquivo Item     Em processamento =', nome_arquivo + " ...")
                    arquivo = os.path.join( diretorio, nome_arquivo )
                    encoding = comum.encodingDoArquivo( arquivo )
                    fd = open(arquivo, 'r', encoding=encoding)
                    
                    for registro in fd:
                        registroItem = layout.quebraRegistro(registro, LayoutItem)                        
                        it                    = batimento()                        
                        it.EMPS_COD           = Empresa_Emitente
                        it.FILI_COD           = Filial_Emitente
                        it.MES_ANO            = Arquivo_ano_mes[2:] + "/20" + Arquivo_ano_mes[:2]
                        it.SERIE              = Arquivo_serie
                        it.AREA               = area                        

                        it.CFOP               = registroItem[dic_campos[LayoutItem]['CFOP']-1]
                        it.VALOR_TOTAL        = int(registroItem[dic_campos[LayoutItem]['VALOR_TOTAL']-1])    
                        it.BASE_ICMS          = int(registroItem[dic_campos[LayoutItem]['BASE_ICMS']-1])     
                        it.VALOR_ICMS         = int(registroItem[dic_campos[LayoutItem]['VALOR_ICMS']-1])    
                        it.ISENTAS_ICMS       = int(registroItem[dic_campos[LayoutItem]['ISENTAS_ICMS']-1])  
                        it.OUTROS_VALORES     = int(registroItem[dic_campos[LayoutItem]['OUTROS_VALORES']-1])
                        it.DESCONTO           = int(registroItem[dic_campos[LayoutItem]['DESCONTO']-1])      
                        it.VALOR_CONTABIL     = (int(registroItem[dic_campos[LayoutItem]['VALOR_TOTAL']-1]) - int(registroItem[dic_campos[LayoutItem]['DESCONTO']-1]))
                        
                        statusDocumento = registroItem[dic_campos[LayoutItem]['SIT_DOC']-1]
                        
                        if  statusDocumento == 'N':
                            achou = 'N'
                            
                            for ix, bt in enumerate(abaBT):                            
                                if (it.FILI_COD          == bt.FILI_COD
                                and it.MES_ANO           == bt.MES_ANO
                                and it.SERIE             == bt.SERIE  
                                and it.AREA              == bt.AREA                   
                                and it.CFOP              == bt.CFOP):                   
                                    achou = 'S'
                                    abaBT[ix].VALOR_TOTAL        = abaBT[ix].VALOR_TOTAL + it.VALOR_TOTAL
                                    abaBT[ix].BASE_ICMS          = abaBT[ix].BASE_ICMS + it.BASE_ICMS
                                    abaBT[ix].VALOR_ICMS         = abaBT[ix].VALOR_ICMS + it.VALOR_ICMS
                                    abaBT[ix].ISENTAS_ICMS       = abaBT[ix].ISENTAS_ICMS + it.ISENTAS_ICMS
                                    abaBT[ix].OUTROS_VALORES     = abaBT[ix].OUTROS_VALORES + it.OUTROS_VALORES
                                    abaBT[ix].DESCONTO           = abaBT[ix].DESCONTO + it.DESCONTO
                                    abaBT[ix].VALOR_CONTABIL     = abaBT[ix].VALOR_CONTABIL + it.VALOR_CONTABIL
                                
                            if  achou == 'N':                           
                                abaBT.append(it)
                        
                if  Arquivo_tipo == 'M':
                    log('Arquivo Mestre   Em processamento =', nome_arquivo + " ...")
                    arquivo = os.path.join( diretorio, nome_arquivo )
                    encoding = comum.encodingDoArquivo( arquivo )
                    fd = open(arquivo, 'r', encoding=encoding)
                    xlinha = 0
                    for registro in fd:
                        xlinha += 1
                        registroMestre = layout.quebraRegistro(registro, LayoutMestre)                        
                        nf                    = batimentoQuantidade()                        
                        nf.EMPS_COD           = Empresa_Emitente
                        nf.FILI_COD           = Filial_Emitente
                        nf.MES_ANO            = Arquivo_ano_mes[2:] + "/20" + Arquivo_ano_mes[:2]
                        nf.SERIE              = Arquivo_serie
                        nf.AREA               = area
                        nf.QTD_NFS            = 1
                        
                        nf.VALOR_TOTAL_MESTRE = int(registroMestre[dic_campos[LayoutMestre]['VALOR_TOTAL']-1])
                        nf.VALOR_TOTAL        = int(registroMestre[dic_campos[LayoutMestre]['VALOR_TOTAL']-1])
                        nf.BASE_ICMS          = int(registroMestre[dic_campos[LayoutMestre]['BASE_ICMS']-1])
                        nf.VALOR_ICMS         = int(registroMestre[dic_campos[LayoutMestre]['VALOR_ICMS']-1])
                        nf.BASE_ICMS_MESTRE          = int(registroMestre[dic_campos[LayoutMestre]['BASE_ICMS']-1])  # PTITES-866
                        nf.VALOR_ICMS_MESTRE         = int(registroMestre[dic_campos[LayoutMestre]['VALOR_ICMS']-1])  # PTITES-866
                        nf.ISENTAS_ICMS_MESTRE       = int(registroMestre[dic_campos[LayoutMestre]['ISENTAS_ICMS']-1])  # PTITES-866
                        nf.OUTROS_VALORES_MESTRE     = int(registroMestre[dic_campos[LayoutMestre]['OUTROS_VALORES']-1])  # PTITES-866

        
                        total_notas[area] = total_notas.get(area, {})
                        total_notas[area][vol] = total_notas[area].get(vol, {})
                        total_notas[area][vol][ xlinha ] = {}
                        total_notas[area][vol][ xlinha ]['VALOR_TOTAL'] = nf.VALOR_TOTAL
                        total_notas[area][vol][ xlinha ]['BASE_ICMS'] = nf.BASE_ICMS
                        total_notas[area][vol][ xlinha ]['VALOR_ICMS'] = nf.VALOR_ICMS
                        statusDocumento = registroMestre[dic_campos[LayoutMestre]['SIT_DOC']-1]
                        
                        if  statusDocumento == 'N':
                            achou = 'N'
                            
                            for ix, bq in enumerate(abaBQ):
                            
                                if (nf.FILI_COD          == bq.FILI_COD
                                and nf.MES_ANO           == bq.MES_ANO
                                and nf.SERIE             == bq.SERIE  
                                and nf.AREA              == bq.AREA):                   
                                    achou = 'S'
                                    abaBQ[ix].QTD_NFS            = abaBQ[ix].QTD_NFS + nf.QTD_NFS
                                    abaBQ[ix].VALOR_TOTAL_MESTRE = abaBQ[ix].VALOR_TOTAL_MESTRE + nf.VALOR_TOTAL_MESTRE                                    
                                    abaBQ[ix].BASE_ICMS_MESTRE      = abaBQ[ix].BASE_ICMS_MESTRE + nf.BASE_ICMS_MESTRE # PTITES-866
                                    abaBQ[ix].VALOR_ICMS_MESTRE     = abaBQ[ix].VALOR_ICMS_MESTRE + nf.VALOR_ICMS_MESTRE # PTITES-866
                                    abaBQ[ix].ISENTAS_ICMS_MESTRE   = abaBQ[ix].ISENTAS_ICMS_MESTRE + nf.ISENTAS_ICMS_MESTRE # PTITES-866
                                    abaBQ[ix].OUTROS_VALORES_MESTRE = abaBQ[ix].OUTROS_VALORES_MESTRE + nf.OUTROS_VALORES_MESTRE # PTITES-866
  
                            if  achou == 'N':                           
                                abaBQ.append(nf)
                 
                if  Arquivo_tipo == 'D':
                    log('Arquivo Cadastro Em processamento =', nome_arquivo + " ...")
                    arquivo = os.path.join( diretorio, nome_arquivo )
                    encoding = comum.encodingDoArquivo( arquivo )
                    fd = open(arquivo, 'r', encoding=encoding)
                    endereco_vivo = retornaEnderecoVivo(uf)
                    nr = 1
                    xlinha = 0
                    for registro in fd:
                        xlinha += 1
                        registroCadastro = layout.quebraRegistro(registro, LayoutCadastro)
                        ic                        = checkListCadastro()
                        ic.EMPS_COD               = Empresa_Emitente
                        ic.FILI_COD               = Filial_Emitente
                        ic.MES_ANO                = Arquivo_ano_mes[2:] + "/20" + Arquivo_ano_mes[:2]
                        ic.SERIE                  = Arquivo_serie
                        ic.AREA                   = area
                        ic.QTD_NFS                = 0
                        ic.QTD_NFS_TOTAL          = 1
                        ic.TOTAL_NOME_CLIENTE     = 1 if registroCadastro[dic_campos[LayoutCadastro]['RAZAOSOCIAL']-1].strip().upper().__contains__('CONSUMIDOR') else 0
                        if  ic.TOTAL_NOME_CLIENTE == 1 :
                            ic.TOTAL_NOME_CLIENTE_VALOR_TOTAL_NF  += total_notas[area][vol][ xlinha ]['VALOR_TOTAL'] 
                            ic.TOTAL_NOME_CLIENTE_BASE_ICMS       += total_notas[area][vol][ xlinha ]['BASE_ICMS']
                            ic.TOTAL_NOME_CLIENTE_VALOR_ICMS      += total_notas[area][vol][ xlinha ]['VALOR_ICMS']
                        ic.TOTAL_NOME_CLIENTE_INV = 0
                        if (registroCadastro[dic_campos[LayoutCadastro]['RAZAOSOCIAL']-1].strip() == ""):
                            ic.TOTAL_NOME_CLIENTE_INV = 1
                        elif (registroCadastro[dic_campos[LayoutCadastro]['RAZAOSOCIAL']-1][0] == " " ):
                            ic.TOTAL_NOME_CLIENTE_INV = 1
                        elif (len(registroCadastro[dic_campos[LayoutCadastro]['RAZAOSOCIAL']-1].strip()) < 4):
                            ic.TOTAL_NOME_CLIENTE_INV = 1
                        elif (registroCadastro[dic_campos[LayoutCadastro]['RAZAOSOCIAL']-1][0] == "0" ):
                            ic.TOTAL_NOME_CLIENTE_INV = 1
                        elif (registroCadastro[dic_campos[LayoutCadastro]['RAZAOSOCIAL']-1][0] == "." ):
                            ic.TOTAL_NOME_CLIENTE_INV = 1
                        elif (registroCadastro[dic_campos[LayoutCadastro]['RAZAOSOCIAL']-1].strip().__contains__('...')):
                            ic.TOTAL_NOME_CLIENTE_INV = 1
                        elif (registroCadastro[dic_campos[LayoutCadastro]['RAZAOSOCIAL']-1].strip().__contains__('---')):
                            ic.TOTAL_NOME_CLIENTE_INV = 1
                        elif (registroCadastro[dic_campos[LayoutCadastro]['RAZAOSOCIAL']-1].strip().__contains__('|')):
                            ic.TOTAL_NOME_CLIENTE_INV = 1
                        elif (registroCadastro[dic_campos[LayoutCadastro]['RAZAOSOCIAL']-1].strip().__contains__('\\')):
                            ic.TOTAL_NOME_CLIENTE_INV = 1
                        elif (registroCadastro[dic_campos[LayoutCadastro]['RAZAOSOCIAL']-1].strip().__contains__('/')):
                            ic.TOTAL_NOME_CLIENTE_INV = 1
                        elif (registroCadastro[dic_campos[LayoutCadastro]['RAZAOSOCIAL']-1].strip().isdigit() ):
                            ic.TOTAL_NOME_CLIENTE_INV = 1
                        if ic.TOTAL_NOME_CLIENTE_INV == 1 :
                            ic.TOTAL_NOME_CLIENTE_INV_VALOR_TOTAL_NF  += total_notas[area][vol][ xlinha ]['VALOR_TOTAL'] 
                            ic.TOTAL_NOME_CLIENTE_INV_BASE_ICMS       += total_notas[area][vol][ xlinha ]['BASE_ICMS']
                            ic.TOTAL_NOME_CLIENTE_INV_VALOR_ICMS      += total_notas[area][vol][ xlinha ]['VALOR_ICMS']
                        ic.TOTAL_NOME_FANTASIA    = 0 # *** Não aparece no arquivo de cadastro
                        ic.TOTAL_CPF_CNPJ         = 0 if registroCadastro[dic_campos[LayoutCadastro]['CNPJ_CPF']-1].lstrip('0') not in ['11111111111', '11111111111111'] else 1
                        if  ic.TOTAL_CPF_CNPJ == 1 :
                            ic.TOTAL_CPF_CNPJ_VALOR_TOTAL_NF  += total_notas[area][vol][ xlinha ]['VALOR_TOTAL'] 
                            ic.TOTAL_CPF_CNPJ_BASE_ICMS       += total_notas[area][vol][ xlinha ]['BASE_ICMS']
                            ic.TOTAL_CPF_CNPJ_VALOR_ICMS      += total_notas[area][vol][ xlinha ]['VALOR_ICMS']
                        ic.TOTAL_CPF_CNPJ_INV = 0
                        if (registroCadastro[dic_campos[LayoutCadastro]['CNPJ_CPF']-1][0:3] == '000'):
                            if registroCadastro[dic_campos[LayoutCadastro]['CNPJ_CPF']-1][3:] != '11111111111' :
                                testecpfcnpj = util.isCpfValid(registroCadastro[dic_campos[LayoutCadastro]['CNPJ_CPF']-1][3:])
                                if not testecpfcnpj :
                                    testecpfcnpj = util.isCnpjValid(registroCadastro[dic_campos[LayoutCadastro]['CNPJ_CPF']-1])
                            else :
                                testecpfcnpj = True
                        else:
                            testecpfcnpj = util.isCnpjValid(registroCadastro[dic_campos[LayoutCadastro]['CNPJ_CPF']-1])
                        if (registroCadastro[dic_campos[LayoutCadastro]['CNPJ_CPF']-1].strip() == ""):
                            ic.TOTAL_CPF_CNPJ_INV = 1
                        elif (not registroCadastro[dic_campos[LayoutCadastro]['CNPJ_CPF']-1].strip().isdigit()):
                            ic.TOTAL_CPF_CNPJ_INV = 1
                        elif (testecpfcnpj == False):
                            ic.TOTAL_CPF_CNPJ_INV = 1
                        if  ic.TOTAL_CPF_CNPJ_INV == 1 :
                            ic.TOTAL_CPF_CNPJ_INV_VALOR_TOTAL_NF  += total_notas[area][vol][ xlinha ]['VALOR_TOTAL'] 
                            ic.TOTAL_CPF_CNPJ_INV_BASE_ICMS       += total_notas[area][vol][ xlinha ]['BASE_ICMS']
                            ic.TOTAL_CPF_CNPJ_INV_VALOR_ICMS      += total_notas[area][vol][ xlinha ]['VALOR_ICMS']
                        ic.TOTAL_TERMINAL         = 0 if registroCadastro[dic_campos[LayoutCadastro]['NUMEROTERMINAL']-1].strip().lstrip('0') != '1135497777' else 1
                        if  ic.TOTAL_TERMINAL == 1 :
                            ic.TOTAL_TERMINAL_VALOR_TOTAL_NF  += total_notas[area][vol][ xlinha ]['VALOR_TOTAL'] 
                            ic.TOTAL_TERMINAL_BASE_ICMS       += total_notas[area][vol][ xlinha ]['BASE_ICMS']
                            ic.TOTAL_TERMINAL_VALOR_ICMS      += total_notas[area][vol][ xlinha ]['VALOR_ICMS']
                        ic.TOTAL_TERMINAL_INV = 0 
                        lterminal = len(registroCadastro[dic_campos[LayoutCadastro]['NUMEROTERMINAL']-1].strip().lstrip('0'))
                        if (lterminal == 0):
                            if (semespacos(ic.SERIE) not in ('CA1','IN1','ASS','TV2','T1')):
                                ic.TOTAL_TERMINAL_INV = 1
                        elif ((lterminal < 10) or (lterminal > 11)):
                            ic.TOTAL_TERMINAL_INV = 1
                        elif (( not registroCadastro[dic_campos[LayoutCadastro]['NUMEROTERMINAL']-1].strip().lstrip('0').isdigit())):
                            ic.TOTAL_TERMINAL_INV = 1
                        elif ((registroCadastro[dic_campos[LayoutCadastro]['NUMEROTERMINAL']-1].strip().lstrip('0')[0:2] not in ('11','12','13','14','15','16','17','18','19','21','22','24','27','28','31','32','33','34','35','37','38','41','42','43','44','45','46','47','48','49','51','53','54','55','61','62','63','64','65','66','67','68','69','71','73','74','75','77','79','81','82','83','84','85','86','87','88','89','91','92','93','94','95','96','97','98','99'))):
                            ic.TOTAL_TERMINAL_INV = 1
                        elif (registroCadastro[dic_campos[LayoutCadastro]['NUMEROTERMINAL']-1].strip().lstrip('0')[2] == '0'):
                            ic.TOTAL_TERMINAL_INV = 1
                        if  ic.TOTAL_TERMINAL_INV == 1 :
                            ic.TOTAL_TERMINAL_INV_VALOR_TOTAL_NF  += total_notas[area][vol][ xlinha ]['VALOR_TOTAL'] 
                            ic.TOTAL_TERMINAL_INV_BASE_ICMS       += total_notas[area][vol][ xlinha ]['BASE_ICMS']
                            ic.TOTAL_TERMINAL_INV_VALOR_ICMS      += total_notas[area][vol][ xlinha ]['VALOR_ICMS']

                        ic.TOTAL_TEL_CONTATO      = 0 if registroCadastro[dic_campos[LayoutCadastro]['TELEFONECONTATO']-1].strip().lstrip('0') != '1135497777' else 1
                        if  ic.TOTAL_TEL_CONTATO == 1 :
                            ic.TOTAL_TEL_CONTATO_VALOR_TOTAL_NF  += total_notas[area][vol][ xlinha ]['VALOR_TOTAL'] 
                            ic.TOTAL_TEL_CONTATO_BASE_ICMS       += total_notas[area][vol][ xlinha ]['BASE_ICMS']
                            ic.TOTAL_TEL_CONTATO_VALOR_ICMS      += total_notas[area][vol][ xlinha ]['VALOR_ICMS']
                        ic.TOTAL_TEL_CONTATO_INV = 0 
                        lterminal = len(registroCadastro[dic_campos[LayoutCadastro]['TELEFONECONTATO']-1].strip().lstrip('0'))
                        if (lterminal == 0):
                            if (semespacos(ic.SERIE) not in ('CA1','IN1','ASS','TV2','T1')):
                                ic.TOTAL_TEL_CONTATO_INV = 1
                        elif ((lterminal < 10) or (lterminal > 11)):
                            ic.TOTAL_TEL_CONTATO_INV = 1
                        elif (( not registroCadastro[dic_campos[LayoutCadastro]['TELEFONECONTATO']-1].strip().lstrip('0').isdigit())):
                            ic.TOTAL_TEL_CONTATO_INV = 1
                        elif ((registroCadastro[dic_campos[LayoutCadastro]['TELEFONECONTATO']-1].strip().lstrip('0')[0:2] not in ('11','12','13','14','15','16','17','18','19','21','22','24','27','28','31','32','33','34','35','37','38','41','42','43','44','45','46','47','48','49','51','53','54','55','61','62','63','64','65','66','67','68','69','71','73','74','75','77','79','81','82','83','84','85','86','87','88','89','91','92','93','94','95','96','97','98','99'))):
                            ic.TOTAL_TEL_CONTATO_INV = 1
                        elif (registroCadastro[dic_campos[LayoutCadastro]['TELEFONECONTATO']-1].strip().lstrip('0')[2] == '0'):
                            ic.TOTAL_TEL_CONTATO_INV = 1
                        if  ic.TOTAL_TEL_CONTATO_INV == 1 :
                            ic.TOTAL_TEL_CONTATO_INV_VALOR_TOTAL_NF  += total_notas[area][vol][ xlinha ]['VALOR_TOTAL'] 
                            ic.TOTAL_TEL_CONTATO_INV_BASE_ICMS       += total_notas[area][vol][ xlinha ]['BASE_ICMS']
                            ic.TOTAL_TEL_CONTATO_INV_VALOR_ICMS      += total_notas[area][vol][ xlinha ]['VALOR_ICMS']
                        ic.TOTAL_ENDERECO_CLIENTE = 0 
                        if endereco_vivo :
                            if registroCadastro[dic_campos[LayoutCadastro]['ENDERECO']-1].strip().upper().__contains__( endereco_vivo[0][endereco_vivo[0].index(' '):].strip().upper() ) :
                                if int(registroCadastro[dic_campos[LayoutCadastro]['NUMERO']-1]) == int(endereco_vivo[1]) or int(registroCadastro[dic_campos[LayoutCadastro]['NUMERO']-1]) == 0 :
                                    ic.TOTAL_ENDERECO_CLIENTE = 1
                        if  ic.TOTAL_ENDERECO_CLIENTE == 1 :
                            ic.TOTAL_ENDERECO_CLIENTE_VALOR_TOTAL_NF  += total_notas[area][vol][ xlinha ]['VALOR_TOTAL'] 
                            ic.TOTAL_ENDERECO_CLIENTE_BASE_ICMS       += total_notas[area][vol][ xlinha ]['BASE_ICMS']
                            ic.TOTAL_ENDERECO_CLIENTE_VALOR_ICMS      += total_notas[area][vol][ xlinha ]['VALOR_ICMS']
                        ic.TOTAL_ENDERECO_CLIENTE_INV = 0
                        if ( len( registroCadastro[dic_campos[LayoutCadastro]['ENDERECO']-1].strip()) < 4 ):
                            ic.TOTAL_ENDERECO_CLIENTE_INV = 1
                        elif ( registroCadastro[dic_campos[LayoutCadastro]['ENDERECO']-1][0] in ('0','.') ):
                            ic.TOTAL_ENDERECO_CLIENTE_INV = 1
                        elif ( registroCadastro[dic_campos[LayoutCadastro]['ENDERECO']-1].strip().upper().__contains__('---') ):
                            ic.TOTAL_ENDERECO_CLIENTE_INV = 1
                        elif ( registroCadastro[dic_campos[LayoutCadastro]['ENDERECO']-1].strip().upper() == 'AVENIDA' ):
                            ic.TOTAL_ENDERECO_CLIENTE_INV = 1    
                        elif ( registroCadastro[dic_campos[LayoutCadastro]['ENDERECO']-1].strip().isdigit() ):
                            ic.TOTAL_ENDERECO_CLIENTE_INV = 1  
                        if  ic.TOTAL_ENDERECO_CLIENTE_INV == 1 :
                            ic.TOTAL_ENDERECO_CLIENTE_INV_VALOR_TOTAL_NF  += total_notas[area][vol][ xlinha ]['VALOR_TOTAL'] 
                            ic.TOTAL_ENDERECO_CLIENTE_INV_BASE_ICMS       += total_notas[area][vol][ xlinha ]['BASE_ICMS']
                            ic.TOTAL_ENDERECO_CLIENTE_INV_VALOR_ICMS      += total_notas[area][vol][ xlinha ]['VALOR_ICMS']
                        ic.TOTAL_INSCRICAO_INV = 0
                        if (registroCadastro[dic_campos[LayoutCadastro]['IE']-1].strip() != 'ISENTO' ):
                            if ( not vie.valida_insc_est(registroCadastro[dic_campos[LayoutCadastro]['UF']-1].strip(), registroCadastro[dic_campos[LayoutCadastro]['IE']-1].strip() ) ) :
                                ic.TOTAL_INSCRICAO_INV = 1
                        if  ic.TOTAL_INSCRICAO_INV == 1 :
                            ic.TOTAL_INSCRICAO_INV_VALOR_TOTAL_NF  += total_notas[area][vol][ xlinha ]['VALOR_TOTAL'] 
                            ic.TOTAL_INSCRICAO_INV_BASE_ICMS       += total_notas[area][vol][ xlinha ]['BASE_ICMS']
                            ic.TOTAL_INSCRICAO_INV_VALOR_ICMS      += total_notas[area][vol][ xlinha ]['VALOR_ICMS']
                        if ic.TOTAL_NOME_CLIENTE > 0 or ic.TOTAL_NOME_FANTASIA > 0 or ic.TOTAL_CPF_CNPJ > 0 \
                                or ic.TOTAL_TERMINAL > 0 or ic.TOTAL_TEL_CONTATO > 0 or ic.TOTAL_ENDERECO_CLIENTE > 0 \
                                or ic.TOTAL_NOME_CLIENTE_INV > 0 or ic.TOTAL_CPF_CNPJ_INV > 0 or ic.TOTAL_INSCRICAO_INV > 0 \
                                or ic.TOTAL_TERMINAL_INV > 0 or ic.TOTAL_TEL_CONTATO_INV > 0 or ic.TOTAL_ENDERECO_CLIENTE_INV > 0:
                            ic.QTD_NFS                = 1
                        achou = 'N'
                            
                        for ix, cc in enumerate(abaCC):
                            if (cc.FILI_COD          == ic.FILI_COD
                            and cc.MES_ANO           == ic.MES_ANO
                            and cc.SERIE             == ic.SERIE  
                            and cc.AREA              == ic.AREA):
                                achou = 'S'
                                abaCC[ix].QTD_NFS                    += ic.QTD_NFS
                                abaCC[ix].QTD_NFS_TOTAL              += ic.QTD_NFS_TOTAL

                                abaCC[ix].TOTAL_NOME_CLIENTE         += ic.TOTAL_NOME_CLIENTE
                                abaCC[ix].TOTAL_NOME_CLIENTE_INV     += ic.TOTAL_NOME_CLIENTE_INV
                                abaCC[ix].TOTAL_NOME_FANTASIA        += ic.TOTAL_NOME_FANTASIA   
                                abaCC[ix].TOTAL_CPF_CNPJ             += ic.TOTAL_CPF_CNPJ 
                                abaCC[ix].TOTAL_CPF_CNPJ_INV         += ic.TOTAL_CPF_CNPJ_INV                                 
                                abaCC[ix].TOTAL_TERMINAL             += ic.TOTAL_TERMINAL
                                abaCC[ix].TOTAL_TERMINAL_INV         += ic.TOTAL_TERMINAL_INV
                                abaCC[ix].TOTAL_TEL_CONTATO          += ic.TOTAL_TEL_CONTATO  
                                abaCC[ix].TOTAL_TEL_CONTATO_INV      += ic.TOTAL_TEL_CONTATO_INV
                                abaCC[ix].TOTAL_ENDERECO_CLIENTE     += ic.TOTAL_ENDERECO_CLIENTE
                                abaCC[ix].TOTAL_ENDERECO_CLIENTE_INV += ic.TOTAL_ENDERECO_CLIENTE_INV
                                abaCC[ix].TOTAL_INSCRICAO_INV        += ic.TOTAL_INSCRICAO_INV      
                                abaCC[ix].TOTAL_NOME_CLIENTE_VALOR_TOTAL_NF         += ic.TOTAL_NOME_CLIENTE_VALOR_TOTAL_NF
                                abaCC[ix].TOTAL_NOME_CLIENTE_INV_VALOR_TOTAL_NF     += ic.TOTAL_NOME_CLIENTE_INV_VALOR_TOTAL_NF
                                abaCC[ix].TOTAL_NOME_FANTASIA_VALOR_TOTAL_NF        += ic.TOTAL_NOME_FANTASIA_VALOR_TOTAL_NF
                                abaCC[ix].TOTAL_CPF_CNPJ_VALOR_TOTAL_NF             += ic.TOTAL_CPF_CNPJ_VALOR_TOTAL_NF
                                abaCC[ix].TOTAL_CPF_CNPJ_INV_VALOR_TOTAL_NF         += ic.TOTAL_CPF_CNPJ_INV_VALOR_TOTAL_NF
                                abaCC[ix].TOTAL_TERMINAL_VALOR_TOTAL_NF             += ic.TOTAL_TERMINAL_VALOR_TOTAL_NF
                                abaCC[ix].TOTAL_TERMINAL_INV_VALOR_TOTAL_NF         += ic.TOTAL_TERMINAL_INV_VALOR_TOTAL_NF
                                abaCC[ix].TOTAL_TEL_CONTATO_VALOR_TOTAL_NF          += ic.TOTAL_TEL_CONTATO_VALOR_TOTAL_NF
                                abaCC[ix].TOTAL_TEL_CONTATO_INV_VALOR_TOTAL_NF      += ic.TOTAL_TEL_CONTATO_INV_VALOR_TOTAL_NF
                                abaCC[ix].TOTAL_ENDERECO_CLIENTE_VALOR_TOTAL_NF     += ic.TOTAL_ENDERECO_CLIENTE_VALOR_TOTAL_NF
                                abaCC[ix].TOTAL_ENDERECO_CLIENTE_INV_VALOR_TOTAL_NF += ic.TOTAL_ENDERECO_CLIENTE_INV_VALOR_TOTAL_NF
                                abaCC[ix].TOTAL_INSCRICAO_INV_VALOR_TOTAL_NF        += ic.TOTAL_INSCRICAO_INV_VALOR_TOTAL_NF
                                abaCC[ix].TOTAL_NOME_CLIENTE_BASE_ICMS         += ic.TOTAL_NOME_CLIENTE_BASE_ICMS
                                abaCC[ix].TOTAL_NOME_CLIENTE_INV_BASE_ICMS     += ic.TOTAL_NOME_CLIENTE_INV_BASE_ICMS
                                abaCC[ix].TOTAL_NOME_FANTASIA_BASE_ICMS        += ic.TOTAL_NOME_FANTASIA_BASE_ICMS
                                abaCC[ix].TOTAL_CPF_CNPJ_BASE_ICMS             += ic.TOTAL_CPF_CNPJ_BASE_ICMS
                                abaCC[ix].TOTAL_CPF_CNPJ_INV_BASE_ICMS         += ic.TOTAL_CPF_CNPJ_INV_BASE_ICMS
                                abaCC[ix].TOTAL_TERMINAL_BASE_ICMS             += ic.TOTAL_TERMINAL_BASE_ICMS
                                abaCC[ix].TOTAL_TERMINAL_INV_BASE_ICMS         += ic.TOTAL_TERMINAL_INV_BASE_ICMS
                                abaCC[ix].TOTAL_TEL_CONTATO_BASE_ICMS          += ic.TOTAL_TEL_CONTATO_BASE_ICMS
                                abaCC[ix].TOTAL_TEL_CONTATO_INV_BASE_ICMS      += ic.TOTAL_TEL_CONTATO_INV_BASE_ICMS
                                abaCC[ix].TOTAL_ENDERECO_CLIENTE_BASE_ICMS     += ic.TOTAL_ENDERECO_CLIENTE_BASE_ICMS
                                abaCC[ix].TOTAL_ENDERECO_CLIENTE_INV_BASE_ICMS += ic.TOTAL_ENDERECO_CLIENTE_INV_BASE_ICMS
                                abaCC[ix].TOTAL_INSCRICAO_INV_BASE_ICMS        += ic.TOTAL_INSCRICAO_INV_BASE_ICMS
                                abaCC[ix].TOTAL_NOME_CLIENTE_VALOR_ICMS         += ic.TOTAL_NOME_CLIENTE_VALOR_ICMS
                                abaCC[ix].TOTAL_NOME_CLIENTE_INV_VALOR_ICMS     += ic.TOTAL_NOME_CLIENTE_INV_VALOR_ICMS
                                abaCC[ix].TOTAL_NOME_FANTASIA_VALOR_ICMS        += ic.TOTAL_NOME_FANTASIA_VALOR_ICMS
                                abaCC[ix].TOTAL_CPF_CNPJ_VALOR_ICMS             += ic.TOTAL_CPF_CNPJ_VALOR_ICMS
                                abaCC[ix].TOTAL_CPF_CNPJ_INV_VALOR_ICMS         += ic.TOTAL_CPF_CNPJ_INV_VALOR_ICMS
                                abaCC[ix].TOTAL_TERMINAL_VALOR_ICMS             += ic.TOTAL_TERMINAL_VALOR_ICMS
                                abaCC[ix].TOTAL_TERMINAL_INV_VALOR_ICMS         += ic.TOTAL_TERMINAL_INV_VALOR_ICMS
                                abaCC[ix].TOTAL_TEL_CONTATO_VALOR_ICMS          += ic.TOTAL_TEL_CONTATO_VALOR_ICMS
                                abaCC[ix].TOTAL_TEL_CONTATO_INV_VALOR_ICMS      += ic.TOTAL_TEL_CONTATO_INV_VALOR_ICMS
                                abaCC[ix].TOTAL_ENDERECO_CLIENTE_VALOR_ICMS     += ic.TOTAL_ENDERECO_CLIENTE_VALOR_ICMS
                                abaCC[ix].TOTAL_ENDERECO_CLIENTE_INV_VALOR_ICMS += ic.TOTAL_ENDERECO_CLIENTE_INV_VALOR_ICMS
                                abaCC[ix].TOTAL_INSCRICAO_INV_VALOR_ICMS        += ic.TOTAL_INSCRICAO_INV_VALOR_ICMS
                        if  achou == 'N':                           
                            abaCC.append(ic)
            else:
                log('Arquivo não é referente ao ID da Série. Não será processado ! > ' + nome_arquivo)

def carregaDadosPlanliha(workbook):
    global strglobal
    global ERRO
    
    
    erroitemp = "" 
    erroitemr = "" 
    errototalp = "" 
    errototalr = "" 
    errototald = "" 
    erromestrep = "" 
    erromestrer = ""     
    
    ERRO     = ""
    newline  = '\n'
    msgerro1 = "Existem valores negativos para CFOP diferente de 0000."
    msgerro2 = "Existem valores em (Base | Valor | Isentas) para CFOP igual a 0000."
    msgerro3 = "VALOR_TOTAL - DESCONTO deve ser igual a VALOR_TOTAL da tabela MESTRE DE NOTA FISCAL."
    msgerro4 = "CFOP não permitido. Primeiro algarismo diferente de 5, 6 ou 7 e segundo algarismo diferente de 3."  
    msgerro5 = "VALOR_CONTABIL deve ser igual a VALOR_TOTAL da tabela MESTRE DE NOTA FISCAL."         
    msgerro6 = "A soma dos valores de BASE_ICMS, ISENTAS_ICMS, e OUTROS_VALORES deve ser igual a VALOR_CONTABIL." 
    msgerro7 = "Diferença BASE_ICM superior a R$0,05."
    msgerro8 = "Diferença VALOR_ICMS superior a R$0,05."
    msgerro9 = "VALOR_TOTAL deve ser igual a VALOR_TOTAL - DESCONTO da tabela TOTALIZADORES PROTOCOLADO E ATUAL_TI - ITEM DE NOTA FISCAL."
    msgerro10 = "QTD_NFS do ATUAL_TI diferente do protocolado."
    msgerro11 = "QTD_NFS do protocolado diferente do ATUAL_TI."
    msgerro12 = "VALOR_TOTAL deve ser igual a VALOR_CONTABIL da tabela TOTALIZADORES"
    msgerro13 = "Diferença VALOR_CONTABIL superior a R$0,05."
   
    totir_valor_total = 0.00 
    totir_base_icms = 0.00
    totir_valor_icms = 0.00
    totir_isentas_icms = 0.00
    totir_outros_valores =  0.00
    totir_desconto =  0.00
    totir_valor_contabil = 0.00
   
    totip_valor_total =  0.00
    totip_base_icms = 0.00
    totip_valor_icms = 0.00
    totip_isentas_icms = 0.00
    totip_outros_valores =  0.00
    totip_desconto =  0.00
    totip_valor_contabil = 0.00
 
    totip_EMPS_COD = totir_EMPS_COD = "" 
    totip_FILI_COD = totir_FILI_COD = ""
    totip_MES_ANO  = totir_MES_ANO  = ""
    totip_SERIE    = totir_SERIE    = ""
 
   
#### - ITEM DE NOTA FISCAL
#### - ITEM DE NOTA FISCAL
    # ------- ABA Batimento --------
    log("- Aba Batimento... ")
    planilha = workbook.active
    planilha.title = "Batimento"
     
    formatarCelulaTituloMasterA(planilha.cell(1,  1, "ITEM DE NOTA FISCAL"))
    formatarCelulaTituloMasterP(planilha.cell(1,  14, "VALIDAÇÃO")) 
    planilha.cell(1,  1).border = bordaB
    planilha.cell(1,  14).border = bordaB
    planilha.merge_cells('A1:M1')    
   
    formatarCelulaTitulo(planilha.cell(2,  1, "EMPS_COD"))
    formatarCelulaTitulo(planilha.cell(2,  2, "FILI_COD"))
    formatarCelulaTitulo(planilha.cell(2,  3, "MES_ANO"))
    formatarCelulaTitulo(planilha.cell(2,  4, "SERIE"))
    formatarCelulaTitulo(planilha.cell(2,  5, "CFOP"))
    formatarCelulaTitulo(planilha.cell(2,  6, "VALOR_TOTAL"))
    formatarCelulaTitulo(planilha.cell(2,  7, "BASE_ICMS"))
    formatarCelulaTitulo(planilha.cell(2,  8, "VALOR_ICMS"))
    formatarCelulaTitulo(planilha.cell(2,  9, "ISENTAS_ICMS"))
    formatarCelulaTitulo(planilha.cell(2, 10, "OUTROS_VALORES"))
    formatarCelulaTitulo(planilha.cell(2, 11, "DESCONTO"))
    formatarCelulaTitulo(planilha.cell(2, 12, "VALOR_CONTABIL"))
    formatarCelulaTitulo(planilha.cell(2, 13, "AREA"))
    formatarCelulaTitulo(planilha.cell(2, 14, ""))

    numLinha = 3
    
    valor_total_item_reg = 0.00
    
    for bt in abaBT:                    

        formatarCelulaDetalhe(planilha.cell(numLinha,  1, bt.EMPS_COD)                  ,bt.AREA, 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  2, bt.FILI_COD)                  ,bt.AREA, 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  3, bt.MES_ANO)                   ,bt.AREA, 'MA')
        formatarCelulaDetalhe(planilha.cell(numLinha,  4, bt.SERIE)                     ,bt.AREA, 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  5, bt.CFOP)                      ,bt.AREA, 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  6, bt.VALOR_TOTAL    / 100 )  ,bt.AREA, 'VL') #***
        formatarCelulaDetalhe(planilha.cell(numLinha,  7, bt.BASE_ICMS      / 100 )  ,bt.AREA, 'VL') #***
        formatarCelulaDetalhe(planilha.cell(numLinha,  8, bt.VALOR_ICMS     / 100 )  ,bt.AREA, 'VL') #***
        formatarCelulaDetalhe(planilha.cell(numLinha,  9, bt.ISENTAS_ICMS   / 100 )  ,bt.AREA, 'VL') #***
        formatarCelulaDetalhe(planilha.cell(numLinha, 10, bt.OUTROS_VALORES / 100 )  ,bt.AREA, 'VL') #***
        formatarCelulaDetalhe(planilha.cell(numLinha, 11, bt.DESCONTO       / 100 )  ,bt.AREA, 'VL') #***
        formatarCelulaDetalhe(planilha.cell(numLinha, 12, bt.VALOR_CONTABIL / 100 )  ,bt.AREA, 'VL') #***
        
        formatarCelulaDetalhe(planilha.cell(numLinha, 13, bt.AREA)                      ,bt.AREA, 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha, 14, "OK")                         ,bt.AREA, 'CH')
        
        #TOTALIZAÇOES PARA ABA TOTALIZADORES PROTOCOLADO E REGERADO - ITEM DE NOTA FISCAL
        if (bt.AREA == "ATUAL_TI" ):
            totir_valor_total    = totir_valor_total    + bt.VALOR_TOTAL
            totir_base_icms      = totir_base_icms      + bt.BASE_ICMS
            totir_valor_icms     = totir_valor_icms     + bt.VALOR_ICMS
            totir_isentas_icms   = totir_isentas_icms   + bt.ISENTAS_ICMS
            totir_outros_valores = totir_outros_valores + bt.OUTROS_VALORES 
            totir_desconto       = totir_desconto       + bt.DESCONTO
            totir_valor_contabil = totir_valor_contabil + bt.VALOR_CONTABIL
            totir_EMPS_COD = bt.EMPS_COD
            totir_FILI_COD = bt.FILI_COD
            totir_MES_ANO  = bt.MES_ANO
            totir_SERIE    = bt.SERIE
        
            
        if ("PROTOCOLADO" in bt.AREA):   # PTITES-1367  
            totip_valor_total    = totip_valor_total    + bt.VALOR_TOTAL
            totip_base_icms      = totip_base_icms      + bt.BASE_ICMS
            totip_valor_icms     = totip_valor_icms     + bt.VALOR_ICMS
            totip_isentas_icms   = totip_isentas_icms   + bt.ISENTAS_ICMS
            totip_outros_valores = totip_outros_valores + bt.OUTROS_VALORES
            totip_desconto       = totip_desconto       + bt.DESCONTO
            totip_valor_contabil = totip_valor_contabil + bt.VALOR_CONTABIL
            totip_EMPS_COD = bt.EMPS_COD
            totip_FILI_COD = bt.FILI_COD
            totip_MES_ANO  = bt.MES_ANO
            totip_SERIE    = bt.SERIE
     
        # VALIDAÇÕES:   RF002 - Aba Batimento 

        msgerro = "" 
        altura = 15
        qtderr = 0  
        
        if (bt.AREA == "ATUAL_TI" ):

            #Regra 1
            if (bt.CFOP != "0000" and ( bt.VALOR_TOTAL < 0.00 or bt.BASE_ICMS < 0.00 or bt.VALOR_ICMS < 0.00 or bt.ISENTAS_ICMS < 0.00 or bt.OUTROS_VALORES < 0.00 or  bt.DESCONTO < 0.00 or bt.VALOR_CONTABIL < 0.00 )):
                if len(msgerro) > 3:
                    qtderr += 1
                    msgerro = msgerro + newline + msgerro1
                else:
                    qtderr = 1
                    msgerro = msgerro1
                    
                erroitemr = erroitemr + "CFOP = "+ bt.CFOP + " => " + msgerro1 + newline                   
            
            #Regra 2
            if (bt.CFOP == "0000" and (bt.BASE_ICMS != 0.00 or bt.VALOR_ICMS != 0.00 or bt.ISENTAS_ICMS!= 0.00) ):
                if len(msgerro) > 3:
                    qtderr += 1
                    msgerro = msgerro + newline + msgerro2
                else:
                    qtderr = 1
                    msgerro = msgerro2
                    
                erroitemr = erroitemr + "CFOP = "+ bt.CFOP + " => " + msgerro2 + newline
                    
            #Regra 4    
            if (bt.CFOP != '0000' and bt.CFOP[0] not in ('5','6','7') and (bt.CFOP[1] != '3') ):                
                if len(msgerro) > 3:
                    qtderr += 1
                    msgerro = msgerro + newline + msgerro4
                else:
                    qtderr = 1
                    msgerro = msgerro4
                    
                erroitemr = erroitemr + "CFOP = "+ bt.CFOP + " => " + msgerro4 + newline
                
            if len(msgerro) > 3:
                planilha.row_dimensions[numLinha].height = qtderr * altura
                formatarCelulaDetalhe(planilha.cell(numLinha, 14,msgerro) ,"ERRO", 'CH')  
                planilha.cell(numLinha,  14).alignment = Alignment(wrap_text = True) 
                strglobal = "_ERRO.xlsx"
 
        addTabelaValidacao(
            bt.EMPS_COD
           ,bt.FILI_COD
           ,bt.MES_ANO
           ,bt.SERIE
           ,bt.CFOP
           ,bt.VALOR_TOTAL
           ,bt.BASE_ICMS
           ,bt.VALOR_ICMS
           ,bt.ISENTAS_ICMS
           ,bt.OUTROS_VALORES
           ,bt.DESCONTO
           ,bt.VALOR_CONTABIL
           ,bt.AREA
           ,bt.VALIDACAO
        )

        numLinha += 1

    if ano <= 16:
       planilha.delete_cols(2)
       planilha.column_dimensions['A'].width = 11
       planilha.column_dimensions['B'].width = 13
       planilha.column_dimensions['C'].width = 9
       planilha.column_dimensions['D'].width = 7
       planilha.column_dimensions['E'].width = 17
       planilha.column_dimensions['F'].width = 17
       planilha.column_dimensions['G'].width = 17
       planilha.column_dimensions['H'].width = 17
       planilha.column_dimensions['I'].width = 17
       planilha.column_dimensions['J'].width = 17
       planilha.column_dimensions['K'].width = 17
       planilha.column_dimensions['L'].width = 17
       planilha.column_dimensions['M'].width = 17
       planilha.column_dimensions['N'].width = 130
    else:
       planilha.column_dimensions['A'].width = 11
       planilha.column_dimensions['B'].width = 9
       planilha.column_dimensions['C'].width = 11
       planilha.column_dimensions['D'].width = 6
       planilha.column_dimensions['E'].width = 17
       planilha.column_dimensions['F'].width = 17
       planilha.column_dimensions['G'].width = 17
       planilha.column_dimensions['H'].width = 17
       planilha.column_dimensions['I'].width = 17
       planilha.column_dimensions['J'].width = 17
       planilha.column_dimensions['K'].width = 17
       planilha.column_dimensions['L'].width = 17
       planilha.column_dimensions['M'].width = 15
       planilha.column_dimensions['N'].width = 130

    if ( (erroitemp != "") or (erroitemr != "") ):
        ERRO = ERRO + ("-"* 100) + newline
        ERRO = ERRO + "===> ERROS DE VALIDACAO - PLANILHA (ITEM DE NOTA FISCAL)\n\n"
        if (erroitemp != ""):  
            ERRO = ERRO + "PROTOCOLADO:\n"
            ERRO = ERRO + erroitemp + "\n"
        if (erroitemr != ""):  
            ERRO = ERRO + "ATUAL_TI:\n"
            ERRO = ERRO + erroitemr + "\n"

    numLinha += 3    

#### - TOTALIZADORES PROTOCOLO E REGERADO - ITEM DE NOTA FISCAL
#### - TOTALIZADORES PROTOCOLO E REGERADO - ITEM DE NOTA FISCAL
    formatarCelulaTituloMasterA(planilha.cell(numLinha,  1, "TOTALIZADORES PROTOCOLADO E ATUAL_TI - ITEM DE NOTA FISCAL"))
    formatarCelulaTituloMasterP(planilha.cell(numLinha,  14, "VALIDAÇÃO"))    
    planilha.cell(numLinha,  1).border = bordaB
    planilha.cell(numLinha,  14).border = bordaB      
    planilha.merge_cells('A'+str(numLinha)+':M'+str(numLinha))
    numLinha += 1
    formatarCelulaTitulo(planilha.cell(numLinha,  1, "EMPS_COD"))
    formatarCelulaTitulo(planilha.cell(numLinha,  2, "FILI_COD"))
    formatarCelulaTitulo(planilha.cell(numLinha,  3, "MES_ANO"))
    formatarCelulaTitulo(planilha.cell(numLinha,  4, "SERIE"))
    formatarCelulaTitulo(planilha.cell(numLinha,  5, "VALOR_TOTAL"))
    formatarCelulaTitulo(planilha.cell(numLinha,  6, "BASE_ICMS"))
    formatarCelulaTitulo(planilha.cell(numLinha,  7, "VALOR_ICMS"))
    formatarCelulaTitulo(planilha.cell(numLinha,  8, "ISENTAS_ICMS"))
    formatarCelulaTitulo(planilha.cell(numLinha,  9, "OUTROS_VALORES"))
    formatarCelulaTitulo(planilha.cell(numLinha,  10, "DESCONTO"))
    formatarCelulaTitulo(planilha.cell(numLinha,  11, "VALOR_CONTABIL"))
    formatarCelulaTitulo(planilha.cell(numLinha,  12, "AREA"))
    planilha.cell(numLinha,  12).alignment = Alignment(horizontal='center')
    formatarCelulaTitulo(planilha.cell(numLinha,  13, ""))
    formatarCelulaTitulo(planilha.cell(numLinha,  14, ""))
    planilha.merge_cells('L'+str(numLinha)+':M'+str(numLinha))
    numLinha +=1
 
    formatarCelulaDetalhe(planilha.cell(numLinha,  1, totip_EMPS_COD)                             ,"PROTOCOLADO", 'CH')
    formatarCelulaDetalhe(planilha.cell(numLinha,  2, totip_FILI_COD)                             ,"PROTOCOLADO", 'CH')
    formatarCelulaDetalhe(planilha.cell(numLinha,  3, totip_MES_ANO)                              ,"PROTOCOLADO", 'CH')
    formatarCelulaDetalhe(planilha.cell(numLinha,  4, totip_SERIE)                                ,"PROTOCOLADO", 'CH')
    formatarCelulaDetalhe(planilha.cell(numLinha,  5, totip_valor_total / 100    , )              ,"PROTOCOLADO", 'VL')
    formatarCelulaDetalhe(planilha.cell(numLinha,  6, totip_base_icms / 100      , )              ,"PROTOCOLADO", 'VL')
    formatarCelulaDetalhe(planilha.cell(numLinha,  7, totip_valor_icms / 100     , )              ,"PROTOCOLADO", 'VL')
    formatarCelulaDetalhe(planilha.cell(numLinha,  8, totip_isentas_icms / 100   , )              ,"PROTOCOLADO", 'VL')
    formatarCelulaDetalhe(planilha.cell(numLinha,  9, totip_outros_valores / 100 , )              ,"PROTOCOLADO", 'VL')
    formatarCelulaDetalhe(planilha.cell(numLinha, 10, totip_desconto / 100       , )              ,"PROTOCOLADO", 'VL')
    formatarCelulaDetalhe(planilha.cell(numLinha, 11, totip_valor_contabil / 100 , )              ,"PROTOCOLADO", 'VL')
    
    formatarCelulaDetalhe(planilha.cell(numLinha, 12, configuracoes.origem_protocolado)           ,"PROTOCOLADO", 'CH') #PTITES-1367
    planilha.cell(numLinha,  12).alignment = Alignment(horizontal='center')
    planilha.merge_cells('L'+str(numLinha)+':M'+str(numLinha))
    formatarCelulaDetalhe(planilha.cell(numLinha, 14, "")                                         ,"PROTOCOLADO", 'CH')
    
    mqtd_nfs_reg = mqtd_nfs_prot = 0
    mvalor_total_reg = mvalor_total_prot = 0.00
    for bq in abaBQ:                            
        if ("PROTOCOLADO" in bq.AREA): #PTITES-1367
            mqtd_nfs_prot = mqtd_nfs_prot + bq.QTD_NFS
            mvalor_total_prot = mvalor_total_prot + bq.VALOR_TOTAL_MESTRE
            
        if (bq.AREA == "ATUAL_TI"):
            mqtd_nfs_reg = mqtd_nfs_reg + bq.QTD_NFS
            mvalor_total_reg = mvalor_total_reg + bq.VALOR_TOTAL_MESTRE
    
    msgerro = "" 
    altura = 15
    qtderr = 0 
    diferenca = abs(round(totip_valor_contabil , 2) - round( mvalor_total_prot , 2)) / 100 #***
    
    
    if ( diferenca > 0.05 ):
        if len(msgerro) > 3:
            msgerro = msgerro + newline + msgerro5
            qtderr += 1
        else:
            msgerro = msgerro5
            qtderr = 1
        errototalp = errototalp + msgerro5 + newline            
         
    if len(msgerro) > 3:
        planilha.row_dimensions[numLinha].height = qtderr * altura
        formatarCelulaDetalhe(planilha.cell(numLinha, 14,msgerro) ,"ERRO", 'CH') 
        planilha.cell(numLinha,  14).alignment = Alignment(wrap_text = True)         
        strglobal = "_ERRO.xlsx"
    numLinha += 1

    formatarCelulaDetalhe(planilha.cell(numLinha,  1, totir_EMPS_COD)                             ,"ATUAL_TI", 'CH')
    formatarCelulaDetalhe(planilha.cell(numLinha,  2, totir_FILI_COD)                             ,"ATUAL_TI", 'CH')
    formatarCelulaDetalhe(planilha.cell(numLinha,  3, totir_MES_ANO)                              ,"ATUAL_TI", 'CH')
    formatarCelulaDetalhe(planilha.cell(numLinha,  4, totir_SERIE)                                ,"ATUAL_TI", 'CH')
    formatarCelulaDetalhe(planilha.cell(numLinha,  5, totir_valor_total / 100    , )              ,"ATUAL_TI", 'VL') 
    formatarCelulaDetalhe(planilha.cell(numLinha,  6, totir_base_icms / 100      , )              ,"ATUAL_TI", 'VL') 
    formatarCelulaDetalhe(planilha.cell(numLinha,  7, totir_valor_icms / 100     , )              ,"ATUAL_TI", 'VL') 
    formatarCelulaDetalhe(planilha.cell(numLinha,  8, totir_isentas_icms / 100   , )              ,"ATUAL_TI", 'VL') 
    formatarCelulaDetalhe(planilha.cell(numLinha,  9, totir_outros_valores / 100 , )              ,"ATUAL_TI", 'VL') 
    formatarCelulaDetalhe(planilha.cell(numLinha, 10, totir_desconto / 100       , )              ,"ATUAL_TI", 'VL') 
    formatarCelulaDetalhe(planilha.cell(numLinha, 11, totir_valor_contabil / 100 , )              ,"ATUAL_TI", 'VL') 

    
    formatarCelulaDetalhe(planilha.cell(numLinha, 12, "ATUAL_TI")                                 ,"ATUAL_TI", 'CH')
    planilha.cell(numLinha,  12).alignment = Alignment(horizontal='center')
    planilha.merge_cells('L'+str(numLinha)+':M'+str(numLinha))
    formatarCelulaDetalhe(planilha.cell(numLinha, 14, "")                                         ,"ATUAL_TI", 'CH')
           
    msgerro = "" 
    altura = 15
    qtderr = 0 

    valtemp = abs(round(totir_valor_contabil , 2) - round( totir_base_icms + totir_isentas_icms + totir_outros_valores , 2) ) / 100 
        

    if ( valtemp > 0.05):

        if len(msgerro) > 3:
            msgerro = msgerro + newline + msgerro6 + "Diferença = " + str(valtemp)
            qtderr += 1
        else:
            qtderr = 1
            msgerro = msgerro6
        errototalr = errototalr + msgerro6 + newline            
  
    if len(msgerro) > 3:
        planilha.row_dimensions[numLinha].height = qtderr * altura
        formatarCelulaDetalhe(planilha.cell(numLinha, 14,msgerro) ,"ERRO", 'CH')  
        planilha.cell(numLinha,  14).alignment = Alignment(wrap_text = True)         
        strglobal = "_ERRO.xlsx"
    
    numLinha += 1

    formatarCelulaTituloMasterA(planilha.cell(numLinha,  1, "DIFERENÇA"))
    planilha.cell(numLinha,  1).border = bordaB
    planilha.merge_cells('A'+str(numLinha)+':D'+str(numLinha)) 
    planilha.cell(numLinha,  1).alignment = Alignment(vertical='center')
    formatarCelulaDetalhe(planilha.cell(numLinha,  5, round(totip_valor_total - totir_valor_total, 2)/100 ) ,"ATUAL_TI", 'VL') #*** ALT VICTOR
    planilha.cell(numLinha,  5).alignment = Alignment(vertical='center')
    formatarCelulaDetalhe(planilha.cell(numLinha,  6, round(totip_base_icms - totir_base_icms           , 2) /100) ,"ATUAL_TI", 'VL') #***
    planilha.cell(numLinha,  6).alignment = Alignment(vertical='center')
    formatarCelulaDetalhe(planilha.cell(numLinha,  7, round(totip_valor_icms - totir_valor_icms         , 2) /100) ,"ATUAL_TI", 'VL') #***
    planilha.cell(numLinha,  7).alignment = Alignment(vertical='center')
    formatarCelulaDetalhe(planilha.cell(numLinha,  8, round(totip_isentas_icms - totir_isentas_icms     , 2) /100) ,"ATUAL_TI", 'VL') #***
    planilha.cell(numLinha,  8).alignment = Alignment(vertical='center')
    formatarCelulaDetalhe(planilha.cell(numLinha,  9, round(totip_outros_valores - totir_outros_valores , 2) /100) ,"ATUAL_TI", 'VL') #***
    planilha.cell(numLinha,  9).alignment = Alignment(vertical='center')
    formatarCelulaDetalhe(planilha.cell(numLinha, 10, round(totip_desconto - totir_desconto             , 2) /100) ,"ATUAL_TI", 'VL') #***
    planilha.cell(numLinha,  10).alignment = Alignment(vertical='center')
    formatarCelulaDetalhe(planilha.cell(numLinha, 11, round(totip_valor_contabil - totir_valor_contabil , 2) /100) ,"ATUAL_TI", 'VL') #***
    planilha.cell(numLinha,  11).alignment = Alignment(vertical='center')
    formatarCelulaDetalhe(planilha.cell(numLinha, 12, "")                                                   ,"ATUAL_TI", 'CH')
    planilha.cell(numLinha,  12).alignment = Alignment(vertical='center')
    planilha.merge_cells('L'+str(numLinha)+':M'+str(numLinha))
    formatarCelulaDetalhe(planilha.cell(numLinha, 14, "")                                                   ,"ATUAL_TI", 'CH')
    planilha.cell(numLinha,  14).alignment = Alignment(vertical='center')
    
    msgerro = "" 
    altura = 15
    qtderr = 0

    if ( abs(round(totir_base_icms , 2) - round( totip_base_icms , 2)) /100 > 0.05 ): #***
        if len(msgerro) > 3:
            qtderr += 1
            msgerro = msgerro + newline + msgerro7
        else:
            msgerro = msgerro7 
            qtderr = 1
        errototald = errototald + msgerro7 + newline            
        
    if ( abs(round(totir_valor_icms , 2) - round( totip_valor_icms , 2)) / 100 > 0.05 ): #****
        if len(msgerro) > 3:
            msgerro = msgerro + newline + msgerro8
            qtderr += 1
        else:
            msgerro = msgerro8 
            qtderr = 1
        errototald = errototald + msgerro8 + newline    

    if ( abs(round(totir_valor_contabil , 2) - round( totip_valor_contabil , 2)) / 100 > 0.05 ): #****
        if len(msgerro) > 3:
            msgerro = msgerro + newline + msgerro13
            qtderr += 1
        else:
            msgerro = msgerro13 
            qtderr = 1
        errototald = errototald + msgerro13 + newline          

    if len(msgerro) > 3:
        planilha.row_dimensions[numLinha].height = qtderr * altura
        formatarCelulaDetalhe(planilha.cell(numLinha, 14,msgerro) ,"ERRO", 'CH')    
        planilha.cell(numLinha,  14).alignment = Alignment(wrap_text = True) 
        strglobal = "_ERRO.xlsx"
       
    numLinha += 4

    if ( (errototalp != "") or (errototalr != "") or (errototald != "") ):
        ERRO = ERRO + ("-"* 100) + newline
        ERRO = ERRO + "===> ERROS DE VALIDACAO - PLANILHA (TOTALIZADORES PROTOCOLADO E ATUAL_TI - ITEM DE NOTA FISCAL)\n\n"
        if (errototalp != ""):  
            ERRO = ERRO + "PROTOCOLADO:\n"
            ERRO = ERRO + errototalp + "\n"
        if (errototalr != ""):  
            ERRO = ERRO + "ATUAL_TI:\n"
            ERRO = ERRO + errototalr + "\n"
        if (errototald != ""):  
            ERRO = ERRO + "DIFERENÇA:\n"
            ERRO = ERRO + errototald + "\n"
    
####  - MESTRE DE NOTA FISCAL
####  - MESTRE DE NOTA FISCAL
    mqtd_nfs_reg = mqtd_nfs_prot = 0
    mvalor_total_reg = mvalor_total_prot = 0.00

    formatarCelulaTituloMasterA(planilha.cell(numLinha,  1, "MESTRE DE NOTA FISCAL"))
    formatarCelulaTituloMasterP(planilha.cell(numLinha,  14, "VALIDAÇÃO")) 
    planilha.cell(numLinha,  1).border = bordaB
    planilha.cell(numLinha,  14).border = bordaB    
    planilha.merge_cells('A'+str(numLinha)+':M'+str(numLinha))    
    numLinha += 1
    formatarCelulaTitulo(planilha.cell(numLinha,  1, "EMPS_COD"))
    formatarCelulaTitulo(planilha.cell(numLinha,  2, "FILI_COD"))
    formatarCelulaTitulo(planilha.cell(numLinha,  3, "MES_ANO"))
    formatarCelulaTitulo(planilha.cell(numLinha,  4, "SERIE"))
    formatarCelulaTitulo(planilha.cell(numLinha,  5, "QTD. NFS"))
    formatarCelulaTitulo(planilha.cell(numLinha,  6, "VALOR_TOTAL"))
    formatarCelulaTitulo(planilha.cell(numLinha,  7, "BASE_ICMS")) # PTITES-866
    formatarCelulaTitulo(planilha.cell(numLinha,  8, "VALOR_ICMS")) # PTITES-866
    formatarCelulaTitulo(planilha.cell(numLinha,  9, "ISENTAS_ICMS")) # PTITES-866
    formatarCelulaTitulo(planilha.cell(numLinha,  10, "OUTROS_VALORES")) # PTITES-866
    formatarCelulaTitulo(planilha.cell(numLinha,  11, "VALOR_CONTABIL")) # PTITES-866
    formatarCelulaTitulo(planilha.cell(numLinha,  12, "AREA")) # PTITES-866
    planilha.cell(numLinha, 12).alignment = Alignment(horizontal='center') # PTITES-866
    formatarCelulaTitulo(planilha.cell(numLinha,  13, "")) # PTITES-866
    formatarCelulaTitulo(planilha.cell(numLinha,  14, "")) # PTITES-866
    planilha.merge_cells('L'+str(numLinha)+':M'+str(numLinha)) # PTITES-866
    numLinha += 1

    temr = False
    temp = False
    
    for bq in abaBQ:
        if ("PROTOCOLADO" in bq.AREA): # PTITES-1367
            temp = True
        if (bq.AREA == "ATUAL_TI"):
            temr = True

    if temp == False:
        formatarCelulaDetalhe(planilha.cell(numLinha,  1, "")            ,"PROTOCOLADO", 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  2, "")            ,"PROTOCOLADO", 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  3, "")            ,"PROTOCOLADO", 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  4, "")            ,"PROTOCOLADO", 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  5, 0 )            ,"PROTOCOLADO", 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  6, 0.00)          ,"PROTOCOLADO", 'VL')
        formatarCelulaDetalhe(planilha.cell(numLinha,  7, 0.00)          ,"PROTOCOLADO", 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  8, 0.00)          ,"PROTOCOLADO", 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  9, 0.00)          ,"PROTOCOLADO", 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  10, 0.00)          ,"PROTOCOLADO", 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  11, 0.00)          ,"PROTOCOLADO", 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  12, configuracoes.origem_protocolado) ,"PROTOCOLADO", 'CH') # PTITES-866 # PTITES-1367
        planilha.cell(numLinha,  12).alignment = Alignment(horizontal='center') # PTITES-866
        planilha.merge_cells('L'+str(numLinha)+':M'+str(numLinha)) # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha, 14, "")            ,"PROTOCOLADO", 'CH')
        numLinhap = numLinha
        numLinha += 1

    for bq in abaBQ:                            
        formatarCelulaDetalhe(planilha.cell(numLinha,  1, bq.EMPS_COD)                     ,bq.AREA, 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  2, bq.FILI_COD)                     ,bq.AREA, 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  3, bq.MES_ANO)                      ,bq.AREA, 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  4, bq.SERIE)                        ,bq.AREA, 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  5, bq.QTD_NFS)                      ,bq.AREA, 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  6, bq.VALOR_TOTAL_MESTRE / 100 )          ,bq.AREA, 'VL') #***
        formatarCelulaDetalhe(planilha.cell(numLinha,  7, bq.BASE_ICMS_MESTRE / 100 )          ,bq.AREA, 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  8, bq.VALOR_ICMS_MESTRE / 100 )          ,bq.AREA, 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  9, bq.ISENTAS_ICMS_MESTRE / 100 )          ,bq.AREA, 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  10, bq.OUTROS_VALORES_MESTRE / 100 )          ,bq.AREA, 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  11, bq.VALOR_TOTAL_MESTRE / 100 )          ,bq.AREA, 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  12, bq.AREA)                         ,bq.AREA, 'CH') # PTITES-866
        planilha.cell(numLinha,  12).alignment = Alignment(horizontal='center') # PTITES-866
        planilha.merge_cells('L'+str(numLinha)+':M'+str(numLinha)) # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha, 14, "")                              ,bq.AREA, 'CH')
        if ("PROTOCOLADO" in bq.AREA): # PTITES-1367
            mqtd_nfs_prot = mqtd_nfs_prot + bq.QTD_NFS
            mvalor_total_prot = mvalor_total_prot + bq.VALOR_TOTAL_MESTRE
            numLinhap = numLinha
        if (bq.AREA == "ATUAL_TI"):
            mqtd_nfs_reg = mqtd_nfs_reg + bq.QTD_NFS
            mvalor_total_reg = mvalor_total_reg + bq.VALOR_TOTAL_MESTRE
            numLinhar = numLinha 
        numLinha += 1            
     
    if temr == False:
        formatarCelulaDetalhe(planilha.cell(numLinha,  1, "")            ,"ATUAL_TI", 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  2, "")            ,"ATUAL_TI", 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  3, "")            ,"ATUAL_TI", 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  4, "")            ,"ATUAL_TI", 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  5, 0 )            ,"ATUAL_TI", 'CH')
        formatarCelulaDetalhe(planilha.cell(numLinha,  6, 0.00)          ,"ATUAL_TI", 'VL')
        formatarCelulaDetalhe(planilha.cell(numLinha,  7, 0.00)          ,"ATUAL_TI", 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  8, 0.00)          ,"ATUAL_TI", 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  9, 0.00)          ,"ATUAL_TI", 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  10, 0.00)         ,"ATUAL_TI", 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  11, 0.00)         ,"ATUAL_TI", 'VL') # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha,  12, "ATUAL_TI")   ,"ATUAL_TI", 'CH') # PTITES-866
        planilha.cell(numLinha,  12).alignment = Alignment(horizontal='center') # PTITES-866
        planilha.merge_cells('L'+str(numLinha)+':M'+str(numLinha)) # PTITES-866
        formatarCelulaDetalhe(planilha.cell(numLinha, 14, "")            ,"ATUAL_TI", 'CH')
        numLinhar = numLinha
        numLinha += 1
        
    msgerro = "" 
    altura = 15
    qtderr = 0
     
    diferenca = abs(round( totir_valor_total - totir_desconto , 2) - round(mvalor_total_reg , 2)) /100 #***

    if ( diferenca > 0.05 ):
        if len(msgerro) > 3:
            qtderr += 1
            msgerro = msgerro + newline + msgerro9
        else:
            msgerro = msgerro9 
            qtderr = 1
        erromestrer = erromestrer + msgerro9 + newline            


    if (mqtd_nfs_reg != mqtd_nfs_prot):
        if len(msgerro) > 3:
            qtderr += 1
            msgerro = msgerro + newline + msgerro10
        else:
            msgerro = msgerro10
            qtderr = 1
        erromestrer = erromestrer + msgerro10 + newline            

    if len(msgerro) > 3:
        planilha.row_dimensions[numLinhar].height = qtderr * altura
        formatarCelulaDetalhe(planilha.cell(numLinhar, 14,msgerro) ,"ERRO", 'CH')  
        planilha.cell(numLinhar,  14).alignment = Alignment(wrap_text = True) 
        strglobal = "_ERRO.xlsx"
    msgerro = "" 
    altura = 15
    qtderr = 0

    if (mqtd_nfs_reg != mqtd_nfs_prot):
        if len(msgerro) > 3:
            qtderr += 1
            msgerro = msgerro + newline + msgerro11
        else:
            msgerro = msgerro11
            qtderr = 1
        erromestrep = erromestrep + msgerro11 + newline   
    diferenca = abs(round( totip_valor_contabil, 2) - round(mvalor_total_prot, 2)) /100 #***
    if ( diferenca > 0.05 ):
        if len(msgerro) > 3:
            qtderr += 1
            msgerro = msgerro + newline + msgerro12
        else:
            msgerro = msgerro12 
            qtderr = 1
        erromestrep = erromestrep + msgerro12 + newline  
    if len(msgerro) > 3:
        planilha.row_dimensions[numLinhap].height = qtderr * altura
        formatarCelulaDetalhe(planilha.cell(numLinhap, 14,msgerro) ,"ERRO", 'CH')
        planilha.cell(numLinhap,  14).alignment = Alignment(wrap_text = True) 
        strglobal = "_ERRO.xlsx"
    if ( (erromestrep != "") or (erromestrer != "") ):
        ERRO = ERRO + ("-"* 100) + newline
        ERRO = ERRO + "===> ERROS DE VALIDACAO - PLANILHA (MESTRE DE NOTA FISCAL)\n\n"
        if (erromestrep != ""):  
            ERRO = ERRO + "PROTOCOLADO:\n"
            ERRO = ERRO + erromestrep + "\n"
        if (erromestrer != ""):  
            ERRO = ERRO + "ATUAL_TI:\n"
            ERRO = ERRO + erromestrer + "\n"
        ERRO = ERRO + ("-"* 100) + newline            
    numLinha += 3    

#### - CHECKLIST CADASTRO
#### - CHECKLIST CADASTRO
    log("- Aba Checklist Cadastro... ")
    planilha = workbook.create_sheet("Checklist_Cadastro", 3)
    formatarCelulaTitulo(planilha.cell(1, 2, "Ultimo_protocolado"))
    planilha.merge_cells('B1:C1')
    formatarCelulaTitulo(planilha.cell(1, 4, "Atual_ti"))
    planilha.merge_cells('D1:E1')
    formatarCelulaTitulo(planilha.cell(2, 1, "Critério de validação"))
    formatarCelulaTitulo(planilha.cell(2, 2, "Qtd."))
    formatarCelulaTitulo(planilha.cell(2, 3, "Percentual"))
    formatarCelulaTitulo(planilha.cell(2, 4, "Qtd."))
    formatarCelulaTitulo(planilha.cell(2, 5, "Percentual"))
    formatarCelulaTitulo(planilha.cell(2, 6, "Valor Total da NF" ))
    formatarCelulaTitulo(planilha.cell(2, 7, "Base ICMS" ))
    formatarCelulaTitulo(planilha.cell(2, 8, "Valor ICMS" ))
    formatarCelulaTitulo(planilha.cell(2, 9, "Descrição" ))
    planilha.freeze_panes = planilha["A2"]
    numLinha = 2
    totalDados = {}
    for cc in abaCC:
        totalDados[cc.AREA] = totalDados.get(cc.AREA, {})
        totalDados[cc.AREA]['QTD_NFS'] = totalDados[cc.AREA].get('QTD_NFS', 0 ) + cc.QTD_NFS
        totalDados[cc.AREA]['QTD_NFS_TOTAL'] = totalDados[cc.AREA].get('QTD_NFS_TOTAL', 0 ) + cc.QTD_NFS_TOTAL
        totalDados[cc.AREA]['TOTAL_NOME_CLIENTE'] = totalDados[cc.AREA].get('TOTAL_NOME_CLIENTE', 0 ) + cc.TOTAL_NOME_CLIENTE
        totalDados[cc.AREA]['TOTAL_NOME_CLIENTE_INV'] = totalDados[cc.AREA].get('TOTAL_NOME_CLIENTE_INV', 0 ) + cc.TOTAL_NOME_CLIENTE_INV
        totalDados[cc.AREA]['TOTAL_CPF_CNPJ'] = totalDados[cc.AREA].get('TOTAL_CPF_CNPJ', 0 ) + cc.TOTAL_CPF_CNPJ
        totalDados[cc.AREA]['TOTAL_CPF_CNPJ_INV'] = totalDados[cc.AREA].get('TOTAL_CPF_CNPJ_INV', 0 ) + cc.TOTAL_CPF_CNPJ_INV
        totalDados[cc.AREA]['TOTAL_TERMINAL'] = totalDados[cc.AREA].get('TOTAL_TERMINAL', 0 ) + cc.TOTAL_TERMINAL
        totalDados[cc.AREA]['TOTAL_TERMINAL_INV'] = totalDados[cc.AREA].get('TOTAL_TERMINAL_INV', 0 ) + cc.TOTAL_TERMINAL_INV
        totalDados[cc.AREA]['TOTAL_TEL_CONTATO'] = totalDados[cc.AREA].get('TOTAL_TEL_CONTATO', 0 ) + cc.TOTAL_TEL_CONTATO
        totalDados[cc.AREA]['TOTAL_TEL_CONTATO_INV'] = totalDados[cc.AREA].get('TOTAL_TEL_CONTATO_INV', 0 ) + cc.TOTAL_TEL_CONTATO_INV
        totalDados[cc.AREA]['TOTAL_ENDERECO_CLIENTE'] = totalDados[cc.AREA].get('TOTAL_ENDERECO_CLIENTE', 0 ) + cc.TOTAL_ENDERECO_CLIENTE
        totalDados[cc.AREA]['TOTAL_ENDERECO_CLIENTE_INV'] = totalDados[cc.AREA].get('TOTAL_ENDERECO_CLIENTE_INV', 0 ) + cc.TOTAL_ENDERECO_CLIENTE_INV
        totalDados[cc.AREA]['TOTAL_INSCRICAO_INV'] = totalDados[cc.AREA].get('TOTAL_INSCRICAO_INV', 0 ) + cc.TOTAL_INSCRICAO_INV        
        totalDados[cc.AREA]['TOTAL_NOME_CLIENTE_VALOR_TOTAL_NF'] = totalDados[cc.AREA].get('TOTAL_NOME_CLIENTE_VALOR_TOTAL_NF', 0 ) + cc.TOTAL_NOME_CLIENTE_VALOR_TOTAL_NF
        totalDados[cc.AREA]['TOTAL_NOME_CLIENTE_INV_VALOR_TOTAL_NF'] = totalDados[cc.AREA].get('TOTAL_NOME_CLIENTE_INV_VALOR_TOTAL_NF', 0 ) + cc.TOTAL_NOME_CLIENTE_INV_VALOR_TOTAL_NF
        totalDados[cc.AREA]['TOTAL_CPF_CNPJ_VALOR_TOTAL_NF'] = totalDados[cc.AREA].get('TOTAL_CPF_CNPJ_VALOR_TOTAL_NF', 0 ) + cc.TOTAL_CPF_CNPJ_VALOR_TOTAL_NF
        totalDados[cc.AREA]['TOTAL_CPF_CNPJ_INV_VALOR_TOTAL_NF'] = totalDados[cc.AREA].get('TOTAL_CPF_CNPJ_INV_VALOR_TOTAL_NF', 0 ) + cc.TOTAL_CPF_CNPJ_INV_VALOR_TOTAL_NF
        totalDados[cc.AREA]['TOTAL_TERMINAL_VALOR_TOTAL_NF'] = totalDados[cc.AREA].get('TOTAL_TERMINAL_VALOR_TOTAL_NF', 0 ) + cc.TOTAL_TERMINAL_VALOR_TOTAL_NF
        totalDados[cc.AREA]['TOTAL_TERMINAL_INV_VALOR_TOTAL_NF'] = totalDados[cc.AREA].get('TOTAL_TERMINAL_INV_VALOR_TOTAL_NF', 0 ) + cc.TOTAL_TERMINAL_INV_VALOR_TOTAL_NF
        totalDados[cc.AREA]['TOTAL_TEL_CONTATO_VALOR_TOTAL_NF'] = totalDados[cc.AREA].get('TOTAL_TEL_CONTATO_VALOR_TOTAL_NF', 0 ) + cc.TOTAL_TEL_CONTATO_VALOR_TOTAL_NF
        totalDados[cc.AREA]['TOTAL_TEL_CONTATO_INV_VALOR_TOTAL_NF'] = totalDados[cc.AREA].get('TOTAL_TEL_CONTATO_INV_VALOR_TOTAL_NF', 0 ) + cc.TOTAL_TEL_CONTATO_INV_VALOR_TOTAL_NF
        totalDados[cc.AREA]['TOTAL_ENDERECO_CLIENTE_VALOR_TOTAL_NF'] = totalDados[cc.AREA].get('TOTAL_ENDERECO_CLIENTE_VALOR_TOTAL_NF', 0 ) + cc.TOTAL_ENDERECO_CLIENTE_VALOR_TOTAL_NF
        totalDados[cc.AREA]['TOTAL_ENDERECO_CLIENTE_INV_VALOR_TOTAL_NF'] = totalDados[cc.AREA].get('TOTAL_ENDERECO_CLIENTE_INV_VALOR_TOTAL_NF', 0 ) + cc.TOTAL_ENDERECO_CLIENTE_INV_VALOR_TOTAL_NF
        totalDados[cc.AREA]['TOTAL_INSCRICAO_INV_VALOR_TOTAL_NF'] = totalDados[cc.AREA].get('TOTAL_INSCRICAO_INV_VALOR_TOTAL_NF', 0 ) + cc.TOTAL_INSCRICAO_INV_VALOR_TOTAL_NF
        totalDados[cc.AREA]['TOTAL_NOME_CLIENTE_BASE_ICMS'] = totalDados[cc.AREA].get('TOTAL_NOME_CLIENTE_BASE_ICMS', 0 ) + cc.TOTAL_NOME_CLIENTE_BASE_ICMS
        totalDados[cc.AREA]['TOTAL_NOME_CLIENTE_INV_BASE_ICMS'] = totalDados[cc.AREA].get('TOTAL_NOME_CLIENTE_INV_BASE_ICMS', 0 ) + cc.TOTAL_NOME_CLIENTE_INV_BASE_ICMS
        totalDados[cc.AREA]['TOTAL_CPF_CNPJ_BASE_ICMS'] = totalDados[cc.AREA].get('TOTAL_CPF_CNPJ_BASE_ICMS', 0 ) + cc.TOTAL_CPF_CNPJ_BASE_ICMS
        totalDados[cc.AREA]['TOTAL_CPF_CNPJ_INV_BASE_ICMS'] = totalDados[cc.AREA].get('TOTAL_CPF_CNPJ_INV_BASE_ICMS', 0 ) + cc.TOTAL_CPF_CNPJ_INV_BASE_ICMS
        totalDados[cc.AREA]['TOTAL_TERMINAL_BASE_ICMS'] = totalDados[cc.AREA].get('TOTAL_TERMINAL_BASE_ICMS', 0 ) + cc.TOTAL_TERMINAL_BASE_ICMS
        totalDados[cc.AREA]['TOTAL_TERMINAL_INV_BASE_ICMS'] = totalDados[cc.AREA].get('TOTAL_TERMINAL_INV_BASE_ICMS', 0 ) + cc.TOTAL_TERMINAL_INV_BASE_ICMS
        totalDados[cc.AREA]['TOTAL_TEL_CONTATO_BASE_ICMS'] = totalDados[cc.AREA].get('TOTAL_TEL_CONTATO_BASE_ICMS', 0 ) + cc.TOTAL_TEL_CONTATO_BASE_ICMS
        totalDados[cc.AREA]['TOTAL_TEL_CONTATO_INV_BASE_ICMS'] = totalDados[cc.AREA].get('TOTAL_TEL_CONTATO_INV_BASE_ICMS', 0 ) + cc.TOTAL_TEL_CONTATO_INV_BASE_ICMS
        totalDados[cc.AREA]['TOTAL_ENDERECO_CLIENTE_BASE_ICMS'] = totalDados[cc.AREA].get('TOTAL_ENDERECO_CLIENTE_BASE_ICMS', 0 ) + cc.TOTAL_ENDERECO_CLIENTE_BASE_ICMS
        totalDados[cc.AREA]['TOTAL_ENDERECO_CLIENTE_INV_BASE_ICMS'] = totalDados[cc.AREA].get('TOTAL_ENDERECO_CLIENTE_INV_BASE_ICMS', 0 ) + cc.TOTAL_ENDERECO_CLIENTE_INV_BASE_ICMS
        totalDados[cc.AREA]['TOTAL_INSCRICAO_INV_BASE_ICMS'] = totalDados[cc.AREA].get('TOTAL_INSCRICAO_INV_BASE_ICMS', 0 ) + cc.TOTAL_INSCRICAO_INV_BASE_ICMS
        totalDados[cc.AREA]['TOTAL_NOME_CLIENTE_VALOR_ICMS'] = totalDados[cc.AREA].get('TOTAL_NOME_CLIENTE_VALOR_ICMS', 0 ) + cc.TOTAL_NOME_CLIENTE_VALOR_ICMS
        totalDados[cc.AREA]['TOTAL_NOME_CLIENTE_INV_VALOR_ICMS'] = totalDados[cc.AREA].get('TOTAL_NOME_CLIENTE_INV_VALOR_ICMS', 0 ) + cc.TOTAL_NOME_CLIENTE_INV_VALOR_ICMS
        totalDados[cc.AREA]['TOTAL_CPF_CNPJ_VALOR_ICMS'] = totalDados[cc.AREA].get('TOTAL_CPF_CNPJ_VALOR_ICMS', 0 ) + cc.TOTAL_CPF_CNPJ_VALOR_ICMS
        totalDados[cc.AREA]['TOTAL_CPF_CNPJ_INV_VALOR_ICMS'] = totalDados[cc.AREA].get('TOTAL_CPF_CNPJ_INV_VALOR_ICMS', 0 ) + cc.TOTAL_CPF_CNPJ_INV_VALOR_ICMS
        totalDados[cc.AREA]['TOTAL_TERMINAL_VALOR_ICMS'] = totalDados[cc.AREA].get('TOTAL_TERMINAL_VALOR_ICMS', 0 ) + cc.TOTAL_TERMINAL_VALOR_ICMS
        totalDados[cc.AREA]['TOTAL_TERMINAL_INV_VALOR_ICMS'] = totalDados[cc.AREA].get('TOTAL_TERMINAL_INV_VALOR_ICMS', 0 ) + cc.TOTAL_TERMINAL_INV_VALOR_ICMS
        totalDados[cc.AREA]['TOTAL_TEL_CONTATO_VALOR_ICMS'] = totalDados[cc.AREA].get('TOTAL_TEL_CONTATO_VALOR_ICMS', 0 ) + cc.TOTAL_TEL_CONTATO_VALOR_ICMS
        totalDados[cc.AREA]['TOTAL_TEL_CONTATO_INV_VALOR_ICMS'] = totalDados[cc.AREA].get('TOTAL_TEL_CONTATO_INV_VALOR_ICMS', 0 ) + cc.TOTAL_TEL_CONTATO_INV_VALOR_ICMS
        totalDados[cc.AREA]['TOTAL_ENDERECO_CLIENTE_VALOR_ICMS'] = totalDados[cc.AREA].get('TOTAL_ENDERECO_CLIENTE_VALOR_ICMS', 0 ) + cc.TOTAL_ENDERECO_CLIENTE_VALOR_ICMS
        totalDados[cc.AREA]['TOTAL_ENDERECO_CLIENTE_INV_VALOR_ICMS'] = totalDados[cc.AREA].get('TOTAL_ENDERECO_CLIENTE_INV_VALOR_ICMS', 0 ) + cc.TOTAL_ENDERECO_CLIENTE_INV_VALOR_ICMS
        totalDados[cc.AREA]['TOTAL_INSCRICAO_INV_VALOR_ICMS'] = totalDados[cc.AREA].get('TOTAL_INSCRICAO_INV_VALOR_ICMS', 0 ) + cc.TOTAL_INSCRICAO_INV_VALOR_ICMS
    formatarCelulaTitulo(planilha.cell(3,  1, 'Nome Cliente Generico'))
    formatarCelulaTitulo(planilha.cell(4,  1, 'Nome Cliente Inválido'))
    formatarCelulaTitulo(planilha.cell(5,  1, 'CPF_CNPJ Generico'))
    formatarCelulaTitulo(planilha.cell(6,  1, 'CPF_CNPJ Inválido'))
    formatarCelulaTitulo(planilha.cell(7,  1, 'Terminal Generico'))
    formatarCelulaTitulo(planilha.cell(8,  1, 'Terminal Inválido'))
    formatarCelulaTitulo(planilha.cell(9,  1, 'Tel. contato Generico'))
    formatarCelulaTitulo(planilha.cell(10, 1, 'Tel. contato Inválido'))
    formatarCelulaTitulo(planilha.cell(11, 1, 'Endereço cliente Generico'))
    formatarCelulaTitulo(planilha.cell(12, 1, 'Endereço cliente Inválido'))
    formatarCelulaTitulo(planilha.cell(13, 1, 'Inscrição Estadual Inválida'))
    formatarCelulaDetalhe(planilha.cell(3,  9, 'Onde o nome do cliente genérico é igual a "CONSUMIDOR"'),'','CH')
    formatarCelulaDetalhe(planilha.cell(4,  9, 'Onde o campo Nome Cliente não está preenchido, ou tenha menos que 4 caracteres, ou possua caracteres inadequados (“...”, ”---”, ”|”, ”\”, ”\”), ou seja iniciado em 0 (zero) ou seja um numérico.'),'','CH')
    formatarCelulaDetalhe(planilha.cell(5,  9, 'Onde CPF genérico (11 dígitos) é igual a "11111111111" ou CNPJ (14 dígitos) igual a "11111111111111"'),'','CH')
    formatarCelulaDetalhe(planilha.cell(6,  9, 'Onde o CPF_CNPJ não é genérico e está preenchido indevidamente (tamanho inválido, não é numérico, digito verificador errado ou está vazio).'),'','CH')
    formatarCelulaDetalhe(planilha.cell(7,  9, 'Onde terminal genérico é igual a "1135497777"'),'','CH')
    formatarCelulaDetalhe(planilha.cell(8,  9, 'Onde o Terminal é inválido (não é numérico ou possui DDD inválido ou possui tamanho diferente de 10 e 11 dígitos ou não foi informado para uma serie Telecom).'),'','CH')
    formatarCelulaDetalhe(planilha.cell(9,  9, 'Onde Telefone de Contato genérico é igual a "1135497777"'),'','CH')
    formatarCelulaDetalhe(planilha.cell(10, 9, 'Onde o Tel de contato é inválido (não é numérico ou possui DDD inexistente ou possui tamanho diferente de 10 e 11 dígitos ou não foi informado para uma serie Telecom).'),'','CH')
    formatarCelulaDetalhe(planilha.cell(11, 9, 'Onde o endereço genérico for igual ao da Vivo da Capital(Endereços-Vivo) com base na UF da emissão da NF.'),'','CH')
    formatarCelulaDetalhe(planilha.cell(12, 9, 'Onde o endereço é vazio ou inválido (iniciado com zero ou espaço ou ponto, está preenchido apenas com a palavra ‘AVENIDA’, é inteiramente numérico, contém a sequência “---”, tenha menos que 4 caracteres)'),'','CH')
    formatarCelulaDetalhe(planilha.cell(13, 9, 'Onde a Inscrição Estadual está invalida para o Estado do cliente.'),'','CH')
    
    linha = 3

    for area in totalDados.keys() :
        if 'PROTOCOLADO' in area.upper(): # PTITES-1367
            col_inicial = 2
        else :
            col_inicial = 4        
        linha = 3
        formatarCelulaDetalhe(planilha.cell(linha,  col_inicial, totalDados[area]['TOTAL_NOME_CLIENTE'])           ,area, 'NU')
        if 'PROTOCOLADO' not in area.upper() :  # PTITES-1367
            formatarCelulaDetalhe(planilha.cell(linha,  6, totalDados[area]['TOTAL_NOME_CLIENTE_VALOR_TOTAL_NF']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  7, totalDados[area]['TOTAL_NOME_CLIENTE_BASE_ICMS']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  8, totalDados[area]['TOTAL_NOME_CLIENTE_VALOR_ICMS']/100)               , area, 'VL')
        if totalDados[area]['QTD_NFS_TOTAL'] > 0 :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, totalDados[area]['TOTAL_NOME_CLIENTE'] / totalDados[area]['QTD_NFS_TOTAL'] )           ,area, 'PC')
        else :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, 0 )           ,area, 'PC')
        linha += 1
        formatarCelulaDetalhe(planilha.cell(linha,  col_inicial, totalDados[area]['TOTAL_NOME_CLIENTE_INV'])           ,area, 'NU')
        if 'PROTOCOLADO' not in area.upper() :  # PTITES-1367
            formatarCelulaDetalhe(planilha.cell(linha,  6, totalDados[area]['TOTAL_NOME_CLIENTE_INV_VALOR_TOTAL_NF']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  7, totalDados[area]['TOTAL_NOME_CLIENTE_INV_BASE_ICMS']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  8, totalDados[area]['TOTAL_NOME_CLIENTE_INV_VALOR_ICMS']/100)               , area, 'VL')
        if totalDados[area]['QTD_NFS_TOTAL'] > 0 :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, totalDados[area]['TOTAL_NOME_CLIENTE_INV'] / totalDados[area]['QTD_NFS_TOTAL'] )           ,area, 'PC')
        else :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, 0 )           ,area, 'PC')
        linha += 1
        formatarCelulaDetalhe(planilha.cell(linha,  col_inicial, totalDados[area]['TOTAL_CPF_CNPJ'])           ,area, 'NU')
        if 'PROTOCOLADO' not in area.upper() :  # PTITES-1367
            formatarCelulaDetalhe(planilha.cell(linha,  6, totalDados[area]['TOTAL_CPF_CNPJ_VALOR_TOTAL_NF']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  7, totalDados[area]['TOTAL_CPF_CNPJ_BASE_ICMS']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  8, totalDados[area]['TOTAL_CPF_CNPJ_VALOR_ICMS']/100)               , area, 'VL')
        if totalDados[area]['QTD_NFS_TOTAL'] > 0 :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, totalDados[area]['TOTAL_CPF_CNPJ'] / totalDados[area]['QTD_NFS_TOTAL'] )           ,area, 'PC')
        else :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, 0 )           ,area, 'PC')
        linha += 1
        formatarCelulaDetalhe(planilha.cell(linha,  col_inicial, totalDados[area]['TOTAL_CPF_CNPJ_INV'])           ,area, 'NU')
        if 'PROTOCOLADO' not in area.upper() :  # PTITES-1367
            formatarCelulaDetalhe(planilha.cell(linha,  6, totalDados[area]['TOTAL_CPF_CNPJ_INV_VALOR_TOTAL_NF']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  7, totalDados[area]['TOTAL_CPF_CNPJ_INV_BASE_ICMS']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  8, totalDados[area]['TOTAL_CPF_CNPJ_INV_VALOR_ICMS']/100)               , area, 'VL')
        if totalDados[area]['QTD_NFS_TOTAL'] > 0 :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, totalDados[area]['TOTAL_CPF_CNPJ_INV'] / totalDados[area]['QTD_NFS_TOTAL'] )           ,area, 'PC')
        else :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, 0 )           ,area, 'PC')
        linha += 1
        formatarCelulaDetalhe(planilha.cell(linha,  col_inicial, totalDados[area]['TOTAL_TERMINAL'])           ,area, 'NU')
        if 'PROTOCOLADO' not in area.upper() :  # PTITES-1367
            formatarCelulaDetalhe(planilha.cell(linha,  6, totalDados[area]['TOTAL_TERMINAL_VALOR_TOTAL_NF']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  7, totalDados[area]['TOTAL_TERMINAL_BASE_ICMS']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  8, totalDados[area]['TOTAL_TERMINAL_VALOR_ICMS']/100)               , area, 'VL')
        if totalDados[area]['QTD_NFS_TOTAL'] > 0 :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, totalDados[area]['TOTAL_TERMINAL'] / totalDados[area]['QTD_NFS_TOTAL'] )           ,area, 'PC')
        else :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, 0 )           ,area, 'PC')
        linha += 1
        formatarCelulaDetalhe(planilha.cell(linha,  col_inicial, totalDados[area]['TOTAL_TERMINAL_INV'])           ,area, 'NU')
        if 'PROTOCOLADO' not in area.upper() :  # PTITES-1367
            formatarCelulaDetalhe(planilha.cell(linha,  6, totalDados[area]['TOTAL_TERMINAL_INV_VALOR_TOTAL_NF']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  7, totalDados[area]['TOTAL_TERMINAL_INV_BASE_ICMS']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  8, totalDados[area]['TOTAL_TERMINAL_INV_VALOR_ICMS']/100)               , area, 'VL')
        if totalDados[area]['QTD_NFS_TOTAL'] > 0 :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, totalDados[area]['TOTAL_TERMINAL_INV'] / totalDados[area]['QTD_NFS_TOTAL'] )           ,area, 'PC')
        else :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, 0 )           ,area, 'PC')
        linha += 1
        formatarCelulaDetalhe(planilha.cell(linha,  col_inicial, totalDados[area]['TOTAL_TEL_CONTATO'])           ,area, 'NU')
        if 'PROTOCOLADO' not in area.upper() :  # PTITES-1367
            formatarCelulaDetalhe(planilha.cell(linha,  6, totalDados[area]['TOTAL_TEL_CONTATO_VALOR_TOTAL_NF']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  7, totalDados[area]['TOTAL_TEL_CONTATO_BASE_ICMS']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  8, totalDados[area]['TOTAL_TEL_CONTATO_VALOR_ICMS']/100)               , area, 'VL')
        if totalDados[area]['QTD_NFS_TOTAL'] > 0 :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, totalDados[area]['TOTAL_TEL_CONTATO'] / totalDados[area]['QTD_NFS_TOTAL'] )           ,area, 'PC')
        else :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, 0 )           ,area, 'PC')
        linha += 1
        formatarCelulaDetalhe(planilha.cell(linha,  col_inicial, totalDados[area]['TOTAL_TEL_CONTATO_INV'])           ,area, 'NU')
        if 'PROTOCOLADO' not in area.upper() :  # PTITES-1367
            formatarCelulaDetalhe(planilha.cell(linha,  6, totalDados[area]['TOTAL_TEL_CONTATO_INV_VALOR_TOTAL_NF']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  7, totalDados[area]['TOTAL_TEL_CONTATO_INV_BASE_ICMS']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  8, totalDados[area]['TOTAL_TEL_CONTATO_INV_VALOR_ICMS']/100)               , area, 'VL')
        if totalDados[area]['QTD_NFS_TOTAL'] > 0 :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, totalDados[area]['TOTAL_TEL_CONTATO_INV'] / totalDados[area]['QTD_NFS_TOTAL'] )           ,area, 'PC')
        else :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, 0 )           ,area, 'PC')
        linha += 1
        formatarCelulaDetalhe(planilha.cell(linha,  col_inicial, totalDados[area]['TOTAL_ENDERECO_CLIENTE'])           ,area, 'NU')
        if 'PROTOCOLADO' not in area.upper() :  # PTITES-1367
            formatarCelulaDetalhe(planilha.cell(linha,  6, totalDados[area]['TOTAL_ENDERECO_CLIENTE_VALOR_TOTAL_NF']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  7, totalDados[area]['TOTAL_ENDERECO_CLIENTE_BASE_ICMS']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  8, totalDados[area]['TOTAL_ENDERECO_CLIENTE_VALOR_ICMS']/100)               , area, 'VL')
        if totalDados[area]['QTD_NFS_TOTAL'] > 0 :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, totalDados[area]['TOTAL_ENDERECO_CLIENTE'] / totalDados[area]['QTD_NFS_TOTAL'] )           ,area, 'PC')
        else :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, 0 )           ,area, 'PC')
        linha += 1
        formatarCelulaDetalhe(planilha.cell(linha,  col_inicial, totalDados[area]['TOTAL_ENDERECO_CLIENTE_INV'])           ,area, 'NU')
        if 'PROTOCOLADO' not in area.upper() :  # PTITES-1367
            formatarCelulaDetalhe(planilha.cell(linha,  6, totalDados[area]['TOTAL_ENDERECO_CLIENTE_INV_VALOR_TOTAL_NF']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  7, totalDados[area]['TOTAL_ENDERECO_CLIENTE_INV_BASE_ICMS']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  8, totalDados[area]['TOTAL_ENDERECO_CLIENTE_INV_VALOR_ICMS']/100)               , area, 'VL')
        if totalDados[area]['QTD_NFS_TOTAL'] > 0 :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, totalDados[area]['TOTAL_ENDERECO_CLIENTE_INV'] / totalDados[area]['QTD_NFS_TOTAL'] )           ,area, 'PC')
        else :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, 0 )           ,area, 'PC')
        linha += 1
        formatarCelulaDetalhe(planilha.cell(linha,  col_inicial, totalDados[area]['TOTAL_INSCRICAO_INV'])           ,area, 'NU')
        if 'PROTOCOLADO' not in area.upper() :  # PTITES-1367
            formatarCelulaDetalhe(planilha.cell(linha,  6, totalDados[area]['TOTAL_INSCRICAO_INV_VALOR_TOTAL_NF']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  7, totalDados[area]['TOTAL_INSCRICAO_INV_BASE_ICMS']/100)               , area, 'VL')
            formatarCelulaDetalhe(planilha.cell(linha,  8, totalDados[area]['TOTAL_INSCRICAO_INV_VALOR_ICMS']/100)               , area, 'VL')
        if totalDados[area]['QTD_NFS_TOTAL'] > 0 :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, totalDados[area]['TOTAL_INSCRICAO_INV'] / totalDados[area]['QTD_NFS_TOTAL'] )           ,area, 'PC')
        else :
            formatarCelulaDetalhe(planilha.cell(linha,  col_inicial+1, 0 )           ,area, 'PC')
        
        if 'PROTOCOLADO' in area.upper(): # PTITES-1367
            dif_linha = 3
        else :
            dif_linha = 4

        linha += 1
        
        formatarCelulaTitulo(planilha.cell(linha + dif_linha,  1, area.capitalize()))
        formatarCelulaDetalhe(planilha.cell(linha + dif_linha,  2, totalDados[area]['QTD_NFS_TOTAL'] ) ,area, 'NU')
        formatarCelulaDetalhe(planilha.cell(linha + dif_linha,  4, totalDados[area]['QTD_NFS_TOTAL'] - totalDados[area]['QTD_NFS'] ) ,area, 'NU')
        formatarCelulaDetalhe(planilha.cell(linha + dif_linha,  3, totalDados[area]['QTD_NFS'] ) ,area, 'NU')
    
    linha += 2

    formatarCelulaTitulo(planilha.cell(linha, 2, 'Total de notas da Série'))
    formatarCelulaTitulo(planilha.cell(linha, 4, 'Total notas Não impactadas'))
    formatarCelulaTitulo(planilha.cell(linha, 3, 'Total de notas Impactadas'))
    planilha.column_dimensions['A'].width = 30
    planilha.column_dimensions['B'].width = 25
    planilha.column_dimensions['C'].width = 25
    planilha.column_dimensions['D'].width = 25
    planilha.column_dimensions['E'].width = 25
    planilha.column_dimensions['F'].width = 25
    planilha.column_dimensions['G'].width = 25
    planilha.column_dimensions['H'].width = 25
    planilha.column_dimensions['I'].width = 195
    return ERRO

def addTabelaValidacao(
                       emps_cod
                      ,fili_cod
                      ,mes_ano
                      ,serie
                      ,cfop
                      ,valor_total
                      ,base_icms
                      ,valor_icms
                      ,isentas_icms
                      ,outros_valores
                      ,desconto                      
                      ,valor_contabil
                      ,area
                      ,validacao
                      ):
    chaveRelatorio = emps_cod+":"+fili_cod+":"+str(mes_ano)+":"+serie+":"+cfop
    info_serie = None

    if chaveRelatorio in seriesRelatorio:
        info_serie = seriesRelatorio[chaveRelatorio]

    if not info_serie:
        info_serie = Serie(emps_cod, fili_cod, mes_ano, serie, cfop)
        seriesRelatorio[chaveRelatorio] = info_serie

    info_serie.areas[area] = LinhaRelatorio(valor_total, base_icms, valor_icms, isentas_icms, outros_valores, desconto, valor_contabil)

fontMasterPreta  = Font(color='00000000', bold=True, size=12)
fontMasterBranca = Font(color='FFFFFFFF', bold=True, size=12)
fontNegrito      = Font(color='00000000', bold=True)
fontAzul         = Font(color='FF0000FF')
fontVermelha     = Font(color='FFFF0000')
fontPreta        = Font(color='00000000')

borda = Border(
            left   = Side(style='thin'),
            right  = Side(style='thin'),
            top    = Side(style='thin'),
            bottom = Side(style='thin')
)

bordaB = Border(
            left   = Side(style='medium'),
            right  = Side(style='medium'),
            top    = Side(style='medium'),
            bottom = Side(style='medium')
)

fundoCelula = PatternFill(
                start_color = 'FFE0E0E0',
                end_color   = 'FFE0E0E0',
                fill_type   = 'solid'
)

fundoCelulaAzul = PatternFill(
                start_color = '0066CCCC',
                end_color   = '0066CCCC',
                fill_type   = 'solid'
)

fundoCelulaBranco = PatternFill(
                start_color = 'FFFFFFFF',
                end_color   = 'FFFFFFFF',
                fill_type   = 'solid'
)

fundoCelulaPreto = PatternFill(
                start_color = '00000000',
                end_color   = '00000000',
                fill_type   = 'solid'
)

def formatarCelulaTituloMasterA(celula):
    celula.font = fontMasterPreta
    celula.fill = fundoCelulaAzul
    celula.alignment = alignment = Alignment(horizontal='center') 
    celula.border = bordaB
    
def formatarCelulaTituloMasterP(celula):
    celula.font = fontMasterBranca
    celula.fill = fundoCelulaPreto  
    celula.alignment = alignment = Alignment(horizontal='center')
    celula.border = bordaB
    
def formatarCelulaTitulo(celula):
    celula.font = fontNegrito
    celula.border = borda

def formatarCelulaDetalhe(celula, area, dataType):
    if  'PROTOCOLADO' in area: # PTITES-1367
        celula.font = fontAzul
        celula.fill = fundoCelula
    else: 
        celula.font = fontPreta
        celula.fill = fundoCelulaBranco
    
    if area == 'ERRO':
        celula.font = fontVermelha
        
    if  dataType == 'MA':
        celula.number_format = "mmmm-yy"
    
    if  dataType == 'VL': 
        celula.number_format = "#,##0.00"
        if celula.value < 0.00:
            celula.font = fontVermelha
    
    if  dataType == 'PC': 
        celula.number_format = "#,##0.00%"
    
    if  dataType == 'NU': 
        celula.number_format = "0"
        
    celula.border = borda  

def adicionarLinhaResumo(serie, planilha, num_linha, resumo):
    planilha.cell(num_linha, 1, serie.emps_cod).border = borda
    planilha.cell(num_linha, 2, serie.fili_cod).border = borda

    celula = planilha.cell(num_linha, 3, serie.mes_ano)
    celula.border = borda
    celula.number_format = "mmmm-yy"

    planilha.cell(num_linha, 4, serie.serie).border = borda
    planilha.cell(num_linha, 5, serie.cfop).border = borda
    planilha.cell(num_linha, 6, resumo).border = borda

def semespacos(frase):
    retorno = "" 
    for l in frase:
        if (l != " "):
            retorno = retorno + l
    return(retorno)

if __name__ == "__main__":
    configuracoes.numero_erro = 0 # PTITES-1367
    configuracoes.descricao_erro = "" # PTITES-1367
    configuracoes.origem_protocolado  = "PROTOCOLADO" # PTITES-1367
    retornou = 0 
    retornou = processar()
    log("Retorno do processar = ", retornou)
    # PTITES-1367
    log(configuracoes.numero_erro)
    log(configuracoes.descricao_erro)
    if configuracoes.numero_erro:
        ret = 1
    else:
        if (retornou[1] == ""):
            ret = 0
        else:
            ret = 99
    log(ret)        
    # PTITES-1367        
    sys.exit(ret)
