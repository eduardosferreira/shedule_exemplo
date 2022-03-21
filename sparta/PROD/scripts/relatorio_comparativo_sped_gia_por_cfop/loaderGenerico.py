#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-
"""
----------------------------------------------------------------------------------------------
MODULO ...: TESHUVA
SCRIPT ...: loaderGenerico.py
CRIACAO ..: 24/03/2021
AUTOR ....: EDUARDO DA SILVA FERREIRA / KYROS TECNOLOGIA
            eduardof@kyros.com.br
DESCRICAO.: geração de um insumo onde sejam confrontados
os valores por CFOP do SPED e da GIA regerados 
e protocolados conforme o layout abaixo:
----------------------------------------------------------------------------------------------
PARAMETROS: 
Parâmetros de entrada:
1)	MESANO: Mês e ano no formato MMAAAA - Obrigatório
2)	UF: UF do estado - Obrigatório
3)	IE: Inscrição estadual - Obrigatório
----------------------------------------------------------------------------------------------
  HISTORICO : 
----------------------------------------------------------------------------------------------
"""

import collections
import datetime
import os
import cx_Oracle
import re
import sys
import shutil
import traceback
import string

# Nome do script
nome_script = os.path.basename( sys.argv[0] ).replace('.py', '')

# Lista de String
gv_lista_string = list(string.ascii_lowercase)

import comum
import sql
variaveis = {'teste': '001'}
comum.variaveis = variaveis
sql.variaveis = variaveis
from layout import *
from pathlib import Path
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook

log.gerar_log_em_arquivo = True

# Parametros Globais
global ret # variavel para controle de retorno de erro 
           # (!= 0 -> ERRO, 0 - SUCESSO)
global txt # Variavel para tratamento de saida
# Entrada
global gv_mes_ano
global gv_mes
global gv_ano
global gv_uf
global gv_ie
# Conexao
global gv_conexao
# Arquivo
global gv_usuario
global gv_senha
global gv_banco
global gv_diretorio_sped_fiscal_enxertado
global gv_diretorio_sped_fiscal_protocolado
global gv_diretorio_gia_enxertado
global gv_diretorio_gia_protocolado
global gv_diretorio_insumo_sped
global gv_arq_sped_fiscal_enxertado
global gv_arq_sped_fiscal_protocolado
global gv_arq_gia_enxertado
global gv_arq_gia_protocolado
global gv_arq_insumo_sped
global gv_lst_arq_sped_fiscal_enxertado
global gv_lst_arq_sped_fiscal_protocolado
global gv_lst_arq_gia_enxertado
global gv_lst_arq_gia_protocolado
global gv_lst_arq_insumo_sped

"""
Retorna lista de alfabeto 
"""
def listAlphabet():
  return list(map(chr, range(97, 123)))

"""
Retorna a data formata 
"""
def dtf():
    return (datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S'))

"""
Funcao para conectar na base de dados 
"""
def conectar_BancoDados():
    try:

        l_connection = cx_Oracle.connect(gv_usuario+"/"+gv_senha+"@"+gv_banco)
        v_SQL="""
        SELECT 1 FROM DUAL
        """
        l_cursor = l_connection.cursor()
        l_cursor.execute(v_SQL)
        l_result = l_cursor.fetchone()
        l_cursor.close()

        if(l_result != None):
            log("SUCESSO CONEXAO BANCO DE DADOS ") 
            return l_connection  
        else:
            log("ERRO CONEXAO BANCO DE DADOS ") 
            ret = 91
            return None


    except Exception as e:
        txt = traceback.format_exc()
        log("ERRO CONEXAO BANCO DE DADOS .: " + str(e) + " - TRACE - " + txt)
        return None

"""
Funcao para fechar Conexao Banco de Dados 
"""
def fechar_ConexaoBancoDados(p_connection):
    l_ret = 0
    
    try:
        p_connection.close()  
    except:
        l_ret = 0

    try:
        p_connection = None  
    except:
        l_ret = 0

    return l_ret

"""
Funcao para retornar UF 
"""
def retornar_UF(p_IE, p_connection):
    l_UF = ""
    l_query="""
    select distinct f.unfe_sig  from openrisow.filial f where f.fili_cod_insest='%s' and rownum < 2
    """%(p_IE)

    l_cursor = p_connection.cursor()
    l_cursor.execute(l_query)
    l_result = l_cursor.fetchone()

    if(l_result != None): 
        for campo in l_result:
            l_UF = campo

    l_cursor.close()
    return(l_UF)

"""
Funcao para retornar arquivos existentes
"""
def ultimo_Arquivo_Diretorio(p_arq_mascara,p_diretorio):
    l_qte = 0
    l_arquivo = ""
    l_diretorio = Path(p_diretorio)
    l_arq = l_diretorio.glob(p_arq_mascara)
    l_procura_arquivos = sorted(l_arq, reverse=False)
    if l_procura_arquivos:        
        for arquivo in l_procura_arquivos:
            # if l_qte == 0:
            #    log("# Arquivos encontrados: ")    
            l_qte = l_qte + 1
            # log("#   " + str(l_qte) + " => " + str(arquivo))
            l_arquivo = str(arquivo)

        if l_qte == 0:
            log('Arquivo ' + p_arq_mascara +' não está na pasta ' + p_diretorio)
            return ""
        else:
            return(l_arquivo)

    else: 
        log('Arquivo ' + p_arq_mascara +' não está na pasta ' + p_diretorio)
        return ""

"""
Retorna a definicao tipo de arquivo
"""
def tipoArquivo(path_arq) :
    try :
        fd = open(path_arq, 'r', encoding='iso-8859-1')
        fd.readline()
        fd.close()
    except :
        return 'utf-8'
    return 'iso-8859-1'

"""
Funcao processa arquivo do tipo SPED REGERADO , PROTOCOLADO
"""
def processa_SPED(l_arquivo):
    l_regrel=[[],[]]
    l_regrel[0]=["CFOP","VLR_OPER","VLR_BASE_ICMS","VLR_ICMS","TIPO"]

    # inicializa a variaveis de controle
    l_flsair = 0 
    l_contador = 0
    l_nregrel = 0   
    l_tp_reg = ""

    l_CFOP = ""
    l_VLR_OPER = float(0)
    l_VLR_BASE_ICMS = float(0)
    l_VLR_ICMS = float(0)
    
    if (len(l_arquivo) > 4 and os.path.isfile(l_arquivo)) :
        
        log("Processando leitura do arquivo : " + l_arquivo)
        
        # realiza a abertura do arquivo
        l_ent = open(l_arquivo,mode="r",encoding=tipoArquivo(l_arquivo))
                
        # realiza a leitura da linha do arquivo
        l_linhalida = l_ent.readline()
                
        # percorrre o arquivo
        while (l_linhalida and l_flsair == 0):
            # inicio while
            l_contador += 1

            # Quebra em vetor        
            l_dados_ent = l_linhalida.split("|")

            l_tp_reg = ""
            if (len(l_dados_ent) >=7):
                l_tp_reg = l_dados_ent[1].upper().strip()

            # Valida se os tipos        
            if (l_tp_reg == "C190"
            or l_tp_reg == "C590"
            or l_tp_reg == "D190"
            or l_tp_reg == "D590"
            or l_tp_reg == "D696"
            ):
                # inicio if
                l_nregrel += 1 

                l_CFOP = l_dados_ent[3].upper().strip()
                l_VLR_OPER = float(l_dados_ent[5].replace(',', '.'))
                l_VLR_BASE_ICMS = float(l_dados_ent[6].replace(',', '.'))
                l_VLR_ICMS = float(l_dados_ent[7].replace(',', '.'))
                        
                l_regrel.append([])
                l_regrel[l_nregrel].append(l_CFOP)
                l_regrel[l_nregrel].append(l_VLR_OPER)
                l_regrel[l_nregrel].append(l_VLR_BASE_ICMS)
                l_regrel[l_nregrel].append(l_VLR_ICMS)
                l_regrel[l_nregrel].append(l_tp_reg)

                # Teste
                # if l_nregrel < 3:
                #    log("regrel: " + str(l_regrel))

                # fim if

            # proxima linha
            l_linhalida = l_ent.readline()
                
            # Fim while
    
        # fechamento do arquivo
        l_ent.close()

        # Verifica se o arquivo teve algum processamentp                
        if l_nregrel > 0:
            return l_regrel
        else:
            log("# " + str(len(l_regrel)) + " >> " + "Não processou nenhum dados do arquivo : " + l_arquivo)
            return None
        
    else:
       return None     

"""
Funcao processa arquivo do tipo GIA >> REGERADO , PROTOCOLADO
"""
def processa_GIA(l_arquivo):
    l_regrel=[[],[]]
    l_regrel[0]=["CFOP","VLR_CONTABIL","VLR_BASE_ICMS","VLR_ICMS","VLR_ISENTAS","VLR_OUTRAS","SUBSTITUTO","SUBSTITUIDO","OUTROS_IMPOSTOS","TIPO"]

    # inicializa a variaveis de controle
    l_flsair = 0 
    l_contador = 0
    l_nregrel = 0   
    l_tp_reg = ""

    l_CFOP = ""
    l_VLR_CONTABIL = float(0)
    l_VLR_BASE_ICMS = float(0)
    l_VLR_ICMS = float(0)
    l_VLR_ISENTAS = float(0)
    l_VLR_OUTRAS = float(0)
    l_SUBSTITUTO = float(0)
    l_SUBSTITUIDO = float(0)
    l_OUTROS_IMPOSTOS = float(0)
    
    if (len(l_arquivo) > 4 and os.path.isfile(l_arquivo)) :
        
        log("Processando leitura do arquivo : " + l_arquivo)
        
        # realiza a abertura do arquivo
        l_ent = open(l_arquivo,mode="r",encoding=tipoArquivo(l_arquivo))
                
        # realiza a leitura da linha do arquivo
        l_linhalida = l_ent.readline()
                
        # percorrre o arquivo
        while (l_linhalida and l_flsair == 0):
            # inicio while
            l_contador += 1

            # Quebra em vetor        
            l_dados_ent = str(l_linhalida).strip()

            l_tp_reg = ""
            if len(l_dados_ent) >=147:
                if l_dados_ent[0:2] == "10":
                    l_tp_reg = "10"

            # Valida se os tipos        
            if (l_tp_reg == "10"):
                # inicio if
                l_nregrel += 1 
                """
				CR	2	1                   [0:2]
				* CFOP:CFOP	4	3                                [2:6]
				  Filler	2	7               [6:8]
				* VLR_CONTABIL:ValorContábil	15	9            [8:23]
				* VLR_BASE_ICMS:BaseCálculo	15	24               [23:38]
				* VLR_ICMS:Imposto	15	39                       [38:53]
				* VLR_ISENTAS:IsentasNãoTrib	15	54           [53:68]
				* VLR_OUTRAS:Outras	15	69                       [68:83] 
				  ImpostoRetidoST	15	84      [83:98]
				* SUBSTITUTO:ImpRetSubstitutoST	15	99           [98:113]
				* SUBSTITUÍDO:ImpRetSubstituído	15	114          [113:128]
				* OUTROS_IMPOSTOS:OutrosImpostos	15	129      [128:143]
				  Q14	4	144
				"""                
                l_CFOP = l_dados_ent[2:6].upper().strip()
                l_VLR_CONTABIL = float(l_dados_ent[8:23])/100
                l_VLR_BASE_ICMS = float(l_dados_ent[23:38])/100
                l_VLR_ICMS = float(l_dados_ent[38:53])/100
                l_VLR_ISENTAS = float(l_dados_ent[53:68])/100
                l_VLR_OUTRAS = float(l_dados_ent[68:83])/100
                l_SUBSTITUTO = float(l_dados_ent[98:113])/100
                l_SUBSTITUIDO = float(l_dados_ent[113:128])/100
                l_OUTROS_IMPOSTOS = float(l_dados_ent[128:143])/100

                l_regrel.append([])
                l_regrel[l_nregrel].append(l_CFOP)
                l_regrel[l_nregrel].append(l_VLR_CONTABIL)
                l_regrel[l_nregrel].append(l_VLR_BASE_ICMS)
                l_regrel[l_nregrel].append(l_VLR_ICMS)
                l_regrel[l_nregrel].append(l_VLR_ISENTAS)
                l_regrel[l_nregrel].append(l_VLR_OUTRAS)
                l_regrel[l_nregrel].append(l_SUBSTITUTO)
                l_regrel[l_nregrel].append(l_SUBSTITUIDO)
                l_regrel[l_nregrel].append(l_OUTROS_IMPOSTOS)
                l_regrel[l_nregrel].append(l_tp_reg)                

                # Teste
                # if l_nregrel < 3:
                #    log("regrel: " + str(l_regrel))

                # fim if

            # proxima linha
            l_linhalida = l_ent.readline()
                
            # Fim while
    
        # fechamento do arquivo
        l_ent.close()

        # Verifica se o arquivo teve algum processamentp                
        if l_nregrel > 0:
            return l_regrel
        else:
            log("# " + str(len(l_regrel)) + " >> " + "Não processou nenhum dados do arquivo : " + l_arquivo)
            return None
        
    else:
       return None     

"""
Funcao principal responsavel por acionamento 
das acoes a serem realizados
"""
def main(p_connection):

    log("\n")
    log("PARAMETRO ARQUIVO SPED FISCAL ENXERTADO: " + gv_lst_arq_sped_fiscal_enxertado)
    log("PARAMETRO ARQUIVO SPED FISCAL PROTOCOLADO: " + gv_lst_arq_sped_fiscal_protocolado)
    log("PARAMETRO ARQUIVO GIA ENXERTADO: " + gv_lst_arq_gia_enxertado)
    log("PARAMETRO ARQUIVO GIA PROTOCOLADO: " + gv_lst_arq_gia_protocolado)
    log("PARAMETRO A SER CRIADO ARQUIVO INSUMO SPED: " + gv_lst_arq_insumo_sped)
    log("\n")

    l_ret = 0 # controle de retorno
    l_a1 = 0
    l_a2 = 0
    l_a3 = 0
    l_a4 = 0
    l_a5 = 0

    # Crias as listas a serem carregadas
    v_sped_enxertado=[]
    v_sped_protocolado=[]
    v_gia_enxertado=[]
    v_gia_protocolado=[]

    # Carrega as listas
    if not l_ret:
        if (len(gv_lst_arq_sped_fiscal_enxertado) > 4 
        and os.path.isfile(gv_lst_arq_sped_fiscal_enxertado)) :
            try:
                v_sped_enxertado = processa_SPED(gv_lst_arq_sped_fiscal_enxertado)
                if v_sped_enxertado is not None:
                    l_a1 = 1 # Processado
                    log("# " + str(len(v_sped_enxertado)) + " >> Processado arquivo : " + gv_lst_arq_sped_fiscal_enxertado)

            except:
                l_ret = 1
                log("Erro leitura do arquivo : " + gv_lst_arq_sped_fiscal_enxertado)
    
    if not l_ret:
        if (len(gv_lst_arq_sped_fiscal_protocolado) > 4 
        and os.path.isfile(gv_lst_arq_sped_fiscal_protocolado)) :
            try:
                v_sped_protocolado = processa_SPED(gv_lst_arq_sped_fiscal_protocolado)
                if v_sped_protocolado is not None:
                    l_a2 = 1 # Processado
                    log("# " + str(len(v_sped_protocolado)) + " >> Processado arquivo : " + gv_lst_arq_sped_fiscal_protocolado)

            except:
                l_ret = 1
                log("Erro leitura do arquivo : " + gv_lst_arq_sped_fiscal_protocolado)
    
    if not l_ret:
        if (len(gv_lst_arq_gia_enxertado) > 4 
        and os.path.isfile(gv_lst_arq_gia_enxertado)) :
            try:
                v_gia_enxertado = processa_GIA(gv_lst_arq_gia_enxertado)
                if v_gia_enxertado is not None:
                    l_a3 = 1 # Processado
                    log("# " + str(len(v_gia_enxertado)) + " >> Processado arquivo : " + gv_lst_arq_gia_enxertado)

            except:
                l_ret = 1
                log("Erro leitura do arquivo : " + gv_lst_arq_gia_enxertado)
    
    if not l_ret:
        if (len(gv_lst_arq_gia_protocolado) > 4 
        and os.path.isfile(gv_lst_arq_gia_protocolado)) :
            try:
                v_gia_protocolado = processa_GIA(gv_lst_arq_gia_protocolado)
                if v_gia_protocolado is not None:
                    l_a4 = 1 # Processado
                    log("# " + str(len(v_gia_protocolado)) + " >> Processado arquivo : " + gv_lst_arq_gia_protocolado)

            except:
                l_ret = 1
                log("Erro leitura do arquivo : " + gv_lst_arq_gia_protocolado)
    

    # Une todas as listas em um dicionario, agrupando por CFOP
    if not l_ret:
        if (len(gv_lst_arq_insumo_sped) > 4
        and (l_a1 == 1 or l_a2 == 1 or l_a3 == 1 or l_a4 == 1) 
        ): 
            # Principal campo do formulario, e aonde guarda todas 
            # as informacoes a serem geradas           
            l_dicionario = {}
            l_cont = 0

            # Chave do dicionario
            l_CFOP = ""
            # Campos do dicionario
            l_REG_SPED = "" # 0
            l_REG_VLR_OPER  = float(0) # 1
            l_REG_VLR_BASE_ICMS = float(0) # 2
            l_REG_VLR_ICMS = float(0) # 3
            l_PROT_SPED = "" # 4
            l_PROT_VLR_OPER  = float(0) # 5
            l_PROT_VLR_BASE_ICMS = float(0) # 6
            l_PROT_VLR_ICMS = float(0) # 7
            l_REG_GIA = "" # 8
            l_REG_GIA_VLR_CONTABIL = float(0) # 9
            l_REG_GIA_VLR_BASE_ICMS = float(0) # 10
            l_REG_GIA_VLR_ICMS = float(0) # 11
            l_REG_GIA_VLR_ISENTAS = float(0) # 12
            l_REG_GIA_VLR_OUTRAS = float(0) # 13
            l_REG_GIA_SUBSTITUTO = float(0) # 14
            l_REG_GIA_SUBSTITUIDO = float(0) # 15
            l_REG_GIA_OUTROS_IMPOSTOS = float(0) # 16
            l_PROT_GIA = "" # 17
            l_PROT_GIA_VLR_CONTABIL = float(0) # 18
            l_PROT_GIA_VLR_BASE_ICMS = float(0) # 19
            l_PROT_GIA_VLR_ICMS = float(0) # 20
            l_PROT_GIA_VLR_ISENTAS = float(0) # 21
            l_PROT_GIA_VLR_OUTRAS = float(0) # 22
            l_PROT_GIA_SUBSTITUTO = float(0) # 23
            l_PROT_GIA_SUBSTITUIDO = float(0) # 24
            l_PROT_GIA_OUTROS_IMPOSTOS = float(0) # 25

            # SPED REGERADO
            if l_a1 == 1 and v_sped_enxertado is not None:
                l_cont = 0
                for linha in v_sped_enxertado:
                    l_cont += 1
                    if l_cont == 1 or len(linha) < 3:
                        continue          
                    # Atribui os valores da lista atual para colocar no dicionario
                    l_CFOP = str(linha[0]).strip().upper()
                    l_REG_SPED = l_CFOP
                    l_REG_VLR_OPER  = linha[1]
                    l_REG_VLR_BASE_ICMS = linha[2]
                    l_REG_VLR_ICMS = linha[3]

                    # Verifica se existe no dicionario, agrupando por CFOP
                    if l_CFOP in l_dicionario:
                        # Atualiza os valores do dicionario
                        l_dicionario[l_CFOP][0] = l_REG_SPED
                        l_dicionario[l_CFOP][1] += l_REG_VLR_OPER
                        l_dicionario[l_CFOP][2] += l_REG_VLR_BASE_ICMS
                        l_dicionario[l_CFOP][3] += l_REG_VLR_ICMS

                    else:
                        # Cria uma nova lista e adiciona no dicionario
                        l_lst_aux = []
                        # A lista deve conter estas seguencias 
                        # Refere a lista atual
                        l_lst_aux.append(l_REG_SPED)
                        l_lst_aux.append(l_REG_VLR_OPER)
                        l_lst_aux.append(l_REG_VLR_BASE_ICMS)
                        l_lst_aux.append(l_REG_VLR_ICMS)
                        # Valores default
                        l_lst_aux.append(l_PROT_SPED)
                        l_lst_aux.append(l_PROT_VLR_OPER)
                        l_lst_aux.append(l_PROT_VLR_BASE_ICMS)
                        l_lst_aux.append(l_PROT_VLR_ICMS)
                        l_lst_aux.append(l_REG_GIA)
                        l_lst_aux.append(l_REG_GIA_VLR_CONTABIL)
                        l_lst_aux.append(l_REG_GIA_VLR_BASE_ICMS)
                        l_lst_aux.append(l_REG_GIA_VLR_ICMS)
                        l_lst_aux.append(l_REG_GIA_VLR_ISENTAS)
                        l_lst_aux.append(l_REG_GIA_VLR_OUTRAS)
                        l_lst_aux.append(l_REG_GIA_SUBSTITUTO)
                        l_lst_aux.append(l_REG_GIA_SUBSTITUIDO)
                        l_lst_aux.append(l_REG_GIA_OUTROS_IMPOSTOS)
                        l_lst_aux.append(l_PROT_GIA)
                        l_lst_aux.append(l_PROT_GIA_VLR_CONTABIL)
                        l_lst_aux.append(l_PROT_GIA_VLR_BASE_ICMS)
                        l_lst_aux.append(l_PROT_GIA_VLR_ICMS)
                        l_lst_aux.append(l_PROT_GIA_VLR_ISENTAS)
                        l_lst_aux.append(l_PROT_GIA_VLR_OUTRAS)
                        l_lst_aux.append(l_PROT_GIA_SUBSTITUTO)
                        l_lst_aux.append(l_PROT_GIA_SUBSTITUIDO)
                        l_lst_aux.append(l_PROT_GIA_OUTROS_IMPOSTOS)
                        # Adiciona a lista no dicionario
                        l_dicionario[l_CFOP] = l_lst_aux                        

                    # Limpa os valores UTILIZADOS para proxima lista:
                    l_CFOP = ""
                    l_REG_SPED = ""
                    l_REG_VLR_OPER  = float(0)
                    l_REG_VLR_BASE_ICMS = float(0)
                    l_REG_VLR_ICMS = float(0)
                    # Fim For    

            # SPED PROTOCOLADO
            if l_a2 == 1 and v_sped_protocolado is not None:
                l_cont = 0
                for linha in v_sped_protocolado:
                    l_cont += 1
                    if l_cont == 1 or len(linha) < 3:
                        continue          
                    # Atribui os valores da lista atual para colocar no dicionario
                    l_CFOP = str(linha[0]).strip().upper()
                    l_PROT_SPED = l_CFOP
                    l_PROT_VLR_OPER  = linha[1]
                    l_PROT_VLR_BASE_ICMS = linha[2]
                    l_PROT_VLR_ICMS = linha[3]

                    # Verifica se existe no dicionario, agrupando por CFOP
                    if l_CFOP in l_dicionario:
                        # Atualiza os valores do dicionario
                        l_dicionario[l_CFOP][4] = l_PROT_SPED
                        l_dicionario[l_CFOP][5] += l_PROT_VLR_OPER
                        l_dicionario[l_CFOP][6] += l_PROT_VLR_BASE_ICMS
                        l_dicionario[l_CFOP][7] += l_PROT_VLR_ICMS

                    else:
                        # Cria uma nova lista e adiciona no dicionario
                        l_lst_aux = []
                        # A lista deve conter estas seguencias 
                        # Valores default
                        l_lst_aux.append(l_REG_SPED)
                        l_lst_aux.append(l_REG_VLR_OPER)
                        l_lst_aux.append(l_REG_VLR_BASE_ICMS)
                        l_lst_aux.append(l_REG_VLR_ICMS)
                        # Refere a lista atual
                        l_lst_aux.append(l_PROT_SPED)
                        l_lst_aux.append(l_PROT_VLR_OPER)
                        l_lst_aux.append(l_PROT_VLR_BASE_ICMS)
                        l_lst_aux.append(l_PROT_VLR_ICMS)
                        # Valores default
                        l_lst_aux.append(l_REG_GIA)
                        l_lst_aux.append(l_REG_GIA_VLR_CONTABIL)
                        l_lst_aux.append(l_REG_GIA_VLR_BASE_ICMS)
                        l_lst_aux.append(l_REG_GIA_VLR_ICMS)
                        l_lst_aux.append(l_REG_GIA_VLR_ISENTAS)
                        l_lst_aux.append(l_REG_GIA_VLR_OUTRAS)
                        l_lst_aux.append(l_REG_GIA_SUBSTITUTO)
                        l_lst_aux.append(l_REG_GIA_SUBSTITUIDO)
                        l_lst_aux.append(l_REG_GIA_OUTROS_IMPOSTOS)
                        l_lst_aux.append(l_PROT_GIA)
                        l_lst_aux.append(l_PROT_GIA_VLR_CONTABIL)
                        l_lst_aux.append(l_PROT_GIA_VLR_BASE_ICMS)
                        l_lst_aux.append(l_PROT_GIA_VLR_ICMS)
                        l_lst_aux.append(l_PROT_GIA_VLR_ISENTAS)
                        l_lst_aux.append(l_PROT_GIA_VLR_OUTRAS)
                        l_lst_aux.append(l_PROT_GIA_SUBSTITUTO)
                        l_lst_aux.append(l_PROT_GIA_SUBSTITUIDO)
                        l_lst_aux.append(l_PROT_GIA_OUTROS_IMPOSTOS)
                        # Adiciona a lista no dicionario
                        l_dicionario[l_CFOP] = l_lst_aux                        
                    
                    # log(str(linha) + " >> k: " +  str(l_CFOP) + " >> v:  " + str(l_dicionario[l_CFOP]) + "")
                    # if l_cont > 4:
                    #    break
                    
                    # Limpa os valores UTILIZADOS para proxima lista:
                    l_CFOP = ""
                    l_PROT_SPED = ""
                    l_PROT_VLR_OPER  = float(0)
                    l_PROT_VLR_BASE_ICMS = float(0)
                    l_PROT_VLR_ICMS = float(0)                   
                    # Fim For    
    
            # GIA REGERADO    
            if l_a3 == 1 and v_gia_enxertado is not None:
                l_cont = 0
                for linha in v_gia_enxertado:
                    l_cont += 1
                    if l_cont == 1 or len(linha) < 3:
                        continue          
                    # Atribui os valores da lista atual para colocar no dicionario
                    l_CFOP = str(linha[0]).strip().upper()
                    l_REG_GIA = l_CFOP
                    l_REG_GIA_VLR_CONTABIL = linha[1]
                    l_REG_GIA_VLR_BASE_ICMS = linha[2]
                    l_REG_GIA_VLR_ICMS = linha[3]
                    l_REG_GIA_VLR_ISENTAS = linha[4]
                    l_REG_GIA_VLR_OUTRAS = linha[5]
                    l_REG_GIA_SUBSTITUTO = linha[6]
                    l_REG_GIA_SUBSTITUIDO = linha[7]
                    l_REG_GIA_OUTROS_IMPOSTOS = linha[8]                  

                    # Verifica se existe no dicionario, agrupando por CFOP
                    if l_CFOP in l_dicionario:
                        # Atualiza os valores do dicionario
                        l_dicionario[l_CFOP][8] = l_REG_GIA
                        l_dicionario[l_CFOP][9] += l_REG_GIA_VLR_CONTABIL
                        l_dicionario[l_CFOP][10] += l_REG_GIA_VLR_BASE_ICMS
                        l_dicionario[l_CFOP][11] += l_REG_GIA_VLR_ICMS
                        l_dicionario[l_CFOP][12] += l_REG_GIA_VLR_ISENTAS
                        l_dicionario[l_CFOP][13] += l_REG_GIA_VLR_OUTRAS
                        l_dicionario[l_CFOP][14] += l_REG_GIA_SUBSTITUTO
                        l_dicionario[l_CFOP][15] += l_REG_GIA_SUBSTITUIDO
                        l_dicionario[l_CFOP][16] += l_REG_GIA_OUTROS_IMPOSTOS

                    else:
                        # Cria uma nova lista e adiciona no dicionario
                        l_lst_aux = []
                        # A lista deve conter estas seguencias 
                        # Valores default
                        l_lst_aux.append(l_REG_SPED)
                        l_lst_aux.append(l_REG_VLR_OPER)
                        l_lst_aux.append(l_REG_VLR_BASE_ICMS)
                        l_lst_aux.append(l_REG_VLR_ICMS)
                        # Refere a lista atual
                        l_lst_aux.append(l_PROT_SPED)
                        l_lst_aux.append(l_PROT_VLR_OPER)
                        l_lst_aux.append(l_PROT_VLR_BASE_ICMS)
                        l_lst_aux.append(l_PROT_VLR_ICMS)
                        # Valores default
                        l_lst_aux.append(l_REG_GIA)
                        l_lst_aux.append(l_REG_GIA_VLR_CONTABIL)
                        l_lst_aux.append(l_REG_GIA_VLR_BASE_ICMS)
                        l_lst_aux.append(l_REG_GIA_VLR_ICMS)
                        l_lst_aux.append(l_REG_GIA_VLR_ISENTAS)
                        l_lst_aux.append(l_REG_GIA_VLR_OUTRAS)
                        l_lst_aux.append(l_REG_GIA_SUBSTITUTO)
                        l_lst_aux.append(l_REG_GIA_SUBSTITUIDO)
                        l_lst_aux.append(l_REG_GIA_OUTROS_IMPOSTOS)
                        l_lst_aux.append(l_PROT_GIA)
                        l_lst_aux.append(l_PROT_GIA_VLR_CONTABIL)
                        l_lst_aux.append(l_PROT_GIA_VLR_BASE_ICMS)
                        l_lst_aux.append(l_PROT_GIA_VLR_ICMS)
                        l_lst_aux.append(l_PROT_GIA_VLR_ISENTAS)
                        l_lst_aux.append(l_PROT_GIA_VLR_OUTRAS)
                        l_lst_aux.append(l_PROT_GIA_SUBSTITUTO)
                        l_lst_aux.append(l_PROT_GIA_SUBSTITUIDO)
                        l_lst_aux.append(l_PROT_GIA_OUTROS_IMPOSTOS)
                        # Adiciona a lista no dicionario
                        l_dicionario[l_CFOP] = l_lst_aux                        
                    
                    # log(str(linha) + " >> k: " +  str(l_CFOP) + " >> v:  " + str(l_dicionario[l_CFOP]) + "")
                    # if l_cont > 4:
                    #    break
                    
                    # Limpa os valores UTILIZADOS para proxima lista:
                    l_CFOP = ""
                    l_REG_GIA = ""
                    l_REG_GIA_VLR_CONTABIL  = float(0)
                    l_REG_GIA_VLR_BASE_ICMS = float(0)
                    l_REG_GIA_VLR_ICMS = float(0) 
                    l_REG_GIA_VLR_ISENTAS = float(0)
                    l_REG_GIA_VLR_OUTRAS = float(0)
                    l_REG_GIA_SUBSTITUTO = float(0)
                    l_REG_GIA_SUBSTITUIDO = float(0)
                    l_REG_GIA_OUTROS_IMPOSTOS = float(0)                                             
                    # Fim For  
              
            # GIA PROTOCOLADO    
            if l_a4 == 1 and v_gia_protocolado is not None:
                l_cont = 0
                for linha in v_gia_protocolado:
                    l_cont += 1
                    if l_cont == 1 or len(linha) < 3:
                        continue          
                    # Atribui os valores da lista atual para colocar no dicionario
                    l_CFOP = str(linha[0]).strip().upper()
                    l_PROT_GIA = l_CFOP
                    l_PROT_GIA_VLR_CONTABIL = linha[1]
                    l_PROT_GIA_VLR_BASE_ICMS = linha[2]
                    l_PROT_GIA_VLR_ICMS = linha[3]
                    l_PROT_GIA_VLR_ISENTAS = linha[4]
                    l_PROT_GIA_VLR_OUTRAS = linha[5]
                    l_PROT_GIA_SUBSTITUTO = linha[6]
                    l_PROT_GIA_SUBSTITUIDO = linha[7]
                    l_PROT_GIA_OUTROS_IMPOSTOS = linha[8]                  

                    # Verifica se existe no dicionario, agrupando por CFOP
                    if l_CFOP in l_dicionario:
                        # Atualiza os valores do dicionario
                        l_dicionario[l_CFOP][17] = l_PROT_GIA
                        l_dicionario[l_CFOP][18] += l_PROT_GIA_VLR_CONTABIL
                        l_dicionario[l_CFOP][19] += l_PROT_GIA_VLR_BASE_ICMS
                        l_dicionario[l_CFOP][20] += l_PROT_GIA_VLR_ICMS
                        l_dicionario[l_CFOP][21] += l_PROT_GIA_VLR_ISENTAS
                        l_dicionario[l_CFOP][22] += l_PROT_GIA_VLR_OUTRAS
                        l_dicionario[l_CFOP][23] += l_PROT_GIA_SUBSTITUTO
                        l_dicionario[l_CFOP][24] += l_PROT_GIA_SUBSTITUIDO
                        l_dicionario[l_CFOP][25] += l_PROT_GIA_OUTROS_IMPOSTOS

                    else:
                        # Cria uma nova lista e adiciona no dicionario
                        l_lst_aux = []
                        # A lista deve conter estas seguencias 
                        # Valores default
                        l_lst_aux.append(l_REG_SPED)
                        l_lst_aux.append(l_REG_VLR_OPER)
                        l_lst_aux.append(l_REG_VLR_BASE_ICMS)
                        l_lst_aux.append(l_REG_VLR_ICMS)
                        # Refere a lista atual
                        l_lst_aux.append(l_PROT_SPED)
                        l_lst_aux.append(l_PROT_VLR_OPER)
                        l_lst_aux.append(l_PROT_VLR_BASE_ICMS)
                        l_lst_aux.append(l_PROT_VLR_ICMS)
                        # Valores default
                        l_lst_aux.append(l_REG_GIA)
                        l_lst_aux.append(l_REG_GIA_VLR_CONTABIL)
                        l_lst_aux.append(l_REG_GIA_VLR_BASE_ICMS)
                        l_lst_aux.append(l_REG_GIA_VLR_ICMS)
                        l_lst_aux.append(l_REG_GIA_VLR_ISENTAS)
                        l_lst_aux.append(l_REG_GIA_VLR_OUTRAS)
                        l_lst_aux.append(l_REG_GIA_SUBSTITUTO)
                        l_lst_aux.append(l_REG_GIA_SUBSTITUIDO)
                        l_lst_aux.append(l_REG_GIA_OUTROS_IMPOSTOS)
                        l_lst_aux.append(l_PROT_GIA)
                        l_lst_aux.append(l_PROT_GIA_VLR_CONTABIL)
                        l_lst_aux.append(l_PROT_GIA_VLR_BASE_ICMS)
                        l_lst_aux.append(l_PROT_GIA_VLR_ICMS)
                        l_lst_aux.append(l_PROT_GIA_VLR_ISENTAS)
                        l_lst_aux.append(l_PROT_GIA_VLR_OUTRAS)
                        l_lst_aux.append(l_PROT_GIA_SUBSTITUTO)
                        l_lst_aux.append(l_PROT_GIA_SUBSTITUIDO)
                        l_lst_aux.append(l_PROT_GIA_OUTROS_IMPOSTOS)
                        # Adiciona a lista no dicionario
                        l_dicionario[l_CFOP] = l_lst_aux                        
                    
                    # log(str(linha) + " >> k: " +  str(l_CFOP) + " >> v:  " + str(l_dicionario[l_CFOP]) + "")
                    # if l_cont > 4:
                    #    break
                    
                    # Limpa os valores UTILIZADOS para proxima lista:
                    l_CFOP = ""
                    l_PROT_GIA = ""
                    l_PROT_GIA_VLR_CONTABIL  = float(0)
                    l_PROT_GIA_VLR_BASE_ICMS = float(0)
                    l_PROT_GIA_VLR_ICMS = float(0) 
                    l_PROT_GIA_VLR_ISENTAS = float(0)
                    l_PROT_GIA_VLR_OUTRAS = float(0)
                    l_PROT_GIA_SUBSTITUTO = float(0)
                    l_PROT_GIA_SUBSTITUIDO = float(0)
                    l_PROT_GIA_OUTROS_IMPOSTOS = float(0)                                             
                    # Fim For  
            
            # Zerando as variaveis para somatorio
            # GIA - Regerado x Protocolado : TOTAL ENTRADA 
            l_ENT_REG_GIA_VLR_CONTABIL  = float(0)
            l_ENT_REG_GIA_VLR_BASE_ICMS = float(0)
            l_ENT_REG_GIA_VLR_ICMS = float(0) 
            l_ENT_REG_GIA_VLR_ISENTAS = float(0)
            l_ENT_REG_GIA_VLR_OUTRAS = float(0)
            l_ENT_REG_GIA_SUBSTITUTO = float(0)
            l_ENT_REG_GIA_SUBSTITUIDO = float(0)
            l_ENT_REG_GIA_OUTROS_IMPOSTOS = float(0)             
            l_ENT_PROT_GIA_VLR_CONTABIL  = float(0)
            l_ENT_PROT_GIA_VLR_BASE_ICMS = float(0)
            l_ENT_PROT_GIA_VLR_ICMS = float(0) 
            l_ENT_PROT_GIA_VLR_ISENTAS = float(0)
            l_ENT_PROT_GIA_VLR_OUTRAS = float(0)
            l_ENT_PROT_GIA_SUBSTITUTO = float(0)
            l_ENT_PROT_GIA_SUBSTITUIDO = float(0)
            l_ENT_PROT_GIA_OUTROS_IMPOSTOS = float(0)    
            # GIA - Regerado x Protocolado TOTAL SAIDA 
            l_SAI_REG_GIA_VLR_CONTABIL  = float(0)
            l_SAI_REG_GIA_VLR_BASE_ICMS = float(0)
            l_SAI_REG_GIA_VLR_ICMS = float(0) 
            l_SAI_REG_GIA_VLR_ISENTAS = float(0)
            l_SAI_REG_GIA_VLR_OUTRAS = float(0)
            l_SAI_REG_GIA_SUBSTITUTO = float(0)
            l_SAI_REG_GIA_SUBSTITUIDO = float(0)
            l_SAI_REG_GIA_OUTROS_IMPOSTOS = float(0)             
            l_SAI_PROT_GIA_VLR_CONTABIL  = float(0)
            l_SAI_PROT_GIA_VLR_BASE_ICMS = float(0)
            l_SAI_PROT_GIA_VLR_ICMS = float(0) 
            l_SAI_PROT_GIA_VLR_ISENTAS = float(0)
            l_SAI_PROT_GIA_VLR_OUTRAS = float(0)
            l_SAI_PROT_GIA_SUBSTITUTO = float(0)
            l_SAI_PROT_GIA_SUBSTITUIDO = float(0)
            l_SAI_PROT_GIA_OUTROS_IMPOSTOS = float(0)   

            # Sped - Regerado x Protocolado : ENTRADA
            l_ENT_REG_VLR_OPER  = float(0) 
            l_ENT_REG_VLR_BASE_ICMS = float(0) 
            l_ENT_REG_VLR_ICMS = float(0) 
            l_ENT_PROT_VLR_OPER  = float(0) 
            l_ENT_PROT_VLR_BASE_ICMS = float(0) 
            l_ENT_PROT_VLR_ICMS = float(0) 
            # Sped - Regerado x Protocolado : SAIDA
            l_SAI_REG_VLR_OPER  = float(0) 
            l_SAI_REG_VLR_BASE_ICMS = float(0) 
            l_SAI_REG_VLR_ICMS = float(0) 
            l_SAI_PROT_VLR_OPER  = float(0) 
            l_SAI_PROT_VLR_BASE_ICMS = float(0) 
            l_SAI_PROT_VLR_ICMS = float(0) 

            # Realizando leitura do dicionario
            if len(l_dicionario) > 0:
                try:
                    log("Criando arquivo : " + gv_lst_arq_insumo_sped)
                    #### Cria a planilha em memória....
                    arquivo_excel = Workbook()
                    planilha1 = arquivo_excel.active
                    planilha1.title = "GIA - Regerado x Protocolado"
                    planilha2 = arquivo_excel.create_sheet("Sped - Regerado x Protocolado", 1)
                    planilha3 = arquivo_excel.create_sheet("Sped x Gia_Regerado", 2)
                    planilha8 = arquivo_excel.create_sheet("Parametros", 3)
                    planilha8.sheet_state = 'hidden'
                    # planilha4 = arquivo_excel.create_sheet("Sped - Regerado", 4)
                    # planilha5 = arquivo_excel.create_sheet("Sped - Protocolado", 5)
                    # planilha6 = arquivo_excel.create_sheet("GIA - Regerado", 6)
                    # planilha7 = arquivo_excel.create_sheet("GIA - Protocolado", 7)
                    
                    
                    # Tratamento das planilhas 
                    l_cont = 0
                    l_linha  = 1
                    l_coluna = 2  

                    # Lista de forma ordenada a lista 
                    sd = collections.OrderedDict(sorted(l_dicionario.items()))
                    for k in sd: # sorted(l_dicionario, key = l_dicionario.get):
                        # inicio for
                        l_cont += 1

                        # Chave
                        l_CFOP = str(k)

                        # log("k: " +  str(k) + " >> v:  " + str(l_dicionario[k]) + "")
                        if l_cont == 1:
                            # INICIO IF
                            # Header - Cabeçalho  ** GIA ***
                            l_linha  = 1 # Linha  1
                            l_coluna = 2 # Coluna B # Gia (Regerado)
                            planilha1.cell(l_linha,l_coluna,"Gia (Regerado)")
                            planilha1.cell(l_linha,l_coluna).font=Font(bold=True)
                            planilha1.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                            planilha1.merge_cells('B'+ str(l_linha) + ':J' + str(l_linha))

                            l_coluna = 12 # Coluna L # Gia (Protocolado)
                            planilha1.cell(l_linha,l_coluna,"Gia (Protocolado)")
                            planilha1.cell(l_linha,l_coluna).font=Font(bold=True)
                            planilha1.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                            planilha1.merge_cells('L'+ str(l_linha) + ':T' + str(l_linha))
                            		
                            l_coluna = 22 # Coluna V # Diferença: Gia (Regerado) versus Gia Protocolado
                            planilha1.cell(l_linha,l_coluna,"Diferença: Gia (Regerado) versus Gia Protocolado")
                            planilha1.cell(l_linha,l_coluna).font=Font(bold=True)
                            planilha1.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                            planilha1.merge_cells('V'+ str(l_linha) + ':AD' + str(l_linha))

                            # Campos - Cabeçalho
                            # Col 2: Gia (Regerado) ; Col 12: Gia (Protocolado) ; Col 22 : Diferença: Gia (Regerado) versus Gia Protocolado
                            for nCol in (2,12,22):
                                l_coluna = nCol # Coluna 
                                l_linha  = 2 # Linha  2
                                for nColuna in ("CFOP", 
                                                "VLR_CONTABIL",	
                                                "VLR_BASE_ICMS",	
                                                "VLR_ICMS",	
                                                "VLR_ISENTAS",	
                                                "VLR_OUTRAS",	
                                                "SUBSTITUTO",
                                                "SUBSTITUÍDO",	
                                                "OUTROS IMPOSTOS"):
                                    planilha1.cell(l_linha,l_coluna,nColuna)
                                    planilha1.cell(l_linha,l_coluna).font=Font(bold=True)
                                    planilha1.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                                    l_coluna = l_coluna + 1                            

                            # Header - Cabeçalho  ** Sped - Regerado x Protocolado ***
                            l_linha  = 1 # Linha  1
                            l_coluna = 2 # Coluna B # PVA SPED FISCAL - REGERADO
                            planilha2.cell(l_linha,l_coluna,"PVA SPED FISCAL - REGERADO")
                            planilha2.cell(l_linha,l_coluna).font=Font(bold=True)
                            planilha2.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                            planilha2.merge_cells('B'+ str(l_linha) + ':E' + str(l_linha))

                            l_coluna = 7 # Coluna G # PVA SPED FISCAL - PROTOCOLADO
                            planilha2.cell(l_linha,l_coluna,"PVA SPED FISCAL - PROTOCOLADO")
                            planilha2.cell(l_linha,l_coluna).font=Font(bold=True)
                            planilha2.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                            planilha2.merge_cells('G'+ str(l_linha) + ':J' + str(l_linha))
                            		
                            l_coluna = 12 # Coluna L # DIFERENÇA: SPED REGERADO X SPED PROTOCOLADO
                            planilha2.cell(l_linha,l_coluna,"DIFERENÇA: SPED REGERADO X SPED PROTOCOLADO")
                            planilha2.cell(l_linha,l_coluna).font=Font(bold=True)
                            planilha2.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                            planilha2.merge_cells('L'+ str(l_linha) + ':O' + str(l_linha))

                            # Header - Cabeçalho  ** Sped x Gia_Regerado ***
                            l_linha  = 1 # Linha  1
                            l_coluna = 2 # Coluna B # PVA SPED FISCAL - REGERADO
                            planilha3.cell(l_linha,l_coluna,"PVA SPED FISCAL - REGERADO")
                            planilha3.cell(l_linha,l_coluna).font=Font(bold=True)
                            planilha3.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                            planilha3.merge_cells('B'+ str(l_linha) + ':E' + str(l_linha))

                            l_coluna = 7 # Coluna G # GIA - REGERADO
                            planilha3.cell(l_linha,l_coluna,"GIA - REGERADO")
                            planilha3.cell(l_linha,l_coluna).font=Font(bold=True)
                            planilha3.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                            planilha3.merge_cells('G'+ str(l_linha) + ':J' + str(l_linha))
                            		
                            l_coluna = 12 # Coluna L # DIFERENÇA: SPED REGERADO X GIA REGERADO
                            planilha3.cell(l_linha,l_coluna,"DIFERENÇA: SPED REGERADO X GIA REGERADO")
                            planilha3.cell(l_linha,l_coluna).font=Font(bold=True)
                            planilha3.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                            planilha3.merge_cells('L'+ str(l_linha) + ':O' + str(l_linha))

                            # Campos - Cabeçalho
                            for nCol in (2,7,12):
                                l_coluna = nCol # Coluna 
                                l_linha  = 2 # Linha  2
                                for nColuna in ("CFOP", 
                                                "VLR_OPER",	
                                                "VLR_BASE_ICMS",	
                                                "VLR_ICMS"                                                
                                                ):                                    
                                    # ** Sped - Regerado x Protocolado ***
                                    planilha2.cell(l_linha,l_coluna,nColuna)
                                    planilha2.cell(l_linha,l_coluna).font=Font(bold=True)
                                    planilha2.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                                    
                                    # ** Sped x Gia_Regerado ***
                                    planilha3.cell(l_linha,l_coluna,nColuna)
                                    planilha3.cell(l_linha,l_coluna).font=Font(bold=True)
                                    planilha3.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                                    
                                    l_coluna = l_coluna + 1       

                            # Preparando proxima linha 
                            l_linha  = 2 # Linha  2    
                            # FIM IF

                        # Nova linha
                        l_linha = 2 + l_cont # Linha  3 (a partir)
                        
                        # Preenchimento da Planilha com CFOP
                        for nCol in (2,7,12,22):
                                l_coluna = nCol # Coluna 
                                if l_coluna in (2,12,22):
                                    # GIA - Regerado x Protocolado    
                                    planilha1.cell(l_linha,l_coluna,l_CFOP)
                                    planilha1.cell(l_linha,l_coluna).font=Font(bold=True)
                                    planilha1.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                                
                                if l_coluna in (2,7,12):
                                    # ** Sped - Regerado x Protocolado ***
                                    planilha2.cell(l_linha,l_coluna,l_CFOP) 
                                    planilha2.cell(l_linha,l_coluna).font=Font(bold=True) 
                                    planilha2.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                                    # ** Sped x Gia_Regerado ***
                                    planilha3.cell(l_linha,l_coluna,l_CFOP) 
                                    planilha3.cell(l_linha,l_coluna).font=Font(bold=True) 
                                    planilha3.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')

                        # Preenchimento valores : GIA - Regerado x Protocolado
                        l_coluna = 3 # Coluna B 
                        l_CFOF_GIA =  l_dicionario[l_CFOP][8].strip()
                        for n in (9,10,11,12,13,14,15,16,18,19,20,21,22,23,24,25):
                            # Inicio For
                            if n == 18:
                                l_coluna = 13 # Protocolado 
                                l_CFOF_GIA =  l_dicionario[l_CFOP][17].strip()

                            if l_CFOF_GIA is not None and len(l_CFOF_GIA) > 0:
                                planilha1.cell(l_linha,l_coluna,round(l_dicionario[l_CFOP][n],2))                             
                            planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"

                            # Preenchimento valores : Sped x Gia_Regerado
                            if n in (9,10,11):
                                planilha3.cell(l_linha,l_coluna+5,round(l_dicionario[l_CFOP][n],2))                             
                                planilha3.cell(l_linha,l_coluna+5).number_format = "#,##0.00"                                

                            if n < 18: 
                                # Diferença
                                if (len(l_dicionario[l_CFOP][8].strip()) > 0 
                                or len(l_dicionario[l_CFOP][17].strip()) > 0 ):
                                    planilha1.cell(l_linha,l_coluna+20,round(l_dicionario[l_CFOP][n]-l_dicionario[l_CFOP][n+9],2)) 
                                planilha1.cell(l_linha,l_coluna+20).number_format = "#,##0.00"

                            l_coluna += 1
                            # Fim For

                        # Preenchimento valores : Sped - Regerado x Protocolado
                        l_coluna = 3 # Coluna B 
                        l_CFOF_SPED = l_dicionario[l_CFOP][0].strip()
                        for n in (1,2,3,5,6,7):
                            # Inicio For
                            if n == 5:
                                l_coluna = 8 # Protocolado 
                                l_CFOF_SPED = l_dicionario[l_CFOP][4].strip()
                            
                            if l_CFOF_SPED is not None and len(l_CFOF_SPED) > 0:
                                planilha2.cell(l_linha,l_coluna,round(l_dicionario[l_CFOP][n],2))                             
                            planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"

                            if n < 5: 
                                # Diferença
                                if (len(l_dicionario[l_CFOP][0].strip()) > 0 
                                or len(l_dicionario[l_CFOP][4].strip()) > 0 ):
                                    planilha2.cell(l_linha,l_coluna+10,round(l_dicionario[l_CFOP][n]-l_dicionario[l_CFOP][n+4],2)) 
                                planilha2.cell(l_linha,l_coluna+10).number_format = "#,##0.00"
                                
                                # Preenchimento valores : Sped x Gia_Regerado
                                planilha3.cell(l_linha,l_coluna,round(l_dicionario[l_CFOP][n],2))                             
                                planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"
                                # Diferença: Preenchimento valores : Sped x Gia_Regerado
                                planilha3.cell(l_linha,l_coluna+10,round(l_dicionario[l_CFOP][n]-l_dicionario[l_CFOP][n+8],2))                             
                                planilha3.cell(l_linha,l_coluna+10).number_format = "#,##0.00"

                            l_coluna += 1                            
                            # Fim For

                        # Somatorios de Entrada
                        if l_CFOP[0] in ('1','2','3'):
                            #  GIA - Regerado x Protocolado 
                            l_ENT_REG_GIA_VLR_CONTABIL += l_dicionario[l_CFOP][9] 
                            l_ENT_REG_GIA_VLR_BASE_ICMS += l_dicionario[l_CFOP][10]
                            l_ENT_REG_GIA_VLR_ICMS += l_dicionario[l_CFOP][11]
                            l_ENT_REG_GIA_VLR_ISENTAS += l_dicionario[l_CFOP][12]
                            l_ENT_REG_GIA_VLR_OUTRAS += l_dicionario[l_CFOP][13]
                            l_ENT_REG_GIA_SUBSTITUTO += l_dicionario[l_CFOP][14]
                            l_ENT_REG_GIA_SUBSTITUIDO += l_dicionario[l_CFOP][15]
                            l_ENT_REG_GIA_OUTROS_IMPOSTOS += l_dicionario[l_CFOP][16]   
                            l_ENT_PROT_GIA_VLR_CONTABIL += l_dicionario[l_CFOP][18]
                            l_ENT_PROT_GIA_VLR_BASE_ICMS += l_dicionario[l_CFOP][19]
                            l_ENT_PROT_GIA_VLR_ICMS += l_dicionario[l_CFOP][20]
                            l_ENT_PROT_GIA_VLR_ISENTAS += l_dicionario[l_CFOP][21]
                            l_ENT_PROT_GIA_VLR_OUTRAS += l_dicionario[l_CFOP][22]
                            l_ENT_PROT_GIA_SUBSTITUTO += l_dicionario[l_CFOP][23]
                            l_ENT_PROT_GIA_SUBSTITUIDO += l_dicionario[l_CFOP][24]
                            l_ENT_PROT_GIA_OUTROS_IMPOSTOS += l_dicionario[l_CFOP][25]
                            # Sped - Regerado x Protocolado
                            l_ENT_REG_VLR_OPER += l_dicionario[l_CFOP][1] 
                            l_ENT_REG_VLR_BASE_ICMS += l_dicionario[l_CFOP][2] 
                            l_ENT_REG_VLR_ICMS += l_dicionario[l_CFOP][3] 
                            l_ENT_PROT_VLR_OPER += l_dicionario[l_CFOP][5] 
                            l_ENT_PROT_VLR_BASE_ICMS += l_dicionario[l_CFOP][6] 
                            l_ENT_PROT_VLR_ICMS += l_dicionario[l_CFOP][7] 

                        # Somatorio de SAÍDA
                        if l_CFOP[0] in ('5','6','7'):
                             #  GIA - Regerado x Protocolado     
                            l_SAI_REG_GIA_VLR_CONTABIL += l_dicionario[l_CFOP][9] 
                            l_SAI_REG_GIA_VLR_BASE_ICMS += l_dicionario[l_CFOP][10]
                            l_SAI_REG_GIA_VLR_ICMS += l_dicionario[l_CFOP][11]
                            l_SAI_REG_GIA_VLR_ISENTAS += l_dicionario[l_CFOP][12]
                            l_SAI_REG_GIA_VLR_OUTRAS += l_dicionario[l_CFOP][13]
                            l_SAI_REG_GIA_SUBSTITUTO += l_dicionario[l_CFOP][14]
                            l_SAI_REG_GIA_SUBSTITUIDO += l_dicionario[l_CFOP][15]
                            l_SAI_REG_GIA_OUTROS_IMPOSTOS += l_dicionario[l_CFOP][16]   
                            l_SAI_PROT_GIA_VLR_CONTABIL += l_dicionario[l_CFOP][18]
                            l_SAI_PROT_GIA_VLR_BASE_ICMS += l_dicionario[l_CFOP][19]
                            l_SAI_PROT_GIA_VLR_ICMS += l_dicionario[l_CFOP][20]
                            l_SAI_PROT_GIA_VLR_ISENTAS += l_dicionario[l_CFOP][21]
                            l_SAI_PROT_GIA_VLR_OUTRAS += l_dicionario[l_CFOP][22]
                            l_SAI_PROT_GIA_SUBSTITUTO += l_dicionario[l_CFOP][23]
                            l_SAI_PROT_GIA_SUBSTITUIDO += l_dicionario[l_CFOP][24]
                            l_SAI_PROT_GIA_OUTROS_IMPOSTOS += l_dicionario[l_CFOP][25]
                            # Sped - Regerado x Protocolado
                            l_SAI_REG_VLR_OPER += l_dicionario[l_CFOP][1] 
                            l_SAI_REG_VLR_BASE_ICMS += l_dicionario[l_CFOP][2] 
                            l_SAI_REG_VLR_ICMS += l_dicionario[l_CFOP][3] 
                            l_SAI_PROT_VLR_OPER += l_dicionario[l_CFOP][5] 
                            l_SAI_PROT_VLR_BASE_ICMS += l_dicionario[l_CFOP][6] 
                            l_SAI_PROT_VLR_ICMS += l_dicionario[l_CFOP][7] 

                        # fim for
                        # if l_cont > 4:
                        #    break

                    # TOTAL DE ENTRADA E SAIDA
                    if l_cont > 0:
                        # TOTAL DE ENTRADA
                        l_linha += 2
                        l_coluna = 2 # Coluna B 
                        planilha1.cell(l_linha,l_coluna,"TOTAL ENTRADA") 
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)
                        planilha1.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                        planilha2.cell(l_linha,l_coluna,"TOTAL ENTRADA") 
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)
                        planilha2.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')                        
                        planilha3.cell(l_linha,l_coluna,"TOTAL ENTRADA") 
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)
                        planilha3.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')

                        # GIA - Regerado x Protocolado
                        l_coluna = 3
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_VLR_CONTABIL,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"    
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                    

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_VLR_BASE_ICMS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_VLR_ICMS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        
                        
                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_VLR_ISENTAS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        
                       
                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_VLR_OUTRAS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_SUBSTITUTO,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        
                        
                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_SUBSTITUIDO,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_OUTROS_IMPOSTOS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna = 13 # PROTOCOLADO
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_PROT_GIA_VLR_CONTABIL,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_PROT_GIA_VLR_BASE_ICMS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_PROT_GIA_VLR_ICMS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_PROT_GIA_VLR_ISENTAS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_PROT_GIA_VLR_OUTRAS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_PROT_GIA_SUBSTITUTO,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_PROT_GIA_SUBSTITUIDO,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_PROT_GIA_OUTROS_IMPOSTOS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna = 23 # DIFERENCA
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_VLR_CONTABIL-l_ENT_PROT_GIA_VLR_CONTABIL,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_VLR_BASE_ICMS-l_ENT_PROT_GIA_VLR_BASE_ICMS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_VLR_ICMS-l_ENT_PROT_GIA_VLR_ICMS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        
                        
                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_VLR_ISENTAS-l_ENT_PROT_GIA_VLR_ISENTAS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        
                       
                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_VLR_OUTRAS-l_ENT_PROT_GIA_VLR_OUTRAS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_SUBSTITUTO-l_ENT_PROT_GIA_SUBSTITUTO,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        
                        
                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_SUBSTITUIDO-l_ENT_PROT_GIA_SUBSTITUIDO,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_OUTROS_IMPOSTOS-l_ENT_PROT_GIA_OUTROS_IMPOSTOS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)    

                        # Sped - Regerado x Protocolado
                        l_coluna = 3
                        planilha2.cell(l_linha,l_coluna,round(l_ENT_REG_VLR_OPER,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"    
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)                    

                        l_coluna += 1
                        planilha2.cell(l_linha,l_coluna,round(l_ENT_REG_VLR_BASE_ICMS,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna += 1
                        planilha2.cell(l_linha,l_coluna,round(l_ENT_REG_VLR_ICMS,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna = 8 # PROTOCOLADO
                        planilha2.cell(l_linha,l_coluna,round(l_ENT_PROT_VLR_OPER,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna += 1
                        planilha2.cell(l_linha,l_coluna,round(l_ENT_PROT_VLR_BASE_ICMS,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna += 1
                        planilha2.cell(l_linha,l_coluna,round(l_ENT_PROT_VLR_ICMS,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna = 13 # DIFERENCA
                        planilha2.cell(l_linha,l_coluna,round(l_ENT_REG_VLR_OPER-l_ENT_PROT_VLR_OPER,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"    
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)                    

                        l_coluna += 1
                        planilha2.cell(l_linha,l_coluna,round(l_ENT_REG_VLR_BASE_ICMS-l_ENT_PROT_VLR_BASE_ICMS,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna += 1
                        planilha2.cell(l_linha,l_coluna,round(l_ENT_REG_VLR_ICMS-l_ENT_PROT_VLR_ICMS,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)  

                        # Sped x Gia_Regerado
                        # SPED
                        l_coluna = 3
                        planilha3.cell(l_linha,l_coluna,round(l_ENT_REG_VLR_OPER,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"    
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)                    

                        l_coluna += 1
                        planilha3.cell(l_linha,l_coluna,round(l_ENT_REG_VLR_BASE_ICMS,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna += 1
                        planilha3.cell(l_linha,l_coluna,round(l_ENT_REG_VLR_ICMS,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)  

                        # GIA
                        l_coluna = 8
                        planilha3.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_VLR_CONTABIL,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"    
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)                    

                        l_coluna += 1
                        planilha3.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_VLR_BASE_ICMS,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha3.cell(l_linha,l_coluna,round(l_ENT_REG_GIA_VLR_ICMS,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        # DIFERENCA
                        l_coluna = 13
                        planilha3.cell(l_linha,l_coluna,round(l_ENT_REG_VLR_OPER-l_ENT_REG_GIA_VLR_CONTABIL,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"    
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)                    

                        l_coluna += 1
                        planilha3.cell(l_linha,l_coluna,round(l_ENT_REG_VLR_BASE_ICMS-l_ENT_REG_GIA_VLR_BASE_ICMS,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna += 1
                        planilha3.cell(l_linha,l_coluna,round(l_ENT_REG_VLR_ICMS-l_ENT_REG_GIA_VLR_ICMS,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)  

                        # TOTAL DE SAIDA
                        l_linha += 1
                        l_coluna = 2 # Coluna B 
                        planilha1.cell(l_linha,l_coluna,"TOTAL SAIDA") # Coluna L # Gia (Protocolado)
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True) 
                        planilha1.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                        planilha2.cell(l_linha,l_coluna,"TOTAL SAIDA") # Coluna L # Gia (Protocolado)
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True) 
                        planilha2.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
                        planilha3.cell(l_linha,l_coluna,"TOTAL SAIDA") # Coluna L # Gia (Protocolado)
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True) 
                        planilha3.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')

                        # GIA - Regerado x Protocolado
                        l_coluna = 3
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_VLR_CONTABIL,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_VLR_BASE_ICMS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_VLR_ICMS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        
                        
                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_VLR_ISENTAS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        
                       
                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_VLR_OUTRAS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_SUBSTITUTO,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        
                        
                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_SUBSTITUIDO,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_OUTROS_IMPOSTOS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna = 13 # PROTOCOLADO
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_PROT_GIA_VLR_CONTABIL,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_PROT_GIA_VLR_BASE_ICMS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_PROT_GIA_VLR_ICMS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00" 
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                       

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_PROT_GIA_VLR_ISENTAS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_PROT_GIA_VLR_OUTRAS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_PROT_GIA_SUBSTITUTO,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_PROT_GIA_SUBSTITUIDO,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_PROT_GIA_OUTROS_IMPOSTOS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00" 
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)         

                        l_coluna = 23 # DIF
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_VLR_CONTABIL-l_SAI_PROT_GIA_VLR_CONTABIL,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00" 
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                       

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_VLR_BASE_ICMS-l_SAI_PROT_GIA_VLR_BASE_ICMS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00" 
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                       

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_VLR_ICMS-l_SAI_PROT_GIA_VLR_ICMS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00" 
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                       
                        
                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_VLR_ISENTAS-l_SAI_PROT_GIA_VLR_ISENTAS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        
                       
                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_VLR_OUTRAS-l_SAI_PROT_GIA_VLR_OUTRAS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_SUBSTITUTO-l_SAI_PROT_GIA_SUBSTITUTO,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                        
                        
                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_SUBSTITUIDO-l_SAI_PROT_GIA_SUBSTITUIDO,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00" 
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)                       

                        l_coluna += 1
                        planilha1.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_OUTROS_IMPOSTOS-l_SAI_PROT_GIA_OUTROS_IMPOSTOS,2)) 
                        planilha1.cell(l_linha,l_coluna).number_format = "#,##0.00" 
                        planilha1.cell(l_linha,l_coluna).font=Font(bold=True)  

                        # Sped - Regerado x Protocolado
                        l_coluna = 3
                        planilha2.cell(l_linha,l_coluna,round(l_SAI_REG_VLR_OPER,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"    
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)                    

                        l_coluna += 1
                        planilha2.cell(l_linha,l_coluna,round(l_SAI_REG_VLR_BASE_ICMS,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna += 1
                        planilha2.cell(l_linha,l_coluna,round(l_SAI_REG_VLR_ICMS,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna = 8 # PROTOCOLADO
                        planilha2.cell(l_linha,l_coluna,round(l_SAI_PROT_VLR_OPER,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna += 1
                        planilha2.cell(l_linha,l_coluna,round(l_SAI_PROT_VLR_BASE_ICMS,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna += 1
                        planilha2.cell(l_linha,l_coluna,round(l_SAI_PROT_VLR_ICMS,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna = 13 # DIFERENCA
                        planilha2.cell(l_linha,l_coluna,round(l_SAI_REG_VLR_OPER-l_SAI_PROT_VLR_OPER,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"    
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)                    

                        l_coluna += 1
                        planilha2.cell(l_linha,l_coluna,round(l_SAI_REG_VLR_BASE_ICMS-l_SAI_PROT_VLR_BASE_ICMS,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna += 1
                        planilha2.cell(l_linha,l_coluna,round(l_SAI_REG_VLR_ICMS-l_SAI_PROT_VLR_ICMS,2)) 
                        planilha2.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha2.cell(l_linha,l_coluna).font=Font(bold=True)  

                        # Sped x Gia_Regerado
                        # SPED
                        l_coluna = 3
                        planilha3.cell(l_linha,l_coluna,round(l_SAI_REG_VLR_OPER,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"    
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)                    

                        l_coluna += 1
                        planilha3.cell(l_linha,l_coluna,round(l_SAI_REG_VLR_BASE_ICMS,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna += 1
                        planilha3.cell(l_linha,l_coluna,round(l_SAI_REG_VLR_ICMS,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)             

                        # GIA
                        l_coluna = 8
                        planilha3.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_VLR_CONTABIL,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha3.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_VLR_BASE_ICMS,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)                        

                        l_coluna += 1
                        planilha3.cell(l_linha,l_coluna,round(l_SAI_REG_GIA_VLR_ICMS,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)     

                        l_coluna = 13 # DIFERENCA                                     
                        planilha3.cell(l_linha,l_coluna,round(l_SAI_REG_VLR_OPER-l_SAI_REG_GIA_VLR_CONTABIL,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"    
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)                    

                        l_coluna += 1
                        planilha3.cell(l_linha,l_coluna,round(l_SAI_REG_VLR_BASE_ICMS-l_SAI_REG_GIA_VLR_BASE_ICMS,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)  

                        l_coluna += 1
                        planilha3.cell(l_linha,l_coluna,round(l_SAI_REG_VLR_ICMS-l_SAI_REG_GIA_VLR_ICMS,2)) 
                        planilha3.cell(l_linha,l_coluna).number_format = "#,##0.00"
                        planilha3.cell(l_linha,l_coluna).font=Font(bold=True)   

                    """
                    if l_a1 == 1 and v_sped_enxertado is not None:
                        for linha in v_sped_enxertado:
                            planilha4.append(linha)

                    if l_a2 == 1 and v_sped_protocolado is not None:
                        for linha in v_sped_protocolado:
                            planilha5.append(linha)

                    if l_a3 == 1 and v_gia_enxertado is not None:
                        for linha in v_gia_enxertado:
                            planilha6.append(linha)

                    if l_a4 == 1 and v_gia_protocolado is not None:
                        for linha in v_gia_protocolado:
                            planilha7.append(linha)
                    """
                    # Lista a planilha de parametros
                    l_linha = 1
                    planilha8.cell(l_linha,1,"PARAMETROS - Insumo do CFOP, SPED GIA = 4989")

                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," DIRETORIO SPED FISCAL ENXERTADO: " + gv_diretorio_sped_fiscal_enxertado)
                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," DIRETORIO SPED FISCAL PROTOCOLADO: " + gv_diretorio_sped_fiscal_protocolado)
                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," DIRETORIO GIA ENXERTADO: " + gv_diretorio_gia_enxertado)
                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," DIRETORIO GIA PROTOCOLADO: " + gv_diretorio_gia_protocolado)
                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," DIRETORIO INSUMO SPED: " + gv_diretorio_insumo_sped)
                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," MASCARA ARQUIVO SPED FISCAL ENXERTADO: " + gv_arq_sped_fiscal_enxertado)
                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," MASCARA ARQUIVO SPED FISCAL PROTOCOLADO: " + gv_arq_sped_fiscal_protocolado)
                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," MASCARA ARQUIVO GIA ENXERTADO: " + gv_arq_gia_enxertado)
                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," MASCARA ARQUIVO GIA PROTOCOLADO: " + gv_arq_gia_protocolado)
                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," MASCARA ARQUIVO INSUMO SPED: " + gv_arq_insumo_sped)
                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," ARQUIVO SPED FISCAL ENXERTADO: " + gv_lst_arq_sped_fiscal_enxertado)
                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," ARQUIVO SPED FISCAL PROTOCOLADO: " + gv_lst_arq_sped_fiscal_protocolado)
                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," ARQUIVO GIA ENXERTADO: " + gv_lst_arq_gia_enxertado)
                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," ARQUIVO GIA PROTOCOLADO: " + gv_lst_arq_gia_protocolado)
                    l_linha = l_linha + 1
                    planilha8.cell(l_linha,1," ARQUIVO INSUMO SPED: " + gv_lst_arq_insumo_sped)


                    # Regulariza a dimensao das planilhas
                    for nColP in gv_lista_string:
                        planilha1.column_dimensions[str(nColP).upper()].width = 20   
                        planilha2.column_dimensions[str(nColP).upper()].width = 20   
                        planilha3.column_dimensions[str(nColP).upper()].width = 20   
                        # planilha4.column_dimensions[str(nColP).upper()].width = 20   
                        # planilha5.column_dimensions[str(nColP).upper()].width = 20   
                        # planilha6.column_dimensions[str(nColP).upper()].width = 20   
                        # planilha7.column_dimensions[str(nColP).upper()].width = 20   
                        planilha8.column_dimensions[str(nColP).upper()].width = 20   
                    
                    planilha1.column_dimensions['A'].width = 5 
                    planilha1.column_dimensions['AA'].width = 20
                    planilha1.column_dimensions['AA'].width = 20
                    planilha1.column_dimensions['AA'].width = 20
                    planilha1.column_dimensions['AA'].width = 20
                    planilha1.column_dimensions['AB'].width = 20
                    planilha1.column_dimensions['AC'].width = 20
                    planilha1.column_dimensions['AD'].width = 20
                    planilha1.column_dimensions['AE'].width = 20
                    planilha2.column_dimensions['A'].width = 5  
                    planilha3.column_dimensions['A'].width = 5  

                    # Grava a planilha Excel
                    arquivo_excel.save(gv_lst_arq_insumo_sped)

                    l_a5 = 1 # Processado
                    log(" >> Processado arquivo : " + gv_lst_arq_insumo_sped)
            
                except Exception as e:
                    txt = traceback.format_exc()
                    log(gv_lst_arq_insumo_sped + " >> ERRO NA GERACAO ARQUIVO.: " + str(e) + " - TRACE : " + txt)
                    l_ret = 1
                
        else:
            l_ret = 1
 
    # Codigo de Retorno
    return l_ret # Sucesso ou Erro

# Ponto de partida
if __name__ == "__main__" :
    
    # Codigo de Retorno
    ret = 0
    # Tratamento de excessao
    txt = ''
    # Conexao
    gv_conexao = None
    
    try:

        # Carrega os parametros do arquivo .cfg 
        config = comum.carregaConfiguracoes() 
        variaveis['config'] = config
        config = variaveis['config']
        log("-"*150)

        ### Cria os parametros do script 
        comum.addParametro( 'MESANO',  None, "MESANO (MMYYYY) dos arquivos", True, '122015' )
        comum.addParametro( 'UF', None, 'Unidade Federativa (UF) dos arquivo  ', True, 'SP')
        comum.addParametro( 'IE', None, 'Inscricao Estadual (IE) dos arquivo  ', True, '108383949112')

        # Validacao dos parametros de entrada
        if not comum.validarParametros() :
            ret = 91
        else:
            gv_mes_ano = comum.getParametro('MESANO').upper().strip()
            gv_uf = comum.getParametro('UF').upper().strip()
            gv_ie = comum.getParametro('IE').upper().strip()

            if (len(gv_mes_ano) != 6 
            ):
                log("PARAMETRO MES ANO: Invalido!") 
                ret = 91

            if not ret :
                gv_mes = gv_mes_ano[0:2]
                gv_ano = gv_mes_ano[2:6]

            if not ret :
                try:
                    if (int(gv_mes) < 1
                    or int(gv_mes) > 12 
                    ):
                        log("PARAMETRO MES : Invalido!") 
                        ret = 91
                except Exception as e:
                    log("PARAMETRO MES : Invalido!") 
                    ret = 91

            if not ret :
                try:
                    if (
                       int(gv_ano) > datetime.datetime.now().year
                    or int(gv_ano) < (datetime.datetime.now().year)-50
                    ):
                        log("PARAMETRO ANO : Invalido!") 
                        ret = 91
                except Exception as e:
                    log("PARAMETRO ANO : Invalido!") 
                    ret = 91
            
            if not ret :
                if len(gv_uf) != 2:
                    log("PARAMETRO UF: Invalido!") 
                    ret = 91

            if not ret :
                l_iei = re.sub('[^0-9]','',gv_ie)
                if ( (l_iei == "") or (l_iei == "''") or (l_iei == '""') or (int("0"+l_iei) == 0)):
                    log("PARAMETRO IE : Invalido!") 
                    ret = 91

        log("\n")

        # Verifica os parametros
        if not ret :
            try:
                gv_usuario = config['usuario'].strip()
                gv_senha = config['senha']
                gv_banco = config['banco'].strip()
                
                gv_diretorio_sped_fiscal_enxertado = config['sped_fiscal_enxertado'].replace("<<MM>>",gv_mes).replace("<<AAAA>>",gv_ano).replace("<<UF>>",gv_uf).replace("<<IE>>",gv_ie).replace("<<MESANO>>",gv_mes_ano).strip()
                gv_diretorio_sped_fiscal_protocolado = config['sped_fiscal_protocolado'].replace("<<MM>>",gv_mes).replace("<<AAAA>>",gv_ano).replace("<<UF>>",gv_uf).replace("<<IE>>",gv_ie).replace("<<MESANO>>",gv_mes_ano).strip()
                gv_diretorio_gia_enxertado = config['gia_enxertado'].replace("<<MM>>",gv_mes).replace("<<AAAA>>",gv_ano).replace("<<UF>>",gv_uf).replace("<<IE>>",gv_ie).replace("<<MESANO>>",gv_mes_ano).strip()
                gv_diretorio_gia_protocolado = config['gia_protocolado'].replace("<<MM>>",gv_mes).replace("<<AAAA>>",gv_ano).replace("<<UF>>",gv_uf).replace("<<IE>>",gv_ie).replace("<<MESANO>>",gv_mes_ano).strip()
                gv_diretorio_insumo_sped = config['insumo_sped'].replace("<<MM>>",gv_mes).replace("<<AAAA>>",gv_ano).replace("<<UF>>",gv_uf).replace("<<IE>>",gv_ie).replace("<<MESANO>>",gv_mes_ano).strip()
                
                gv_arq_sped_fiscal_enxertado = config['arq_sped_fiscal_enxertado'].replace("<<MM>>",gv_mes).replace("<<AAAA>>",gv_ano).replace("<<UF>>",gv_uf).replace("<<IE>>",gv_ie).replace("<<MESANO>>",gv_mes_ano).replace("<<NNN>>","*").strip()
                gv_arq_sped_fiscal_protocolado = config['arq_sped_fiscal_protocolado'].replace("<<MM>>",gv_mes).replace("<<AAAA>>",gv_ano).replace("<<UF>>",gv_uf).replace("<<IE>>",gv_ie).replace("<<MESANO>>",gv_mes_ano).replace("<<NNN>>","*").strip()
                gv_arq_gia_enxertado = config['arq_gia_enxertado'].replace("<<MM>>",gv_mes).replace("<<AAAA>>",gv_ano).replace("<<UF>>",gv_uf).replace("<<IE>>",gv_ie).replace("<<MESANO>>",gv_mes_ano).replace("<<NNN>>","*").strip()
                gv_arq_gia_protocolado = config['arq_gia_protocolado'].replace("<<MM>>",gv_mes).replace("<<AAAA>>",gv_ano).replace("<<UF>>",gv_uf).replace("<<IE>>",gv_ie).replace("<<MESANO>>",gv_mes_ano).replace("<<NNN>>","*").strip()
                gv_arq_insumo_sped = config['arq_insumo_sped'].replace("<<MM>>",gv_mes).replace("<<AAAA>>",gv_ano).replace("<<UF>>",gv_uf).replace("<<IE>>",gv_ie).replace("<<MESANO>>",gv_mes_ano).replace("<<NNN>>","*").strip()

            except Exception as e:
                log("PARAMETRO DOS ARQUIVOS NAO ENCONTRADOS!") 
                ret = 91        

        # Funcao conectar banco de dados
        if not ret :
            gv_conexao = conectar_BancoDados()
            if gv_conexao is None:
                log("<<ERRO>> CONEXAO BANCO DE DADOS") 
                ret = 91        
        
        # Funcao conectar banco de dados
        if not ret :
            l_str_uf = retornar_UF(gv_ie,gv_conexao) 
            log("UF BD:"+l_str_uf)
            if (gv_uf.upper().strip() != l_str_uf.upper().strip()):
                log("PARAMETRO UF Invalido de acordo IE!") 
                ret = 91

        # Verifica a existencia dos diretorios
        
        if not ret :
            try:
                l_a1 = 0
                l_a2 = 0
                l_a3 = 0
                l_a4 = 0
                l_a5 = 0

                gv_lst_arq_sped_fiscal_enxertado = ""
                gv_lst_arq_sped_fiscal_protocolado = ""
                gv_lst_arq_gia_enxertado = ""
                gv_lst_arq_gia_protocolado = ""
                gv_lst_arq_insumo_sped = ""
                    
                if not os.path.isdir(gv_diretorio_sped_fiscal_enxertado):
                    log("Diretório não localizado : " + gv_diretorio_sped_fiscal_enxertado)        
                    l_a1 = 1
                else:    
                    gv_lst_arq_sped_fiscal_enxertado = ultimo_Arquivo_Diretorio(gv_arq_sped_fiscal_enxertado,gv_diretorio_sped_fiscal_enxertado).strip()
                
                if not os.path.isdir(gv_diretorio_sped_fiscal_protocolado):
                    log("Diretório não localizado : " + gv_diretorio_sped_fiscal_protocolado)        
                    l_a2 = 1
                else:
                    gv_lst_arq_sped_fiscal_protocolado = ultimo_Arquivo_Diretorio(gv_arq_sped_fiscal_protocolado,gv_diretorio_sped_fiscal_protocolado).strip()
                
                if not os.path.isdir(gv_diretorio_gia_enxertado):
                    log("Diretório não localizado : " + gv_diretorio_gia_enxertado)        
                    l_a3 = 1
                else:
                    gv_lst_arq_gia_enxertado = ultimo_Arquivo_Diretorio(gv_arq_gia_enxertado,gv_diretorio_gia_enxertado).strip()
                
                if not os.path.isdir(gv_diretorio_gia_protocolado):
                    log("Diretório não localizado : " + gv_diretorio_gia_protocolado)        
                    l_a4 = 1
                else:
                    gv_lst_arq_gia_protocolado = ultimo_Arquivo_Diretorio(gv_arq_gia_protocolado,gv_diretorio_gia_protocolado).strip()
                
                if not os.path.isdir(gv_diretorio_insumo_sped):
                    log("Diretório não localizado : " + gv_diretorio_insumo_sped)        
                    l_a5 = 1
                else:
                    gv_lst_arq_insumo_sped = ultimo_Arquivo_Diretorio(gv_arq_insumo_sped,gv_diretorio_insumo_sped).strip()
                    if len(gv_lst_arq_insumo_sped) < 4:
                        gv_lst_arq_insumo_sped = os.path.join(gv_diretorio_insumo_sped, gv_arq_insumo_sped.replace("*","001"))
                    else:
                        if (gv_lst_arq_insumo_sped.strip().upper().endswith(".TXT")
                        or gv_lst_arq_insumo_sped.strip().upper().find(".")):
                            gv_lst_arq_insumo_sped = gv_lst_arq_insumo_sped.split(".")[0][:-3]+ str(int(gv_lst_arq_insumo_sped.split(".")[0][-3:])+1).rjust(3,'0') +  "." +gv_lst_arq_insumo_sped.split(".")[1] 
                        else:
                            gv_lst_arq_insumo_sped = os.path.join(gv_diretorio_insumo_sped, gv_arq_insumo_sped.replace("*","001"))

            except Exception as e:
                log("INVALIDO DIRETORIO DOS ARQUIVOS CONFIGURADOS!") 
                ret = 91     

            if not ret :
                if ((l_a1 == 1 or len(gv_lst_arq_sped_fiscal_enxertado) < 4) 
                and (l_a2 == 1 or len(gv_lst_arq_sped_fiscal_protocolado) < 4)
                and (l_a3 == 1 or len(gv_lst_arq_gia_enxertado) < 4)
                and (l_a4 == 1 or len(gv_lst_arq_gia_protocolado) < 4)
                ):
                    ret = 91
                    log("NÃO EXISTEM ARQUIVOS A SEREM PROCESSADOS!") 
                else:
                    if l_a5 == 1:
                        os.makedirs(gv_diretorio_insumo_sped)
                        log("Diretório criado : " + gv_diretorio_insumo_sped)        
                        gv_lst_arq_insumo_sped = os.path.join(gv_diretorio_insumo_sped, gv_arq_insumo_sped.replace("*","001"))

                    try:    
                        log("FUNCAO PRINCIPAL:INICIO")    
                        ret = main(gv_conexao)        
                        log("FUNCAO PRINCIPAL:FIM >> " + str(ret))
                    except Exception as e:
                        txt = traceback.format_exc()
                        log("ERRO FUNCAO PRINCIPAL .: " + str(e))
                        ret = 91

        log("\n")

        # Fechar a Conexao
        fechar = fechar_ConexaoBancoDados(gv_conexao)
        log("STATUS FECHADO:" + str(fechar))

        # Finalizacao
        log("\n")            
        
        if not ret :
            log("SUCESSO")
        else:
            log("ERRO")

        log("\n")
    
    except Exception as e:
        txt = traceback.format_exc()
        log("ERRO .: " + str(e))
        ret = 93
    
    sys.exit(ret if ret >= log.ret else log.ret )
