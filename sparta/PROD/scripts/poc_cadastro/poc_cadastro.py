#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-
"""
----------------------------------------------------------------------------------------------
MODULO ...: TESHUVA
SCRIPT ...: poc_ba.py
CRIACAO ..: 02/02/2022
AUTOR ....: Victor Santos / Welber Pena - KYROS TECNOLOGIA
DESCRICAO.: 
----------------------------------------------------------------------------------------------
  HISTORICO : 
----------------------------------------------------------------------------------------------
"""
from calendar import month
import os
import sys

global SD, dir_base
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)

import configuracoes
import comum
import sql
import util
from openpyxl import Workbook, load_workbook
import openpyxl
import datetime
import shutil
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook 
from copy import copy

import converte_csv_to_json

SD = ('/' if os.name == 'posix' else '\\')

global g_relatorio_saida

g_dic_resumo = {}

def preencher_aba_percent_dif(p_registros):

    g_ix_linha = 1

    global g_dic_resumo

    v_aba = g_relatorio_saida["Percentuais diferença"]

    g_layout_aba_dif = {
         'TIPO DE BATIMENTO'                   :1                         
        ,'EMPRESA'                             :2 
        ,'FILIAL'                              :3
        ,'SERIE'                               :4
        ,'MES REFERENCIA'                      :5      
        ,'TOTAL NFs ULTIMO PROT'               :6                     
        ,'ARQUIVO'                             :7 
        ,'DESCRICAO CAMPO'                     :8         
        ,'CATEGORIA'                           :9   
        ,'INDICADOR_OUTLIER'                   :10             
        ,'TOTAL DIFERENCA'                     :11         
        ,'DIFERENCA PERC.'                     :12        
        ,'TOTAL BASE CALCULO ICMS'             :13                 
        ,'BASE CALCULO ICMS DIFERENCA'         :14                     
        ,'BASE CALCULO ICMS DIF.PERCENTUAL'    :15                         
        ,'TOTAL DE ICMS'                       :16         
        ,'ICMS DIFERENCA'                      :17         
        ,'ICMS DIFERENCA PERCENTUAL'           :18                     
        ,'TOTAL NFs DO CONV39'                 :19            
        ,'TOTAL NFs CONV39 DIFERENTE'          :20                     
        ,'CONV39 DIF PERCENTUAL'               :21                     
        ,'TOTAL DO ESTORNO DO CONV39'          :22                     
        ,'CONV39 DIFERENCA'                    :23         
        ,'CONV39 DIFERENCA PERCENTUAL'         :24                     
    }


    for registro in p_registros:
        
        g_ix_linha += 1
        
        chave = {}

        valor = 0

        indicador = ''
        
        for coluna in g_layout_aba_dif.keys():

            if coluna == 'DIFERENCA PERC.':
                valor = float(registro[coluna].replace('%','').replace("'",'').replace(",",'.'))

            if coluna == 'INDICADOR_OUTLIER':
                indicador = registro[coluna]
            
            if coluna in ['FILIAL','SERIE','MES REFERENCIA','CATEGORIA']:
                chave[coluna] = registro[coluna]

            col = g_layout_aba_dif[coluna]

            celula = v_aba.cell(g_ix_linha,col,registro[coluna])                    

            if g_ix_linha > 2:
                column_letter = get_column_letter(col)
                cell = v_aba[column_letter + str(2)]

                celula._style = copy(cell._style)               

        dados_chave = (chave['FILIAL'],chave['SERIE'],chave['CATEGORIA'],chave['MES REFERENCIA'])

        if indicador == 'N':  

            if g_dic_resumo.get(dados_chave, -1) < valor:
                g_dic_resumo[dados_chave] = valor
        
        else:

            if g_dic_resumo.get(dados_chave, -1):
                g_dic_resumo[dados_chave] = 0.0

            
            
            


    return True

def preencher_aba_amostra(p_registros):
    
    g_ix_linha = 1

    v_aba = g_relatorio_saida["Amostragem"]

    g_layout_aba_amostra = {
         'EMPRESA'                                   :1            
        ,'FILIAL'                                    :2          
        ,'SERIE'                                     :3          
        ,'DATA_EMISSAO'                              :4                 
        ,'NUMERO_NF'                                 :5              
        ,'BASE DE CALCULO ICMS'                      :6                         
        ,'VALOR ICMS'                                :7               
        ,'VALOR ESTORNO CONV39'                      :8                         
        ,'FLAG_ENCONTROU_BILLING'                    :9                           
        ,'CNPJ_IMPRESSAO'                            :10                   
        ,'CNPJ_ULTIMO_PROTOCOLADO'                   :11                            
        ,'FLAG_CNPJ'                                 :12         
        ,'IE_IMPRESSAO'                              :13                 
        ,'IE_ULTIMO_PROTOCOLADO'                     :14                          
        ,'FLAG_IE'                                   :15         
        ,'RAZAOSOCIAL_IMPRESSAO'                     :16                          
        ,'RAZAOSOCIAL_ULTIMO_PROTOCOLADO'            :17                                   
        ,'FLAG_RAZAO'                                :18         
        ,'ENDERECO_IMPRESSAO'                        :19                       
        ,'ENDERECO_ULTIMO_PROTOCOLADO'               :20                               
        ,'FLAG_END'                                  :21         
        ,'NUMERO_IMPRESSAO'                          :22                     
        ,'NUMERO_ULTIMO_PROTOCOLADO'                 :23                             
        ,'FLAG_NUM'                                  :24         
        ,'COMPLEMENTO_IMPRESSAO'                     :25                          
        ,'COMPLEMENTO_ULTIMO_PROTOCOLADO'            :26                                   
        ,'FLAG_COMPL'                                :27         
        ,'CEP_IMPRESSAO'                             :28                  
        ,'CEP_ULTIMO_PROTOCOLADO'                    :29                           
        ,'FLAG_CEP'                                  :30         
        ,'BAIRRO_IMPRESSAO'                          :31                     
        ,'BAIRRO_ULTIMO_PROTOCOLADO'                 :32                              
        ,'FLAG_BAIRRO'                               :33         
        ,'MUNICIPIO_IMPRESSAO'                       :34                        
        ,'MUNICIPIO_ULTIMO_PROTOCOLADO'              :35                                 
        ,'FLAG_MUN'                                  :36         
        ,'UF_IMPRESSAO'                              :37                 
        ,'UF_ULTIMO_PROTOCOLADO'                     :38                          
        ,'FLAG_UF'                                   :39         
        ,'TELEFONECONTATO_IMPRESSAO'                 :40                              
        ,'TELEFONECONTATO_ULTIMO_PROTOCOLADO'        :41                                       
        ,'FLAG_TEL'                                  :42         
        ,'CODIDENTCONSUMIDOR_IMPRESSAO'              :43                                 
        ,'CODIDENTCONSUMIDOR_ULTIMO_PROTOCOLADO'     :44                                          
        ,'FLAG_CODI'                                 :45         
        ,'NUMEROTERMINAL_IMPRESSAO'                  :46                             
        ,'NUMEROTERMINAL_ULTIMO_PROTOCOLADO'         :47                                      
        ,'FLAG_NUMT'                                 :48         
        ,'UFHABILITACAO_IMPRESSAO'                   :49                            
        ,'UFHABILITACAO_ULTIMO_PROTOCOLADO'          :50                                     
        ,'FLAG_UFH'                                  :51         
        ,'CODIGOMUNICIPIO_IMPRESSAO'                 :52                              
        ,'CODIGOMUNICIPIO_ULTIMO_PROTOCOLADO'        :53                                       
        ,'FLAGCODMUN'                                :54         
        ,'CNPJ_IMPRESSAO_C'                          :55                   
        ,'CNPJ_ULTIMO_PROTOCOLADO_C'                 :56                           
        ,'FLAG_CNPJC'                                :57        
        ,'IE_IMPRESSAO_C'                            :58                 
        ,'IE_ULTIMO_PROTOCOLADO_C'                   :59                          
        ,'FLAG_IEC'                                  :60         
        ,'RAZAOSOCIAL_IMPRESSAO_C'                   :61                          
        ,'RAZAOSOCIAL_ULTIMO_PROTOCOLADO_C'          :62                                   
        ,'FLAG_RAZAOC'                               :63         
        ,'UF_IMPRESSAO_C'                            :64                 
        ,'UF_ULTIMO_PROTOCOLADO_C'                   :65                          
        ,'FLAG_UFC'                                  :66         
        ,'CADG_COD_IMPRESSAO'                        :67                     
        ,'CADG_COD_ULTIMO_PROTOCOLADO'               :68                                
        ,'FLAG_CADG'                                 :69         
        ,'TERMINAL_TELEF_IMPRESSAO'                  :70                             
        ,'TERMINAL_TELEF_ULTIMO_PROTOCOLADO'         :71                                      
        ,'FLAG_TERMINAL'                             :72         
        ,'TIPO_PESSOA_IMPRESSAO'                     :73                          
        ,'TIPO_PESSOA_ULTIMO_PROTOCOLADO'            :74                                   
        ,'FLAG_TIPO'                                 :75         
        ,'TERMINAL_PRINC_IMPRESSAO'                  :76                             
        ,'TERMINAL_PRINC_ULTIMO_PROTOCOLADO'         :77                                      
        ,'FLAG_TERMINALP'                            :78                 
    }

    for registro in p_registros:

        g_ix_linha += 1

        for coluna in g_layout_aba_amostra.keys():

            col = g_layout_aba_amostra[coluna]
            
            celula = v_aba.cell(g_ix_linha,col,registro[coluna])                    

            if g_ix_linha > 2:
                column_letter = get_column_letter(col)
                cell = v_aba[column_letter + str(2)]
                celula._style = copy(cell._style)               
               
    return True

def preencher_aba_resumo():
    
    global g_dic_resumo

    v_aba = g_relatorio_saida["Resumo"]  

    g_ix_linha = 2

    for dados_chave in g_dic_resumo:

        g_ix_linha += 1

        for col in range(1,6):

            if col < 5:
                celula = v_aba.cell(g_ix_linha,col,dados_chave[col - 1])                    
            else:
                celula = v_aba.cell(g_ix_linha,col,'%.2f%%'%g_dic_resumo[dados_chave])
            
            if g_ix_linha > 2:
                column_letter = get_column_letter(col)
                cell = v_aba[column_letter + str(3)]
                celula._style = copy(cell._style)               
               
    return True

def ultimodia(ano,mes):
       return(31 if mes == 12 else datetime.date.fromordinal((datetime.date(ano,mes+1,1)).toordinal()-1).day)

def processa_relatorio(p_uf,p_dt_ini,p_dt_fim,p_serie):
     
    connection = sql.geraCnxBD(configuracoes)
    cd_erro = connection.var(int)
    ds_erro = connection.var(str)
    p_limite = 'S'
    p_origem = ''
    procedure  = "gfcadastro.TSH_SANTL_REL_POC_CADASTRO" 
    parametros = [  
                p_uf,     
                p_dt_ini, 
                p_dt_fim, 
                p_limite, 
                p_origem, 
                p_serie,  
                cd_erro,
                ds_erro,
            ]
    connection.executaProcedure(procedure, *parametros)
    return [cd_erro.getvalue(), ds_erro.getvalue()]

def processar() :

    global g_dic_resumo
    
    global g_relatorio_saida

    uf = comum.getParametro('UF')
    if util.validauf(uf):
        data_ini = comum.getParametro('DATA_INI')
        data_fim = comum.getParametro('DATA_FIM')
        serie    = comum.getParametro('SERIE')

        datai = datetime.datetime(int(data_ini[4:]),int(data_ini[2:4]),int(data_ini[:2]))
        dataf = datetime.datetime(int(data_fim[4:]),int(data_fim[2:4]),int(data_fim[:2]))

    else:
        comum.imprimeHelp()
        ret = 91
        return ret

    data_atua = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    
    log(' Iniciando a geração do relatório, aguarde... '.center(100,'='))
    
    cd_erro,ds_erro = processa_relatorio(uf,datai,dataf,serie)
    
    log ('CODIGO      -> ', cd_erro)
    log ('DESCRIÇÃO   -> ', ds_erro)

    # cd_erro = 0 
    # ds_erro = 'SUCESSO'    

    if cd_erro == 0 and ds_erro == 'SUCESSO':
        
        dic_ret = converte_csv_to_json.converter(uf, data_ini, configuracoes.arq_resumo)

        if dic_ret['status'] == 'Ok':
            v_regs_arquivo_resumo = dic_ret['dados']
            dic_ret = converte_csv_to_json.converter(uf, data_ini, configuracoes.arq_amostra)
        
        if dic_ret['status'] == 'Ok' : 
            v_regs_arquivo_amostra = dic_ret['dados']    

        if dic_ret['status'] == 'Ok' :
            g_relatorio_saida = load_workbook(filename='Template_relatorio_POC_v2.xlsx',keep_vba=True) 
            
            log(' Escrevendo dados na planilha - Percentual de Diferenças '.center(100,'='))
            log('Quantidade de linhas:', len(v_regs_arquivo_resumo))
            preencher_aba_percent_dif(v_regs_arquivo_resumo)

            log(' Escrevendo dados na planilha - Amostras '.center(100,'='))
            log('Quantidade de linhas:', len(v_regs_arquivo_amostra))
            preencher_aba_amostra(v_regs_arquivo_amostra)

            log(' Escrevendo dados na planilha - Resumo '.center(100,'='))
            log('Quantidade de linhas:', len(g_dic_resumo))
            preencher_aba_resumo()
            
            # v_diretorio = configuracoes.v_caminho_relatorio
            v_diretorio = configuracoes.dir_geracao_arquivos
            if not os.path.isdir(v_diretorio) :
                os.makedirs(v_diretorio)

            v_dir_base  = os.path.join(v_diretorio)    
            v_nome_rel = 'Qualidade_Dados_Cadastrais_'+ uf + '_' + data_ini + '_' + data_fim + '_' + data_atua + '.xlsm'
            
            if not os.path.isdir(v_dir_base) :
                log(' Criando diretótrio'.center(100,'='))
                log(v_dir_base)
                os.makedirs(v_dir_base)

            v_nome_relatorio = os.path.join(v_dir_base, v_nome_rel)

            log('Salvando PLANILHA'.center(100,'='))
            g_relatorio_saida.save(v_nome_relatorio)
            log(' Caminho ', v_nome_relatorio.center(100,'='))

        else:
            log('ERRO')
            return False

    else:
        log('ERRO NA CHAMADA DA PROCEDURE')
        return False    

    return True

if __name__ == "__main__" :
    
    ret = 0
    
    comum.addParametro( 'UF'       , None, "UNIDADE FEDERATIVA" , True , 'SP' )
    comum.addParametro( 'DATA_INI' , None, "DATA INICIAL"       , True , '01/01/2015')
    comum.addParametro( 'DATA_FIM' , None, "DATA FINAL"         , True , '31/01/2015')
    comum.addParametro( 'SERIE'    , None, "SERIE PROCESSADA"   ,False , 'C')

    if not comum.validarParametros() :
        log('### ERRO AO VALIDAR OS PARÂMETROS')
        ret = 91
    else:
        # configuracoes.uf       = comum.getParametro('UF').upper()
        # configuracoes.data_ini = comum.getParametro('DATA_INI')
        # configuracoes.data_fim = comum.getParametro('DATA_FIM')
        # configuracoes.serie    = comum.getParametro('SERIE').upper()

        if not processar() :
            log('ERRO no processamento do relatorio !')
            ret = 92

    sys.exit(ret)