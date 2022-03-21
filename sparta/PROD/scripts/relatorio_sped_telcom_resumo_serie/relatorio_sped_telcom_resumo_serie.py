#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-
"""
----------------------------------------------------------------------------------------------
MODULO ...: TESHUVA
SCRIPT ...: relatorio_sped_telcom_resumo_serie.py
CRIACAO ..: 06/10/2020
AUTOR ....: Airton Borges da Silva Filho / KYROS Consultoria
DESCRICAO.: Execução da exportacao do levantamento realizado para a mercadoria
----------------------------------------------------------------------------------------------
PARAMETROS: 
Parâmetros de entrada:
1 ) P_CC_UF              : [OBRIGATÓRIO - FORMATO AA    ] Unidade Federativa do Brasil. Exemplo. SP
2 ) P_CC_MESANO          : [OBRIGATÓRIO - FORMATO MMYYYY] Mes e Ano de processamento. Exemplo: 012015
3 ) P_CC_IE              : [OPCIONAL] Inscricao Estadual . Exemplo: 108383949112

UF: Formato AA (Obrigatório)
MÊS/ANO: Formato MMAAAA (Obrigatório)
IE: (Opcional)

----------------------------------------------------------------------------------------------
    HISTORICO : 
        * 06/10/2020 - Airton Borges da Silva Filho / KYROS Consultoria
        - Criacao do script.
        * 15/10/2020 - Airton /portaloptrib/TESHUVA/sparta/PROD/scripts/Insumos_SPED/relatorio_sped_telcom_resumo_serie/Borges da Silva Filho / KYROS Consultoria
        - Acertando arredondamento da aba resumo
        * 09/12/2020 
        - Incluido verificação de aspas no IE, incluido a biblioteca re.
        * 09/09/2021 - EDUARDO DA SILVA FERREIRA / KYROS TECNOLOGIA (eduardof@kyros.com.br)
        - PTITES-144: DV - Novo Padrão: INSUMOS SPED - Relatório SPED Telecom – Resumo por Série
        - Alterado nome para : relatorio_sped_telcom_resumo_serie.py
        Relatório SPED Telecom – Resumo por Série    
        
        - ALT001 - Flavio Teixeira - 03/01/2022:
          Ajustes para: - Quando uma serie possuir mais de um codigo de filial para mesma 
                          uf filial, considerar a versao com indicacao de retificacao S
----------------------------------------------------------------------------------------------
"""

import sys
import os
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)
import configuracoes
import comum
import util
import sql
import datetime
import traceback

from openpyxl import Workbook
from pathlib import Path
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_MEDIUM, BORDER_DOUBLE
from openpyxl.utils import get_column_letter

# Demais configuração
comum.carregaConfiguracoes(configuracoes)
log.gerar_log_em_arquivo = True


def retiraespacos(frase):
    retorno = "" 
    for l in frase:
        if (l != " "):
            retorno = retorno + l
    return(retorno)

    
def buscasomaCFOP(serie,valoresCFOP):
    elemento = 0
    for linha in valoresCFOP:
        if (retiraespacos(linha[0]) == retiraespacos(serie)):
            return(linha[1],elemento)
        elemento = elemento + 1    
    return(0.00,-1)


def adjust_column(ws, min_row, min_col, max_col):
    
    column_widths = []

    for i, col in \
        enumerate(
            ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
        ):

        for cell in col:
            value = cell.value
            if value is not None:

                if isinstance(value, str) is False:
                    value = str(value)

                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))

    for i, width in enumerate(column_widths):

        col_name = get_column_letter(min_col + i)
        value = column_widths[i] + 2
        value = int(round(value + value * 0.2,0))
        ws.column_dimensions[col_name].width = value


def formata_resumo(planilha):
    
    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
    uc  = planilha.max_column
    ul  = planilha.max_row
    sul = str(ul)
    
    # LINHA 1 = TITULO
    for colatu in range(1,uc+1):
        planilha.cell(1,colatu).font = fontMasterPreta
        planilha.cell(1,colatu).alignment = Alignment(horizontal='center')

        planilha.cell(2,colatu).font = fontMasterPreta
        planilha.cell(2,colatu).alignment = Alignment(horizontal='center')
    
    
    planilha.merge_cells('C1:E1')
    planilha.merge_cells('F1:H1')
    planilha.merge_cells('I1:K1')
    planilha.merge_cells('L1:N1')
    planilha.merge_cells('O1:Q1')

    planilha.freeze_panes = 'A3'
    # GERAL

#   adjust_column(planilha, primeira_linha,primeira_coluna, ultima_colunac)
    adjust_column(planilha, 2,1, uc)
    
    # LINHA 3 Até ÚLTIMA = DADOS
    pl = 3
    pc = 3
    
    for linha in range(pl,ul+1):
        for coluna in range(pc,uc+1):
            planilha.cell(linha, coluna).number_format = "#,##0.00"
  
    # LINHA FINAL = TOTAIS
    for coluna in range(1,uc+1):
        planilha.cell(ul,  coluna).alignment = Alignment(horizontal='center')
        planilha.cell(ul,  coluna).font = fontMasterPreta

#    planilha.merge_cells('A'+sul+':B'+sul)

    
     #BORDAS    

    set_border_edsi(planilha, 'B1:Q2' , 'mmmm')
    set_border_edsi(planilha, 'B3:Q'+str(ul-1) , 'ffff')
    set_border_edsi(planilha, 'B'+sul+':Q'+sul , 'mmmm')
    
    
    #COR DE FUNDO
    
    #http://2.bp.blogspot.com/-t8bYkpBVvVs/UF_7KatWLWI/AAAAAAAAAEc/Pt-ZZ1zYDEk/s1600/color_chart_prnt.gif
    #https://pbs.twimg.com/media/A4KckTBCAAAK0mh.png
    azul1='66FFFF'
    azul2='66CCFF'
    azul3='6699FF'
    
    verde1='99FFCC'
    verde2='99FF99'
    verde3='99CC33'
    
    amarelo1='FFFFCC'
    amarelo2='FFFF66'
    amarelo3='FFCC66'

    vermelho1='FFCCCC'
    vermelho2='FF9999'
    vermelho3='FF6666'
    
#   cor_fundo(planilha, c_i, l_i, c_f, l_f, cor)
    
    
    cor_fundo(planilha,2,2,3,planilha.max_row+1,'FF99CC')
    
    cor_fundo(planilha,3,2,6,planilha.max_row+1,'00CC99')
    
    cor_fundo(planilha,6,2,9,planilha.max_row+1,'66CCCC')
    
    cor_fundo(planilha,9,2,12,planilha.max_row+1,'99CC99')

    cor_fundo(planilha,12,2,15,planilha.max_row+1,'CCCCFF')

    cor_fundo(planilha,15,2,18,planilha.max_row+1,'FF6666')


def formata_valor(planilha,coluna):
    for linha in range(1,planilha.max_row+1):    
        planilha.cell(linha, coluna).number_format = "#,##0.00"

def cor_fundo(planilha,colini,linini,colfim,linfim,cor):
    corfundo = PatternFill(start_color=cor,end_color=cor,fill_type='solid')
    for linha in range(linini,linfim):  
        for coluna in range(colini,colfim):
            planilha.cell(linha, coluna).fill = corfundo

def formata_comum(planilha):
    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
    uc  = planilha.max_column
    ul  = planilha.max_row
    sul = str(ul)
    lucol = get_column_letter(planilha.max_column)
    
    # LINHA 1 = TITULO
# =============================================================================
#     for colatu in range(1,uc):
#         for linatu in (1,2):
#             planilha.cell(linatu,colatu).font = fontMasterPreta
#             planilha.cell(linatu,colatu).alignment = Alignment(horizontal='center')
#     planilha.freeze_panes = 'A3'
#     for colatu in range(1,uc):
# =============================================================================
    for colatu in range(1,uc+1):
        planilha.cell(1,colatu).font = fontMasterPreta
        planilha.cell(1,colatu).alignment = Alignment(horizontal='center')
    planilha.freeze_panes = 'A2'

    # TAMANHO DAS COLUNAS
    #   adjust_column(planilha, primeira_linha,primeira_coluna, ultima_colunac)
    adjust_column(planilha, 1,1, uc)

    # BORDAS
    set_border_edsi(planilha, 'A1:'+lucol+str(ul) , 'ffff')


def set_border_edsi( ws, cell_range, edsi):
    b=[]
 
    n=Side(border_style= None    , color='00000000')
    f=Side(border_style='thin'   , color='00000000')
    m=Side(border_style='medium' , color='00000000')
    d=Side(border_style='double' , color='00000000')
    p=Side(border_style='dotted' , color='00000000')
    t=Side(border_style='dashed' , color='00000000')
    g=Side(border_style='thick'  , color='00000000')
    
    for x in (0,1,2,3):
        if (edsi[x]=="n"):
            b.append(n)
        elif (edsi[x]=='f'):
            b.append(f)
        elif (edsi[x]=='m'):
            b.append(m)
        elif (edsi[x]=='d'):
            b.append(d)
        elif (edsi[x]=='p'):
            b.append(p)
        elif (edsi[x]=='t'):
            b.append(t)
        elif (edsi[x]=='g'):
            b.append(g)
        else:
            return
    
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = Border(left=b[0],right=b[1],top=b[2],bottom=b[3])

def ob_global():
    """
        Funcao falsa apenas para sergir de apoio para armazenar as variaveis globais 
    """
    pass

def fnc_log():
    """
        Funcao apenas para logar as variaveis principais
    """
    try:
        log("-"*150)
      
        try:
            log("dir_base                   " , dir_base                   )    
        except:
            pass
        
        try:
            if ob_global.gv_ob_conexao is not None:
                log("conexao              " , "ATIVO"              )           
        except:
            pass
        
        try:
            log("uf                   " , ob_global.gv_cc_uf                   )    
        except:
            pass
        
        try:
            log("mesano               " , ob_global.gv_cc_mesano               )
        except:
            pass
        
        try:
            log("ie                   " , ob_global.gv_cc_ie                   )
        except:
            pass
        
        try:
            log("dir_protocolados     " , ob_global.dir_protocolados           )
        except:
            pass
        
        try:
            log("dir_regerados        " , ob_global.dir_regerados              )
        except:
            pass
        
        try:
            log("dir_relatorios       " , ob_global.dir_relatorios             )
        except:
            pass
        
        try:
            log("mascara_regerado     " , ob_global.mascara_regerado           )
        except:
            pass
        
        try:
            log("mascara_protocolado  " , ob_global.mascara_protocolado        )
        except:
            pass
        
        try:
            log("nome_relatorio       " , ob_global.nome_relatorio             )
        except:
            pass
            
        try:
            log("lista_ie             " , str(ob_global.gv_ob_lista_ie)        )
        except:
            pass
        
        try:
            log("dic_arquivos         " , str(ob_global.gv_ob_dic_arquivo)     )
        except:
            pass
            
        log("-"*150)    
            
    except:
        pass

def processa_SPED_Regerado(arquivo,ie,uf,dtini,dtfim):
    """
        Funcao processa dados SPED Regerado
    """
    regrel=[[],[]]
    regrel[0]=["Modelo",
                "Série",
                "NF_Ini",
                "NF_Fim",
                "Dt_Ini",
                "Dt_Fim",
                "Nome_Arq",
                "Hashcode",
                # "Contábil_SPED",
                "Liquido_SPED",
                # "ContábilCFOP0000",
                "LiquidoCFOP0000",
                # "Valor_contábil_total",
                "Valor_liquido_total",
                "Base",
                "ICMS"]

    query = """
    SELECT  /*+ parallel(8) */
        infst_serie,
        SUM(
            CASE
                WHEN infst_dtemiss < TO_DATE('01/01/2017', 'dd/mm/yyyy') THEN
                    nvl(infst_val_serv, 0) - nvl(infst_val_desc, 0)
                ELSE
                    nvl(infst_val_cont, 0)
            END
        ) AS valor_contabil
    FROM
        openrisow.item_nftl_serv
    WHERE
            1 = 1
        AND infst_dtemiss >= to_date('%s','ddmmyyyy')--MANTER PARAMETRO ANTERIOR
        AND infst_dtemiss <= to_date('%s','ddmmyyyy')--MANTER PARAMETRO ANTERIOR
        AND emps_cod = 'TBRA'
        AND fili_cod IN (
            SELECT fili_cod
              FROM openrisow.filial
             WHERE 1 = 1
               AND emps_cod = 'TBRA'
               AND fili_cod_insest = '%s' --MANTER PARAMETRO ANTERIOR
               AND unfe_sig        = '%s')--MANTER PARAMETRO ANTERIOR
        AND infst_ind_canc = 'N'
        AND cfop = '0000'
    GROUP BY
        infst_serie
    """ %(dtini,dtfim,ie,uf)
    
    ob_global.gv_ob_conexao.executa(query)
    somaCFOP=[]
    result = ob_global.gv_ob_conexao.fetchone()

    if result == None or len(str(result)) == 0:
        log("#### ATENÇÃO: Nenhum Resultado para SPED_Regerado")
        log("####     Query = ")
        log("####")
        log(query)
        log("####")
        return somaCFOP
  
    while result:
        somaCFOP.append(result)
        result = ob_global.gv_ob_conexao.fetchone()
        
    nregrel = 0   
    ent = open(arquivo,mode="r",encoding=comum.encodingDoArquivo(arquivo))
    linhalida = ent.readline()
    while linhalida:
        dados_ent = linhalida.split("|")
        tp_reg = ""
        if (len(dados_ent) >=2):
            tp_reg = dados_ent[1]
        if tp_reg == "D695":
            nregrel = nregrel + 1
            regrel.append([])
            regrel[nregrel].append(dados_ent[2])
            regrel[nregrel].append(dados_ent[3])
            regrel[nregrel].append(dados_ent[4])
            regrel[nregrel].append(dados_ent[5])
            regrel[nregrel].append(dados_ent[6])
            regrel[nregrel].append(dados_ent[7])
            regrel[nregrel].append(dados_ent[8])
            regrel[nregrel].append(dados_ent[9])
            regrel[nregrel].append(0.00)
            regrel[nregrel].append(0.00)
            regrel[nregrel].append(0.00)
            regrel[nregrel].append(0.00)
            regrel[nregrel].append(0.00)
        elif tp_reg == "D696":
            regrel[nregrel][8]=float(regrel[nregrel][8])+float(dados_ent[5].replace(',', '.'))
            valx = buscasomaCFOP(regrel[nregrel][1],somaCFOP)
            if valx[1]>-1:
                regrel[nregrel][9]=valx[0]
                del(somaCFOP[valx[1]])
            regrel[nregrel][10]=float(regrel[nregrel][8])+float(regrel[nregrel][9])
            regrel[nregrel][11]=float(regrel[nregrel][11])+float(dados_ent[6].replace(',', '.'))
            regrel[nregrel][12]=float(regrel[nregrel][12])+float(dados_ent[7].replace(',', '.'))
        linhalida = ent.readline()
    ent.close()
    return(regrel)


def processa_SPED_Protocolado(arquivo):
    """
        Funcao processa dados SPED Protocolado
    """
    regrel=[]
    nregrel = 0   
    ent = open(arquivo,mode="r",encoding=comum.encodingDoArquivo(arquivo))
    linhalida = ent.readline()
    while linhalida:
        dados_ent = linhalida.split("|")
        tp_reg = ""
        if (len(dados_ent) >=2):
            tp_reg = dados_ent[1]
        if tp_reg == "D695":
            if nregrel == 0:
                regrel=[[],[]]
                regrel[0]=["Modelo",
                "Série",
                "NF_Ini",
                "NF_Fim",
                "Dt_Ini",
                "Dt_Fim",
                "Nome_Arq",
                "Hashcode",
                # "Contábil_SPED",
                "Liquido_SPED",
                "Base",
                "ICMS"]            
            nregrel = nregrel + 1
            regrel.append([])
            regrel[nregrel].append(dados_ent[2])
            regrel[nregrel].append(dados_ent[3])
            regrel[nregrel].append(dados_ent[4])
            regrel[nregrel].append(dados_ent[5])
            regrel[nregrel].append(dados_ent[6])
            regrel[nregrel].append(dados_ent[7])
            regrel[nregrel].append(dados_ent[8])
            regrel[nregrel].append(dados_ent[9])
            regrel[nregrel].append(0.00)
            regrel[nregrel].append(0.00)
            regrel[nregrel].append(0.00)
        elif tp_reg == "D696":
            regrel[nregrel][8]=float(regrel[nregrel][8])+float(dados_ent[5].replace(',', '.'))
            regrel[nregrel][9]=float(regrel[nregrel][9])+float(dados_ent[6].replace(',', '.'))
            regrel[nregrel][10]=float(regrel[nregrel][10])+float(dados_ent[7].replace(',', '.'))
        linhalida = ent.readline()
    ent.close()
    return(regrel)

def processa_controle_convenio(ie,dtini):
    """
        Funcao processa dados controle convenio
    """
    rel=[[]]
    rel[0]=["EMPS_COD",
         "FILI_COD",
         "FILI_COD_IE",
         "CTR_APUR_DTINI",
         "CTR_APUR_DTFIN",
         "CTR_SERIE",
         "CTR_MODELO",
         "CTR_VOLUME",
         "CTR_IND_RETIF",
         "CTR_TIP_MIDIA",
         "CTR_QTD_MESTRE",
         "CTR_QTD_NFCANC",
         "CTR_DTA_NFINI",
         "CTR_DTA_NFFIN",
         "CTR_NUM_NFINI",
         "CTR_NUM_NFFIN",
         "CTR_NF_VLRTOTAL",
         "CTR_NF_VLRBASE",
         "CTR_NF_VLRICMS",
         "CTR_NF_VLRISEN",
         "CTR_NF_VLROUTRAS",
         "CTR_NF_NOMARQ",
         "CTR_CODH_ARQNF",
         "CTR_QTD_ITEM",
         "CTR_QTD_ITEMCANC",
         "CTR_DTA_ITEMINI",
         "CTR_DTA_ITEMFIN",
         "CTR_NUM_ITEMINI",
         "CTR_NUM_ITEMFIN",
         "CTR_ITEM_VLRTOTAL",
         "CTR_ITEM_VLRDESC",
         "CTR_ITEM_VLRDESP",
         "CTR_ITEM_VLRBASE",
         "CTR_ITEM_VLRICMS",
         "CTR_ITEM_VLRISEN",
         "CTR_ITEM_VLROUTR",
         "CTR_ITEM_NOMARQ",
         "CTR_CODH_ARQITEM",	
         "CTR_QTD_CLI",
         "CTR_CLI_NOMARQ",
         "CTR_CODH_ARQCLI",
         "CTR_CODH_REG",
         "CTR_SER_ORI",
         "CTR_VAL_RED",
         "CTR_DT_GER",
         "CTR_USUA_GER"]
   
    
    query = """SELECT 
         c.EMPS_COD, --ALT001
         c.FILI_COD, --ALT001
         FILI_COD_IE,
         CTR_APUR_DTINI,
         CTR_APUR_DTFIN,
         CTR_SERIE,
         CTR_MODELO,
         TO_NUMBER(CTR_VOLUME),
         CTR_IND_RETIF,
         CTR_TIP_MIDIA,
         CTR_QTD_MESTRE,
         CTR_QTD_NFCANC,
         CTR_DTA_NFINI,
         CTR_DTA_NFFIN,
         CTR_NUM_NFINI,
         CTR_NUM_NFFIN,
         CTR_NF_VLRTOTAL,
         CTR_NF_VLRBASE,
         CTR_NF_VLRICMS,
         CTR_NF_VLRISEN,
         CTR_NF_VLROUTRAS,
         CTR_NF_NOMARQ,
         CTR_CODH_ARQNF,
         CTR_QTD_ITEM,
         CTR_QTD_ITEMCANC,
         CTR_DTA_ITEMINI,
         CTR_DTA_ITEMFIN,
         CTR_NUM_ITEMINI,
         CTR_NUM_ITEMFIN,
         CTR_ITEM_VLRTOTAL,
         CTR_ITEM_VLRDESC,
         CTR_ITEM_VLRDESP,
         CTR_ITEM_VLRBASE,
         CTR_ITEM_VLRICMS,
         CTR_ITEM_VLRISEN,
         CTR_ITEM_VLROUTR,
         CTR_ITEM_NOMARQ,
         CTR_CODH_ARQITEM,	
         CTR_QTD_CLI,
         CTR_CLI_NOMARQ,
         CTR_CODH_ARQCLI,
         CTR_CODH_REG,
         CTR_SER_ORI,
         CTR_VAL_RED,
         CTR_DT_GER,
         CTR_USUA_GER
        ,f.UNFE_SIG UF_FILIAL--ALT001
      FROM  
          --ALT001 OPENRISOW.CTR_IDENT_CNV115
           OPENRISOW.CTR_IDENT_CNV115 c --ALT001
      join openrisow.FILIAL f           --ALT001
        on c.emps_cod = f.emps_cod      --ALT001
       and c.fili_cod = f.fili_cod      --ALT001
      WHERE   
          1=1
          AND c.EMPS_COD  = 'TBRA' --ALT001
          AND FILI_COD_IE = '%s'
          AND CTR_APUR_DTINI = to_date('%s','ddmmyyyy')
      ORDER BY
         c.EMPS_COD,             -- 0  --ALT001
         CTR_APUR_DTINI,         -- 3  --ALT001
         CTR_SERIE,              -- 5  --ALT001
         CTR_MODELO,             -- 7  --ALT001
         UF_FILIAL,              -- 46 --ALT001
         TO_NUMBER(CTR_VOLUME),  -- 8  --ALT001
         CTR_IND_RETIF                 
      --ALT001    CTR_SERIE,
      --ALT001    TO_NUMBER(CTR_VOLUME),
      --ALT001    FILI_COD_IE,
      --ALT001    EMPS_COD,
      --ALT001    CTR_APUR_DTINI,
      --ALT001    CTR_IND_RETIF,
      --ALT001    CTR_DT_GER
    """%(ie,dtini)

    ob_global.gv_ob_conexao.executa(query)
    result = ob_global.gv_ob_conexao.fetchone()
    
    if result == None or len(str(result)) == 0:
        log("#### ATENÇÃO: Nenhum Resultado para Controle Convênio")
        log("####     Query = ")
        log("####")
        log(query)
        log("####")
        rel=[]
        return(rel)
    else:
        a5=result[5]
        a7=result[7]
        #ALT001 a1=result[1]
        a46=result[46] #ALT001 - Alterando Fili_cod para Uf_filial na chave de agrupamento
        a0=result[0]
        a3=result[3]
        a8=result[8]
        regt=result
        while result:
            if  ( 
                (result[0] != a0) or
                #ALT001 (result[1] != a1) or
                (result[46] != a46) or #ALT001 - Alterando Fili_cod para Uf_filial na chave de agrupamento
                (result[3] != a3) or
                (result[5] != a5) or
                (result[7] != a7) or
                (result[8] != a8)
                ):       
            
                if  ( 
                    (result[0] != a0) or
                    #ALT001 (result[1] != a1) or
                    (result[46] != a46) or #ALT001 - Alterando Fili_cod para Uf_filial na chave de agrupamento
                    (result[3] != a3) or
                    (result[5] != a5) or
                    (result[7] != a7)
                    ):       
                    rel.append([])
                    for campo in regt:
                        rel[len(rel)-1].append(campo)
                        a0=result[0]
                        #ALT001 (a1=result[1]
                        a46=result[46] #ALT001 - Alterando Fili_cod para Uf_filial na chave de agrupamento
                        a3=result[3]
                        a5=result[5]
                        a7=result[7]
                        a8=result[8]
                else:
                    a8=result[8]
            regt=result
            result = ob_global.gv_ob_conexao.fetchone()  
        rel.append([])
        for campo in regt:
            rel[len(rel)-1].append(campo)    
            
    return(rel)


def processa_CFOP_0000(ie,uf,dtini,dtfim):
    """
        Funcao processa dados CFOP_0000
    """
    regrel=[[],[]]
    regrel[0]=["EMPRESA",
    	"FILIAL",
        "SERIE",
        # "VLR_CONTABIL",
        "VLR_LIQUIDO",
        "VLR_BASE_ICMS",
        "VLR_ICMS",
        "VLR_ISENTAS",
        "VLR_OUTRAS",
        "QDADE_REGISTROS"]

    # query = """
    # select  /*+ parallel(8) */ 
    #     EMPS_COD                as EMPRESA, 
    #     FILI_COD                as FILIAL, 
    #     INFST_serie             as SERIE,
    #     SUM(infst_val_cont)     as VALOR_CONTABIL,
    #     SUM(infst_base_icms)    as BASE_ICMS, 
    #     SUM(infst_val_icms)     as VALOR_ICMS, 
    #     SUM(infst_isenta_icms)  as ISENTAS,
    #     SUM(infst_outras_icms)  as OUTRAS, 
    #     COUNT(1)
    # FROM  openrisow.ITEM_NFTL_SERV
    # WHERE 1 = 1
    #     AND INFST_dtemiss >= to_date('%s','ddmmyyyy')
    #     AND INFST_dtemiss <= to_date('%s','ddmmyyyy')
    #     AND EMPS_COD = 'TBRA'
    #     AND FILI_COD IN (SELECT fili_cod 
    #                         FROM openrisow.filial
    #                         WHERE 1=1
    #                             AND EMPS_COD = 'TBRA'
    #                             AND fili_cod_insest = '%s'
    #                             AND UNFE_SIG = '%s')
    #     AND infst_ind_canc = 'N'
    #     AND cfop = '0000'
    # GROUP BY EMPS_COD, FILI_COD, INFST_serie
    # """%(dtini,dtfim,ie,uf)

    query = """
    SELECT  /*+ parallel(8) */
        emps_cod                   AS empresa,
        fili_cod                   AS filial,
        infst_serie                AS serie,
        SUM(
            CASE
                WHEN infst_dtemiss < TO_DATE('01/01/2017', 'dd/mm/yyyy') THEN
                    nvl(infst_val_serv, 0) - nvl(infst_val_desc, 0)
                ELSE
                    nvl(infst_val_cont, 0)
            END
        )                          AS valor_contabil,
        SUM(infst_base_icms)       AS base_icms,
        SUM(infst_val_icms)        AS valor_icms,
        SUM(infst_isenta_icms)     AS isentas,
        SUM(infst_outras_icms)     AS outras,
        COUNT(1)
    FROM
        openrisow.item_nftl_serv
    WHERE 1 = 1
    AND infst_dtemiss >= to_date('%s','ddmmyyyy') --MANTER PARAMETRO ANTERIOR
    AND infst_dtemiss <= to_date('%s','ddmmyyyy') --MANTER PARAMETRO ANTERIOR
    AND emps_cod = 'TBRA'
    AND fili_cod IN ( SELECT fili_cod
                        FROM
                            openrisow.filial
                        WHERE 1 = 1
                        AND emps_cod        = 'TBRA'
                        AND fili_cod_insest = '%s'  --MANTER PARAMETRO ANTERIOR
                        AND unfe_sig        = '%s') --MANTER PARAMETRO ANTERIOR

    AND infst_ind_canc = 'N'
    AND cfop = '0000'
    GROUP BY
        emps_cod,fili_cod,infst_serie      
    """%(dtini,dtfim,ie,uf)

    ob_global.gv_ob_conexao.executa(query)
    result = ob_global.gv_ob_conexao.fetchone()
    nregrel = 0
    
    if result == None or len(str(result)) == 0:
        log("#### ATENÇÃO: Nenhum Resultado para a aba CFOP_0000 do relatório")
        log("####     Query = ")
        log("####")
        log(query)
        log("####")
        rel=[]
        return(rel)
    else:
        while result:
            nregrel = nregrel + 1
            regrel.append([])
            for campo in result:
                regrel[nregrel].append(campo)
            result = ob_global.gv_ob_conexao.fetchone()

    return(regrel)


def obtemseries(preg,pprot,pconv):
    """
        Funcao processa OBTENCAO DE SERIES
    """
    lseries = []
    maxpreg=preg.max_row
    maxpprot=pprot.max_row
    maxpconv=pconv.max_row

    for i in range(2,maxpreg + 1):
        serie = preg.cell(row=i,column=2).value
        existe = False
        for eserie in lseries:
            if (eserie == serie):
                existe = True
        if (existe == False and serie):
            lseries.append(serie)

    for i in range(2,maxpprot + 1):
        serie = pprot.cell(row=i,column=2).value
        existe = False
        for eserie in lseries:
            if (eserie == serie):
                existe = True
        if (existe == False and serie):
            lseries.append(serie)

    for i in range(2,maxpconv + 1):
        serie = pconv.cell(row=i,column=6).value
        existe = False
        for eserie in lseries:
            if (eserie == serie):
                existe = True
        if (existe == False and serie):
            lseries.append(serie)
    return(lseries)


def processa_Resumo(pr,pp,pc):
    """
        Funcao processa dados RESUMO
    """

    regrel=[]
        
    lseries=obtemseries(pr,pp,pc)
    nregrel = 1
    
    maxpr=pr.max_row
    maxpp=pp.max_row
    maxpc=pc.max_row

    for campo in lseries:
        if nregrel == 1:
            regrel=[[],[]]
            regrel[0]=["",
                "",
                "Sped Protocolado",
                "",
                "",
                "Sped Regerado (Telecom + CFOP 0000)",
                "",
                "",
                "Conv115 Regerado",
                "",
                "",
                "Sped Protocolado X Conv115",
                "",
                "",
                "Sped Regerado X Conv115",
                ""]
            regrel[1]=["",
                "Série",
                # "Contábil",
                "Liquido",
                "Base",
                "ICMS",
                # "Contábil",
                "Liquido",
                "Base",
                "ICMS",
                # "Contábil",
                "Liquido",
                "Base",
                "ICMS",
                "Dif_Cont",
                "Dif_Base",
                "Dif_ICMS",
                "Dif_Cont",
                "Dif_Base",
                "Dif_ICMS"]       
        
        regrel.append([])
        nregrel = nregrel + 1
        regrel[nregrel].append("")
        regrel[nregrel].append(campo)

        valorcpr = 0.00
        valorbpr = 0.00
        valoripr = 0.00
    
        valorcpp = 0.00
        valorbpp = 0.00
        valoripp = 0.00
    
        valorcpc = 0.00
        valorbpc = 0.00
        valoripc = 0.00

        for i in range(2,maxpp + 1):
            if (campo == (pp.cell(row=i,column=2).value)):
                valorcpp = valorcpp + ((pp.cell(row=i,column=9).value) or 0.00)
                valorbpp = valorbpp + ((pp.cell(row=i,column=10).value) or 0.00)
                valoripp = valoripp + ((pp.cell(row=i,column=11).value) or 0.00)

        for i in range(2,maxpr + 1):
            if (campo == (pr.cell(row=i,column=2).value)):
                valorcpr = valorcpr + ((pr.cell(row=i,column=11).value) or 0.00)
                valorbpr = valorbpr + ((pr.cell(row=i,column=12).value) or 0.00)
                valoripr = valoripr + ((pr.cell(row=i,column=13).value) or 0.00)
      
        for i in range(2,maxpc + 1):
            if (campo == (pc.cell(row=i,column=6).value)):
                valorcpc = valorcpc + ((pc.cell(row=i,column=17).value) or 0.00)
                valorbpc = valorbpc + ((pc.cell(row=i,column=18).value) or 0.00)
                valoripc = valoripc + ((pc.cell(row=i,column=19).value) or 0.00)
      
        regrel[nregrel].append(valorcpp)
        regrel[nregrel].append(valorbpp)
        regrel[nregrel].append(valoripp)

        regrel[nregrel].append(valorcpr)
        regrel[nregrel].append(valorbpr)
        regrel[nregrel].append(valoripr)
  
        regrel[nregrel].append(valorcpc)
        regrel[nregrel].append(valorbpc)
        regrel[nregrel].append(valoripc)

        lin = str(nregrel+1)
        difl="=ROUND(SUM(I" + lin + "-C" + lin + "),2)"
        difm="=ROUND(SUM(J" + lin + "-D" + lin + "),2)"
        difn="=ROUND(SUM(K" + lin + "-E" + lin + "),2)"
        difo="=ROUND(SUM(I" + lin + "-F" + lin + "),2)"
        difp="=ROUND(SUM(J" + lin + "-G" + lin + "),2)"
        difq="=ROUND(SUM(K" + lin + "-H" + lin + "),2)"

        regrel[nregrel].append(difl)
        regrel[nregrel].append(difm)
        regrel[nregrel].append(difn)
        regrel[nregrel].append(difo)
        regrel[nregrel].append(difp)
        regrel[nregrel].append(difq)

    if nregrel > 1:
        regrel.append([])
        nregrel = nregrel + 1
        lin = str(nregrel)
        somac="=ROUND(SUM(C3:C" + lin + "),2)"
        somad="=ROUND(SUM(D3:D" + lin + "),2)"
        somae="=ROUND(SUM(E3:E" + lin + "),2)"
        somaf="=ROUND(SUM(F3:F" + lin + "),2)"
        somag="=ROUND(SUM(G3:G" + lin + "),2)"
        somah="=ROUND(SUM(H3:H" + lin + "),2)"
        somai="=ROUND(SUM(I3:I" + lin + "),2)"
        somaj="=ROUND(SUM(J3:J" + lin + "),2)"
        somak="=ROUND(SUM(K3:K" + lin + "),2)"
        somal="=ROUND(SUM(L3:L" + lin + "),2)"
        somam="=ROUND(SUM(M3:M" + lin + "),2)"
        soman="=ROUND(SUM(N3:N" + lin + "),2)"
        somao="=ROUND(SUM(O3:O" + lin + "),2)"
        somap="=ROUND(SUM(P3:P" + lin + "),2)"
        somaq="=ROUND(SUM(Q3:Q" + lin + "),2)"
    
        regrel[nregrel].append("")
        regrel[nregrel].append("Total Geral")
        regrel[nregrel].append(somac)
        regrel[nregrel].append(somad)
        regrel[nregrel].append(somae)
        regrel[nregrel].append(somaf)
        regrel[nregrel].append(somag)
        regrel[nregrel].append(somah)
        regrel[nregrel].append(somai)
        regrel[nregrel].append(somaj)
        regrel[nregrel].append(somak)
        regrel[nregrel].append(somal)
        regrel[nregrel].append(somam)
        regrel[nregrel].append(soman)
        regrel[nregrel].append(somao)
        regrel[nregrel].append(somap)
        regrel[nregrel].append(somaq)
    
    return(regrel)
    
def fnc_retorna_id_arquivo(p_cc_path, p_cc_uf, p_cc_ie, p_cc_mes, p_cc_ano) :
    """
        Funcao para retornar identificacao id arquivo 
    """
    # try:
    if p_cc_path:
        try :
            fd = open(p_cc_path,'r') #, encoding=comum.encodingDoArquivo(p_cc_path))
            lin = fd.readline()
        except :
            fd = open(p_cc_path,'r', encoding=comum.encodingDoArquivo(p_cc_path))
            lin = fd.readline()
        
        fd.close()
        
        if lin and lin.startswith('|0000|') :
            ano = lin.split('|')[4][4:] 
            mes = lin.split('|')[4][2:4]
            uf = lin.split('|')[9]
            insc = lin.split('|')[10]
            compet_i = lin.split('|')[4]
            compet_f = lin.split('|')[5]
            
            if uf and insc and compet_i and compet_f and mes and ano:
                if (str(p_cc_uf).strip() == str(uf).strip() \
                and str(p_cc_ie).strip() == str(insc).strip() \
                and str(p_cc_mes).strip() == str(mes).strip() \
                and str(p_cc_ano).strip() == str(ano).strip()):
                    return [True, uf, insc, compet_i, compet_f, mes, ano]
                else:
                    return [False, uf, insc, compet_i, compet_f, mes, ano]
            else:
                return False, uf, insc, compet_i, compet_f, mes, ano
                
        return False, "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"
        
    else: #except:
        return False, "FALHA", "FALHA", "FALHA", "FALHA", "FALHA", "FALHA"

def fnc_conectar_banco_dados():
    """
        Funcao para conectar na base de dados 
    """
    try:
        
        ob_global.gv_ob_conexao = sql.geraCnxBD(configuracoes)
        v_ds_sql="""
        SELECT 'PAINEL_' || TO_CHAR(SYSDATE,'YYYYMMDD_HH24MISS') NM_JOB FROM DUAL
        """
        ob_global.gv_ob_conexao.executa(v_ds_sql)
        
        v_ob_cursor = ob_global.gv_ob_conexao.fetchone()
        if (v_ob_cursor):    
            for campo in v_ob_cursor:
                log(str(campo) + " >> SUCESSO CONEXAO BANCO DE DADOS") 
                ob_global.gv_nm_job = str(campo)
                return 0
                break
        
        else:
            log("ERRO CONEXAO BANCO DE DADOS ") 
            return 91
        
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO CONEXAO BANCO DE DADOS .: " + str(e) + " - TRACE - " + v_ds_trace)
        return 91


def fnc_validar_entrada():
    """
        Retorna a validação de entrada e dos arquivos de configuração
    """
    try:
        v_nr_retorno = 0        
         
        log("-"*150)

        comum.addParametro('P_CC_UF', None, 'Unidade Federativa do Brasil.', True,'SP')
        comum.addParametro('P_CC_MESANO', None, 'Mes e Ano de processamento.', True,'012015')
        comum.addParametro('P_CC_IE', None, 'Inscricao Estadual.', False,'108383949112')
        
        # Validacao dos parametros de entrada
        if not comum.validarParametros() :
            v_nr_retorno = 91
        
        else:
        
            # INICIO ELSE
            ob_global.gv_cc_uf                           = comum.getParametro('P_CC_UF').upper().strip()   
            ob_global.gv_cc_mesano                       = comum.getParametro('P_CC_MESANO').upper().strip()
            
            try:
                ob_global.gv_cc_ie                       = comum.getParametro('P_CC_IE').strip()
            except:                                      
                ob_global.gv_cc_ie                       = ""
            
            if not v_nr_retorno:
                try:
                    if len(ob_global.gv_cc_uf) != 2:
                        log("PARAMETRO UF: Invalido! " + ob_global.gv_cc_uf) 
                        v_nr_retorno = 91 
                    else:
                        if not util.validauf(ob_global.gv_cc_uf):
                            log("PARAMETRO UF: Invalido! " + ob_global.gv_cc_uf) 
                            v_nr_retorno = 91 
                except:
                    log("PARAMETRO UF: Invalido! " + ob_global.gv_cc_uf) 
                    v_nr_retorno = 91            
                    
            if not v_nr_retorno:
                try:
                    if (len(ob_global.gv_cc_mesano) != 6):
                        log("PARAMETRO MESANO: Invalido! " + ob_global.gv_cc_mesano) 
                        v_nr_retorno = 91           
                    else:
                        if (
                           int(ob_global.gv_cc_mesano[0:2]) > 12
                        or int(ob_global.gv_cc_mesano[0:2]) < 1
                        ):
                            log("PARAMETRO MES [MESANO] : Invalido! " + ob_global.gv_cc_mesano) 
                            v_nr_retorno = 91                         
                except:
                    log("PARAMETRO MESANO : Invalido! " + ob_global.gv_cc_mesano) 
                    v_nr_retorno = 91            

            if not v_nr_retorno:
                try:
                    if len(ob_global.gv_cc_ie) == 0:
                        ob_global.gv_cc_ie = "*"
                    else:
                        if util.valida_ie(ob_global.gv_cc_ie).strip() == "#":
                            log("PARAMETRO I.E: Invalido! " + ob_global.gv_cc_ie \
                            + ", \n porem sera atribuido * e continuar o processamento .... ")
                            ob_global.gv_cc_ie = "*" 
                except:
                    log("PARAMETRO I.E: Invalido! " + ob_global.gv_cc_ie) 
                    v_nr_retorno = 91            
            
            if not v_nr_retorno:
                try:
                    ob_global.dir_protocolados = configuracoes.dir_protocolados.strip()
                    ob_global.dir_regerados = configuracoes.dir_regerados.strip()
                    ob_global.dir_relatorios = configuracoes.dir_relatorios.strip()
                    ob_global.mascara_regerado = configuracoes.mascara_regerado.strip()
                    ob_global.mascara_protocolado = configuracoes.mascara_protocolado.strip()
                    ob_global.nome_relatorio = configuracoes.nome_relatorio.strip()
                    
                    if len(ob_global.dir_protocolados) == 0:
                        log("PARAMETRO CONFIGURACOES [DIR_PROTOCOLADOS]: Invalido! " + ob_global.dir_protocolados) 
                        v_nr_retorno = 91                    
                    elif len(ob_global.dir_regerados) == 0:
                        log("PARAMETRO CONFIGURACOES [DIR_REGERADOS]: Invalido! " + ob_global.dir_regerados) 
                        v_nr_retorno = 91                    
                    elif len(ob_global.dir_relatorios) == 0:
                        log("PARAMETRO CONFIGURACOES [DIR_RELATORIOS]: Invalido! " + ob_global.dir_relatorios) 
                        v_nr_retorno = 91                    
                    elif len(ob_global.mascara_regerado) == 0:
                        log("PARAMETRO CONFIGURACOES [MASCARA_REGERADO]: Invalido! " + ob_global.mascara_regerado) 
                        v_nr_retorno = 91                    
                    elif len(ob_global.mascara_protocolado) == 0:
                        log("PARAMETRO CONFIGURACOES [MASCARA_PROTOCOLADO]: Invalido! " + ob_global.mascara_protocolado) 
                        v_nr_retorno = 91    
                    elif len(ob_global.nome_relatorio) == 0:
                        log("PARAMETRO CONFIGURACOES [NOME_RELATORIO]: Invalido! " + ob_global.nome_relatorio) 
                        v_nr_retorno = 91    
                        
                except:
                    log("PARAMETRO de CONFIGURACOES : Invalido! Favor validar ... ") 
                    v_nr_retorno = 91            

            if not v_nr_retorno:
                try:
                    ob_global.dir_protocolados = os.path.join(ob_global.dir_protocolados \
                                                            , ob_global.gv_cc_uf \
                                                            , ob_global.gv_cc_mesano[2:] \
                                                            , ob_global.gv_cc_mesano[:2])
                                                            
                    ob_global.dir_regerados = os.path.join(ob_global.dir_regerados \
                                                            , ob_global.gv_cc_uf \
                                                            , ob_global.gv_cc_mesano[2:] \
                                                            , ob_global.gv_cc_mesano[:2])
                                                            
                    ob_global.dir_relatorios = os.path.join(ob_global.dir_relatorios \
                                                            , ob_global.gv_cc_uf \
                                                            , ob_global.gv_cc_mesano[2:] \
                                                            , ob_global.gv_cc_mesano[:2])
                    
                    ob_global.dir_Apoio = os.path.join(ob_global.dir_relatorios, 'Apoio')
                    
                    ob_global.mascara_regerado = ob_global.mascara_regerado.replace('<<UF>>',ob_global.gv_cc_uf).replace('<<MESANO>>',ob_global.gv_cc_mesano).strip()
                    ob_global.mascara_protocolado = ob_global.mascara_protocolado.replace('<<UF>>',ob_global.gv_cc_uf).replace('<<MESANO>>',ob_global.gv_cc_mesano).strip()
                    ob_global.nome_relatorio = ob_global.nome_relatorio.replace('<<UF>>',ob_global.gv_cc_uf).replace('<<MESANO>>',ob_global.gv_cc_mesano).strip()
                    if ob_global.gv_cc_ie != "*":
                        ob_global.mascara_regerado = ob_global.mascara_regerado.replace('<<IE>>',ob_global.gv_cc_ie).strip()
                        ob_global.mascara_protocolado = ob_global.mascara_protocolado.replace('<<IE>>',ob_global.gv_cc_ie).strip()
                        ob_global.nome_relatorio = ob_global.nome_relatorio.replace('<<IE>>',ob_global.gv_cc_ie).strip()
                        
                except Exception as e:
                    v_ds_trace = traceback.format_exc()
                    log("ERRO JOIN DOS DIRETORIOS DE CONFIGURACOES: " + str(e)+ " >> " + v_ds_trace)         
                    v_nr_retorno = 91
                    
            if not v_nr_retorno:
                try:
                    if dir_base.strip().find('DEV') >= 0:
                        if not os.path.isdir(ob_global.dir_protocolados) :
                            os.makedirs(ob_global.dir_protocolados)
                        if not os.path.isdir(ob_global.dir_regerados) :
                            os.makedirs(ob_global.dir_regerados)
                        if not os.path.isdir(ob_global.dir_relatorios) :
                            os.makedirs(ob_global.dir_relatorios)
                        if not os.path.isdir(ob_global.dir_Apoio) :
                            os.makedirs(ob_global.dir_Apoio)
                    elif not os.path.isdir(ob_global.dir_protocolados) :
                        log("NAO EXISTE DIRETORIO [PROTOCOLADOS]: Invalido! " + ob_global.dir_protocolados) 
                        v_nr_retorno = 91           
                    elif not os.path.isdir(ob_global.dir_regerados) :
                        log("NAO EXISTE DIRETORIO [REGERADOS]: Invalido! " + ob_global.dir_regerados) 
                        v_nr_retorno = 91                
                    else:
                        if not os.path.isdir(ob_global.dir_relatorios) :
                            os.makedirs(ob_global.dir_relatorios)
                        if not os.path.isdir(ob_global.dir_Apoio) :
                            os.makedirs(ob_global.dir_Apoio)
                        
                except Exception as e:
                    v_ds_trace = traceback.format_exc()
                    log("ERRO CRIACAO DOS DIRETORIOS DE CONFIGURACOES: " + str(e)+ " >> " + v_ds_trace)         
                    v_nr_retorno = 91
                    
            if not v_nr_retorno:
                try:                    
                    v_mascara_regerado = ob_global.mascara_regerado
                    if ob_global.gv_cc_ie == "*":
                        v_mascara_regerado = v_mascara_regerado.replace('<<IE>>',ob_global.gv_cc_ie).strip()
                    
                    v_lista_ie_regerado = util.lista_ies_existentes(v_mascara_regerado,ob_global.dir_regerados)
                    if len(v_lista_ie_regerado) == 0:
                        log("NAO EXISTE ARQUIVO [REGERADOS]: Invalido! " + v_mascara_regerado + " >> DIR: " + ob_global.dir_regerados) 
                        v_nr_retorno = 91           
                    elif not isinstance(v_lista_ie_regerado,list):
                        log("ARQUIVOs [REGERADOS]: Invalido! " + str(v_lista_ie_regerado)) 
                        v_nr_retorno = 91           
                    else:
                        ob_global.gv_ob_lista_ie = list(set(v_lista_ie_regerado))
                        if len(ob_global.gv_ob_lista_ie) == 0:
                            log("NAO EXISTE I.E [REGERADOS]: Invalido! >> DIR: " + ob_global.dir_regerados) 
                            v_nr_retorno = 91 
                            
                except Exception as e:
                    v_ds_trace = traceback.format_exc()
                    log("ERRO BUSCA AS I.ES DOS DIRETORIOS REGERADOS: " + str(e)+ " >> " + v_ds_trace)         
                    v_nr_retorno = 91
            
            if not v_nr_retorno:
                try:                    
           
                    ob_global.gv_ob_dic_arquivo = dict()
                    
                    for campo_ie in ob_global.gv_ob_lista_ie:
                    
                        if len(campo_ie.strip()) == 0:
                            continue
                            
                        v_arquivo_protocolado = ob_global.mascara_protocolado.replace('<<IE>>',campo_ie.strip()).strip()
                        v_arquivo_regerado = ob_global.mascara_regerado.replace('<<IE>>',campo_ie.strip()).strip()
                       
                        v_arquivo_regerado = str(util.nome_arquivo(v_arquivo_regerado,ob_global.dir_regerados)).strip()
                        if len(v_arquivo_regerado) == 0:
                            log("NAO EXISTE ARQUIVO [REGERADOS]: Invalido! " + ob_global.mascara_regerado.replace('<<IE>>',campo_ie).strip() + " >> DIR: " + ob_global.dir_regerados) 
                            v_nr_retorno = 91    
                            break    
                        
                        v_arquivo_protocolado = str(util.nome_arquivo(v_arquivo_protocolado,ob_global.dir_protocolados)).strip()
                        if len(v_arquivo_protocolado) == 0:
                            log("NAO EXISTE ARQUIVO [PROTOCOLADOS]: Invalido! " + ob_global.mascara_protocolado.replace('<<IE>>',campo_ie).strip() + " >> DIR: " + ob_global.dir_protocolados) 
                            v_nr_retorno = 91    
                            break    
                        
                        val_r, uf_r,insc_r,compet_i_r,compet_f_r,mes_r, ano_r = fnc_retorna_id_arquivo(p_cc_path=v_arquivo_regerado, p_cc_uf=ob_global.gv_cc_uf, p_cc_ie=campo_ie.strip(), p_cc_mes=ob_global.gv_cc_mesano[:2], p_cc_ano=ob_global.gv_cc_mesano[2:])
                                                
                        if (val_r and uf_r and insc_r and compet_i_r and compet_f_r and mes_r and ano_r):
                            pass
                        else:
                            log('-'*100)
                            log('#### Erro - Arquivo na pasta de entrada REGERADOS está com dados diferentes dos esperados - REMOVA ou CORRIJA')
                            log('#### Nome do arquivo = ',v_arquivo_regerado)
                            log('####')
                            log('####    UF INFORMADO ......... = ',str(ob_global.gv_cc_uf))
                            log('####    UF NO ARQUIVO ........ = ',str(uf_r))
                            log('####')
                            log('####    INCRIÇÃO DO NOME ..... = ',str(campo_ie).strip())
                            log('####    INSCRIÇÃO NO ARQUIVO . = ',str(insc_r))
                            log('####')
                            log('####    ANO INFORMADO ........ = ',str(ob_global.gv_cc_mesano[2:]))
                            log('####    ANO NO ARQUIVO ....... = ',str(ano_r))
                            log('####')
                            log('####    MÊS INFORMADO ........ = ',str(ob_global.gv_cc_mesano[:2]))
                            log('####    MÊS NO ARQUIVO ....... = ',str(mes_r))
                            log('-'*100)                        
                            v_nr_retorno = 91    
                            break                             
                        
                        
                        val_p, uf_p,insc_p,compet_i_p,compet_f_p,mes_p, ano_p = fnc_retorna_id_arquivo(p_cc_path=v_arquivo_protocolado, p_cc_uf=ob_global.gv_cc_uf, p_cc_ie=campo_ie.strip(), p_cc_mes=ob_global.gv_cc_mesano[:2], p_cc_ano=ob_global.gv_cc_mesano[2:])
                                                
                        if (val_p and uf_p and insc_p and compet_i_p and compet_f_p and mes_p and ano_p):
                            pass
                        else:
                            log('-'*100)
                            log('#### Erro - Arquivo na pasta de entrada PROTOCOLADO está com dados diferentes dos esperados - REMOVA ou CORRIJA')
                            log('#### Nome do arquivo = ',v_arquivo_protocolado)
                            log('####')
                            log('####    UF INFORMADO ......... = ',str(ob_global.gv_cc_uf))
                            log('####    UF NO ARQUIVO ........ = ',str(uf_p))
                            log('####')
                            log('####    INCRIÇÃO DO NOME ..... = ',str(campo_ie).strip())
                            log('####    INSCRIÇÃO NO ARQUIVO . = ',str(insc_p))
                            log('####')
                            log('####    ANO INFORMADO ........ = ',str(ob_global.gv_cc_mesano[2:]))
                            log('####    ANO NO ARQUIVO ....... = ',str(ano_p))
                            log('####')
                            log('####    MÊS INFORMADO ........ = ',str(ob_global.gv_cc_mesano[:2]))
                            log('####    MÊS NO ARQUIVO ....... = ',str(mes_p))
                            log('-'*100)                        
                            v_nr_retorno = 91    
                            break                             
                        
                        ob_global.gv_ob_dic_arquivo[campo_ie.strip()] = (v_arquivo_regerado,v_arquivo_protocolado,val_r, uf_r,insc_r,compet_i_r,compet_f_r,mes_r, ano_r,val_p, uf_p,insc_p,compet_i_p,compet_f_p,mes_p, ano_p)    
                    
                    if len(ob_global.gv_ob_dic_arquivo) == 0:
                        log("FALHA BUSCA DOS ARQUIVOS DOS DIRETORIOS PROTOCOLADOS/REGERADOS: " + str(ob_global.gv_ob_dic_arquivo))         
                        v_nr_retorno = 91
                                    
                except Exception as e:
                    v_ds_trace = traceback.format_exc()
                    log("ERRO BUSCA DOS ARQUIVOS DOS DIRETORIOS: " + str(e)+ " >> " + v_ds_trace)         
                    v_nr_retorno = 91
                
            # FIM ELSE
            
        return v_nr_retorno
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO VALIDAÇÃO DOS PARAMETROS DE ENTRADA: " + str(e)+ " >> " + v_ds_trace)
        v_nr_retorno = 93
        return v_nr_retorno


def fnc_processar():
    """
        Funcao principal para processar as informacoes
    """
    try:
        
        fnc_log()
        
        v_nr_retorno = 0
        v_ds_erro = ""
        v_ob_lista_relatorio = []
        
        for campo_ie,campo_valor in ob_global.gv_ob_dic_arquivo.items():
            # INICIO FOR
            v_arquivo_regerado,v_arquivo_protocolado,val_r, uf_r,insc_r,compet_i_r,compet_f_r,mes_r, ano_r,val_p, uf_p,insc_p,compet_i_p,compet_f_p,mes_p, ano_p = campo_valor
            
            log('-'*100)
            log('# REGERADO...')
            log('# Nome do arquivo = ',v_arquivo_regerado)
            log('#    UF .................. = ',uf_r)
            log('#    ANO ................. = ',ano_r)
            log('#    MÊS ................. = ',mes_r)
            log('#    INSCRIÇÃO ESTADUAL .. = ',insc_r)
            log('#    COMPETÊNCIA INICIAL . = ',compet_i_r)
            log('#    COMPETÊNCIA FINAL ... = ',compet_f_r)
            log('-'*100)

            log('-'*100)
            log('# PROTOCOLADO...')
            log('# Nome do arquivo = ',v_arquivo_protocolado)
            log('#    UF .................. = ',uf_p)
            log('#    ANO ................. = ',ano_p)
            log('#    MÊS ................. = ',mes_p)
            log('#    INSCRIÇÃO ESTADUAL .. = ',insc_p)
            log('#    COMPETÊNCIA INICIAL . = ',compet_i_p)
            log('#    COMPETÊNCIA FINAL ... = ',compet_f_p)
            log('-'*100)            
            
            v_versao_arquivo = str(str(str(v_arquivo_regerado).split("_")[-1]).split(".")[0]).strip().upper()
            if len(v_versao_arquivo) > 0: 
                v_versao_arquivo = "_" + v_versao_arquivo
            v_nome_relatorio = ob_global.nome_relatorio.replace('<<IE>>',str(campo_ie)).strip()
            v_nome_relatorio = v_nome_relatorio.replace('<<VERSAO_RELATORIO>>',str(v_versao_arquivo)).strip()
            v_nome_relatorio = os.path.join(ob_global.dir_relatorios,v_nome_relatorio)
            
            log('- INICIO [processa_SPED_Regerado] ',str(v_arquivo_regerado),str(insc_r), str(uf_r), str(compet_i_r), str(compet_f_r))
            v_ob_dados_processa_sped_regerado = []
            v_ob_dados_processa_sped_regerado = processa_SPED_Regerado(v_arquivo_regerado,insc_r, uf_r, compet_i_r, compet_f_r)
            if len(v_ob_dados_processa_sped_regerado) == 0:
                log("ERRO: NÃO EXISTE DADOS PARA SPED REGERADO!")
                v_nr_retorno = 93
                v_ds_erro = "ERRO: NÃO EXISTE DADOS PARA SPED REGERADO!" + " \n " + v_ds_erro
                try:
                    v_ds_erro =  " [ " + str(str(v_arquivo_regerado).split(SD)[-1]) + " ] " + v_ds_erro
                except:
                    pass
                # return v_nr_retorno                    
                # break
            log('- FIM [processa_SPED_Regerado] ',str(len(v_ob_dados_processa_sped_regerado)))
            
            log('- INICIO [processa_SPED_Protocolado] ',str(v_arquivo_protocolado))
            v_ob_dados_processa_sped_protocolado = []
            v_ob_dados_processa_sped_protocolado = processa_SPED_Protocolado(v_arquivo_protocolado)
            if len(v_ob_dados_processa_sped_protocolado) == 0:
                log("ERRO: NÃO EXISTE DADOS PARA SPED PROTOCOLADO!")
                v_nr_retorno = 93
                v_ds_erro = "ERRO: NÃO EXISTE DADOS PARA SPED PROTOCOLADO!" + " \n " + v_ds_erro
                try:
                    v_ds_erro =  " [ " + str(str(v_arquivo_protocolado).split(SD)[-1]) + " ] " + v_ds_erro
                except:
                    pass
                # return v_nr_retorno                    
                # break
            log('- FIM [processa_SPED_Protocolado] ',str(len(v_ob_dados_processa_sped_protocolado)))
            
            log('- INICIO [processa_controle_convenio] ',str(insc_r),str(compet_i_r))
            v_ob_dados_processa_controle_convenio = []
            v_ob_dados_processa_controle_convenio = processa_controle_convenio(insc_r, compet_i_r)
            if len(v_ob_dados_processa_controle_convenio) == 0:
                log("ERRO: NÃO EXISTE DADOS PARA Controle_Convenio!")
                v_nr_retorno = 93
                v_ds_erro = "ERRO: NÃO EXISTE DADOS PARA Controle_Convenio!" + " \n " + v_ds_erro
                try:
                    v_ds_erro =  " [ " + str(str(v_arquivo_regerado).split(SD)[-1]) + " ] " + " >> I.E : " + str(insc_r)  + " >> COMP.INI : " +  str(compet_i_r) + " " + v_ds_erro
                except:
                    pass
                # return v_nr_retorno                    
                # break
            log('- FIM [processa_controle_convenio] ',str(len(v_ob_dados_processa_controle_convenio)))
            
            log('- INICIO [processa_CFOP_0000] ',str(insc_r), str(uf_r), str(compet_i_r), str(compet_f_r))
            v_ob_dados_processa_cfop_0000 = []
            v_ob_dados_processa_cfop_0000 = processa_CFOP_0000(insc_r, uf_r, compet_i_r, compet_f_r)
            if len(v_ob_dados_processa_cfop_0000) == 0:
                log("ERRO: NÃO EXISTE DADOS PARA CFOP_0000!")
                v_nr_retorno = 93
                v_ds_erro = "ERRO: NÃO EXISTE DADOS PARA CFOP_0000!" + " \n " + v_ds_erro
                try:
                    v_ds_erro =  " [ " + str(str(v_arquivo_regerado).split(SD)[-1]) + " ] " + " >> I.E : " + str(insc_r)  + " >> COMP.INI : " +  str(compet_i_r)  + " >> COMP.FIM : " +  str(compet_f_r) + " " + v_ds_erro
                except:
                    pass                
                # return v_nr_retorno                    
                # break
            log('- FIM [processa_CFOP_0000] ',str(len(v_ob_dados_processa_cfop_0000)))
            
            log("#### Início da criacao da planilha. #### ",str(v_nome_relatorio))
            #### Cria a planilha em memória....
            arquivo_excel = Workbook()
            planilha0 = arquivo_excel.active
            planilha0.title = "SPED_Regerado"
            planilha1 = arquivo_excel.create_sheet("SPED_Protocolado", 1)
            planilha2 = arquivo_excel.create_sheet("Controle_Convenio", 2)
            planilha3 = arquivo_excel.create_sheet("CFOP_0000", 3)
            planilha4 = arquivo_excel.create_sheet("Resumo", 4)   
        
            log('-'*100)
            log("- Início do processamento da aba SPED_Regerado.")
            if (v_ob_dados_processa_sped_regerado):
                for linha in v_ob_dados_processa_sped_regerado:
                    planilha0.append(linha)                
                for c in range(9,14):    
                    formata_valor(planilha0,c)            
                formata_comum(planilha0)
                arquivo_excel.save(v_nome_relatorio)            
            log("- Fim do processamento da aba SPED_Regerado.",str(len(v_ob_dados_processa_sped_regerado)))
            log('-'*100)            
            
            log('-'*100)
            log("- Início do processamento da aba SPED_Protocolado.")
            if (v_ob_dados_processa_sped_protocolado):
                for linha in v_ob_dados_processa_sped_protocolado:
                    planilha1.append(linha)
                for c in range(9,12):    
                    formata_valor(planilha1,c)
                formata_comum(planilha1)
                arquivo_excel.save(v_nome_relatorio)
            log("- Fim do processamento da aba SPED_Protocolado.",str(len(v_ob_dados_processa_sped_protocolado)))
            log('-'*100)            
            
            log('-'*100)
            log("- Início do processamento da aba Controle_Convenio.")
            if (v_ob_dados_processa_controle_convenio):
                for linha in v_ob_dados_processa_controle_convenio:
                    planilha2.append(linha)
                for c in range(17,22):    
                    formata_valor(planilha2,c)
                formata_comum(planilha2)
                arquivo_excel.save(v_nome_relatorio)
            log("- Fim do processamento da aba Controle_Convenio.",str(len(v_ob_dados_processa_controle_convenio)))
            log('-'*100)

            log('-'*100)
            log("- Início do processamento da aba CFOP_0000.")
            if (v_ob_dados_processa_cfop_0000):
                for linha in v_ob_dados_processa_cfop_0000:
                    planilha3.append(linha)
                for c in range(4,9):    
                    formata_valor(planilha3,c)
                formata_comum(planilha3)
                arquivo_excel.save(v_nome_relatorio)
            log("- Fim do processamento da aba CFOP_0000.",str(len(v_ob_dados_processa_cfop_0000)))
            log('-'*100)

            log('- INICIO [processa_Resumo] / aba Resumo. ')
            v_ob_dados_processa_resumo = []
            v_ob_dados_processa_resumo = processa_Resumo(planilha0,planilha1,planilha2)
            if len(v_ob_dados_processa_resumo) == 0:
                log("ERRO: NÃO EXISTE DADOS PARA RESUMO!")
                v_nr_retorno = 93
                v_ds_erro = "ERRO: NÃO EXISTE DADOS PARA RESUMO!" + " \n " + v_ds_erro                
                # return v_nr_retorno                    
                # break
            else:    
                for linha in v_ob_dados_processa_resumo:
                    planilha4.append(linha)
                formata_resumo(planilha4)    
                arquivo_excel.save(v_nome_relatorio)    
            log('- FIM [processa_Resumo] / aba Resumo. ',str(len(v_ob_dados_processa_resumo)))            
            
            arquivo_excel.save(v_nome_relatorio)
            log("#### FIM da criacao da planilha. #### ",str(v_nome_relatorio))
            v_ob_lista_relatorio.append(v_nome_relatorio)
            
            # FIM FOR
        
        try:
            if len(v_ob_lista_relatorio) > 0:
                log('-'*100)
                log('# RELATORIOS CRIADOS...')
                log('# Lista de arquivos = ',str(v_ob_lista_relatorio))
                log('-'*100)     
        except:
            pass                
        
        try:
            if v_nr_retorno > 0 and v_ds_erro:
                log('-'*100)
                log('ERRO >> ', " \n ", v_ds_erro)
                log('-'*100)  
        except:
            pass
        
        return v_nr_retorno
        
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO PROCESSAMENTO: " + str(e)+ " >> " + v_ds_trace)
        v_nr_retorno = 93
        return v_nr_retorno        


if __name__ == "__main__" :
    """
        Ponto de partida
    """    
    # Codigo de Retorno
    log.ret = 0
    v_nr_ret = 0

    # Tratamento de excessao
    v_ds_trace = ''

    # Tratamento de variaveis globais
    ob_global.gv_ob_conexao = None
    # Parametros do arquivo de configuração
    ob_global.gv_nm_job ="" 

    try:

        log("-"*100)
        log(" INICIO DA EXECUÇÃO ".center(120,'#'))
            
        log(" ")
        
        # Validacao dos parametros de entrada
        if not v_nr_ret :
            v_nr_ret = fnc_validar_entrada()
        
        log(" ")
        
        # Verificar conexao com o banco
        if not v_nr_ret :
            v_nr_ret = fnc_conectar_banco_dados()   
        
        log(" ")

        # Processar         
        if not v_nr_ret :
            v_nr_ret = fnc_processar()                    

        #fnc_log()
        
        # Finalizacao
        log(" ")            
        
        if not v_nr_ret :
            log("SUCESSO NA EXECUÇÃO")
        else:
            log("ERRO NA EXECUÇÃO")
                        
        log(" ")
    
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO .: " + str(e) + " >> " + v_ds_trace)
        v_nr_ret = 93
    
    
    try:
        if v_nr_ret is not None:
            v_nr_ret = int(v_nr_ret)
        elif log.ret is not None:
            v_nr_ret = int(log.ret)
        else:
            v_nr_ret = int(0)
    except:    
        v_nr_ret = int(0)

    try:
        if log.ret is not None:
            log.ret = int(log.ret)
        else:
            log.ret = int(0)
    except:    
        log.ret = int(0)
        
    sys.exit(v_nr_ret if v_nr_ret >= log.ret else log.ret )
