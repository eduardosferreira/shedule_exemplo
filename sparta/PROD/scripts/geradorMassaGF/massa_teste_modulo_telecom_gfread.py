#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-

"""
--------------------------------------------------------------------------------------------------------------------------
  SISTEMA ..: 
  MODULO ...: 
  SCRIPT ...: Criação Massa de Teste
  CRIACAO ..: 13/08/2021
  AUTOR ....: Victor Santos - Kyros Consultoria 
  DESCRICAO : Desenvolver um processo python que gere massa de dados para teste.
  ALTERACAO :
--------------------------------------------------------------------------------------------------------------------------
    Exemplo de comando: ./criacao_massa_teste.py <MESANO> <SERIE> <UF> <FILI_COD> <NUM_NOTA_INI> <NUM_NOTA_FIM> <OWNER> 
--------------------------------------------------------------------------------------------------------------------------
"""

import sys
import datetime
import os
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.reader.excel import load_workbook 

dir_base = os.path.join( os.path.realpath('.').split('/PROD/')[0], 'PROD') if os.path.realpath('.').__contains__('/PROD/') else os.path.join( os.path.realpath('.').split('/DEV/')[0], 'DEV')
sys.path.append(dir_base)
import configuracoes
import comum
import sql

log.gerar_log_em_arquivo = True

config = comum.carregaConfiguracoes(configuracoes) 

SD = ('/' if os.name == 'posix' else '\\')

def processar():

    vDataIni    = ""
    vDataFim    = ""
    vSerie      = ""
    vUF         = ""
    vFiliCod    = ""
    vNumNotaIni = ""   
    vNumNotaFim = "" 
    vOwner      = ""  
    ret         = 0
    
    conexao = sql.geraCnxBD(configuracoes)

    if ( len(sys.argv) == 8 ):
        vDataIni    = sys.argv[1]
        vDataFim    = sys.argv[2]
        vSerie      = sys.argv[3] 
        # vUF         = sys.argv[4].upper()
        vFiliCod    = sys.argv[4] 
        vNumNotaIni = sys.argv[5] 
        vNumNotaFim = sys.argv[6] 
        vOwner      = sys.argv[7] 

        if vOwner == '':
            log('#### ERRO - O PARAMETRO OWNER É OBRIGATÓRIO...')
            ret = 99
            return ret

        log("-"* 100)
        log('# - Data Inicial.................................:', vDataIni)
        log('# - Data Fim.....................................:', vDataFim)
        log('# - Série........................................:', vSerie)
        # log('# - UF...........................................:', vUF)
        log('# - Fili_cod.....................................:', vFiliCod)
        log('# - Número de Nota Inicial.......................:', vNumNotaIni)
        log('# - Número de Nota Final.........................:', vNumNotaFim)
        log('# - Owner........................................:', vOwner)
        log("-"* 100)
        
    else:
        log("-" * 100)
        log("#### ")
        log('#### ERRO - Erro nos parametros do script.')
        log("Retorno = 99") 
        ret = 99
        return ret  

#### DADOS
#### DADOS
#### DADOS

    vWhere = '1 = 1'
    
    if vSerie != '':
        result = vSerie.split(',')
        serie = "('"
        for row in result:
            serie = serie + row.replace(' ', '') + "','"
        serie = serie[:len(serie)-2]
        serie = serie + ")"

        vWhere = vWhere + " AND replace(serie,' ','') in " + serie

    # if vUF != '':
    #     valida = validauf(vUF)
        
    #     if valida == False:
    #         log('#### ERRO, UF INVÁLIDA')
    #         ret = 99
    #         return ret
        
    #     result = vUF.split(',')
    #     uf = "('"
    #     for row in result:
    #         uf = uf + row.replace(' ', '') + "','"
    #     uf = uf[:len(uf)-2]
    #     uf = uf + ")"
    #     vWhere = vWhere + " AND l.uf_filial in " + uf

    if vDataIni != '':
        vWhere = vWhere + " AND data_ini >= TO_DATE ('" + vDataIni + "', 'dd/mm/yyyy') "
        vWhere = vWhere + " AND data_fim <= TO_DATE ('" + vDataFim + "', 'dd/mm/yyyy') "

    if vFiliCod != '':
        result = vFiliCod.split(',')
        fili = "('"
        for row in result:
            fili = fili + row.replace(' ', '') + "','"
        fili = fili[:len(fili)-2]
        fili = fili + ")"
        vWhere = vWhere + " AND fili_cod  in " + fili

    vQueryCli = vWhere
    if vNumNotaIni != '':
        vWhere = vWhere + " AND num_nota >= '" + vNumNotaIni.zfill(9) + "'"

    if vNumNotaFim != '':
        vWhere = vWhere + " AND num_nota <=  '" + vNumNotaFim.zfill(9) + "'"
    
    dados = create_table(vOwner,vWhere,vQueryCli)

    if dados != 0:
        return (dados)
    
    return(dados)

def create_table(vOwner,pWhere,vQueryCli):

############ MESTRE_NFTL_SERV
############ MESTRE_NFTL_SERV
############ MESTRE_NFTL_SERV
    ret = 0

    tabela_mestre = 'MESTRE_NFTL_SERV'

    vWhere = pWhere + " AND nf.EMPS_COD = 'TBRA'"
    vWhere = re.sub( r'serie'   , 'nf.MNFST_SERIE'  ,  vWhere)
    vWhere = re.sub( r'data_ini', 'nf.MNFST_DTEMISS',  vWhere)
    vWhere = re.sub( r'data_fim', 'nf.MNFST_DTEMISS',  vWhere)
    vWhere = re.sub( r'fili_cod', 'nf.FILI_COD'     ,  vWhere)
    vWhere = re.sub( r'num_nota', 'nf.MNFST_NUM'    ,  vWhere)

    con=sql.geraCnxBD(configuracoes)

    log( '# INICIANDO O PROCESSO DE CRIAÇÃO DA TABELA', tabela_mestre)
    
    log( '# TENTANDO DROPAR A TABELA', tabela_mestre)
    

    query="""
        DROP TABLE %s.%s PURGE
    """%(vOwner, tabela_mestre)
    try:
        con.executa(query)        
        log( '# TABELA DROPADA COM SUCESSO', tabela_mestre)
        
    except Exception as e:
        log( '#### TABELA NÃO EXISTE NO OWNER INFORMADO...')
            

    log( '# INICIANDO CRIAÇÃO DA TABELA', tabela_mestre)
    
    query="""
        CREATE TABLE %s.%s TABLESPACE CADASTRO_DATA as 
            SELECT /*+ PARALLEL (16) */ nf.* 
              FROM openrisow.mestre_nftl_serv@gfread nf
             WHERE %s
    """%(vOwner,tabela_mestre,vWhere)

    log(query)
    try:
        con.executa(query)
        log( '# TABELA CRIADA COM SUCESSO', tabela_mestre)
        
    except Exception as e:
        log('ERRO ASSINALADO: ', str(e))
        ret = 99
        return ret
    
    log( '# INICIANDO CRIAÇÃO DE PK DA TABELA', tabela_mestre)
    
    query="""ALTER TABLE %s.%s ADD CONSTRAINT PK_%s PRIMARY KEY (MNFST_DTEMISS,MNFST_SERIE,MNFST_NUM,FILI_COD,MDOC_COD,EMPS_COD)
          """%(vOwner,tabela_mestre,tabela_mestre)
    try:
        con.executa(query)
        log( '# PK CRIADA COM SUCESSO PARA A TABELA', tabela_mestre)
        
    except Exception as e:
        log('ERRO', str(e))
        ret = 99
        return ret

############ ITEM_NFTL_SERV
############ ITEM_NFTL_SERV
############ ITEM_NFTL_SERV

    tabela_item = 'ITEM_NFTL_SERV'

    vWhere = pWhere + " AND item.EMPS_COD = 'TBRA'"
    vWhere = re.sub( r'serie'   , 'item.INFST_SERIE'  ,  vWhere)
    vWhere = re.sub( r'data_ini', 'item.INFST_DTEMISS',  vWhere)
    vWhere = re.sub( r'data_fim', 'item.INFST_DTEMISS',  vWhere)
    vWhere = re.sub( r'fili_cod', 'item.FILI_COD'     ,  vWhere)
    vWhere = re.sub( r'num_nota', 'item.INFST_NUM'    ,  vWhere)

    query="""
        DROP TABLE %s.%s PURGE
    """%(vOwner, tabela_item)
    try:
        con.executa(query)
        log( '# TABELA DROPADA COM SUCESSO', tabela_item)
        
    except Exception as e:
        log( '#### TABELA NÃO EXISTE NO OWNER INFORMADO...')
            

    log( '# INICIANDO O CREATE DA TABELA', tabela_item)
    
    query="""
        CREATE TABLE %s.%s TABLESPACE CADASTRO_DATA as 
            SELECT /*+ PARALLEL (16) */ item.* 
              FROM openrisow.ITEM_NFTL_SERV@gfread item
             WHERE %s
    """%(vOwner,tabela_item,vWhere)

    log(query)
    try:
        con.executa(query)
        log( '# TABELA CRIADA COM SUCESSO', tabela_item)
        
    except Exception as e:
        log('ERRO ASSINALADO: ', str(e))
        ret = 99
        return ret
    
    log( '# INICIANDO CRIAÇÃO DE PK DA TABELA', tabela_item)
    
    query="""ALTER TABLE %s.%s ADD CONSTRAINT PK_%s PRIMARY KEY (INFST_NUM,INFST_DTEMISS,INFST_SERIE,INFST_NUM_SEQ,EMPS_COD,FILI_COD)
          """%(vOwner,tabela_item,tabela_item)
    try:
        con.executa(query)
        log( '# PK CRIADA COM SUCESSO PARA A TABELA', tabela_item)
        
    except Exception as e:
        log('ERRO', str(e))
        ret = 99
        return ret

############ CLI_FORNEC_TRANSP
############ CLI_FORNEC_TRANSP
############ CLI_FORNEC_TRANSP

    tabela_cli = 'CLI_FORNEC_TRANSP'
    vQueryCli = vQueryCli.replace('1 = 1', '')
    vWhere = vQueryCli + " AND nf.EMPS_COD = 'TBRA'"
    vWhere = re.sub( r'serie'   , 'nf.MNFST_SERIE'  ,  vWhere)
    vWhere = re.sub( r'data_ini', 'nf.MNFST_DTEMISS',  vWhere)
    vWhere = re.sub( r'data_fim', 'nf.MNFST_DTEMISS',  vWhere)
    vWhere = re.sub( r'fili_cod', 'nf.FILI_COD'     ,  vWhere)

    query="""
        DROP TABLE %s.%s PURGE
    """%(vOwner, tabela_cli)
    try:
        con.executa(query)
        log( '# TABELA DROPADA COM SUCESSO', tabela_cli)
        
    except Exception as e:
        log( '#### TABELA NÃO EXISTE NO OWNER INFORMADO...')
            

    log( '# INICIANDO O CREATE DA TABELA', tabela_cli)
    
    query="""
        CREATE TABLE %s.%s TABLESPACE CADASTRO_DATA as 
            SELECT /*+ PARALLEL(16) */  p.* 
              FROM %s.%s nf,
                   openrisow.cli_fornec_transp@gfread p
             WHERE p.cadg_cod = nf.cadg_cod 
               AND p.catg_cod = nf.catg_cod 
               AND p.cadg_dat_atua = ( SELECT MAX(temp.cadg_dat_atua) 
                                         FROM openrisow.cli_fornec_transp@gfread temp 
                                        WHERE temp.cadg_cod       = nf.cadg_cod 
                                          AND temp.CATG_COD       = nf.CATG_COD 
                                          AND temp.cadg_dat_atua <= nf.mnfst_dtemiss)
            %s
    """%(vOwner,tabela_cli,vOwner,tabela_mestre,vWhere)

    log(query)
    try:
        con.executa(query)
        log( '# TABELA CRIADA COM SUCESSO', tabela_cli)
        
    except Exception as e:
        log('ERRO ASSINALADO: ', str(e))
        ret = 99
        return ret
    
    log( '# INICIANDO CRIAÇÃO DE PK DA TABELA', tabela_cli)
    
    query="""
        ALTER TABLE %s.%s ADD CONSTRAINT PK_%s PRIMARY KEY (CADG_COD, CATG_COD, CADG_DAT_ATUA)
    """%(vOwner,tabela_cli,tabela_cli)
    try:
        con.executa(query)
        log( '# PK CRIADA COM SUCESSO PARA A TABELA', tabela_cli)
        
    except Exception as e:
        log('ERRO', str(e))
        ret = 99
        return ret


############ COMPLVU_CLIFORNEC
############ COMPLVU_CLIFORNEC
############ COMPLVU_CLIFORNEC

    tabela_cplvu = 'COMPLVU_CLIFORNEC'
    vQueryCli = vQueryCli.replace('1 = 1', '')
    vWhere = vQueryCli + " AND nf.EMPS_COD = 'TBRA'"
    vWhere = re.sub( r'serie'   , 'nf.MNFST_SERIE'  ,  vWhere)
    vWhere = re.sub( r'data_ini', 'nf.MNFST_DTEMISS',  vWhere)
    vWhere = re.sub( r'data_fim', 'nf.MNFST_DTEMISS',  vWhere)
    vWhere = re.sub( r'fili_cod', 'nf.FILI_COD'     ,  vWhere)

    query="""
        DROP TABLE %s.%s PURGE
    """%(vOwner, tabela_cplvu)
    try:
        con.executa(query)
        log( '# TABELA DROPADA COM SUCESSO', tabela_cplvu)
        
    except Exception as e:
        log( '#### TABELA NÃO EXISTE NO OWNER INFORMADO...')
            

    log( '# INICIANDO O CREATE DA TABELA', tabela_cplvu)
    
    query="""
        CREATE TABLE %s.%s TABLESPACE CADASTRO_DATA as 
            SELECT /*+ PARALLEL(16) */  p.* 
              FROM %s.%s nf,
                   openrisow.complvu_clifornec@gfread p
             WHERE p.cadg_cod = nf.cadg_cod 
               AND p.catg_cod = nf.catg_cod 
               AND p.cadg_dat_atua = ( SELECT MAX(temp.cadg_dat_atua) 
                                         FROM openrisow.complvu_clifornec@gfread temp 
                                        WHERE temp.cadg_cod       = nf.cadg_cod 
                                          AND temp.CATG_COD       = nf.CATG_COD 
                                          AND temp.cadg_dat_atua <= nf.mnfst_dtemiss)
            %s
    """%(vOwner,tabela_cplvu,vOwner,tabela_mestre,vWhere)

    log(query)
    try:
        con.executa(query)
        log( '# TABELA CRIADA COM SUCESSO', tabela_cplvu)
        
    except Exception as e:
        log('ERRO ASSINALADO: ', str(e))
        ret = 99
        return ret
    
    log( '# INICIANDO CRIAÇÃO DE PK DA TABELA', tabela_cplvu)
    
    query="""
        ALTER TABLE %s.%s ADD CONSTRAINT PK_%s PRIMARY KEY (CADG_COD, CATG_COD, CADG_DAT_ATUA)
    """%(vOwner,tabela_cplvu,tabela_cplvu)
    try:
        con.executa(query)
        log( '# PK CRIADA COM SUCESSO PARA A TABELA', tabela_cplvu)
        
    except Exception as e:
        log('ERRO', str(e))
        ret = 99
        return ret

############ SERVICO_TELCOM
############ SERVICO_TELCOM
############ SERVICO_TELCOM

    tabela_serv = 'SERVICO_TELCOM'

    query="""
        DROP TABLE %s.%s PURGE
    """%(vOwner, tabela_serv)
    try:
        con.executa(query)
        log( '# TABELA DROPADA COM SUCESSO', tabela_serv)
        
    except Exception as e:
        log( '#### TABELA NÃO EXISTE NO OWNER INFORMADO...')
            

    log( '# INICIANDO O CREATE DA TABELA', tabela_serv)
    
    query="""CREATE TABLE %s.%s TABLESPACE CADASTRO_DATA as 
                SELECT /*+ first_rows(1)*/
              DISTINCT st.*
                  FROM openrisow.servico_telcom@gfread st,
                       %s.%s tt
                 WHERE st.emps_cod        =   tt.emps_cod
                   AND st.fili_cod        =   tt.fili_cod
                   AND st.servtl_cod      =   tt.serv_cod
                   AND st.servtl_dat_atua = (SELECT MIN(st2.servtl_dat_atua)
                                               FROM openrisow.servico_telcom@gfread st2
                                              WHERE st2.emps_cod        = st.emps_cod
                                                AND st2.fili_cod        = st.fili_cod
                                                AND st2.servtl_cod      = st.servtl_cod)
          """%(vOwner,tabela_serv,vOwner,tabela_item)

    log(query)
    try:
        con.executa(query)
        log( '# TABELA CRIADA COM SUCESSO', tabela_serv)
        
    except Exception as e:
        log('ERRO ASSINALADO: ', str(e))
        ret = 99
        return ret
    
    log( '# INICIANDO CRIAÇÃO DE PK DA TABELA', tabela_serv)
    
    query="""
        ALTER TABLE %s.%s ADD CONSTRAINT PK_%s PRIMARY KEY (EMPS_COD, FILI_COD, SERVTL_DAT_ATUA, SERVTL_COD)
    """%(vOwner,tabela_serv,tabela_serv)
    try:
        con.executa(query)
        log( '# PK CRIADA COM SUCESSO PARA A TABELA', tabela_serv)
        
    except Exception as e:
        log('ERRO', str(e))
        ret = 99
        return ret

############ TEFA
############ TEFA
############ TEFA

    tabela_tefa = 'TEFA'

    vWhere = pWhere + " AND nf.EMPS_COD = 'TBRA'"
    vWhere = re.sub( r'serie'   , 'nf.MNFST_SERIE'  ,  vWhere)
    vWhere = re.sub( r'data_ini', 'nf.MNFST_DTEMISS',  vWhere)
    vWhere = re.sub( r'data_fim', 'nf.MNFST_DTEMISS',  vWhere)
    vWhere = re.sub( r'fili_cod', 'nf.FILI_COD'     ,  vWhere)
    vWhere = re.sub( r'num_nota', 'nf.MNFST_NUM'    ,  vWhere)

    con=sql.geraCnxBD(configuracoes)

    log( '# INICIANDO O PROCESSO DE CRIAÇÃO DA TABELA', tabela_tefa)
    
    log( '# TENTANDO DROPAR A TABELA', tabela_tefa)
    

    query="""
        DROP TABLE %s.%s PURGE
    """%(vOwner, tabela_tefa)
    try:
        con.executa(query)        
        log( '# TABELA DROPADA COM SUCESSO', tabela_tefa)
        
    except Exception as e:
        log( '#### TABELA NÃO EXISTE NO OWNER INFORMADO...')
            

    log( '# INICIANDO CRIAÇÃO DA TABELA', tabela_tefa)
    
    query="""
        CREATE TABLE %s.%s TABLESPACE CADASTRO_DATA as 
            SELECT /*+ PARALLEL (16) */ nf.* 
              FROM openrisow.tefa@gfread nf
             WHERE %s
    """%(vOwner,tabela_tefa,vWhere)

    log(query)
    try:
        con.executa(query)
        log( '# TABELA CRIADA COM SUCESSO', tabela_tefa)
        
    except Exception as e:
        log('ERRO ASSINALADO: ', str(e))
        ret = 99
        return ret
    
    log( '# INICIANDO CRIAÇÃO DE PK DA TABELA', tabela_tefa)
    
    query="""ALTER TABLE %s.%s ADD CONSTRAINT PK_%s PRIMARY KEY (EMPS_COD, FILI_COD, MNFST_SERIE, MNFST_NUM, MNFST_DTEMISS, TEFA_COD_AREA, TEFA_TERMINAL, TEFA_IND_SERV, TEFA_PER_FISCAL, TEFA_COD_AREA_T, TEFA_TERMINAL_T)
          """%(vOwner,tabela_tefa,tabela_tefa)
    try:
        con.executa(query)
        log( '# PK CRIADA COM SUCESSO PARA A TABELA', tabela_tefa)
        
    except Exception as e:
        log('ERRO', str(e))
        ret = 99
        return ret

    return ret

if __name__ == "__main__":
    ret = processar()
    if (ret > 0) :
        log("#### Código de execução = ", ret)
    sys.exit(ret) 
