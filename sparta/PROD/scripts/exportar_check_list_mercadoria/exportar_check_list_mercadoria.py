#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-
"""
----------------------------------------------------------------------------------------------
MODULO ...: TESHUVA
SCRIPT ...: exportar_check_list_mercadoria.py
CRIACAO ..: 28/07/2021
AUTOR ....: EDUARDO DA SILVA FERREIRA / KYROS TECNOLOGIA
            eduardof@kyros.com.br
DESCRICAO.: Execução da exportacao do levantamento realizado para a mercadoria
----------------------------------------------------------------------------------------------
PARAMETROS: 
Parâmetros de entrada:
1 ) P_CC_CONTROLE_GAP              : [OBRIGATÓRIO  ] CONTROLE de GAP informado no levantamento do check list [LOG]. Exemplo. PAINEL_123.
2 ) P_CC_GAPS                      : [OBRIGATÓRIO  ] Nome dos gaps cadastradas na tabela [GFCADASTRO.GAP_PROCEDIMENTO.GAP] para exportacao. Exemplo: EXPORTA_MESTRE_NFEN_MERC
3 ) P_DT_FILTRO_INICIO             : [OBRIGATÓRIO - FORMATO DD/MM/YYYY] Data de inicio para realizacao da pesquisa . Exemplo: 01/01/0001
4 ) P_DT_FILTRO_FIM                : [OBRIGATÓRIO - FORMATO DD/MM/YYYY] Data de fim    para realizacao da pesquisa . Exemplo: 31/12/9999

----------------------------------------------------------------------------------------------
    HISTORICO : 
        * 28/07/2021 - EDUARDO DA SILVA FERREIRA / KYROS TECNOLOGIA (eduardof@kyros.com.br)
        - Criacao do script.
        * 01/09/2021 - EDUARDO DA SILVA FERREIRA / KYROS TECNOLOGIA (eduardof@kyros.com.br)
        - PTITES-139: DV - Novo Padrão: SANEAMENTO / Mercadoria - Exportar Checklist
        - Alterado nome para : exportar_check_list_mercadoria.py
        
        * 20/09/2021 - AIRTON BORGES - KYROS
        - Formatadas todas as colunas da planilha excel gerada em tipo "texto"
        
----------------------------------------------------------------------------------------------
"""

import sys
import os
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)
import configuracoes
import comum
import sql
import datetime
import traceback
# Caso especifico devido ao EXCEL
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

# Demais configuração
comum.carregaConfiguracoes(configuracoes)
log.gerar_log_em_arquivo = True

import string
# Lista de String
gv_ob_lista_string = list(string.ascii_lowercase)

def ob_global():
    """
        Funcao falsa apenas para sergir de apoio para armazenar as variaveis globais 
    """
    pass

def fnc_conectar_banco_dados():
    """
        Funcao para conectar na base de dados 
    """
    try:
        
        ob_global.gv_ob_conexao = sql.geraCnxBD(configuracoes)
        v_ds_sql="""
        SELECT 'EXPORTACAO_' || TO_CHAR(SYSDATE,'YYYYMMDD_HH24MISS') NM_JOB FROM DUAL
        """
        ob_global.gv_ob_conexao.executa(v_ds_sql)
        v_ob_cursor = ob_global.gv_ob_conexao.fetchone()
        if (v_ob_cursor):    
            for campo in v_ob_cursor:
                log(str(campo) + " >> SUCESSO CONEXAO BANCO DE DADOS") 
                ob_global.gv_nm_job = str(campo)
                return 0
                break
        
        else:
            log("ERRO CONEXAO BANCO DE DADOS ") 
            return 91
        
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO CONEXAO BANCO DE DADOS .: " + str(e) + " - TRACE - " + v_ds_trace)
        return 91


def fnc_validar_entrada():
    """
        Retorna a validação de entrada e dos arquivos de configuração
    """
    try:
        v_nr_retorno = 0        
         
        log("-"*150)

        #--comum.addParametro('P_CC_ORIGEM_UF', None, 'Referente se sera acessado as tabelas oficiaos ou saneadas.', True,'FULL')
        comum.addParametro('P_CC_CONTROLE_GAP', None, 'CONTROLE de GAP informado no levantamento do check list.', True,'PAINEL_123')
        comum.addParametro('P_CC_GAPS', None, 'Nome dos gaps cadastradas na tabela [GFCADASTRO.GAP_PROCEDIMENTO.GAP] para exportacao.', True,'EXPORTA_MESTRE_NFEN_MERC')
        comum.addParametro('P_DT_FILTRO_INICIO', None, '[FORMATO DD/MM/YYYY] Data de inicio para realizacao da pesquisa.', True,'01/01/2015')
        comum.addParametro('P_DT_FILTRO_FIM', None, '[FORMATO DD/MM/YYYY] Data de fim    para realizacao da pesquisa.', True,'31/01/2015')
        
        # Validacao dos parametros de entrada
        if not comum.validarParametros() :
            v_nr_retorno = 91
        
        else:
        
            # INICIO ELSE
            ob_global.gv_cc_uf                           = 'FULL'#--comum.getParametro('P_CC_ORIGEM_UF').upper().strip()   
            ob_global.gv_cc_controle_gap                 = comum.getParametro('P_CC_CONTROLE_GAP').upper().strip()
            ob_global.gv_cc_gaps                         = comum.getParametro('P_CC_GAPS').upper().strip()
            ob_global.gv_dt_filtro_inicio                = comum.getParametro('P_DT_FILTRO_INICIO').upper().strip()
            ob_global.gv_dt_filtro_fim                   = comum.getParametro('P_DT_FILTRO_FIM').upper().strip()
            
            if not v_nr_retorno :
                if ob_global.gv_cc_uf.strip().upper() != 'FULL' \
                and len(ob_global.gv_cc_uf.strip()) == 2:
                    ob_global.gv_owner_gap                 = "MERC_" + str(ob_global.gv_cc_uf)  + "_"
                    ob_global.gv_owner_cadastro            = "MERC_" + str(ob_global.gv_cc_uf)  + "_RD_"
                    ob_global.gv_owner_prefixo             = "MERC_" + str(ob_global.gv_cc_uf)  + "_"
                    ob_global.gv_owner_prd_cadastro        = "PRD_"  + str(ob_global.gv_cc_uf)  + "_RD_"
                    ob_global.gv_owner_prd_prefixo         = "PRD_"  + str(ob_global.gv_cc_uf)  + "_"
                    
                elif ob_global.gv_cc_uf.strip().upper() != 'FULL':
                    log("PARAMETRO ORIGEM: Invalido! " + ob_global.gv_cc_uf) 
                    v_nr_retorno = 91 
                    
            if not v_nr_retorno :
                if (ob_global.gv_cc_controle_gap.count("TODO") > 0
                or ob_global.gv_cc_controle_gap.count("TODA") > 0 
                ):
                    ob_global.gv_cc_controle_gap = "TODOS"        

            if not v_nr_retorno :
                if (ob_global.gv_cc_gaps.count("TODO") > 0
                or ob_global.gv_cc_gaps.count("TODA") > 0 
                ):
                    ob_global.gv_cc_gaps = "TODOS"        
                    
            if not v_nr_retorno :
                try:
                    if (len(ob_global.gv_dt_filtro_inicio) != 10):
                        log("PARAMETRO DATA [INICIAL]: Invalido! " + ob_global.gv_dt_filtro_inicio) 
                        v_nr_retorno = 91           
                    else:
                        if (
                           int(ob_global.gv_dt_filtro_inicio[0:2]) > 31
                        or int(ob_global.gv_dt_filtro_inicio[0:2]) < 1
                        ):
                            log("PARAMETRO DIA [DATA INICIAL] : Invalido! " + ob_global.gv_dt_filtro_inicio[0:2]) 
                            v_nr_retorno = 91                         
                        elif (
                           int(ob_global.gv_dt_filtro_inicio[3:5]) > 12
                        or int(ob_global.gv_dt_filtro_inicio[3:5]) < 1
                        ):
                            log("PARAMETRO MES [DATA INICIAL] : Invalido! " + ob_global.gv_dt_filtro_inicio[3:5]) 
                            v_nr_retorno = 91                         
                        else:
                            try:
                                newDate = datetime.datetime(int(ob_global.gv_dt_filtro_inicio[6:10]) \
                                                          , int(ob_global.gv_dt_filtro_inicio[3:5]) \
                                                          , int(ob_global.gv_dt_filtro_inicio[0:2]))
                            except:
                                log("PARAMETRO DATA [INICIAL]: Invalido! " + ob_global.gv_dt_filtro_inicio) 
                                v_nr_retorno = 91                                                       

                except:
                    log("PARAMETRO DATA : Invalido [INICIAL]! " + ob_global.gv_dt_filtro_inicio) 
                    v_nr_retorno = 91            
                    
            if not v_nr_retorno :
                try:
                    if (len(ob_global.gv_dt_filtro_fim) != 10):
                        log("PARAMETRO DATA [FINAL]: Invalido! " + ob_global.gv_dt_filtro_fim) 
                        v_nr_retorno = 91           
                    else:
                        if (
                           int(ob_global.gv_dt_filtro_fim[0:2]) > 31
                        or int(ob_global.gv_dt_filtro_fim[0:2]) < 1
                        ):
                            log("PARAMETRO DIA [DATA FINAL] : Invalido! " + ob_global.gv_dt_filtro_fim[0:2]) 
                            v_nr_retorno = 91                         
                        elif (
                           int(ob_global.gv_dt_filtro_fim[3:5]) > 12
                        or int(ob_global.gv_dt_filtro_fim[3:5]) < 1
                        ):
                            log("PARAMETRO MES [DATA FINAL] : Invalido! " + ob_global.gv_dt_filtro_fim[3:5]) 
                            v_nr_retorno = 91                         
                        else:
                            try:
                                newDate = datetime.datetime(int(ob_global.gv_dt_filtro_fim[6:10]) \
                                                            , int(ob_global.gv_dt_filtro_fim[3:5]) \
                                                            , int(ob_global.gv_dt_filtro_fim[0:2]))
                            except:
                                log("PARAMETRO DATA [FINAL]: Invalido! " + ob_global.gv_dt_filtro_fim) 
                                v_nr_retorno = 91          
                except:
                    log("PARAMETRO DATA : Invalido [FINAL]! " + ob_global.gv_dt_filtro_fim) 
                    v_nr_retorno = 91

            if not v_nr_retorno :
                try:
                    l_dt_ini_yyyymmdd = str(ob_global.gv_dt_filtro_inicio[6:10]) \
                                      + str(ob_global.gv_dt_filtro_inicio[3:5]) \
                                      + str(ob_global.gv_dt_filtro_inicio[0:2])
                    l_dt_fim_yyyymmdd   = str(ob_global.gv_dt_filtro_fim[6:10]) \
                                        + str(ob_global.gv_dt_filtro_fim[3:5]) \
                                        + str(ob_global.gv_dt_filtro_fim[0:2])    
                    if int(l_dt_fim_yyyymmdd) < int(l_dt_ini_yyyymmdd):
                        log("PARAMETRO DATA [FINAL] : Invalido! NÃO PODE SER MENOR QUE O INICIAL !") 
                        v_nr_retorno = 91                         
                except:
                    log("PARAMETRO DATA [FINAL]: Invalido! " + ob_global.gv_dt_filtro_fim) 
                    v_nr_retorno = 91
            
            if not v_nr_retorno :
                ob_global.gv_cc_caminho_relatorio = configuracoes.dir_geracao
                if len(ob_global.gv_cc_caminho_relatorio) == 0:
                    log("PARAMETRO RELATORIO: INVALIDO! " + str(gv_cc_caminho_relatorio)) 
                    v_nr_retorno = 91

            if not v_nr_retorno :
                try:
                    if not os.path.isdir(ob_global.gv_cc_caminho_relatorio):
                        log("Diretório não ecxistente : " + ob_global.gv_cc_caminho_relatorio)        
                        os.makedirs(ob_global.gv_cc_caminho_relatorio)
                        log("Diretório criado : " + ob_global.gv_cc_caminho_relatorio)
                
                except Exception as e:
                    v_ds_trace = traceback.format_exc()
                    log(ob_global.gv_cc_caminho_relatorio + " >> PARAMETRO DO DIRETÓRIO INVÁLIDO! " + str(e) + " >> " + v_ds_trace) 
                    v_nr_retorno = 91   

            if not v_nr_retorno :
                try:

                    v_cc_caminho = ""
                
                    v_cc_resultado = ob_global.gv_cc_controle_gap.strip().upper()
                    if v_cc_resultado.count(".") > 0:
                        v_cc_lista = v_cc_resultado.split(".")
                        v_cc_caminho = str(v_cc_lista[0])
                    else:
                        v_cc_caminho = v_cc_resultado       
                    
                    v_cc_caminho = v_cc_caminho + "_" + ob_global.gv_cc_gaps + "_" + str(datetime.datetime.now().strftime('%Y%m%d_%H%M%S')) + ".xlsx"
                
                    ob_global.gv_cc_caminho_relatorio = os.path.join(ob_global.gv_cc_caminho_relatorio, v_cc_caminho)
                
                except Exception as e:
                    v_ds_trace = traceback.format_exc()
                    log(ob_global.gv_cc_caminho_relatorio + " >> PARAMETRO DO DIRETÓRIO INVÁLIDO! " + str(e) + " >> " + v_ds_trace) 
                    v_nr_retorno = 91     
                
            # FIM ELSE
            
        return v_nr_retorno
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO VALIDAÇÃO DOS PARAMETROS DE ENTRADA: " + str(e)+ " >> " + v_ds_trace)
        v_nr_retorno = 93
        return v_nr_retorno


def fnc_processar():
    """
        Funcao principal para processar as informacoes
    """
    try:
        
        v_nr_retorno = 0

        log("-"*150)
      
        if ob_global.gv_ob_conexao is not None:
            log("conexao              " , "ATIVO"              )           
        
        #--log("origem [uf]          " , ob_global.gv_cc_uf                   )    
        #--log("job                  " , ob_global.gv_nm_job                  )
        log("controle_gap         " , ob_global.gv_cc_controle_gap         )     
        log("gaps                 " , ob_global.gv_cc_gaps                 )     
        log("filtro_inicio        " , ob_global.gv_dt_filtro_inicio        )     
        log("filtro_fim           " , ob_global.gv_dt_filtro_fim           )     
        
        log("caminho_relatorio    " , ob_global.gv_cc_caminho_relatorio            ) 
        
        #--log("owner_cadastro       " , ob_global.gv_owner_cadastro                  ) 
        #--log("owner_prefixo        " , ob_global.gv_owner_prefixo                   )
        #--log("owner_prd_cadastro   " , ob_global.gv_owner_prd_cadastro              )
        #--log("owner_prd_prefixo    " , ob_global.gv_owner_prd_prefixo               )
        #--log("owner_gap            " , ob_global.gv_owner_gap                       )
                
        log("-"*150)    
              
        v_cd_erro = ob_global.gv_ob_conexao.var(int)
        v_ds_erro = ob_global.gv_ob_conexao.var(str)
        v_cc_retorno = ob_global.gv_ob_conexao.var('CLOB')
        
        ob_global.gv_ob_conexao.executa("""
                                        BEGIN
                                        
                                        :P_RETORNO := GFCADASTRO.TSH_SANMC_RETORNA_CONSULTA(P_COD_ERRO=>:P_COD_ERRO
                                                                                         , P_DESC_ERRO=>:P_DESC_ERRO
                                                                                         , P_CONTROLE_GAP_LEVANTAMENTO=>:P_CONTROLE_GAP_LEVANTAMENTO     
                                                                                         , P_GAP_EXPORTACAO=>:P_GAP_EXPORTACAO
                                                                                         , P_DT_INICIO=>:P_DT_INICIO
                                                                                         , P_DT_DIM=>:P_DT_DIM
                                                                                         );
                                        
                                        END;    
                                        """
                                        , P_RETORNO=v_cc_retorno         
                                        , P_COD_ERRO=v_cd_erro
                                        , P_DESC_ERRO=v_ds_erro
                                        , P_CONTROLE_GAP_LEVANTAMENTO=ob_global.gv_cc_controle_gap     
                                        , P_GAP_EXPORTACAO=ob_global.gv_cc_gaps
                                        , P_DT_INICIO=ob_global.gv_dt_filtro_inicio
                                        , P_DT_DIM=ob_global.gv_dt_filtro_fim
                                        )  
        log(v_cd_erro.getvalue()) 
        log(v_ds_erro.getvalue())
        
        try:
            v_nr_retorno = v_cd_erro.getvalue()
        except:
            v_nr_retorno = 1

        ob_global.gv_ob_conexao.commit()
        log("")
        
        if v_nr_retorno < 0:
            v_nr_retorno = 1
        else:
            ob_global.gv_cc_consulta_relatorio = v_cc_retorno.getvalue()	
            
            log("")
            try:
                log(ob_global.gv_cc_consulta_relatorio)
            except:
                pass
            log("")
            
        return v_nr_retorno
        
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO PROCESSAMENTO: " + str(e)+ " >> " + v_ds_trace)
        v_nr_retorno = 93
        return v_nr_retorno        

def fnc_gerar_relatorio():
    """
        Funcao principal para gerar o relatorio
    """
    try:
        
        log("Iniciando a geração de relatorio ... ")
        
        # Inicializa as variaveis
        v_nr_retorno = 0
        v_nr_linha = 0
        v_nr_coluna = 1
        v_ob_lista_retorno=[]
        
        log("Executa a consulta principal ... ")
        # Abrir o cursor
        ob_global.gv_ob_conexao.executa(str(ob_global.gv_cc_consulta_relatorio))
        v_ob_resultado_cursor = ob_global.gv_ob_conexao.fetchone()
        log("Iniciando a carregamento dos dados na memoria ... ")
        if v_ob_resultado_cursor:
            
            # Nomes das colunas
            v_ob_lista_retorno.append([])
            for cc_linha in ob_global.gv_ob_conexao.description():
                v_ob_lista_retorno[v_nr_linha].append(str(cc_linha[0]))
                v_nr_coluna += 1    
            v_nr_linha += 1
            
            # Valores das colunas
            while v_ob_resultado_cursor:
                v_ob_lista_retorno.append([])
                for cc_linha in v_ob_resultado_cursor:
                    v_ob_lista_retorno[v_nr_linha].append(cc_linha)                
                v_nr_linha += 1
                v_ob_resultado_cursor = ob_global.gv_ob_conexao.fetchone()
            
        if v_nr_linha < 2:
            log("")
            log("NÃO FOI ENCONTRADO NENHUM DADO!")
            log("")
            v_nr_retorno = 1
            return v_nr_retorno
        
        if not v_nr_retorno and v_nr_linha > 1:
            try:
                log("Criando o arquivo : " + ob_global.gv_cc_caminho_relatorio)        
                v_ob_arq_excel = Workbook()
                v_ob_wks_01_arq_excel = v_ob_arq_excel.active
                v_ob_wks_01_arq_excel.title = ob_global.gv_cc_gaps
                
                log("Atribui os dados do arquivo ... ")   
                for cc_dado in v_ob_lista_retorno:
                   
                    v_ob_wks_01_arq_excel.append(cc_dado)           
                
                log("# Regulariza a dimensao das planilhas ... ")   
                for ds_col in gv_ob_lista_string:                
                    v_ob_wks_01_arq_excel.column_dimensions[str(ds_col).upper()].width = 30   
                    v_ob_wks_01_arq_excel.column_dimensions["A"+str(ds_col).upper()].width = 30  
                
                log("# Regulariza o fonte das planilhas ... ")  
                #for nr_col in range(1,v_nr_coluna):
                #    v_ob_wks_01_arq_excel.cell(1,nr_col).font=Font(bold=True)
                



# =============================================================================
# #AIRTON 29/09/2021 - INICIO               
#                 #altera o tipo de dado dos registros da planilha:
#                 m_coluna = v_ob_wks_01_arq_excel.max_column
#                 m_linha  = v_ob_wks_01_arq_excel.max_row
# 
#                 for l in range(1,m_linha+1):
#                     for c in range(1,m_coluna+1):
#                         tipo = type(v_ob_wks_01_arq_excel.cell(row=l,column=c).value) 
#                         if (v_ob_wks_01_arq_excel.cell(row=l,column=c).value == None):
#                             v_ob_wks_01_arq_excel.cell(row=l,column=c).value == ' '
#                         elif (tipo in (datetime, datetime.date, datetime.datetime, datetime.time)):
#                             v_ob_wks_01_arq_excel.cell(row=l,column=c).value = v_ob_wks_01_arq_excel.cell(row=l,column=c).value.strftime("%d/%m/%Y, %H:%M:%S")
# #                        elif (tipo in (datetime, datetime.date, datetime.datetime, datetime.time)):
# #                            v_ob_wks_01_arq_excel.cell(row=l,column=c).value = v_ob_wks_01_arq_excel.cell(row=l,column=c).value.strftime("%d/%m/%Y, %H:%M:%S")
#                         else:
#                            v_ob_wks_01_arq_excel.cell(row=l,column=c).value = str(v_ob_wks_01_arq_excel.cell(row=l,column=c).value)
#                         v_ob_wks_01_arq_excel.cell(row=l,column=c).number_format = '@'
#  #AIRTON 29/09/2021 - FINAL               
#          
# =============================================================================

        
#AIRTON 29/09/2021 - INICIO               
                #altera o tipo de dado dos registros da planilha:
                m_coluna = v_ob_wks_01_arq_excel.max_column
                m_linha  = v_ob_wks_01_arq_excel.max_row

                for l in range(1,m_linha+1):
                    for c in range(1,m_coluna+1):

                        tipo = type(v_ob_wks_01_arq_excel.cell(row=l,column=c).value)
                        
                        if (tipo == float):
#                            print("tipo identificado = float")
#                            v_ob_wks_01_arq_excel.cell(row=l,column=c).number_format = "#,##0.00"
                            None                            

                        elif (tipo == str):
#                            print("tipo identificado = str")
                            v_ob_wks_01_arq_excel.cell(row=l,column=c).value = str(v_ob_wks_01_arq_excel.cell(row=l,column=c).value)
                            v_ob_wks_01_arq_excel.cell(row=l,column=c).number_format = '@'

                        elif (tipo == int):
#                            print("tipo identificado = int")
                            None
                            
                        elif (tipo == tuple):
#                            print("tipo identificado = tuple")
                            v_ob_wks_01_arq_excel.cell(row=l,column=c).value = str(v_ob_wks_01_arq_excel.cell(row=l,column=c).value)
                            v_ob_wks_01_arq_excel.cell(row=l,column=c).number_format = '@'

                        elif (tipo == list):
#                            print("tipo identificado = list")
                            v_ob_wks_01_arq_excel.cell(row=l,column=c).value = str(v_ob_wks_01_arq_excel.cell(row=l,column=c).value)
                            v_ob_wks_01_arq_excel.cell(row=l,column=c).number_format = '@'

                        elif (tipo in (datetime, datetime.date, datetime.datetime, datetime.time)):
#                            print("tipo identificado = datetime")
                            v_ob_wks_01_arq_excel.cell(row=l,column=c).value = v_ob_wks_01_arq_excel.cell(row=l,column=c).value.strftime("%d/%m/%Y, %H:%M:%S")
                            v_ob_wks_01_arq_excel.cell(row=l,column=c).number_format = '@'

                        else:
#                            print("tipo NÃO identificado = ", tipo)
                            v_ob_wks_01_arq_excel.cell(row=l,column=c).value = str(v_ob_wks_01_arq_excel.cell(row=l,column=c).value)
                            v_ob_wks_01_arq_excel.cell(row=l,column=c).number_format = '@'
                        
                        if(v_ob_wks_01_arq_excel.cell(row=l,column=c).value in (" None", "None", None, "") ):
                            v_ob_wks_01_arq_excel.cell(row=l,column=c).value = " "
                            
                       
                            
#AIRTON 29/09/2021 - FINAL 





                # Grava a planilha Excel
                log("Grava o arquivo ... ") 
                v_ob_arq_excel.save(ob_global.gv_cc_caminho_relatorio)
                
            except Exception as e:
                v_ds_trace = traceback.format_exc()
                log("ERRO GERAÇÃO DO RELATORIO: " + str(e)+ " >> " + v_ds_trace)
                v_nr_retorno = 93
                
        return v_nr_retorno
        
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO PROCESSAMENTO DO RELATORIO: " + str(e)+ " >> " + v_ds_trace)
        v_nr_retorno = 93

        
if __name__ == "__main__" :
    """
        Ponto de partida
    """    
    # Codigo de Retorno
    v_nr_ret = 0

    # Tratamento de excessao
    v_ds_trace = ''

    # Tratamento de variaveis globais
    ob_global.gv_ob_conexao = None
    # Parametros do arquivo de configuração
    ob_global.gv_nm_job = "" 
    ob_global.gv_cc_consulta_relatorio = ""
    ob_global.gv_cc_caminho_relatorio = ""  


    ob_global.gv_owner_cadastro            = "OPENRISOW."
    ob_global.gv_owner_prefixo             = "OPENRISOW."
    ob_global.gv_owner_prd_cadastro        = "OPENRISOW."
    ob_global.gv_owner_prd_prefixo         = "OPENRISOW."
    ob_global.gv_owner_openrisow           = "OPENRISOW."
    ob_global.gv_owner_gfcarga             = "GFCARGA."
    ob_global.gv_owner_gfcadastro          = "GFCADASTRO."
    ob_global.gv_owner_gap                 = "GFCADASTRO."
    
    try:

        log("-"*100)
        log(" INICIO DA EXECUÇÃO ".center(120,'#'))
            
        log(" ")
        
        # Validacao dos parametros de entrada
        if not v_nr_ret :
            v_nr_ret = fnc_validar_entrada()
        
        log(" ")
        
        # Verificar conexao com o banco
        if not v_nr_ret :
            v_nr_ret = fnc_conectar_banco_dados()   
        
        log(" ")

        # Processar         
        if not v_nr_ret :
            v_nr_ret = fnc_processar()                    

        log(" ")
        
        # Gerar relatorio
        if not v_nr_ret :
            v_nr_ret = fnc_gerar_relatorio()    
        
        # Finalizacao
        log(" ")            
        
        if not v_nr_ret :
            log("SUCESSO NA EXECUÇÃO")
        else:
            log("ERRO NA EXECUÇÃO")
                        
        log(" ")
    
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO .: " + str(e) + " >> " + v_ds_trace)
        v_nr_ret = 93
    
    sys.exit(v_nr_ret if v_nr_ret >= log.ret else log.ret )

