#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-

"""
-----------------------------------------------------------------------------------------------------------------------
  SISTEMA ..: SPARTA
  MODULO ...: TESHUVA
  SCRIPT ...: relatorio_comparativo_rj.py
  CRIACAO ..: 22/11/2021
  AUTOR ....: Eduardo da Silva Ferreira (eduardof@kyros.com.br) 
              / KYROS Consultoria
  DESCRICAO : Este relatório tem por finalidade gerar um comparativo do SPED entre o Arquivo de PROTOCOLADO e o Atual . 
-----------------------------------------------------------------------------------------------------------------------
  HISTORICO :
    * 22/11/2021 - Eduardo da Silva Ferreira (eduardof@kyros.com.br) 
                   / KYROS Consultoria - Criacao do script.
------------------------------------------------------------------------------------------------------------------------
"""
#### PATRONIZACAO PARA O PAINEL DE EXECUCOES....
import sys
import os
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)
import configuracoes
import comum
import sql
import layout
import util
import vie 

#
import datetime
import traceback
import re
import string

#
import operator
import glob
from pathlib import Path
import openpyxl
import copy
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# Demais configuração
comum.carregaConfiguracoes(configuracoes)
#layout.carregaLayout()
log.gerar_log_em_arquivo = True

class cls_sped(object):
    """
        Classe para guardar as informações de SPED 
    """
    nr_count = 0
    def __init__(self):
        cls_sped.nr_count += 1
        self.CODIGO             = cls_sped.nr_count
        self.TIPO_NF            = ''
        self.NUMERO_NF          = ''
        self.SERIE              = ''
        self.MODELO             = ''
        self.DATA_EMISSAO       = ''
        self.DATA_ENTRADA       = ''
        self.CODIGO_PF_PJ       = ''
        self.CHAVE_NFE          = ''
        self.CFOP               = ''
        self.NATUREZA_OPERACAO  = ''
        self.CST                = ''
        self.PRECO_TOTAL        = float(0)
        self.VALOR_CONTABIL     = float(0)
        self.BASE_ICMS          = float(0)
        self.ALIQUOTA_ICMS      = float(0)
        self.VALOR_ICMS         = float(0)
        self.VALOR_IPI          = float(0)
        self.VALOR_DESCONTO     = float(0)
        self.VALOR_ISENTAS      = float(0)
        self.VALOR_OUTRAS       = float(0)
        self.BASE_ICMS_ST       = float(0)
        self.ALIQUOTA_ICMS_ST   = float(0)
        self.VALOR_ICMS_ST      = float(0)
        self.CODIGO_MATERIAL    = ''
        self.SEQ_ITEM           = '0' 
        self.SEQ                = int(0)
        self.TIPO_REG           = ''
        self.GRUPO_REG_SPED     = ''
        self.TIPO_SPED          = ''
        self.MES_ANO_SPED       = ''
        self.UF_SPED            = ''
        self.IE_SPED            = ''
        self.VALIDACAO          = ''
        self.NUMERO             = int(0)
        self.NR_SEQUENCIAL_CONTROLE     = int(0)

def as_text(value):
    if value is None:
        return ""
    return str(value)        
def fnc_column_dimensions(p_ob_ws):

    for column_cells in p_ob_ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        p_ob_ws.column_dimensions[column_cells[0].column_letter].width = length+15        
    
def ob_global():
    """
        Funcao falsa apenas para sergir de apoio para armazenar as variaveis globais 
    """
    pass


def fnc_ultimo_arquivo_diretorio(p_ds_mascara,p_ds_diretorio):
    """
        Funcao para retornar arquivos existentes
    """
    v_nr_qt = 0
    v_nm_arquivo = ""
    v_ds_dir = Path(p_ds_diretorio)
    v_nm_arq = v_ds_dir.glob(p_ds_mascara)
    v_cc_procura_arquivos = sorted(v_nm_arq, reverse=False)
    if v_cc_procura_arquivos:        
        for arquivo in v_cc_procura_arquivos:
            v_nr_qt = v_nr_qt + 1
            v_nm_arquivo = str(arquivo)

    return(v_nm_arquivo)

def fnc_conectar_banco_dados():
    """
        Funcao para conectar na base de dados 
    """
    try:
        
        ob_global.gv_ob_conexao = sql.geraCnxBD(configuracoes)
        v_ds_sql="""
        SELECT 'PAINELEXECUCAO_'||TO_CHAR(SYSDATE,'YYYYMMDD_HH24MISS') NM_JOB FROM DUAL
        """
        ob_global.gv_ob_conexao.executa(v_ds_sql)
        
        v_ob_cursor = ob_global.gv_ob_conexao.fetchone()
        if (v_ob_cursor):    
            for campo in v_ob_cursor:
                log(str(campo) + " SUCESSO CONEXAO BANCO DE DADOS") 
                ob_global.gv_nm_job = str(campo)
                break
        
        else:
            log("ERRO CONEXAO BANCO DE DADOS ") 
            return 91
            
        return 0
                
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO CONEXAO BANCO DE DADOS .: " + str(e) + " - TRACE - " + v_ds_trace)
        return 91

def fnc_processa_sped(p_nm_arquivo,p_tp_arquivo="PROTOCOLADO"):
    """
        Funcao processa arquivo do tipo SPED
    """
    
    # inicializa a variaveis de controle
    v_cc_reg_rel=[]
    v_cc_reg_rel_saida=[]
    v_nr_contador_nf = 0
    v_fl_nf = 0
    v_ds_linha = ""
    try:
    
        if os.path.isfile(p_nm_arquivo):
            
            log(p_tp_arquivo + " >> Processando leitura do arquivo : " + p_nm_arquivo)
            
            v_ob_encoding = comum.encodingDoArquivo( p_nm_arquivo )
            v_ob_fd = open(p_nm_arquivo, 'r', encoding=v_ob_encoding)
                        
            for v_cc_linha in v_ob_fd:
                #if len(v_cc_reg_rel) > 150:
                #    break               
                # Quebra em vetor        
                v_ds_linha = v_cc_linha
                v_cc_dados = v_ds_linha.split("|")
    
                if (len(v_cc_dados) < 7):
                    continue
                
                # Valida se os tipos        
                if v_cc_dados[1].upper().strip() \
                in ('C170','C100','C190','C500','C590','D100','D190','D500','D590'):
                    # inicio if
                    
                    if v_cc_dados[1].upper().strip() \
                    in ('C100','C500','D100','D500'):
                        # if v_fl_nf > 0:
                        #     v_cc_reg_rel.append(ob_sped)
                        ob_sped = cls_sped()                    
                        v_nr_contador_nf += 1
                        v_fl_nf = 0 
                        ob_sped.NR_SEQUENCIAL_CONTROLE = int(0)
                        ob_sped.CODIGO_MATERIAL = ''
                        ob_sped.TIPO_REG = str(v_cc_dados[1].upper().strip())
                        ob_sped.VALIDACAO = 'MESTRE'
                        ob_sped.TIPO_NF = 'ENTRADA'
                        ob_sped.SEQ = int(0)
                        ob_sped.SEQ_ITEM = '0'
                        ob_sped.CODIGO_PF_PJ = v_cc_dados[4].upper().strip()
                        ob_sped.MODELO = v_cc_dados[5].upper().strip()
                        ob_sped.SERIE = v_cc_dados[7].upper().strip()                    
                        if v_cc_dados[2].upper().strip() == '1':
                            ob_sped.TIPO_NF = 'SAIDA'
                        if v_cc_dados[1].upper().strip() == 'C100':
                            ob_sped.NUMERO_NF = v_cc_dados[8].upper().strip()
                            ob_sped.CHAVE_NFE = v_cc_dados[9].upper().strip()
                            ob_sped.DATA_EMISSAO = v_cc_dados[10].upper().strip()                        
                            ob_sped.DATA_ENTRADA = v_cc_dados[11].upper().strip()                        
                            try:
                                ob_sped.VALOR_DESCONTO = float(str(v_cc_dados[14].upper().strip()).replace(',', '.'))
                            except:
                                ob_sped.VALOR_DESCONTO = float(0)
                            try:
                                ob_sped.PRECO_TOTAL = float(str(v_cc_dados[16].upper().strip()).replace(',', '.'))
                            except:
                                ob_sped.PRECO_TOTAL = float(0)                               
                        elif v_cc_dados[1].upper().strip() == 'C500':
                            ob_sped.NUMERO_NF = v_cc_dados[10].upper().strip()
                            ob_sped.DATA_EMISSAO = v_cc_dados[11].upper().strip()
                            ob_sped.DATA_ENTRADA = v_cc_dados[12].upper().strip()
                            try:
                                ob_sped.VALOR_DESCONTO = float(str(v_cc_dados[14].upper().strip()).replace(',', '.'))
                            except:
                                ob_sped.VALOR_DESCONTO = float(0)
                            try:
                                ob_sped.PRECO_TOTAL = float(str(v_cc_dados[15].upper().strip()).replace(',', '.'))
                            except:
                                ob_sped.PRECO_TOTAL = float(0)                               
                        elif v_cc_dados[1].upper().strip() == 'D100':
                            ob_sped.NUMERO_NF = v_cc_dados[9].upper().strip()
                            ob_sped.CHAVE_NFE = v_cc_dados[10].upper().strip()                        
                            ob_sped.DATA_EMISSAO = v_cc_dados[11].upper().strip()
                            ob_sped.DATA_ENTRADA = v_cc_dados[12].upper().strip()
                            try:
                                ob_sped.VALOR_DESCONTO = float(str(v_cc_dados[16].upper().strip()).replace(',', '.'))
                            except:
                                ob_sped.VALOR_DESCONTO = float(0)
                            try:
                                ob_sped.PRECO_TOTAL = float(str(v_cc_dados[18].upper().strip()).replace(',', '.'))
                            except:
                                ob_sped.PRECO_TOTAL = float(0)                                       
                        elif v_cc_dados[1].upper().strip() == 'D500':
                            ob_sped.NUMERO_NF = v_cc_dados[9].upper().strip()    
                            ob_sped.DATA_EMISSAO = v_cc_dados[10].upper().strip()
                            ob_sped.DATA_ENTRADA = v_cc_dados[11].upper().strip()
                            try:
                                ob_sped.VALOR_DESCONTO = float(str(v_cc_dados[13].upper().strip()).replace(',', '.'))
                            except:
                                ob_sped.VALOR_DESCONTO = float(0)
                            try:
                                ob_sped.PRECO_TOTAL = float(str(v_cc_dados[14].upper().strip()).replace(',', '.'))
                            except:
                                ob_sped.PRECO_TOTAL = float(0)                                   
                        ob_sped.NUMERO = int(ob_sped.NUMERO_NF)
                        
                    ob_sped.TIPO_SPED = p_tp_arquivo
                    ob_sped.MES_ANO_SPED = ob_global.gv_mes_ano
                    ob_sped.UF_SPED = ob_global.gv_uf
                    ob_sped.IE_SPED = ob_global.gv_ie                    
                    ob_sped.GRUPO_REG_SPED = str(v_cc_dados[1].upper().strip())[0:1]                                  
                    if v_cc_dados[1].upper().strip() == 'C170':
                        ob_sped.TIPO_REG = ob_sped.TIPO_REG + '>>' + str(v_cc_dados[1].upper().strip())  
                        ob_sped.NR_SEQUENCIAL_CONTROLE += 1 
                        ob_sped_item = copy.deepcopy(ob_sped)
                        ob_sped_item.VALIDACAO = 'ITEM'
                        try:
                            ob_sped_item.CST = str(v_cc_dados[10].upper().strip())
                        except:
                            ob_sped_item.CST = ''         
                        try:
                            ob_sped_item.CODIGO_MATERIAL = str(v_cc_dados[3].upper().strip())
                        except:
                            ob_sped_item.CODIGO_MATERIAL = ''         
                        ob_sped_item.CFOP = v_cc_dados[11].upper().strip()
                        ob_sped_item.NATUREZA_OPERACAO = v_cc_dados[12].upper().strip()  
                        ob_sped_item.SEQ_ITEM = str(v_cc_dados[2].upper().strip())
                        try:
                            ob_sped_item.SEQ = int(str(str(v_cc_dados[2].upper().strip()).replace(',', '')).replace('.', ''))
                        except:
                            ob_sped_item.SEQ = int(1)
                        try:
                            ob_sped_item.PRECO_TOTAL = float(str(v_cc_dados[7].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_item.PRECO_TOTAL = float(0)                                
                        try:
                            ob_sped_item.VALOR_DESCONTO = float(str(v_cc_dados[8].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_item.VALOR_DESCONTO = float(0)                             
                        try:
                            ob_sped_item.BASE_ICMS = float(str(v_cc_dados[13].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_item.BASE_ICMS = float(0)                          
                        try:
                            ob_sped_item.ALIQUOTA_ICMS = float(str(v_cc_dados[14].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_item.ALIQUOTA_ICMS = float(0)                          
                        try:
                            ob_sped_item.VALOR_ICMS = float(str(v_cc_dados[15].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_item.VALOR_ICMS = float(0)                      
                        try:
                            ob_sped_item.VALOR_IPI = float(str(v_cc_dados[24].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_item.VALOR_IPI = float(0)                        
                        try:
                            ob_sped_item.BASE_ICMS_ST = float(str(v_cc_dados[16].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_item.BASE_ICMS_ST = float(0)                          
                        try:
                            ob_sped_item.ALIQUOTA_ICMS_ST = float(str(v_cc_dados[17].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_item.ALIQUOTA_ICMS_ST = float(0)                          
                        try:
                            ob_sped_item.VALOR_ICMS_ST = float(str(v_cc_dados[18].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_item.VALOR_ICMS_ST = float(0)
                        if ob_sped_item.TIPO_NF == 'SAIDA':
                            v_cc_reg_rel_saida.append(ob_sped_item)
                        else:    
                            v_cc_reg_rel.append(ob_sped_item)
                        
                    if v_cc_dados[1].upper().strip() in ('C190','C590','D190','D590'):  
                        if v_cc_dados[1].upper().strip() == 'C190' \
                        and 'C170' in ob_sped.TIPO_REG:
                            continue
                        v_fl_nf += 1 
                        ob_sped.TIPO_REG = ob_sped.TIPO_REG + '>>' + str(v_cc_dados[1].upper().strip())  
                        ob_sped.NR_SEQUENCIAL_CONTROLE += 1 
                        ob_sped_resumo = copy.deepcopy(ob_sped)
                        ob_sped_resumo.VALIDACAO = 'CONSOLIDADO'
                        ob_sped_resumo.SEQ = int(0)
                        ob_sped_resumo.SEQ_ITEM = '0'
                        try:
                            ob_sped_resumo.CST = str(v_cc_dados[2].upper().strip())
                        except:
                            ob_sped_resumo.CST = ''                              
                        ob_sped_resumo.CFOP = v_cc_dados[3].upper().strip()
                        try:
                            ob_sped_resumo.ALIQUOTA_ICMS = float(str(v_cc_dados[4].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_resumo.ALIQUOTA_ICMS = float(0)
                        try:
                            ob_sped_resumo.VALOR_CONTABIL = float(str(v_cc_dados[5].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_resumo.VALOR_CONTABIL = float(0)
                        try:
                            ob_sped_resumo.BASE_ICMS = float(str(v_cc_dados[6].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_resumo.BASE_ICMS = float(0)
                        try:
                            ob_sped_resumo.VALOR_ICMS = float(str(v_cc_dados[7].upper().strip()).replace(',', '.'))
                        except:
                            ob_sped_resumo.VALOR_ICMS = float(0)                                       
                        if v_cc_dados[1].upper().strip() == 'C190':
                            try:
                                ob_sped_resumo.VALOR_IPI = float(str(v_cc_dados[11].upper().strip()).replace(',', '.'))
                            except:
                                ob_sped_resumo.VALOR_IPI = float(0)      
                            if str(v_cc_dados[2].upper().strip())[-2:] in ("20", "40", "41"):   
                                try:
                                    ob_sped_resumo.VALOR_ISENTAS = float(str(v_cc_dados[5].upper().strip()).replace(',', '.'))
                                except:
                                    ob_sped_resumo.VALOR_ISENTAS = float(0)    
                            if str(v_cc_dados[2].upper().strip())[-2:] in ("30", "50", "51", "60", "70", "90"):         
                                try:
                                    ob_sped_resumo.VALOR_OUTRAS = float(str(v_cc_dados[5].upper().strip()).replace(',', '.'))
                                except:
                                    ob_sped_resumo.VALOR_OUTRAS = float(0)                                
                            try:
                                ob_sped_resumo.BASE_ICMS_ST = float(str(v_cc_dados[8].upper().strip()).replace(',', '.'))
                            except:
                                ob_sped_resumo.BASE_ICMS_ST = float(0)                          
                            try:
                                ob_sped_resumo.VALOR_ICMS_ST = float(str(v_cc_dados[9].upper().strip()).replace(',', '.'))
                            except:
                                ob_sped_resumo.VALOR_ICMS_ST = float(0)
                        if ob_sped_resumo.TIPO_NF == 'SAIDA':
                            v_cc_reg_rel_saida.append(ob_sped_resumo)
                        else:    
                            v_cc_reg_rel.append(ob_sped_resumo)
                        
            # Verifica se o arquivo teve algum processamentp                
            if (not v_cc_reg_rel and not v_cc_reg_rel_saida):
                log("# " + str(len(v_cc_reg_rel)) + " >> " + " -> Não processou nenhum dados do arquivo : " + p_nm_arquivo)
                log("# " + str(len(v_cc_reg_rel_saida)) + " >> " + " -> Não processou nenhum dados do arquivo [SAIDA] : " + p_nm_arquivo)
                return None, None
            else:
                if not v_cc_reg_rel:
                    log("# " + str(len(v_cc_reg_rel)) + " >> " + " -> Não processou nenhum dados do arquivo : " + p_nm_arquivo)
                else:
                    log("# " + str(len(v_cc_reg_rel)) + " >> " + " -> Dados processados com sucesso do arquivo : " + p_nm_arquivo)
                
                if not v_cc_reg_rel_saida:
                    log("# " + str(len(v_cc_reg_rel_saida)) + " >> " + " -> Não processou nenhum dados do arquivo [SAIDA] : " + p_nm_arquivo)
                else:
                    log("# " + str(len(v_cc_reg_rel_saida)) + " >> " + " -> Dados processados com sucesso do arquivo [SAIDA] : " + p_nm_arquivo)
                                   
            return v_cc_reg_rel,v_cc_reg_rel_saida
        
        else:
            return None, None    

    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO LEITURA DO ARQUIVO .: " + str(e) + " - TRACE - " + v_ds_trace + " >> ARQUIVO: " + p_nm_arquivo + " >> TIPO: " + p_tp_arquivo + " >> LINHA: " + v_ds_linha)
        return None,None

       
def fnc_validar_entrada():
    """
        Retorna a validação de entrada e dos arquivos de configuração
    """
    try:
        v_nr_retorno = 0        
        
        comum.addParametro( 'MESANO',  None, "MESANO (MMYYYY) dos arquivos", True, '122015' )
        comum.addParametro( 'UF', None, 'Unidade Federativa (UF) dos arquivo  ', True, 'SP')
        comum.addParametro( 'IE', None, 'Inscricao Estadual (IE) dos arquivo  ', True, '108383949112')
        
        # Validacao dos parametros de entrada
        if not comum.validarParametros() :
            v_nr_retorno = 91
            return v_nr_retorno     
            
        else:
            ob_global.gv_mes_ano = comum.getParametro('MESANO').upper().strip()
            ob_global.gv_uf = comum.getParametro('UF').upper().strip()
            ob_global.gv_ie = comum.getParametro('IE').upper().strip()
        
            if (len(ob_global.gv_mes_ano) != 6):
                log("PARAMETRO MES ANO: Invalido!") 
                v_nr_retorno = 91

            if not v_nr_retorno:
                ob_global.gv_mes = ob_global.gv_mes_ano[0:2]
                ob_global.gv_ano = ob_global.gv_mes_ano[2:6]

            if not v_nr_retorno:
                try:
                    if (int(ob_global.gv_mes) < 1
                    or int(ob_global.gv_mes) > 12 
                    ):
                        log("PARAMETRO MES : Invalido!") 
                        v_nr_retorno = 91
                except:
                    log("PARAMETRO MES : Invalido!") 
                    v_nr_retorno = 91

            if not v_nr_retorno :
                try:
                    if (
                       int(ob_global.gv_ano) > datetime.datetime.now().year
                    or int(ob_global.gv_ano) < (datetime.datetime.now().year)-100
                    ):
                        log("PARAMETRO ANO : Invalido!") 
                        v_nr_retorno = 91
                except:
                    log("PARAMETRO ANO : Invalido!") 
                    v_nr_retorno = 91
            
            if not v_nr_retorno :
                if len(ob_global.gv_uf) != 2:
                    log("PARAMETRO UF: Invalido!") 
                    v_nr_retorno = 91

            if not v_nr_retorno :
                try:
                    v_nr_iei = re.sub('[^0-9]','',ob_global.gv_ie)
                    if ( (v_nr_iei == "") or (v_nr_iei == "''") or (v_nr_iei == '""') or (int("0"+v_nr_iei) == 0)):
                        log("PARAMETRO IE : Invalido!") 
                        v_nr_retorno = 91        
                except:
                    log("PARAMETRO IE : Invalido!") 
                    v_nr_retorno = 91    
        
        return v_nr_retorno            
        
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO VALIDAÇÃO DOS PARAMETROS DE ENTRADA: " + str(e)+ " >> " + v_ds_trace)
        return 93        

def fnc_validar_configuracao():
    """
        Retorna a validação dos arquivos de configuração
    """
    try:
        v_nr_retorno = 0        

        ob_global.gv_diretorio_sped_fiscal_atual = configuracoes.dir_sped_fiscal_atual.replace("<<MM>>",ob_global.gv_mes).replace("<<AAAA>>",ob_global.gv_ano).replace("<<UF>>",ob_global.gv_uf).replace("<<IE>>",ob_global.gv_ie).replace("<<MESANO>>",ob_global.gv_mes_ano).strip()
        ob_global.gv_diretorio_sped_fiscal_protocolado = configuracoes.dir_sped_fiscal_protocolado.replace("<<MM>>",ob_global.gv_mes).replace("<<AAAA>>",ob_global.gv_ano).replace("<<UF>>",ob_global.gv_uf).replace("<<IE>>",ob_global.gv_ie).replace("<<MESANO>>",ob_global.gv_mes_ano).strip()
        ob_global.gv_diretorio_insumo_sped = configuracoes.dir_insumo_sped.replace("<<MM>>",ob_global.gv_mes).replace("<<AAAA>>",ob_global.gv_ano).replace("<<UF>>",ob_global.gv_uf).replace("<<IE>>",ob_global.gv_ie).replace("<<MESANO>>",ob_global.gv_mes_ano).strip()
        
        ob_global.gv_arq_sped_fiscal_atual = configuracoes.arq_sped_fiscal_atual.replace("<<MM>>",ob_global.gv_mes).replace("<<AAAA>>",ob_global.gv_ano).replace("<<UF>>",ob_global.gv_uf).replace("<<IE>>",ob_global.gv_ie).replace("<<MESANO>>",ob_global.gv_mes_ano).replace("<<NNN>>","*").strip()
        ob_global.gv_arq_sped_fiscal_protocolado = configuracoes.arq_sped_fiscal_protocolado.replace("<<MM>>",ob_global.gv_mes).replace("<<AAAA>>",ob_global.gv_ano).replace("<<UF>>",ob_global.gv_uf).replace("<<IE>>",ob_global.gv_ie).replace("<<MESANO>>",ob_global.gv_mes_ano).replace("<<NNN>>","*").strip()
        ob_global.gv_arq_insumo_sped = configuracoes.arq_insumo_saida.replace("<<MM>>",ob_global.gv_mes).replace("<<AAAA>>",ob_global.gv_ano).replace("<<UF>>",ob_global.gv_uf).replace("<<IE>>",ob_global.gv_ie).replace("<<MESANO>>",ob_global.gv_mes_ano).replace("<<NNN>>","*").strip()

        if not os.path.isdir(ob_global.gv_diretorio_sped_fiscal_atual):
            log("Diretório não localizado sped atual : " + ob_global.gv_diretorio_sped_fiscal_atual)        
            return 1
        else:    
            ob_global.gv_lst_arq_sped_fiscal_atual = fnc_ultimo_arquivo_diretorio(ob_global.gv_arq_sped_fiscal_atual,ob_global.gv_diretorio_sped_fiscal_atual).strip()
            if not ob_global.gv_lst_arq_sped_fiscal_atual:
                log(ob_global.gv_lst_arq_sped_fiscal_atual + " Arquivo não localizado sped atual : " + ob_global.gv_diretorio_sped_fiscal_atual)        
                return 1
        
        if not os.path.isdir(ob_global.gv_diretorio_sped_fiscal_protocolado):
            log("Diretório não localizado sped protocolado : " + ob_global.gv_diretorio_sped_fiscal_protocolado)        
            return 1
        else:    
            ob_global.gv_lst_arq_sped_fiscal_protocolado = fnc_ultimo_arquivo_diretorio(ob_global.gv_arq_sped_fiscal_protocolado,ob_global.gv_diretorio_sped_fiscal_protocolado).strip()
            if not ob_global.gv_lst_arq_sped_fiscal_protocolado:
                log(ob_global.gv_lst_arq_sped_fiscal_protocolado + " Arquivo não localizado sped protocolado : " + ob_global.gv_diretorio_sped_fiscal_protocolado)        
                return 1
                
        if not os.path.isdir(ob_global.gv_diretorio_insumo_sped):
            log("Diretório não localizado insumo, porem está sendo criado : " + ob_global.gv_diretorio_insumo_sped)     
            try:    
                os.mkdir(ob_global.gv_diretorio_insumo_sped)
            except Exception as e:
                v_ds_trace = traceback.format_exc()
                log("ERRO CRIAÇÃO DO DIRETÓRIO: " + str(e)+ " >> " + v_ds_trace)
                return 93     
        
        if not ob_global.gv_arq_insumo_sped:
            log("Arquivo invalido insumo : " + ob_global.gv_arq_insumo_sped)        
            return 1        
        else:    
            ob_global.gv_lst_arq_insumo_sped = fnc_ultimo_arquivo_diretorio(ob_global.gv_arq_insumo_sped,ob_global.gv_diretorio_insumo_sped).strip()
            if len(ob_global.gv_lst_arq_insumo_sped) < 4:
                ob_global.gv_lst_arq_insumo_sped = os.path.join(ob_global.gv_diretorio_insumo_sped, ob_global.gv_arq_insumo_sped.replace("*","001"))
            else:
                if (ob_global.gv_lst_arq_insumo_sped.strip().upper().endswith(".TXT")
                or ob_global.gv_lst_arq_insumo_sped.strip().upper().find(".")):
                    ob_global.gv_lst_arq_insumo_sped = ob_global.gv_lst_arq_insumo_sped.split(".")[0][:-3]+ str(int(ob_global.gv_lst_arq_insumo_sped.split(".")[0][-3:])+1).rjust(3,'0') +  "." +ob_global.gv_lst_arq_insumo_sped.split(".")[1] 
                else:
                    ob_global.gv_lst_arq_insumo_sped = os.path.join(ob_global.gv_diretorio_insumo_sped, ob_global.gv_arq_insumo_sped.replace("*","001"))


        return v_nr_retorno
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO VALIDAÇÃO DOS PARAMETROS DE CONFIGURACAO: " + str(e)+ " >> " + v_ds_trace)
        return 93     
 
def fnc_imprimi_variaveis():
    """
        Imprimi variaveis
    """
    try:
    
        if ob_global.gv_ob_conexao:
            log("conexao  : ATIVO") 
        else:
            log("conexao  : DESATIVO")
            
        log("job      : " + str(ob_global.gv_nm_job)) 
        log("mes_ano  : " + str(ob_global.gv_mes_ano)) 
        log("uf       : " + str(ob_global.gv_uf)) 
        log("ie       : " + str(ob_global.gv_ie)) 
        log("mes      : " + str(ob_global.gv_mes)) 
        log("ano      : " + str(ob_global.gv_ano)) 
        
        log("diretorio_sped_fiscal_atual       : " + str(ob_global.gv_diretorio_sped_fiscal_atual)) 
        log("diretorio_sped_fiscal_protocolado : " + str(ob_global.gv_diretorio_sped_fiscal_protocolado))
        log("diretorio_insumo_sped             : " + str(ob_global.gv_diretorio_insumo_sped))
        
        log("arq_sped_fiscal_atual             : " + str(ob_global.gv_arq_sped_fiscal_atual))
        log("arq_sped_fiscal_protocolado       : " + str(ob_global.gv_arq_sped_fiscal_protocolado))
        log("arq_insumo_sped                   : " + str(ob_global.gv_arq_insumo_sped))
        
        log("lst_arq_sped_fiscal_atual         : " + str(ob_global.gv_lst_arq_sped_fiscal_atual))
        log("lst_arq_sped_fiscal_protocolado   : " + str(ob_global.gv_lst_arq_sped_fiscal_protocolado))
        log("lst_arq_insumo_sped               : " + str(ob_global.gv_lst_arq_insumo_sped))
        
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO IMPRESSAO DE VARIAVEIS: " + str(e)+ " >> " + v_ds_trace)
 
def fnc_validar_leitura_carregamento_arquivo():
    """
        Retorna a validação e leitura do carregamento do arquivo
    """
    v_nr_retorno = 0     
    try:
        ob_global.gv_ob_lista_sped_atual, ob_global.gv_ob_lista_sped_atual_saida = fnc_processa_sped(ob_global.gv_lst_arq_sped_fiscal_atual,"ATUAL")
        if not ob_global.gv_ob_lista_sped_atual and not ob_global.gv_ob_lista_sped_atual_saida:
            log("FALHA NA LEITURA DO ARQUIVO DE SPED ATUAL: " + ob_global.gv_lst_arq_sped_fiscal_atual)
            return 93     

        ob_global.gv_ob_lista_sped_protocolado,ob_global.gv_ob_lista_sped_protocolado_saida = fnc_processa_sped(ob_global.gv_lst_arq_sped_fiscal_protocolado,"PROTOCOLADO")
        if not ob_global.gv_ob_lista_sped_protocolado and not ob_global.gv_ob_lista_sped_protocolado_saida:
            log("FALHA NA LEITURA DO ARQUIVO DE SPED PROTOCOLADO: " + ob_global.gv_lst_arq_sped_fiscal_protocolado)
            return 93    
        
        if ob_global.gv_ob_lista_sped_atual:
            ob_global.gv_ob_lista_sped_atual.sort(key=operator.attrgetter('TIPO_NF','NUMERO','SEQ','SERIE','MODELO','CODIGO_PF_PJ','DATA_EMISSAO','PRECO_TOTAL','VALOR_CONTABIL'))
        if ob_global.gv_ob_lista_sped_atual_saida:
            ob_global.gv_ob_lista_sped_atual_saida.sort(key=operator.attrgetter('TIPO_NF','NUMERO','SEQ','SERIE','MODELO','CODIGO_PF_PJ','DATA_EMISSAO','PRECO_TOTAL','VALOR_CONTABIL'))
        
        if ob_global.gv_ob_lista_sped_protocolado:
            ob_global.gv_ob_lista_sped_protocolado.sort(key=operator.attrgetter('TIPO_NF','NUMERO','SEQ','SERIE','MODELO','CODIGO_PF_PJ','DATA_EMISSAO','PRECO_TOTAL','VALOR_CONTABIL'))
        if ob_global.gv_ob_lista_sped_protocolado_saida:
            ob_global.gv_ob_lista_sped_protocolado_saida.sort(key=operator.attrgetter('TIPO_NF','NUMERO','SEQ','SERIE','MODELO','CODIGO_PF_PJ','DATA_EMISSAO','PRECO_TOTAL','VALOR_CONTABIL'))
        
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO LEITURA DO CARREGAMENTO DO ARQUIVO: " + str(e)+ " >> " + v_ds_trace)
        v_nr_retorno = 93
    
    return v_nr_retorno         

def fnc_processar_tipo_nota(p_ob_wks_excel,p_tipo_nf,p_ob_lista_sped_protocolado,p_ob_lista_sped_atual):
    """
        Processa relatorio de comparativo
    """
    v_ob_existe_lista_sped_atual = []
    def fnc_processa_wks_comparacao(p_nr_linha,p_ds_grupo,p_ob_wks_excel_aux,p_ob_sped_prot,p_ob_sped_atual):
        p_ob_wks_excel_aux.cell(p_nr_linha,1 ,p_ob_sped_prot.TIPO_NF if p_ob_sped_prot else (p_ob_sped_atual.TIPO_NF if p_ob_sped_atual else ""))
        p_ob_wks_excel_aux.cell(p_nr_linha,2 ,int(p_ob_sped_prot.NUMERO_NF) if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,3 ,int(p_ob_sped_atual.NUMERO_NF) if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,4 ,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and int(p_ob_sped_prot.NUMERO_NF) == int(p_ob_sped_atual.NUMERO_NF) else "FALSO")
        ob_global.gv_dic_resumo["Numero NF"] += (0 if p_ob_sped_prot and p_ob_sped_atual and int(p_ob_sped_prot.NUMERO_NF) == int(p_ob_sped_atual.NUMERO_NF) else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,5 ,p_ob_sped_prot.SERIE if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,6 ,p_ob_sped_atual.SERIE if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,7 ,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.SERIE == p_ob_sped_atual.SERIE else "FALSO")
        ob_global.gv_dic_resumo["Serie"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.SERIE == p_ob_sped_atual.SERIE else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,8 ,p_ob_sped_prot.MODELO if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,9 ,p_ob_sped_atual.MODELO if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,10,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.MODELO == p_ob_sped_atual.MODELO else "FALSO")
        ob_global.gv_dic_resumo["Modelo Documento"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.MODELO == p_ob_sped_atual.MODELO else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,11,p_ob_sped_prot.DATA_EMISSAO if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,12,p_ob_sped_atual.DATA_EMISSAO if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,13,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.DATA_EMISSAO == p_ob_sped_atual.DATA_EMISSAO else "FALSO")
        ob_global.gv_dic_resumo["Data Emissão"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.DATA_EMISSAO == p_ob_sped_atual.DATA_EMISSAO else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,14,p_ob_sped_prot.DATA_ENTRADA if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,15,p_ob_sped_atual.DATA_ENTRADA if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,16,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.DATA_ENTRADA == p_ob_sped_atual.DATA_ENTRADA else "FALSO")
        ob_global.gv_dic_resumo["Data Entrada"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.DATA_ENTRADA == p_ob_sped_atual.DATA_ENTRADA else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,17,p_ob_sped_prot.CODIGO_PF_PJ if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,18,p_ob_sped_atual.CODIGO_PF_PJ if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,19,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.CODIGO_PF_PJ == p_ob_sped_atual.CODIGO_PF_PJ else "FALSO")
        ob_global.gv_dic_resumo["Codigo PF/PJ"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.CODIGO_PF_PJ == p_ob_sped_atual.CODIGO_PF_PJ else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,20,p_ob_sped_prot.CHAVE_NFE if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,21,p_ob_sped_atual.CHAVE_NFE if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,22,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.CHAVE_NFE == p_ob_sped_atual.CHAVE_NFE else "FALSO")
        ob_global.gv_dic_resumo["Chave NFe"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.CHAVE_NFE == p_ob_sped_atual.CHAVE_NFE else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,23,p_ob_sped_prot.CFOP if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,24,p_ob_sped_atual.CFOP if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,25,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.CFOP == p_ob_sped_atual.CFOP else "FALSO")
        ob_global.gv_dic_resumo["CFOP"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.CFOP == p_ob_sped_atual.CFOP else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,26,p_ob_sped_prot.CST if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,27,p_ob_sped_atual.CST if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,28,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.CST == p_ob_sped_atual.CST else "FALSO")
        ob_global.gv_dic_resumo["CST"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.CST == p_ob_sped_atual.CST else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,29,p_ob_sped_prot.PRECO_TOTAL if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,30,p_ob_sped_atual.PRECO_TOTAL if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,31,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.PRECO_TOTAL == p_ob_sped_atual.PRECO_TOTAL else "FALSO")
        ob_global.gv_dic_resumo["Preço Total"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.PRECO_TOTAL == p_ob_sped_atual.PRECO_TOTAL else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,32,p_ob_sped_prot.VALOR_CONTABIL if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,33,p_ob_sped_atual.VALOR_CONTABIL if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,34,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.VALOR_CONTABIL == p_ob_sped_atual.VALOR_CONTABIL else "FALSO")
        ob_global.gv_dic_resumo["Valor Contabil"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.VALOR_CONTABIL == p_ob_sped_atual.VALOR_CONTABIL else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,35,p_ob_sped_prot.BASE_ICMS if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,36,p_ob_sped_atual.BASE_ICMS if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,37,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.BASE_ICMS == p_ob_sped_atual.BASE_ICMS else "FALSO")
        ob_global.gv_dic_resumo["Base ICMS"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.BASE_ICMS == p_ob_sped_atual.BASE_ICMS else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,38,p_ob_sped_prot.ALIQUOTA_ICMS if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,39,p_ob_sped_atual.ALIQUOTA_ICMS if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,40,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.ALIQUOTA_ICMS == p_ob_sped_atual.ALIQUOTA_ICMS else "FALSO")
        ob_global.gv_dic_resumo["Aliquota ICMS"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.ALIQUOTA_ICMS == p_ob_sped_atual.ALIQUOTA_ICMS else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,41,p_ob_sped_prot.VALOR_ICMS if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,42,p_ob_sped_atual.VALOR_ICMS if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,43,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.VALOR_ICMS == p_ob_sped_atual.VALOR_ICMS else "FALSO")
        ob_global.gv_dic_resumo["Valor ICMS"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.VALOR_ICMS == p_ob_sped_atual.VALOR_ICMS else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,44,p_ob_sped_prot.VALOR_IPI if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,45,p_ob_sped_atual.VALOR_IPI if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,46,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.VALOR_IPI == p_ob_sped_atual.VALOR_IPI else "FALSO")
        ob_global.gv_dic_resumo["Valor IPI"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.VALOR_IPI == p_ob_sped_atual.VALOR_IPI else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,47,p_ob_sped_prot.VALOR_DESCONTO if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,48,p_ob_sped_atual.VALOR_DESCONTO if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,49,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.VALOR_DESCONTO == p_ob_sped_atual.VALOR_DESCONTO else "FALSO")
        ob_global.gv_dic_resumo["Valor Desconto"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.VALOR_DESCONTO == p_ob_sped_atual.VALOR_DESCONTO else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,50,p_ob_sped_prot.VALOR_ISENTAS if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,51,p_ob_sped_atual.VALOR_ISENTAS if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,52,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.VALOR_ISENTAS == p_ob_sped_atual.VALOR_ISENTAS else "FALSO")
        ob_global.gv_dic_resumo["Valor Isentas"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.VALOR_ISENTAS == p_ob_sped_atual.VALOR_ISENTAS else  1)      
        p_ob_wks_excel_aux.cell(p_nr_linha,53,p_ob_sped_prot.VALOR_OUTRAS if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,54,p_ob_sped_atual.VALOR_OUTRAS if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,55,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.VALOR_OUTRAS == p_ob_sped_atual.VALOR_OUTRAS else "FALSO")
        ob_global.gv_dic_resumo["Valor Outras"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.VALOR_OUTRAS == p_ob_sped_atual.VALOR_OUTRAS else  1)             
        p_ob_wks_excel_aux.cell(p_nr_linha,56,p_ob_sped_prot.BASE_ICMS_ST if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,57,p_ob_sped_atual.BASE_ICMS_ST if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,58,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.BASE_ICMS_ST == p_ob_sped_atual.BASE_ICMS_ST else "FALSO")
        ob_global.gv_dic_resumo["Base ICMS ST"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.BASE_ICMS_ST == p_ob_sped_atual.BASE_ICMS_ST else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,59,p_ob_sped_prot.ALIQUOTA_ICMS_ST if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,60,p_ob_sped_atual.ALIQUOTA_ICMS_ST if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,61,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.ALIQUOTA_ICMS_ST == p_ob_sped_atual.ALIQUOTA_ICMS_ST else "FALSO")
        ob_global.gv_dic_resumo["Aliquota ICMS ST"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.ALIQUOTA_ICMS_ST == p_ob_sped_atual.ALIQUOTA_ICMS_ST else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,62,p_ob_sped_prot.VALOR_ICMS_ST if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,63,p_ob_sped_atual.VALOR_ICMS_ST if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,64,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.VALOR_ICMS_ST == p_ob_sped_atual.VALOR_ICMS_ST else "FALSO")
        ob_global.gv_dic_resumo["Valor ICMS ST"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.VALOR_ICMS_ST == p_ob_sped_atual.VALOR_ICMS_ST else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,65,p_ob_sped_prot.CODIGO_MATERIAL if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,66,p_ob_sped_atual.CODIGO_MATERIAL if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,67,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.CODIGO_MATERIAL == p_ob_sped_atual.CODIGO_MATERIAL else "FALSO")
        ob_global.gv_dic_resumo["Codigo Material"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.CODIGO_MATERIAL == p_ob_sped_atual.CODIGO_MATERIAL else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,68,p_ob_sped_prot.NATUREZA_OPERACAO if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,69,p_ob_sped_atual.NATUREZA_OPERACAO if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,70,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.NATUREZA_OPERACAO == p_ob_sped_atual.NATUREZA_OPERACAO else "FALSO")
        ob_global.gv_dic_resumo["Natureza Operação"] += (0 if p_ob_sped_prot and p_ob_sped_atual and p_ob_sped_prot.NATUREZA_OPERACAO == p_ob_sped_atual.NATUREZA_OPERACAO else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,71,int(p_ob_sped_prot.SEQ_ITEM) if p_ob_sped_prot else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,72,int(p_ob_sped_atual.SEQ_ITEM) if p_ob_sped_atual else "")
        p_ob_wks_excel_aux.cell(p_nr_linha,73,"VERDADEIRO" if p_ob_sped_prot and p_ob_sped_atual and int(p_ob_sped_prot.SEQ_ITEM) == int(p_ob_sped_atual.SEQ_ITEM) else "FALSO")
        ob_global.gv_dic_resumo["Numero Sequencial"] += (0 if p_ob_sped_prot and p_ob_sped_atual and int(p_ob_sped_prot.SEQ_ITEM) == int(p_ob_sped_atual.SEQ_ITEM) else  1)
        p_ob_wks_excel_aux.cell(p_nr_linha,74,p_ds_grupo)     
        p_ob_wks_excel_aux.cell(p_nr_linha,75,p_ob_sped_prot.NR_SEQUENCIAL_CONTROLE if p_ob_sped_prot else (p_ob_sped_atual.NR_SEQUENCIAL_CONTROLE if p_ob_sped_atual else ""))    
        for idx,nr_col in enumerate(range(29,64)):
            if ((idx+1) % 3) != 0:
                p_ob_wks_excel_aux.cell(p_nr_linha,nr_col).number_format = "#,##0.00"
    log("Criando cabeçalho : " + p_ob_wks_excel.title)
    p_ob_wks_excel.cell(1,1 ,"Tipo NF")
    p_ob_wks_excel.cell(1,2 ,"Numero NF  - Protocolado \n(C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,3 ,"Numero NF  - Base Atual \n(C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,4 ,"Numero NF  - Comparativo")
    p_ob_wks_excel.cell(1,5 ,"Serie - Protocolado \n(C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,6 ,"Serie - Base Atual \n(C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,7 ,"Serie - Comparativo")
    p_ob_wks_excel.cell(1,8 ,"Modelo Documento - Protocolado \n(C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,9 ,"Modelo Documento - Base Atual \n(C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,10,"Modelo Documento - Comparativo")
    p_ob_wks_excel.cell(1,11,"Data Emissão - Protocolado \n(C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,12,"Data Emissão - Base Atual \n(C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,13,"Data Emissão - Comparativo")
    p_ob_wks_excel.cell(1,14,"Data Entrada - Protocolado \n(C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,15,"Data Entrada - Base Atual \n(C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,16,"Data Entrada - Comparativo")
    p_ob_wks_excel.cell(1,17,"Codigo PF/PJ - Protocolado \n(C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,18,"Codigo PF/PJ - Base Atual \n(C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,19,"Codigo PF/PJ - Comparativo")
    p_ob_wks_excel.cell(1,20,"Chave NFe - Protocolado \n(C100 / D100)")
    p_ob_wks_excel.cell(1,21,"Chave NFe - Base Atual \n(C100 / D100)")
    p_ob_wks_excel.cell(1,22,"Chave NFe - Comparativo")
    p_ob_wks_excel.cell(1,23,"CFOP - Protocolado \n(C170 / C190 / C590 / D190 / D590)")
    p_ob_wks_excel.cell(1,24,"CFOP - Base Atual \n(C170 / C190 / C590 / D190 / D590)")
    p_ob_wks_excel.cell(1,25,"CFOP - Comparativo")
    p_ob_wks_excel.cell(1,26,"CST - Protocolado \n(C170 / C190 / C590 / D190 / D590)")
    p_ob_wks_excel.cell(1,27,"CST - Base Atual \n(C170 / C190 / C590 / D190 / D590)")
    p_ob_wks_excel.cell(1,28,"CST - Comparativo")
    p_ob_wks_excel.cell(1,29,"Preço Total - Protocolado \n(C170 / C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,30,"Preço Total - Base Atual \n(C170 / C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,31,"Preço Total - Comparativo")
    p_ob_wks_excel.cell(1,32,"Valor Contabil - Protocolado \n(C190 / C590 / D190 / D590)")
    p_ob_wks_excel.cell(1,33,"Valor Contabil - Base Atual \n(C190 / C590 / D190 / D590)")
    p_ob_wks_excel.cell(1,34,"Valor Contabil - Comparativo")
    p_ob_wks_excel.cell(1,35,"Base ICMS - Protocolado \n(C170 / C190 / C590 / D190 / D590)")
    p_ob_wks_excel.cell(1,36,"Base ICMS - Base Atual \n(C170 / C190 / C590 / D190 / D590)")
    p_ob_wks_excel.cell(1,37,"Base ICMS - Comparativo")
    p_ob_wks_excel.cell(1,38,"Aliquota ICMS - Protocolado \n(C170 / C190 / C590 / D190 / D590)")
    p_ob_wks_excel.cell(1,39,"Aliquota ICMS - Base Atual \n(C170 / C190 / C590 / D190 / D590)")
    p_ob_wks_excel.cell(1,40,"Aliquota ICMS - Comparativo")
    p_ob_wks_excel.cell(1,41,"Valor ICMS - Protocolado \n(C170 / C190 / C590 / D190 / D590)")
    p_ob_wks_excel.cell(1,42,"Valor ICMS - Base Atual \n(C170 / C190 / C590 / D190 / D590)")
    p_ob_wks_excel.cell(1,43,"Valor ICMS - Comparativo")
    p_ob_wks_excel.cell(1,44,"Valor IPI - Protocolado \n(C170 / C190)")
    p_ob_wks_excel.cell(1,45,"Valor IPI - Base Atual \n(C170 / C190)")
    p_ob_wks_excel.cell(1,46,"Valor IPI - Comparativo")
    p_ob_wks_excel.cell(1,47,"Valor Desconto - Protocolado \n(C170 / C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,48,"Valor Desconto - Base Atual \n(C170 / C100 / C500 / D100 / D500)")
    p_ob_wks_excel.cell(1,49,"Valor Desconto - Comparativo") 
    p_ob_wks_excel.cell(1,50,"Valor Isentas - Protocolado \n(C170)")
    p_ob_wks_excel.cell(1,51,"Valor Isentas - Base Atual \n(C170)")
    p_ob_wks_excel.cell(1,52,"Valor Isentas - Comparativo") 
    p_ob_wks_excel.cell(1,53,"Valor Outras - Protocolado \n(C170)")
    p_ob_wks_excel.cell(1,54,"Valor Outras - Base Atual \n(C170)")
    p_ob_wks_excel.cell(1,55,"Valor Outras - Comparativo")
    p_ob_wks_excel.cell(1,56,"Base ICMS ST - Protocolado \n(C170 / C190)")
    p_ob_wks_excel.cell(1,57,"Base ICMS ST - Base Atual \n((C170 / C190)")
    p_ob_wks_excel.cell(1,58,"Base ICMS ST - Comparativo")
    p_ob_wks_excel.cell(1,59,"Aliquota ICMS ST - Protocolado \n(C170)")
    p_ob_wks_excel.cell(1,60,"Aliquota ICMS ST - Base Atual \n(C170)")
    p_ob_wks_excel.cell(1,61,"Aliquota ICMS ST - Comparativo")
    p_ob_wks_excel.cell(1,62,"Valor ICMS ST - Protocolado \n(C170 / C190)")
    p_ob_wks_excel.cell(1,63,"Valor ICMS ST - Base Atual \n(C170 / C190)")
    p_ob_wks_excel.cell(1,64,"Valor ICMS ST - Comparativo")
    p_ob_wks_excel.cell(1,65,"Codigo Material - Protocolado \n(C170)")
    p_ob_wks_excel.cell(1,66,"Codigo Material - Base Atual \n(C170)")
    p_ob_wks_excel.cell(1,67,"Codigo Material - Comparativo")
    p_ob_wks_excel.cell(1,68,"Natureza Operação - Protocolado \n(C170)")
    p_ob_wks_excel.cell(1,69,"Natureza Operação - Base Atual \n(C170)")
    p_ob_wks_excel.cell(1,70,"Natureza Operação - Comparativo")
    p_ob_wks_excel.cell(1,71,"Numero Sequencial - Protocolado \n(C170)")
    p_ob_wks_excel.cell(1,72,"Numero Sequencial - Base Atual \n(C170)")
    p_ob_wks_excel.cell(1,73,"Numero Sequencial - Comparativo")   
    p_ob_wks_excel.cell(1,74,"Grupo NF")
    p_ob_wks_excel.cell(1,75,"Nro. Sequencial Controle")
    
    log("Criando cabeçalho... : " + p_ob_wks_excel.title)
    for nr_col in range(1,76):
        p_ob_wks_excel.cell(1,nr_col).font=Font(bold=True)
        p_ob_wks_excel.cell(1,nr_col).alignment = Alignment(horizontal='center')
        if (nr_col > 1 and nr_col < 74) and ((nr_col-1) % 3 == 0):
            p_ob_wks_excel.cell(1,nr_col).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")        
    
    log("Processando... : " + p_ob_wks_excel.title)
    v_nr_linha = 1
    for reg in p_ob_lista_sped_protocolado:    
        if reg.TIPO_NF != p_tipo_nf:
            continue                
        ob_sped    = None
        v_fl_achou = 0 
        v_ds_grupo = "NAO LOCALIZADO DE: PROTOCOLADO -> PARA: ATUAL"
        for rega in p_ob_lista_sped_atual:
            if reg.TIPO_NF != rega.TIPO_NF:
                continue
            if reg.TIPO_NF == rega.TIPO_NF \
            and rega.NUMERO > reg.NUMERO:
                break
            elif reg.TIPO_NF == rega.TIPO_NF \
            and reg.NUMERO == rega.NUMERO \
            and reg.SEQ == rega.SEQ \
            and rega not in v_ob_existe_lista_sped_atual:
                v_fl_achou_aux = v_fl_achou
                if v_fl_achou < 1:
                    v_fl_achou = 1
                    v_ds_grupo = "1-Numero NF"
                    ob_sped = rega
                if reg.SERIE == rega.SERIE:
                    if v_fl_achou < 200:
                        v_fl_achou = 200
                        v_ds_grupo = "200-Numero NF + Serie"
                        ob_sped = rega
                    if reg.MODELO == rega.MODELO:    
                        if v_fl_achou < 300:
                            v_fl_achou = 300
                            v_ds_grupo = "300-Numero NF + Serie + Modelo Doc"
                            ob_sped = rega
                        if reg.CODIGO_PF_PJ == rega.CODIGO_PF_PJ:    
                            if v_fl_achou < 400:
                                v_fl_achou = 400
                                v_ds_grupo = "400-Numero NF + Serie + Modelo Doc + Codigo PF/PJ"
                                ob_sped = rega
                            if reg.DATA_EMISSAO == rega.DATA_EMISSAO:    
                                if v_fl_achou < 500:
                                    v_fl_achou = 500
                                    v_ds_grupo = "500-Numero NF + Serie + Modelo Doc + Codigo PF/PJ + Data Emissao"
                                    ob_sped = rega    
                                if reg.PRECO_TOTAL == rega.PRECO_TOTAL:    
                                    if v_fl_achou < 600:
                                        v_fl_achou = 600
                                        v_ds_grupo = "600-Numero NF + Serie + Modelo Doc + Codigo PF/PJ + Data Emissao + Preço Total" 
                                        ob_sped = rega    
                                    if reg.VALOR_CONTABIL == rega.VALOR_CONTABIL:    
                                        if v_fl_achou < 700:
                                            v_fl_achou = 700
                                            v_ds_grupo = "700-Numero NF + Serie + Modelo Doc + Codigo PF/PJ + Data Emissao + Preço Total + Valor Contabil"
                                            ob_sped = rega
                                            break    
                else:
                    if reg.MODELO == rega.MODELO:    
                        if v_fl_achou < 30:
                            v_fl_achou = 30
                            v_ds_grupo = "30-Numero NF + Modelo Doc"
                            ob_sped = rega    
                        if reg.CODIGO_PF_PJ == rega.CODIGO_PF_PJ:    
                            if v_fl_achou < 40:
                                v_fl_achou = 40
                                v_ds_grupo = "40-Numero NF + Modelo Doc + Codigo PF/PJ" 
                                ob_sped = rega
                            if reg.DATA_EMISSAO == rega.DATA_EMISSAO:    
                                if v_fl_achou < 50:
                                    v_fl_achou = 50
                                    v_ds_grupo = "50-Numero NF + Modelo Doc + Codigo PF/PJ + Data Emissao" 
                                    ob_sped = rega
                                if reg.PRECO_TOTAL == rega.PRECO_TOTAL:    
                                    if v_fl_achou < 60:
                                        v_fl_achou = 60
                                        v_ds_grupo = "60-Numero NF + Modelo Doc + Codigo PF/PJ + Data Emissao + Preço Total" 
                                        ob_sped = rega
                                    if reg.VALOR_CONTABIL == rega.VALOR_CONTABIL:    
                                        if v_fl_achou < 70:
                                            v_fl_achou = 70
                                            v_ds_grupo = "70-Numero NF + Modelo Doc + Codigo PF/PJ + Data Emissao + Preço Total + Valor Contabil"    
                                            ob_sped = rega
                    else:
                        if reg.CODIGO_PF_PJ == rega.CODIGO_PF_PJ:    
                            if v_fl_achou < 20:
                                v_fl_achou = 20
                                v_ds_grupo = "20-Numero NF + Codigo PF/PJ"
                                ob_sped = rega    
                            if reg.DATA_EMISSAO == rega.DATA_EMISSAO:    
                                if v_fl_achou < 21:
                                    v_fl_achou = 21
                                    v_ds_grupo = "21-Numero NF + Codigo PF/PJ + Data Emissao"
                                    ob_sped = rega    
                                if reg.PRECO_TOTAL == rega.PRECO_TOTAL:    
                                    if v_fl_achou < 22:
                                        v_fl_achou = 22
                                        v_ds_grupo = "22-Numero NF + Codigo PF/PJ + Data Emissao + Preço Total"
                                        ob_sped = rega    
                                    if reg.VALOR_CONTABIL == rega.VALOR_CONTABIL:    
                                        if v_fl_achou < 23:
                                            v_fl_achou = 23
                                            v_ds_grupo = "23-Numero NF + Codigo PF/PJ + Data Emissao + Preço Total + Valor Contabil"    
                                            ob_sped = rega
                        else:
                            if reg.DATA_EMISSAO == rega.DATA_EMISSAO:    
                                if v_fl_achou < 15:
                                    v_fl_achou = 15
                                    v_ds_grupo = "15-Numero NF + Data Emissao"
                                    ob_sped = rega    
                                if reg.PRECO_TOTAL == rega.PRECO_TOTAL:    
                                    if v_fl_achou < 16:
                                        v_fl_achou = 16
                                        v_ds_grupo = "16-Numero NF + Data Emissao + Preço Total"
                                        ob_sped = rega    
                                    if reg.VALOR_CONTABIL == rega.VALOR_CONTABIL:    
                                        if v_fl_achou < 17:
                                            v_fl_achou = 17
                                            v_ds_grupo = "17-Numero NF + Data Emissao + Preço Total + Valor Contabil"   
                                            ob_sped = rega    
                            else:
                                if reg.PRECO_TOTAL == rega.PRECO_TOTAL:    
                                    if v_fl_achou < 11:
                                        v_fl_achou = 11
                                        v_ds_grupo = "11-Numero NF + Preço Total"    
                                        ob_sped = rega
                                    if reg.VALOR_CONTABIL == rega.VALOR_CONTABIL:    
                                        if v_fl_achou < 12:
                                            v_fl_achou = 12
                                            v_ds_grupo = "12-Numero NF + Preço Total + Valor Contabil"                                                
                                            ob_sped = rega
        if ob_sped and ob_sped not in v_ob_existe_lista_sped_atual:
            v_ob_existe_lista_sped_atual.append(ob_sped)
        v_nr_linha += 1
        fnc_processa_wks_comparacao(p_nr_linha=v_nr_linha, \
                                    p_ds_grupo=v_ds_grupo, \
                                    p_ob_wks_excel_aux=p_ob_wks_excel, \
                                    p_ob_sped_prot=reg, \
                                    p_ob_sped_atual=ob_sped)
    
    log("Processando dados nao existente... : " + p_ob_wks_excel.title)
    v_ob_lista_aux = []
    if p_ob_lista_sped_atual:
        v_ob_lista_aux = list(set(p_ob_lista_sped_atual).difference(set(v_ob_existe_lista_sped_atual))) 
    ob_sped    = None             
    for rega in v_ob_lista_aux:
        if rega.TIPO_NF != p_tipo_nf:
            continue                
        v_ds_grupo = "NAO LOCALIZADO DE: ATUAL -> PARA: PROTOCOLADO"
        v_nr_linha += 1
        fnc_processa_wks_comparacao(p_nr_linha=v_nr_linha, \
                                    p_ds_grupo=v_ds_grupo, \
                                    p_ob_wks_excel_aux=p_ob_wks_excel, \
                                    p_ob_sped_prot=ob_sped, \
                                    p_ob_sped_atual=rega)


    try:
        p_ob_wks_excel.auto_filter.ref = "A1:" + get_column_letter(p_ob_wks_excel.max_column) \
                            + str(p_ob_wks_excel.max_row)        
    except:
        pass
    
    try:
        log("Redimensiona a planilha Excel : " + p_ob_wks_excel.title)
        fnc_column_dimensions(p_ob_wks_excel) 
    except:
        pass
             
    try:
        log("freeze_panes a planilha Excel : " + p_ob_wks_excel.title)
        p_ob_wks_excel.freeze_panes = 'A2'
    except:
        pass

    try:    
        for col in ['AX', 'AY', 'AZ','BA','BB','BC','BW']:
            p_ob_wks_excel.column_dimensions[col].hidden= True
    except:
        pass

def fnc_processar_wks(p_ob_wks_excel,p_ob_lista_sped):
    """
        Processa informacoes de detalhamento do relatorio
    """
    p_ob_wks_excel.cell(1,1 ,'TIPO_NF') 
    p_ob_wks_excel.cell(1,2 ,'NUMERO_NF')
    p_ob_wks_excel.cell(1,3 ,'SERIE')
    p_ob_wks_excel.cell(1,4 ,'MODELO')
    p_ob_wks_excel.cell(1,5 ,'DATA_EMISSAO')
    p_ob_wks_excel.cell(1,6 ,'DATA_ENTRADA')
    p_ob_wks_excel.cell(1,7 ,'CODIGO_PF_PJ')
    p_ob_wks_excel.cell(1,8 ,'CHAVE_NFE')
    p_ob_wks_excel.cell(1,9 ,'CFOP')
    p_ob_wks_excel.cell(1,10,'NATUREZA_OPERACAO')
    p_ob_wks_excel.cell(1,11,'CST')
    p_ob_wks_excel.cell(1,12,'PRECO_TOTAL')
    p_ob_wks_excel.cell(1,13,'VALOR_CONTABIL')
    p_ob_wks_excel.cell(1,14,'BASE_ICMS')
    p_ob_wks_excel.cell(1,15,'ALIQUOTA_ICMS')
    p_ob_wks_excel.cell(1,16,'VALOR_ICMS')
    p_ob_wks_excel.cell(1,17,'VALOR_IPI')
    p_ob_wks_excel.cell(1,18,'VALOR_DESCONTO')
    p_ob_wks_excel.cell(1,19,'VALOR_ISENTAS')
    p_ob_wks_excel.cell(1,20,'VALOR_OUTRAS')
    p_ob_wks_excel.cell(1,21,'BASE_ICMS_ST')
    p_ob_wks_excel.cell(1,22,'ALIQUOTA_ICMS_ST')
    p_ob_wks_excel.cell(1,23,'VALOR_ICMS_ST')
    p_ob_wks_excel.cell(1,24,'CODIGO_MATERIAL')
    p_ob_wks_excel.cell(1,25,'SEQ_ITEM')
    p_ob_wks_excel.cell(1,26,'TIPO_REG')
    p_ob_wks_excel.cell(1,27,'GRUPO_REG_SPED')
    p_ob_wks_excel.cell(1,28,'TIPO_SPED')
    p_ob_wks_excel.cell(1,29,'MES_ANO_SPED')
    p_ob_wks_excel.cell(1,30,'UF_SPED')
    p_ob_wks_excel.cell(1,31,'IE_SPED')
    p_ob_wks_excel.cell(1,32,'VALIDACAO')   
    p_ob_wks_excel.cell(1,33,'NR_SEQUENCIAL_CONTROLE')                 
    for nr_col in range(1,34):
        p_ob_wks_excel.cell(1,nr_col).font=Font(bold=True)
        p_ob_wks_excel.cell(1,nr_col).alignment = Alignment(horizontal='center')
        p_ob_wks_excel.cell(1,nr_col).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")                    
    for idx, reg in enumerate(p_ob_lista_sped):    
        p_ob_wks_excel.cell(idx+2,1 ,reg.TIPO_NF)
        p_ob_wks_excel.cell(idx+2,2 ,reg.NUMERO_NF)
        p_ob_wks_excel.cell(idx+2,3 ,reg.SERIE)
        p_ob_wks_excel.cell(idx+2,4 ,reg.MODELO)
        p_ob_wks_excel.cell(idx+2,5 ,reg.DATA_EMISSAO)
        p_ob_wks_excel.cell(idx+2,6 ,reg.DATA_ENTRADA)
        p_ob_wks_excel.cell(idx+2,7 ,reg.CODIGO_PF_PJ)
        p_ob_wks_excel.cell(idx+2,8 ,reg.CHAVE_NFE)
        p_ob_wks_excel.cell(idx+2,9 ,reg.CFOP)
        p_ob_wks_excel.cell(idx+2,10,reg.NATUREZA_OPERACAO)
        p_ob_wks_excel.cell(idx+2,11,reg.CST)
        p_ob_wks_excel.cell(idx+2,12,reg.PRECO_TOTAL)
        p_ob_wks_excel.cell(idx+2,13,reg.VALOR_CONTABIL)
        p_ob_wks_excel.cell(idx+2,14,reg.BASE_ICMS)
        p_ob_wks_excel.cell(idx+2,15,reg.ALIQUOTA_ICMS)
        p_ob_wks_excel.cell(idx+2,16,reg.VALOR_ICMS)
        p_ob_wks_excel.cell(idx+2,17,reg.VALOR_IPI)
        p_ob_wks_excel.cell(idx+2,18,reg.VALOR_DESCONTO)
        p_ob_wks_excel.cell(idx+2,19,reg.VALOR_ISENTAS)
        p_ob_wks_excel.cell(idx+2,20,reg.VALOR_OUTRAS)
        p_ob_wks_excel.cell(idx+2,21,reg.BASE_ICMS_ST)
        p_ob_wks_excel.cell(idx+2,22,reg.ALIQUOTA_ICMS_ST)
        p_ob_wks_excel.cell(idx+2,23,reg.VALOR_ICMS_ST)
        p_ob_wks_excel.cell(idx+2,24,reg.CODIGO_MATERIAL)
        p_ob_wks_excel.cell(idx+2,25,reg.SEQ)
        p_ob_wks_excel.cell(idx+2,26,reg.TIPO_REG)
        p_ob_wks_excel.cell(idx+2,27,reg.GRUPO_REG_SPED)
        p_ob_wks_excel.cell(idx+2,28,reg.TIPO_SPED)
        p_ob_wks_excel.cell(idx+2,29,reg.MES_ANO_SPED)
        p_ob_wks_excel.cell(idx+2,30,reg.UF_SPED)
        p_ob_wks_excel.cell(idx+2,31,reg.IE_SPED)
        p_ob_wks_excel.cell(idx+2,32,reg.VALIDACAO)    
        p_ob_wks_excel.cell(idx+2,33,reg.NR_SEQUENCIAL_CONTROLE)             
        for idx,nr_col in enumerate(range(12,24)):
            p_ob_wks_excel.cell(idx+2,nr_col).number_format = "#,##0.00"             
    try:
        p_ob_wks_excel.auto_filter.ref = "A1:" + get_column_letter(p_ob_wks_excel.max_column) \
                            + str(p_ob_wks_excel.max_row)        
    except:
        pass
    
    try:
        log("Redimensiona a planilha Excel : " + p_ob_wks_excel.title)
        fnc_column_dimensions(p_ob_wks_excel) 
    except:
        pass
                 
    try:
        log("freeze_panes a planilha Excel : " + p_ob_wks_excel.title)
        p_ob_wks_excel.freeze_panes = 'A2'
    except:
        pass
    
    try:    
        # OCULTA AS COLUNAS, OUTRAS, ISENTAS, SEQ    
        for col in ['S', 'T','AG']:
            p_ob_wks_excel.column_dimensions[col].hidden= True
    except:
        pass
    
def fnc_processar():
    """
        Funcao principal para processar as informacoes
    """
    try:
        v_nr_retorno = 0
        log("Criando arquivo : " + ob_global.gv_lst_arq_insumo_sped)
        #### Cria a planilha em memória....
        v_obj_arq_excel = Workbook()
        v_ob_wks_excel_nr = v_obj_arq_excel.active
        v_ob_wks_excel_nr.title = "RESUMO_FALSO"
        log("Criando wks : " + v_ob_wks_excel_nr.title)
        
        v_ob_wks_excel_nr_0 = v_obj_arq_excel.create_sheet("ATUAL x PROTOCOLADO - ENTRADA", 1)
        log("Criando wks : " + v_ob_wks_excel_nr_0.title)
        fnc_processar_tipo_nota(p_ob_wks_excel=v_ob_wks_excel_nr_0, \
                                p_tipo_nf="ENTRADA", \
                                p_ob_lista_sped_protocolado=ob_global.gv_ob_lista_sped_protocolado, \
                                p_ob_lista_sped_atual=ob_global.gv_ob_lista_sped_atual)
            
        v_ob_wks_excel_nr_0_1 = v_obj_arq_excel.create_sheet("ATUAL x PROTOCOLADO - SAIDA", 2)
        log("Criando wks : " + v_ob_wks_excel_nr_0_1.title)
        fnc_processar_tipo_nota(p_ob_wks_excel=v_ob_wks_excel_nr_0_1, \
                                p_tipo_nf="SAIDA", \
                                p_ob_lista_sped_protocolado=ob_global.gv_ob_lista_sped_protocolado_saida, \
                                p_ob_lista_sped_atual=ob_global.gv_ob_lista_sped_atual_saida)
        
        v_ob_wks_excel_nr_2 = v_obj_arq_excel.create_sheet("ATUAL_ENTRADA", 3)
        log("Planilha : " + v_ob_wks_excel_nr_2.title)
        fnc_processar_wks(p_ob_wks_excel=v_ob_wks_excel_nr_2,p_ob_lista_sped=ob_global.gv_ob_lista_sped_atual)
 
        v_ob_wks_excel_nr_2_0 = v_obj_arq_excel.create_sheet("ATUAL_SAIDA", 4)
        log("Planilha : " + v_ob_wks_excel_nr_2_0.title)
        fnc_processar_wks(p_ob_wks_excel=v_ob_wks_excel_nr_2_0,p_ob_lista_sped=ob_global.gv_ob_lista_sped_atual_saida)
 
        v_ob_wks_excel_nr_3 = v_obj_arq_excel.create_sheet("PROTOCOLADO_ENTRADA", 5)  
        log("Planilha : " + v_ob_wks_excel_nr_3.title)    
        fnc_processar_wks(p_ob_wks_excel=v_ob_wks_excel_nr_3,p_ob_lista_sped=ob_global.gv_ob_lista_sped_protocolado)
        
        v_ob_wks_excel_nr_3_0 = v_obj_arq_excel.create_sheet("PROTOCOLADO_SAIDA", 6)  
        log("Planilha : " + v_ob_wks_excel_nr_3_0.title)    
        fnc_processar_wks(p_ob_wks_excel=v_ob_wks_excel_nr_3_0,p_ob_lista_sped=ob_global.gv_ob_lista_sped_protocolado_saida)

        v_nr_linha = 0        
        log("Planilha : " + v_ob_wks_excel_nr.title)        
        v_nr_linha += 1
        v_ob_wks_excel_nr.cell(v_nr_linha,1 ,"TIPO")
        v_ob_wks_excel_nr.cell(v_nr_linha,2 ,"VALOR") 
        v_ob_wks_excel_nr.cell(v_nr_linha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid") 
        v_ob_wks_excel_nr.cell(v_nr_linha,2).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid") 
        for i,v in ob_global.gv_dic_resumo.items(): 
            v_nr_linha += 1
            v_ob_wks_excel_nr.cell(v_nr_linha,1 ,str(i))
            v_ob_wks_excel_nr.cell(v_nr_linha,2 ,str(v))
            v_ob_wks_excel_nr.cell(v_nr_linha,1).font=Font(bold=True)
            #v_ob_wks_excel_nr.cell(v_nr_linha,1).alignment = Alignment(horizontal='center')
            v_ob_wks_excel_nr.cell(v_nr_linha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")               
        try:
            log("Redimensiona a planilha Excel : " + p_ob_wks_excel.title)
            fnc_column_dimensions(v_ob_wks_excel_nr) 
        except:
            pass                
        try:
            log("freeze_panes a planilha Excel : " + v_ob_wks_excel_nr.title)
            v_ob_wks_excel_nr.freeze_panes = 'A2'
        except:
            pass
        try:
            log("auto_filter a planilha Excel : " + v_ob_wks_excel_nr.title)
            v_ob_wks_excel_nr.auto_filter.ref = "A1:" + get_column_letter(v_ob_wks_excel_nr.max_column) \
                                + str(v_ob_wks_excel_nr.max_row)        
        except:
            pass

        # Grava a planilha Excel
        log("Grava a planilha Excel : " + ob_global.gv_lst_arq_insumo_sped)
        v_obj_arq_excel.save(ob_global.gv_lst_arq_insumo_sped)
        log(str(v_nr_retorno) + " Salvo a planilha Excel : " + ob_global.gv_lst_arq_insumo_sped)
        
        return v_nr_retorno
        
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO PROCESSAMENTO: " + str(e)+ " >> " + v_ds_trace)
        v_nr_retorno = 93
        return v_nr_retorno        

if __name__ == "__main__" :
    """
        Ponto de partida
    """    
    # Codigo de Retorno
    v_nr_ret = 0

    # Tratamento de excessao
    v_ds_trace = ''

    # Tratamento de variaveis globais
    ob_global.gv_ob_conexao = None
    # Parametros do arquivo de configuração
    ob_global.gv_nm_job = "" 
    ob_global.gv_dic_resumo = {}
    ob_global.gv_dic_resumo["Numero NF"] = 0
    ob_global.gv_dic_resumo["Serie"] = 0
    ob_global.gv_dic_resumo["Modelo Documento"] = 0
    ob_global.gv_dic_resumo["Data Emissão"] = 0
    ob_global.gv_dic_resumo["Data Entrada"] = 0
    ob_global.gv_dic_resumo["Codigo PF/PJ"] = 0
    ob_global.gv_dic_resumo["Chave NFe"] = 0
    ob_global.gv_dic_resumo["CFOP"] = 0
    ob_global.gv_dic_resumo["CST"] = 0
    ob_global.gv_dic_resumo["Preço Total"] = 0
    ob_global.gv_dic_resumo["Valor Contabil"] = 0
    ob_global.gv_dic_resumo["Base ICMS"] = 0
    ob_global.gv_dic_resumo["Aliquota ICMS"] = 0
    ob_global.gv_dic_resumo["Valor ICMS"] = 0
    ob_global.gv_dic_resumo["Valor IPI"] = 0
    ob_global.gv_dic_resumo["Valor Desconto"] = 0                
    ob_global.gv_dic_resumo["Valor Isentas"] = 0      
    ob_global.gv_dic_resumo["Valor Outras"] = 0 
    ob_global.gv_dic_resumo["Base ICMS ST"] = 0
    ob_global.gv_dic_resumo["Aliquota ICMS ST"] = 0
    ob_global.gv_dic_resumo["Valor ICMS ST"] = 0
    ob_global.gv_dic_resumo["Codigo Material"] = 0 
    ob_global.gv_dic_resumo["Natureza Operação"] = 0 
    ob_global.gv_dic_resumo["Numero Sequencial"] = 0

    # Lista de String
    ob_global.gv_ob_lista_string = list(string.ascii_lowercase)

    try:

        log("-"*100)
        log("INICIO DA EXECUÇÃO".center(120,'#'))
       
        # Validacao dos parametros de entrada
        if not v_nr_ret:
            v_nr_ret = fnc_validar_entrada()
        
        # Verificar conexao com o banco
        if not v_nr_ret:
            v_nr_ret = fnc_conectar_banco_dados()   

        # Validacao dos parametros de configuracoes
        if not v_nr_ret:
            v_nr_ret = fnc_validar_configuracao()

        # Validacao da leitura e carregamento do arquivo
        if not v_nr_ret:
            fnc_imprimi_variaveis()
            v_nr_ret = fnc_validar_leitura_carregamento_arquivo()
            
        # Processar         
        if not v_nr_ret:            
            v_nr_ret = fnc_processar()                    
        
        log("-"*100)
        # Finalizacao
        if not v_nr_ret:
            log("SUCESSO NA EXECUÇÃO")
        else:
            log("ERRO NA EXECUÇÃO")
    
    except Exception as e:
        v_ds_trace = traceback.format_exc()
        log("ERRO .: " + str(e) + " >> " + v_ds_trace)
        v_nr_ret = 93
    
    sys.exit(v_nr_ret if v_nr_ret >= log.ret else log.ret )