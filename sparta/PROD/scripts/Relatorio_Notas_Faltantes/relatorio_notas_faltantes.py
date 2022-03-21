#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-
"""
----------------------------------------------------------------------------------------------
MODULO ...: TESHUVA
SCRIPT ...: loaderRelatorioNotasNaoEncontradasArquivoImpressao.py
CRIACAO ..: 14/07/2021
AUTOR ....: EDUARDO DA SILVA FERREIRA / KYROS TECNOLOGIA
            eduardof@kyros.com.br
DESCRICAO.: Geração de relatórios NOTAS não encontradas nos 
            arquivos de impressão 
----------------------------------------------------------------------------------------------
PARAMETROS: 
Parâmetros de entrada:
1)	MES_ANO_INICIAL: Mês e Ano (MMYYYY) Inicial - Obrigatório
2)	MES_ANO_FINAL: Mês e Ano (MMYYYY) Final - Obrigatório
3)	SERIE: Série da NOTA - Obrigatório
4)	UF: Unidade Federativa do Brasil 
    (Estado - Ex. SP) - Obrigatório

----------------------------------------------------------------------------------------------
  HISTORICO : 
        Adequação para novo formato de script 
        SCRIPT ......: loader_sped_registro_O150.py
        AUTOR .......: Victor Santos
----------------------------------------------------------------------------------------------
"""
import os
import sys

global SD, dir_base
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)

import configuracoes

import datetime
import cx_Oracle
import traceback
import string

gv_lista_string = list(string.ascii_lowercase)

import comum
import sql
import layout
import util
from pathlib import Path
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook

log.gerar_log_em_arquivo = True

def set_border_edsi( ws, cell_range, edsi):
    b=[]
 
    n=Side(border_style= None    , color='00000000')
    f=Side(border_style='thin'   , color='00000000')
    m=Side(border_style='medium' , color='00000000')
    d=Side(border_style='double' , color='00000000')
    p=Side(border_style='dotted' , color='00000000')
    t=Side(border_style='dashed' , color='00000000')
    g=Side(border_style='thick'  , color='00000000')
    
    for x in (0,1,2,3):
        if (edsi[x]=="n"):
            b.append(n)
        elif (edsi[x]=='f'):
            b.append(f)
        elif (edsi[x]=='m'):
            b.append(m)
        elif (edsi[x]=='d'):
            b.append(d)
        elif (edsi[x]=='p'):
            b.append(p)
        elif (edsi[x]=='t'):
            b.append(t)
        elif (edsi[x]=='g'):
            b.append(g)
        else:
            return
    
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = Border(left=b[0],right=b[1],top=b[2],bottom=b[3])

"""
Retorna dados
"""
def retorna_Dados(p_mes_ano_inicial,
                  p_mes_ano_final,  
                  p_serie,
                  p_uf,
                  p_conexao,
                  p_opc=0):

    log("Consultando os dados ... " + " <<" + str(p_opc) + ">> ")  
    log("Periodo ... " + " <<" + str(p_mes_ano_inicial) + ">> " + " <<" + str(p_mes_ano_final) + ">> " )  
    
    l_cabecalho_01 = ["CNPJ_FILIAL",
"IE_FILIAL",
"NUMERO_NF",
"SERIE",
"MODELO",
"VALOR_ESTORNO_CONV39",
"CPF_CNPJ_MES_ANTERIOR",
"CPF_CNPJ_MES_SEGUINTE",
"CPF_CNPJ_CV115ORIGINAL",
"IE_IMPMES_ANTERIOR",
"IE_IMPMES_SEGUINTE",
"IE_CV115ORIGINAL",
"RAZAOSOCIAL_IMPMES_ANTERIOR",
"RAZAOSOCIAL_IMPMES_SEGUINTE",
"RAZAOSOCIAL_CV115ORIGINAL",
"LOGRADOURO_IMPMES_ANTERIOR",
"LOGRADOURO_IMPMES_SEGUINTE",
"LOGRADOURO_CV115ORIGINAL",
"NUMERO_IMPMES_ANTERIOR",
"NUMERO_IMPMES_SEGUINTE",
"NUMERO_CV115ORIGINAL",
"COMPLEMENTO_IMPMES_ANTERIOR",
"COMPLEMENTO_IMPMES_SEGUINTE",
"COMPLEMENTO_CV115ORIGINAL",
"CEP_IMPMES_ANTERIOR",
"CEP_IMPMES_SEGUINTE",
"CEP_CV115ORIGINAL",
"BAIRRO_IMPMES_ANTERIOR",
"BAIRRO_IMPMES_SEGUINTE",
"BAIRRO_CV115ORIGINAL",
"MUNICIPIO_IMPMES_ANTERIOR",
"MUNICIPIO_IMPMES_SEGUINTE",
"MUNICIPIO_CV115ORIGINAL",
"UF_IMPMES_ANTERIOR",
"UF_IMPMES_SEGUINTE",
"UF_CV115ORIGINAL",
"TELEFONE_CONTATO_IMPMES_ANTERIOR",
"TELEFONE_CONTATO_IMPMES_SEGUINTE",
"TELEFONE_CONTATO_CV115ORIGINAL",
"CODIGO_CLIENTE_IMPMES_ANTERIOR",
"CODIGO_CLIENTE_IMPMES_SEGUINTE",
"CODIGO_CLIENTE_CV115ORIGINAL",
"TERMINAL_IMPMES_ANTERIOR",
"TERMINAL_IMPMES_SEGUINTE",
"TERMINAL_CV115ORIGINAL",
"UF_HABILITACAO_IMPMES_ANTERIOR",
"UF_HABILITACAO_IMPMES_SEGUINTE",
"UF_HABILITACAO_CV115ORIGINAL",
"DATA_EMISSAO_IMPMES_ANTERIOR",
"DATA_EMISSAO_IMPMES_SEGUINTE",
"DATA_EMISSAO_CV115ORIGINAL"]

    l_cabecalho_02 = ["SERIE", 
"MES_ANO_EMISSAO", 
"UF_FILIAL", 
"TOTAL_NOTAS_NAO_ENCONTRADAS", 
"VALOR_ESTORNO_NF_TOTAL", 
"VALOR_TOTAL_NOTAS", 
"TOTAL_BASE_ICMS", 
"VALOR_ICMS_TOTAL"]

    l_retorno=[[],[]]
    if p_opc == 0:
        l_retorno[0] = l_cabecalho_01
    else:
        l_retorno[0] = l_cabecalho_02

    l_data_inicial = "01" + "/" + str(p_mes_ano_inicial[0:2]) +  "/" +  str(p_mes_ano_inicial[2:6])
    l_data_final = "01" + "/" + str(p_mes_ano_final[0:2]) +  "/" +  str(p_mes_ano_final[2:6])

    try:

        l_query = ""    
        if p_opc == 0:
            l_query = """ 
                    WITH PARAMETROS AS
                    (SELECT to_date('%s','DD/MM/YYYY') PARAMETRO_DATA,
                        '%s'  AS PARAMETRO_SERIE,
                        '%s' AS PARAMETRO_UF
                    FROM DUAL
                    ),
                    IMPRESSAO_ANTERIOR AS
                    (SELECT L.EMPS_COD EMPS_COD_IMPR_ANTERIOR ,
                        L.FILI_COD FILI_COD_IMPR_ANTERIOR ,
                        L.SERIE MNFST_SERIE_IMPR_ANTERIOR ,
                        M.MNFST_DTEMISS MNFST_DTEMISS_IMPR_ANTERIOR ,
                        M.MNFST_NUM MNFST_NUM_IMPR_ANTERIOR ,
                        NULL TIPO_ASSINANTE_IMPR_ANTERIOR ,
                        NULL TIPO_UTILIZ_IMPR_ANTERIOR ,
                        NULL GRUPO_TENSAO_IMPR_ANTERIOR ,
                        M.CADG_COD CADG_COD_IMPR_ANTERIOR ,
                        M.CADG_NUM_CONTA TERMINAL_TELEF_IMPR_ANTERIOR ,
                        M.CADG_TIP_BILLING IND_CAMPO_01_IMPR_ANTERIOR ,
                        DECODE(CADG_TIP,'X','F',CADG_TIP) CADG_TIP_IMPR_ANTERIOR ,
                        NULL CADG_TIP_CLI_IMPR_ANTERIOR ,
                        NULL SUB_CLASSE_IMPR_ANTERIOR ,
                        NULL TERMINAL_PRINC_IMPR_ANTERIOR ,
                        NULL CNPJ_EMIT_IMPR_ANTERIOR ,
                        M.CADG_COD_CGCCPF CNPJ_CPF_IMPR_ANTERIOR ,
                        M.CADG_COD_INSEST IE_IMPR_ANTERIOR ,
                        M.CADG_NOM RAZAO_SOCIAL_IMPR_ANTERIOR ,
                        M.CADG_END ENDEREC_IMPR_ANTERIOR ,
                        M.CADG_END_NUM_BILLING NUMERO_IMPR_ANTERIOR ,
                        M.CADG_END_COMP COMPLEMENTO_IMPR_ANTERIOR ,
                        M.CADG_END_CEP CEP_IMPR_ANTERIOR ,
                        M.CADG_END_BAIRRO BAIRRO_IMPR_ANTERIOR ,
                        M.CADG_END_MUNIC MUNICIPIO_IMPR_ANTERIOR ,
                        M.UNFE_SIG UF_IMPR_ANTERIOR ,
                        M.CADG_TEL_CONTATO CADG_TEL_CONTATO_IMPR_ANTERIOR ,
                        M.CADG_COD COD_CONSUMIDOR_IMPR_ANTERIOR ,
                        M.CADG_NUM_CONTA NUMEROTERMINAL_IMPR_ANTERIOR ,
                        m.CADG_UF_HABILIT UFHABILITACAO_IMPR_ANTERIOR ,
                        M.origem ORIGEM_IMPR_ANTERIOR ,
                        M.MIBGE_COD_MUN CODIGOMUNICIPIO_IMPR_ANTERIOR ,
                        VALIDA_DOCUMENTO(M.CADG_COD_CGCCPF) DOC_VALIDO_IMPR_ANTERIOR ,
                        M.var05 VAR05_IMPR_ANTERIOR
                    FROM PARAMETROS P,
                        gfcadastro.billing_combinado_final M
                    JOIN gfcarga.tsh_serie_levantamento l
                    ON l.uf_filial              = M.uf_filial
                    AND REPLACE(l.serie,' ','') = M.mnfst_serie
                    AND l.mes_ano               = to_date('01'
                        || TO_CHAR(M.mnfst_dtemiss,'MMYYYY'),'DDMMYYYY')
                    WHERE REPLACE(l.serie,' ','') = REPLACE(P.PARAMETRO_SERIE,' ','')
                    AND m.mnfst_dtemiss          >= ADD_MONTHS(TRUNC(P.PARAMETRO_DATA,'MM'),         -1)
                    AND m.mnfst_dtemiss          <= LAST_DAY(ADD_MONTHS(TRUNC(P.PARAMETRO_DATA,'MM'),-1))
                    AND m.UF_FILIAL               = P.PARAMETRO_UF
                    ),
                    IMPRESSAO_DEPOIS AS
                    (SELECT L.EMPS_COD EMPS_COD_IMPR_DEPOIS ,
                        L.FILI_COD FILI_COD_IMPR_DEPOIS ,
                        L.SERIE MNFST_SERIE_IMPR_DEPOIS ,
                        M.MNFST_DTEMISS MNFST_DTEMISS_IMPR_DEPOIS ,
                        M.MNFST_NUM MNFST_NUM_IMPR_DEPOIS ,
                        NULL TIPO_ASSINANTE_IMPR_DEPOIS ,
                        NULL TIPO_UTILIZ_IMPR_DEPOIS ,
                        NULL GRUPO_TENSAO_IMPR_DEPOIS ,
                        M.CADG_COD CADG_COD_IMPR_DEPOIS ,
                        M.CADG_NUM_CONTA TERMINAL_TELEF_IMPR_DEPOIS ,
                        M.CADG_TIP_BILLING IND_CAMPO_01_IMPR_DEPOIS ,
                        DECODE(CADG_TIP,'X','F',CADG_TIP) CADG_TIP_IMPR_DEPOIS ,
                        NULL CADG_TIP_CLI_IMPR_DEPOIS ,
                        NULL SUB_CLASSE_IMPR_DEPOIS ,
                        NULL TERMINAL_PRINC_IMPR_DEPOIS ,
                        NULL CNPJ_EMIT_IMPR_DEPOIS ,
                        M.CADG_COD_CGCCPF CNPJ_CPF_IMPR_DEPOIS ,
                        M.CADG_COD_INSEST IE_IMPR_DEPOIS ,
                        M.CADG_NOM RAZAO_SOCIAL_IMPR_DEPOIS ,
                        M.CADG_END ENDEREC_IMPR_DEPOIS ,
                        M.CADG_END_NUM_BILLING NUMERO_IMPR_DEPOIS ,
                        M.CADG_END_COMP COMPLEMENTO_IMPR_DEPOIS ,
                        M.CADG_END_CEP CEP_IMPR_DEPOIS ,
                        M.CADG_END_BAIRRO BAIRRO_IMPR_DEPOIS ,
                        M.CADG_END_MUNIC MUNICIPIO_IMPR_DEPOIS ,
                        M.UNFE_SIG UF_IMPR_DEPOIS ,
                        M.CADG_TEL_CONTATO CADG_TEL_CONTATO_IMPR_DEPOIS ,
                        M.CADG_COD COD_CONSUMIDOR_IMPR_DEPOIS ,
                        M.CADG_NUM_CONTA NUMEROTERMINAL_IMPR_DEPOIS ,
                        m.CADG_UF_HABILIT UFHABILITACAO_IMPR_DEPOIS ,
                        M.origem ORIGEM_IMPR_DEPOIS ,
                        M.MIBGE_COD_MUN CODIGOMUNICIPIO_IMPR_DEPOIS ,
                        VALIDA_DOCUMENTO(M.CADG_COD_CGCCPF) DOC_VALIDO_IMPR_DEPOIS ,
                        M.var05 VAR05_IMPR_DEPOIS
                    FROM PARAMETROS P,
                        gfcadastro.billing_combinado_final M
                    JOIN gfcarga.tsh_serie_levantamento l
                    ON l.uf_filial              = M.uf_filial
                    AND REPLACE(l.serie,' ','') = M.mnfst_serie
                    AND l.mes_ano               = to_date('01'
                        || TO_CHAR(M.mnfst_dtemiss,'MMYYYY'),'DDMMYYYY')
                    WHERE REPLACE(l.serie,' ','') = REPLACE(P.PARAMETRO_SERIE,' ','')
                    AND m.mnfst_dtemiss          >= ADD_MONTHS(TRUNC(P.PARAMETRO_DATA,'MM'),1)
                    AND m.mnfst_dtemiss          <= LAST_DAY(ADD_MONTHS(TRUNC(P.PARAMETRO_DATA,'MM'),1))
                    AND m.UF_FILIAL               = P.PARAMETRO_UF
                    ),
                    IMPRESSAO AS
                    (SELECT L.EMPS_COD EMPS_COD_IMPRESSAO ,
                        L.FILI_COD FILI_COD_IMPRESSAO ,
                        L.SERIE MNFST_SERIE_IMPRESSAO ,
                        M.MNFST_DTEMISS MNFST_DTEMISS_IMPRESSAO ,
                        M.MNFST_NUM MNFST_NUM_IMPRESSAO ,
                        NULL TIPO_ASSINANTE_IMPRESSAO ,
                        NULL TIPO_UTILIZ_IMPRESSAO ,
                        NULL GRUPO_TENSAO_IMPRESSAO ,
                        M.CADG_COD CADG_COD_IMPRESSAO ,
                        M.CADG_NUM_CONTA TERMINAL_TELEF_IMPRESSAO ,
                        M.CADG_TIP_BILLING IND_CAMPO_01_IMPRESSAO ,
                        DECODE(CADG_TIP,'X','F',CADG_TIP) CADG_TIP_IMPRESSAO ,
                        NULL CADG_TIP_CLI_IMPRESSAO ,
                        NULL SUB_CLASSE_IMPRESSAO ,
                        NULL TERMINAL_PRINC_IMPRESSAO ,
                        NULL CNPJ_EMIT_IMPRESSAO ,
                        M.CADG_COD_CGCCPF CNPJ_CPF_IMPRESSAO ,
                        M.CADG_COD_INSEST IE_IMPRESSAO ,
                        M.CADG_NOM RAZAO_SOCIAL_IMPRESSAO ,
                        M.CADG_END ENDEREC_IMPRESSAO ,
                        M.CADG_END_NUM_BILLING NUMERO_IMPRESSAO ,
                        M.CADG_END_COMP COMPLEMENTO_IMPRESSAO ,
                        M.CADG_END_CEP CEP_IMPRESSAO ,
                        M.CADG_END_BAIRRO BAIRRO_IMPRESSAO ,
                        M.CADG_END_MUNIC MUNICIPIO_IMPRESSAO ,
                        M.UNFE_SIG UF_IMPRESSAO ,
                        M.CADG_TEL_CONTATO CADG_TEL_CONTATO_IMPRESSAO ,
                        M.CADG_COD COD_CONSUMIDOR_IMPRESSAO ,
                        M.CADG_NUM_CONTA NUMEROTERMINAL_IMPRESSAO ,
                        m.CADG_UF_HABILIT UFHABILITACAO_IMPRESSAO ,
                        M.origem ORIGEM_IMPRESSAO ,
                        M.MIBGE_COD_MUN CODIGOMUNICIPIO_IMPRESSAO ,
                        VALIDA_DOCUMENTO(M.CADG_COD_CGCCPF) DOC_VALIDO_IMPRESSAO ,
                        M.var05 VAR05_IMPRESSAO
                    FROM PARAMETROS P,
                        gfcadastro.billing_combinado_final M
                    JOIN gfcarga.tsh_serie_levantamento l
                    ON l.uf_filial              = M.uf_filial
                    AND REPLACE(l.serie,' ','') = M.mnfst_serie
                    AND l.mes_ano               = to_date('01'
                        || TO_CHAR(M.mnfst_dtemiss,'MMYYYY'),'DDMMYYYY')
                    WHERE REPLACE(l.serie,' ','') = REPLACE(P.PARAMETRO_SERIE,' ','')
                    AND m.mnfst_dtemiss          >= TRUNC(P.PARAMETRO_DATA,'MM')
                    AND m.mnfst_dtemiss          <= LAST_DAY(TRUNC(P.PARAMETRO_DATA,'MM'))
                    AND m.UF_FILIAL               = P.PARAMETRO_UF
                    ),
                    PROTOCOLADO AS
                    (SELECT l.id_serie_levantamento,
                        l.EMPS_COD EMPS_COD_ORIGINAL ,
                        l.FILI_COD FILI_COD_ORIGINAL ,
                        l.SERIE MNFST_SERIE_ORIGINAL ,
                        mp.DATA_EMISSAO MNFST_DTEMISS_ORIGINAL ,
                        mp.MODELO MODELO_ORIGINAL ,
                        mp.NUMERO_NF MNFST_NUM_ORIGINAL ,
                        MP.CLASSE_CONS TIPO_ASSINANTE_ORIGINAL ,
                        MP.TIPO_UTILIZ TIPO_UTILIZ_ORIGINAL ,
                        MP.GRUPO_TENSAO GRUPO_TENSAO_ORIGINAL ,
                        MP.CADG_COD CADG_COD_ORIGINAL ,
                        MP.TERMINAL_TELEF TERMINAL_TELEF_ORIGINAL ,
                        MP.IND_CAMPO_01 IND_CAMPO_01_ORIGINAL ,
                        MP.TIPO_CLIENTE CADG_TIP_CLI_ORIGINAL ,
                        MP.SUB_CLASSE SUB_CLASSE_ORIGINAL ,
                        MP.TERMINAL_PRINC TERMINAL_PRINC_ORIGINAL ,
                        MP.CNPJ_EMIT CNPJ_EMIT_ORIGINAL ,
                        D.CNPJ_CPF CNPJ_CPF_ORIGINAL ,
                        D.IE IE_ORIGINAL ,
                        D.RAZAOSOCIAL RAZAO_SOCIAL_ORIGINAL ,
                        D.ENDERECO ENDEREC_ORIGINAL ,
                        TO_CHAR(lpad(D.NUMERO,5,'0')) NUMERO_ORIGINAL ,
                        D.COMPLEMENTO COMPLEMENTO_ORIGINAL ,
                        lpad(D.CEP, 8, '0') CEP_ORIGINAL ,
                        D.BAIRRO BAIRRO_ORIGINAL ,
                        D.MUNICIPIO MUNICIPIO_ORIGINAL ,
                        D.UF UF_ORIGINAL ,
                        D.TELEFONECONTATO CADG_TEL_CONTATO_ORIGINAL ,
                        D.CODIDENTCONSUMIDOR COD_CONSUMIDOR_ORIGINAL ,
                        D.NUMEROTERMINAL NUMEROTERMINAL_ORIGINAL ,
                        D.UFHABILITACAO UFHABILITACAO_ORIGINAL ,
                        D.CODIGOMUNICIPIO CODIGOMUNICIPIO_ORIGINAL ,
                        VALIDA_DOCUMENTO(D.CNPJ_CPF) DOC_VALIDO_ORIGINAL ,
                        CASE
                        WHEN TO_CHAR(mp.data_emissao, 'YYYY') <= 2016
                        THEN
                            CASE
                            WHEN valida_documento(mp.cnpj_cpf) = 'S'
                            AND LENGTH(mp.cnpj_cpf)            = 14
                            THEN 'J'
                            ELSE 'F'
                            END
                        ELSE DECODE(mp.ind_campo_01, 1, 'J', 2, 'F', 3, 'E', 4, 'I', 'F')
                        END CADG_TIP_ORIGINAL
                    FROM PARAMETROS P,
                        GFCARGA.TSH_MESTRE_CONV_115 MP
                    JOIN GFCARGA.TSH_DESTINATARIO_CONV_115 D
                    ON MP.ID_SERIE_LEVANTAMENTO = D.ID_SERIE_LEVANTAMENTO
                    AND MP.VOLUME               = D.VOLUME
                    AND MP.LINHA                = D.LINHA
                    AND MP.UF_FILIAL            = D.UF_FILIAL
                    JOIN GFCARGA.TSH_SERIE_LEVANTAMENTO L
                    ON MP.ID_SERIE_LEVANTAMENTO   = L.ID_SERIE_LEVANTAMENTO
                    WHERE REPLACE(l.serie,' ','') = REPLACE(P.PARAMETRO_SERIE,' ','')
                    AND L.mes_ano                 = TRUNC(P.PARAMETRO_DATA,'MM')
                    AND MP.UF_FILIAL              = P.PARAMETRO_UF
                    ),
                    CONVENIO_39 AS
                    (SELECT SERIE ,
                        EMPS_COD ,
                        FILI_COD ,
                        DT_EMISSAO ,
                        NU_NF ,
                        SUM(val_icms_item)/100 VALOR_ESTORNO_NF
                    FROM PARAMETROS P,
                        gfcadastro.CONV39_ESTORNO_IMPOSTO_NF_VW e
                    WHERE e.SERIE     = P.PARAMETRO_SERIE
                    AND e.DT_EMISSAO >= TRUNC(P.PARAMETRO_DATA,'MM')
                    AND e.DT_EMISSAO <= LAST_DAY(TRUNC(P.PARAMETRO_DATA,'MM'))
                    GROUP BY SERIE,
                        EMPS_COD,
                        FILI_COD,
                        DT_EMISSAO,
                        NU_NF
                    )
                    SELECT
                    /*+ parallel(8) */
                    FILI_COD_CGC CNPJ_FILIAL ,
                    FILI_COD_INSEST IE_FILIAL ,
                    MNFST_NUM_ORIGINAL NUMERO_NF ,
                    MNFST_SERIE_ORIGINAL SERIE ,
                    MODELO_ORIGINAL MODELO ,
                    VALOR_ESTORNO_NF VALOR_ESTORNO_CONV39 ,
                    CNPJ_CPF_IMPR_ANTERIOR CPF_CNPJ_IMPANTERIOR,
                    CNPJ_CPF_IMPR_DEPOIS CPF_CNPJ_IMPDEPOIS,
                    CNPJ_CPF_ORIGINAL CPF_CNPJ_CV115ORIGINAL,
                    IE_IMPR_ANTERIOR IE_IMPANTERIOR,
                    IE_IMPR_DEPOIS IE_IMPDEPOIS,
                    IE_ORIGINAL IE_CV115ORIGINAL,
                    RAZAO_SOCIAL_IMPR_ANTERIOR RAZAOSOCIAL_IMPANTERIOR,
                    RAZAO_SOCIAL_IMPR_DEPOIS RAZAOSOCIAL_IMPDEPOIS,
                    RAZAO_SOCIAL_ORIGINAL RAZAOSOCIAL_CV115ORIGINAL,
                    ENDEREC_IMPR_ANTERIOR LOGRADOURO_IMPANTERIOR,
                    ENDEREC_IMPR_DEPOIS LOGRADOURO_IMPDEPOIS,
                    ENDEREC_ORIGINAL LOGRADOURO_CV115ORIGINAL,
                    NUMERO_IMPR_ANTERIOR NUMERO_IMPANTERIOR,
                    NUMERO_IMPR_DEPOIS NUMERO_IMPDEPOIS,
                    NUMERO_ORIGINAL NUMERO_CV115ORIGINAL,
                    COMPLEMENTO_IMPR_ANTERIOR COMPLEMENTO_IMPANTERIOR,
                    COMPLEMENTO_IMPR_DEPOIS COMPLEMENTO_IMPDEPOIS,
                    COMPLEMENTO_ORIGINAL COMPLEMENTO_CV115ORIGINAL,
                    CEP_IMPR_ANTERIOR CEP_IMPANTERIOR,
                    CEP_IMPR_DEPOIS CEP_IMPDEPOIS,
                    CEP_ORIGINAL CEP_CV115ORIGINAL,
                    BAIRRO_IMPR_ANTERIOR BAIRRO_IMPANTERIOR,
                    BAIRRO_IMPR_DEPOIS BAIRRO_IMPDEPOIS,
                    BAIRRO_ORIGINAL BAIRRO_CV115ORIGINAL,
                    MUNICIPIO_IMPR_ANTERIOR MUNICIPIO_IMPANTERIOR,
                    MUNICIPIO_IMPR_DEPOIS MUNICIPIO_IMPDEPOIS,
                    MUNICIPIO_ORIGINAL MUNICIPIO_CV115ORIGINAL,
                    UF_IMPR_ANTERIOR UF_IMPANTERIOR,
                    UF_IMPR_DEPOIS UF_IMPDEPOIS,
                    UF_ORIGINAL UF_CV115ORIGINAL,
                    CADG_TEL_CONTATO_IMPR_ANTERIOR TELEFONE_CONTATO_IMPANTERIOR,
                    CADG_TEL_CONTATO_IMPR_DEPOIS TELEFONE_CONTATO_IMPDEPOIS,
                    CADG_TEL_CONTATO_ORIGINAL TELEFONE_CONTATO_CV115ORIGINAL,
                    COD_CONSUMIDOR_IMPR_ANTERIOR CODIGO_CLIENTE_IMPANTERIOR,
                    COD_CONSUMIDOR_IMPR_DEPOIS CODIGO_CLIENTE_IMPDEPOIS,
                    COD_CONSUMIDOR_ORIGINAL CODIGO_CLIENTE_CV115ORIGINAL,
                    NUMEROTERMINAL_IMPR_ANTERIOR TERMINAL_IMPANTERIOR,
                    NUMEROTERMINAL_IMPR_DEPOIS TERMINAL_IMPDEPOIS,
                    NUMEROTERMINAL_ORIGINAL TERMINAL_CV115ORIGINAL,
                    UFHABILITACAO_IMPR_ANTERIOR UF_HABILITACAO_IMPANTERIOR,
                    UFHABILITACAO_IMPR_DEPOIS UF_HABILITACAO_IMPDEPOIS,
                    UFHABILITACAO_ORIGINAL UF_HABILITACAO_CV115ORIGINAL,
                    TO_CHAR(MNFST_DTEMISS_IMPR_ANTERIOR,'DD/MM/YYYY') DATA_EMISSAO_IMPANTERIOR,
                    TO_CHAR(MNFST_DTEMISS_IMPR_DEPOIS,'DD/MM/YYYY') DATA_EMISSAO_IMPDEPOIS,
                    TO_CHAR(MNFST_DTEMISS_ORIGINAL,'DD/MM/YYYY') DATA_EMISSAO_CV115ORIGINAL
                    FROM protocolado p
                    LEFT JOIN IMPRESSAO I
                    ON p.EMPS_COD_ORIGINAL       = I.EMPS_COD_IMPRESSAO
                    AND p.FILI_COD_ORIGINAL      = I.FILI_COD_IMPRESSAO
                    AND p.MNFST_SERIE_ORIGINAL   = I.MNFST_SERIE_IMPRESSAO
                    AND p.MNFST_NUM_ORIGINAL     = I.MNFST_NUM_IMPRESSAO
                    AND p.MNFST_DTEMISS_ORIGINAL = I.MNFST_DTEMISS_IMPRESSAO
                    LEFT JOIN IMPRESSAO_ANTERIOR j
                    ON p.EMPS_COD_ORIGINAL        = j.EMPS_COD_IMPR_ANTERIOR
                    AND p.FILI_COD_ORIGINAL       = j.FILI_COD_IMPR_ANTERIOR
                    AND p.MNFST_SERIE_ORIGINAL    = j.MNFST_SERIE_IMPR_ANTERIOR
                    AND p.COD_CONSUMIDOR_ORIGINAL = j.COD_CONSUMIDOR_IMPR_ANTERIOR
                    LEFT JOIN IMPRESSAO_DEPOIS a
                    ON p.EMPS_COD_ORIGINAL        = a.EMPS_COD_IMPR_DEPOIS
                    AND p.FILI_COD_ORIGINAL       = a.FILI_COD_IMPR_DEPOIS
                    AND p.MNFST_SERIE_ORIGINAL    = a.MNFST_SERIE_IMPR_DEPOIS
                    AND p.COD_CONSUMIDOR_ORIGINAL = a.COD_CONSUMIDOR_IMPR_DEPOIS
                    LEFT JOIN CONVENIO_39 e
                    ON p.MNFST_SERIE_ORIGINAL    = e.SERIE
                    AND p.EMPS_COD_ORIGINAL      = e.EMPS_COD
                    AND p.FILI_COD_ORIGINAL      = e.FILI_COD
                    AND p.MNFST_DTEMISS_ORIGINAL = e.DT_EMISSAO
                    AND p.MNFST_NUM_ORIGINAL     = e.NU_NF
                    LEFT JOIN openrisow.FILIAL c
                    ON p.FILI_COD_ORIGINAL  = c.FILI_COD
                    AND p.EMPS_COD_ORIGINAL = C.EMPS_COD
                    WHERE ORIGEM_IMPRESSAO IS NULL
                    ORDER BY p.MNFST_NUM_ORIGINAL 
        """%(l_data_inicial,p_serie,p_uf)
        else:
            l_query = """ 
            WITH PARAMETROS AS
  (SELECT TRUNC(TO_DATE('%s','DD/MM/YYYY'),'MM') AS PARAMETRO_DATA_INICIAL,
    LAST_DAY(TO_DATE('%s','DD/MM/YYYY'))         AS PARAMETRO_DATA_FINAL,
    '%s'                                         AS PARAMETRO_SERIE,
    '%s'                                         AS PARAMETRO_UF
  FROM DUAL
  ),
  BILLING AS
  (SELECT L.EMPS_COD EMPS_COD_M,
    L.FILI_COD FILI_COD_M,
    M.MNFST_SERIE MNFST_SERIE_M,
    M.MNFST_DTEMISS MNFST_DTEMISS_M,
    M.MNFST_NUM MNFST_NUM_M,
    M.CADG_COD_CGCCPF CNPJ_CPF_M,
    M.CADG_COD_INSEST IE_M,
    REPLACE(TRIM(SUBSTR(M.CADG_NOM, 1, 35)),CHR(160),'') RAZAO_SOCIAL_M,
    M.UNFE_SIG UF_M,
    SUBSTR(M.CADG_COD, 1, 12) CADG_COD_M,
    M.CADG_NUM_CONTA TERMINAL_TELEF_M,
    DECODE(NVL(TRIM(M.CADG_TIP),'X'), 'J', '1', 'F', '2', 'X') IND_CAMPO_01_M,
    NULL TERMINAL_PRINC_M,
    M.CADG_COD_CGCCPF CNPJ_CPF_C,
    M.CADG_COD_INSEST IE_C,
    REPLACE(TRIM(SUBSTR(M.CADG_NOM, 1, 35)),CHR(160),'') RAZAO_SOCIAL_C,
    SUBSTR(M.CADG_END, 1, 45) ENDEREC_C,
    CASE
      WHEN EHNUMERICO(M.CADG_END_NUM_BILLING) = 'S'
      THEN TO_CHAR(TO_NUMBER(M.CADG_END_NUM_BILLING))
      ELSE M.CADG_END_NUM_BILLING
    END NUMERO_C,
    TRIM(SUBSTR(M.CADG_END_COMP, 1, 15)) COMPLEMENTO_C,
    M.CADG_END_CEP CEP_C,
    M.CADG_END_BAIRRO BAIRRO_C,
    M.CADG_END_MUNIC MUNICIPIO_C,
    M.UNFE_SIG UF_C,
    M.CADG_TEL_CONTATO CADG_TEL_CONTATO_C,
    SUBSTR(M.CADG_COD, 1, 12) CODIDENTCONSUMIDOR_C,
    M.CADG_NUM_CONTA NUMEROTERMINAL_C,
    M.CADG_UF_HABILIT UFHABILITACAO_C,
    M.MIBGE_COD_MUN CODIGOMUNICIPIO_C,
    M.ORIGEM SISTEMA_ORIGEM_C
  FROM PARAMETROS P ,
    GFCADASTRO.BILLING_COMBINADO_FINAL M
  JOIN GFCARGA.TSH_SERIE_LEVANTAMENTO L
  ON L.UF_FILIAL              = M.UF_FILIAL
  AND REPLACE(L.SERIE,' ','') = REPLACE(M.MNFST_SERIE,' ','')
  AND L.MES_ANO               = TO_DATE('01'
    || TO_CHAR(M.MNFST_DTEMISS,'MMYYYY'),'DDMMYYYY')
  WHERE M.MNFST_DTEMISS >= P.PARAMETRO_DATA_INICIAL
  AND M.MNFST_DTEMISS   <= P.PARAMETRO_DATA_FINAL
  AND M.UF_FILIAL        = P.PARAMETRO_UF
  AND M.MNFST_SERIE      = P.PARAMETRO_SERIE
  ),
  PROTOCOLADO AS
  (SELECT 
    MP.UF_FILIAL,
    REPLACE(L.SERIE, ' ', '') SERIE,
    L.EMPS_COD,
    L.FILI_COD,
    L.ID_SERIE_LEVANTAMENTO,
    MP.DATA_EMISSAO,
    MP.NUMERO_NF,
    MP.CNPJ_CPF CNPJ_CPF_MP,
    MP.IE IE_MP,    
    MP.UF UF_MP,
    MP.RAZAO_SOCIAL RAZAO_SOCIAL_MP,
    MP.CADG_COD CADG_COD_MP,
    MP.TERMINAL_TELEF TERMINAL_TELEF_MP,
    MP.IND_CAMPO_01 IND_CAMPO_01_MP,
    MP.TERMINAL_PRINC TERMINAL_PRINC_MP,
    D.CNPJ_CPF CNPJ_CPF_CP,
    D.IE IE_CP,
    D.RAZAOSOCIAL RAZAOSOCIAL_CP,
    D.ENDERECO ENDERECO_CP,
    TO_CHAR(D.NUMERO) NUMERO_CP,
    D.COMPLEMENTO COMPLEMENTO_CP,
    D.CEP CEP_CP,
    D.BAIRRO BAIRRO_CP,
    D.MUNICIPIO MUNICIPIO_CP,
    D.UF UF_CP,
    D.TELEFONECONTATO TELEFONECONTATO_CP,
    D.CODIDENTCONSUMIDOR CODIDENTCONSUMIDOR_CP,
    D.NUMEROTERMINAL NUMEROTERMINAL_CP,
    D.UFHABILITACAO UFHABILITACAO_CP,
    D.CODIGOMUNICIPIO CODIGOMUNICIPIO_CP ,
    MP.VALOR_TOTAL ,
    MP.BASE_ICMS ,
    MP.VALOR_ICMS ,
    MP.SIT_DOC
  FROM PARAMETROS P ,
    GFCARGA.TSH_MESTRE_CONV_115 MP
  JOIN GFCARGA.TSH_DESTINATARIO_CONV_115 D
  ON MP.ID_SERIE_LEVANTAMENTO = D.ID_SERIE_LEVANTAMENTO
  AND MP.VOLUME               = D.VOLUME
  AND MP.LINHA                = D.LINHA
  AND MP.UF_FILIAL            = D.UF_FILIAL
  JOIN GFCARGA.TSH_SERIE_LEVANTAMENTO L
  ON MP.ID_SERIE_LEVANTAMENTO = L.ID_SERIE_LEVANTAMENTO
  WHERE L.MES_ANO            >= P.PARAMETRO_DATA_INICIAL
  AND L.MES_ANO              <= P.PARAMETRO_DATA_FINAL
  AND MP.UF_FILIAL            = P.PARAMETRO_UF
  AND REPLACE(L.SERIE,' ','') = REPLACE(P.PARAMETRO_SERIE,' ','')
  ),
  CONVENIO_39 AS
  (SELECT SERIE ,
    EMPS_COD ,
    FILI_COD ,
    DT_EMISSAO ,
    NU_NF ,
    SUM(VAL_ICMS_ITEM) VALOR_ESTORNO_NF
  FROM PARAMETROS P ,
    GFCADASTRO.CONV39_ESTORNO_IMPOSTO_NF_VW E
  WHERE E.SERIE     = P.PARAMETRO_SERIE
  AND E.DT_EMISSAO >= P.PARAMETRO_DATA_INICIAL
  AND E.DT_EMISSAO <= P.PARAMETRO_DATA_FINAL
  GROUP BY SERIE,
    EMPS_COD,
    FILI_COD,
    DT_EMISSAO,
    NU_NF
  )
SELECT /*+ parallel(8) */
  P.SERIE ,
  TO_CHAR(P.DATA_EMISSAO, 'MM/YYYY') MES_ANO_EMISSAO ,
  P.UF_FILIAL ,
  COUNT(1) AS TOTAL_NOTAS_NAO_ENCONTRADAS,
  SUM(E.VALOR_ESTORNO_NF) AS VALOR_ESTORNO_NF_TOTAL ,
  SUM(P.VALOR_TOTAL) AS VALOR_TOTAL_NOTAS ,
  SUM(P.BASE_ICMS) TOTAL_BASE_ICMS,
  SUM(P.VALOR_ICMS) VALOR_ICMS_TOTAL
FROM PROTOCOLADO P
LEFT JOIN BILLING R
ON P.SERIE         = R.MNFST_SERIE_M
AND P.EMPS_COD     = R.EMPS_COD_M
AND P.FILI_COD     = R.FILI_COD_M
AND P.DATA_EMISSAO = R.MNFST_DTEMISS_M
AND P.NUMERO_NF    = R.MNFST_NUM_M
LEFT JOIN CONVENIO_39 E
ON P.SERIE           = E.SERIE
AND P.EMPS_COD       = E.EMPS_COD
AND P.FILI_COD       = E.FILI_COD
AND P.DATA_EMISSAO   = E.DT_EMISSAO
AND P.NUMERO_NF      = E.NU_NF
WHERE 1              = 1
AND R.MNFST_SERIE_M IS NULL
GROUP BY P.SERIE,
  TO_CHAR(P.DATA_EMISSAO, 'MM/YYYY'),
  P.UF_FILIAL 
ORDER BY SERIE, SUBSTR(MES_ANO_EMISSAO,4,4), SUBSTR(MES_ANO_EMISSAO,1,2)  
        """%(l_data_inicial,l_data_final,p_serie,p_uf)
        
        log("Executando a consulta ... " + " <<" + str(p_opc) + ">> ")  
        l_cursor = sql.geraCnxBD(configuracoes)
        l_cursor.executa(l_query)
        l_result = l_cursor.fetchone()
        l_lin = 0

        if l_result is not None:
            log("Carregando na memoria os dados da consulta ... " + " <<" + str(p_opc) + ">> ")      
            while l_result:
                # + 1
                l_lin += 1
                #Cria uma linha para informar as colunas
                l_retorno.append([])
                # busca as colunas
                for field in l_result:
                    l_retorno[l_lin].append(field)
                # prox. registro
                l_result = l_cursor.fetchone()
        
        log("Finalizando a consulta ... " + " <<" + str(p_opc) + ">> ")    

    except Exception as e:
        txt = traceback.format_exc()
        log(" >> FALHA NA GERAÇÃO DOS DADOS ! " + str(txt) + " >> " +  str(e))    
        l_retorno = None
        
    return l_retorno    

"""
Processamento de relatório Analitico
"""
def processarRelatorioAnalitico(p_mes_ano,
                                p_serie,
                                p_uf,
                                p_relatorio,
                                p_conexao):
    l_ret = 0
    try:
        l_dados = []
        if not l_ret:
            l_dados = retorna_Dados(p_mes_ano_inicial=p_mes_ano,
                                    p_mes_ano_final=p_mes_ano, 
                                    p_serie=p_serie,
                                    p_uf=p_uf,
                                    p_conexao=p_conexao)
            if l_dados is None:
                log("Não foi encontrado nenhum dados [Analitico] ! (1) " + " << " + str(p_mes_ano) + " >> ")
                l_ret = 1
            elif len(l_dados) <= 2:
                log("Não foi encontrado nenhum dados [Analitico] ! (2) " + " << " + str(p_mes_ano) + " >> ")
                l_ret = 1
            else:
                log("Quantidade de dados [Analitico] : " + str(len(l_dados)-1)  + " << " + str(p_mes_ano) + " >> "  )
        if not l_ret:
            l_relatorio = str(p_relatorio).replace("<<MES_ANO>>", str(p_mes_ano[2:6])+ str(p_mes_ano[0:2]))    
            log("\n")
            log("Iniciando a criacao do relatório " + str(l_relatorio)) 
            arquivo_excel = Workbook()
            
            log("Criando os worksheet ... ")  
            planilha0 = arquivo_excel.active
            planilha0.title = p_mes_ano + "_" + p_uf + "_" + p_serie
         
            log("Carregando os dados para planilha ... ")
            l_linha:int = 0
            for dado in l_dados:
                planilha0.append(dado)
                l_linha += 1

            log("Aumentando as colunas ... ")
            for nColP in gv_lista_string:
                planilha0.column_dimensions[str(nColP).upper()].width = 25   
                planilha0.column_dimensions["A"+str(nColP).upper()].width = 25   

            log("Colocando cor nas linhas e colunas ... ")
            fontMasterPreta    = Font(color='00000000', bold=True, size=12)
            for nCol in range(len(l_dados[0])):
                planilha0.cell(1,nCol+1).font=fontMasterPreta
                planilha0.cell(1,nCol+1).alignment = Alignment(horizontal='center')
                planilha0.cell(1,nCol+1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
  
            log("Salvando a planilha : " + str(l_relatorio))
            arquivo_excel.save(l_relatorio)

    except Exception as e:
        l_ret = 1    
        txt = traceback.format_exc()
        log("ERRO NA GERAÇÃO DO RELATÓRIO.: " + txt +  " >> " +  str(e))

    return l_ret

"""
Processamento de relatório 
"""
def processarRelatorio(p_mes_ano_inicial,
                       p_mes_ano_final,  
                       p_serie,
                       p_uf,
                       p_relatorio,
                       p_resumo,
                       p_conexao,
                       p_analitico="N"):
    l_ret = 0
    l_dados = []
    try:
        if not l_ret:
            l_dados = retorna_Dados(p_mes_ano_inicial=p_mes_ano_inicial,
                                    p_mes_ano_final=p_mes_ano_final, 
                                    p_serie=p_serie,
                                    p_uf=p_uf,
                                    p_conexao=p_conexao,
                                    p_opc=1)
            if l_dados is None:
                log("Não foi encontrado nenhum dados [RESUMO] ! (1) ")
                l_ret = 1
            elif len(l_dados) <= 2:
                log("Não foi encontrado nenhum dados [RESUMO] ! (2) ")
                l_ret = 1
            else:
                log("Quantidade de dados [RESUMO] : " + str(len(l_dados)-1) )

        if not l_ret:
            l_relatorio = str(p_resumo).replace("<<MES_ANO>>",p_resumo)    
            log("\n")
            log("Iniciando a criacao do relatório " + str(l_relatorio)) 
            arquivo_excel = Workbook()
            
            log("Criando os worksheet ... ")  
            planilha0 = arquivo_excel.active
            planilha0.title = p_mes_ano_inicial + "_" +  p_mes_ano_final + "_" + p_uf + "_" + p_serie
         
            log("Carregando os dados para planilha ... ")
            l_linha:int = 0
            for dado in l_dados:
                planilha0.append(dado)
                l_linha += 1

            log("Aumentando as colunas ... ")
            for nColP in gv_lista_string:
                planilha0.column_dimensions[str(nColP).upper()].width = 25   

            log("Colocando cor nas linhas e colunas ... ")
            fontMasterPreta    = Font(color='00000000', bold=True, size=12)
            for nCol in range(len(l_dados[0])):
                planilha0.cell(1,nCol+1).font=fontMasterPreta
                planilha0.cell(1,nCol+1).alignment = Alignment(horizontal='center')
                planilha0.cell(1,nCol+1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

            for nLin in range(2,l_linha+1):
                planilha0.cell(nLin,5).number_format = "#,##0.00"
                planilha0.cell(nLin,6).number_format = "#,##0.00"
                planilha0.cell(nLin,7).number_format = "#,##0.00"
                planilha0.cell(nLin,8).number_format = "#,##0.00"

            log("Colocando bordas nas linhas e colunas ... ")  
            l_linhas_colunas0='A1:H'+str(planilha0.max_row)
            set_border_edsi(planilha0, l_linhas_colunas0, 'ffff')

            log("Congelando a primeira linha ... ")  
            planilha0.freeze_panes = 'A2'

            log("Salvando a planilha : " + str(l_relatorio))
            arquivo_excel.save(l_relatorio)

    except Exception as e:
        l_ret = 1    
        txt = traceback.format_exc()
        log("ERRO NA GERAÇÃO DO RELATÓRIO RESUMO.: " + txt +  " >> " +  str(e))

    try:
        if not l_ret and p_analitico == "S":
            l_linha:int = 0
            for dado in l_dados:
                l_linha += 1
                if l_linha == 1:
                    continue
                try:
                    l_serie = str(dado[0]).strip()
                    l_mes_ano = str(dado[1]).replace("/","").strip()
                    l_uf = str(dado[2]).strip().upper()
                    log("Relatorio : " + l_serie + " >> " + l_mes_ano + "  >> " + l_uf )
                    l_ret = processarRelatorioAnalitico(p_mes_ano=l_mes_ano, 
                                              p_serie=l_serie,
                                              p_uf=l_uf,
                                              p_relatorio=p_relatorio,
                                              p_conexao=p_conexao)    
                    if l_ret is None:
                        log("Falha! " + l_serie + " >> " + l_mes_ano + "  >> " + l_uf )
                        break
                    elif str(l_ret) != "0":
                        log("Falha! "  + l_serie + " >> " + l_mes_ano + "  >> " + l_uf)
                        break
                except:
                    pass
    except Exception as e:
        l_ret = 1    
        txt = traceback.format_exc()
        log("ERRO NA GERAÇÃO DO RELATÓRIO ANALITICO.: " + txt +  " >> " +  str(e))          
    return l_ret

"""
Retorna a validação de entrada e dos arquivos de configuração
"""
def validacaoEntrada():
    try:
        global gv_mes_ano_inicial
        global gv_mes_ano_final
        global gv_serie
        global gv_uf
        global gv_analitico
        global gv_usuario
        global gv_senha
        global gv_banco
        global gv_caminho
        global gv_relatorio
        global gv_resumo

        l_ret = 0
        comum.carregaConfiguracoes(configuracoes)
        comum.addParametro( 'MES_ANO_INICIAL', None, 'Mês e Ano Inicial - Formato MMYYYY', True, '122015')
        comum.addParametro( 'MES_ANO_FINAL', None, 'Mês e Ano Final - Formato MMYYYY', True, '122015')
        comum.addParametro( 'SERIE',  None, "Serie da NOTA", True, '01' )
        comum.addParametro( 'UF', None, 'Unidade Federativa do Brasil (Estado)', True, 'SP')
        comum.addParametro( 'ANALITICO', None, 'Relatório detalhado com os dados (S/N)', False, 'S')
        # Validacao dos parametros de entrada
        if not comum.validarParametros() :
            l_ret = 91
        else:
            gv_mes_ano_inicial = comum.getParametro('MES_ANO_INICIAL').upper().strip()
            gv_mes_ano_final = comum.getParametro('MES_ANO_FINAL').upper().strip()
            gv_serie = comum.getParametro('SERIE').upper().strip()
            gv_uf = comum.getParametro('UF').upper().strip()
            gv_analitico = comum.getParametro('ANALITICO').upper().strip()

            if not l_ret :
                try:
                    if (len(gv_mes_ano_inicial) != 6):
                        log("PARAMETRO MES e ANO [INICIAL]: Invalido! " + gv_mes_ano_inicial) 
                        l_ret = 91           
                    else:
                        if (
                           int(gv_mes_ano_inicial[0:2]) > 12
                        or int(gv_mes_ano_inicial[0:2]) < 1
                        ):
                            log("PARAMETRO MES [INICIAL] : Invalido! " + gv_mes_ano_inicial[0:2]) 
                            l_ret = 91                         
                        elif (
                           int(gv_mes_ano_inicial[2:6]) > datetime.datetime.now().year
                        or int(gv_mes_ano_inicial[2:6]) < (datetime.datetime.now().year)-50
                        ):
                            log("PARAMETRO ANO [INICIAL] : Invalido! " + gv_mes_ano_inicial[2:6]) 
                            l_ret = 91
                except:
                    log("PARAMETRO MES e ANO [INICIAL]: Invalido! " + gv_mes_ano_inicial) 
                    l_ret = 91
            
            if not l_ret :
                try:
                    if (len(gv_mes_ano_final) != 6):
                        log("PARAMETRO MES e ANO [FINAL]: Invalido! " + gv_mes_ano_final) 
                        l_ret = 91           
                    else:
                        if (
                           int(gv_mes_ano_final[0:2]) > 12
                        or int(gv_mes_ano_final[0:2]) < 1
                        ):
                            log("PARAMETRO MES [FINAL] : Invalido! " + gv_mes_ano_final[0:2]) 
                            l_ret = 91                         
                        elif (
                           int(gv_mes_ano_final[2:6]) > datetime.datetime.now().year
                        or int(gv_mes_ano_final[2:6]) < (datetime.datetime.now().year)-50
                        ):
                            log("PARAMETRO ANO [FINAL] : Invalido! " + gv_mes_ano_final[2:6]) 
                            l_ret = 91
                except:
                    log("PARAMETRO MES e ANO [FINAL]: Invalido! " + gv_mes_ano_final) 
                    l_ret = 91
            if not l_ret :
                try:
                    l_ano_mes_inicial = str(gv_mes_ano_inicial[2:6])+ str(gv_mes_ano_inicial[0:2])
                    l_ano_mes_final   = str(gv_mes_ano_final[2:6])  + str(gv_mes_ano_final[0:2])    
                    if int(l_ano_mes_final) < int(l_ano_mes_inicial):
                        log("PARAMETRO ANO e MES [FINAL] : Invalido! NÃO PODE SER MENOR QUE O INICIAL !") 
                        l_ret = 91                         
                except:
                    log("PARAMETRO MES e ANO [FINAL]: Invalido! " + gv_mes_ano_final) 
                    l_ret = 91            

            if not l_ret :
                try:
                    if (len(str(gv_serie).strip()) == 0):
                        log("PARAMETRO SERIE: Invalido! " + gv_serie) 
                        l_ret = 91       
                except:
                    log("PARAMETRO SERIE: Invalido! " + gv_serie) 
                    l_ret = 91

            if not l_ret :
                try:
                    if (len(str(gv_uf).strip()) != 2):
                        log("PARAMETRO UF: Invalido! " + gv_uf) 
                        l_ret = 91       
                except:
                    log("PARAMETRO UF: Invalido! " + gv_uf) 
                    l_ret = 91         

            if not l_ret :
                try:
                    if gv_analitico is not None:
                        if (len(str(gv_analitico).strip()) > 0):
                            if gv_analitico not in ("S","N"):
                                log("PARAMETRO ANALITICO: Invalido! " + gv_analitico) 
                                l_ret = 91
                        else:
                            gv_analitico = "N"              
                    else:
                        gv_analitico = "N"      
                except:
                    log("PARAMETRO UF: Invalido! " + gv_uf) 
                    l_ret = 91                                    
        # Verifica os parametros
        if not l_ret :
            try:
                gv_caminho   = configuracoes.caminho
                gv_relatorio = configuracoes.relatorio.replace("<<UF>>",gv_uf).replace("<<SERIE>>",gv_serie).replace("<<DATA_HORA>>",datetime.datetime.now().strftime('%Y%m%d_%H%M%S')).replace("<<MES_ANO_INICIAL>>",str(gv_mes_ano_inicial[2:6])+ str(gv_mes_ano_inicial[0:2])).replace("<<MES_ANO_FINAL>>",str(gv_mes_ano_final[2:6])+ str(gv_mes_ano_final[0:2])).replace(" ","").strip()
                gv_resumo    = configuracoes.resumo.replace("<<UF>>",gv_uf).replace("<<SERIE>>",gv_serie).replace("<<DATA_HORA>>",datetime.datetime.now().strftime('%Y%m%d_%H%M%S')).replace("<<MES_ANO_INICIAL>>",str(gv_mes_ano_inicial[2:6])+ str(gv_mes_ano_inicial[0:2])).replace("<<MES_ANO_FINAL>>",str(gv_mes_ano_final[2:6])+ str(gv_mes_ano_final[0:2])).replace(" ","").strip()
                if not l_ret :
                    if (len(gv_relatorio) <= 5
                        or not gv_relatorio.upper().endswith(".XLSX")
                        ):
                        log("PARAMETRO DO ARQUIVO  RELATORIO: INVALIDO! " + gv_relatorio) 
                        l_ret = 91
                    else:
                        log("Arquivo do relatorio : " + gv_relatorio)
                if not l_ret :
                    if (len(gv_resumo) <= 5
                        or not gv_resumo.upper().endswith(".XLSX")
                        ):
                        log("PARAMETRO DO ARQUIVO  RESUMO: INVALIDO! " + gv_resumo) 
                        l_ret = 91
                    else:
                        log("Arquivo do resumo : " + gv_resumo)
                        
                if not l_ret :
                    try:
                        if not os.path.isdir(gv_caminho):
                            log("Diretório não existente : " + gv_caminho)        
                            os.makedirs(gv_caminho)
                            log("Diretório criado : " + gv_caminho)
                        gv_resumo = os.path.join(gv_caminho, gv_resumo)
                        if os.path.isfile(gv_resumo):
                            log("Arquivo já existente !!!  " + gv_resumo)        
                            l_ret = 91
                        gv_relatorio = os.path.join(gv_caminho, gv_relatorio)
                    except Exception as e:
                        txt = traceback.format_exc()
                        log(gv_caminho + " >> PARAMETRO DO RELATÓRIO " + gv_resumo +  " INVÁLIDO! " + str(e) + " >> " + txt) 
                        l_ret = 91              
            except Exception as e:
                txt = traceback.format_exc()
                log("PARAMETRO DOS ARQUIVOS NAO ENCONTRADOS! " + str(e)+ " >> " + txt) 
                l_ret = 91        
        return l_ret
    except Exception as e:
        txt = traceback.format_exc()
        log("ERRO VALIDAÇÃO DOS PARAMETROS DE ENTRADA: " + str(e)+ " >> " + txt)
        l_ret = 93
        return l_ret

if __name__ == "__main__" :
    ret = 0
    txt = ''
    gv_conexao = None
    gv_mes_ano_inicial = ""
    gv_mes_ano_final = ""
    gv_serie = ""
    gv_uf = ""
    gv_analitico = "N"
    gv_usuario = ""
    gv_senha = ""
    gv_banco = ""
    gv_caminho=""
    gv_relatorio=""
    gv_resumo=""

    try:
        log("\n")
        if not ret :
            ret = validacaoEntrada()
        log("\n")
        # Processar o relatório
        if not ret :            
            ret = processarRelatorio(p_mes_ano_inicial=gv_mes_ano_inicial,
                       p_mes_ano_final=gv_mes_ano_final,  
                       p_serie=gv_serie,
                       p_uf=gv_uf,
                       p_relatorio=gv_relatorio,
                       p_resumo=gv_resumo,
                       p_conexao=gv_conexao,
                       p_analitico=gv_analitico)
            
        log("\n")
        if not ret :
            log("***SUCESSO***")
        else:
            log("<<<ERRO>>>")
        log("\n")
    except Exception as e:
        txt = traceback.format_exc()
        log("ERRO .: " + str(e) + " >> " + txt)
        ret = 93
    sys.exit(ret if ret >= log.ret else log.ret )