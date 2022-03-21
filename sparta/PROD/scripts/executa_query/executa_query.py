#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-
"""
----------------------------------------------------------------------------------------------
MODULO ...: TESHUVA
SCRIPT ...: loaderRelatorioNotasNaoEncontradasArquivoImpressao.py
CRIACAO ..: 14/07/2021
AUTOR ....: EDUARDO DA SILVA FERREIRA / KYROS TECNOLOGIA
            eduardof@kyros.com.br
DESCRICAO.: Geração de relatórios NOTAS não encontradas nos 
            arquivos de impressão 
----------------------------------------------------------------------------------------------
  HISTORICO : 
        Adequação para novo formato de script 
        SCRIPT ......: loader_sped_registro_O150.py
        AUTOR .......: Victor Santos
----------------------------------------------------------------------------------------------
"""
import os
import sys

global SD, dir_base
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)

import configuracoes
import agente_sparta
import comum
import sql
from openpyxl import Workbook, load_workbook
import openpyxl
import datetime
import shutil
from pathlib import Path
from openpyxl.utils import get_column_letter


NoneType = type(None)
StrType = type("string")
IntType = type(88)
FloatType = type(88.00)
DateType = type(datetime.datetime.now())
msize = 1.1
SD = ('/' if os.name == 'posix' else '\\')

log.gerar_log_em_arquivo = True
comum.carregaConfiguracoes(configuracoes)
con=sql.geraCnxBD(configuracoes)

def retornaQuery():

    ips = agente_sparta.retornaIpsLocal()

    ip = ips['bond0']
    port = 65502

    log('IP:...............', ip)
    log('PORTA:............', port)

    dados_server = { 'HOST' : ip, 'PORT': port }
    msg = agente_sparta.execFncXmlRpc( dados_server, 'buscaParametro', configuracoes.id_exec, 'SQL' )
    return msg['SQL']

def datavalida(valor):
    try:
        datetime.datetime.strptime(valor, '%d/%m/%Y')
        return True
    except:
        return False
        
def formataAba(aba, tipocampo):

    size = []
    for col in aba[1]:
        v_col = col.value
        size.append((len(v_col)*msize)+4)     
    
    col_idx = 0
    lin_idx = 0

    for lin in aba:
        col_idx  = 0
        lin_idx += 1
        for col in lin:
            col_idx += 1
            campo = col.value
            tamcampo = 1
            
            if (tipocampo == NoneType or campo == None or campo == 'None'):
                campo = ""
            
            if (tipocampo == DateType):
                log(campo) 
                campo = campo.strftime('%d/%m/%Y, %H%M%S')
                    
            tamcampo = len(str(campo))
            if (tipocampo != IntType and tipocampo != FloatType):   
                campo = str(campo)

            tipocampo = type(campo)
            letra = get_column_letter(col_idx)

            aba.cell(row=lin_idx, column=col_idx, value=campo)
            if( type(campo) == StrType ):
                aba.cell(row=lin_idx, column=col_idx).number_format = '@'
                tamcampo = len(campo)

            tamaju = (tamcampo*msize) + 4
                
            if ( tamaju > size[col_idx-1] ):
                size[col_idx-1] = tamaju
            
            aba.column_dimensions[letra].width = size[col_idx-1]
    return

def valida_nome_arq(string_nome_arquivo):
    carac_invalidos = [ '.', ',', '-', ';', '@', '!', '?']
    arquivo_ok = True
    for c in string_nome_arquivo :
        if c in carac_invalidos :
            arquivo_ok = False
    return arquivo_ok

def valida_sql(sql):
    carac_invalidos = [ 'DECLARE', 'BEGIN', 'END', 'TRUNCATE', 'CREATE', 'DROP', 'GRANT' ]
    sql_ok = True
    for c in sql :
        if c in carac_invalidos :
            sql_ok = False
    return sql_ok

def processar():
    
    ret = 0
    now = datetime.datetime.now()
    dt_string = now.strftime("%d-%m-%Y %H:%M:%S")
    valida_nome = valida_nome_arq(configuracoes.nome_arq)
    dir_base = SD + 'portaloptrib' + SD + 'TESHUVA' + SD + 'sparta' + SD +  'DEV' + SD + 'scripts' + SD + 'Tecnico' + SD + 'execucao' + SD + 'executa_query' + SD

#### Monta caminho e nome do destino
    if not os.path.isdir(dir_base) :
        os.makedirs(dir_base) 
    
    arquivo_destino = configuracoes.nome_arq + '_' + str(dt_string) + '.xlsx'
    nome_relatorio = os.path.join(dir_base,arquivo_destino)
    # arquivo = open(nome_relatorio, 'w')
    # arquivo.close() 

    if valida_nome == False:
       log('ERRO, COLOQUE UM NOME DE ARQUIVO SEM CARACTERES ESPECIAIS.')
       ret = 99
       return ret 

    query_exec = retornaQuery()

    query      = query_exec.upper()
   
    valida     = valida_sql(query)

    if valida == False:
        log('ERRO, SQL INVALIDO, INSTRUÇÕES QUE COMEÇAM COM DECLARE, BEGIN, END, TRUNCATE, CREATE, DROP, GRANT, -, ;, / NÃO SÃO PERMITIDAS.')
        ret = 99
        return ret 
    
    if query.split()[0] != configuracoes.tipo:
        log('ERRO, O TIPO SELECIONADO NÃO CONFERE COM A INSTRUÇÃO SOLICITADA...')
        ret = 99
        return ret
    else:
        if query.startswith('UPDATE') or query.startswith('DELETE') or query.startswith('SELECT') or query.startswith('INSERT') or query.startswith('WITH'):
            if query.startswith('SELECT') or query.startswith('WITH'):
                
                if not query.__contains__('WHERE'):
                    log('ERRO, CONFIRA SUA QUERY, NÃO FOI ENCONTRADA A CLÁUSULA WHERE E ISTO GERA RISCO AOS DADOS. \nCASO QUEIRA EXECUTAR, ADICIONE WHERE 1 = 1') #mudar frase
                    ret=99
                    return ret

                log('Executando SELECT, aguarde...')
                log('***QUERY***', query_exec)
                retorno   = []
                cabecalho = []
                tipo      = []
                con.executa(query_exec)
                colunas = con.description()
                result  = con.fetchone()
                count   = 0
                if not result:
                    log("#### ATENÇÃO: Nenhum Resultado para query")
                    log("####     Query = ")
                    log("####")
                    log(query)
                    log("####")
                    ret=99
                    return ret
                else:                    
                    try:     
                        for col in colunas:
                            cabecalho.append(col[0])
                            tipo.append(str(col[1]).split('.')[1].split("'")[0])
                        retorno.append(cabecalho)
                        while result:
                            retorno.append(result)
                            result = con.fetchone()
                            count += 1
                    except Exception as e:
                        log('ERRO - ' , e)
                        ret=99

                arq_excel = Workbook()
                aba0 = arq_excel.active
                aba0.title = 'RESULTADO'
                lin_num = 0
                for col in retorno:
                    lin_num += 1
                    col_num = 0
                    for cel in col:
                        col_num += 1
                        aba0.cell(row=lin_num, column=col_num, value=cel)

                formataAba(aba0, tipo)
                if configuracoes.nome_arq:                    
                    nome_rel = dir_base + configuracoes.nome_arq + '_' + str(dt_string) + '.xlsx'
                    arq_excel.save(nome_relatorio)
                else:
                    nome_rel = dir_base + 'SAIDA_' + dt_string + '.xlsx' 
                    # log('') -- printar somente no log
                
                log('Quantidade de linhas: ', count)
                log('Arquivo de saída:     ', nome_rel)
                log('SUCESSO')

            if query.startswith('UPDATE') or query.startswith('DELETE'):
                if not query.__contains__('WHERE'):
                    log('ERRO, CONFIRA SUA QUERY, NÃO FOI ENCONTRADA A CLÁUSULA WHERE E ISTO GERA RISCO AOS DADOS. \nCASO QUEIRA EXECUTAR, ADICIONE WHERE 1 = 1')
                    ret = 1
                    return ret
                else:
                    try:
                        log('Executando instrução, aguarde...')
                        log('***QUERY***', query_exec)
                        con.executa(query_exec)
                        con.commit()
                        log('Quantidade de linhas alteradas...:', str(con.rowcount()))
                        log('SUCESSO')                        

                    except Exception as e:
                        log('ERRO - ' , e)
                        con.rollback()
                        ret=99
                        return ret

            if query.startswith('INSERT'):
                if query.__contains__('SELECT') and not query.__contains__('WHERE'):
                    log('COMANDO INVÁLIDO PARA INSERÇÃO! ISTO GERA RISCO AOS DADOS, CASO QUEIRA EXECUTAR, ADICIONE CRITÉRIO FALSO COMO WHERE 1 = 1')
                    ret=99
                    return ret
                else:
                    try:
                        log('Executando instrução, aguarde...')
                        log('***QUERY***', query_exec)
                        con.executa(query_exec)
                        con.commit()
                        log('Quantidade de linhas alteradas...:', str(con.rowcount()))
                        log('SUCESSO')
                    except Exception as e:
                        log('ERRO - ' , e)
                        con.rollback()
                        ret=99
                        return ret    
        else:
            log('ERRO, CONFIRA SUA QUERY, NÃO FOI ENCONTRADA A CLÁUSULA ( SELECT OU UPDDATE OU INSERT OU DELETE OU WITH ) NO INÍCIO DA INSTRUÇÃO.')
            ret  = 99
            return ret         

    return ret

if __name__ == "__main__" :
    
    ret = 0
    
    comum.addParametro( 'ID_EXEC' , None, "ID DA EXECUÇÃO (1234)"         , True , '1234' )
    comum.addParametro( 'TIPO'    , None, "SELECT, UPDATE, INSERT, DELETE", True , 'SELECT')
    comum.addParametro( 'NOME_ARQ', None, 'NOME DO ARQUIVO DE SAÍDA'      , False, 'sucesso.txt')

    if not comum.validarParametros() :
        log('ERRO AO VALIDAR OS PARAMETROS')
        ret = 91
    else:
        configuracoes.id_exec  = int(comum.getParametro('ID_EXEC'))
        configuracoes.tipo     =     comum.getParametro('TIPO').upper()
        configuracoes.nome_arq =     comum.getParametro('NOME_ARQ')    

        ret = processar()            
    sys.exit(ret)