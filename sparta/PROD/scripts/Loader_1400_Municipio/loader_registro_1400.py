#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-
"""
----------------------------------------------------------------------------------------------
  MODULO ...: TESHUVA
  SCRIPT ...: loader_1400_municipio.py
  CRIACAO ..: 26/05/2021
  AUTOR ....: Victor Santos / KYROS TECNOLOGIA
  DESCRICAO : 

----------------------------------------------------------------------------------------------
  HISTORICO : 
          * 02/06/2021 - AIRTON BORGES DA SILVA FILHO - Kyros Tecnologia
      - Incluido a possibilidade de carga para outros estados em tabelas diferentes
        Documentacao : 01 - Teshuva_RMSV0_Registro 1400 SP_V2.docx
    * 14/03/2022 - Eduardo da Silva Ferreira - Kyros Tecnologia
                 - [PTITES-1698] Padrão de diretórios do SPARTA                    

----------------------------------------------------------------------------------------------
"""

import os
import sys

global SD, dir_base
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)
import configuracoes

import datetime
import cx_Oracle
import shutil
from pathlib import Path
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

import comum
import sql
import layout
import util

log.gerar_log_em_arquivo = True

ufi = ""

def dtf():
    return (datetime.datetime.now().strftime('%Y%m%d%H%M%S'))

def lerXlsx(diretorio, arquivo) :
    global ufi
    log('-'*120)
    ret = True
    path_arq = os.path.join( diretorio, arquivo )
    log('Processando o arquivo:', path_arq)
    
    xls = load_workbook(path_arq)
    p1 = xls.sheetnames[0]
    sh = xls[p1]
    nlinhas  = sh.max_row
    ncolunas = sh.max_column
    
    planilha = xls.active
    dados = []

    niei = ""
    iei  = ""
    iano = ""
    imes = ""

#### Captura a IE da planilha
    try:
        niei = planilha.cell(row=3, column=2).value.split(':')[0]
        iei  = int(planilha.cell(row=3, column=2).value.split(':')[1])
        nmesano = planilha.cell(row=4, column=2).value.split(':')[0]
        mesano  = planilha.cell(row=4, column=2).value.split(':')[1]

        f"{1:04d}"
        
        imes =  f"{int(mesano.split('/')[0]):02d}"
        iano =  f"{int(mesano.split('/')[1]):04d}"


        log("IE  = ", iei)
        log("MES = ", imes)
        log("ANO = ", iano)


    except Exception as e :
        log("#### - ERRO - Planilha com dados de entrada não está no formato padrão.")
        log("#### - ERRO obtino na tentativa de extrair os dados: ", str(e))   
        return 13

    if (( ncolunas != 5) or 
        ( nlinhas < 5) or 
        ( niei.upper() != "IE") or 
        ( nmesano.upper() != "MÊS/ANO") or 
        ( int(iano) > datetime.datetime.now().year) or
        ( int(iano) < ((datetime.datetime.now().year)-30)) or
        ( int(imes) < 1 ) or
        ( int(imes) > 12) ):
        log("ERRO - Formatação da planilha 'Parametro' não está no padrão correto.")
        return False

    r = 0
    for row in planilha.rows :
        dados.append([])
        for col in row :
            dados[r].append(col.value)
        r += 1 
    cabecalho = dados[4]
   
    dic_colunas = {}
    dic_colunas['MUNICIPIO'] = False
    dic_colunas['CODIGO MUNICIPIO (IBGE)'] = False
    dic_colunas['%RATEIO'] = False
    dic_colunas['CODIGO ITEM (TABELA PVA)'] = False

    colunas = dic_colunas.keys()
    
    for i in range(len(cabecalho)) :
        if (cabecalho[i] != None):
            if cabecalho[i].upper() in colunas :
                dic_colunas[cabecalho[i].upper()] = i+1
    
    for k in colunas :
        if not dic_colunas[k] :
            if (k != 'MUNICIPIO'):
                log('ERRO - Não encontrada a coluna < %s > no xlsx .. Verifique !'%(k))
                return 30

    con=sql.geraCnxBD(configuracoes)
    reg = 0
    
    datafim = datetime.datetime.now().strftime('%d/%m/%Y')

#### Determina a UF através da IE da planilha excell
    query="""
    select distinct f.unfe_sig  from openrisow.filial f where f.fili_cod_insest='%s'
    """%(iei)

    try :
        con.executa(query)
    except Exception as e :
        log(query)
        log(str(e))   
        return 31
        
    result = con.fetchone()
    ufi = ""
    if (result):    
        for campo in result:
            ufi = campo

    if (util.validauf(ufi)==False):
        log("ERRO - Não foi possível determinar a UF pela IE informada na planilha de dados.")
        return 33

    log("UF   =", ufi)

    if (ufi == 'RJ'):
        tabela = "gfcadastro.reg_1400_municipio"
    elif (ufi == 'SP'):
        tabela = "gfcadastro.reg_1400_municipio_sp"

#### Atualiza a data final de validade dos registros existentes.

    cmd_sql1=""" 
      update %s
      set data_fim = TO_DATE('%s', 'DD/MM/YYYY')
      where ano = '%s'
          and mes = '%s'
          and uf = '%s'
          """%(tabela,datafim,iano,imes,ufi)
    try : 
        con.executa(cmd_sql1)
    except Exception as e :
        log(cmd_sql1)
        log(str(e))
  
#### Insere os novos registros    
  
    for row in dados[5:] :
        municipio = row[1].strip()
        codigo_ibge = (0 if row[2] == None else row[2])
        rateio = row[3]
        codigo_pva = row[4]
        uf = ufi
        if municipio[:5].upper() not in ('','TOTAL', 'MUNIC'):    
            municipio = municipio.replace("'"," ")
            cmd_sql2=""" 
                insert into %s (
                                descricao,
                                uf,
                                codigo_ibge,
                                indice,
                                codigo_pva,
                                ano,
                                mes,
                                data_inicio)
                values ('%s','%s','%s','%s','%s','%s','%s',TO_DATE('%s','DD/MM/YYYY')) 
                """%(tabela,municipio,ufi,codigo_ibge,rateio,codigo_pva,iano,imes,datafim) 
            reg = reg + 1
            try : 
                con.executa(cmd_sql2)
            except Exception as e :
                log('Erro ao executar o Sql ...')
                log(str(e))
                log(cmd_sql2)
                ret = False
                break
            if reg % 1000 == 0 :
                log('- Processados %s registros'%(reg))

            
    log("Foram incluidos", reg, "novos registros.", ufi,imes,iano)        
    con.commit()
    
    return ret

def processar() :
    global ufi
    comum.carregaConfiguracoes(configuracoes)
    dir_arquivos = os.path.join(os.path.dirname(configuracoes.dir_entrada), 'REGISTRO_1400', 'municipio') # [PTITES-1698] # configuracoes.diretorio_arquivos
    diretorio_processados = os.path.join(os.path.dirname(configuracoes.dir_geracao_arquivos), 'REGISTRO_1400', 'municipio', 'processados') # [PTITES-1698] # configuracoes.diretorio_processados

    log("Verificando arquivos para processar no diretorio de entrada ...")
    log('- Diretorio : %s'%(dir_arquivos))
    if not os.path.isdir(dir_arquivos) :
        os.makedirs(dir_arquivos)
    
    if not os.path.isdir(diretorio_processados) :
        os.makedirs(diretorio_processados)
    
    qtd_arqs = 0
    ret = 99
    for arq in os.listdir(dir_arquivos) :
        ret = 0
        log(arq)
        if os.path.isfile( os.path.join(dir_arquivos, arq)) :
            if arq.endswith('.xlsx') :
                log('-'*120)
                qtd_arqs += 1
                log('Verificando o arquivo :',arq)
                
                if not lerXlsx(dir_arquivos, arq) :
                    ret = 99
                else :
                    vdataArq= dtf()
                    os.rename(os.path.join(dir_arquivos, arq),os.path.join(dir_arquivos, arq)+ufi+vdataArq)
                    shutil.move(os.path.join(dir_arquivos, arq)+ufi+vdataArq, diretorio_processados)
 
    if (qtd_arqs == 0) :
        log("ERRO - Nao foi encontrado nenhum arquivo xlsx a ser processado.")
        ret = 99

    log('-'*120)

    return ret

if __name__ == "__main__" :
    ret = 0
    txt = ''

    log('Processando ... ')

    ret = processar() 
    
    log("Codigo de retorno =", ret)
    if (ret != 0 ):
        log('ERRO no processamento.')
        ret = 92