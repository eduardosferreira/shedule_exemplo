#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-
"""
----------------------------------------------------------------------------------------------
MODULO ...: TESHUVA
SCRIPT ...: check_list_telecom.py
CRIACAO ..: 06/01/2022
AUTOR ....: EDUARDO DA SILVA FERREIRA 
            / KYROS TECNOLOGIA (eduardof@kyros.com.br)
DESCRICAO.: Este script possibilita a execução 
            dos objetos de banco (procedures) que 
-           executam check list do saneador de telecom, 
            através do painel de execuções.
----------------------------------------------------------------------------------------------

----------------------------------------------------------------------------------------------
  HISTORICO : 
  06/01/2022 : EDUARDO DA SILVA FERREIRA 
            / KYROS TECNOLOGIA (eduardof@kyros.com.br) 
        - Criação do Script.
                 
----------------------------------------------------------------------------------------------
"""
#### PATRONIZACAO PARA O PAINEL DE EXECUCOES....
import sys
import os
import datetime
gv_cc_sep_dir = ('/' if os.name == 'posix' else '\\')
gv_cc_dir_bse = os.path.join( \
    os.path.realpath('.').split( \
    gv_cc_sep_dir+'PROD'+gv_cc_sep_dir)[0], 'PROD') \
    if os.path.realpath('.').__contains__( \
        gv_cc_sep_dir+'PROD'+gv_cc_sep_dir) \
    else os.path.join( os.path.realpath('.').split( \
        gv_cc_sep_dir+'DEV'+gv_cc_sep_dir)[0], 'DEV')
sys.path.append(gv_cc_dir_bse)
# imports do sparta
import configuracoes
from comum import \
    log \
    , carregaConfiguracoes \
    , getParametro \
    , addParametro \
    , validarParametros        
import sql
configuracoes.gv_ds_log = " <||> "
log.gerar_log_em_arquivo = True
carregaConfiguracoes(configuracoes)

# demais imports
import traceback
from calendar import monthrange
import string
import random
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook

def __fnc_msg(p_cc_dado):
    """
        Funcao apenas para carregamento de msg
    """
    try:
        configuracoes.gv_ds_log =  " <||> " 
        + str(p_cc_dado) 
        + str(configuracoes.gv_ds_log)
    except:
        pass
    try:
        log(p_cc_dado)
    except:
        pass
        
def __fnc_tratar_entrada_dados( \
          p_nm_job = None \
        , p_cc_filtro_regra_execucao = None \
        , p_dt_filtro_inicio = None \
        , p_dt_filtro_fim = None \
        , p_cc_filtro_uf = None \
        , p_cc_filtro_filial = None \
        , p_cc_filtro_serie = None \
        , p_qt_filtro_relatorio = None \
        , p_tp_filtro_relatorio = None \
    ):
    """
        Funcao trata a entrada de dados
    """
    def fnc_get_parametro(p_nm_parametro):
        try:
            if getParametro(p_nm_parametro):
                return str(getParametro(p_nm_parametro))
            else:
                return ""
        except:
            return ""
    
    def fnc_valida_data(p_dt_data):
        try:
            v_dt_data = str(p_dt_data).replace("/", "").strip()

            if len(v_dt_data) != 8:
                __fnc_msg("ERRO >> PARAMETRO DATA: Inválido! " \
                + str(p_dt_data))
                return None, None

            v_dt_data = str(v_dt_data[0:2]) \
            + "/" + str(v_dt_data[2:4]) \
            + "/" + str(v_dt_data[4:8])
            
            if (int(v_dt_data[0:2]) > 31 \
                or int(v_dt_data[0:2]) < 1): 
                __fnc_msg("PARAMETRO DATA [DIA]: Inválido! " \
                + str(v_dt_data))
                return None, None
            


            if (int(v_dt_data[3:5]) > 12 \
                or int(v_dt_data[3:5]) < 1): 
                __fnc_msg("PARAMETRO DATA [MES]: Inválido! " \
                + str(v_dt_data))
                return None, None
                
            v_dt_aux = datetime.datetime(int(v_dt_data[6:10]) \
                                       , int(v_dt_data[3:5]) \
                                       , int(v_dt_data[0:2]))   
  
            v_dt_yyyymmdd = (str(v_dt_data[6:10]) \
                            + str(v_dt_data[3:5]) \
                            + str(v_dt_data[0:2]))
                                       
            return v_dt_aux, v_dt_yyyymmdd
        except:
            __fnc_msg("PARAMETRO DATA: Inválido! " + str(p_dt_data))
            return None, None
    
    configuracoes.nm_job += "" if not p_nm_job else str(p_nm_job)
    log("-"*150)
    log("Tratar a entrada de dados")
    if not p_dt_filtro_inicio \
    or not p_dt_filtro_fim \
    or not p_cc_filtro_uf \
    or not p_cc_filtro_filial \
    or not p_cc_filtro_serie \
    or not p_cc_filtro_regra_execucao \
    or not p_qt_filtro_relatorio \
    or not p_tp_filtro_relatorio \
    :
        
        addParametro('P_CC_FILTRO_REGRA_EXECUCAO', None, \
        'Regra de execução. '\
        , True,'TODOS')
        addParametro('P_DT_FILTRO_INICIO', None, \
        'Periodo inicio para realizacao da pesquisa no banco. ' \
        + 'FORMATO (MMYYYY).', True,'012015')        
        addParametro('P_DT_FILTRO_FIM', None, \
        'Periodo fim para realizacao da pesquisa no banco.' \
        + 'FORMATO (MMYYYY).', True,'122015')
        addParametro('P_CC_FILTRO_UF', None, \
        'Unidade Federativa do Brasil. Aceita apenas um valor.', True,'SP')
        addParametro('P_CC_FILTRO_FILIAL', None, \
        'Codigo da Filial, campo para pesquisa no banco. ' \
        + 'Aceita mais de um valor separado por virgula.' \
            , False,'0001, 0002')
        addParametro('P_CC_FILTRO_SERIE', None, \
        'Código da Série da NF, campo para pesquisa no banco. ' \
        + 'Aceita mais de um valor separado por virgula..' \
            , False,'U  T, 06, 1')
        addParametro('P_TP_FILTRO_RELATORIO', None, \
        'Tipo de Relatório (R-RESUMO ou D-DETALHADO). ' \
            , False,'R')
        addParametro('P_QT_FILTRO_RELATORIO', None, \
        'Quantidade de registros ' , False,'1')    
        # Validacao dos parametros de entrada
        if not validarParametros() :
            return 1
        else:        
            configuracoes.filtro_regra_execucao = \
            fnc_get_parametro('P_CC_FILTRO_REGRA_EXECUCAO').strip().upper()            
            configuracoes.filtro_dt_inicio = \
            "01" + fnc_get_parametro('P_DT_FILTRO_INICIO').strip().upper()
            configuracoes.filtro_dt_fim = \
            "01" + fnc_get_parametro('P_DT_FILTRO_FIM').strip().upper()
            configuracoes.filtro_uf = \
            fnc_get_parametro('P_CC_FILTRO_UF').strip().upper()
            configuracoes.filtro_filial = \
            fnc_get_parametro('P_CC_FILTRO_FILIAL').strip().upper()            
            configuracoes.filtro_serie = \
            fnc_get_parametro('P_CC_FILTRO_SERIE').strip().upper()
            configuracoes.filtro_qt_relatorio=\
            fnc_get_parametro('P_QT_FILTRO_RELATORIO').strip().upper()
            configuracoes.filtro_relatorio=\
            fnc_get_parametro('P_TP_FILTRO_RELATORIO').strip().upper()
            try:
                if not p_nm_job:
                    configuracoes.nm_job = sys.argv[0].upper()
            except:
                pass
    else:
        configuracoes.filtro_dt_inicio = \
        "01" + str(p_dt_filtro_inicio).strip().upper()
        configuracoes.filtro_dt_fim = \
        "01" + str(p_dt_filtro_fim).strip().upper()
        configuracoes.filtro_uf = \
        str(p_cc_filtro_uf).strip().upper()
        configuracoes.filtro_filial = \
        str(p_cc_filtro_filial).strip().upper()          
        configuracoes.filtro_serie = \
        str(p_cc_filtro_serie).strip().upper()
        configuracoes.filtro_regra_execucao = \
        str(p_cc_filtro_regra_execucao).strip().upper()
        configuracoes.filtro_qt_relatorio=\
        str(p_qt_filtro_relatorio).strip().upper()    
        configuracoes.filtro_relatorio = \
        str(p_tp_filtro_relatorio).strip().upper()

    configuracoes.dt_inicio, configuracoes.dt_inicio_yyyymmdd = \
    fnc_valida_data(configuracoes.filtro_dt_inicio)    
    if not configuracoes.dt_inicio: 
        return 1
        
    configuracoes.dt_fim, configuracoes.dt_fim_yyyymmdd = \
    fnc_valida_data(configuracoes.filtro_dt_fim)    
    if not configuracoes.dt_fim: 
        return 1
 
    if int(configuracoes.dt_fim_yyyymmdd) \
     < int(configuracoes.dt_inicio_yyyymmdd):
        __fnc_msg("PARAMETRO DATA FINAL: Inválido! ' \
        + 'A data fim não pode ser menor que a data início.")
        return 1  

    if abs((configuracoes.dt_fim\
        -configuracoes.dt_inicio).days) > 366:
        __fnc_msg("PARAMETRO DATA INICIO/FINAL: Inválido! ' \
        + 'A datas não podem ter diferença 366 dias.")
        return 1  
    
    
    if (not configuracoes.filtro_qt_relatorio\
    and str(configuracoes.filtro_qt_relatorio).strip() != "0")\
    or not configuracoes.filtro_qt_relatorio.isnumeric()\
    or configuracoes.filtro_relatorio == "R":
        configuracoes.filtro_qt_relatorio = "0"
    else:
        try:
            if int(configuracoes.filtro_qt_relatorio) < 1:
                __fnc_msg("PARAMETRO QUANTIDADE de RELATÓRIO: Inválido! Valor deve ser maior que zero (0) ")
                return 1
        except:
            __fnc_msg("PARAMETRO QUANTIDADE de RELATÓRIO: Inválido! ")
            return 1  
    
    if not configuracoes.filtro_relatorio:
        configuracoes.filtro_relatorio = "R"
    
    if configuracoes.filtro_relatorio not in ('R','D'):
        __fnc_msg("PARAMETRO TIPO de RELATÓRIO: Inválido! ")
        return 1  

    log("-"*150)
    return 0

def __fnc_tratar_parametro_configuracoes():
    """
        Funcao trata arquivo de configuracoes
    """    
    v_nr_ret = 0
    configuracoes.var_owner_tab_openrisow  = "OPENRISOW."
    configuracoes.var_owner_tab_gfcadastro = "GFCADASTRO."
    configuracoes.var_owner_tab_gfcarga    = "GFCARGA."     
    configuracoes.var_diretorio_relatorio = ""
    configuracoes.var_mascarara_arquivo_diretorio    = ""         
    
    if not v_nr_ret:
        try:
            owner_tab_openrisow = getattr(configuracoes, 'owner_openrisow' \
                , configuracoes.var_owner_tab_openrisow)
            if (len(str(owner_tab_openrisow).strip()) > 0):
                configuracoes.var_owner_tab_openrisow = owner_tab_openrisow
        except:
            __fnc_msg("ERRO >> PARAMETRO OWNER_OPENRISOW: Inválido! ")
            v_nr_ret = 1 
    
    if not v_nr_ret:
        try:
            owner_tab_gfcadastro = getattr(configuracoes, 'owner_gfcadastro' \
                , configuracoes.var_owner_tab_gfcadastro)
            if (len(str(owner_tab_gfcadastro).strip()) > 0):
                configuracoes.var_owner_tab_gfcadastro = owner_tab_gfcadastro
        except:
            __fnc_msg("ERRO >> PARAMETRO OWNER_GFCADASTRO: Inválido! ")
            v_nr_ret = 1
    
    if not v_nr_ret:
        try:
            owner_tab_gfcarga = getattr(configuracoes, 'owner_gfcarga' \
                , configuracoes.var_owner_tab_gfcarga)
            if (len(str(owner_tab_gfcarga).strip()) > 0):
                configuracoes.var_owner_tab_gfcarga = owner_tab_gfcarga
        except:
            __fnc_msg("ERRO >> PARAMETRO OWNER_GFCARGA: Inválido! ")
            v_nr_ret = 1

    #if not v_nr_ret:
    #    try:
    #        var_diretorio_relatorio = getattr(configuracoes, 'diretorio_relatorio' \
    #            , configuracoes.var_diretorio_relatorio)
    #        if (len(str(var_diretorio_relatorio).strip()) > 0):
    #            configuracoes.var_diretorio_relatorio = var_diretorio_relatorio
    #            if not os.path.isdir(configuracoes.var_diretorio_relatorio) :
    #                os.makedirs(configuracoes.var_diretorio_relatorio)
    #        else:                
    #            __fnc_msg("ERRO >> PARAMETRO DIRETORIO_RELATORIO: Inexistente! ")
    #            v_nr_ret = 1
    #    except:
    #        __fnc_msg("ERRO >> PARAMETRO DIRETORIO_RELATORIO: Inválido! ")
    #        v_nr_ret = 1   
    #
    #if not v_nr_ret:
    #    try:
    #        var_mascarara_arquivo_diretorio = getattr(configuracoes, 'mascarara_arquivo_diretorio' \
    #            , configuracoes.var_mascarara_arquivo_diretorio)
    #        if (len(str(var_mascarara_arquivo_diretorio).strip()) > 0):
    #            configuracoes.var_mascarara_arquivo_diretorio = var_mascarara_arquivo_diretorio
    #        else:                
    #            __fnc_msg("ERRO >> PARAMETRO MASCARA_ARQUIVO: Inexistente! ")
    #            v_nr_ret = 1
    #    except:
    #        __fnc_msg("ERRO >> PARAMETRO MASCARA_ARQUIVO: Inválido! ")
    #        v_nr_ret = 1

    return v_nr_ret

def __fnc_processar():
    """
        Funcao trata o processamento
    """     
    v_nr_ret = 0  
    try:   
        configuracoes.lista_id_exec = list()
        configuracoes.id_exec = ""
        dt_inicio = configuracoes.dt_inicio        
        log("inicio: " + str(configuracoes.dt_inicio))
        log("fim: " + str(configuracoes.dt_fim))  
        while dt_inicio <= configuracoes.dt_fim:
            nr_mes = int(monthrange(dt_inicio.year,dt_inicio.month)[1])
            dt_fim = dt_inicio.replace(day=nr_mes)
            # Busca ID DE EXECUCAO
            configuracoes.id_exec = ""            
            while not configuracoes.id_exec:
                configuracoes.id_exec =\
                configuracoes.dt_atual_banco_dados_texto + ''.join(random.SystemRandom()\
                .choice(string.ascii_letters + string.digits)\
                for _ in range(4))

                log("id_execucao >> ",configuracoes.id_exec)
                v_ds_sql = \
                """ 
                SELECT 1 AS CAMPO FROM DUAL 
                WHERE NOT EXISTS (SELECT 1 
                    FROM GFCADASTRO.TSH_SAN_CONTROLE
                    WHERE ID_EXEC = '%s')
                """%(configuracoes.id_exec)      
                
                configuracoes.connection.executa(v_ds_sql)
                configuracoes.cursor = \
                configuracoes.connection.fetchone()        
                if not (configuracoes.cursor):
                    configuracoes.id_exec = ""

            log("id_exec [tsh_san_controle]: " + str(configuracoes.id_exec))                            
            v_ds_job = configuracoes.nm_job + "_"\
                 + configuracoes.dt_atual_banco_dados_texto + " "\
                 + "(" + configuracoes.id_exec + ")"

            v_ob_dicionario_controle = dict()
            v_ob_dicionario_controle['ID_EXEC'] = configuracoes.id_exec
            v_ob_dicionario_controle['NOME_JOB'] = configuracoes.nm_job + "_"\
                 + configuracoes.dt_atual_banco_dados_texto + " "
    
            log("job [tsh_san_controle]: " + v_ds_job)
            log("filtro_dt_inicio: " + str(configuracoes.filtro_dt_inicio))
            log("filtro_dt_fim: " + str(configuracoes.filtro_dt_fim))
            log("filtro_uf: " + str(configuracoes.filtro_uf))
            log("filtro_filial: " + str(configuracoes.filtro_filial))
            log("filtro_serie: " + str(configuracoes.filtro_serie))
            log("filtro_regra_execucao: " + str(configuracoes.filtro_regra_execucao))
            log("filtro_relatorio: " + str(configuracoes.filtro_relatorio))
            log("filtro_qt_relatorio: " + str(configuracoes.filtro_qt_relatorio))            
            log("data_inicio: " + str(dt_inicio))
            log("data_fim: " + str(dt_fim))  
            
            configuracoes.ds_erro = configuracoes.connection.var(str)
            configuracoes.cd_erro = configuracoes.connection.var(int)
            configuracoes.connection.executaProcedure("GFCADASTRO.TSH_SANTL_40230_RELATORIO",
                                                P_NM_JOB=v_ds_job,
                                                P_UF_FILTRO=\
                                                configuracoes.filtro_uf,
                                                P_DT_INI=dt_inicio,
                                                P_DT_FIM=dt_fim,
                                                P_CD_ERRO=\
                                                configuracoes.cd_erro,
                                                P_DS_ERRO=\
                                                configuracoes.ds_erro,
                                                P_CC_FILI_COD=\
                                                configuracoes.filtro_filial,
                                                P_CC_SERIE=\
                                                configuracoes.filtro_serie,
                                                P_CC_REGRA_EXECUCAO=\
                                                configuracoes.filtro_regra_execucao,
                                                P_CC_OWNER_OPENRISOW=\
                                                configuracoes.var_owner_tab_openrisow,
                                                P_TP_RELATORIO=\
                                                configuracoes.filtro_relatorio,
                                                P_QT_RELATORIO=\
                                                int(configuracoes.filtro_qt_relatorio)                                                                                           
                                                )
            v_ob_dicionario_controle['CODIGO_RETORNO']=None
            v_ob_dicionario_controle['RETORNO']=None    
            try:
                log("")
                log("")
                log('*'*100)      
                v_ob_dicionario_controle['CODIGO_RETORNO']=\
                configuracoes.cd_erro.getvalue()   
                v_ob_dicionario_controle['RETORNO']=\
                str(configuracoes.ds_erro.getvalue())
                log("CODIGO RETORNO >> " \
                    + str(configuracoes.cd_erro.getvalue()))   
                if configuracoes.cd_erro.getvalue():
                    v_nr_ret = 1
                    log("")
                    v_ob_dicionario_controle['RETORNO']=\
                    "ERRO NESTA EXECUÇÃO. " \
                            + " EXISTE A POSSIBILIDADE DE NÃO EXISTIR DADOS " \
                            + " NA TABELA ou RELATÓRIO PARA ESTA EXECUÇÃO. "\
                            + str(configuracoes.ds_erro.getvalue()).upper()
                    __fnc_msg("ERRO NESTA EXECUÇÃO. " \
                            + " EXISTE A POSSIBILIDADE DE NÃO EXISTIR DADOS " \
                            + " NA TABELA ou RELATÓRIO PARA ESTA EXECUÇÃO. "\
                            + str(configuracoes.ds_erro.getvalue()))     
                    log("")                    
                else:
                    log("")
                    v_ob_dicionario_controle['RETORNO']=\
                    "SUCESSO PARA ESTA EXECUÇÃO! SEGUE O RETORNO >> " \
                    + str(configuracoes.ds_erro.getvalue()).upper()
                
                    log("SUCESSO PARA ESTA EXECUÇÃO! SEGUE O RETORNO >> " \
                        + str(configuracoes.ds_erro.getvalue()).upper())                
                log('*'*100)
                log("")                    
                log("")
            except:
                pass     
            
            log(" ")
            log(" ")
            log('*'*100)
            nr_linha_controle = 0
            cc_retorno_controle = [[]]
            cc_retorno_controle[0]=["IDCONTROLE"
                        , "JOB"
                        , "MODULO"
                        , "UF"
                        , "INICIO_PERIODO"
                        , "FIM_PERIODO"
                        , "STATUS"
                        , "MSG_ERRO"
                        , "INICIO_EXECUCAO"
                        , "FIM_EXECUCAO"
                        , "QUANTIDADE_INPUT"
                        , "QUANTIDADE_OUTPUT"
                        , "ID_EXEC"
                        , "IDCONTROLE_PAI"
                        , "IDCONTROLE_REP"
                        , "ID_EXEC_PAI"
                        , "NUM_TENTATIVA_EXEC"]
            log("CONSULTA DOS STATUS DA EXECUÇÃO: ")            
            cc_consulta_controle="""
                SELECT 
                  IDCONTROLE
                , JOB
                , MODULO
                , UF
                , TO_CHAR(INICIO_PERIODO,'DD/MM/YYYY HH24:MI:SS') AS INICIO_PERIODO
                , TO_CHAR(FIM_PERIODO,'DD/MM/YYYY HH24:MI:SS') AS FIM_PERIODO
                , STATUS
                , MSG_ERRO
                , TO_CHAR(INICIO_EXECUCAO,'DD/MM/YYYY HH24:MI:SS') AS INICIO_EXECUCAO
                , TO_CHAR(FIM_EXECUCAO,'DD/MM/YYYY HH24:MI:SS') AS FIM_EXECUCAO
                , QUANTIDADE_INPUT
                , QUANTIDADE_OUTPUT
                , ID_EXEC
                , IDCONTROLE_PAI
                , IDCONTROLE_REP
                , ID_EXEC_PAI
                , NUM_TENTATIVA_EXEC  
                FROM GFCADASTRO.tsh_san_controle
            WHERE ID_EXEC = '%s'
            ORDER BY IDCONTROLE, INICIO_EXECUCAO, MODULO
            """%(configuracoes.id_exec)
            log("CONSULTA DO CHECK LIST: ")
            log(cc_consulta_controle)
            configuracoes.connection.executa(cc_consulta_controle)
            configuracoes.cursor = configuracoes.connection.fetchone()
            log(" ")
            log(" ")
            v_ob_dicionario_controle['SQL_TSH_SAN_CONTROLE'] =\
            cc_consulta_controle    
            v_lista_id_controles = set()            
            if (configuracoes.cursor):
                log("CONFIRA O RESULTADO : ")
                #log(str(cc_retorno_controle[nr_linha_controle]))
                while configuracoes.cursor:
                    nr_linha_controle += 1
                    cc_retorno_controle.append([])
                    for campo in configuracoes.cursor:
                        cc_retorno_controle[nr_linha_controle].append(campo)
                    if cc_retorno_controle[nr_linha_controle]:
                        #log(str(cc_retorno_controle[nr_linha_controle]))    
                        try:
                            v_id_controle = cc_retorno_controle[nr_linha_controle][0]
                            if v_id_controle:
                                v_lista_id_controles.add(str(v_id_controle))
                        except:
                            pass    
                    configuracoes.cursor = configuracoes.connection.fetchone()        
            else:
                log(" ** NAO EXISTE DADOS na tabela tsh_san_controle.*** ")
            
            v_ob_dicionario_controle['TSH_SAN_CONTROLE']=\
            cc_retorno_controle
            
            v_ob_dicionario_controle['LISTA_CONTROLE']=\
            v_lista_id_controles
            
            # nr_linha = 0
            # cc_retorno = [[]]
            # cc_retorno[0]=["ID_CONTROLE",	
            #             "DS_ERRO",	
            #             "DS_OBSERVACAO",	
            #             "FILI_COD",	
            #             "DT_REFERENCIA",	
            #             "INFST_SERIE",	
            #             "QTD_ERROS"]    
            #-- if v_lista_id_controles:
            #--     log('*'*100)                
            #--     cc_consulta="""
            #--     SELECT 
            #--         ID_CONTROLE, 
            #--         DS_ERRO, 
            #--         DS_OBSERVACAO, 
            #--         FILI_COD, 
            #--         DT_REFERENCIA, 
            #--         INFST_SERIE, 
            #--         QTD_ERROS
            #--     FROM GFCADASTRO.THS_VALNFTEL_CHECKLIST
            #--     WHERE ID_CONTROLE IN (%s)
            #--     ORDER BY ID_CONTROLE, FILI_COD, INFST_SERIE, DT_REFERENCIA , DS_ERRO
            #--     """%(",".join(v_lista_id_controles))
            #--     log("CONSULTA: ")
            #--     log(cc_consulta)
            #--     configuracoes.connection.executa(cc_consulta)
            #--     configuracoes.cursor = configuracoes.connection.fetchone()
            #--     log(" ")
            #--     log(" ")            
            #--     if (configuracoes.cursor):
            #--         log("CONFIRA O RESULTADO : ")
            #--         log(str(cc_retorno[nr_linha]))
            #--         while configuracoes.cursor:
            #--             nr_linha += 1
            #--             cc_retorno.append([])
            #--             for campo in configuracoes.cursor:
            #--                 cc_retorno[nr_linha].append(campo)
            #--             log(str(cc_retorno[nr_linha]))    
            #--             configuracoes.cursor = configuracoes.connection.fetchone()        
            #--     else:
            #--         log(" ** NAO EXISTE DADOS na tabela THS_VALNFTEL_CHECKLIST.*** ")
            #--     log(" ")
            #--     log(" ")
            #--     log('*'*100)
            #--     log(" ")
            #--     log(" ")
            #--     log('*'*100)
            #--     log("Verificar relatório do checklist " \
            #--         + " na tabela THS_VALNFTEL_CHECKLIST. " \
            #--         + " >> ID_CONTROLE : ", \
            #--         ",".join(v_lista_id_controles))    
            #--     log('*'*100)
            #--     log("")                    
            #--     log("")
            #-- 
            #-- v_cc_relatorio = ""
            #-- if (cc_retorno_controle\
            #-- or cc_retorno)\
            #-- and configuracoes.var_diretorio_relatorio\
            #-- and configuracoes.var_mascarara_arquivo_diretorio:
            #--     pass
            #--     v_cc_relatorio =\
            #--     str(configuracoes.var_mascarara_arquivo_diretorio)\
            #--     .replace('<<ID_CONTROLE>>'\
            #--     ,('REGRA_'\
            #--     + str(configuracoes.filtro_regra_execucao).replace(" ","").replace(",", "_").strip()\
            #--     + '_PERIODO_'\
            #--     + str(dt_inicio.year).rjust(4,'0')\
            #--     + str(dt_inicio.month).rjust(2,'0') \
            #--     + str(dt_inicio.day).rjust(2,'0')\
            #--     + '_'\
            #--     + str(dt_fim.year).rjust(4,'0')\
            #--     + str(dt_fim.month).rjust(2,'0')\
            #--     + str(dt_fim.day).rjust(2,'0')\
            #--     + '_UF_'\
            #--     + str(configuracoes.filtro_uf).strip() + '_' \
            #--     + (('FILIAL_' + str(configuracoes.filtro_filial).replace(" ","").replace(",", "_").strip()\
            #--         +'_') if configuracoes.filtro_filial else '')\
            #--     + (('SERIE_' + str(configuracoes.filtro_serie).replace(" ","").replace(",", "_").strip()\
            #--         +'_') if configuracoes.filtro_filial else '')\
            #--     + configuracoes.dt_atual_banco_dados_texto))
            #--     v_cc_relatorio = \
            #--     os.path.join(configuracoes.var_diretorio_relatorio\
            #--         , v_cc_relatorio)
            #--     log('Criando o relatorio',v_cc_relatorio)
            #--     #### Cria a planilha em memória....
            #--     arquivo_excel = Workbook()
            #--     planilha1 = arquivo_excel.active
            #--     planilha1.title = "STATUS_PROCESSAMENTO"
            #--     for linha in cc_retorno_controle:
            #--         planilha1.append(linha)
            #--     planilha2 = arquivo_excel.create_sheet("RESULTADO_PROCESSAMENTO", 1)
            #--     for linha in cc_retorno:
            #--         planilha2.append(linha)                
            #--     # Grava a planilha Excel
            #--     arquivo_excel.save(v_cc_relatorio)
            #-- else:
            #--     log(" ** NAO FOI POSSIVEL DE CRIAR PLANILHA COM OS DADOS.** ")            
            #-- v_ob_dicionario_controle['RELATORIO'] = v_cc_relatorio,
            configuracoes.lista_id_exec.append(v_ob_dicionario_controle)
            dt_inicio = dt_fim + datetime.timedelta(days=1)

        log('*'*100)
        if configuracoes.lista_id_exec:
            log(" ")
            log(" ")
            log('*'*100)
            try:
                log(" **** DADOS GERADOS (TSH_SAN_CONTROLE) **** ")
                for numero,dicionario in enumerate(configuracoes.lista_id_exec):
                    try:    
                        log(str(numero+1),\
                        ") ID_EXEC: ",\
                        dicionario['ID_EXEC'])
                        log("CODIGO: ",dicionario['CODIGO_RETORNO'])
                        log("RETORNO: ",\
                                    " **ATENCAO** "\
                                    if dicionario['CODIGO_RETORNO']
                                    else ' '
                                    ,dicionario['RETORNO'])    
                        try:
                            for index, tsh_san_controle \
                            in enumerate(dicionario['TSH_SAN_CONTROLE']):
                                if index == 0:
                                    continue
                                log(str(numero+1),\
                                    ".",\
                                    str(index),\
                                    ") IDCONTROLE: ",\
                                    tsh_san_controle[0])
                                log("-JOB: ",tsh_san_controle[1])
                                log("-MODULO: ",tsh_san_controle[2])
                                log("-UF: ",tsh_san_controle[3])
                                if tsh_san_controle[4]:
                                    log("-INICIO_PERIODO: ",tsh_san_controle[4])
                                if tsh_san_controle[4]:
                                    log("-FIM_PERIODO: ",tsh_san_controle[5])
                                log("-STATUS: ",tsh_san_controle[6])
                                if tsh_san_controle[7]:
                                    log("-MSG: ",\
                                        " **ATENCAO** "\
                                        if 'ERRO' in str(tsh_san_controle[7]).upper()
                                        else ' '
                                        , tsh_san_controle[7]
                                    )
                                log("-INICIO_EXECUCAO: ",tsh_san_controle[8])
                                log("-FIM_EXECUCAO: ",tsh_san_controle[9])
                                log("-QUANTIDADE_INPUT: ",tsh_san_controle[10])
                                log("-QUANTIDADE_OUTPUT: ",tsh_san_controle[11])
                                log(" ")
                        except:
                            pass    
                        if dicionario['CODIGO_RETORNO']:
                            log("***ATENÇÃO***")
                            log("***SQL***",dicionario['SQL_TSH_SAN_CONTROLE'])
                            log(" ")                        
                        log(" ")
                        log(" ")
                    except:
                        pass    
            except:
                pass
                
            log('*'*100)
            log(" ")
            log(" ")



            log('*'*100)

    except Exception as err:
        err_desc_trace = traceback.format_exc()
        __fnc_msg(" ERRO AO PROCESSAR: " 
        + str(err) + " - TRACE - " + err_desc_trace)
        v_nr_ret = 1
        
    return v_nr_ret
    
def __fnc_verifica_database():
    """
        Função que conecta no banco de dados. 
    """
    try:
        configuracoes.dt_atual_banco_dados_texto = ""

        configuracoes.connection = sql.geraCnxBD(configuracoes)

        cc_test_sql = """ SELECT 
        TO_CHAR(SYSDATE, 'YYYYMMDDHH24MISS')  AS DATA_ATUAL 
                       FROM DUAL       
                   """

        configuracoes.connection.executa(cc_test_sql)

        configuracoes.cursor = configuracoes.connection.fetchone()
        
        if (configuracoes.cursor):
            for campo in configuracoes.cursor:
                configuracoes.dt_atual_banco_dados_texto = campo
                __fnc_msg(" CONEXAO DE SUCESSO COM O BANCO DE DADOS! ")
                return 0 
        else:
            __fnc_msg(" ERRO AO CONECTAR COM O BANCO DE DADOS! ")
            return 1 
            
    except Exception as err:
        err_desc_trace = traceback.format_exc()
        __fnc_msg(" ERRO AO CONECTADO COM O BANCO DE DADOS: " 
        + str(err) + " - TRACE - " + err_desc_trace)
        return 91 
        
def fnc_main(**kwargs):
    """
        Funcao que executa as funcionalidades principais
    """
    v_nr_ret = 0 
    configuracoes.nm_job = ""
    configuracoes.gv_dt_data_e_hora_atuais =\
    datetime.datetime.now()
    configuracoes.gv_ds_data_e_hora_atuais_texto = \
    configuracoes.gv_dt_data_e_hora_atuais.strftime('%d/%m/%Y %H:%M')
    configuracoes.gv_cc_diferenca_fuso_horario_sao_paulo = \
    datetime.timedelta(hours=-3) #timezone('America/Sao_Paulo')
    configuracoes.gv_cc_fuso_horario_sao_paulo = \
    datetime.timezone(configuracoes.gv_cc_diferenca_fuso_horario_sao_paulo)
    configuracoes.gv_dt_data_e_hora_sao_paulo = \
    configuracoes.gv_dt_data_e_hora_atuais.astimezone(configuracoes.gv_cc_fuso_horario_sao_paulo)
    configuracoes.gv_ds_data_e_hora_sao_paulo_em_texto = \
    configuracoes.gv_dt_data_e_hora_sao_paulo.strftime('%d/%m/%Y %H:%M')
    log("Tratamento de entrada de dados!")
    v_nr_ret = __fnc_tratar_entrada_dados( \
                p_nm_job=kwargs.get('p_nm_job',None)
                , p_cc_filtro_regra_execucao=kwargs.get('p_cc_filtro_regra_execucao',None)
                , p_dt_filtro_inicio=kwargs.get('p_dt_filtro_inicio',None)
                , p_dt_filtro_fim=kwargs.get('p_dt_filtro_fim',None)
                , p_cc_filtro_uf=kwargs.get('p_cc_filtro_uf',None)
                , p_cc_filtro_filial=kwargs.get('p_cc_filtro_filial',None)
                , p_cc_filtro_serie=kwargs.get('p_cc_filtro_serie',None)
                , p_qt_filtro_relatorio=kwargs.get('p_qt_filtro_relatorio',None)               
                , p_tp_filtro_relatorio=kwargs.get('p_tp_filtro_relatorio',None)
                )    
    if v_nr_ret:
        return v_nr_ret, configuracoes.gv_ds_log      
        
    log("Tratamento de parametros de configuracoes!")  
    v_nr_ret = __fnc_tratar_parametro_configuracoes()    
    if v_nr_ret:
        return v_nr_ret, configuracoes.gv_ds_log    
     
    log("Verifica a base de dados!")  
    v_nr_ret = __fnc_verifica_database()
    if v_nr_ret:
        return v_nr_ret, configuracoes.gv_ds_log    
        
    log("Processar os dados!")  
    v_nr_ret = __fnc_processar()    
    if v_nr_ret:
        return v_nr_ret, configuracoes.gv_ds_log           
        
    return v_nr_ret, configuracoes.gv_ds_log
    
if __name__ == "__main__":
    """
        Ponto de partida
    """
    v_nr_ret,v_ds_log = fnc_main() 
    log("retorno ", "->", v_nr_ret)
    sys.exit(v_nr_ret)   
