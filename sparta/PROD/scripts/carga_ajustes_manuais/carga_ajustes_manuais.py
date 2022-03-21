#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-
"""
----------------------------------------------------------------------------------------------
  SISTEMA ..: 
  MODULO ...: 
  SCRIPT ...: carga_ajustes_manuais.py
  CRIACAO ..: 21/10/2020
  AUTOR ....: Airton Borges da Silva Filho / KYROS Consultoria
  DESCRICAO : 
      Listar todos os arquivos *.xlsx do diretório : /arquivos/RETORNO_TRIBUTARIO/Ajustes_manuais/A_Processar
      O nome da tabela onde os dados serão inseridos é obtido na coluna "VALORES" 
      Verificar na coluna "Acao" se o valor dos 4 primeiros caracteres da mesma, transformados em maiúsculas, inicia-se com [ "ALTE" ou  "UPDA" ou "EXCL" ou "DELE" ou "INSE" ou "INCL" ] 
      O nome dos campos da tabela, são os mesmos das colunas no cabeçalho da aba.
      Apresentar no log do processamento o número de registros inseridos para o arquivo.
      Mover o arquivo processado para o diretorio de processados : /arquivos/RETORNO_TRIBUTARIO/Ajustes_manuais/Processado
      O arquivo movido deve ser renomeado, acrescentando em seu nome a tag : "_Processado_YYYYMMDD_HHMi" antes da extensão do mesmo.
      Exemplo : Retorno_Tributario_GAP_42.xlsx renomeado para Retorno_Tributario_GAP_42_Processado_20210908_0922.xlsx

----------------------------------------------------------------------------------------------
  HISTORICO : 
    * 21/10/2021 - Airton Borges da Silva Filho / KYROS Consultoria - Criacao do script.
----------------------------------------------------------------------------------------------
            
"""
import sys
import os
SD = '/' if os.name == 'posix' else '\\'
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV')[0], 'DEV')
sys.path.append(dir_base)
import configuracoes
from comum import log
import comum
import sql
import util
import json

from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook

from openpyxl.reader.excel import load_workbook
import datetime
import shutil
from pathlib import Path
global ret
global nomeowner

nomeowner = "gfcadastro."

NoneType = type(None)
StrType = type("string")
IntType = type(88)
DecType = type(88.88)
DatType = type(datetime.datetime.now())

comum.carregaConfiguracoes(configuracoes)

disco = ('' if os.name == 'posix' else 'D:')

def nome_arquivo(mascara,diretorio):
    qdade = 0
    nomearq = "" 
    directory = Path(diretorio)
    files = directory.glob(mascara)
    sorted_files = sorted(files, reverse=False)
    if sorted_files:
        for f in sorted_files:
            qdade = qdade + 1
            nomearq = f
            log("# ", qdade ," - ",f )
    else: 
        nomearq=""
    return(nomearq)

def lista_arquivos(mascara,diretorio):
    #data_criacao = lambda f: f.stat().st_ctime removido, por mim
    data_modificacao = lambda f: f.stat().st_mtime
    qdade = 0
    nomearq = [] 
    directory = Path(diretorio)
    files = directory.glob(mascara)
    sorted_files = sorted(files, key=data_modificacao, reverse=False)
    if sorted_files:
        for f in sorted_files:
            qdade = qdade + 1
            nomearq.append(f)
            log("# ",qdade ," - ",f )
    else: 
        nomearq = []
    return(nomearq)

     
def dtf():
    return (datetime.datetime.now().strftime('%Y%m%d_%H%M%S'))

def datavalida(valor):
    try:
        datetime.datetime.strptime(valor, '%d/%m/%Y')
        return True
    except:
        return False



def processaXlsx(diretorio, arquivo) :
#    print('-'*100)
    ok = 0
    ret = 0
    path_arq = os.path.join( diretorio, arquivo )
    log('Processando o arquivo:', path_arq)
    xls = load_workbook(path_arq)
    log("Quantidade de abas da planilha = ", len(xls.sheetnames))
    log("Nomes das abas da planilha     = ", xls.sheetnames)
    
    #PROCESSA ABAS DA PLANILHA  
    for p1 in xls.sheetnames:
        print('-'*150)
        log("Processando a aba '", p1, "' do arquivo '", path_arq + "'")
        planilha = xls[p1]
        ncolunas = planilha.max_column
        log("Quantidade de colunas da aba = ", ncolunas)
        if (ncolunas < 4):
            log("ATENÇÃO - aba desprezada por conter somente ", ncolunas , " colunas." )
            continue
        dados = []
        dtabelas = json.loads(configuracoes.tabelas)
        linha = -1
        coluna = -1
        col_tabela = -1
        col_acao = -1
        erroatual = 0
        pulou = 0 
        inseriu = 0
        queries = []
        
        
        #PROCESSA LINHAS DA ABA
        for row in planilha.rows :
            if (erroatual > 0):
                print("erroatual > 0 = ",erroatual)
                break
            linha = linha + 1

            #PROCESSA COLUNAS DAS LINHAS
            dadost = []
            if (linha == 0):
                dadosi = []
                for col in row :
                    comcolatu=("" if type(col.value)==NoneType else col.value.upper())
                    dadosi.append( comcolatu )
                try:
                   col_tabela = dadosi.index("TABELA")
                except:
                   col_tabela = -1
                   erroatual = -1
                try:
                   col_acao = dadosi.index("ACAO")
                except:
                   col_acao = -1
                   erroatual = -1
                if ( col_acao != -1 and col_tabela != -1 ):
                    nomecampos = " (" 
                    for icol in dadosi:
                        nomecampos = nomecampos + icol + ","
                    nomecampos = nomecampos[0:len(nomecampos)-1] + ") "
                else:
                    log("ERRO - Aba desprezada por não conter cabecalho com nomes da tabela e/ou ação")
                    erroatual = -1
                    break         
                #print("Nome de todos os campos da tabela = ", nomecampos)
            else: # LINHA != 0
                valcampos = "("
                catutemp=0
                for col in row :
                    catutemp=catutemp + 1
                    valcamatu = ("Null" if (type(col.value)==NoneType or col.value == "") else col.value)
                    valcamatu = (str(col.value) if (type(col.value) in (IntType,DecType,DatType)) else valcamatu )
                    valcamatu = ("Null" if valcamatu.upper() == "NONE" else valcamatu )
                    valcamatu = ("TO_DATE('"+valcamatu+"','"+ "DD/MM/YYYY')" if datavalida(valcamatu) else valcamatu )
#                    print("valcamatu = ", valcamatu)

                    dadost.append(valcamatu) 
                    if (valcamatu.strip() == ""):
                        valcamatu = "Null"
                    if (len(valcamatu) > 0 and valcamatu != "Null" and valcamatu[0:7].upper() != "TO_DATE"):
                        valcamatu = "'" + valcamatu + "'"
                        
                    valcampos = valcampos + valcamatu + ","
                valcampos = valcampos[0:len(valcampos)-1] + ")"
                nometabela = (dadost[col_tabela].upper() if type(dadost[col_tabela]) != NoneType else "")
                nometabela = nomeowner + dtabelas[nometabela]
                nomeacao   = (dadost[col_acao].upper()[0:4] if type(dadost[col_acao]) !=  NoneType else "")
#                print("========================>>>>>>> NOME DA ACAO = ", nomeacao)
                if ( nomeacao in ("ALTE", "UPDA", "EXCL", "DELE", "INSE", "INCL") ):
                    comando = "insert into " + nometabela + nomecampos + " VALUES " + valcampos               
                    queries.append(comando)  
                    inseriu = inseriu + 1
#                    print("Vai inserir o comando ", comando)
#                    input("CONTINUA?")
                   
                elif ( nomeacao in ("MANT", "", "NONE", "NULL") ):
#                    print("...PULANDO... nome da acao está em (MANT, NONE, NULL, ''). Nome da acao = ", nomeacao)
                    pulou = pulou + 1
                    erroatual = -1
                else:
                    erroatual = 99
                    ok = ok - 1
                
            if ( erroatual > 0 ):
                if (erroatual < 0):
                    log("ATENÇÃO - Aba desprezada por não conter as colunas 'TABELA' e/ou 'ACAO'.")
                else:
                    log("ATENÇÃO - Aba desprezada por conter coluna com valor inválido no campo 'ACAO'. Valor encontrado = ", nomeacao.upper(), ". Erro = ", erroatual)
                break
        ok = ok + 1
        log("="*100)                    
        log("==========>>>>> PROCURANDO NOVA ABA.....")
        
        if (erroatual > 0):
            break

    log("-"*100)
    
#    if ( erroatual == 0 ):
#        print("... COMMIT ...")
#        input("CONT?")
#    elif ( erroatual < 0) :
#        print("... IGNORANDO ...")
#        input("CONT?")
#    else:
#        print("... ROLLBACK ...")
#        input("CONT?")
            
    if ok == 1:
       log("RESUMO DE REGISTROS ENCONTRADOS NO ARQUIVO PROCESSADO:")
       log("==========>>>>> Inseriu = ", inseriu)
       log("==========>>>>> Pulou   = ", pulou)
       log("-"*100)
       return(True,queries)
    else:
        if ok == 0:
            log("ERRO - O arquivo não possui abas com valores corretos a serem inseridos nas tabelas")
        else:
            log("ERRO - O arquivo possui linhas com valores inválidos a serem inseridos nas tabelas, por isso foi desprezado")
            
    return(False,[])


def processar():
    
    log ('')
    log("Verificando arquivos para processar no diretorio de entrada : ", configuracoes.dir_aprocessar)
    
    if not os.path.isdir(configuracoes.dir_aprocessar) :
        os.makedirs(configuracoes.dir_aprocessar)
    
    if not os.path.isdir(configuracoes.dir_processado) :
        os.makedirs(configuracoes.dir_processado)
    
    qtd_arqs = 0
    ret = 99
    
    con=sql.geraCnxBD(configuracoes)
    
    for arq in os.listdir(configuracoes.dir_aprocessar) :
        print (" ")
        print (" ")
        print (" ")
        print("#"* 150)
        log("NOVO ARQUIVO A PROCESSAR.....", arq)
#        input("continua?")
        ret = 0
        if os.path.isfile( os.path.join(configuracoes.dir_aprocessar, arq)) :
            if arq.endswith('.xlsx') :
#                print('-'*100)
                qtd_arqs += 1
#                log('Verificando o arquivo  :',arq)
                
                retorno,queries = processaXlsx(configuracoes.dir_aprocessar, arq)
                
                if(retorno):
                    log("INSERINDO DADOS NA TABELA...")
                    try:
                        for query in queries:
#                            print("Vai executar a query : ",query)
#                            input("Vai executar....")
                            con.executa(query)
#                            input("Executou")
                    except:
                        log("Erro ao inserir os dados na tabela.")
                        retorno = False
                else:
                    log("ARQUIVO DESPREZADO !")
                    
                if (retorno == False):
                    log("ERRO no processamento do Arquivo        :", ' "'+os.path.join(configuracoes.dir_aprocessar, arq)+'"' )
                    ret = 99
                else :
                    vdataArq= dtf()
                    nomearq = arq.split(".xlsx")
                    arqprocessado = nomearq[0]+"_Processado_"+ vdataArq + ".xlsx"
                    log("Arquivo processado     :", '"'+os.path.join(configuracoes.dir_aprocessar, arq)+'"' )
                    log("Movido e renomeado para:", '"'+os.path.join(configuracoes.dir_processado , arqprocessado )+'"') 
                    os.rename(os.path.join(configuracoes.dir_aprocessar, arq),arqprocessado)
                    shutil.move(arqprocessado, configuracoes.dir_processado)
                    con.commit()
                    print("#"*100)
 
    if (qtd_arqs == 0) :
        log("ERRO - Nao foi encontrado nenhum arquivo xlxs a ser processado.")
        ret = 99

    print('-'*100)

    return ret

if __name__ == "__main__" :

    #con=sql.geraCnxBD(configuracoes)
    
    ret = 0
    txt = ''
    ret = processar() 
    log("Codigo de retorno =", ret)
    if (ret != 0 ):
        log('ERRO no processamento.')
        ret = 92

    







