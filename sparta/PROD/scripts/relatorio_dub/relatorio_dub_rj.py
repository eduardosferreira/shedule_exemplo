#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-

"""
----------------------------------------------------------------------------------------------
  SISTEMA ..: 
  MODULO ...: 
  SCRIPT ...: DUB RJ
  CRIACAO ..: 16/07/2021
  AUTOR ....: Victor Santos - Kyros Consultoria 
  DESCRICAO : 
  ALTERACAO :
----------------------------------------------------------------------------------------------
    Exemplo de comando: ./dub_RJ.py <INSCRICAO_ESTADUAL> <ANO> <MẼS INICIO> <MÊS FIM>
    Diretório: /arquivos/DUB/RELATORIOS
    Exemplo: arquivos/DUB/relatório/77452443 – DUB – 032016.xlsx
    
    31/08/2021 - Adquações para novo painel
----------------------------------------------------------------------------------------------
"""

import sys
import os
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)
import configuracoes

import datetime
#import time
import os
from typing import Pattern
import cx_Oracle
#import glob
#import atexit
#import shutil
#import re
#from pathlib import Path
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_MEDIUM, BORDER_DOUBLE
from openpyxl.utils import get_column_letter

sys.path.append( os.path.join( os.path.realpath('..'), "modulosPython" ) )
import comum
import sql

global variaveis
global db

fontMasterPreta  = Font(color='00000000', bold=True, size=12)
fontNegrito      = Font(color='00000000', bold=True)

 
ret = 0
nome_relatorio = "" 
dir_destino = "" 
dir_base = "" 

SD = ('/' if os.name == 'posix' else '\\')
name_script = os.path.basename(__file__).split('.')[0]
variaveis = {}

def processar():
    global variaveis


    vDataIni =""
    vDataFim =""
    IE   = ""
    flag = "" 
    ret = 0

    log(len(sys.argv))
    if ( 
        len(sys.argv) == 5 
        and len(sys.argv[1])==4 
        and len(sys.argv[2])==2 
        and len(sys.argv[3])==2  
        and int(sys.argv[2][0:2])>0 
        and int(sys.argv[2][0:2])<13
        and int(sys.argv[3][0:2])>0 
        and int(sys.argv[3][0:2])<13
        and int(sys.argv[1])<=datetime.datetime.now().year
        and int(sys.argv[1])>(datetime.datetime.now().year)-50
        ):
        
        vAno     = sys.argv[1]
        vMesIni  = sys.argv[2]
        vMesFim  = sys.argv[3]
        IE       = sys.argv[4] 

        vDataIni ='01/' + vMesIni +'/'+ vAno
        UltDiaMes =ultimodia(int(vAno), int(vMesFim))
        vDataFim =str(UltDiaMes)+ '/' + vMesFim + '/' + vAno 

        log("-"* 100)
        log('# - Ano................:', vAno)
        log('# - Mẽs Inicial........:', vMesIni)
        log('# - Mẽs FInal..........:', vMesFim)
        log('# - Inscricao estadual.:', IE)
        log("-"* 100)
        log('# - DataIni............:',vDataIni)
        log('# - DataFim............:',vDataFim)
        log("-"* 100)
        
    else:
        log("-" * 100)
        log("#### ")
        log('#### ERRO - Erro nos parametros do script.')
        log("#### ")
        log('#### Exemplo de como deve ser :')
        log('####      %s  <ANO> <MÊS INICIO> <MÊS FIM> <IE>'%(sys.argv[0]))
        log("#### ")
        log('#### Onde')
        log('####      < ANO >        = 2021')
        log('####      < MÊS INICIO > = 01')
        log('####      < MÊS FIM >    = 03')
        log('####      < IE >         = 77452443')
        log('#### segue um exemplo %s 2021 01 03 77452443'%(sys.argv[0]))
        log("#### ")
        log("-" * 100)
        log("")
        log("Retorno = 99") 
        ret = 99
        return(ret)  

    log("# - Conectando no banco de dados")
    vUF=retornaUF(IE)
    if (vUF == ""):
        log("#### ERRO - Não foi possível determinar a UF pela IE informada.")
        ret = 99
        return(ret)
    if (vUF != "RJ"):
        log("#### ERRO - A IE informada não é de RJ.")
        ret = 99
        return(ret)
    

#### Monta caminho e nome do destino
    dir_arq = configuracoes.dir_arquivos
    dir_arq =  os.path.join(dir_arq, vUF)   
    dir_destino = os.path.join(dir_arq, vAno, vMesIni)  

    if not os.path.isdir(dir_destino) :
        os.makedirs(dir_destino) 
    
    vPeriodo = vMesIni + "_a_" + vMesFim + "_de_" + vAno
    arquivo_destino = IE+'_DUB_RJ_'+vPeriodo+'.xlsx'
    nome_relatorio = os.path.join(dir_destino,arquivo_destino)
    log("-"* 100)
    log('#### - Planilha Relatório Destino = ',nome_relatorio)
    log("-"* 100)

    arquivo = open(nome_relatorio, 'w')
    arquivo.close() 

    #### Cria a planilha em memória....
    arquivo_excel = Workbook()
    planilha0 = arquivo_excel.active
    planilha0.title = "ABA 1.DADOS MENSAL"

###################################################################################
####Aba 2.  Levantamento Município Saídas – Telecom
####Aba 2.  Levantamento Município Saídas – Telecom
####Aba 2.  Levantamento Município Saídas – Telecom
###################################################################################
    log("")
    log("# ",dtf() , " - Início do processamento da ABA 1: 'Dados Mensal'.")

#### CABEÇALHO 
#### CABEÇALHO 
#### CABEÇALHO 

    vLinha = 1
    planilha0.cell(vLinha,1,"DUB-ICMS – Documento de Utilização de Benefício Fiscal (Mensal)")
    planilha0.cell(vLinha,1).font=Font(bold=True,size=14)
    planilha0.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha0.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.merge_cells('A'+ str(vLinha) + ':N' + str(vLinha))

    vLinha = vLinha + 1

    planilha0.cell(vLinha,1,"Insc Estadual...: "+IE)
    planilha0.cell(vLinha,1).font=Font(bold=True)
    planilha0.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.merge_cells('A'+ str(vLinha) + ':N' + str(vLinha))

    vLinha = vLinha + 1

    planilha0.cell(vLinha,1,"Mês/Ano........: " + vPeriodo)
    planilha0.cell(vLinha,1).font=Font(bold=True)
    planilha0.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.merge_cells('A'+ str(vLinha) + ':N' + str(vLinha))

    vLinha = vLinha + 1

    vColuna = 1

    for nColuna in ("Código","Razão Social","CNPJ","Data Emissão","Serie","Número NF","Terminal","Código Serviço",
                    "Descrição Serviço","Valor Total","Base ICMS","Valor ICMS","Valor Isentas","Valor Outras"):
        planilha0.cell(vLinha,vColuna,nColuna)
        planilha0.cell(vLinha,vColuna).font=Font(bold=True)
        planilha0.cell(vLinha,vColuna).alignment = Alignment(horizontal='center')
        vColuna = vColuna + 1

#### DADOS
#### DADOS
#### DADOS
    
    dados = []
    dados = aba_dados_mensal(vDataIni,vDataFim,IE)  
    plinhaP0 = vLinha + 1 ## primeira linha com dados a serem somados.
    
    if not dados:
        return(99)

    for linha in dados:
        vLinha = vLinha + 1
        planilha0.cell(vLinha, 1,linha[0]).alignment = Alignment(horizontal='center')
        planilha0.cell(vLinha, 2,linha[1])
        planilha0.cell(vLinha, 3,linha[2])
        planilha0.cell(vLinha, 4,linha[3]).alignment = Alignment(horizontal='center')
        planilha0.cell(vLinha, 5,linha[4]).alignment = Alignment(horizontal='center')
        planilha0.cell(vLinha, 6,linha[5]).alignment = Alignment(horizontal='center')
        planilha0.cell(vLinha, 7,linha[6]).alignment = Alignment(horizontal='center')
        planilha0.cell(vLinha, 8,linha[7]).alignment = Alignment(horizontal='center')
        planilha0.cell(vLinha, 9,linha[8])
        planilha0.cell(vLinha,10,linha[9]).alignment = Alignment(horizontal='center')
        planilha0.cell(vLinha,11,linha[10]).alignment = Alignment(horizontal='center')
        planilha0.cell(vLinha,12,linha[11]).alignment = Alignment(horizontal='center')
        planilha0.cell(vLinha,13,linha[12]).alignment = Alignment(horizontal='center')
        planilha0.cell(vLinha,14,linha[13]).alignment = Alignment(horizontal='center')

        planilha0.cell(vLinha,10).number_format = "#,##0.00"
        planilha0.cell(vLinha,11).number_format = "#,##0.00"
        planilha0.cell(vLinha,12).number_format = "#,##0.00"
        planilha0.cell(vLinha,13).number_format = "#,##0.00"
        planilha0.cell(vLinha,14).number_format = "#,##0.00"
    
    ulinhaP0 = vLinha 

#### TOTAIS
#### TOTAIS
#### TOTAIS
    
    vLinha = vLinha + 2
    planilha0.cell(vLinha,1,"TOTAIS:")
    planilha0.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha0.cell(vLinha,1).font=Font(bold=True)
    planilha0.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.merge_cells('A'+ str(vLinha) + ':I' + str(vLinha))
    

    planilha0.cell(vLinha,10,"=SUM(J"+str(plinhaP0)+":J"+str(ulinhaP0)+")")
    planilha0.cell(vLinha,10).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.cell(vLinha,10).font=Font(bold=True)
    planilha0.cell(vLinha,10).number_format = "#,##0.00"

    planilha0.cell(vLinha,11,"=SUM(K"+str(plinhaP0)+":K"+str(ulinhaP0)+")")
    planilha0.cell(vLinha,11).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.cell(vLinha,11).font=Font(bold=True)
    planilha0.cell(vLinha,11).number_format = "#,##0.00"

    planilha0.cell(vLinha,12,"=SUM(L"+str(plinhaP0)+":L"+str(ulinhaP0)+")")
    planilha0.cell(vLinha,12).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.cell(vLinha,12).font=Font(bold=True)
    planilha0.cell(vLinha,12).number_format = "#,##0.00"

    planilha0.cell(vLinha,13,"=SUM(M"+str(plinhaP0)+":M"+str(ulinhaP0)+")")
    planilha0.cell(vLinha,13).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.cell(vLinha,13).font=Font(bold=True)
    planilha0.cell(vLinha,13).number_format = "#,##0.00"

    planilha0.cell(vLinha,14,"=SUM(N"+str(plinhaP0)+":N"+str(ulinhaP0)+")")
    planilha0.cell(vLinha,14).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.cell(vLinha,14).font=Font(bold=True)
    planilha0.cell(vLinha,14).number_format = "#,##0.00"
    
#### FORMATAÇAO
#### FORMATAÇAO
#### FORMATAÇAO

    adjust_column(planilha0, 1,1, planilha0.max_column)

    planilha0.column_dimensions['A'].width = 20 

    set_border_edsi(planilha0, 'A4:N'+str(planilha0.max_row), 'ffff')
    set_border_edsi(planilha0, 'A'+str(planilha0.max_row)+':N'+str(planilha0.max_row), 'gggg')


#### GRAVA A PLANILHA
#### GRAVA A PLANILHA
#### GRAVA A PLANILHA

    arquivo_excel.save(nome_relatorio)   

    log("")
    log("# ",dtf() , " - Fim do processamento da ABA 1: 'Dados Mensal'.") 

    return(ret)

def adjust_column(ws, min_row, min_col, max_col):
    
    column_widths = []

    for i, col in \
        enumerate(
            ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
        ):

        for cell in col:
            value = cell.value
            if value is not None:

                if isinstance(value, str) is False:
                    value = str(value)

                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))

    for i, width in enumerate(column_widths):

        col_name = get_column_letter(min_col + i)
        value = column_widths[i] + 2
        ws.column_dimensions[col_name].width = value

def dtf():
    return (datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S'))

def ultimodia(ano,mes):
   return(31 if mes == 12 else datetime.date.fromordinal((datetime.date(ano,mes+1,1)).toordinal()-1).day)

def retornaUF(IE):
    vUF = ""
    query="""
    select distinct f.unfe_sig  from openrisow.filial f where f.fili_cod_insest='%s'
    """%(IE)

    con = sql.geraCnxBD(configuracoes)
    con.executa(query)
    result = con.fetchone()

    if(result != None): 
        for campo in result:
            vUF = campo
    return(vUF)


def aba_dados_mensal(vDataIni,vDataFim,IE):
    query="""WITH consulado AS (SELECT DISTINCT RAZAO_SOCIAL, CNPJ FROM gfcadastro.tb_clientes_consulados)   
                SELECT /*+ parallel(12) */
                    it.CADG_COD,
                    cc.razao_social,
                    it.CGC_CPF,
                    to_char(it.INFST_DTEMISS,'dd/mm/yyyy') AS DTEmissao,
                    it.INFST_serie,
                    it.INFST_NUM,
                        (SELECT CADG_NUM_CONTA--CADG_TEL_CONTATO
                                FROM openrisow.COMPLVU_CLIFORNEC cli
                            WHERE cli.cadg_cod       = it.cadg_cod
                                AND cli.catg_cod       = it.catg_cod
                                AND cli.cadg_dat_atua = (select max(cli2.cadg_dat_atua) cadg_dat_atua
                                                    FROM openrisow.COMPLVU_CLIFORNEC cli2
                                                    WHERE cli2.cadg_cod       = it.cadg_cod
                                                        AND cli2.catg_cod       = it.catg_cod
                                                        AND cli2.cadg_dat_atua  <= it.infst_dtemiss)) AS CADG_NUM_CONTA,
                    it.SERV_COD,
                    it.INFST_DSC_COMPL,
                    it.INFST_VAL_CONT,
                    it.INFST_BASE_ICMS,
                    it.INFST_VAL_ICMS,
                    it.INFST_ISENTA_ICMS,
                    it.INFST_OUTRAS_ICMS       
                    
                FROM openrisow.ITEM_NFTL_SERV it,
                    consulado cc
                WHERE it.EMPS_COD       = 'TBRA'   --fixo
                AND it.FILI_COD       IN (select fili_cod from openrisow.filial f where f.emps_cod = 'TBRA' AND f.fili_cod_insest= '%s')--Parametro: IE
                AND it.INFST_SERIE   IS NOT NULL --fixo
                AND it.INFST_IND_CANC <> 'S'     --fixo
                AND it.INFST_dtemiss >= TO_DATE('%s', 'DD/MM/YYYY')--Parametro: <Data de acordo com a aba que esta sendo gerada>
                AND it.INFST_dtemiss <= TO_DATE('%s', 'DD/MM/YYYY')--Parametro: <Data de acordo com a aba que esta sendo gerada>
                AND it.CGC_CPF = cc.cnpj
                order by CGC_CPF,to_char(it.INFST_DTEMISS,'dd/mm/yyyy'),  INFST_NUM, INFST_DSC_COMPL

    """%(IE,vDataIni,vDataFim)
    
    retorno=[]
    con = sql.geraCnxBD(configuracoes)
    con.executa(query)
    result = con.fetchone()
    lin = 0

    if result == None:
        log("#### ATENÇÃO: Nenhum Resultado para aba 0")
        log("####     Query = ")
        log("####")
        log(query)
        log("####")
        return(retorno)
    else:
        while result:
            retorno.append([])
            for campo in result:
                retorno[lin].append(campo)
            lin = lin + 1
            result = con.fetchone()
    
    return(retorno)

def set_border_edsi( ws, cell_range, edsi):
    b=[]
 
    n=Side(border_style= None    , color='00000000')
    f=Side(border_style='thin'   , color='00000000')
    m=Side(border_style='medium' , color='00000000')
    d=Side(border_style='double' , color='00000000')
    p=Side(border_style='dotted' , color='00000000')
    t=Side(border_style='dashed' , color='00000000')
    g=Side(border_style='thick'  , color='00000000')
    
    for x in (0,1,2,3):
        if (edsi[x]=="n"):
            b.append(n)
        elif (edsi[x]=='f'):
            b.append(f)
        elif (edsi[x]=='m'):
            b.append(m)
        elif (edsi[x]=='d'):
            b.append(d)
        elif (edsi[x]=='p'):
            b.append(p)
        elif (edsi[x]=='t'):
            b.append(t)
        elif (edsi[x]=='g'):
            b.append(g)
        else:
            return
    
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = Border(left=b[0],right=b[1],top=b[2],bottom=b[3])



if __name__ == "__main__":
    log("-"*100)
    log("#### ",dtf(), " INICIO DO RELATORIO DUB RJ... ####")
    variaveis = comum.carregaConfiguracoes(configuracoes)
    log('usuario de banco',configuracoes.userBD)

    ret = processar()
    if (ret > 0) :
        log("#### Código de execução = ", ret)
    log("#### ",dtf(), " FIM DO RELATORIO DUB RJ... ####")
    sys.exit(ret)   
