#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-

"""
----------------------------------------------------------------------------------------------
  SISTEMA ..: 
  MODULO ...: 
  SCRIPT ...: Relatório gia totais por cfop
  CRIACAO ..: 12/01/2021
  AUTOR ....:  Airton Borges da Silva Filho - Kyros Consultoria 
  DESCRICAO : Agendar e acompanhar a execução do SPED
                

----------------------------------------------------------------------------------------------
    Exemplo de comando: ./relatorio_gia_totais_por_cfop.py SP 102020 108383949112
    Diretório: /arquivos/GIA<UF>/RELATORIOS/<ANO>/<MES>
    Exemplo: /arquivos/GIASP/RELATORIOS/2020/10
----------------------------------------------------------------------------------------------
    20210519 - Airton
      Alterar nome de colunas 
    20210811 - Gremonesi
        Adquacoes para o novo painel 

"""

import sys
import os
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)
import configuracoes

import datetime
import time
import cx_Oracle
import glob
import shutil
import re
from pathlib import Path
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook

global variaveis
global db
global arquivo_destino

import comum
import sql


log.gerar_log_em_arquivo = True


#exit(0)


#diretorio_carga = configuracoes.dir

ret = 0
nome_relatorio = "" 
dir_destino = "" 


SD = ('/' if os.name == 'posix' else '\\')
name_script = os.path.basename(__file__).split('.')[0]
variaveis = {}

fontMasterPreta  = Font(color='00000000', bold=True, size=12)


def validauf(uf):
    return(True if (uf.upper() in ('AC','AL','AM','AP','BA','CE','DF','ES','GO','MA','MG','MS','MT','PA','PB','PE','PI','PR','RJ','RN','RO','RR','RS','SC','SE','SP','TO')) else False)
          
def dtf():
    return (datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S'))

def ultimodia(ano,mes):
   return(31 if mes == 12 else datetime.date.fromordinal((datetime.date(ano,mes+1,1)).toordinal()-1).day)

# def carregaConfiguracoes(logar = True) :
#     arq_cfg = '%s.cfg'%(os.path.basename(sys.argv[0]).replace('.py',''))
#     path_arq_cfg = os.path.join( '.', arq_cfg )
#     if not os.path.isfile( path_arq_cfg ) :
#         arq_cfg = 'scripts_rpa.cfg'
#         path_arq_cfg = os.path.join( os.path.realpath('.'), arq_cfg )
#         if not os.path.isfile( path_arq_cfg ) :
#             path_arq_cfg = os.path.join( os.path.realpath('..'), 'unificado', arq_cfg )
#             if not os.path.isfile( path_arq_cfg ) :
#                 arq_cfg = False
#     config = {}

#     if arq_cfg :
#         log('## Carregando dados do arquivo de configuracao .:', arq_cfg)
#         fd = open(path_arq_cfg, 'r')
#         for item in fd.readlines() :
#             if not item.startswith('#') :
#                 if item.__contains__('=') :
#                     linha = item.replace('\n','').split('=')
#                     if len(linha) >= 2 :
#                         config[linha[0]] = linha[1]
#         fd.close()
#     else :
#         return False
#     return config

def proximo_arquivo(mascara,diretorio):
    qdade = 0
    nomearq = "" 
    directory = Path(diretorio)
    files = directory.glob(mascara)
    sorted_files = sorted(files, reverse=False)    
   
    nomearq = mascara.replace("*", "V000")
    proximo = "001"
    
    if sorted_files:
        for f in sorted_files:
            qdade = qdade + 1
            nomearq = f
        proximo = '{:03d}'.format(int((str(nomearq).split(".")[0]).split("_")[-1][1:]) + 1)
    nomearq = mascara.replace("*", "V"+proximo)
    return(nomearq)


def processar():
    global variaveis
    global db
    global arquivo_destino
    ufi = ""
    mesanoi = ""
    mesi = ""
    anoi = "" 
    ret = 0
    
#### Recebe, verifica e formata os argumentos de entrada.  
    if (len(sys.argv) == 4 ): 
        ufi = sys.argv[1].upper()
    if (len(sys.argv) == 4 
        and validauf(ufi)
        and len(sys.argv[2])==6  
        and int(sys.argv[2][:2])>0 
        and int(sys.argv[2][:2])<13
        and int(sys.argv[2][2:])<=datetime.datetime.now().year
        and int(sys.argv[2][2:])>(datetime.datetime.now().year)-50
        ):
     
        mesanoi = sys.argv[2].upper()
        mesi  = sys.argv[2][:2].upper()
        anoi  = sys.argv[2][2:].upper()
        datai = "01/"+mesi+"/"+anoi
        dataf = str(ultimodia(int(anoi),int(mesi)))+"/"+str(mesi)+"/"+str(anoi)
        iei   = sys.argv[3]
        iei = re.sub('[^0-9]','',iei)
        if ( (iei == "") or (iei == "''") or (iei == '""') or (int("0"+iei) == 0)):
            iei = "*"

    else :
        log("-" * 100)
        log("#### ")
        log('#### ERRO - Erro nos parametros do script.')
        log("#### ")
        log('#### Exemplo de como deve ser :')
        log('####      %s <UF> <MMAAAA> <IE> '%(sys.argv[0]))
        log("#### ")
        log('#### Onde')
        log('####      <UF>     = estado. Ex: SP')
        log('####      <MMAAAA> = mês e ano. Ex: Para junho de 2020 informe 062020')
        log('####      <IE>     = Inscrição Estadual')
        log("#### ")
        log('#### Portanto, se o estado = SP, o mes = 06 e o ano = 2020, e IE = 108383949112 o comando correto deve ser :')  
        log('####      %s SP 062020 108383949112'%(sys.argv[0]))  
        log("#### ")
        log('#### ')
        log("-" * 100)
        log("")
        log("Retorno = 99") 
        ret = 99
        return(ret)
    
#### Monta caminho e nome do destino
#    dir_base = SD + 'arquivos' + SD + 'RELATORIOS' + SD + 'MENSAIS' 
#    dir_destino = os.path.join(dir_base, ufi, anoi, mesi)
    #dir_base = SD + 'arquivos' + SD + 'GIA' + ufi + SD + 'RELATORIOS' + SD  
    dir_arquivos = configuracoes.diretorio_arquivos
    dir_arquivos = dir_arquivos + SD + ufi
    log("diretorio onde sera gravado", dir_arquivos)
    dir_destino = os.path.join(dir_arquivos, anoi, mesi)
    
#### Se a pasta do relatório não existir, cria
    if not os.path.isdir(dir_destino) :
        os.makedirs(dir_destino)     
    
#### Monta o nome do próximo arquivo (VERSÃO V001 ou última + 1 )    
    maskfile =  "relatorio_gia_totais_por_cfop_" + mesanoi + "_" + ufi + "_"+ iei +"_*.xlsx"
    arquivo_destino = proximo_arquivo(maskfile,dir_destino)
    arquivo_destino = os.path.join(dir_destino,arquivo_destino)
    nome_relatorio = arquivo_destino
    arquivo = open(arquivo_destino, 'w')
    arquivo.close()   
    
#### Carrega o nome do banco 
 ##   db = (variaveis.get('banco') if variaveis.get('banco') else db)    
    
#### Cria a planilha em memória....
    arquivo_excel = Workbook()
    planilha0 = arquivo_excel.active
    planilha0.title = "Detalhado"
    planilha1 = arquivo_excel.create_sheet("Resumo", 1)
    
    
    #### define o dia inicial e final no formato invertido AAAAMMDD     
    pdiai = str(anoi)+str(mesi)+"01"
    udiai = str(anoi)+str(mesi)+str(ultimodia(int(anoi),int(mesi)))

    
    dhip = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')      
    amdhm = datetime.datetime.now().strftime('%Y%m%d%H%M')
    dip = datetime.datetime.now().strftime('%d/%m/%Y')

    if ( 1==2):
        log(" = ", )
        log("###### debug ######")
        log("dir_arquivos        = ", dir_arquivos)    
        log("dir_destino         = ", dir_destino)
        log("maskfile            = ", maskfile)
        log("arquivo_destino     = ", arquivo_destino)
        log("db                  = ", db)
        
        log("IE                  = ", iei)
        log("data_início         = ", datai)
        log("data_fim            = ", dataf)
        log("data_inicio_invert  = ", pdiai)
        log("data_fim_invert     = ", udiai)
        log("MêsAno              = ", mesanoi)
        log("Data_inicio_Proc    = ", dip)
        log("Data_Hora_Inic_Proc = ", dhip)
        log("AAAAMMDDHHMM        = ", amdhm)
        log("###### debug ######")
        log(" = ", )

    
    log("# ") 
    log("# ",dtf() , " - Início do processamento da aba 'Detalhado'. Script relatorio_gia_totais_por_cfop.py")
    dados = []
    dados = aba_detalhado(iei,pdiai,udiai)
    for linha in dados:
#        print ("linha = ", linha)
        planilha0.append(linha)
    formata_planilha0(planilha0)
#    print ("nome_relatorio = ",nome_relatorio )
    arquivo_excel.save(nome_relatorio)
    log("# ",dtf() , " - Fim do processamento da aba 'Detalhado'. Script relatorio_gia_totais_por_cfop.py")
    log("# ")

    
    log("# ") 
    log("# ",dtf() , " - Início do processamento da aba 'Resumo'. Script relatorio_gia_totais_por_cfop.py")
    dadosr = []
    dadosr = aba_resumo(dados)
    for linha in dadosr:
#        print ("linha = ", linha)
        planilha1.append(linha)
    formata_planilha1(planilha1)
#    print ("nome_relatorio = ",nome_relatorio )
    arquivo_excel.save(nome_relatorio)
    log("# ",dtf() , " - Fim do processamento da aba 'Resumo'. Script relatorio_gia_totais_por_cfop.py")
    log("# ")
    
    log("# ")
    print ("# ",dtf() , "nome_relatorio = ",nome_relatorio )
    log("# ")
     
     
    return(0)



 
def procura(valor,lista):
    linp = 0     
    qlin = len(lista)
    
    
    while linp < qlin:
        
 #       log("lista loop = ", lista)
 #       print ("linp      = ", linp)
 #       print ("qlin      = ", qlin)
        
        
        
        if (lista[linp][0] == valor):
            return(linp)
        linp = linp + 1
    return(-1)    
    
        


def aba_resumo(dados):

    retorno = []
    lin     = 0
    qdade = len(dados)
    ins = 0
    
    
    while lin < qdade:
#        log("lin while  = ", lin)
#        log("qdade      = ", qdade)
        if (lin == 0 ):
            retorno.append([])
            retorno[lin].append("CFOP")
            retorno[lin].append("VLR_LIQUIDO")
            retorno[lin].append("BC_ICMS")
            retorno[lin].append("VLR_ICMS")
            retorno[lin].append("VLR_ISENTAS")
            retorno[lin].append("VLR_OUTRAS")
#            log("retorno[0][0]", retorno[0][0])
#            log("retorno[0][1]", retorno[0][1])
#            log("retorno[0][2]", retorno[0][2])
#            log("retorno[0][3]", retorno[0][3])
#            log("retorno[0][4]", retorno[0][4])
#            log("retorno[0][5]", retorno[0][5])
            
        else:    
            if (dados[lin][2] != "0000"):
                indice = procura(dados[lin][2], retorno)
#                log("indice = ", indice)
#                log("linha  = ", lin)
                if ( indice == -1 ):
                    retorno.append([])
                    ins = ins + 1
#                    log("lin xxxxxxxxxxxx = ", lin)

#                    log("dados[lin][0] xx = ", dados[lin][0])
                    
                    retorno[ins].append(dados[lin][2])
                    retorno[ins].append(dados[lin][5])
                    retorno[ins].append(dados[lin][6])
                    retorno[ins].append(dados[lin][7])
                    retorno[ins].append(dados[lin][8])
                    retorno[ins].append(dados[lin][9])
                else:
                    retorno[indice][1] = (0.00 if (retorno[indice][1] is None) else retorno[indice][1]) + (0.00 if (dados[lin][5] is None) else dados[lin][5]) 
                    retorno[indice][2] = (0.00 if (retorno[indice][2] is None) else retorno[indice][2]) + (0.00 if (dados[lin][6] is None) else dados[lin][6]) 
                    retorno[indice][3] = (0.00 if (retorno[indice][3] is None) else retorno[indice][3]) + (0.00 if (dados[lin][7] is None) else dados[lin][7]) 
                    retorno[indice][4] = (0.00 if (retorno[indice][4] is None) else retorno[indice][4]) + (0.00 if (dados[lin][8] is None) else dados[lin][8]) 
                    retorno[indice][5] = (0.00 if (retorno[indice][5] is None) else retorno[indice][5]) + (0.00 if (dados[lin][9] is None) else dados[lin][9]) 
        lin = lin + 1
    ins = ins + 1
    retorno.append([])
    retorno[ins].append("")
    retorno[ins].append("=SUM(B2:B" + str(ins) + ")")
    retorno[ins].append("=SUM(C2:C" + str(ins) + ")")
    retorno[ins].append("=SUM(D2:D" + str(ins) + ")")
    retorno[ins].append("=SUM(E2:E" + str(ins) + ")")
    retorno[ins].append("=SUM(F2:F" + str(ins) + ")")

    

    return(retorno)



def aba_detalhado(ie,dtini,dtfim):
    
    query = """
        SELECT /*+ PARALLEL (8) */
            I.EMPS_COD               AS EMPRESA, 
            I.FILI_COD               AS FILI_COD,
            I.CFOP                   AS CFOP,
            I.UF                     AS UF_NOTA,
            i.infst_serie            as SERIE,
            SUM(CASE WHEN i.infst_dtemiss < to_date('01/01/2017','dd/mm/yyyy') THEN NVL(INFST_VAL_SERV,0) - NVL(INFST_VAL_DESC,0)
                ELSE NVL(INFST_VAL_CONT,0)
                END)                 AS TOTAL_LIQUIDO,--Antes VLR_CONTABIL
            SUM(I.INFST_BASE_ICMS)   AS VLR_BASE_ICMS, 
            SUM(I.INFST_VAL_ICMS)    AS VLR_ICMS,
            SUM(I.INFST_ISENTA_ICMS) AS VLR_ISENTAS, 
            SUM(I.INFST_OUTRAS_ICMS) AS VLR_OUTRAS
        FROM openrisow.ITEM_NFTL_SERV I
        WHERE I.EMPS_COD = 'TBRA'
            AND I.FILI_COD in (select f.fili_cod from openrisow.filial f where f.emps_cod = 'TBRA' AND f.FILI_COD_INSEST = '%s')
            AND I.infst_dtemiss   >= TO_DATE('%s', 'YYYYMMDD')
            AND i.infst_dtemiss   <= TO_DATE('%s', 'YYYYMMDD')
            AND I.INFST_IND_CANC = 'N'
        GROUP BY I.EMPS_COD,I.FILI_COD,I.infst_serie,I.CFOP,I.UF,EXTRACT(month FROM i.INFST_DTEMISS) 
        order by infst_serie, I.CFOP
    """%(ie,dtini,dtfim)

    #query= comum.troca_owner(query,configuracoes)
       
    retorno = [[]]
    lin = 0 
    retorno[0]=["EMPRESA",
                "FILI_COD",
                "CFOP",
                "UF_NOTA",
                "SERIE",
                "VLR_LIQUIDO",
                "VLR_BASE_ICMS",
                "VLR_ICMS",
                "VLR_ISENTAS",
                "VLR_OUTRAS"]
   
####DEBUG####
####DEBUG####
#    log("query = ", query)    
####DEBUG####
####DEBUG####
   
    ##connection = cx_Oracle.connect(uid+"/"+pwd+"@"+db)
    con=sql.geraCnxBD(configuracoes)
    con.executa(query)
    result = con.fetchone()

    log(result)
    

    if result == None:
        log("#### ATENÇÃO: Nenhum Resultado para aba detalhado")
        log("####     Query = ")
        log("####")
        log(query)
        log("####")
        ret=99
        return(retorno)
    else:
        while result:
            lin = lin + 1
            retorno.append([])
            for campo in result:
                retorno[lin].append(campo)
            result = con.fetchone()
    #cursor.close()
    #connection.close()

    return(retorno)








def formata_planilha0(planilha):
    planilha.column_dimensions['A'].width = 11
    planilha.column_dimensions['B'].width = 11
    planilha.column_dimensions['C'].width = 7
    planilha.column_dimensions['D'].width = 11
    planilha.column_dimensions['E'].width = 7
    planilha.column_dimensions['F'].width = 18
    planilha.column_dimensions['G'].width = 18
    planilha.column_dimensions['H'].width = 18
    planilha.column_dimensions['I'].width = 18
    planilha.column_dimensions['J'].width = 18
    planilha.cell(1,  1).font = fontMasterPreta
    planilha.cell(1,  2).font = fontMasterPreta
    planilha.cell(1,  3).font = fontMasterPreta
    planilha.cell(1,  4).font = fontMasterPreta
    planilha.cell(1,  5).font = fontMasterPreta
    planilha.cell(1,  6).font = fontMasterPreta
    planilha.cell(1,  7).font = fontMasterPreta
    planilha.cell(1,  8).font = fontMasterPreta
    planilha.cell(1,  9).font = fontMasterPreta
    planilha.cell(1,  10).font = fontMasterPreta
    planilha.cell(1,  1).alignment = Alignment(horizontal='center')
    planilha.cell(1,  2).alignment = Alignment(horizontal='center')
    planilha.cell(1,  3).alignment = Alignment(horizontal='center')
    planilha.cell(1,  4).alignment = Alignment(horizontal='center')
    planilha.cell(1,  5).alignment = Alignment(horizontal='center')
    planilha.cell(1,  6).alignment = Alignment(horizontal='center')
    planilha.cell(1,  7).alignment = Alignment(horizontal='center')
    planilha.cell(1,  8).alignment = Alignment(horizontal='center')
    planilha.cell(1,  9).alignment = Alignment(horizontal='center')
    planilha.cell(1,  10).alignment = Alignment(horizontal='center')
#    log("Quantidade de linhas = ",planilha.get_highest_column())

    linha = 1
    for row in planilha.rows:
        if (linha > 1):
            planilha.cell(linha,  6).number_format = "#,##0.00"
            planilha.cell(linha,  7).number_format = "#,##0.00"
            planilha.cell(linha,  8).number_format = "#,##0.00"
            planilha.cell(linha,  9).number_format = "#,##0.00"
            planilha.cell(linha,  10).number_format = "#,##0.00"  
        linha = linha + 1    


def formata_planilha1(planilha):
    planilha.column_dimensions['A'].width = 7
    planilha.column_dimensions['B'].width = 18
    planilha.column_dimensions['C'].width = 18
    planilha.column_dimensions['D'].width = 18
    planilha.column_dimensions['E'].width = 18
    planilha.column_dimensions['F'].width = 18
    planilha.cell(1,  1).font = fontMasterPreta
    planilha.cell(1,  2).font = fontMasterPreta
    planilha.cell(1,  3).font = fontMasterPreta
    planilha.cell(1,  4).font = fontMasterPreta
    planilha.cell(1,  5).font = fontMasterPreta
    planilha.cell(1,  6).font = fontMasterPreta
    planilha.cell(1,  1).alignment = Alignment(horizontal='center')
    planilha.cell(1,  2).alignment = Alignment(horizontal='center')
    planilha.cell(1,  3).alignment = Alignment(horizontal='center')
    planilha.cell(1,  4).alignment = Alignment(horizontal='center')
    planilha.cell(1,  5).alignment = Alignment(horizontal='center')
    planilha.cell(1,  6).alignment = Alignment(horizontal='center')   
    
    linha = 1
    for row in planilha.rows:
        if (linha > 1):
            planilha.cell(linha,  2).number_format = "#,##0.00"
            planilha.cell(linha,  3).number_format = "#,##0.00"
            planilha.cell(linha,  4).number_format = "#,##0.00"
            planilha.cell(linha,  5).number_format = "#,##0.00"
            planilha.cell(linha,  6).number_format = "#,##0.00"   
        linha = linha + 1    
     
    planilha.cell(linha-1,  1).font = fontMasterPreta
    planilha.cell(linha-1,  2).font = fontMasterPreta
    planilha.cell(linha-1,  3).font = fontMasterPreta
    planilha.cell(linha-1,  4).font = fontMasterPreta
    planilha.cell(linha-1,  5).font = fontMasterPreta
    planilha.cell(linha-1,  6).font = fontMasterPreta
    
    
    
    
    
    

"""

fontMasterBranca = Font(color='FFFFFFFF', bold=True, size=12)
fontNegrito      = Font(color='00000000', bold=True)
fontAzul         = Font(color='FF0000FF')
fontVermelha     = Font(color='FFFF0000')
fontPreta        = Font(color='00000000')

planilha.cell(numLinha,  14).alignment = Alignment(wrap_text = True)  

planilha.cell(numLinha,  1).border = bordaB
planilha.cell(numLinha,  14).border = bordaB    
planilha.merge_cells('A'+str(numLinha)+':M'+str(numLinha))

planilha.cell(numLinha,  7).alignment = Alignment(horizontal='center')
planilha.merge_cells('G'+str(numLinha)+':M'+str(numLinha))
formatarCelulaDetalhe(planilha.cell(numLinha, 14, "")                              ,bq.AREA, 'CH')


formatarCelulaTitulo(planilha.cell(linha, 2, 'Total de notas da Série'))
formatarCelulaTitulo(planilha.cell(linha, 4, 'Total notas Não impactadas'))
formatarCelulaTitulo(planilha.cell(linha, 3, 'Total de notas Impactadas'))

planilha.column_dimensions['A'].width = 19
planilha.column_dimensions['B'].width = 19
planilha.column_dimensions['C'].width = 22
planilha.column_dimensions['D'].width = 21
planilha.column_dimensions['E'].width = 11
planilha.column_dimensions['F'].width = 90


fontMasterPreta  = Font(color='00000000', bold=True, size=12)
fontMasterBranca = Font(color='FFFFFFFF', bold=True, size=12)
fontNegrito      = Font(color='00000000', bold=True)
fontAzul         = Font(color='FF0000FF')
fontVermelha     = Font(color='FFFF0000')
fontPreta        = Font(color='00000000')

borda = Border(
            left   = Side(style='thin'),
            right  = Side(style='thin'),
            top    = Side(style='thin'),
            bottom = Side(style='thin')
)

bordaB = Border(
            left   = Side(style='medium'),
            right  = Side(style='medium'),
            top    = Side(style='medium'),
            bottom = Side(style='medium')
)

fundoCelula = PatternFill(
                start_color = 'FFE0E0E0',
                end_color   = 'FFE0E0E0',
                fill_type   = 'solid'
)

fundoCelulaAzul = PatternFill(
                start_color = '0066CCCC',
                end_color   = '0066CCCC',
                fill_type   = 'solid'
)

fundoCelulaBranco = PatternFill(
                start_color = 'FFFFFFFF',
                end_color   = 'FFFFFFFF',
                fill_type   = 'solid'
)

fundoCelulaPreto = PatternFill(
                start_color = '00000000',
                end_color   = '00000000',
                fill_type   = 'solid'
)


def formatarCelulaTituloMasterA(celula):
    celula.font = fontMasterPreta
    celula.fill = fundoCelulaAzul
    celula.alignment = alignment = Alignment(horizontal='center') 
    celula.border = bordaB
    
def formatarCelulaTituloMasterP(celula):
    celula.font = fontMasterBranca
    celula.fill = fundoCelulaPreto  
    celula.alignment = alignment = Alignment(horizontal='center')
    celula.border = bordaB
    
def formatarCelulaTitulo(celula):
    celula.font = fontNegrito
    celula.border = borda

def formatarCelulaDetalhe(celula, area, dataType):
    if  area == 'PROTOCOLADO': 
        celula.font = fontAzul
        celula.fill = fundoCelula
    else: 
        celula.font = fontPreta
        celula.fill = fundoCelulaBranco
    
    if area == 'ERRO':
        celula.font = fontVermelha
        
    if  dataType == 'MA':
        celula.number_format = "mmmm-yy"
    
    if  dataType == 'VL': 
        celula.number_format = "#,##0.00"
        if celula.value < 0.00:
            celula.font = fontVermelha
    
    if  dataType == 'PC': 
        celula.number_format = "#,##0.00%"
    
    if  dataType == 'NU': 
        celula.number_format = "0"
        
    celula.border = borda  
"""

if __name__ == "__main__":
    global arquivo_destino
    arquivo_destino= ""
    comum.log("-"*100)
    log("#### ",dtf(), " INICIO DO RELATORIO RELATORIO_GIA_TOTAIS_POR_CFOP ####")
    variaveis = comum.carregaConfiguracoes(configuracoes)

    ret = processar()
    if (ret > 0) :
        if(arquivo_destino):
            if os.path.isfile(arquivo_destino):
                os.remove(arquivo_destino)
    log("## Código de execução = ", ret)
    log("#### ",dtf(), " FIM DO RELATORIO RELATORIO_GIA_TOTAIS_POR_CFOP  ####")
    sys.exit(ret)







