#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-

"""
----------------------------------------------------------------------------------------------
  SISTEMA ..: 
  MODULO ...: 
  SCRIPT ...: Relatório gera relatorio REGISTRO 1400
  CRIACAO ..: 06/04/2021
  AUTOR ....: Airton Borges - Kyros Consultoria 
  DESCRICAO : Gera relatorio em excel por intervalo de data
  ALTERACAO :
  (1)
  TAG: EDUARDOF20210623
  DATA: 2021-06-23
  RESP: EDUARDOF@KYROS.COM.BR
  DESRICAO: ADICIONAR INFORMACAO NA TABELA INVA
----------------------------------------------------------------------------------------------
    Exemplo de comando: ./registro1400SP.py 202103 <INSCRICAO_ESTADUAL>
    Diretório: /arquivos/registro_1400/RELATORIOS/
    Exemplo: /arquivos/registro_1400/RELATORIOS/registro_032021_inscricao_estadual.xlsx
----------------------------------------------------------------------------------------------
    2021/06/03 - Airton.  Adaptado para cálculo do 1400 de SP. Baseado no registro1400 do RJ.
    Documentação:  "01 - Teshuva_RMSV0_Registro 1400 SP_V2.docx"
----------------------------------------------------------------------------------------------
    20210831 - Airton Borges - Kyros
    Adaptação para o novo Painel de execuções.
----------------------------------------------------------------------------------------------    
    2021/11/17 - Airton.  Alterado a forma de cálculo
    Documentação:  PTITES-985
----------------------------------------------------------------------------------------------
    28/12/2021 - Welber Pena de Sousa - Kyros Tecnologia
        ALT-001 - Alterado os diretorios de arquivos  e o diretorio de Protocolados
    10/03/2022  - Eduardo da Silva Ferreira - Kyros Tecnologia
                - [PTITES-1688] Padrão de diretórios do SPARTA                 
----------------------------------------------------------------------------------------------
"""

#### PATRONIZACAO PARA O PAINEL DE EXECUCOES....
import sys
import os
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)
import configuracoes
import comum
from comum import log
import util
import sql
comum.log.gerar_log_em_arquivo = False
comum.carregaConfiguracoes(configuracoes)
banco=sql.geraCnxBD(configuracoes)

#### PATRONIZACAO PARA O PAINEL DE EXECUCOES....

import datetime
import cx_Oracle
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook
from pathlib import Path
sys.path.append( os.path.join( os.path.realpath('..'), "modulosPython" ) )
nome_script = os.path.basename( sys.argv[0] ).replace('.py', '')
#name_script = os.path.basename(__file__).split('.')[0]

fontMasterPreta  = Font(color='00000000', bold=True, size=12)
fontNegrito      = Font(color='00000000', bold=True)

### ALT-001 - Alterado o valor do diretorio_arquivos no arquivo de cfg. (buscando os arquivo no /portaloptrib )
# INICIO PTITES-1688 
dir_arquivos        = os.path.join(os.path.dirname(configuracoes.dir_geracao_arquivos), 'REGISTRO_1400', 'RELATORIOS') # PTITES-1688 # configuracoes.diretorio_arquivos
dir_protocolados    = os.path.join(os.path.dirname(configuracoes.dir_entrada), 'SPED_FISCAL', 'PROTOCOLADOS') # PTITES-1688 # configuracoes.diretorio_protocolados
log("-"* 14)
log('# - DIRETORIO ARQUIVOS REG. 1400  .:', dir_arquivos)
log('# - DIRETORIO ARQUIVOS PROTOCOLADOS:', dir_protocolados)
log("-"* 14)   
# FIM PTITES-1688 

ret = 0
dir_destino = "" 


def nome_arquivo(mascara, diretorio):
    qdade = 0
    nomearq = "" 
    directory = Path(diretorio)
    files = directory.glob(mascara)
    sorted_files = sorted(files, reverse=False)
    if sorted_files:
        for f in sorted_files:
            qdade = qdade + 1
            nomearq = f
    else: 
        log("-"*100)
        log('ERRO:    Arquivo %s não está na pasta %s'%(mascara,diretorio))
        log("-"*100)
    return(nomearq)

def ultimodia(ano,mes):
   return(31 if mes == 12 else datetime.date.fromordinal((datetime.date(ano,mes+1,1)).toordinal()-1).day)    

def processar():
    vDataIni = ""
    vDataFim = ""
    IE       = "" 
    ufi      = "SP"
    ret = 0

    if (    len(sys.argv) == 3 
        and len(sys.argv[1])==6  
        and int(sys.argv[1][0:2])>0 
        and int(sys.argv[1][0:2])<13
        and int(sys.argv[1][2:6])<=datetime.datetime.now().year
        and int(sys.argv[1][2:6])>(datetime.datetime.now().year)-50 ):
        vPeriodo = sys.argv[1]
        IE       = sys.argv[2] 
    elif (  len(sys.argv) == 5 
        and len(sys.argv[2])==6  
        and int(sys.argv[2][0:2])>0 
        and int(sys.argv[2][0:2])<13
        and int(sys.argv[2][2:6])<=datetime.datetime.now().year
        and int(sys.argv[2][2:6])>(datetime.datetime.now().year)-50 ):
        vPeriodo = sys.argv[2]
        IE       = sys.argv[3] 
    else:
        log("-" * 100)
        log("#### ")
        log('#### ERRO - Erro nos parametros do script.')
        log("#### ")
        log('#### Exemplo de como deve ser :')
        log('####      %s  <MMYYYY>  <IE>'%(sys.argv[0]))
        log("#### ")
        log('#### Onde')
        log('####      <MMYYYY> = 032021')
        log('####       <IE>     = INSCRICAO ESTADUAL  ')
        log('#### segue um exemplo %s 032021 108383949112'%(sys.argv[0]))
        log("#### ")
        log('#### ')
        log("-" * 100)
        log("")
        log("Retorno = 99") 
        ret = 99
        return(99)  
    
    vDataIni  ='01' + str(vPeriodo)
    vAno      = str(vPeriodo[2:6])
    vMes      = str(vPeriodo[0:2])
    UltDiaMes = ultimodia(int(vPeriodo[2:6]), int(vPeriodo[0:2]))
    vDataFim  = str(UltDiaMes)+str(vPeriodo) 
    vUF       = "SP"
    
#    vUF=retornaUF(IE,banco)
#    
#    if (vUF == ""):
#        log("#### ERRO - Não foi possível determinar a UF pela IE informada.")
#        return(99)
#    if (vUF != "SP"):
#        log("#### ERRO - A IE informada não é de SP.")
#        return(99)

    ### INICIO ALT-001 - alterado o diretorio alvo onde os arquivos PROTOCOLADOS estao .
    dir_busca = os.path.join(dir_protocolados, vUF, vAno, vMes )
    log("Diretório onde deve estar o arquivo protocolado: ", dir_busca)
    arq_protocolado=""
    mascara_protocolado = "SPED_"+vPeriodo+"_"+vUF+"_"+IE+"_PROT*.txt"
    arq_protocolado = nome_arquivo(mascara_protocolado, dir_busca)
    ### FIM ALT-001 ### 
    
    log("-"* 140)
    log('# - Periodo.............:', vPeriodo)
    log('# - UF..................:', vUF)
    log('# - Inscricao estadual..:', IE)
    log('# - vDataIni............:',vDataIni)
    log('# - vDataFim............:',vDataFim)
    log('# - Arquivo PROTOCOLADO.:',arq_protocolado)
    log("-"* 140)   


    if (arq_protocolado == ""):
        log("-"*100)
        log("ERRO:    Não foi encontrado um arquivo PROTOCOLADO que atende a mascara: ", mascara_protocolado)
        log("-"*100)
        return(99)


    #### Monta caminho e nome do destino
    dir_base =  os.path.join(dir_arquivos, vUF)   
    dir_destino = os.path.join(dir_base, vAno, vMes) # PTITES-1688 , 'REGISTRO_1400')  

    if not os.path.isdir(dir_destino) :
        os.makedirs(dir_destino) 
    
    arquivo_destino = IE+'_Valores_Agregados_1400_'+vPeriodo+'.xlsx'
    nome_relatorio = os.path.join(dir_destino,arquivo_destino)
 
    log("Arquivo do relatório de saída = ", nome_relatorio)
   
    arquivo = open(nome_relatorio, 'w')
    arquivo.close()  


    #### Cria a planilha em memória....
    arquivo_excel = Workbook()
    planilha0 = arquivo_excel.active
    planilha0.title = "PARAMETRO"
    
######################################################################################################################################################################################################
#aba 0 "Parametro"
#aba 0 "Parametro"
#aba 0 "Parametro"
#aba 0 "Parametro"
#aba 0 "Parametro"
#aba 0 "Parametro"
#######################################################################################################################################################################################################
    log("")
    log("# - Início do processamento da aba 0: '1-Parametro'.")

    #### CABEÇALHO 
    #### CABEÇALHO 
    #### CABEÇALHO 

    vLinha = 1
    planilha0.cell(vLinha,1,"PARÂMETRO - REGISTRO 1400")
    planilha0.cell(vLinha,1).font=Font(bold=True)
    planilha0.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha0.merge_cells('A'+ str(vLinha) + ':D' + str(vLinha))
    
    vLinha = vLinha + 1
    planilha0.cell(vLinha,1,"Insc Estadual...: "+IE)
    planilha0.cell(vLinha,1).font=Font(bold=True)
    planilha0.merge_cells('A'+ str(vLinha) + ':D' + str(vLinha))

    vLinha = vLinha + 1
    planilha0.cell(vLinha,1,"Mes/Ano........: "+str(vDataIni[0:2])+"/"+str(vDataIni[2:4])+"/"+str(vDataIni[4:8])+" a "+str(vDataFim[0:2])+"/"+str(vDataFim[2:4])+"/"+str(vDataFim[4:8]))
    planilha0.cell(vLinha,1).font=Font(bold=True)
    planilha0.merge_cells('A'+ str(vLinha) + ':D' + str(vLinha))
    
    vLinha = vLinha + 1
    vColuna = 1
    for nColuna in ("MUNICIPIO","COD_MUNICIPIO","RATEIO","CODIGO_PVA"):
        planilha0.cell(vLinha,vColuna,nColuna)
        planilha0.cell(vLinha,vColuna).font=Font(bold=True)
        planilha0.cell(vLinha,vColuna).alignment = Alignment(horizontal='center')
        vColuna = vColuna + 1
 
    #### DADOS
    #### DADOS
    #### DADOS
    
    dados = []
    dados = aba_resumo(vUF,vAno,vMes,IE,banco)  
    plinhaP0 = vLinha + 1 ## primeira linha com dados a serem somados.
    
    for linha in dados:
        vLinha = vLinha + 1
        planilha0.cell(vLinha,1,linha[0])
        planilha0.cell(vLinha,2,linha[1])
        planilha0.cell(vLinha,3,linha[2])
        planilha0.cell(vLinha,4,linha[3])
        planilha0.cell(vLinha,3).number_format = "#,###############0.000000000000000"
        
    ulinhaP0 = vLinha

    #### TOTAIS
    #### TOTAIS
    #### TOTAIS
    
    vLinha = vLinha + 1
    planilha0.cell(vLinha,1,"TOTAIS:")
    planilha0.cell(vLinha,1).font=Font(bold=True)

    planilha0.cell(vLinha,3,"=SUM(C"+str(plinhaP0)+":C"+str(ulinhaP0)+")")
    planilha0.cell(vLinha,3).font=Font(bold=True)
    planilha0.cell(vLinha,3).number_format = "#,##0.00 %"

    #### FORMATAÇAO
    #### FORMATAÇAO
    #### FORMATAÇAO

    planilha0.column_dimensions['A'].width = 65  
    planilha0.column_dimensions['B'].width = 20   
    planilha0.column_dimensions['C'].width = 20  
    planilha0.column_dimensions['D'].width = 20  

    #### GRAVA A PLANILHA
    #### GRAVA A PLANILHA
    #### GRAVA A PLANILHA

    arquivo_excel.save(nome_relatorio)

    log("")
    log("# - Final do processamento da aba 0: '1-Parametro'.")
#######################################################################################################################################################################################################
####aba 1 Resumo CFOP - TELECOM SAIDA"
####aba 1 Resumo CFOP - TELECOM SAIDA"
####aba 1 Resumo CFOP - TELECOM SAIDA"
####aba 1 Resumo CFOP - TELECOM SAIDA"
####aba 1 Resumo CFOP - TELECOM SAIDA"
####aba 1 Resumo CFOP - TELECOM SAIDA"
#######################################################################################################################################################################################################
    log("")
    log("# - Início do processamento da aba 1: '2A-Resumo CFOP - Telecom Saída'.")
    planilha1 = arquivo_excel.create_sheet("RESUMO CFOP TELECOM", 1)

    #### CABEÇALHO 
    #### CABEÇALHO 
    #### CABEÇALHO 

    vLinha = 1
    planilha1.cell(vLinha,1,"RESUMO CFOP - TELECOM SAÍDA")
    planilha1.cell(vLinha,1).font=Font(bold=True)
    planilha1.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha1.merge_cells('A'+ str(vLinha) + ':F' + str(vLinha))
    
    vLinha = vLinha + 1
    planilha1.cell(vLinha,1,"Insc Estadual...: "+IE)
    planilha1.cell(vLinha,1).font=Font(bold=True)
    planilha1.merge_cells('A'+ str(vLinha) + ':F' + str(vLinha))

    vLinha = vLinha + 1
    planilha1.cell(vLinha,1,"Mes/Ano........: "+str(vDataIni[0:2])+"/"+str(vDataIni[2:4])+"/"+str(vDataIni[4:8])+" a "+str(vDataFim[0:2])+"/"+str(vDataFim[2:4])+"/"+str(vDataFim[4:8]))
#    planilha1.cell(vLinha,1,"Mes/Ano........: "+str(vDataIni)+" a "+ str(vDataFim))
    planilha1.cell(vLinha,1).font=Font(bold=True)
    planilha1.merge_cells('A'+ str(vLinha) + ':F' + str(vLinha))
    
    vLinha = vLinha + 1
    vColuna = 1
    for nColuna in ("CFOP","VLR_CONTABIL","VLR_BASE_ICMS","VLR_ICMS","VLR_ISENTAS","VLR_OUTRAS"):
        planilha1.cell(vLinha,vColuna,nColuna)
        planilha1.cell(vLinha,vColuna).font=Font(bold=True)
        planilha1.cell(vLinha,vColuna).alignment = Alignment(horizontal='center')
        vColuna = vColuna + 1

    #### DADOS
    #### DADOS
    #### DADOS
    
    dadosCFOP = []
    dadosCFOP = aba_resumoCFOP(vDataIni,vDataFim,IE,banco)
    
    plinhaP1A = vLinha + 1 ## primeira linha com dados a serem somados.
    
    for linha in dadosCFOP:
        vLinha = vLinha + 1
        planilha1.cell(vLinha,1,linha[0])
        planilha1.cell(vLinha,2,linha[1])
        planilha1.cell(vLinha,3,linha[2])
        planilha1.cell(vLinha,4,linha[3])
        planilha1.cell(vLinha,5,linha[4])
        planilha1.cell(vLinha,6,linha[5])
        planilha1.cell(vLinha,2).number_format = "#,##0.00"
        planilha1.cell(vLinha,3).number_format = "#,##0.00"
        planilha1.cell(vLinha,4).number_format = "#,##0.00"
        planilha1.cell(vLinha,5).number_format = "#,##0.00"
        planilha1.cell(vLinha,6).number_format = "#,##0.00"
        
    ulinhaP1A = vLinha  ## última linha com dados a serem somados.    
 
    #### TOTAIS
    #### TOTAIS
    #### TOTAIS
    
    if(ulinhaP1A >= 5):
        planilha1.append(["TOTAIS","=SUM(B5:B"+str(vLinha)+")","=SUM(C5:C"+str(vLinha)+")","=SUM(D5:D"+str(vLinha)+")","=SUM(E5:E"+str(vLinha)+")","=SUM(F5:F"+str(vLinha)+")"])
    else:
        planilha1.append(["TOTAIS",0.00,0.00,0.00,0.00,0.00])

    vLinha = vLinha + 1
    ls1 = vLinha

    #### FORMATAÇAO
    #### FORMATAÇAO
    #### FORMATAÇAO
    planilha1.merge_cells('A1:F1')
    planilha1.merge_cells('A2:F2')
    planilha1.merge_cells('A3:F3')

    planilha1.column_dimensions['A'].width = 15  
    planilha1.column_dimensions['B'].width = 20   
    planilha1.column_dimensions['C'].width = 20  
    planilha1.column_dimensions['D'].width = 20  
    planilha1.column_dimensions['E'].width = 20  
    planilha1.column_dimensions['F'].width = 20   

    planilha1.cell(vLinha,1).font=Font(bold=True)
    planilha1.cell(vLinha,2).font=Font(bold=True)
    planilha1.cell(vLinha,3).font=Font(bold=True)
    planilha1.cell(vLinha,4).font=Font(bold=True)
    planilha1.cell(vLinha,5).font=Font(bold=True)
    planilha1.cell(vLinha,6).font=Font(bold=True)
   
    for lin in range(5,vLinha+1):
        planilha1.cell(lin,2).number_format = "#,##0.00"
        planilha1.cell(lin,3).number_format = "#,##0.00"
        planilha1.cell(lin,4).number_format = "#,##0.00"
        planilha1.cell(lin,5).number_format = "#,##0.00"
        planilha1.cell(lin,6).number_format = "#,##0.00"

    #### GRAVA A PLANILHA
    #### GRAVA A PLANILHA
    #### GRAVA A PLANILHA

    arquivo_excel.save(nome_relatorio)
 
    log("")
    log("# - Final do processamento da aba 1: '2A-Resumo CFOP - Telecom Saída'.")
#####################################################




#######################################################################################################################################################################################################
####aba 1-2 Resumo CFOP - TELECOM ENTRADA"
####aba 1-2 Resumo CFOP - TELECOM ENTRADA"
####aba 1-2 Resumo CFOP - TELECOM ENTRADA"
####aba 1-2 Resumo CFOP - TELECOM ENTRADA"
####aba 1-2 Resumo CFOP - TELECOM ENTRADA"
####aba 1-2 Resumo CFOP - TELECOM ENTRADA"
#######################################################################################################################################################################################################
    log("")
    log("# - Início do processamento da aba 1: '2B-Resumo CFOP - Telecom Entrada'.")
    vLinha = vLinha + 4
    
    #### CABEÇALHO 
    #### CABEÇALHO 
    #### CABEÇALHO 
    planilha1.cell(vLinha,1,"RESUMO CFOP - TELECOM ENTRADA")
    planilha1.cell(vLinha,1).font=Font(bold=True)
    planilha1.cell(vLinha,1).alignment = Alignment(horizontal='center')
#20211108    planilha1.merge_cells('A'+ str(vLinha) + ':F' + str(vLinha))
    planilha1.merge_cells('A'+ str(vLinha) + ':D' + str(vLinha))
    
    vLinha = vLinha + 1
    planilha1.cell(vLinha,1,"Insc Estadual...: "+IE)
    planilha1.cell(vLinha,1).font=Font(bold=True)
#20211108    planilha1.merge_cells('A'+ str(vLinha) + ':F' + str(vLinha))
    planilha1.merge_cells('A'+ str(vLinha) + ':D' + str(vLinha))

    vLinha = vLinha + 1
    planilha1.cell(vLinha,1,"Mes/Ano........: "+str(vDataIni[0:2])+"/"+str(vDataIni[2:4])+"/"+str(vDataIni[4:8])+" a "+str(vDataFim[0:2])+"/"+str(vDataFim[2:4])+"/"+str(vDataFim[4:8]))
#    planilha1.cell(vLinha,1,"Mes/Ano........: "+str(vDataIni)+" a "+ str(vDataFim))
    planilha1.cell(vLinha,1).font=Font(bold=True)
#20211108    planilha1.merge_cells('A'+ str(vLinha) + ':F' + str(vLinha))
    planilha1.merge_cells('A'+ str(vLinha) + ':D' + str(vLinha))
    
    vLinha = vLinha + 1
    vColuna = 1
#20211108    for nColuna in ("CFOP","VLR_CONTABIL","VLR_BASE_ICMS","VLR_ICMS","VLR_ISENTAS","VLR_OUTRAS"):
    for nColuna in ("CFOP","VLR_CONTABIL","VLR_BASE_ICMS","VLR_ICMS"):
        planilha1.cell(vLinha,vColuna,nColuna)
        planilha1.cell(vLinha,vColuna).font=Font(bold=True)
        planilha1.cell(vLinha,vColuna).alignment = Alignment(horizontal='center')
        vColuna = vColuna + 1
 
    #### DADOS
    #### DADOS
    #### DADOS
    
    plinhaP1B = vLinha + 1 ## primeira linha com dados a serem somados.
    
    dadosCFOP2 = []
    dadosCFOP2 = aba_resumoCFOP2(arq_protocolado)
   
    for linha in dadosCFOP2:
        vLinha = vLinha + 1
        planilha1.cell(vLinha,1,linha[0])
        planilha1.cell(vLinha,2,linha[1])
        planilha1.cell(vLinha,3,linha[2])
        planilha1.cell(vLinha,4,linha[3])

        planilha1.cell(vLinha,2).number_format = "#,##0.00"
        planilha1.cell(vLinha,3).number_format = "#,##0.00"
        planilha1.cell(vLinha,4).number_format = "#,##0.00"
        
    ulinhaP1B = vLinha ## última linha com dados a serem somados.    
 
    #### TOTAIS
    #### TOTAIS
    #### TOTAIS

    vLinha = vLinha + 1
    planilha1.cell(vLinha,1,"TOTAIS")
    planilha1.cell(vLinha,2,"=SUM(B"+str(plinhaP1B)+":B"+str(ulinhaP1B)+")")
    planilha1.cell(vLinha,3,"=SUM(C"+str(plinhaP1B)+":C"+str(ulinhaP1B)+")")
    planilha1.cell(vLinha,4,"=SUM(D"+str(plinhaP1B)+":D"+str(ulinhaP1B)+")")
  #  planilha1.cell(vLinha,5,"=SUM(E"+str(plinhaP1B)+":E"+str(ulinhaP1B)+")")
  #  planilha1.cell(vLinha,6,"=SUM(F"+str(plinhaP1B)+":F"+str(ulinhaP1B)+")")
    planilha1.cell(vLinha,2).number_format = "#,##0.00"
    planilha1.cell(vLinha,3).number_format = "#,##0.00"
    planilha1.cell(vLinha,4).number_format = "#,##0.00"
  #  planilha1.cell(vLinha,5).number_format = "#,##0.00"
  #  planilha1.cell(vLinha,6).number_format = "#,##0.00"
    planilha1.cell(vLinha,1).font=Font(bold=True)
    planilha1.cell(vLinha,2).font=Font(bold=True)
    planilha1.cell(vLinha,3).font=Font(bold=True)
    planilha1.cell(vLinha,4).font=Font(bold=True)
  #  planilha1.cell(vLinha,5).font=Font(bold=True)
  #  planilha1.cell(vLinha,6).font=Font(bold=True)
    
    vLinha = vLinha + 4
    planilha1.cell(vLinha,1,"VLR_CALCULO")
    planilha1.cell(vLinha,2,"=B"+str(ulinhaP1A+1)+"-B"+str(ulinhaP1B+1))
    planilha1.cell(vLinha,1).font=Font(bold=True)
    planilha1.cell(vLinha,2).font=Font(bold=True)
    planilha1.cell(vLinha,2).number_format = "#,##0.00"

    ####SAíDAS - ENTRADAS
    vLinha = vLinha +3
    ts = 0.00
    te = 0.00
    for qs in range(plinhaP1A,ulinhaP1A+1):
#        log("plinhaP1A,ulinhaP1A,ts,qs,float(planilha1.cell(qs,2).value)",plinhaP1A, ulinhaP1A, ts, qs, float(planilha1.cell(qs,2).value))
        ts = ts + float(planilha1.cell(qs,2).value)
    for qe in range(plinhaP1B,ulinhaP1B+1):
#        log("plinhaP1B,ulinhaP1B,te,qe,float(planilha1.cell(qe,2).value)",plinhaP1B, ulinhaP1B, te, qe, float(planilha1.cell(qe,2).value))
        te = te + float(planilha1.cell(qe,2).value)
    vc = ts - te
    #planilha1.cell(vLinha,1,"S - E = ")
    #planilha1.cell(vLinha,2,ts - te)
    #planilha1.cell(vLinha,1).font=Font(bold=True)
    #planilha1.cell(vLinha,2).font=Font(bold=True)
    #planilha1.cell(vLinha,2).number_format = "#,##0.00"

    #### GRAVA A PLANILHA
    #### GRAVA A PLANILHA
    #### GRAVA A PLANILHA
    arquivo_excel.save(nome_relatorio)
 
    log("")
    log("# - Final do processamento da aba 1: '2B-Resumo por CFOP - Telecom Entrada'.")

#####################################################




#    vc = 2007776952.48





#######################################################################################################################################################################################################
####aba 2 CALCULO DE VALORES AGREGADOS
####aba 2 CALCULO DE VALORES AGREGADOS
####aba 2 CALCULO DE VALORES AGREGADOS
####aba 2 CALCULO DE VALORES AGREGADOS
####aba 2 CALCULO DE VALORES AGREGADOS
####aba 2 CALCULO DE VALORES AGREGADOS
#######################################################################################################################################################################################################
    planilha2 = arquivo_excel.create_sheet("VALOR AGREGADO", 2) 
    log("")
    log("# - Início do processamento da aba 2: '3-Cálculo de valores agregados'.")
    vLinha = 1
    
    #### CABEÇALHO 
    #### CABEÇALHO 
    #### CABEÇALHO 
    planilha2.cell(vLinha,1,"CALCULO DE VALORES AGREGADOS")
    planilha2.cell(vLinha,1).font=Font(bold=True)
    planilha2.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha2.merge_cells('A'+ str(vLinha) + ':E' + str(vLinha))
    
    vLinha = vLinha + 1
    planilha2.cell(vLinha,1,"Insc Estadual...: "+IE)
    planilha2.cell(vLinha,1).font=Font(bold=True)
    planilha2.merge_cells('A'+ str(vLinha) + ':E' + str(vLinha))

    vLinha = vLinha + 1
    planilha2.cell(vLinha,1,"Mes/Ano........: "+str(vDataIni[0:2])+"/"+str(vDataIni[2:4])+"/"+str(vDataIni[4:8])+" a "+str(vDataFim[0:2])+"/"+str(vDataFim[2:4])+"/"+str(vDataFim[4:8]))
    planilha2.cell(vLinha,1).font=Font(bold=True)
    planilha2.merge_cells('A'+ str(vLinha) + ':E' + str(vLinha))
    
    vLinha = vLinha + 1
    vColuna = 1
    for nColuna in ("MUNICIPIO","CODIGO_MUNICIPIO","RATEIO","CODIGO_PVA","VALOR"):
        planilha2.cell(vLinha,vColuna,nColuna)
        planilha2.cell(vLinha,vColuna).font=Font(bold=True)
        planilha2.cell(vLinha,vColuna).alignment = Alignment(horizontal='center')
        vColuna = vColuna + 1
 

    #### DADOS
    #### DADOS
    #### DADOS
    plinhap2 = vLinha + 1
    for linha in dados:
        vLinha = vLinha + 1
        planilha2.cell(vLinha,1,linha[0])
        planilha2.cell(vLinha,2,linha[1])
        planilha2.cell(vLinha,3,linha[2])
        planilha2.cell(vLinha,3).number_format = "#,###############0.000000000000000"
        planilha2.cell(vLinha,4,linha[3])
        planilha2.cell(vLinha,5,float(linha[2]) * vc)
        planilha2.cell(vLinha,5).number_format = "#,##0.00"

    #### TOTAIS
    #### TOTAIS
    #### TOTAIS

    vLinha = vLinha + 1

    planilha2.cell(vLinha,1,"TOTAIS")
    planilha2.cell(vLinha,1).font=Font(bold=True)    
    planilha2.cell(vLinha,3,"=SUM(C"+str(plinhap2)+":C"+str(vLinha-1)+")")
    planilha2.cell(vLinha,3).font=Font(bold=True)
    planilha2.cell(vLinha,3).number_format = "#,##0.00 %"    
    planilha2.cell(vLinha,5,"=SUM(E"+str(plinhap2)+":E"+str(vLinha-1)+")")
    planilha2.cell(vLinha,5).font=Font(bold=True)
    planilha2.cell(vLinha,5).number_format = "#,##0.00"
    
    planilha2.column_dimensions['A'].width = 60  
    planilha2.column_dimensions['B'].width = 20   
    planilha2.column_dimensions['C'].width = 20  
    planilha2.column_dimensions['D'].width = 20  
    planilha2.column_dimensions['E'].width = 20  

    #### GRAVA A PLANILHA
    #### GRAVA A PLANILHA
    #### GRAVA A PLANILHA
    arquivo_excel.save(nome_relatorio)
 
    log("")
    log("# - Final do processamento da aba 2: '3-Cálculo de valores agregados'.")

#####################################################


#######################################################################################################################################################################################################
####aba 3 REGISTRO 1400
####aba 3 REGISTRO 1400
####aba 3 REGISTRO 1400
####aba 3 REGISTRO 1400
####aba 3 REGISTRO 1400
####aba 3 REGISTRO 1400
#######################################################################################################################################################################################################
    planilha3 = arquivo_excel.create_sheet("REGISTRO 1400", 3) 
    log("")
    log("# - Início do processamento da aba 3: '4-Registro 1400'.")
    vLinha = 1
    
    #### CABEÇALHO 
    #### CABEÇALHO 
    #### CABEÇALHO 
    planilha3.cell(vLinha,1,"REGISTRO 1400")
    planilha3.cell(vLinha,1).font=Font(bold=True)
    planilha3.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha3.merge_cells('A'+ str(vLinha) + ':E' + str(vLinha))
    
    vLinha = vLinha + 1
    planilha3.cell(vLinha,1,"Insc Estadual...: "+IE)
    planilha3.cell(vLinha,1).font=Font(bold=True)
    planilha3.merge_cells('A'+ str(vLinha) + ':E' + str(vLinha))

    vLinha = vLinha + 1
    planilha3.cell(vLinha,1,"Mes/Ano........: "+str(vDataIni[0:2])+"/"+str(vDataIni[2:4])+"/"+str(vDataIni[4:8])+" a "+str(vDataFim[0:2])+"/"+str(vDataFim[2:4])+"/"+str(vDataFim[4:8]))
    planilha3.cell(vLinha,1).font=Font(bold=True)
    planilha3.merge_cells('A'+ str(vLinha) + ':E' + str(vLinha))
    
    vLinha = vLinha + 1
    vColuna = 1
    for nColuna in ("REGISTRO","CODIGO_ITEM","MUNICIPIO","VALOR","REGISTRO"):
        planilha3.cell(vLinha,vColuna,nColuna)
        planilha3.cell(vLinha,vColuna).font=Font(bold=True)
        planilha3.cell(vLinha,vColuna).alignment = Alignment(horizontal='center')
        vColuna = vColuna + 1

    #### DADOS
    #### DADOS
    #### DADOS
    plinhap3 = vLinha + 1
    for linha in dados:
        vLinha = vLinha + 1
        planilha3.cell(vLinha,1,"1400")
        planilha3.cell(vLinha,2,linha[3])
        planilha3.cell(vLinha,3,linha[1])
        planilha3.cell(vLinha,4,round(float(linha[2]) * vc,2))
        planilha3.cell(vLinha,4).number_format = "#,##0.00"
        planilha3.cell(vLinha,5,"|1400|"+linha[3]+"|"+str(linha[1])+"|"+str(   round(float(linha[2]) * vc,2)   ).replace(".", ",")    +"|")
        
    planilha3.column_dimensions['A'].width = 10  
    planilha3.column_dimensions['B'].width = 20   
    planilha3.column_dimensions['C'].width = 20  
    planilha3.column_dimensions['D'].width = 20  
    planilha3.column_dimensions['E'].width = 60  
    
 
    #### GRAVA A PLANILHA
    #### GRAVA A PLANILHA
    #### GRAVA A PLANILHA
    arquivo_excel.save(nome_relatorio)
 
    log("")
    log("# - Final do processamento da aba 3: '4-Registro 1400'.")
#####################################################

    # INICIO <<EDUARDOF20210623>>
    if vLinha >= plinhap3:
        try:
            log("# Inicializa inva")
            cmd_sql1 = """
                DELETE FROM openrisow.inva
                WHERE emps_cod  = 'TBRA'
                AND   fili_cod  = '0001'
                AND   inva_data >= TO_DATE('%s','dd/mm/yyyy')
                AND   inva_data <= TO_DATE('%s','dd/mm/yyyy')
                """ % (str(vDataIni[0:2])+"/"
                       + str(vDataIni[2:4])+"/"
                       + str(vDataIni[4:8]),
                       str(vDataFim[0:2])+"/"
                       + str(vDataFim[2:4])+"/"
                       + str(vDataFim[4:8]))
            vLinha = 1
            for linha in dados:
                if vLinha == 1:
#                    log("-"*100)
#                    log("query = ", cmd_sql1)
#                    log("-"*100)
#                    input("Deleta ?")


                    banco.executa(cmd_sql1)



                cmd_sql2 = """
                    INSERT INTO openrisow.inva (emps_cod,
                    fili_cod,
                    inva_tipo,
                    inva_cod,
                    inva_ori_calc,
                    inva_data,
                    inva_valor,
                    mibge_cod_mun)
                    VALUES ('TBRA',
                    '0001',
                    'U',
                    'SPDIPAM24',
                    'M',
                    to_date('%s','dd/mm/yyyy'),
                    %.2f,
                    '%s'
                    )
                    """ % (str(vDataFim[0:2])+"/"
                           + str(vDataFim[2:4])+"/"
                           + str(vDataFim[4:8]),
                           round(float(linha[2]) * vc, 2),
                           linha[1]
                           )
#                log("-"*100)
#                log("query = ", cmd_sql2)
#                log(round(float(linha[2]) * vc, 2))
#                log("-"*100)
#                input("Insere ?")


                banco.executa(cmd_sql2)


                vLinha += 1
            banco.commit()
            log("# Gravação realizado com sucesso no inva")
            return(0)
        except Exception as e:
            log("Falha na gravação no inva " + str(e))
            try:
                banco.rollback()
            except log(0):
                pass
            return(1)
    else:
        return(0)

    # FIM <<EDUARDOF20210623>>

def retornaUF(IE,banco):
    vUF = ""
    query="""
    select distinct f.unfe_sig  from openrisow.filial f where f.fili_cod_insest='%s'
    """%(IE)

    banco.executa(query)
    result = banco.fetchone()
    #log(result)
    if(result != None): 
        for campo in result:
            vUF = campo
    return(vUF)

def sumColumn(dados, coluna,aPartir):
    total = 0
    if aPartir==None:
        aPartir=0
    for row in range(len(dados)):
        if str(type(dados[row][coluna])).__contains__('float'):
            total += float(dados[row][coluna])
            #log(float(dados[row][coluna]))
        if str(type(dados[row][coluna])).__contains__('int'):
            total += int(dados[row][coluna])
            #log(int(dados[row][coluna]))
    return total


def aba_resumo(vUF,vAno,vMes,IE,banco):
    #log('Gerando query aba_resumo')

    query=""" SELECT 
                m.descricao municipio,
                m.codigo_ibge codigo_municipio,
                m.indice rateio,
                m.codigo_pva
              FROM gfcadastro.reg_1400_municipio_sp m 
              where 1=1
                and m.ano='%s'
                and m.mes='%s'
                and m.uf = '%s'
                and m.data_fim is null
               order by m.descricao
    """%(vAno,vMes,vUF)
    
    retorno=[]

    banco.executa(query)
    result = banco.fetchone()
    lin = 0

    if result == None:
        log("#### ATENÇÃO: Nenhum Resultado para aba detalhado")
        log("####     Query = ")
        log("####")
        log(query)
        log("####")
        ret=99
        return(retorno)
    else:
        while result:
            retorno.append([])
            for campo in result:
                retorno[lin].append(campo)
            lin = lin + 1
            result = banco.fetchone()
    
    return(retorno)
    
def aba_resumoCFOP2(arqprot):

    encoding = comum.encodingDoArquivo(arqprot)
    ap = open(arqprot, 'r', encoding=encoding)

    CFOP1301VC = float(0)
    CFOP1301VB = float(0)
    CFOP1301VI = float(0)

    CFOP2301VC = float(0)
    CFOP2301VB = float(0)
    CFOP2301VI = float(0)

    CFOP3301VC = float(0)
    CFOP3301VB = float(0)
    CFOP3301VI = float(0)

    for linha in ap:
        if linha.startswith('|D590|') :
            registros = linha.split('|')
            if (registros[3] == "1301"):
                CFOP1301VC = CFOP1301VC + float(registros[5].replace(',','.'))
                CFOP1301VB = CFOP1301VB + float(registros[6].replace(',','.'))
                CFOP1301VI = CFOP1301VI + float(registros[7].replace(',','.'))
            if (registros[3] == "2301"):
                CFOP2301VC = CFOP2301VC + float(registros[5].replace(',','.'))
                CFOP2301VB = CFOP2301VB + float(registros[6].replace(',','.'))
                CFOP2301VI = CFOP2301VI + float(registros[7].replace(',','.'))
            if (registros[3] == "3301"):
                CFOP3301VC = CFOP3301VC + float(registros[5].replace(',','.'))
                CFOP3301VB = CFOP3301VB + float(registros[6].replace(',','.'))
                CFOP3301VI = CFOP3301VI + float(registros[7].replace(',','.'))
    ap.close()



    retorno=[]

    retorno.append([])
    retorno[0].append("1301")
    retorno[0].append(CFOP1301VC)
    retorno[0].append(CFOP1301VB)
    retorno[0].append(CFOP1301VI)

    retorno.append([])
    retorno[1].append("2301")
    retorno[1].append(CFOP2301VC)
    retorno[1].append(CFOP2301VB)
    retorno[1].append(CFOP2301VI)
    
    retorno.append([])
    retorno[2].append("3301")
    retorno[2].append(CFOP3301VC)
    retorno[2].append(CFOP3301VB)
    retorno[2].append(CFOP3301VI)

    return(retorno)

def aba_resumoCFOP(vDataIni,vDataFim,IE,banco):

    query ="""
            select  CFOP,  VLR_CONTABIL,  VLR_BASE_ICMS, VLR_ICMS,  VLR_ISENTAS, VLR_OUTRAS
            from(
                SELECT /*+ PARALLEL (8) */
                       EXTRACT(year FROM isd.INFST_DTEMISS) AS ANO,
                       EXTRACT(month FROM isd.INFST_DTEMISS)AS MES,
                       Isd.CFOP                   AS CFOP,
                       SUM(Isd.INFST_VAL_CONT)    AS VLR_CONTABIL,
                       SUM(Isd.INFST_BASE_ICMS)   AS VLR_BASE_ICMS, 
                       SUM(Isd.INFST_VAL_ICMS)    AS VLR_ICMS,
                       SUM(Isd.INFST_ISENTA_ICMS) AS VLR_ISENTAS, 
                       SUM(Isd.INFST_OUTRAS_ICMS) AS VLR_OUTRAS
                  FROM openrisow.ITEM_NFTL_SERV Isd
                 WHERE Isd.EMPS_COD = 'TBRA'
                   AND Isd.INFST_DTEMISS BETWEEN TO_DATE('%s', 'DDMMYYYY') AND TO_DATE('%s', 'DDMMYYYY')
                   AND Isd.FILI_COD in (select distinct f.fili_cod from openrisow.filial f where f.fili_cod_insest = '%s')
                   AND Isd.INFST_IND_CANC = 'N'
                   AND Isd.CFOP <> '0000'
                GROUP BY Isd.CFOP,
                         EXTRACT(year FROM ISD.INFST_DTEMISS),
                         EXTRACT(month FROM Isd.INFST_DTEMISS)
                ORDER BY EXTRACT(month FROM Isd.INFST_DTEMISS), Isd.CFOP)
    """%(vDataIni,vDataFim,IE)

    retorno=[]
      
    lin=0
   
    banco.executa(query)
    result = banco.fetchone()

    if result == None:
        log("#### ATENÇÃO: Nenhum Resultado para aba detalhado")
        log("####     Query = ")
        log("####")
        log(query)
        log("####")
        ret=99
        return(retorno)
    else:
        while result:
            retorno.append([])
            for campo in result:
                retorno[lin].append(campo)
            lin = lin + 1
            result = banco.fetchone()

    return(retorno)

#######inicio##########    

    
def main():    
    log("-"*100)
    log("#### INICIO DO RELATORIO REGISTRO 1400 VERSÃO 20210831 ####")
    ret = processar()
    if (ret > 0) :
        log("ERRO - Verifique as mensagens anteriores...")
        log("#### Código de execução = ", ret)
    log("#### FIM DO RELATORIO REGISTRO 1400 VERSÃO 20210831 ####")
    sys.exit(ret)


if __name__ == "__main__":
    main()


