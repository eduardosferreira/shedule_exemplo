#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-

"""
----------------------------------------------------------------------------------------------
  SISTEMA ..: 
  MODULO ...: 
  SCRIPT ...: Relatório REGISTRO 1400 RJ
  CRIACAO ..: 05/07/2021
  AUTOR ....: Victor Santos - Kyros Consultoria 
  DESCRICAO : 
  ALTERACAO :
----------------------------------------------------------------------------------------------
    Exemplo de comando: ./registro_1400_RJ.py 042016 <INSCRICAO_ESTADUAL>
    Diretório: /arquivos/registro_1400/RELATORIOS/
    Exemplo: /arquivos/registro_1400/RELATORIOS/registro_032021_inscricao_estadual.xlsx
----------------------------------------------------------------------------------------------
    2021/08/31 - Airton Borges - Kyros
                Adaptação para novo Painel de execuções.
    10/03/2022  - Eduardo da Silva Ferreira - Kyros Tecnologia
                - [PTITES-1688] Padrão de diretórios do SPARTA                 
----------------------------------------------------------------------------------------------
                
"""
#### PATRONIZACAO PARA O PAINEL DE EXECUCOES....
import sys
import os
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)
import configuracoes
import comum
import util
import sql
comum.log.gerar_log_em_arquivo = False
comum.carregaConfiguracoes(configuracoes)
banco=sql.geraCnxBD(configuracoes)

#### PATRONIZACAO PARA O PAINEL DE EXECUCOES....
from typing import Pattern
import datetime
import cx_Oracle
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_MEDIUM, BORDER_DOUBLE

fontMasterPreta  = Font(color='00000000', bold=True, size=12)
fontNegrito      = Font(color='00000000', bold=True)

nome_relatorio = "" 
dir_destino = "" 
dir_base = "" 
# INICIO PTITES-1688 
dir_arquivos        = os.path.join(os.path.dirname(configuracoes.dir_geracao_arquivos), 'REGISTRO_1400', 'RELATORIOS') # PTITES-1688 # configuracoes.diretorio_arquivos
log("-"* 14)
log('# - DIRETORIO ARQUIVOS REG. 1400  .:', dir_arquivos)
log("-"* 14)   
# FIM PTITES-1688 

toti = {}
totv = {}
ret = 0 

def ultimodia(ano,mes):
   return(31 if mes == 12 else datetime.date.fromordinal((datetime.date(ano,mes+1,1)).toordinal()-1).day)

def processar():
    vDataIni =""
    vDataFim =""
    IE   = ""
    flag = "" 
    ret = 0
    
    
    log(len(sys.argv))

    for x in range(1,len(sys.argv)):
        log("parametro ", x , "= ", sys.argv[x])
    

    if (    len(sys.argv) == 4 
        and len(sys.argv[1])==6  
        and int(sys.argv[1][0:2])>0 
        and int(sys.argv[1][0:2])<13
        and int(sys.argv[1][2:6])<=datetime.datetime.now().year
        and int(sys.argv[1][2:6])>(datetime.datetime.now().year)-50 ):
        vPeriodo = sys.argv[1]
        IE       = sys.argv[2] 
        flag     = sys.argv[3] 
    elif (  len(sys.argv) == 5 
        and len(sys.argv[2])==6  
        and int(sys.argv[2][0:2])>0 
        and int(sys.argv[2][0:2])<13
        and int(sys.argv[2][2:6])<=datetime.datetime.now().year
        and int(sys.argv[2][2:6])>(datetime.datetime.now().year)-50 ):
        vPeriodo = sys.argv[2]
        IE       = sys.argv[3] 
        flag     = sys.argv[4] 
    else:
        log("-" * 100)
        log("#### ")
        log('#### ERRO - Erro nos parametros do script.')
        log("#### ")
        log('#### Exemplo de como deve ser :')
        log('####      %s  <MMYYYY> <IE> <S/N>'%(sys.argv[0]))
        log("#### ")
        log('#### Onde')
        log('####      <MMYYYY> = 032021')
        log('####       <IE>     = INSCRICAO ESTADUAL  ')
        log('####       <S/N>    = INSERE DADOS NA TABELA INVA  ')
        log('#### segue um exemplo %s 032021 77452443'%(sys.argv[0]))
        log("#### ")
        log('#### ')
        log("-" * 100)
        log("")
        log("Retorno = 99") 
        ret = 99
        return(99)  
    
    vDataIni ='01' + str(vPeriodo)
    vAno = str(vPeriodo[2:6])
    vMes = str(vPeriodo[0:2])
    vDataIni ='01/' + vMes +'/'+ vAno
    UltDiaMes =ultimodia(int(vPeriodo[2:6]), int(vPeriodo[0:2]))
    vDataFim =str(UltDiaMes)+ '/' + vMes + '/' + vAno 
    
    log("-"* 100)
    log('# - Periodo............:', vPeriodo)
    log('# - Inscricao estadual.:', IE)
    log('# - DataIni............:',vDataIni)
    log('# - DataFim............:',vDataFim)
    log('# - Flag Insert INVA...:',flag)
    log("-"* 100)    
    
    

    vUF=retornaUF(IE,banco)
    if (vUF == ""):
        log("ERRO - Não foi possível determinar a UF pela IE informada.")
        return(99)
    if (vUF != "RJ"):
        log("ERRO - A IE informada não é de RJ.")
        return(99)

#### Monta caminho e nome do destino
    dir_base =  os.path.join(dir_arquivos, vUF)   
    dir_destino = os.path.join(dir_base, vAno, vMes)  

    if not os.path.isdir(dir_destino) :
        os.makedirs(dir_destino) 
    
    arquivo_destino = IE+'_Valores_Agregados_1400_'+vPeriodo+'.xlsx'
    nome_relatorio = os.path.join(dir_destino,arquivo_destino)

    log("-"* 100)
    log('#### - Planilha Relatório Destino = ',nome_relatorio)
    log("-"* 100)

    arquivo = open(nome_relatorio, 'w')
    arquivo.close() 

    #### Cria a planilha em memória....
    arquivo_excel = Workbook()
    planilha0 = arquivo_excel.active
    planilha0.title = "ABA 1.RESUMO POR CFOP"

###################################################################################
###Aba 1. Resumo por CFOP
###Aba 1. Resumo por CFOP
###Aba 1. Resumo por CFOP
###################################################################################
    log("")
    log("# - Início do processamento da ABA 1: 'Resumo por CFOP'.")

#### CABEÇALHO 
#### CABEÇALHO 
#### CABEÇALHO 

    vLinha = 1
    planilha0.cell(vLinha,1,"Quadro 1.Resumo CFOP - Entrada Mercadoria")
    planilha0.cell(vLinha,1).font=Font(bold=True)
    planilha0.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha0.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.merge_cells('A'+ str(vLinha) + ':F' + str(vLinha))
    
    vLinha = vLinha + 1
    planilha0.cell(vLinha,1,"Insc Estadual...: "+IE)
    planilha0.cell(vLinha,1).font=Font(bold=True)
    planilha0.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.merge_cells('A'+ str(vLinha) + ':F' + str(vLinha))

    vLinha = vLinha + 1
    planilha0.cell(vLinha,1,"Mês/Ano........: " + vPeriodo)
    planilha0.cell(vLinha,1).font=Font(bold=True)
    planilha0.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.merge_cells('A'+ str(vLinha) + ':F' + str(vLinha))
    
    vLinha = vLinha + 1
    vColuna = 1
    for nColuna in ("CFOP","Valor Contábil","Base ICMS","Valor ICMS","Valor Isentas","Valor Outras"):
        planilha0.cell(vLinha,vColuna,nColuna)
        planilha0.cell(vLinha,vColuna).font=Font(bold=True)
        planilha0.cell(vLinha,vColuna).alignment = Alignment(horizontal='center')
        vColuna = vColuna + 1
 
#### DADOS
#### DADOS
#### DADOS
    
    dados = []
    dados = aba_resumo_cfop(vDataIni,vDataFim,IE,banco)  
    plinhaP0 = vLinha + 1 ## primeira linha com dados a serem somados.
    
    for linha in dados:
        vLinha = vLinha + 1
        planilha0.cell(vLinha,1,linha[0])
        planilha0.cell(vLinha,2,linha[1])
        planilha0.cell(vLinha,3,linha[2])
        planilha0.cell(vLinha,4,linha[3])
        planilha0.cell(vLinha,5,linha[4])
        planilha0.cell(vLinha,6,linha[5])
        
        planilha0.cell(vLinha,2).number_format = "#,##0.00"
        planilha0.cell(vLinha,3).number_format = "#,##0.00"
        planilha0.cell(vLinha,4).number_format = "#,##0.00"
        planilha0.cell(vLinha,5).number_format = "#,##0.00"
        planilha0.cell(vLinha,6).number_format = "#,##0.00"
        
    ulinhaP0 = vLinha

#### TOTAIS
#### TOTAIS
#### TOTAIS
    
    vLinha = vLinha + 1
    planilha0.cell(vLinha,1,"TOTAL:")
    planilha0.cell(vLinha,1).font=Font(bold=True)

    planilha0.cell(vLinha,2,"=SUM(B"+str(plinhaP0)+":B"+str(ulinhaP0)+")")
    planilha0.cell(vLinha,3,"=SUM(C"+str(plinhaP0)+":C"+str(ulinhaP0)+")")
    planilha0.cell(vLinha,4,"=SUM(D"+str(plinhaP0)+":D"+str(ulinhaP0)+")")
    planilha0.cell(vLinha,5,"=SUM(E"+str(plinhaP0)+":E"+str(ulinhaP0)+")")
    planilha0.cell(vLinha,6,"=SUM(F"+str(plinhaP0)+":F"+str(ulinhaP0)+")")
    
    planilha0.cell(vLinha,2).font=Font(bold=True)
    planilha0.cell(vLinha,3).font=Font(bold=True)
    planilha0.cell(vLinha,4).font=Font(bold=True)
    planilha0.cell(vLinha,5).font=Font(bold=True)
    planilha0.cell(vLinha,6).font=Font(bold=True)

    planilha0.cell(vLinha,2).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.cell(vLinha,3).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.cell(vLinha,4).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.cell(vLinha,5).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.cell(vLinha,6).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

    planilha0.cell(vLinha,2).number_format = "#,##0.00"
    planilha0.cell(vLinha,3).number_format = "#,##0.00"
    planilha0.cell(vLinha,4).number_format = "#,##0.00"
    planilha0.cell(vLinha,5).number_format = "#,##0.00"
    planilha0.cell(vLinha,6).number_format = "#,##0.00"

#### CABEÇALHO  2
#### CABEÇALHO  2
#### CABEÇALHO  2
   
    vLinha = vLinha + 3
    planilha0.cell(vLinha,1,"Quadro 2.Resumo CFOP - Telecom Saidas")
    planilha0.cell(vLinha,1).font=Font(bold=True)
    planilha0.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha0.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.merge_cells('A'+ str(vLinha) + ':F' + str(vLinha))
    
    vLinha = vLinha + 1
    planilha0.cell(vLinha,1,"Insc Estadual...: "+IE)
    planilha0.cell(vLinha,1).font=Font(bold=True)
    planilha0.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.merge_cells('A'+ str(vLinha) + ':F' + str(vLinha))

    vLinha = vLinha + 1
    planilha0.cell(vLinha,1,"Mês/Ano........: " + vPeriodo)
    planilha0.cell(vLinha,1).font=Font(bold=True)
    planilha0.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha0.merge_cells('A'+ str(vLinha) + ':F' + str(vLinha))
    
    vLinha = vLinha + 1
    vColuna = 1
    for nColuna in ("CFOP","Valor Contábil","Base ICMS","Valor ICMS","Valor Isentas","Valor Outras"):
        planilha0.cell(vLinha,vColuna,nColuna)
        planilha0.cell(vLinha,vColuna).font=Font(bold=True)
        planilha0.cell(vLinha,vColuna).alignment = Alignment(horizontal='center')
        vColuna = vColuna + 1
 
#### DADOS
#### DADOS
#### DADOS
    
    dados = []
    dados = aba_resumo_cfop_2(vDataIni,vDataFim,IE,banco)  
    plinhaP0 = vLinha + 1 ## primeira linha com dados a serem somados.
    
    for linha in dados:
        vLinha = vLinha + 1
        planilha0.cell(vLinha,1,linha[0])
        planilha0.cell(vLinha,2,linha[1])
        planilha0.cell(vLinha,3,linha[2])
        planilha0.cell(vLinha,4,linha[3])
        planilha0.cell(vLinha,5,linha[4])
        planilha0.cell(vLinha,6,linha[5])

        planilha0.cell(vLinha,2).number_format = "#,##0.00"
        planilha0.cell(vLinha,3).number_format = "#,##0.00"
        planilha0.cell(vLinha,4).number_format = "#,##0.00"
        planilha0.cell(vLinha,5).number_format = "#,##0.00"
        planilha0.cell(vLinha,6).number_format = "#,##0.00"
        
    ulinhaP0 = vLinha

#### TOTAIS
#### TOTAIS
#### TOTAIS
    
    vLinha = vLinha + 1
    planilha0.cell(vLinha,1,"TOTAL:")
    planilha0.cell(vLinha,1).font=Font(bold=True)
    
    if (ulinhaP0 >= plinhaP0):
        planilha0.cell(vLinha,2,"=SUM(B"+str(plinhaP0)+":B"+str(ulinhaP0)+")")
        planilha0.cell(vLinha,3,"=SUM(C"+str(plinhaP0)+":C"+str(ulinhaP0)+")")
        planilha0.cell(vLinha,4,"=SUM(D"+str(plinhaP0)+":D"+str(ulinhaP0)+")")
        planilha0.cell(vLinha,5,"=SUM(E"+str(plinhaP0)+":E"+str(ulinhaP0)+")")
        planilha0.cell(vLinha,6,"=SUM(F"+str(plinhaP0)+":F"+str(ulinhaP0)+")")
    else:
        planilha0.cell(vLinha,2,0.00)
        planilha0.cell(vLinha,3,0.00)
        planilha0.cell(vLinha,4,0.00)
        planilha0.cell(vLinha,5,0.00)
        planilha0.cell(vLinha,6,0.00)
        
    planilha0.cell(vLinha,2).font=Font(bold=True)
    planilha0.cell(vLinha,3).font=Font(bold=True)
    planilha0.cell(vLinha,4).font=Font(bold=True)
    planilha0.cell(vLinha,5).font=Font(bold=True)
    planilha0.cell(vLinha,6).font=Font(bold=True)

    planilha0.cell(vLinha,2).number_format = "#,##0.00"
    planilha0.cell(vLinha,3).number_format = "#,##0.00"
    planilha0.cell(vLinha,4).number_format = "#,##0.00"
    planilha0.cell(vLinha,5).number_format = "#,##0.00"
    planilha0.cell(vLinha,6).number_format = "#,##0.00"

#### FORMATAÇAO
#### FORMATAÇAO
#### FORMATAÇAO

    planilha0.column_dimensions['A'].width = 20  
    planilha0.column_dimensions['B'].width = 20   
    planilha0.column_dimensions['C'].width = 20  
    planilha0.column_dimensions['D'].width = 20  
    planilha0.column_dimensions['E'].width = 20  
    planilha0.column_dimensions['F'].width = 20  
    planilha0.column_dimensions['G'].width = 20

#### GRAVA A PLANILHA
#### GRAVA A PLANILHA
#### GRAVA A PLANILHA

    arquivo_excel.save(nome_relatorio)

    log("")
    log("# - Fim do processamento da ABA 1: 'Resumo por CFOP'.")

###################################################################################
####Aba 2.  Levantamento Município Saídas – Telecom
####Aba 2.  Levantamento Município Saídas – Telecom
####Aba 2.  Levantamento Município Saídas – Telecom
###################################################################################
    log("")
    log("# - Início do processamento da ABA 2: 'Levantamento Saídas'.")
    planilha1 = arquivo_excel.create_sheet("ABA 2.LEVANTAMENTO SAÍDAS", 1)

#### CABEÇALHO 
#### CABEÇALHO 
#### CABEÇALHO 

    vLinha = 1
    planilha1.cell(vLinha,1,"Quadro 1.Levantamento Município Saidas - Telecom")
    planilha1.cell(vLinha,1).font=Font(bold=True,size=14)
    planilha1.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha1.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha1.merge_cells('A'+ str(vLinha) + ':D' + str(vLinha))

    planilha1.cell(vLinha,6,"Quadro 2.Resumo por Município ")
    planilha1.cell(vLinha,6).font=Font(bold=True,size=14)
    planilha1.cell(vLinha,6).alignment = Alignment(horizontal='center')
    planilha1.cell(vLinha,6).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha1.merge_cells('F'+ str(vLinha) + ':G' + str(vLinha))
    
    vLinha = vLinha + 1
    planilha1.cell(vLinha,1,"Insc Estadual...: "+IE)
    planilha1.cell(vLinha,1).font=Font(bold=True)
    planilha1.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha1.merge_cells('A'+ str(vLinha) + ':D' + str(vLinha))

    planilha1.cell(vLinha,6,"Insc Estadual...: "+IE)
    planilha1.cell(vLinha,6).font=Font(bold=True)
    planilha1.cell(vLinha,6).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha1.merge_cells('F'+ str(vLinha) + ':G' + str(vLinha))

    vLinha = vLinha + 1
    planilha1.cell(vLinha,1,"Mês/Ano........: " + vPeriodo)
    planilha1.cell(vLinha,1).font=Font(bold=True)
    planilha1.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha1.merge_cells('A'+ str(vLinha) + ':D' + str(vLinha))

    planilha1.cell(vLinha,6,"Mês/Ano........: " + vPeriodo)
    planilha1.cell(vLinha,6).font=Font(bold=True)
    planilha1.cell(vLinha,6).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha1.merge_cells('F'+ str(vLinha) + ':G' + str(vLinha))
    
    vLinha = vLinha + 1
    vColuna = 1
    for nColuna in ("UF","Desc. Município","Município IBGE","Valor Contábil"):
        planilha1.cell(vLinha,vColuna,nColuna)
        planilha1.cell(vLinha,vColuna).font=Font(bold=True)
        planilha1.cell(vLinha,vColuna).alignment = Alignment(horizontal='center')
        vColuna = vColuna + 1

    vColuna2 = 6
    
    for nColuna in ("Descrição","Valor"):
        planilha1.cell(vLinha,vColuna2,nColuna)
        planilha1.cell(vLinha,vColuna2).font=Font(bold=True)
        planilha1.cell(vLinha,vColuna2).alignment = Alignment(horizontal='center')
        vColuna2 = vColuna2 + 1
    
#### DADOS
#### DADOS
#### DADOS
    
    dados = []
    dados = aba_levant_saidas(vDataIni,vDataFim,IE,banco)  
    plinhaP0 = vLinha + 1 ## primeira linha com dados a serem somados.
    
    soma_RJ     = 0
    soma_outros = 0
    for linha in dados:
        vLinha = vLinha + 1
        planilha1.cell(vLinha,1,linha[0]).alignment = Alignment(horizontal='center')
        planilha1.cell(vLinha,2,linha[1])
        planilha1.cell(vLinha,3,linha[2])
        planilha1.cell(vLinha,4,linha[3])
        planilha1.cell(vLinha,4).number_format = "#,##0.00"

        if linha[0] == 'RJ':
            soma_RJ += linha[3] 
        else:
            soma_outros += linha[3]
    
    vLinha2  = 5
    planilha1.cell(vLinha2,6,"Municípios RJ")
    planilha1.cell(vLinha2,7,soma_RJ)
    planilha1.cell(vLinha2,7).number_format = "#,##0.00"
    
    vLinha2 += 1
    planilha1.cell(vLinha2,6,"Demais Municípios")
    planilha1.cell(vLinha2,7,soma_outros)
    planilha1.cell(vLinha2,7).number_format = "#,##0.00"
    
    vLinha2 += 1
    planilha1.cell(vLinha2,6,"TOTAL")
    planilha1.cell(vLinha2,6).font=Font(bold=True)
    planilha1.cell(vLinha2,6).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

    planilha1.cell(vLinha2,7,(soma_RJ + soma_outros))
    planilha1.cell(vLinha2,7).number_format = "#,##0.00"
    planilha1.cell(vLinha2,7).font=Font(bold=True)
    planilha1.cell(vLinha2,7).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
        
    ulinhaP0 = vLinha 

#### TOTAIS
#### TOTAIS
#### TOTAIS
    
    vLinha = vLinha + 1
    planilha1.cell(vLinha,1,"TOTAL:")
    planilha1.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha1.cell(vLinha,1).font=Font(bold=True)
    planilha1.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha1.merge_cells('A'+ str(vLinha) + ':C' + str(vLinha))
    

    planilha1.cell(vLinha,4,"=SUM(D"+str(plinhaP0)+":D"+str(ulinhaP0)+")")
    planilha1.cell(vLinha,4).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha1.cell(vLinha,4).font=Font(bold=True)
    planilha1.cell(vLinha,4).number_format = "#,##0.00"
    
#### FORMATAÇAO
#### FORMATAÇAO
#### FORMATAÇAO

    planilha1.column_dimensions['A'].width = 20  
    planilha1.column_dimensions['B'].width = 20   
    planilha1.column_dimensions['C'].width = 20  
    planilha1.column_dimensions['D'].width = 20  
    planilha1.column_dimensions['E'].width = 5 
    planilha1.column_dimensions['F'].width = 20  
    planilha1.column_dimensions['G'].width = 20


#### GRAVA A PLANILHA
#### GRAVA A PLANILHA
#### GRAVA A PLANILHA

    arquivo_excel.save(nome_relatorio)   

    log("")
    log("# - Fim do processamento da ABA 2: 'Levantamento Saídas'.") 


###################################################################################
####Aba 3.  Levantamento Município Entradas - Mercadoria
####Aba 3.  Levantamento Município Entradas - Mercadoria
####Aba 3.  Levantamento Município Entradas - Mercadoria
###################################################################################
    log("")
    log("# - Início do processamento da ABA 3: 'Levantamento Entradas'.")
    planilha2 = arquivo_excel.create_sheet("ABA 3.LEVANTAMENTO ENTRADAS", 2)

#### CABEÇALHO 
#### CABEÇALHO 
#### CABEÇALHO 

    vLinha = 1
    planilha2.cell(vLinha,1,"Quadro 1. Levantamento Município Entradas - Mercadoria")
    planilha2.cell(vLinha,1).font=Font(bold=True,size=14)
    planilha2.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha2.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha2.merge_cells('A'+ str(vLinha) + ':D' + str(vLinha))

    planilha2.cell(vLinha,6,"Quadro 2.Resumo por Município ")
    planilha2.cell(vLinha,6).font=Font(bold=True,size=14)
    planilha2.cell(vLinha,6).alignment = Alignment(horizontal='center')
    planilha2.cell(vLinha,6).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha2.merge_cells('F'+ str(vLinha) + ':G' + str(vLinha))
    
    vLinha = vLinha + 1
    planilha2.cell(vLinha,1,"Insc Estadual...: "+IE)
    planilha2.cell(vLinha,1).font=Font(bold=True)
    planilha2.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha2.merge_cells('A'+ str(vLinha) + ':D' + str(vLinha))

    planilha2.cell(vLinha,6,"Insc Estadual...: "+IE)
    planilha2.cell(vLinha,6).font=Font(bold=True)
    planilha2.cell(vLinha,6).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha2.merge_cells('F'+ str(vLinha) + ':G' + str(vLinha))

    vLinha = vLinha + 1
    planilha2.cell(vLinha,1,"Mês/Ano........: " + vPeriodo)
    planilha2.cell(vLinha,1).font=Font(bold=True)
    planilha2.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha2.merge_cells('A'+ str(vLinha) + ':D' + str(vLinha))

    planilha2.cell(vLinha,6,"Mês/Ano........: " + vPeriodo)
    planilha2.cell(vLinha,6).font=Font(bold=True)
    planilha2.cell(vLinha,6).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha2.merge_cells('F'+ str(vLinha) + ':G' + str(vLinha))
    
    vLinha = vLinha + 1
    vColuna = 1
    for nColuna in ("UF","Desc. Município","Município IBGE","Valor Contábil"):
        planilha2.cell(vLinha,vColuna,nColuna)
        planilha2.cell(vLinha,vColuna).font=Font(bold=True)
        planilha2.cell(vLinha,vColuna).alignment = Alignment(horizontal='center')
        vColuna = vColuna + 1

    vColuna2 = 6
    
    for nColuna in ("Descrição","Valor"):
        planilha2.cell(vLinha,vColuna2,nColuna)
        planilha2.cell(vLinha,vColuna2).font=Font(bold=True)
        planilha2.cell(vLinha,vColuna2).alignment = Alignment(horizontal='center')
        vColuna2 = vColuna2 + 1
    
#### DADOS
#### DADOS
#### DADOS
    
    dados = []
    dados = aba_levant_entrada(vDataIni,vDataFim,IE,banco)  
    plinhaP0 = vLinha + 1 ## primeira linha com dados a serem somados.
    
    soma_RJ     = 0
    soma_outros = 0

    for linha in dados:
        vLinha = vLinha + 1
        planilha2.cell(vLinha,1,linha[0]).alignment = Alignment(horizontal='center')
        planilha2.cell(vLinha,2,linha[1])
        planilha2.cell(vLinha,3,linha[2])
        planilha2.cell(vLinha,4,linha[3])
        planilha2.cell(vLinha,4).number_format = "#,##0.00"

        if linha[0] == 'RJ':
            soma_RJ += linha[3] 
        else:
            soma_outros += linha[3]
        
    ulinhaP0 = vLinha 

    vLinha2  = 5
    planilha2.cell(vLinha2,6,"Municípios RJ")
    planilha2.cell(vLinha2,7,soma_RJ)
    planilha2.cell(vLinha2,7).number_format = "#,##0.00"
    
    vLinha2 += 1
    planilha2.cell(vLinha2,6,"Demais Municípios")
    planilha2.cell(vLinha2,7,soma_outros)
    planilha2.cell(vLinha2,7).number_format = "#,##0.00"
    
    vLinha2 += 1
    planilha2.cell(vLinha2,6,"TOTAL")
    planilha2.cell(vLinha2,6).font=Font(bold=True)
    planilha2.cell(vLinha2,6).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha2.cell(vLinha2,7,(soma_RJ + soma_outros))
    planilha2.cell(vLinha2,7).number_format = "#,##0.00"
    planilha2.cell(vLinha2,7).font=Font(bold=True)
    planilha2.cell(vLinha2,7).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

#### TOTAIS
#### TOTAIS
#### TOTAIS
    
    vLinha = vLinha + 1
    planilha2.cell(vLinha,1,"TOTAL:")
    planilha2.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha2.cell(vLinha,1).font=Font(bold=True)
    planilha2.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha2.merge_cells('A'+ str(vLinha) + ':C' + str(vLinha))
    

    planilha2.cell(vLinha,4,"=SUM(D"+str(plinhaP0)+":D"+str(ulinhaP0)+")")
    planilha2.cell(vLinha,4).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha2.cell(vLinha,4).font=Font(bold=True)
    planilha2.cell(vLinha,4).number_format = "#,##0.00"
    
#### FORMATAÇAO
#### FORMATAÇAO
#### FORMATAÇAO

    planilha2.column_dimensions['A'].width = 20  
    planilha2.column_dimensions['B'].width = 20   
    planilha2.column_dimensions['C'].width = 20  
    planilha2.column_dimensions['D'].width = 20  
    planilha2.column_dimensions['E'].width = 10 
    planilha2.column_dimensions['F'].width = 20  
    planilha2.column_dimensions['G'].width = 20

#### GRAVA A PLANILHA
#### GRAVA A PLANILHA
#### GRAVA A PLANILHA

    arquivo_excel.save(nome_relatorio)   

    log("")
    log("# - Fim do processamento da ABA 3: 'Levantamento Entrada'.") 

###################################################################################
####Aba 4. Calculo para Rateio por Município – Saída Telecom
####Aba 4. Calculo para Rateio por Município – Saída Telecom
####Aba 4. Calculo para Rateio por Município – Saída Telecom
###################################################################################
    log("")
    log("# - Início do processamento da ABA 4: 'Cálculo Rateio Saídas'.")
    planilha3 = arquivo_excel.create_sheet("ABA 4.CÁLCULO RATEIO SAÍDAS", 3)

#### CABEÇALHO 
#### CABEÇALHO 
#### CABEÇALHO 

    vLinha = 1
    planilha3.cell(vLinha,1,"Quadro 1. Calculo para Rateio por Município Saidas (Telecom)")
    planilha3.cell(vLinha,1).font=Font(bold=True,size=14)
    planilha3.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha3.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha3.merge_cells('A'+ str(vLinha) + ':G' + str(vLinha))

    planilha3.cell(vLinha,9,"Quadro 2.Resumo por Município ")
    planilha3.cell(vLinha,9).font=Font(bold=True,size=14)
    planilha3.cell(vLinha,9).alignment = Alignment(horizontal='center')
    planilha3.cell(vLinha,9).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha3.merge_cells('I'+ str(vLinha) + ':J' + str(vLinha))
    
    vLinha = vLinha + 1
    planilha3.cell(vLinha,1,"Insc Estadual...: "+IE)
    planilha3.cell(vLinha,1).font=Font(bold=True)
    planilha3.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha3.merge_cells('A'+ str(vLinha) + ':G' + str(vLinha))

    planilha3.cell(vLinha,9,"Insc Estadual...: "+IE)
    planilha3.cell(vLinha,9).font=Font(bold=True)
    planilha3.cell(vLinha,9).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha3.merge_cells('I'+ str(vLinha) + ':J' + str(vLinha))

    vLinha = vLinha + 1
    planilha3.cell(vLinha,1,"Mês/Ano........: " + vPeriodo)
    planilha3.cell(vLinha,1).font=Font(bold=True)
    planilha3.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha3.merge_cells('A'+ str(vLinha) + ':G' + str(vLinha))

    planilha3.cell(vLinha,9,"Mês/Ano........: " + vPeriodo)
    planilha3.cell(vLinha,9).font=Font(bold=True)
    planilha3.cell(vLinha,9).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha3.merge_cells('I'+ str(vLinha) + ':J' + str(vLinha))
    
    vLinha = vLinha + 1
    vColuna = 1
    for nColuna in ("UF","Descrição Município","Município IBGE","Valor Contábil","Percentual Calculado","Valor Diferença","Calculo Valor Agregado"):
        planilha3.cell(vLinha,vColuna,nColuna)
        planilha3.cell(vLinha,vColuna).font=Font(bold=True)
        planilha3.cell(vLinha,vColuna).alignment = Alignment(horizontal='center')
        vColuna = vColuna + 1

    vColuna2 = 9
    
    for nColuna in ("Descrição","Valor"):
        planilha3.cell(vLinha,vColuna2,nColuna)
        planilha3.cell(vLinha,vColuna2).font=Font(bold=True)
        planilha3.cell(vLinha,vColuna2).alignment = Alignment(horizontal='center')
        vColuna2 = vColuna2 + 1
    
#### DADOS
#### DADOS
#### DADOS
    
    dados = []
    dados = aba_rateio_saida(vDataIni,vDataFim,IE,banco)  
    plinhaP0 = vLinha + 1 ## primeira linha com dados a serem somados.
    
    soma_RJ     = 0
    soma_outros = 0

    for linha in dados:
        vLinha = vLinha + 1
        planilha3.cell(vLinha,1,linha[0]).alignment = Alignment(horizontal='center')
        planilha3.cell(vLinha,2,linha[1])
        planilha3.cell(vLinha,3,linha[2])
        planilha3.cell(vLinha,4,linha[3])
        planilha3.cell(vLinha,5,linha[4])
        planilha3.cell(vLinha,6,linha[5])
        planilha3.cell(vLinha,7,linha[6])
        planilha3.cell(vLinha,4).number_format = "#,##0.00"
        planilha3.cell(vLinha,5).number_format = "#,###############0.000000000000000"
        planilha3.cell(vLinha,6).number_format = "#,##0.00"
        planilha3.cell(vLinha,7).number_format = "#,##0.00"

        soma_RJ += linha[3] 
        soma_outros += linha[5]
        
    ulinhaP0 = vLinha 

    vLinha2  = 5
    planilha3.cell(vLinha2,9,"Municípios RJ")
    planilha3.cell(vLinha2,10,soma_RJ)
    planilha3.cell(vLinha2,10).number_format = "#,##0.00"
    
    vLinha2 += 1
    planilha3.cell(vLinha2,9,"Demais Municípios")
    planilha3.cell(vLinha2,10,soma_outros)
    planilha3.cell(vLinha2,10).number_format = "#,##0.00"
    
    vLinha2 += 1
    planilha3.cell(vLinha2,9,"TOTAL")
    planilha3.cell(vLinha2,9).font=Font(bold=True)
    planilha3.cell(vLinha2,9).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha3.cell(vLinha2,10,(soma_RJ + soma_outros))
    planilha3.cell(vLinha2,10).number_format = "#,##0.00"
    planilha3.cell(vLinha2,10).font=Font(bold=True)
    planilha3.cell(vLinha2,10).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

#### TOTAIS
#### TOTAIS
#### TOTAIS
    
    vLinha = vLinha + 1
    planilha3.cell(vLinha,1,"TOTAL:")
    planilha3.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha3.cell(vLinha,1).font=Font(bold=True)
    planilha3.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha3.merge_cells('A'+ str(vLinha) + ':C' + str(vLinha))
    
    planilha3.cell(vLinha,4,"=SUM(D"+str(plinhaP0)+":D"+str(ulinhaP0)+")")
    planilha3.cell(vLinha,4).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha3.cell(vLinha,4).font=Font(bold=True)
    planilha3.cell(vLinha,4).number_format = "#,##0.00"

    planilha3.cell(vLinha,5,"=SUM(E"+str(plinhaP0)+":E"+str(ulinhaP0)+")")
    planilha3.cell(vLinha,5).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha3.cell(vLinha,5).font=Font(bold=True)

    planilha3.cell(vLinha,6,"=SUM(F"+str(plinhaP0)+":F"+str(ulinhaP0)+")")
    planilha3.cell(vLinha,6).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha3.cell(vLinha,6).font=Font(bold=True)
    planilha3.cell(vLinha,6).number_format = "#,##0.00"

    planilha3.cell(vLinha,7,"=SUM(G"+str(plinhaP0)+":G"+str(ulinhaP0)+")")
    planilha3.cell(vLinha,7).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha3.cell(vLinha,7).font=Font(bold=True)
    planilha3.cell(vLinha,7).number_format = "#,##0.00"
    
#### FORMATAÇAO
#### FORMATAÇAO
#### FORMATAÇAO

    planilha3.column_dimensions['A'].width = 25  
    planilha3.column_dimensions['B'].width = 25   
    planilha3.column_dimensions['C'].width = 25  
    planilha3.column_dimensions['D'].width = 25  
    planilha3.column_dimensions['E'].width = 25 
    planilha3.column_dimensions['F'].width = 25  
    planilha3.column_dimensions['G'].width = 25
    planilha3.column_dimensions['H'].width = 5
    planilha3.column_dimensions['I'].width = 25
    planilha3.column_dimensions['J'].width = 25

#### GRAVA A PLANILHA
#### GRAVA A PLANILHA
#### GRAVA A PLANILHA

    arquivo_excel.save(nome_relatorio)   

    log("")
    log("# - Fim do processamento da ABA 4: 'Cálculo Rateio Saídas'.") 
    
    
###################################################################################
####Aba 5. Calculo para Rateio por Município – Entrada Telecom
####Aba 5. Calculo para Rateio por Município – Entrada Telecom
####Aba 5. Calculo para Rateio por Município – Entrada Telecom
###################################################################################
    log("")
    log("# - Início do processamento da ABA 5: 'Cálculo Rateio Entrada'.")
    planilha4 = arquivo_excel.create_sheet("ABA 5.CÁLCULO RATEIO ENTRADAS", 4)

#### CABEÇALHO 
#### CABEÇALHO 
#### CABEÇALHO 

    vLinha = 1
    planilha4.cell(vLinha,1,"Quadro 1. Calculo para Rateio por Municipio Entradas (Mercadoria)")
    planilha4.cell(vLinha,1).font=Font(bold=True,size=14)
    planilha4.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha4.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha4.merge_cells('A'+ str(vLinha) + ':G' + str(vLinha))

    planilha4.cell(vLinha,9,"Quadro 2.Resumo por Município ")
    planilha4.cell(vLinha,9).font=Font(bold=True,size=14)
    planilha4.cell(vLinha,9).alignment = Alignment(horizontal='center')
    planilha4.cell(vLinha,9).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha4.merge_cells('I'+ str(vLinha) + ':J' + str(vLinha))
    
    vLinha = vLinha + 1
    planilha4.cell(vLinha,1,"Insc Estadual...: "+IE)
    planilha4.cell(vLinha,1).font=Font(bold=True)
    planilha4.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha4.merge_cells('A'+ str(vLinha) + ':G' + str(vLinha))

    planilha4.cell(vLinha,9,"Insc Estadual...: "+IE)
    planilha4.cell(vLinha,9).font=Font(bold=True)
    planilha4.cell(vLinha,9).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha4.merge_cells('I'+ str(vLinha) + ':J' + str(vLinha))

    vLinha = vLinha + 1
    planilha4.cell(vLinha,1,"Mês/Ano........: " + vPeriodo)
    planilha4.cell(vLinha,1).font=Font(bold=True)
    planilha4.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha4.merge_cells('A'+ str(vLinha) + ':G' + str(vLinha))

    planilha4.cell(vLinha,9,"Mês/Ano........: " + vPeriodo)
    planilha4.cell(vLinha,9).font=Font(bold=True)
    planilha4.cell(vLinha,9).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha4.merge_cells('I'+ str(vLinha) + ':J' + str(vLinha))
    
    vLinha = vLinha + 1
    vColuna = 1
    for nColuna in ("UF","Descrição Município","Município IBGE","Valor Contábil","Percentual Calculado","Valor Diferença","Calculo Valor Agregado"):
        planilha4.cell(vLinha,vColuna,nColuna)
        planilha4.cell(vLinha,vColuna).font=Font(bold=True)
        planilha4.cell(vLinha,vColuna).alignment = Alignment(horizontal='center')
        vColuna = vColuna + 1

    vColuna2 = 9
    
    for nColuna in ("Descrição","Valor"):
        planilha4.cell(vLinha,vColuna2,nColuna)
        planilha4.cell(vLinha,vColuna2).font=Font(bold=True)
        planilha4.cell(vLinha,vColuna2).alignment = Alignment(horizontal='center')
        vColuna2 = vColuna2 + 1
    
#### DADOS
#### DADOS
#### DADOS
    
    dados = []
    dados = aba_rateio_entrada(vDataIni,vDataFim,IE,banco)  
    plinhaP0 = vLinha + 1 ## primeira linha com dados a serem somados.
    
    soma_RJ     = 0
    soma_outros = 0

    for linha in dados:
        vLinha = vLinha + 1
        planilha4.cell(vLinha,1,linha[0]).alignment = Alignment(horizontal='center')
        planilha4.cell(vLinha,2,linha[1])
        planilha4.cell(vLinha,3,linha[2])
        planilha4.cell(vLinha,4,linha[3])
        planilha4.cell(vLinha,5,linha[4])
        planilha4.cell(vLinha,6,linha[5])
        planilha4.cell(vLinha,7,linha[6])
        planilha4.cell(vLinha,4).number_format = "#,##0.00"
        planilha4.cell(vLinha,5).number_format = "#,###############0.000000000000000"
        planilha4.cell(vLinha,6).number_format = "#,##0.00"
        planilha4.cell(vLinha,7).number_format = "#,##0.00"

        if linha[0] == 'RJ':
            soma_RJ += linha[3] 
        else:
            soma_outros += linha[3]
        
    ulinhaP0 = vLinha 

    vLinha2  = 5
    planilha4.cell(vLinha2,9,"Municípios RJ")
    planilha4.cell(vLinha2,10,soma_RJ)
    planilha4.cell(vLinha2,10).number_format = "#,##0.00"
    
    vLinha2 += 1
    planilha4.cell(vLinha2,9,"Demais Municípios")
    planilha4.cell(vLinha2,10,soma_outros)
    planilha4.cell(vLinha2,10).number_format = "#,##0.00"
    
    vLinha2 += 1
    planilha4.cell(vLinha2,9,"TOTAL")
    planilha4.cell(vLinha2,9).font=Font(bold=True)
    planilha4.cell(vLinha2,9).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha4.cell(vLinha2,10,(soma_RJ + soma_outros))
    planilha4.cell(vLinha2,10).number_format = "#,##0.00"
    planilha4.cell(vLinha2,10).font=Font(bold=True)
    planilha4.cell(vLinha2,10).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

#### TOTAIS
#### TOTAIS
#### TOTAIS
    
    vLinha = vLinha + 1
    planilha4.cell(vLinha,1,"TOTAL:")
    planilha4.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha4.cell(vLinha,1).font=Font(bold=True)
    planilha4.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha4.merge_cells('A'+ str(vLinha) + ':C' + str(vLinha))
    
    planilha4.cell(vLinha,4,"=SUM(D"+str(plinhaP0)+":D"+str(ulinhaP0)+")")
    planilha4.cell(vLinha,4).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha4.cell(vLinha,4).font=Font(bold=True)
    planilha4.cell(vLinha,4).number_format = "#,##0.00"

    planilha4.cell(vLinha,5,"=SUM(E"+str(plinhaP0)+":E"+str(ulinhaP0)+")")
    planilha4.cell(vLinha,5).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha4.cell(vLinha,5).font=Font(bold=True)

    planilha4.cell(vLinha,6,"=SUM(F"+str(plinhaP0)+":F"+str(ulinhaP0)+")")
    planilha4.cell(vLinha,6).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha4.cell(vLinha,6).font=Font(bold=True)
    planilha4.cell(vLinha,6).number_format = "#,##0.00"

    planilha4.cell(vLinha,7,"=SUM(G"+str(plinhaP0)+":G"+str(ulinhaP0)+")")
    planilha4.cell(vLinha,7).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha4.cell(vLinha,7).font=Font(bold=True)
    planilha4.cell(vLinha,7).number_format = "#,##0.00"
    
#### FORMATAÇAO
#### FORMATAÇAO
#### FORMATAÇAO

    planilha4.column_dimensions['A'].width = 25  
    planilha4.column_dimensions['B'].width = 25   
    planilha4.column_dimensions['C'].width = 25  
    planilha4.column_dimensions['D'].width = 25  
    planilha4.column_dimensions['E'].width = 25 
    planilha4.column_dimensions['F'].width = 25  
    planilha4.column_dimensions['G'].width = 25
    planilha4.column_dimensions['H'].width = 5
    planilha4.column_dimensions['I'].width = 25
    planilha4.column_dimensions['J'].width = 25

#### GRAVA A PLANILHA
#### GRAVA A PLANILHA
#### GRAVA A PLANILHA

    arquivo_excel.save(nome_relatorio)   

    log("")
    log("# - Fim do processamento da ABA 5: 'Cálculo Rateio Entradas'.")     
    

###################################################################################
####Aba 6.  Composição Registro 1400 – Tabela INVA (banco GF)
####Aba 6.  Composição Registro 1400 – Tabela INVA (banco GF)
####Aba 6.  Composição Registro 1400 – Tabela INVA (banco GF)
###################################################################################
    log("")
    log("# - Início do processamento da ABA 6: 'Composição registro 1400'.")
    planilha5 = arquivo_excel.create_sheet("ABA 6.COMPOSIÇÃO REGISTRO 1400", 5)

#### CABEÇALHO 
#### CABEÇALHO 
#### CABEÇALHO 

    vLinha = 1
    planilha5.cell(vLinha,1,"Composição Registro 1400 - Tabela INVA (banco GF)")
    planilha5.cell(vLinha,1).font=Font(bold=True,size=14)
    planilha5.cell(vLinha,1).alignment = Alignment(horizontal='center')
    planilha5.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha5.merge_cells('A'+ str(vLinha) + ':H' + str(vLinha))
   
    vLinha = vLinha + 1
    planilha5.cell(vLinha,1,"Insc Estadual...: "+IE)
    planilha5.cell(vLinha,1).font=Font(bold=True)
    planilha5.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha5.merge_cells('A'+ str(vLinha) + ':H' + str(vLinha))

    vLinha = vLinha + 1
    planilha5.cell(vLinha,1,"Mês/Ano........: " + vPeriodo)
    planilha5.cell(vLinha,1).font=Font(bold=True)
    planilha5.cell(vLinha,1).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
    planilha5.merge_cells('A'+ str(vLinha) + ':H' + str(vLinha))

    vLinha = vLinha + 1
    vColuna = 1
    for nColuna in ("Empresa","Filial","Data","Tipo","Codigo Item","Municipio IBGE","Valor Agregado","Origem Calculo"):
        planilha5.cell(vLinha,vColuna,nColuna)
        planilha5.cell(vLinha,vColuna).font=Font(bold=True)
        planilha5.cell(vLinha,vColuna).alignment = Alignment(horizontal='center')
        vColuna = vColuna + 1

#### DADOS
#### DADOS
#### DADOS
    
    dados = []
    dados = aba_composicao(vDataIni,vDataFim,IE,banco)  
    plinhaP0 = vLinha + 1 ## primeira linha com dados a serem somados.
    
    soma_RJ     = 0
    soma_outros = 0

    for linha in dados:
        vLinha = vLinha + 1
        planilha5.cell(vLinha,1,linha[0]).alignment = Alignment(horizontal='center')
        planilha5.cell(vLinha,2,linha[1]).alignment = Alignment(horizontal='center')
        planilha5.cell(vLinha,3,linha[2]).alignment = Alignment(horizontal='center')
        planilha5.cell(vLinha,4,linha[3]).alignment = Alignment(horizontal='center')
        planilha5.cell(vLinha,5,linha[4]).alignment = Alignment(horizontal='center')
        planilha5.cell(vLinha,6,linha[5]).alignment = Alignment(horizontal='center')
        planilha5.cell(vLinha,7,linha[6]).alignment = Alignment(horizontal='center')
        planilha5.cell(vLinha,8,linha[7]).alignment = Alignment(horizontal='center')
        planilha5.cell(vLinha,7).number_format = "#,##0.00"

#### FORMATAÇAO
#### FORMATAÇAO
#### FORMATAÇAO

    planilha5.column_dimensions['A'].width = 20  
    planilha5.column_dimensions['B'].width = 20   
    planilha5.column_dimensions['C'].width = 20  
    planilha5.column_dimensions['D'].width = 20  
    planilha5.column_dimensions['E'].width = 20 
    planilha5.column_dimensions['F'].width = 20  
    planilha5.column_dimensions['G'].width = 20
    planilha5.column_dimensions['H'].width = 20

#### GRAVA A PLANILHA
#### GRAVA A PLANILHA
#### GRAVA A PLANILHA

    arquivo_excel.save(nome_relatorio)   

    log("")
    log("# - Fim do processamento da ABA 6: 'Composição registro 1400'.")     

#### SALVAR DADOS NA TABELA INVA
#### SALVAR DADOS NA TABELA INVA
#### SALVAR DADOS NA TABELA INVA

    if flag == 'S':
        log("")
        log("# - Início do delete / Insert tabela INVA'.") 
        log("")
        for x in range(0,len(dados)):
            fili_cod      = (dados[x][1])

            cmd_sql = """ DELETE FROM openrisow.inva
                                WHERE emps_cod   = 'TBRA'
                                  AND fili_cod   = '%s'
                                  AND inva_data >= TO_DATE('%s','dd/mm/yyyy')
                                  AND inva_data <= TO_DATE('%s','dd/mm/yyyy')
                """ % (fili_cod, vDataIni, vDataFim)  
              
#            banco.executa(cmd_sql)
            log(cmd_sql) 
#        conexao.commit() 

        for x in range(0,len(dados)):
            emps_cod      = (dados[x][0])
            fili_cod      = (dados[x][1])
            inva_data     = (dados[x][2])
            inva_tipo     = (dados[x][3])
            inva_cod      = (dados[x][4])
            mibge_cod_mun = (dados[x][5])
            inva_valor    = (dados[x][6])
            inva_ori_calc = (dados[x][7])

            cmd_sql2 = """ INSERT INTO openrisow.inva 
                                   (emps_cod,
                                    fili_cod,
                                    inva_tipo,
                                    inva_cod,
                                    inva_ori_calc,
                                    inva_data,
                                    inva_valor,
                                    mibge_cod_mun)
                                VALUES 
                                    ('%s',
                                     '%s',
                                     '%s',
                                     '%s',
                                     '%s',
                                     to_date('%s','YYYY-MM-DD HH24:MI:SS'),
                                     '%s',
                                     '%s')
                       """ % (emps_cod,fili_cod,inva_tipo,inva_cod,inva_ori_calc,inva_data,inva_valor,mibge_cod_mun)  
            
#            banco.execute(cmd_sql2)   
            log(cmd_sql2)   
#        conexao.commit() 
    
    return(0)


def retornaUF(IE,cursor):
    vUF = ""
    query="""
    select distinct f.unfe_sig  from openrisow.filial f where f.fili_cod_insest='%s'
    """%(IE)

    cursor.executa(query)
    result = cursor.fetchone()
    if(result != None): 
        for campo in result:
            vUF = campo
    return(vUF)


def aba_resumo_cfop(vDataIni,vDataFim,IE,cursor):
    query=""" SELECT
                i.cfop_cod,
                SUM(CASE WHEN((i.ind_canc = 'N'
                        OR i.ind_canc IS NULL)
                        AND(m.mnfem_den_in NOT IN('D', 'I')
                        OR m.mnfem_den_in IS NULL)
                        AND(i.ind_sit NOT IN('B', 'D', 'W', 'F')
                        OR i.ind_sit IS NULL)) THEN infem_val_cont
                        ELSE 0 END
                )  sd_vl_cont_p,
                SUM(CASE WHEN(i.infem_tribicm = 'S'
                        AND(i.ind_canc = 'N'
                        OR i.ind_canc IS NULL)
                        AND(m.mnfem_den_in NOT IN('D', 'I')
                        OR m.mnfem_den_in IS NULL)
                        AND(i.ind_sit NOT IN('B', 'D', 'W', 'F')
                        OR i.ind_sit IS NULL)) THEN i.infem_bas_icms
                        ELSE 0 END
                )  sd_vl_bc_icms_p,
                SUM(CASE WHEN(i.infem_tribicm = 'S'
                        AND(i.ind_canc = 'N'
                        OR i.ind_canc IS NULL)
                        AND(m.mnfem_den_in NOT IN('D', 'I')
                        OR m.mnfem_den_in IS NULL)
                        AND(i.ind_sit NOT IN('B', 'D', 'W', 'F')
                        OR i.ind_sit IS NULL)) THEN round(i.infem_val_icms, 2)
                        ELSE 0 END
                )  sd_vl_icms_st_p,
                SUM(CASE WHEN((i.infem_val_redicms > 0
                        OR(i.infem_tribicm = 'N'
                        OR i.infem_tribicm IS NULL))
                        AND(i.ind_canc = 'N'
                        OR i.ind_canc IS NULL)
                        AND(m.mnfem_den_in NOT IN('D', 'I')
                        OR m.mnfem_den_in IS NULL)
                        AND(i.ind_sit NOT IN('B', 'D', 'W', 'F')
                        OR i.ind_sit IS NULL)) THEN i.infem_isenta_icms
                        ELSE 0 END
                )  sd_vl_isnt_icms_p,
                SUM(CASE WHEN((i.infem_val_redicms > 0
                        OR i.infem_tribicm = 'P')
                        AND(i.ind_canc = 'N'
                        OR i.ind_canc IS NULL)
                        AND(m.mnfem_den_in NOT IN('D', 'I')
                        OR m.mnfem_den_in IS NULL)
                        AND(i.ind_sit NOT IN('B', 'D', 'W', 'F')
                        OR i.ind_sit IS NULL)) THEN round(i.infem_outra_icms, 2)
                        ELSE 0 END
                )  sd_vl_out_icms_p
            FROM openrisow.mestre_nfen_merc     m
            INNER JOIN openrisow.item_nfem_merc i 
                    ON ( i.emps_cod = m.emps_cod
                AND i.fili_cod = m.fili_cod
                AND i.infem_serie = m.mnfem_serie
                AND i.infem_num = m.mnfem_num
                AND i.infem_dtemis = m.mnfem_dtemis
                AND i.catg_cod = m.catg_cod
                AND i.cadg_cod = m.cadg_cod )
            INNER JOIN openrisow.filial         f 
                    ON f.emps_cod = i.emps_cod
                AND f.fili_cod = i.fili_cod
            WHERE 1 = 1
            AND ( m.mnfem_ind_canc IS NULL
            OR m.mnfem_ind_canc != 'S' )
            AND ( i.ind_canc IS NULL
            OR i.ind_canc != 'S' )
            AND i.cfop_cod = '1301'
            AND i.infem_dtentr >= TO_DATE('%s', 'DD/MM/YYYY')   
            AND i.infem_dtentr <= TO_DATE('%s', 'DD/MM/YYYY')    
            AND f.fili_cod_insest = '%s'                        
            GROUP BY i.cfop_cod
    """%(vDataIni,vDataFim,IE)
    
    retorno=[]
    cursor.executa(query)
    result = cursor.fetchone()
    lin = 0

    if result == None:
        log("#### ATENÇÃO: Nenhum Resultado para aba detalhado")
        log("####     Query = ")
        log("####")
        log(query)
        log("####")
        ret=99
        return(retorno)
    else:
        while result:
            retorno.append([])
            for campo in result:
                retorno[lin].append(campo)
            lin = lin + 1
            result = cursor.fetchone()
    
    return(retorno)

def aba_resumo_cfop_2(vDataIni,vDataFim,IE,cursor):
    query=""" SELECT /*+ PARALLEL(16) */ cfop
                    ,sum(INFST_VAL_SERV - INFST_VAL_DESC)  INFST_VAL_CONT
                    ,sum(INFST_BASE_ICMS) INFST_BASE_ICMS
                    ,sum(INFST_VAL_ICMS) INFST_VAL_ICMS 
                    ,sum(INFST_ISENTA_ICMS) INFST_ISENTA_ICMS
                    ,sum(INFST_OUTRAS_ICMS) INFST_OUTRAS_ICMS 
                FROM  OPENRISOW.MESTRE_NFTL_SERV NF
                inner join openrisow.ITEM_NFTL_SERV inf on ( INF.EMPS_COD = NF.EMPS_COD
                                                        AND INF.FILI_COD = NF.FILI_COD
                                                        AND INF.INFST_SERIE = NF.MNFST_SERIE
                                                        AND INF.INFST_NUM = NF.MNFST_NUM
                                                        AND INF.INFST_DTEMISS = NF.mnfst_dtemiss
                                                        AND INF.MDOC_COD = NF.MDOC_COD),
                    openrisow.filial f 
                WHERE 1 = 1 
                AND nf.emps_cod = f.emps_cod 
                AND nf.fili_cod  = f.fili_cod
                AND f.fili_cod_insest = '%s'--PARAMETRO
                AND nf.MNFST_DTEMISS >= TO_DATE('%s','DD/MM/YYYY')--PARAMETRO
                AND nf.MNFST_DTEMISS <= TO_DATE('%s','DD/MM/YYYY')--PARAMETRO
                AND NF.EMPS_COD = 'TBRA'--FIXO
                and MNFST_IND_CANC = 'N'--FIXO
                and cfop in ('5301','5302','5303','5304','5305','5306','5307')--FIXO
                group by   cfop
                ORDER BY CFOP
    """%(IE,vDataIni,vDataFim)
    
    retorno=[]
    cursor.executa(query)
    result = cursor.fetchone()
    lin = 0

    if result == None:
        log("#### ATENÇÃO: Nenhum Resultado para aba 1")
        log("####     Query = ")
        log("####")
        log(query)
        log("####")
        ret=99
        return(retorno)
    else:
        while result:
            retorno.append([])
            for campo in result:
                retorno[lin].append(campo)
            lin = lin + 1
            result = cursor.fetchone()
   
    return(retorno)

def aba_levant_saidas(vDataIni,vDataFim,IE,cursor):
    query="""WITH TMP_VAL AS (
                SELECT /*+ materialize,PARALLEL(8) */
                    (SELECT MIBGE_COD_MUN
                        FROM openrisow.cli_fornec_transp cli
                    WHERE cli.cadg_cod       = nf.cadg_cod
                        AND cli.catg_cod       = nf.catg_cod
                        AND cli.cadg_dat_atua = (select max(cli2.cadg_dat_atua) cadg_dat_atua
                                            FROM openrisow.cli_fornec_transp cli2
                                            WHERE cli2.cadg_cod       = nf.cadg_cod
                                                AND cli2.catg_cod       = nf.catg_cod
                                                AND cli2.cadg_dat_atua  <= nf.MNFST_DTEMISS)) MIBGE_COD_MUN
                                                ,INFST_VAL_SERV - INFST_VAL_DESC  INFST_VAL_CONT
                FROM  OPENRISOW.MESTRE_NFTL_SERV NF
                inner join openrisow.ITEM_NFTL_SERV inf on ( INF.EMPS_COD = NF.EMPS_COD
                AND INF.FILI_COD = NF.FILI_COD
                                                        AND INF.INFST_SERIE = NF.MNFST_SERIE
                                                        AND INF.INFST_NUM = NF.MNFST_NUM
                                                        AND INF.INFST_DTEMISS = NF.mnfst_dtemiss
                                                        AND INF.MDOC_COD = NF.MDOC_COD),
                    openrisow.filial f 
                WHERE 1 = 1 
                AND nf.emps_cod = f.emps_cod 
                AND nf.fili_cod  = f.fili_cod
                AND f.fili_cod_insest = '%s'--PARAMETRO
                AND nf.MNFST_DTEMISS >= TO_DATE('%s','DD/MM/YYYY')--PARAMETRO
                AND nf.MNFST_DTEMISS <= TO_DATE('%s','DD/MM/YYYY')--PARAMETRO
                AND NF.EMPS_COD = 'TBRA'--FIXO
                and MNFST_IND_CANC = 'N'--FIXO
                and cfop in ('5301','5302','5303','5304','5305','5306','5307')--FIXO
                )
                SELECT IB.UNFE_SIG,
                    IB.MIBGE_DESC_MUN,
                    IB.MIBGE_COD_MUN,
                    sum(AUX.INFST_VAL_CONT) contabil 
                FROM TMP_VAL AUX  INNER JOIN OPENRISOW.MIBGE IB ON (IB.MIBGE_COD_MUN = AUX.MIBGE_COD_MUN)
                GROUP BY IB.UNFE_SIG,
                        IB.MIBGE_DESC_MUN,
                        IB.MIBGE_COD_MUN
                ORDER BY IB.UNFE_SIG ASC
    """%(IE,vDataIni,vDataFim)
    
    retorno=[]
    cursor.executa(query)
    result = cursor.fetchone()
    lin = 0

    if result == None:
        log("#### ATENÇÃO: Nenhum Resultado para aba 2")
        log("####     Query = ")
        log("####")
        log(query)
        log("####")
        ret=99
        return(retorno)
    else:
        while result:
            retorno.append([])
            for campo in result:
                retorno[lin].append(campo)
            lin = lin + 1
            result = cursor.fetchone()
    
    return(retorno)

def aba_levant_entrada(vDataIni,vDataFim,IE,cursor):
    query="""WITH tmp_val AS (
                SELECT /*+ materialize,PARALLEL(8) */
                    coalesce(
                        CASE
                            WHEN(c.mibge_cod_mun IS NOT NULL) THEN
                                c.mibge_cod_mun
                            WHEN(c.tp_loc = '13') THEN
                                c.loca_cod
                            WHEN(c.tp_loc = '01') THEN
                                (SELECT l.loca_cod
                                   FROM openrisow.livro_local l
                                  WHERE l.loca_cor = c.loca_cod
                                    AND l.tp_loc = '13')
                            ELSE(SELECT l.loca_cod
                                   FROM openrisow.livro_local l
                                  WHERE l.loca_cor =(SELECT l2.loca_cor
                                                       FROM openrisow.livro_local l2
                                                      WHERE l2.loca_cod = c.loca_cod
                                                        AND l2.tp_loc = c.tp_loc)
                                    AND l.tp_loc = '13')
                        END, f.fili_mun_ibge)     mibge_cod_mun,
                            i.infem_val_cont      infem_val_cont
                  FROM openrisow.mestre_nfen_merc m
                 INNER JOIN openrisow.item_nfem_merc       i ON ( i.emps_cod = m.emps_cod
                                                            AND i.fili_cod = m.fili_cod
                                                            AND i.infem_serie = m.mnfem_serie
                                                            AND i.infem_num = m.mnfem_num
                                                            AND i.infem_dtemis = m.mnfem_dtemis
                                                            AND i.catg_cod = m.catg_cod
                                                            AND i.cadg_cod = m.cadg_cod )
                    INNER JOIN openrisow.filial               f ON ( f.emps_cod = i.emps_cod
                                                    AND f.fili_cod = i.fili_cod )
                    LEFT OUTER JOIN openrisow.cli_fornec_transp    c ON ( c.cadg_cod = i.cadg_cod
                                                                    AND c.catg_cod = i.catg_cod
                                                                    AND c.cadg_dat_atua <= m.mnfem_dtentr )
                WHERE
                    ( c.cadg_dat_atua IS NULL
                    OR c.cadg_dat_atua = (
                        SELECT
                            MAX(c2.cadg_dat_atua)
                        FROM
                            openrisow.cli_fornec_transp c2
                        WHERE
                                c2.cadg_cod = m.cadg_cod
                            AND c2.catg_cod = m.catg_cod
                            AND c2.cadg_dat_atua <= i.infem_dtentr
                    ) )
                    AND i.cfop_cod = '1301'
                    AND ( m.mnfem_ind_canc IS NULL
                        OR m.mnfem_ind_canc != 'S' )
                    AND ( i.ind_canc IS NULL
                        OR i.ind_canc != 'S' )
                    AND i.infem_dtentr >= TO_DATE('%s', 'DD/MM/YYYY')--PARAMETRO
                    AND i.infem_dtentr <= TO_DATE('%s', 'DD/MM/YYYY')--PARAMETRO
                    AND f.fili_cod_insest = '%s'--PARAMETRO
            )
            SELECT
                ib.unfe_sig,
                ib.mibge_desc_mun,
                ib.mibge_cod_mun,
                SUM(aux.infem_val_cont) contabil
            FROM
                    tmp_val aux
                INNER JOIN openrisow.mibge ib ON ( ib.mibge_cod_mun = aux.mibge_cod_mun )
            GROUP BY ib.unfe_sig,ib.mibge_desc_mun,ib.mibge_cod_mun
            ORDER BY 1, 2
    """%(vDataIni,vDataFim,IE)
    
    retorno=[]
    cursor.executa(query)
    result = cursor.fetchone()
    lin = 0

    if result == None:
        log("#### ATENÇÃO: Nenhum Resultado para aba 3")
        log("####     Query = ")
        log("####")
        log(query)
        log("####")
        ret=99
        return(retorno)
    else:
        while result:
            retorno.append([])
            for campo in result:
                retorno[lin].append(campo)
            lin = lin + 1
            result = cursor.fetchone()
    
    return(retorno)


def aba_rateio_saida(vDataIni,vDataFim,IE,cursor):
    query="""with tmp_val as (
                select  /*+ materialize,PARALLEL(8) */
                        MIBGE_COD_MUN,
                        ROUND(valor_corrigido,2)val_agregado,
                        'M'INVA_ORI_CALC,
                        total_contabil,
                        indice,
                        ROUND(valor_diff,2)valor_diff
                        ,contabil
                from (
                select emps_cod,
                    fili_cod,
                    inva_data,
                    MIBGE_COD_MUN,
                    COD_UF,
                    UNFE_SIG,
                    contabil,
                    sum(contabil) over (partition by 'x') total_contabil,
                    sum(decode(substr(MIBGE_COD_MUN,1,2),COD_UF,0.0,contabil)) over (partition by 'x') erro,
                    sum(decode(substr(MIBGE_COD_MUN,1,2),COD_UF,contabil,0.0)) over (partition by 'x') total_rio,
                    decode(substr(MIBGE_COD_MUN,1,2),COD_UF,contabil,0.0) / sum(decode(substr(MIBGE_COD_MUN,1,2),COD_UF,contabil,0.0)) over (partition by 'x') indice,
                    decode(substr(MIBGE_COD_MUN,1,2),COD_UF,contabil,0.0) + sum(decode(substr(MIBGE_COD_MUN,1,2),COD_UF,0.0,contabil)) over (partition by 'x') * (decode(substr(MIBGE_COD_MUN,1,2),COD_UF,contabil,0.0) / sum(decode(substr(MIBGE_COD_MUN,1,2),COD_UF,contabil,0.0)) over (partition by 'x')) valor_corrigido,
                    sum(decode(substr(MIBGE_COD_MUN,1,2),COD_UF,0.0,contabil)) over (partition by 'x') * (decode(substr(MIBGE_COD_MUN,1,2),COD_UF,contabil,0.0) / sum(decode(substr(MIBGE_COD_MUN,1,2),COD_UF,contabil,0.0)) over (partition by 'x')) valor_diff
                from (
                SELECT emps_cod,
                    fili_cod,
                    inva_data,
                    MIBGE_COD_MUN,
                    COD_UF,
                    UNFE_SIG,
                    sum(INFST_VAL_CONT) contabil
                FROM (
                SELECT /*+ PARALLEL(8) */
                    NF.emps_cod,
                    NF.fili_cod,
                    last_day(NF.mnfst_dtemiss)inva_data,
                    F.UNFE_SIG,
                    (SELECT MIBGE_COD_MUN
                        FROM openrisow.cli_fornec_transp cli
                    WHERE cli.cadg_cod       = nf.cadg_cod
                        AND cli.catg_cod       = nf.catg_cod
                        AND cli.cadg_dat_atua = (select max(cli2.cadg_dat_atua) cadg_dat_atua
                                            FROM openrisow.cli_fornec_transp cli2
                                            WHERE cli2.cadg_cod       = nf.cadg_cod
                                                AND cli2.catg_cod       = nf.catg_cod
                                                AND cli2.cadg_dat_atua  <= nf.MNFST_DTEMISS)) MIBGE_COD_MUN
                                                ,inf.cadg_cod
                                                ,inf.catg_cod
                                                ,INFST_VAL_SERV - INFST_VAL_DESC  INFST_VAL_CONT
                                                ,nf.rowid row_id_mestre
                                                ,SUBSTR(f.fili_mun_ibge, 1, 2) AS COD_UF
                FROM  OPENRISOW.MESTRE_NFTL_SERV NF
                inner join openrisow.ITEM_NFTL_SERV inf on ( INF.EMPS_COD = NF.EMPS_COD
                                                        AND INF.FILI_COD = NF.FILI_COD
                                                        AND INF.INFST_SERIE = NF.MNFST_SERIE
                                                        AND INF.INFST_NUM = NF.MNFST_NUM
                                                        AND INF.INFST_DTEMISS = NF.mnfst_dtemiss
                                                        AND INF.MDOC_COD = NF.MDOC_COD),
                    openrisow.filial f 
                WHERE 1 = 1 
                AND nf.emps_cod = f.emps_cod 
                AND nf.fili_cod  = f.fili_cod
                AND f.fili_cod_insest = '%s'--parametro
                AND nf.MNFST_DTEMISS >= TO_DATE('%s','DD/MM/YYYY')--parametro
                AND nf.MNFST_DTEMISS <= TO_DATE('%s','DD/MM/YYYY')--parametro
                AND NF.EMPS_COD = 'TBRA'
                and MNFST_IND_CANC = 'N'
                and cfop in ('5301','5302','5303','5304','5305','5306','5307')
                )
                GROUP BY emps_cod,
                        fili_cod,
                        inva_data,
                        MIBGE_COD_MUN,
                        COD_UF,
                        UNFE_SIG
                )
                ) a
                where SUBSTR(MIBGE_COD_MUN, 1, 2) = COD_UF
                ) 
                SELECT IB.UNFE_SIG,
                    IB.MIBGE_DESC_MUN,
                    IB.MIBGE_COD_MUN,
                    aux.contabil,
                    aux.indice,
                    valor_diff,
                    val_agregado
                FROM TMP_VAL AUX  INNER JOIN OPENRISOW.MIBGE IB ON (IB.MIBGE_COD_MUN = AUX.MIBGE_COD_MUN)
                order by 1,2
    """%(IE,vDataIni,vDataFim)
    
    retorno=[]
    cursor.executa(query)
    result = cursor.fetchone()
    lin = 0

    if result == None:
        log("#### ATENÇÃO: Nenhum Resultado para aba 4")
        log("####     Query = ")
        log("####")
        log(query)
        log("####")
        ret=0
        return(retorno)
    else:
        while result:
            retorno.append([])
            for campo in result:
                retorno[lin].append(campo)
            lin = lin + 1
            result = cursor.fetchone()
    
    return(retorno)



def aba_rateio_entrada(vDataIni,vDataFim,IE,cursor):
    query="""with tmp_val as (
                select  /*+ parallel(8) */
                        SF_EMPRESA,
                        SF_EMISSAO,
                        sf_cod_item,
                        sf_mun,
                        SF_VALOR,
                        sum(SF_VALOR) over (partition by 'x') total_SF_VALOR,
                        sum(decode(substr(SF_MUN,1,2),COD_UF,0.0,SF_VALOR)) over (partition by 'x') erro,
                        sum(decode(substr(SF_MUN,1,2),COD_UF,SF_VALOR,0.0)) over (partition by 'x') total_rio,
                        decode(substr(SF_MUN,1,2),COD_UF,SF_VALOR,0.0) / sum(decode(substr(SF_MUN,1,2),COD_UF,SF_VALOR,0.0)) over (partition by 'x') indice,
                        sum(decode(substr(SF_MUN,1,2),COD_UF,0.0,SF_VALOR)) over (partition by 'x') * (decode(substr(SF_MUN,1,2),COD_UF,SF_VALOR,0.0) / sum(decode(substr(SF_MUN,1,2),COD_UF,SF_VALOR,0.0)) over (partition by 'x')) valor_diff,
                        decode(substr(SF_MUN,1,2),COD_UF,SF_VALOR,0.0) + sum(decode(substr(SF_MUN,1,2),COD_UF,0.0,SF_VALOR)) over (partition by 'x') * (decode(substr(SF_MUN,1,2),COD_UF,SF_VALOR,0.0) / sum(decode(substr(SF_MUN,1,2),COD_UF,SF_VALOR,0.0)) over (partition by 'x')) valor_corrigido
                        ,COD_UF
                from (
                SELECT T.SF_EMPRESA,
                        last_day(T.SF_EMISSAO) SF_EMISSAO, --ANTES:T.SF_EMISSAO
                        T.SF_IE,
                        T.SF_COD_ITEM,
                        T.SF_MUN,
                        SUM(T.SF_VALOR) SF_VALOR,
                        'TC'            sf_origem
                        ,COD_UF --ACRESCENTADO
                FROM   (SELECT /*+ PARALLEL(8) */
                                M.EMPS_COD                     sf_empresa,
                                M.FILI_COD                     sf_filial,
                                M.MNFEM_DTENTR                 sf_emissao,
                                F.FILI_COD_INSEST              sf_ie,
                                CASE
                                WHEN F.UNFE_SIG = 'AC' THEN 'ACIPME07'
                                WHEN F.UNFE_SIG = 'BA' THEN 'BAE02'
                                WHEN F.UNFE_SIG = 'MA' THEN 'MAVAF002'
                                WHEN F.UNFE_SIG = 'MG' THEN 'Outras_Entradas_a_Detalhar_por_Municipio'
                                WHEN F.UNFE_SIG = 'RJ' THEN 'RJVAF00005'
                                WHEN F.UNFE_SIG = 'RN' THEN 'IPM 4.6'
                                WHEN F.UNFE_SIG = 'SP' THEN 'SPDIPAM24'
                                WHEN F.UNFE_SIG = 'TO' THEN 'TOIPME08'
                                END                            sf_cod_item,
                                COALESCE(CASE
                                            WHEN ( C.MIBGE_COD_MUN IS NOT NULL ) THEN C.MIBGE_COD_MUN
                                            WHEN ( C.TP_LOC = '13' ) THEN C.LOCA_COD
                                            WHEN ( C.TP_LOC = '01' ) THEN (SELECT L.LOCA_COD
                                                                        FROM   openrisow.LIVRO_LOCAL L
                                                                        WHERE  L.LOCA_COR = C.LOCA_COD
                                                                            AND L.TP_LOC = '13')
                                            ELSE (SELECT L.LOCA_COD
                                                FROM   openrisow.LIVRO_LOCAL L
                                                WHERE  L.LOCA_COR = (SELECT L2.LOCA_COR
                                                                    FROM   openrisow.LIVRO_LOCAL L2
                                                                    WHERE  L2.LOCA_COD = C.LOCA_COD
                                                                        AND L2.TP_LOC = C.TP_LOC)
                                                    AND L.TP_LOC = '13')
                                        END, F.FILI_MUN_IBGE) sf_mun,
                                I.INFEM_VAL_CONT               sf_valor,
                                SUBSTR(f.fili_mun_ibge, 1, 2) AS COD_UF
                        FROM   openrisow.MESTRE_NFEN_MERC M
                                INNER JOIN openrisow.ITEM_NFEM_MERC I
                                ON (I.EMPS_COD = M.EMPS_COD
                                AND I.FILI_COD = M.FILI_COD
                                AND I.INFEM_SERIE = M.MNFEM_SERIE
                                AND I.INFEM_NUM = M.MNFEM_NUM
                                AND I.INFEM_DTEMIS = M.MNFEM_DTEMIS--ORIGINAL:I.INFEM_DTENTR = M.MNFEM_DTENTR
                                AND I.CATG_COD = M.CATG_COD
                                AND I.CADG_COD = M.CADG_COD)
                                INNER JOIN openrisow.FILIAL F
                                ON (F.EMPS_COD = I.EMPS_COD
                                AND F.FILI_COD = I.FILI_COD)
                                LEFT OUTER JOIN openrisow.CLI_FORNEC_TRANSP C
                                ON (C.CADG_COD = I.CADG_COD
                                AND C.CATG_COD = I.CATG_COD
                                AND C.CADG_DAT_ATUA <= M.MNFEM_DTENTR )
                        WHERE  (C.CADG_DAT_ATUA IS NULL
                                OR C.CADG_DAT_ATUA = (SELECT MAX(C2.CADG_DAT_ATUA)
                                                    FROM   openrisow.CLI_FORNEC_TRANSP C2
                                                    WHERE  C2.CADG_COD = M.CADG_COD
                                                    AND C2.CATG_COD = M.CATG_COD
                                                    AND C2.CADG_DAT_ATUA <= I.INFEM_DTENTR))
                            AND ( M.MNFEM_IND_CANC IS NULL
                                    OR M.MNFEM_IND_CANC != 'S' )
                            AND ( I.IND_CANC IS NULL
                                    OR I.IND_CANC != 'S' )
                            and i.cfop_cod = '1301'  
                            AND f.fili_cod_insest = '%s'--ACRESCENTADO     
                            and I.INFEM_DTENTR >= TO_DATE('%s','DD/MM/YYYY')--ACRESCENTADO
                            AND I.INFEM_DTENTR <= TO_DATE('%s','DD/MM/YYYY')--ACRESCENTADO                            
                            ) T
                GROUP  BY T.SF_EMPRESA,
                            last_day(T.SF_EMISSAO),
                            T.SF_IE,
                            T.SF_COD_ITEM,
                            T.SF_MUN
                            ,t.COD_UF
                )
                ) 
                SELECT IB.UNFE_SIG,
                    IB.MIBGE_DESC_MUN,
                    IB.MIBGE_COD_MUN,
                    aux.SF_VALOR,
                    aux.indice,
                    valor_diff,
                    valor_corrigido val_agregado
                FROM TMP_VAL AUX  INNER JOIN OPENRISOW.MIBGE IB ON (IB.MIBGE_COD_MUN = AUX.sf_mun)
                where SUBSTR(sf_mun, 1, 2) = COD_UF
                order by 1,2
    """%(IE,vDataIni,vDataFim)
    
    retorno=[]
    cursor.executa(query)
    result = cursor.fetchone()
    lin = 0

    if result == None:
        log("#### ATENÇÃO: Nenhum Resultado para aba 5")
        log("####     Query = ")
        log("####")
        log(query)
        log("####")
        ret=0
        return(retorno)
    else:
        while result:
            retorno.append([])
            for campo in result:
                retorno[lin].append(campo)
            lin = lin + 1
            result = cursor.fetchone()
    
    return(retorno)


def aba_composicao(vDataIni,vDataFim,IE,cursor):
    query="""SELECT
                    SF_EMPRESA EMPS_COD,
                    SF_FILIAL FILI_COD,
                    SF_EMISSAO INVA_DATA,
                    'U' INVA_TIPO,
                    sf_cod_item INVA_COD,
                    sf_mun MIBGE_COD_MUN,
                    ROUND(valor_corrigido,2) INVA_VALOR,
                    'M' INVA_ORI_CALC
            from (
            select  SF_EMPRESA,
                    SF_FILIAL,
                    SF_EMISSAO,
                    sf_cod_item,
                    sf_mun,
                    SF_VALOR,
                    sum(SF_VALOR) over (partition by 'x') total_SF_VALOR,
                    sum(decode(substr(SF_MUN,1,2),COD_UF,0.0,SF_VALOR)) over (partition by 'x') erro,
                    sum(decode(substr(SF_MUN,1,2),COD_UF,SF_VALOR,0.0)) over (partition by 'x') total_rio,
                    decode(substr(SF_MUN,1,2),COD_UF,SF_VALOR,0.0) / sum(decode(substr(SF_MUN,1,2),COD_UF,SF_VALOR,0.0)) over (partition by 'x') indice,
                    decode(substr(SF_MUN,1,2),COD_UF,SF_VALOR,0.0) + sum(decode(substr(SF_MUN,1,2),COD_UF,0.0,SF_VALOR)) over (partition by 'x') * (decode(substr(SF_MUN,1,2),COD_UF,SF_VALOR,0.0) / sum(decode(substr(SF_MUN,1,2),COD_UF,SF_VALOR,0.0)) over (partition by 'x')) valor_corrigido
                    ,COD_UF
            from (
            SELECT T.SF_EMPRESA,
                    T.SF_FILIAL,
                    last_day(T.SF_EMISSAO) SF_EMISSAO, --ANTES:T.SF_EMISSAO
                    T.SF_IE,
                    T.SF_COD_ITEM,
                    T.SF_MUN,
                    SUM(T.SF_VALOR) SF_VALOR,
                    'TC'            sf_origem
                    ,COD_UF --ACRESCENTADO
            FROM   (SELECT /*+ PARALLEL(8) */
                            M.EMPS_COD                     sf_empresa,
                            '1710'                         sf_filial,
                            M.MNFEM_DTENTR                 sf_emissao,
                            F.FILI_COD_INSEST              sf_ie,
                            CASE
                            WHEN F.UNFE_SIG = 'AC' THEN 'ACIPME07'
                            WHEN F.UNFE_SIG = 'BA' THEN 'BAE02'
                            WHEN F.UNFE_SIG = 'MA' THEN 'MAVAF002'
                            WHEN F.UNFE_SIG = 'MG' THEN 'Outras_Entradas_a_Detalhar_por_Municipio'
                            WHEN F.UNFE_SIG = 'RJ' THEN 'RJVAF00005'
                            WHEN F.UNFE_SIG = 'RN' THEN 'IPM 4.6'
                            WHEN F.UNFE_SIG = 'SP' THEN 'SPDIPAM24'
                            WHEN F.UNFE_SIG = 'TO' THEN 'TOIPME08'
                            END                            sf_cod_item,
                            COALESCE(CASE
                                        WHEN ( C.MIBGE_COD_MUN IS NOT NULL ) THEN C.MIBGE_COD_MUN
                                        WHEN ( C.TP_LOC = '13' ) THEN C.LOCA_COD
                                        WHEN ( C.TP_LOC = '01' ) THEN (SELECT L.LOCA_COD
                                                                    FROM   openrisow.LIVRO_LOCAL L
                                                                    WHERE  L.LOCA_COR = C.LOCA_COD
                                                                        AND L.TP_LOC = '13')
                                        ELSE (SELECT L.LOCA_COD
                                            FROM   openrisow.LIVRO_LOCAL L
                                            WHERE  L.LOCA_COR = (SELECT L2.LOCA_COR
                                                                FROM   openrisow.LIVRO_LOCAL L2
                                                                WHERE  L2.LOCA_COD = C.LOCA_COD
                                                                    AND L2.TP_LOC = C.TP_LOC)
                                                AND L.TP_LOC = '13')
                                    END, F.FILI_MUN_IBGE) sf_mun,
                            I.INFEM_VAL_CONT               sf_valor,
                            SUBSTR(f.fili_mun_ibge, 1, 2) AS COD_UF
                    FROM   openrisow.MESTRE_NFEN_MERC M
                            INNER JOIN openrisow.ITEM_NFEM_MERC I
                            ON (I.EMPS_COD = M.EMPS_COD
                            AND I.FILI_COD = M.FILI_COD
                            AND I.INFEM_SERIE = M.MNFEM_SERIE
                            AND I.INFEM_NUM = M.MNFEM_NUM
                            AND I.INFEM_DTEMIS = M.MNFEM_DTEMIS--ORIGINAL:I.INFEM_DTENTR = M.MNFEM_DTENTR
                            AND I.CATG_COD = M.CATG_COD
                            AND I.CADG_COD = M.CADG_COD)
                            INNER JOIN openrisow.FILIAL F
                            ON (F.EMPS_COD = I.EMPS_COD
                            AND F.FILI_COD = I.FILI_COD)
                            LEFT OUTER JOIN openrisow.CLI_FORNEC_TRANSP C
                            ON (C.CADG_COD = I.CADG_COD
                            AND C.CATG_COD = I.CATG_COD
                            AND C.CADG_DAT_ATUA <= M.MNFEM_DTENTR )
                    WHERE  (C.CADG_DAT_ATUA IS NULL
                            OR C.CADG_DAT_ATUA = (SELECT MAX(C2.CADG_DAT_ATUA)
                                                FROM   openrisow.CLI_FORNEC_TRANSP C2
                                                WHERE  C2.CADG_COD = M.CADG_COD
                                                AND C2.CATG_COD = M.CATG_COD
                                                AND C2.CADG_DAT_ATUA <= I.INFEM_DTENTR))
                        AND ( M.MNFEM_IND_CANC IS NULL
                                OR M.MNFEM_IND_CANC != 'S' )
                        AND ( I.IND_CANC IS NULL
                                OR I.IND_CANC != 'S' )
                        AND i.cfop_cod = '1301'  
                        AND f.fili_cod_insest = '%s'--Parametro     
                        AND I.INFEM_DTENTR >= TO_DATE('%s','DD/MM/YYYY')--Parametro
                        AND I.INFEM_DTENTR <= TO_DATE('%s','DD/MM/YYYY')--Parametro
                        
                        ) T
            GROUP  BY T.SF_EMPRESA,
                        T.SF_FILIAL,
                        last_day(T.SF_EMISSAO),
                        T.SF_IE,
                        T.SF_COD_ITEM,
                        T.SF_MUN
                        ,t.COD_UF
            )
            )where SUBSTR(sf_mun, 1, 2) = COD_UF

            union all

            select  EMPS_COD,
                    FILI_COD,
                    INVA_DATA,
                    'U'INVA_TIPO,
                    CASE
                                WHEN UNFE_SIG = 'AC' THEN 'ACIPMS07'
                                WHEN UNFE_SIG = 'BA' THEN 'BAS02'
                                WHEN UNFE_SIG = 'MA' THEN 'MAVAF002'
                                WHEN UNFE_SIG = 'MG' THEN 'Outras_Entradas_a_Detalhar_por_Municipio'
                                WHEN UNFE_SIG = 'PE' THEN 'PEIPMS1'
                                WHEN UNFE_SIG = 'RJ' THEN 'RJVAF10005'
                                WHEN UNFE_SIG = 'RN' THEN 'IPM 4.6'
                                WHEN UNFE_SIG = 'RS' THEN '03'
                                WHEN UNFE_SIG = 'SP' THEN 'SPDIPAM24'
                                WHEN UNFE_SIG = 'TO' THEN 'TOIPMS08'
                    END INVA_COD,
                    MIBGE_COD_MUN,
                    ROUND(valor_corrigido,2)INVA_VALOR,
                    'M'INVA_ORI_CALC
            from (
            select emps_cod,
                fili_cod,
                inva_data,
                MIBGE_COD_MUN,
                COD_UF,
                UNFE_SIG,
                contabil,
                sum(contabil) over (partition by 'x') total_contabil,
                sum(decode(substr(MIBGE_COD_MUN,1,2),COD_UF,0.0,contabil)) over (partition by 'x') erro,
                sum(decode(substr(MIBGE_COD_MUN,1,2),COD_UF,contabil,0.0)) over (partition by 'x') total_rio,
                decode(substr(MIBGE_COD_MUN,1,2),COD_UF,contabil,0.0) / sum(decode(substr(MIBGE_COD_MUN,1,2),COD_UF,contabil,0.0)) over (partition by 'x') indice,
                decode(substr(MIBGE_COD_MUN,1,2),COD_UF,contabil,0.0) + sum(decode(substr(MIBGE_COD_MUN,1,2),COD_UF,0.0,contabil)) over (partition by 'x') * (decode(substr(MIBGE_COD_MUN,1,2),COD_UF,contabil,0.0) / sum(decode(substr(MIBGE_COD_MUN,1,2),COD_UF,contabil,0.0)) over (partition by 'x')) valor_corrigido
            from (
            SELECT emps_cod,
                fili_cod,
                inva_data,
                MIBGE_COD_MUN,
                COD_UF,
                UNFE_SIG,
                sum(INFST_VAL_CONT) contabil
            FROM (
            SELECT /*+ PARALLEL(8) */
                NF.emps_cod,
                NF.fili_cod,
                last_day(NF.mnfst_dtemiss)inva_data,
                F.UNFE_SIG,
                (SELECT MIBGE_COD_MUN
                    FROM openrisow.cli_fornec_transp cli
                WHERE cli.cadg_cod       = nf.cadg_cod
                    AND cli.catg_cod       = nf.catg_cod
                    AND cli.cadg_dat_atua = (select max(cli2.cadg_dat_atua) cadg_dat_atua
                                        FROM openrisow.cli_fornec_transp cli2
                                        WHERE cli2.cadg_cod       = nf.cadg_cod
                                            AND cli2.catg_cod       = nf.catg_cod
                                            AND cli2.cadg_dat_atua  <= nf.MNFST_DTEMISS)) MIBGE_COD_MUN
                                            ,inf.cadg_cod
                                            ,inf.catg_cod
                                            ,INFST_VAL_SERV - INFST_VAL_DESC  INFST_VAL_CONT
                                            ,nf.rowid row_id_mestre
                                            ,SUBSTR(f.fili_mun_ibge, 1, 2) AS COD_UF
            FROM  OPENRISOW.MESTRE_NFTL_SERV NF
            inner join openrisow.ITEM_NFTL_SERV inf on ( INF.EMPS_COD = NF.EMPS_COD
                                                    AND INF.FILI_COD = NF.FILI_COD
                                                    AND INF.INFST_SERIE = NF.MNFST_SERIE
                                                    AND INF.INFST_NUM = NF.MNFST_NUM
                                                    AND INF.INFST_DTEMISS = NF.mnfst_dtemiss
                                                    AND INF.MDOC_COD = NF.MDOC_COD),
                openrisow.filial f 
            WHERE 1 = 1 
            AND nf.emps_cod = f.emps_cod 
            AND nf.fili_cod  = f.fili_cod
            AND f.fili_cod_insest = '%s'--Parametro
            AND nf.MNFST_DTEMISS >= TO_DATE('%s','DD/MM/YYYY')--Parametro
            AND nf.MNFST_DTEMISS <= TO_DATE('%s','DD/MM/YYYY')--Parametro
            AND NF.EMPS_COD = 'TBRA'
            AND nf.FILI_COD = '1710'
            and MNFST_IND_CANC = 'N'
            and cfop in ('5301','5302','5303','5304','5305','5306','5307')
            )
            GROUP BY emps_cod,
                    fili_cod,
                    inva_data,
                    MIBGE_COD_MUN,
                    COD_UF,
                    UNFE_SIG
            )) a 
            where SUBSTR(MIBGE_COD_MUN, 1, 2) = COD_UF


    """%(IE,vDataIni,vDataFim,IE,vDataIni,vDataFim)
    
    retorno=[]
    cursor.executa(query)
    result = cursor.fetchone()
    lin = 0

    if result == None:
        log("#### ATENÇÃO: Nenhum Resultado para aba 6")
        log("####     Query = ")
        log("####")
        log(query)
        log("####")
        ret=0
        return(retorno)
    else:
        while result:
            retorno.append([])
            for campo in result:
                retorno[lin].append(campo)
            lin = lin + 1
            result = cursor.fetchone()
    
    return(retorno)



def main(): 
    log("-"*100)
    log("#### INICIO DO RELATORIO REGISTRO 1400 RJ VERSÃO 20210831 ... ####")
    ret = processar()
    if (ret > 0) :
        log("ERRO - Verifique as mensagens anteriores...")
        log("#### Código de execução = ", ret)
    log("#### FIM DO RELATORIO REGISTRO 1400 RJ VERSÃO 20210831 ... ####")
    sys.exit(ret)   
    
    
if __name__ == "__main__":
    main()
