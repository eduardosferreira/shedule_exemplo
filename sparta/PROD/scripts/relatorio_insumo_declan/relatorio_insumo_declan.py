#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-

"""
----------------------------------------------------------------------------------------------
  SISTEMA ..: 
  MODULO ...: 
  SCRIPT ...: Relatório Insumo_SPED_Fiscal
  CRIACAO ..: 16/07/2021
  AUTOR ....: Airton Borges da Silva Filho - Kyros Consultoria 
  DESCRICAO : Relatório de análises dos insumos SPED
                

----------------------------------------------------------------------------------------------
    Exemplo de comando: ./insumo_declan.py SP 102020 108383949112
    Diretório saída ..: /arquivos/RELATORIOS/SPED/<UF>/<AAAA>/<MM>/
    Arquivo saída ....: Insumo_Consolidado_SPED_Fiscal_<MESANO>_<UF>_<IE>_V<NNN>.xlsx
    Documentação......: CLONE1 /arquivos/TESHUVA/melhorias/XX - Fase 2 - Relatorio SPED Mercadoria/Teshuva_RMSV0_Insumo_SPED_FISCAL_V2.docx
----------------------------------------------------------------------------------------------


  HISTORICO : 
https://www.letscode.com.br/blog/aprenda-a-integrar-python-e-excel
-cd /arquivos/TESHUVA/scripts_rpa/insumo_DECLAN/
-echo "" > insumo_declan.py & nano insumo_declan.py
-./insumo_declan.py rj 122016 77452443 

 31/08/2021 - Adquações para novo painel
----------------------------------------------------------------------------------------------
"""
import sys
import os
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)
import configuracoes

import datetime

import re
import cx_Oracle
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_MEDIUM, BORDER_DOUBLE
from openpyxl.reader.excel import load_workbook 
from openpyxl.utils import get_column_letter

sys.path.append( os.path.join( os.path.realpath('..'), "modulosPython" ) )
nome_script = os.path.basename( sys.argv[0] ).replace('.py', '')

import comum
import sql

global variaveis
global db
global listadeabas
global qtdabas

DEBUG = False
#DEBUG = True


#(01    "Aba1.Dados DECLAN"              # 0) RF01
#(02    "Aba2.Resumo CFOP Entradas"      , 1) RF02
#(03    "Aba3.Resumo CFOP Saidas"        , 2) RF03
#(04    "Aba4.Valor Adicionado por Munic", 3) RF04



listadeabas = ('01','02','03','04')
#listadeabas = ('04')


qtdabas = 4

fontMasterPreta  = Font(color='00000000', bold=True, size=12)
fontNegrito      = Font(color='00000000', bold=True)

SD = ('/' if os.name == 'posix' else '\\')

ret = 0



def processar():
    ret     = 0
    ufi     = ""
    mesanoi = ""
    mesi    = ""
    anoi    = ""
    mesanof = ""
    mesf    = ""
    anof    = ""
    global listadeabas
    try :
        if (len(sys.argv) == 4 and len(sys.argv[1]) > 3 and len(sys.argv[2]) == 6 and len(sys.argv[3]) == 6 and sys.argv[2] <= sys.argv[3]) :
            ufi     = "RJ"
            iei     = sys.argv[1].upper()
            mesanoi = sys.argv[2].upper()
            mesi    = sys.argv[2][:2].upper()
            anoi    = sys.argv[2][2:].upper()
            iei     = re.sub('[^0-9]','',iei)
            mesanof = sys.argv[3].upper()
            mesf    = sys.argv[3][:2].upper()
            anof    = sys.argv[3][2:].upper()

            if (anoi != anof):
                ret = 99

        else :
            ret = 99
    except Exception as e :
        log('#### Erro encontrado :', e)
        ret = 99
    if ret == 99:
        log("-" * 100)
        log("#### ")
        log('#### ERRO - Erro nos parametros do script.')
        log("#### ")
        log('#### Exemplo de como deve ser :')
        log('####      %s <IE> <mmaaaa> <MMAAAA>'%(sys.argv[0]))
        log("#### ")
        log('#### Onde')
        log('####      <IE> = inscrição desejada. Ex: 77452443')
        log('####      <mmaaaa> = mês e ano inicial. Ex: para janeiro de 2020 informe 012020')
        log('####      <MMAAAA> = mês e ano final. Ex: Para junho de 2020 informe 062020')
        log("#### ")
        log('####      <aaaa> e <AAAA> devem ser iguais, ou seja: mesmo ano.')
        log("#### ")
        log("#### Exemplo:  Para 01/01/2020 a 31/01/2020 e ie = 77452443")
        log("#### %s 77452443 012020 012020"%(sys.argv[0]))
        log("#### ")
        log('#### A planilha resultado será:')
        log('####      %sarquivos%sdeclan%srelatorio%s77452443-DECLAN-2020.xlsx'%(SD,SD,SD,SD))
        log('#### ')
        log("-" * 100)
        print ("")
        print ("Retorno = 99")
        return(ret)

    #SETA O ARQUIVO DE SAIDA INSUMOS (PLANILHA EXCEL)
    dir_arq = configuracoes.diretorio_arquivos
    dir_rel_INSUMOS     = os.path.join(SD,dir_arq,ufi,anoi)
    rel_INSUMOS         = iei+"-DECLAN-"+str(anoi)+".xlsx"
    ARQ_INSUMOS         = os.path.join(dir_rel_INSUMOS,rel_INSUMOS)

    
    log("-" * 100)    
    log(" ====> INSUMO = ",ARQ_INSUMOS)  
    log("-" * 100)
  
    #### Se a pasta do relatório INSUMO não existir, cria
    if not os.path.isdir(dir_rel_INSUMOS) :
        os.makedirs(dir_rel_INSUMOS)

    #### Cria a planilha excel INSUMO em memória....
    arquivo_excel = Workbook()
    ABA_DD       = arquivo_excel.active
    ABA_DD.title =                                "Aba1.Dados DECLAN"              #0)
    if ('02' in listadeabas):
        ABA_RCE      = arquivo_excel.create_sheet("Aba2.Resumo CFOP Entradas"      ,1)
    if ('03' in listadeabas):
        ABA_RCF     = arquivo_excel.create_sheet("Aba3.Resumo CFOP Saidas"         ,2)
    if ('04' in listadeabas):
        ABA_VAM     = arquivo_excel.create_sheet("Aba4.Valor Adicionado por Munic", 3)

    ##########  ABA 1-   ABA_DD = Aba1.Dados DECLAN 0
    ##########  ABA 1-   ABA_DD = Aba1.Dados DECLAN 0
    ##########  ABA 1-   ABA_DD = Aba1.Dados DECLAN 0
    ##########  ABA 1-   ABA_DD = Aba1.Dados DECLAN 0
    ##########  ABA 1-   ABA_DD = Aba1.Dados DECLAN 0
    ##########  ABA 1-   ABA_DD = Aba1.Dados DECLAN 0

    if ('01' in listadeabas):
        log("ABA 01 / ", qtdabas , " -> Aba1.Dados DECLAN")

        linhaABA_DD = {}

        ABA_DD.cell(row=1, column=1,  value="Declaração Anual para o IPM")
        ABA_DD.cell(row=2, column=1,  value = "IE:  " + str(iei))
        ABA_DD.cell(row=3, column=1,  value = "Ano:  " + str(anoi))


        ##### QUADRO 1
        ABA_DD.cell(row=5, column=1,  value="Quadro 1 Resumo Geral Operações e Prestações")

        ##### Quadro 1.1
        estado11   = 0.00
        outro11    = 0.00
        exterior11 = 0.00
        total11    = 0.00
        dadosABA_DD = Busca_DD11(iei,"01/"+mesi+"/"+anoi, "01/"+mesf+"/"+anof)
        for dados in dadosABA_DD:
            if (dados[0] == '1.Estado'):
                estado11 = dados[1]
            elif (dados[0] == '2.Outro Estado'):
                outro11 = dados[1]
            elif (dados[0] == '3.Exterior'):
                exterior11 = dados[1]
            total11 = total11 + dados[1]
        ABA_DD.cell(row=7, column=1,  value="Quadro 1.1 - Entradas Ano-Base")
        ABA_DD.cell(row=7, column=2,  value="Valor Contábil")
        ABA_DD.cell(row=8, column=1,  value="1. Estado")
        ABA_DD.cell(row=8, column=2,  value=estado11)
        ABA_DD.cell(row=9, column=1,  value="2. Outro Estado")
        ABA_DD.cell(row=9, column=2,  value=outro11)
        ABA_DD.cell(row=10, column=1,  value="3. Exterior")
        ABA_DD.cell(row=10, column=2,  value=exterior11)
        ABA_DD.cell(row=12, column=1,  value="Total - Quadro 1.1")
        ABA_DD.cell(row=12, column=2,  value=total11)

        ##### Quadro 1.2
        estado12   = 0.00
        outro12    = 0.00
        exterior12 = 0.00
        total12    = 0.00
        dadosABA_DD = Busca_DD12(iei,"01/"+mesi+"/"+anoi,"01/"+mesf+"/"+anof)
        for dados in dadosABA_DD:
            if (dados[0] == '1.Estado'):
                estado12 = dados[1]
            elif (dados[0] == '2.Outro Estado'):
                outro12 = dados[1]
            elif (dados[0] == '3.Exterior'):
                exterior12 = dados[1]
            total12 = total12 + dados[1]
        lin=14
        ABA_DD.cell(row=14, column=1,  value="Quadro 1.2 - Saias Ano-Base")
        ABA_DD.cell(row=14, column=2,  value="Valor Contábil")
        ABA_DD.cell(row=15, column=1,  value="1. Estado")
        ABA_DD.cell(row=15, column=2,  value=estado12)
        ABA_DD.cell(row=16, column=1,  value="2. Outro Estado")
        ABA_DD.cell(row=16, column=2,  value=outro12)
        ABA_DD.cell(row=17, column=1,  value="3. Exterior")
        ABA_DD.cell(row=17, column=2,  value=exterior12)
        ABA_DD.cell(row=19, column=1,  value="Total - Quadro 1.2")
        ABA_DD.cell(row=19, column=2,  value=total12)



        ##### QUADRO 2
        ABA_DD.cell(row=21, column=1,  value="Quadro 2 Resumo Especifico de Operações com Mercadorias")

        ##### Quadro 2.1
        estado21   = 0.00
        outro21    = 0.00
        exterior21 = 0.00
        total21    = 0.00
        dadosABA_DD = Busca_DD21(iei,"01/"+mesi+"/"+anoi,"01/"+mesf+"/"+anof)
        for dados in dadosABA_DD:
            if (dados[0] == '1.Estado'):
                estado21 = dados[1]
            elif (dados[0] == '2.Outro Estado'):
                outro21 = dados[1]
            elif (dados[0] == '3.Exterior'):
                exterior21 = dados[1]
            total21 = total21 + dados[1]
        ABA_DD.cell(row=23, column=1,  value="Quadro 2.1 - Entradas Ano-Base")
        ABA_DD.cell(row=23, column=2,  value="Valor Contábil")
        ABA_DD.cell(row=24, column=1,  value="1. Estado")
        ABA_DD.cell(row=24, column=2,  value=estado21)
        ABA_DD.cell(row=25, column=1,  value="2. Outro Estado")
        ABA_DD.cell(row=25, column=2,  value=outro21)
        ABA_DD.cell(row=26, column=1,  value="3. Exterior")
        ABA_DD.cell(row=26, column=2,  value=exterior21)
        ABA_DD.cell(row=28, column=1,  value="Total - Quadro 2.1")
        ABA_DD.cell(row=28, column=2,  value=total21)

        ##### Quadro 2.2
        estado22   = 0.00
        outro22    = 0.00
        exterior22 = 0.00
        total22    = 0.00
        dadosABA_DD = Busca_DD22(iei,"01/"+mesi+"/"+anoi,"01/"+mesf+"/"+anof)
        for dados in dadosABA_DD:
            if (dados[0] == '1.Estado'):
                estado22 = dados[1]
            elif (dados[0] == '2.Outro Estado'):
                outro22 = dados[1]
            elif (dados[0] == '3.Exterior'):
                exterior22 = dados[1]
            total22 = total22 + dados[1]
        ABA_DD.cell(row=30, column=1,  value="Quadro 2.1 - Entradas Ano-Base")
        ABA_DD.cell(row=30, column=2,  value="Valor Contábil")
        ABA_DD.cell(row=31, column=1,  value="1. Estado")
        ABA_DD.cell(row=31, column=2,  value=estado22)
        ABA_DD.cell(row=32, column=1,  value="2. Outro Estado")
        ABA_DD.cell(row=32, column=2,  value=outro22)
        ABA_DD.cell(row=33, column=1,  value="3. Exterior")
        ABA_DD.cell(row=33, column=2,  value=exterior22)
        ABA_DD.cell(row=35, column=1,  value="Total - Quadro 2.2")
        ABA_DD.cell(row=35, column=2,  value=total22)



        ##### QUADRO 3
        ABA_DD.cell(row=37, column=1,  value="Quadro 3 Ajuste do Valor Adicionado e Outras Informações Economico - Fiscais")

        ##### Quadro 2.1
        ativo31   = 0.00
        consumo31 = 0.00
        nutil31   = 0.00
        icms31    = 0.00
        total31   = 0.00
        dadosABA_DD = Busca_DD31(iei,"01/"+mesi+"/"+anoi,"01/"+mesf+"/"+anof)
        for dados in dadosABA_DD:
            if (dados[0][0]   == '1'):
                ativo31        = dados[1]
            elif (dados[0][0] == '2'):
                consumo31      = dados[1]
            elif (dados[0][0] == '3'):
                nutil31        = dados[1]
            elif (dados[0][0] == '4'):
                icms31         = dados[1]
            total31 = total31 + dados[1]
        ABA_DD.cell(row=39, column=1,  value="3.1 Entradas")
        ABA_DD.cell(row=39, column=2,  value="Valor Contábil")
        ABA_DD.cell(row=40, column=1,  value="1. Operações Relativas ao Ativo Imobilizado")
        ABA_DD.cell(row=40, column=2,  value=ativo31)
        ABA_DD.cell(row=41, column=1,  value="2. Operações Relativas ao Uso e Consumo")
        ABA_DD.cell(row=41, column=2,  value=consumo31)
        ABA_DD.cell(row=42, column=1,  value="3. Operações / Prestações que não são Fato Gerador do ICMS ou nao Utilizadas no VA")
        ABA_DD.cell(row=42, column=2,  value=nutil31)
        ABA_DD.cell(row=43, column=1,  value="4. ICMS Retido por Substituição Tributária")
        ABA_DD.cell(row=43, column=2,  value=icms31)
        ABA_DD.cell(row=45, column=1,  value="Total - Quadro 3.1")
        ABA_DD.cell(row=45, column=2,  value=total31)

        ##### Quadro 3.2
        ativo32   = 0.00
        consumo32 = 0.00
        nutil32   = 0.00
        icms32    = 0.00
        total32   = 0.00
        dadosABA_DD = Busca_DD32(iei,"01/"+mesi+"/"+anoi,"01/"+mesf+"/"+anof)
        for dados in dadosABA_DD:
            if (dados[0][0]   == '1'):
                ativo32        = dados[1]
            elif (dados[0][0] == '2'):
                consumo32      = dados[1]
            elif (dados[0][0] == '3'):
                nutil32        = dados[1]
            elif (dados[0][0] == '4'):
                icms32         = dados[1]
            total32 = total32 + dados[1]
        ABA_DD.cell(row=47, column=1,  value="3.2 Saidas")
        ABA_DD.cell(row=47, column=2,  value="Valor Contábil")
        ABA_DD.cell(row=48, column=1,  value="1. Operações Relativas ao Ativo Imobilizado")
        ABA_DD.cell(row=48, column=2,  value=ativo32)
        ABA_DD.cell(row=49, column=1,  value="2. Operações Relativas ao Uso e Consumo")
        ABA_DD.cell(row=49, column=2,  value=consumo32)
        ABA_DD.cell(row=50, column=1,  value="3. Operações / Prestações que não são Fato Gerador do ICMS ou nao Utilizadas no VA")
        ABA_DD.cell(row=50, column=2,  value=nutil32)
        ABA_DD.cell(row=51, column=1,  value="4. ICMS Retido por Substituição Tributária")
        ABA_DD.cell(row=51, column=2,  value=icms32)
        ABA_DD.cell(row=53, column=1,  value="Total - Quadro 3.2")
        ABA_DD.cell(row=53, column=2,  value=total32)

        formata_DD(ABA_DD)
        arquivo_excel.save(ARQ_INSUMOS)


    ##########  ABA 2 -   ABA_RCE      = Aba2.Resumo CFOP Entradas      ,1
    ##########  ABA 2 -   ABA_RCE      = Aba2.Resumo CFOP Entradas      ,1
    ##########  ABA 2 -   ABA_RCE      = Aba2.Resumo CFOP Entradas      ,1
    ##########  ABA 2 -   ABA_RCE      = Aba2.Resumo CFOP Entradas      ,1
    ##########  ABA 2 -   ABA_RCE      = Aba2.Resumo CFOP Entradas      ,1
    ##########  ABA 2 -   ABA_RCE      = Aba2.Resumo CFOP Entradas      ,1


    if ('02' in listadeabas):
        log("ABA 02 / ", qtdabas , " -> Aba2.Resumo CFOP Entradas")
        fundoclaro  = 'B7DFE6'
        fundoescuro = '92ABB0'
        # TAMANHO DAS COLUNAS:
        ABA_RCE.column_dimensions['A'].width = 2
        ABA_RCE.column_dimensions['B'].width = 20
        ABA_RCE.column_dimensions['C'].width = 20
        ABA_RCE.column_dimensions['D'].width = 20
        ABA_RCE.column_dimensions['E'].width = 20
        ABA_RCE.column_dimensions['F'].width = 20
        ABA_RCE.column_dimensions['G'].width = 20
        ABA_RCE.column_dimensions['H'].width = 20
        ABA_RCE.column_dimensions['I'].width = 20
        ABA_RCE.column_dimensions['J'].width = 20
        ABA_RCE.column_dimensions['K'].width = 20
        ABA_RCE.column_dimensions['L'].width = 20
        ABA_RCE.column_dimensions['M'].width = 20
        ABA_RCE.column_dimensions['N'].width = 20
        ABA_RCE.column_dimensions['O'].width = 20
        ABA_RCE.column_dimensions['P'].width = 20
        # TITULO
        ABA_RCE.cell(row=2, column=2,  value="Quadro 1.Resumo CFOP - Entradas Mercadoria e Telecom")
        ABA_RCE.cell(row=3, column=2,  value = "IE:  " + str(iei))
        # FORMATACAO
        fontMasterPretaG    = Font(color='00000000', bold=True, size=15)
        fontMasterPretaM    = Font(color='00000000', bold=True, size=13)
        fontMasterPretaP    = Font(color='00000000', bold=True, size=11)
        ABA_RCE.cell(2,  2).font = fontMasterPretaG
        ABA_RCE.cell(2,  2).alignment = Alignment(horizontal='center')
        ABA_RCE.merge_cells('B2:O2')
        ABA_RCE.cell(3,  2).font = fontMasterPretaG
        ABA_RCE.merge_cells('B3:O3')
        ABA_RCE.merge_cells('B4:O4')
        cor_fundo_p_ci_li_cf_lf_c(ABA_RCE,2,2,15,3, fundoescuro)
        ABA_RCE.freeze_panes = 'A4'
        # BORDAS
        set_border_edsi(ABA_RCE, 'B2:O2' , 'fffn')
        set_border_edsi(ABA_RCE, 'B3:O3' , 'ffnf')
        lin=4
        ##### QUADRO 1.1
        ##### QUADRO 1.1
        ##### QUADRO 1.1
        ##### CABEÇALHO
        cabecalho_RCE(ABA_RCE,2,lin,"Quadro 1.1.Estado",str(anoi))
        linhaABA_RCE = {}
        dadosABA_RCE = Busca_RCE11(iei,"01/"+mesi+"/"+anoi,"01/"+mesf+"/"+anof)
        dad=0
        lin = lin + 4
        lini1 = lin
        lini = lin
        for dados in dadosABA_RCE:
            ABA_RCE.cell(row=lin, column=2,  value=dados[0])
            for dad in range(1,len(dados)):
                ABA_RCE.cell(row=lin, column=dad+2,value=dados[dad]).number_format = "#,##0.00"
            lin=lin+1
        linf1 = lin
        lin=lin+1
        total_RCE(ABA_RCE,lini,lin)
        lin=lin+2
        arquivo_excel.save(ARQ_INSUMOS)

        ##### QUADRO 1.2
        ##### QUADRO 1.2
        ##### QUADRO 1.2
        ##### CABEÇALHO
        cabecalho_RCE(ABA_RCE,2,lin,"Quadro 1.2.Outro Estado",str(anoi))
        linhaABA_RCE = {}
        dadosABA_RCE = Busca_RCE12(iei,"01/"+mesi+"/"+anoi,"01/"+mesf+"/"+anof)
        dad=0
        lin = lin + 4
        lini = lin
        lini2=lin
        for dados in dadosABA_RCE:
            ABA_RCE.cell(row=lin, column=2,  value=dados[0])
            for dad in range(1,len(dados)):
                ABA_RCE.cell(row=lin, column=dad+2,value=dados[dad]).number_format = "#,##0.00"
            lin=lin+1
        linf2=lin
        lin=lin+1
        total_RCE(ABA_RCE,lini,lin)
        lin=lin+2
        arquivo_excel.save(ARQ_INSUMOS)

        ##### QUADRO 1.3
        ##### QUADRO 1.3
        ##### QUADRO 1.3
        ##### CABEÇALHO
        cabecalho_RCE(ABA_RCE,2,lin,"Quadro 1.3.Exterior",str(anoi))
        linhaABA_RCE = {}
        dadosABA_RCE = Busca_RCE13(iei,"01/"+mesi+"/"+anoi,"01/"+mesf+"/"+anof)
        dad=0
        lin = lin + 4
        lini = lin
        lini3=lin
        for dados in dadosABA_RCE:
            ABA_RCE.cell(row=lin, column=2,  value=dados[0])
            for dad in range(1,len(dados)):
                ABA_RCE.cell(row=lin, column=dad+2,value=dados[dad]).number_format = "#,##0.00"
            lin=lin+1
        linf3=lin
        lin=lin+1
        total_RCE(ABA_RCE,lini,lin)
        lin=lin+3
        arquivo_excel.save(ARQ_INSUMOS)

        c=3
        geral = 0.00
        for catu in ('C','D','E','F','G','H','I','J','K','L','M','N'):
            soma = 0.00
            for l in range(lini1,linf1):
                soma=soma+(0.00 if (ABA_RCE.cell(row=l, column=c).value == None) else float(ABA_RCE.cell(row=l, column=c).value))
            for l in range(lini2,linf2):
                soma=soma+(0.00 if (ABA_RCE.cell(row=l, column=c).value == None) else float(ABA_RCE.cell(row=l, column=c).value))
            for l in range(lini3,linf3):
                soma=soma+(0.00 if (ABA_RCE.cell(row=l, column=c).value == None) else float(ABA_RCE.cell(row=l, column=c).value))
            ABA_RCE.cell(row=lin, column=c,  value=soma).font = fontMasterPretaM
            ABA_RCE.cell(row=lin, column=c).number_format = "#,##0.00"
            c=c+1
            geral=geral+soma
        ABA_RCE.cell(row=lin, column=c,  value=geral).font = fontMasterPretaM
        ABA_RCE.cell(row=lin, column=c).number_format = "#,##0.00"
        ABA_RCE.cell(row=lin, column=c).font = fontMasterPretaM

        ABA_RCE.cell(row=lin, column=2,  value="TOTAL QUADRO 1" ).font = fontMasterPretaM
        set_border_edsi(ABA_RCE, 'B'+str(lin)+':O'+str(lin),'gggg')
        contorno_cli_clf( ABA_RCE, 2,2,15,lin, 'g')

        arquivo_excel.save(ARQ_INSUMOS)


    ##########  ABA 3 -   ABA_RCF      = Aba2.Resumo CFOP Saidas        ,2
    ##########  ABA 3 -   ABA_RCF      = Aba2.Resumo CFOP Saidas        ,2
    ##########  ABA 3 -   ABA_RCF      = Aba2.Resumo CFOP Saidas        ,2
    ##########  ABA 3 -   ABA_RCF      = Aba2.Resumo CFOP Saidas        ,2
    ##########  ABA 3 -   ABA_RCF      = Aba2.Resumo CFOP Saidas        ,2
    
    
    if ('03' in listadeabas):
        log("ABA 03 / ", qtdabas , " -> Aba3.Resumo CFOP Saidas")
        fundoclaro  = 'B7DFE6'
        fundoescuro = '92ABB0'
        # TAMANHO DAS COLUNAS:
        ABA_RCF.column_dimensions['A'].width = 2
        ABA_RCF.column_dimensions['B'].width = 20
        ABA_RCF.column_dimensions['C'].width = 20
        ABA_RCF.column_dimensions['D'].width = 20
        ABA_RCF.column_dimensions['E'].width = 20
        ABA_RCF.column_dimensions['F'].width = 20
        ABA_RCF.column_dimensions['G'].width = 20
        ABA_RCF.column_dimensions['H'].width = 20
        ABA_RCF.column_dimensions['I'].width = 20
        ABA_RCF.column_dimensions['J'].width = 20
        ABA_RCF.column_dimensions['K'].width = 20
        ABA_RCF.column_dimensions['L'].width = 20
        ABA_RCF.column_dimensions['M'].width = 20
        ABA_RCF.column_dimensions['N'].width = 20
        ABA_RCF.column_dimensions['O'].width = 20
        ABA_RCF.column_dimensions['P'].width = 20
        # TITULO
        ABA_RCF.cell(row=2, column=2,  value="Quadro 1.Resumo CFOP - Saídas Mercadoria e Telecom")
        ABA_RCF.cell(row=3, column=2,  value = "IE:  " + str(iei))
        # FORMATACAO
        fontMasterPretaG    = Font(color='00000000', bold=True, size=15)
        fontMasterPretaM    = Font(color='00000000', bold=True, size=13)
        fontMasterPretaP    = Font(color='00000000', bold=True, size=11)
        ABA_RCF.cell(2,  2).font = fontMasterPretaG
        ABA_RCF.cell(2,  2).alignment = Alignment(horizontal='center')
        ABA_RCF.merge_cells('B2:O2')
        ABA_RCF.cell(3,  2).font = fontMasterPretaG
        ABA_RCF.merge_cells('B3:O3')
        ABA_RCF.merge_cells('B4:O4')
        cor_fundo_p_ci_li_cf_lf_c(ABA_RCF,2,2,15,3, fundoescuro)
        ABA_RCF.freeze_panes = 'A4'
        # BORDAS
        set_border_edsi(ABA_RCF, 'B2:O2' , 'fffn')
        set_border_edsi(ABA_RCF, 'B3:O3' , 'ffnf')
        lin=4
        ##### QUADRO 1.1
        ##### QUADRO 1.1
        ##### QUADRO 1.1
        ##### CABEÇALHO
        cabecalho_RCE(ABA_RCF,2,lin,"Quadro 1.1.Estado",str(anoi))
        linhaABA_RCF = {}
        dadosABA_RCF = Busca_RCF11(iei,"01/"+mesi+"/"+anoi,"01/"+mesf+"/"+anof)
        dad=0
        lin = lin + 4
        lini1 = lin
        lini = lin
        for dados in dadosABA_RCF:
            ABA_RCF.cell(row=lin, column=2,  value=dados[0])
            for dad in range(1,len(dados)):
                ABA_RCF.cell(row=lin, column=dad+2,value=dados[dad]).number_format = "#,##0.00"
            lin=lin+1
        linf1 = lin
        lin=lin+1
        total_RCE(ABA_RCF,lini,lin)
        lin=lin+2
        arquivo_excel.save(ARQ_INSUMOS)
    
        ##### QUADRO 1.2
        ##### QUADRO 1.2
        ##### QUADRO 1.2
        ##### CABEÇALHO
        cabecalho_RCE(ABA_RCF,2,lin,"Quadro 1.2.Outro Estado",str(anoi))
        linhaABA_RCF = {}
        dadosABA_RCF = Busca_RCF12(iei,"01/"+mesi+"/"+anoi,"01/"+mesf+"/"+anof)
        dad=0
        lin = lin + 4
        lini = lin
        lini2=lin
        for dados in dadosABA_RCF:
            ABA_RCF.cell(row=lin, column=2,  value=dados[0])
            for dad in range(1,len(dados)):
                ABA_RCF.cell(row=lin, column=dad+2,value=dados[dad]).number_format = "#,##0.00"
            lin=lin+1
        linf2=lin
        lin=lin+1
        total_RCE(ABA_RCF,lini,lin)
        lin=lin+2
        arquivo_excel.save(ARQ_INSUMOS)
    
        ##### QUADRO 1.3
        ##### QUADRO 1.3
        ##### QUADRO 1.3
        ##### CABEÇALHO
        cabecalho_RCE(ABA_RCF,2,lin,"Quadro 1.3.Exterior",str(anoi))
        linhaABA_RCF = {}
        dadosABA_RCF = Busca_RCF13(iei,"01/"+mesi+"/"+anoi,"01/"+mesf+"/"+anof)
        dad=0
        lin = lin + 4
        lini = lin
        lini3=lin
        for dados in dadosABA_RCF:
            ABA_RCF.cell(row=lin, column=2,  value=dados[0])
            for dad in range(1,len(dados)):
                ABA_RCF.cell(row=lin, column=dad+2,value=dados[dad]).number_format = "#,##0.00"
            lin=lin+1
        linf3=lin
        lin=lin+1
        total_RCE(ABA_RCF,lini,lin)
        lin=lin+3
        arquivo_excel.save(ARQ_INSUMOS)
    
        c=3
        geral = 0.00
        for catu in ('C','D','E','F','G','H','I','J','K','L','M','N'):
            soma = 0.00
            for l in range(lini1,linf1):
                soma=soma+(0.00 if (ABA_RCF.cell(row=l, column=c).value == None) else float(ABA_RCF.cell(row=l, column=c).value))
            for l in range(lini2,linf2):
                soma=soma+(0.00 if (ABA_RCF.cell(row=l, column=c).value == None) else float(ABA_RCF.cell(row=l, column=c).value))
            for l in range(lini3,linf3):
                soma=soma+(0.00 if (ABA_RCF.cell(row=l, column=c).value == None) else float(ABA_RCF.cell(row=l, column=c).value))
            ABA_RCF.cell(row=lin, column=c,  value=soma).font = fontMasterPretaM
            ABA_RCF.cell(row=lin, column=c).number_format = "#,##0.00"
            c=c+1
            geral=geral+soma
        ABA_RCF.cell(row=lin, column=c,  value=geral).font = fontMasterPretaM
        ABA_RCF.cell(row=lin, column=c).number_format = "#,##0.00"
        ABA_RCF.cell(row=lin, column=c).font = fontMasterPretaM
    
        ABA_RCF.cell(row=lin, column=2,  value="TOTAL QUADRO 1" ).font = fontMasterPretaM
        set_border_edsi(ABA_RCF, 'B'+str(lin)+':O'+str(lin),'gggg')
        contorno_cli_clf( ABA_RCF, 2,2,15,lin, 'g')
    
        arquivo_excel.save(ARQ_INSUMOS)



    ##########  ABA 4 -   ABA_VAM      = Quadro 1. Valor Adicionado por Município (Telecom)      ,3
    ##########  ABA 4 -   ABA_VAM      = Quadro 1. Valor Adicionado por Município (Telecom)      ,3
    ##########  ABA 4 -   ABA_VAM      = Quadro 1. Valor Adicionado por Município (Telecom)      ,3
    ##########  ABA 4 -   ABA_VAM      = Quadro 1. Valor Adicionado por Município (Telecom)      ,3
    ##########  ABA 4 -   ABA_VAM      = Quadro 1. Valor Adicionado por Município (Telecom)      ,3
    ##########  ABA 4 -   ABA_VAM      = Quadro 1. Valor Adicionado por Município (Telecom)      ,3
    if ('04' in listadeabas):
        log("ABA 04 / ", qtdabas , " -> Aba4.Valor Adicionado por Município")
        fontMasterPretaG    = Font(color='00000000', bold=True, size=15)
        fontMasterPretaM    = Font(color='00000000', bold=True, size=13)
        fontMasterPretaP    = Font(color='00000000', bold=True, size=11)
        fundoclaro          = 'B7DFE6'
        fundoescuro         = '92ABB0'
        linhaABA_VAM = {}

        # TAMANHO DAS COLUNAS:
        ABA_VAM.column_dimensions['A'].width = 2
        ABA_VAM.column_dimensions['B'].width = 4
        ABA_VAM.column_dimensions['C'].width = 50
        ABA_VAM.column_dimensions['D'].width = 20
        ABA_VAM.column_dimensions['E'].width = 25
        ABA_VAM.column_dimensions['F'].width = 25
        ABA_VAM.column_dimensions['G'].width = 25
        ABA_VAM.column_dimensions['H'].width = 25
        ABA_VAM.column_dimensions['I'].width = 2
        ABA_VAM.column_dimensions['J'].width = 20
        ABA_VAM.column_dimensions['K'].width = 30



        # Quadro 1
        ABA_VAM.cell(row=2, column=2,  value="Quadro 1. Valor Adicionado por Município (Telecom)").font = fontMasterPretaG
        ABA_VAM.cell(row=3, column=2,  value = "IE:  " + str(iei)).font = fontMasterPretaM
        ABA_VAM.cell(row=4, column=2,  value = "Ano:  " + str(anoi)).font = fontMasterPretaM
        ABA_VAM.cell(row=5, column=2,  value="UF").font = fontMasterPretaM
        ABA_VAM.cell(row=5, column=3,  value="Descrição Municipio").font = fontMasterPretaM
        ABA_VAM.cell(row=5, column=4,  value="Municipio IBGE").font = fontMasterPretaM
        ABA_VAM.cell(row=5, column=5,  value="Valor Contábil").font = fontMasterPretaM
        ABA_VAM.cell(row=5, column=6,  value="Percentual Calculado ").font = fontMasterPretaM
        ABA_VAM.cell(row=5, column=7,  value="Valor Diferença").font = fontMasterPretaM
        ABA_VAM.cell(row=5, column=8,  value="Valor Adicionado").font = fontMasterPretaM

        ABA_VAM.cell(2,  2).alignment = Alignment(horizontal='center')
        ABA_VAM.cell(5,  2).alignment = Alignment(horizontal='center')
        ABA_VAM.cell(5,  3).alignment = Alignment(horizontal='center')
        ABA_VAM.cell(5,  4).alignment = Alignment(horizontal='center')
        ABA_VAM.cell(5,  5).alignment = Alignment(horizontal='center')
        ABA_VAM.cell(5,  6).alignment = Alignment(horizontal='center')
        ABA_VAM.cell(5,  7).alignment = Alignment(horizontal='center')
        ABA_VAM.cell(5,  8).alignment = Alignment(horizontal='center')

        cor_fundo_p_ci_li_cf_lf_c(ABA_VAM,2,2,8,4, fundoescuro)
        ABA_VAM.merge_cells('B2:H2')
        ABA_VAM.merge_cells('B3:H3')
        ABA_VAM.merge_cells('B4:H4')
        contorno_cli_clf( ABA_VAM, 2,2,8,4, 'g')
        set_border_edsi(ABA_VAM,'B5:H5','gggg')

        ABA_VAM.freeze_panes = 'A6'

        ##### Dados Quadro 1

        dadosABA_VAM = Busca_VAM(iei,"01/"+mesi+"/"+anoi, "01/"+mesf+"/"+anof)

        lin = 6
        for dados in dadosABA_VAM:
            for dad in range(0,len(dados)):
                if (dad > 2):
                    if (dad == 4):
                        ABA_VAM.cell(row=lin, column=dad+2,value=dados[dad]).number_format = "#,##0.000000000"
                    else:
                        ABA_VAM.cell(row=lin, column=dad+2,value=dados[dad]).number_format = "#,##0.00"
                else:
                    ABA_VAM.cell(row=lin, column=dad+2,value=dados[dad])
            lin=lin+1


        set_border_edsi(ABA_VAM, 'B6:H'+str(lin+1),'ffff')
        contorno_cli_clf( ABA_VAM, 2,6,8,lin-1, 'g')
        set_border_edsi(ABA_VAM, 'B5:H5','gggg')


        ABA_VAM.merge_cells('B'+str(lin)+':H'+str(lin+1))
        set_border_edsi(ABA_VAM, 'B'+str(lin+2)+':H'+str(lin+2),'gggg')
        lin=lin+2
        ABA_VAM.cell(row=lin, column=2,  value="TOTAL")
        ABA_VAM.cell(row=lin, column=2).font = fontMasterPretaM
        ABA_VAM.cell(row=lin, column=2).alignment = Alignment(horizontal='center')
        ABA_VAM.merge_cells('B'+str(lin)+':D'+str(lin))

        ABA_VAM.cell(row=lin, column=5,  value="=SUM(E6:E"+str(lin-3)+")" ).font = fontMasterPretaM
        ABA_VAM.cell(row=lin, column=5).number_format = "#,##0.00"
        ABA_VAM.cell(row=lin, column=6,  value="=SUM(F6:F"+str(lin-3)+")" ).font = fontMasterPretaM
        ABA_VAM.cell(row=lin, column=6).number_format = "#,##0.000000000"
        ABA_VAM.cell(row=lin, column=7,  value="=SUM(G6:G"+str(lin-3)+")" ).font = fontMasterPretaM
        ABA_VAM.cell(row=lin, column=7).number_format = "#,##0.00"
        ABA_VAM.cell(row=lin, column=8,  value="=SUM(H6:H"+str(lin-3)+")" ).font = fontMasterPretaM
        ABA_VAM.cell(row=lin, column=8).number_format = "#,##0.00"


        ##### QUADRO 2
        ##### QUADRO 2
        ##### QUADRO 2
        ##### QUADRO 2
        ABA_VAM.cell(row=2, column=10,  value="Quadro 2. Resumo Telecom por Município").font = fontMasterPretaM
        ABA_VAM.cell(row=2, column=10).alignment = Alignment(horizontal='center')
        cor_fundo_p_ci_li_cf_lf_c(ABA_VAM,10,2,11,2, fundoescuro)
        ABA_VAM.merge_cells('J2:K2')

        ABA_VAM.cell(row=3, column=10,  value="Municípios RJ")
        ABA_VAM.cell(row=3, column=11,  value="=SUM(E6:E"+str(lin-3)+")" ).number_format = "#,##0.00"
        ABA_VAM.cell(row=4, column=10,  value="Demais UF's")
        ABA_VAM.cell(row=4, column=11,  value="=SUM(G6:G"+str(lin-3)+")" ).number_format = "#,##0.00"
        ABA_VAM.cell(row=5, column=10,  value="Total").font = fontMasterPretaM
        ABA_VAM.cell(row=5, column=11,  value="=SUM(H6:H"+str(lin-3)+")" ).number_format = "#,##0.00"
        ABA_VAM.cell(row=5, column=11).font = fontMasterPretaM
        set_border_edsi(ABA_VAM, 'J2:K5','gggg')
        set_border_edsi(ABA_VAM, 'J3:J3','gfgf')
        set_border_edsi(ABA_VAM, 'J4:J4','gffg')
        set_border_edsi(ABA_VAM, 'K3:K3','fggf')
        set_border_edsi(ABA_VAM, 'K4:K4','fgfg')

        arquivo_excel.save(ARQ_INSUMOS)

    return(0)

def total_RCE(planilha,li,lf):
    fontMasterPretaG    = Font(color='00000000', bold=True, size=15)
    fontMasterPretaM    = Font(color='00000000', bold=True, size=13)
    fontMasterPretaP    = Font(color='00000000', bold=True, size=11)

    for latu in range(li,lf):
        soma = "=SUM(C"+str(latu)+":N"+str(latu)+")"
        planilha.cell(row=latu, column=15, value=soma).number_format = "#,##0.00"
        planilha.cell(row=latu, column=15).font = fontMasterPretaM

    c=3
    for catu in ('C','D','E','F','G','H','I','J','K','L','M','N','O'):
        soma = "=SUM("+catu+str(li)+":"+catu+str(lf-2)+")"
        planilha.cell(row=lf, column=c,  value=soma).number_format = "#,##0.00"
        planilha.cell(row=lf, column=c).font = fontMasterPretaM
        c=c+1

    planilha.cell(row=lf, column=2, value="TOTAL").font = fontMasterPretaM
    planilha.merge_cells('B'+str(lf-1)+':O'+str(lf-1))
    planilha.merge_cells('B'+str(lf+1)+':O'+str(lf+2))
    set_border_edsi(planilha, 'B'+str(li-3)+':O'+str(lf),'ffff')


########## FORMATACOES ###########
########## FORMATACOES ###########
########## FORMATACOES ###########
########## FORMATACOES ###########
########## FORMATACOES ###########
########## FORMATACOES ###########
def cabecalho_RCE(planilha,col,lin,titulo,ano):
    fontMasterPretaG    = Font(color='00000000', bold=True, size=15)
    fontMasterPretaM    = Font(color='00000000', bold=True, size=13)
    fontMasterPretaP    = Font(color='00000000', bold=True, size=11)
    fundoclaro          = 'B7DFE6'
    fundoescuro         = '92ABB0'

#    planilha.merge_cells('B'+str(lin)+':O'+str(lin))
    lin=lin+1
    planilha.cell(row=lin, column=col, value=titulo)
    planilha.cell(lin, col).font = fontMasterPretaM
    cor_fundo_p_ci_li_cf_lf_c(planilha,2,lin,15,lin, fundoclaro)
    planilha.merge_cells('B'+str(lin)+':O'+str(lin))

    lin=lin+1
    planilha.cell(row=lin, column=3,  value ="Valor Contábil" ).font = fontMasterPretaP
    planilha.cell(row=lin, column=4,  value ="Valor Contábil" ).font = fontMasterPretaP
    planilha.cell(row=lin, column=5,  value ="Valor Contábil" ).font = fontMasterPretaP
    planilha.cell(row=lin, column=6,  value ="Valor Contábil" ).font = fontMasterPretaP
    planilha.cell(row=lin, column=7,  value ="Valor Contábil" ).font = fontMasterPretaP
    planilha.cell(row=lin, column=8,  value ="Valor Contábil" ).font = fontMasterPretaP
    planilha.cell(row=lin, column=9,  value ="Valor Contábil" ).font = fontMasterPretaP
    planilha.cell(row=lin, column=10, value ="Valor Contábil" ).font = fontMasterPretaP
    planilha.cell(row=lin, column=11, value ="Valor Contábil" ).font = fontMasterPretaP
    planilha.cell(row=lin, column=12, value ="Valor Contábil" ).font = fontMasterPretaP
    planilha.cell(row=lin, column=13, value ="Valor Contábil" ).font = fontMasterPretaP
    planilha.cell(row=lin, column=14, value ="Valor Contábil" ).font = fontMasterPretaP
    planilha.cell(row=lin, column=15, value ="Total Anual" ).font = fontMasterPretaM
    cor_fundo_p_ci_li_cf_lf_c(planilha,2,lin,15,lin, fundoclaro)

    lin=lin +1
    planilha.cell(row=lin, column=2,  value ="CFOP" ).font = fontMasterPretaP
    planilha.cell(row=lin, column=3,  value ="jan/"+ano ).font = fontMasterPretaP
    planilha.cell(row=lin, column=4,  value ="fev/"+ano ).font = fontMasterPretaP
    planilha.cell(row=lin, column=5,  value ="mar/"+ano ).font = fontMasterPretaP
    planilha.cell(row=lin, column=6,  value ="abr/"+ano ).font = fontMasterPretaP
    planilha.cell(row=lin, column=7,  value ="mai/"+ano ).font = fontMasterPretaP
    planilha.cell(row=lin, column=8,  value ="jun/"+ano ).font = fontMasterPretaP
    planilha.cell(row=lin, column=9,  value ="jul/"+ano ).font = fontMasterPretaP
    planilha.cell(row=lin, column=10, value ="ago/"+ano ).font = fontMasterPretaP
    planilha.cell(row=lin, column=11, value ="set/"+ano ).font = fontMasterPretaP
    planilha.cell(row=lin, column=12, value ="out/"+ano ).font = fontMasterPretaP
    planilha.cell(row=lin, column=13, value ="nov/"+ano ).font = fontMasterPretaP
    planilha.cell(row=lin, column=14, value ="dez/"+ano ).font = fontMasterPretaP
    cor_fundo_p_ci_li_cf_lf_c(planilha,2,lin,2,lin, fundoclaro)
    cor_fundo_p_ci_li_cf_lf_c(planilha,3,lin,15,lin, fundoescuro)

    set_border_edsi(planilha, 'B2:O2' , 'fffn')

    for col in range(2,16):
        planilha.cell(lin-1, col).alignment = Alignment(horizontal='center')
        planilha.cell(lin,   col).alignment = Alignment(horizontal='center')

    set_border_edsi(planilha, 'B2:O2' , 'fffn')

    return(0)





def formata_DD(planilha):
    fontMasterPretaG    = Font(color='00000000', bold=True, size=15)
    fontMasterPretaM    = Font(color='00000000', bold=True, size=13)
    fontMasterPretaP    = Font(color='00000000', bold=True, size=11)
    # TAMANHO DAS COLUNAS:
    planilha.column_dimensions['A'].width = 100
    planilha.column_dimensions['B'].width = 20

    # LINHA 1 = TITULO
    planilha.cell(1,  1).font = fontMasterPretaG
    planilha.cell(1,  1).alignment = Alignment(horizontal='center')
    planilha.merge_cells('A1:B1')
    planilha.cell(1,  1).font = fontMasterPretaG
    planilha.merge_cells('A2:B2')
    planilha.cell(1,  1).font = fontMasterPretaG
    planilha.merge_cells('A3:B3')
    planilha.freeze_panes = 'A4'

    planilha.cell(2,  1).font = fontMasterPretaM
    planilha.cell(3,  1).font = fontMasterPretaM




    # Quadro 1
    planilha.cell(5,  1).font = fontMasterPretaM
    planilha.cell(5,  1).alignment = Alignment(horizontal='center')
    planilha.merge_cells('A5:B5')

    # Quadro 1.1
    planilha.cell(7,  1).font = fontMasterPretaP
    planilha.cell(7,  2).font = fontMasterPretaP
    planilha.cell(7,  2).alignment = Alignment(horizontal='center')

    planilha.cell(8,  2).number_format = "#,##0.00"
    planilha.cell(9,  2).number_format = "#,##0.00"
    planilha.cell(10, 2).number_format = "#,##0.00"

    planilha.cell(12,  1).font = fontMasterPretaP
    planilha.cell(12,  2).font = fontMasterPretaP
    planilha.cell(12, 2).number_format = "#,##0.00"

    cor_fundo_p_ci_li_cf_lf_c(planilha,1,1,2,3,'A3BBC0')
    cor_fundo_p_ci_li_cf_lf_c(planilha,1,5,2,5,'A3BBC0')
    cor_fundo_p_ci_li_cf_lf_c(planilha,1,7,2,7,'C3D4D7')



    # Quadro 1.2

    planilha.cell(14, 1).font = fontMasterPretaP
    planilha.cell(14, 2).font = fontMasterPretaP
    planilha.cell(14, 2).alignment = Alignment(horizontal='center')
    cor_fundo_p_ci_li_cf_lf_c(planilha,1,14,2,14,'C3D4D7')

    planilha.cell(15, 2).number_format = "#,##0.00"
    planilha.cell(16, 2).number_format = "#,##0.00"
    planilha.cell(17, 2).number_format = "#,##0.00"

    planilha.cell(19, 1).font = fontMasterPretaP
    planilha.cell(19, 2).font = fontMasterPretaP
    planilha.cell(19, 2).number_format = "#,##0.00"




    # Quadro 2
    planilha.cell(21,  1).font = fontMasterPretaM
    planilha.cell(21,  1).alignment = Alignment(horizontal='center')
    planilha.merge_cells('A21:B21')
    cor_fundo_p_ci_li_cf_lf_c(planilha,1,21,2,21,'A3BBC0')

    # Quadro 2.1
    planilha.cell(23, 1).font = fontMasterPretaP
    planilha.cell(23, 2).font = fontMasterPretaP
    planilha.cell(23, 2).alignment = Alignment(horizontal='center')
    cor_fundo_p_ci_li_cf_lf_c(planilha,1,23,2,23,'C3D4D7')

    planilha.cell(24, 2).number_format = "#,##0.00"
    planilha.cell(25, 2).number_format = "#,##0.00"
    planilha.cell(26, 2).number_format = "#,##0.00"

    planilha.cell(28, 1).font = fontMasterPretaP
    planilha.cell(28, 2).font = fontMasterPretaP
    planilha.cell(28, 2).number_format = "#,##0.00"


    # Quadro 2.2

    planilha.cell(30, 1).font = fontMasterPretaP
    planilha.cell(30, 2).font = fontMasterPretaP
    planilha.cell(30, 2).alignment = Alignment(horizontal='center')
    cor_fundo_p_ci_li_cf_lf_c(planilha,1,30,2,30,'C3D4D7')

    planilha.cell(31, 2).number_format = "#,##0.00"
    planilha.cell(32, 2).number_format = "#,##0.00"
    planilha.cell(33, 2).number_format = "#,##0.00"

    planilha.cell(35, 1).font = fontMasterPretaP
    planilha.cell(35, 2).font = fontMasterPretaP
    planilha.cell(35, 2).number_format = "#,##0.00"



    # Quadro 3
    planilha.cell(37,  1).font = fontMasterPretaM
    planilha.cell(37,  1).alignment = Alignment(horizontal='center')
    planilha.merge_cells('A37:B37')
    cor_fundo_p_ci_li_cf_lf_c(planilha,1,37,2,37,'A3BBC0')

    # Quadro 3.1
    planilha.cell(39, 1).font = fontMasterPretaP
    planilha.cell(39, 2).font = fontMasterPretaP
    planilha.cell(39, 2).alignment = Alignment(horizontal='center')
    cor_fundo_p_ci_li_cf_lf_c(planilha,1,39,2,39,'C3D4D7')

    planilha.cell(40, 2).number_format = "#,##0.00"
    planilha.cell(41, 2).number_format = "#,##0.00"
    planilha.cell(42, 2).number_format = "#,##0.00"
    planilha.cell(43, 2).number_format = "#,##0.00"

    planilha.cell(45, 1).font = fontMasterPretaP
    planilha.cell(45, 2).font = fontMasterPretaP
    planilha.cell(45, 2).number_format = "#,##0.00"

    # Quadro 3.2
    planilha.cell(47, 1).font = fontMasterPretaP
    planilha.cell(47, 2).font = fontMasterPretaP
    planilha.cell(47, 2).alignment = Alignment(horizontal='center')
    cor_fundo_p_ci_li_cf_lf_c(planilha,1,47,2,47,'C3D4D7')

    planilha.cell(48, 2).number_format = "#,##0.00"
    planilha.cell(49, 2).number_format = "#,##0.00"
    planilha.cell(50, 2).number_format = "#,##0.00"
    planilha.cell(51, 2).number_format = "#,##0.00"

    planilha.cell(53, 1).font = fontMasterPretaP
    planilha.cell(53, 2).font = fontMasterPretaP
    planilha.cell(53, 2).number_format = "#,##0.00"

    set_border_edsi(planilha, 'A1:B1' , 'fffn')
    set_border_edsi(planilha, 'A2:B2' , 'ffnn')
    set_border_edsi(planilha, 'A3:B3' , 'ffnf')
    set_border_edsi(planilha, 'A4:B53' , 'ffff')


    planilha.merge_cells('A4:B4')
    planilha.merge_cells('A6:B6')
    planilha.merge_cells('A11:B11')
    planilha.merge_cells('A13:B13')
    planilha.merge_cells('A18:B18')
    planilha.merge_cells('A20:B20')
    planilha.merge_cells('A22:B22')
    planilha.merge_cells('A27:B27')
    planilha.merge_cells('A29:B29')
    planilha.merge_cells('A34:B34')
    planilha.merge_cells('A36:B36')
    planilha.merge_cells('A38:B38')
    planilha.merge_cells('A44:B44')
    planilha.merge_cells('A46:B46')
    planilha.merge_cells('A52:B52')


def formata_VAM(planilha):
    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
    # TAMANHO DAS COLUNAS:
    adjust_column(planilha, 5,2,planilha.max_column)
    # LINHA 1 = TITULO
    planilha.cell(2,  2).font = fontMasterPreta
    planilha.cell(2,  2).alignment = Alignment(horizontal='center')
    planilha.merge_cells('A2:H2')
    planilha.freeze_panes = 'A6'
    # LINHA 2 = CABEÇALHO
    for c in range(2,planilha.max_column+1):
        planilha.cell(2,  c).font = fontMasterPreta
        planilha.cell(2,  c).alignment = Alignment(horizontal='center')              
    return

def formata_M(planilha):
    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
    # TAMANHO DAS COLUNAS:
    adjust_column(planilha, 1,1,planilha.max_column) 
    # LINHA 1 = TITULO
    planilha.cell(1,  1).font = fontMasterPreta
    planilha.cell(1,  1).alignment = Alignment(horizontal='center')
    planilha.merge_cells('A1:P1')
    planilha.freeze_panes = 'A3'
    # LINHA 2 = CABEÇALHO
    for c in range(1,planilha.max_column+1):
        planilha.cell(2,  c).font = fontMasterPreta
        planilha.cell(2,  c).alignment = Alignment(horizontal='center')              
    return

def Busca_SMI(vIE,vDataIni):
    global ret
    query = """SELECT /*+ PARALLEL(8) */ 
                    ''   								AS GAP,
                    'ALTERAR/MANTER/EXCLUIR/INCLUIR'    AS ACAO,
                    'item_nfsd_merc'            		AS TABELA,
                    I.ROWID                    			AS ROW_ID,
                    I.*
                FROM  OPENRISOW.ITEM_NFSD_MERC I,
                      OPENRISOW.FILIAL F
                WHERE  F.EMPS_COD = I.EMPS_COD
                  AND F.FILI_COD = I.FILI_COD
                  AND I.EMPS_COD = 'TBRA'
                  AND F.FILI_COD_INSEST = '%s'
                  AND I.INFSM_DTEMISS   >= TO_DATE('%s','DD/MM/YYYY')
                  AND I.INFSM_DTEMISS   <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1)
    """%(vIE,vDataIni,vDataIni)
    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchall()
    return(result)



def Busca_VAM(vIE,vDataIni,vDataFim):
    global ret
    query ="""with tmp_val as (
                select /*+ materilize,parallel(8) */ 
                       emps_cod,
                       fili_cod,
                       inva_data,
                       MIBGE_COD_MUN,
                       COD_UF,
                       UNFE_SIG,
                       contabil,
                       sum(contabil) over (partition by 'x') total_contabil,
                       contabil / sum(case when substr(MIBGE_COD_MUN,1,2) = COD_UF then contabil else 0 end ) over (partition by 'x') indice,
                       sum(case when substr(MIBGE_COD_MUN,1,2) <> COD_UF then contabil else 0 end ) over (partition by 'x') fora_estado,
                       (sum(case when substr(MIBGE_COD_MUN,1,2) <> COD_UF then contabil else 0 end ) over (partition by 'x') *(contabil / sum(case when substr(MIBGE_COD_MUN,1,2) = COD_UF then contabil else 0 end ) over (partition by 'x'))) ERRO,
                       contabil +(sum(case when substr(MIBGE_COD_MUN,1,2) <> COD_UF then contabil else 0 end ) over (partition by 'x') * (contabil / sum(case when substr(MIBGE_COD_MUN,1,2) = COD_UF then contabil else 0 end ) over (partition by 'x'))) valor_agregado
                from (
                SELECT emps_cod,
                       fili_cod,
                       inva_data,
                       MIBGE_COD_MUN,
                       COD_UF,
                       UNFE_SIG,
                       sum(INFST_VAL_CONT) contabil
                FROM (
                SELECT /*+ PARALLEL(8) */
                      NF.emps_cod,
                      NF.fili_cod,
                      last_day(NF.mnfst_dtemiss)inva_data,
                      F.UNFE_SIG,
                      (SELECT MIBGE_COD_MUN
                        FROM openrisow.cli_fornec_transp cli
                       WHERE cli.cadg_cod       = nf.cadg_cod
                         AND cli.catg_cod       = nf.catg_cod
                         AND cli.cadg_dat_atua = (select max(cli2.cadg_dat_atua) cadg_dat_atua
                                               FROM openrisow.cli_fornec_transp cli2
                                               WHERE cli2.cadg_cod       = nf.cadg_cod
                                                 AND cli2.catg_cod       = nf.catg_cod
                                                 AND cli2.cadg_dat_atua  <= nf.MNFST_DTEMISS)) MIBGE_COD_MUN
                                                 ,inf.cadg_cod
                                                 ,inf.catg_cod
                                                 ,INFST_VAL_SERV - INFST_VAL_DESC  INFST_VAL_CONT
                                                 ,nf.rowid row_id_mestre
                                                 ,SUBSTR(f.fili_mun_ibge, 1, 2) AS COD_UF
                FROM  OPENRISOW.MESTRE_NFTL_SERV NF
                 inner join openrisow.ITEM_NFTL_SERV inf on ( INF.EMPS_COD = NF.EMPS_COD
                                                          AND INF.FILI_COD = NF.FILI_COD
                                                          AND INF.INFST_SERIE = NF.MNFST_SERIE
                                                          AND INF.INFST_NUM = NF.MNFST_NUM
                                                          AND INF.INFST_DTEMISS = NF.mnfst_dtemiss
                                                          AND INF.MDOC_COD = NF.MDOC_COD),
                    openrisow.filial f 
                WHERE 1 = 1 
                  AND nf.emps_cod = f.emps_cod 
                  AND nf.fili_cod  = f.fili_cod
                  AND f.fili_cod_insest = '%s'--parametro
                  AND nf.MNFST_DTEMISS >= TO_DATE('%s','DD/MM/YYYY')--parametro
                  AND nf.MNFST_DTEMISS <   ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),12)--parametro
                  AND NF.EMPS_COD = 'TBRA'
                  and MNFST_IND_CANC = 'N'
                  and cfop in ('5301','5302','5303','5304','5305','5306','5307','6301', '6302', '6303', '6304', '6305', '6306', '6307')
                  )
                GROUP BY emps_cod,
                         fili_cod,
                         inva_data,
                         MIBGE_COD_MUN,
                         COD_UF,
                         UNFE_SIG
                )
                )
                SELECT IB.UNFE_SIG UF,
                       IB.MIBGE_DESC_MUN "descrção municí­pio",
                       IB.MIBGE_COD_MUN "municí­pio IBGE" ,
                       sum(contabil) "valor contábil",
                       sum(indice) "percentual calculado",
                       round(sum(ERRO),2) "diferença",
                       round(sum(valor_agregado),2) "valor agregado"
                FROM TMP_VAL AUX  INNER JOIN OPENRISOW.MIBGE IB ON (IB.MIBGE_COD_MUN = AUX.MIBGE_COD_MUN)
                where substr(aux.MIBGE_COD_MUN,1,2) = aux.COD_UF
                group by TO_CHAR(inva_data,'yyyy'),
                         IB.UNFE_SIG,
                         IB.MIBGE_DESC_MUN ,
                         IB.MIBGE_COD_MUN 
                order by UF,MIBGE_DESC_MUN
    """%(vIE,vDataIni,vDataIni)

    #log(query)
    
    cursor =  sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchall()
    
    return(result)

def Busca_RCE11(vIE,vDataIni,vDataFim):
    global ret
    query ="""select * from (
                SELECT 
                    CFOP_COD,
                    TO_CHAR(resdata,'MM') MES,/*
                    DECODE(TO_CHAR(resdata,'MM'),'01','jan/'
                                                ,'02','fev/'
                                                ,'03','mar/'
                                                ,'04','abr/'
                                                ,'05','mai/'
                                                ,'06','jun/'
                                                ,'07','jul/'
                                                ,'08','ago/'
                                                ,'09','set/'
                                                ,'10','out/'
                                                ,'11','nov/'
                                                ,'12','dez/'
                                                ,'invalido')||TO_CHAR(resdata,'YYYY') MES_ANO,*/
                    SUM(nvl(res.val_cont, 0)) AS VLR_CONTABIL
                FROM
                    openrisow.resumo_fiscal res
                WHERE
                        1 = 1
                    AND (   (ind_es = 'S' AND origem = 'T')
                        OR  (ind_es = 'S' AND origem = 'M')
                        OR  (ind_es = 'E' AND origem = 'T')
                        OR  (ind_es = 'E' AND origem = 'M')
                        )
                    AND resdata         >= TO_DATE('%s', 'DD/MM/YYYY')
                    AND resdata         <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1)
                    AND FILI_COD_INSEST = '%s'
                    AND SUBSTR(CFOP_COD,1,1) IN ('1')
                    AND CFOP_COD        <> '0000'
                GROUP BY
                    --res.cfop_cod,
                    CFOP_COD,
                    TO_CHAR(resdata,'MM'),
                    DECODE(TO_CHAR(resdata,'MM'),'01','jan/'
                                                ,'02','fev/'
                                                ,'03','mar/'
                                                ,'04','abr/'
                                                ,'05','mai/'
                                                ,'06','jun/'
                                                ,'07','jul/'
                                                ,'08','ago/'
                                                ,'09','set/'
                                                ,'10','out/'
                                                ,'11','nov/'
                                                ,'12','dez/'
                                                ,'invalido')||TO_CHAR(resdata,'YYYY')
                )
                pivot 
                (
                   SUM(VLR_CONTABIL)
                   for MES in ('01','02','03','04','05','06','07','08','09','10','11','12')
                )
                order by CFOP_COD
    """%(vDataIni,vDataFim,vIE)

    cursor =sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchall()
    return(result)

def Busca_RCF11(vIE,vDataIni,vDataFim):
    global ret
    query ="""select * from (
                SELECT 
                    CFOP_COD,
                    TO_CHAR(resdata,'MM') MES,/*
                    DECODE(TO_CHAR(resdata,'MM'),'01','jan/'
                                                ,'02','fev/'
                                                ,'03','mar/'
                                                ,'04','abr/'
                                                ,'05','mai/'
                                                ,'06','jun/'
                                                ,'07','jul/'
                                                ,'08','ago/'
                                                ,'09','set/'
                                                ,'10','out/'
                                                ,'11','nov/'
                                                ,'12','dez/'
                                                ,'invalido')||TO_CHAR(resdata,'YYYY') MES_ANO,*/
                    SUM(nvl(res.val_cont, 0)) AS VLR_CONTABIL
                FROM
                    openrisow.resumo_fiscal res
                WHERE
                        1 = 1
                    AND (   (ind_es = 'S' AND origem = 'T')
                        OR  (ind_es = 'S' AND origem = 'M')
                        OR  (ind_es = 'E' AND origem = 'T')
                        OR  (ind_es = 'E' AND origem = 'M')
                        )
                    AND resdata         >= TO_DATE('%s', 'DD/MM/YYYY')               
                    AND resdata         <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1)
                    AND FILI_COD_INSEST = '%s'
                    AND SUBSTR(CFOP_COD,1,1) IN ('5')
                    AND CFOP_COD        <> '0000'
                GROUP BY
                    --res.cfop_cod,
                    CFOP_COD,
                    TO_CHAR(resdata,'MM'),
                    DECODE(TO_CHAR(resdata,'MM'),'01','jan/'
                                                ,'02','fev/'
                                                ,'03','mar/'
                                                ,'04','abr/'
                                                ,'05','mai/'
                                                ,'06','jun/'
                                                ,'07','jul/'
                                                ,'08','ago/'
                                                ,'09','set/'
                                                ,'10','out/'
                                                ,'11','nov/'
                                                ,'12','dez/'
                                                ,'invalido')||TO_CHAR(resdata,'YYYY')
                )
                pivot 
                (
                   SUM(VLR_CONTABIL)
                   for MES in ('01','02','03','04','05','06','07','08','09','10','11','12')
                )
                order by CFOP_COD
    """%(vDataIni,vDataFim,vIE)

    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchall()
    return(result)

def Busca_RCE12(vIE,vDataIni,vDataFim):
    global ret
    query ="""select * from (
                SELECT 
                    CFOP_COD,
                    TO_CHAR(resdata,'MM') MES,/*
                    DECODE(TO_CHAR(resdata,'MM'),'01','jan/'
                                                ,'02','fev/'
                                                ,'03','mar/'
                                                ,'04','abr/'
                                                ,'05','mai/'
                                                ,'06','jun/'
                                                ,'07','jul/'
                                                ,'08','ago/'
                                                ,'09','set/'
                                                ,'10','out/'
                                                ,'11','nov/'
                                                ,'12','dez/'
                                                ,'invalido')||TO_CHAR(resdata,'YYYY') MES_ANO,*/
                    SUM(nvl(res.val_cont, 0)) AS VLR_CONTABIL
                FROM
                    openrisow.resumo_fiscal res
                WHERE
                        1 = 1
                    AND (   (ind_es = 'S' AND origem = 'T')
                        OR  (ind_es = 'S' AND origem = 'M')
                        OR  (ind_es = 'E' AND origem = 'T')
                        OR  (ind_es = 'E' AND origem = 'M')
                        )
                    AND resdata         >= TO_DATE('%s', 'DD/MM/YYYY')               --<<PARÂMETRO: DATA_INI>>
                    AND resdata         <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1) --<<PARÂMETRO: DATA_FIM>>
                    AND FILI_COD_INSEST = '%s'
                    AND SUBSTR(CFOP_COD,1,1) IN ('2')--<<PARÂMETRO: INSCRIÇÃO ESTADUAL>>
                    AND CFOP_COD        <> '0000'
                GROUP BY
                    --res.cfop_cod,
                    CFOP_COD,
                    TO_CHAR(resdata,'MM'),
                    DECODE(TO_CHAR(resdata,'MM'),'01','jan/'
                                                ,'02','fev/'
                                                ,'03','mar/'
                                                ,'04','abr/'
                                                ,'05','mai/'
                                                ,'06','jun/'
                                                ,'07','jul/'
                                                ,'08','ago/'
                                                ,'09','set/'
                                                ,'10','out/'
                                                ,'11','nov/'
                                                ,'12','dez/'
                                                ,'invalido')||TO_CHAR(resdata,'YYYY')
                )
                pivot 
                (
                   SUM(VLR_CONTABIL)
                   for MES in ('01','02','03','04','05','06','07','08','09','10','11','12')
                )
                order by CFOP_COD
    """%(vDataIni,vDataFim,vIE)

    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchall()
    return(result)

def Busca_RCF12(vIE,vDataIni,vDataFim):
    global ret
    query ="""select * from (
                SELECT 
                    CFOP_COD,
                    TO_CHAR(resdata,'MM') MES,/*
                    DECODE(TO_CHAR(resdata,'MM'),'01','jan/'
                                                ,'02','fev/'
                                                ,'03','mar/'
                                                ,'04','abr/'
                                                ,'05','mai/'
                                                ,'06','jun/'
                                                ,'07','jul/'
                                                ,'08','ago/'
                                                ,'09','set/'
                                                ,'10','out/'
                                                ,'11','nov/'
                                                ,'12','dez/'
                                                ,'invalido')||TO_CHAR(resdata,'YYYY') MES_ANO,*/
                    SUM(nvl(res.val_cont, 0)) AS VLR_CONTABIL
                FROM
                    openrisow.resumo_fiscal res
                WHERE
                        1 = 1
                    AND (   (ind_es = 'S' AND origem = 'T')
                        OR  (ind_es = 'S' AND origem = 'M')
                        OR  (ind_es = 'E' AND origem = 'T')
                        OR  (ind_es = 'E' AND origem = 'M')
                        )
                    AND resdata         >= TO_DATE('%s', 'DD/MM/YYYY') 
                    AND resdata         <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1) 
                    AND FILI_COD_INSEST = '%s'
                    AND SUBSTR(CFOP_COD,1,1) IN ('6')
                    AND CFOP_COD        <> '0000'
                GROUP BY
                    --res.cfop_cod,
                    CFOP_COD,
                    TO_CHAR(resdata,'MM'),
                    DECODE(TO_CHAR(resdata,'MM'),'01','jan/'
                                                ,'02','fev/'
                                                ,'03','mar/'
                                                ,'04','abr/'
                                                ,'05','mai/'
                                                ,'06','jun/'
                                                ,'07','jul/'
                                                ,'08','ago/'
                                                ,'09','set/'
                                                ,'10','out/'
                                                ,'11','nov/'
                                                ,'12','dez/'
                                                ,'invalido')||TO_CHAR(resdata,'YYYY')
                )
                pivot 
                (
                   SUM(VLR_CONTABIL)
                   for MES in ('01','02','03','04','05','06','07','08','09','10','11','12')
                )
                order by CFOP_COD
    """%(vDataIni,vDataFim,vIE)

    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchall()
    return(result)


def Busca_RCE13(vIE,vDataIni,vDataFim):
    global ret
    query ="""select * from (
                SELECT 
                    CFOP_COD,
                    TO_CHAR(resdata,'MM') MES,/*
                    DECODE(TO_CHAR(resdata,'MM'),'01','jan/'
                                                ,'02','fev/'
                                                ,'03','mar/'
                                                ,'04','abr/'
                                                ,'05','mai/'
                                                ,'06','jun/'
                                                ,'07','jul/'
                                                ,'08','ago/'
                                                ,'09','set/'
                                                ,'10','out/'
                                                ,'11','nov/'
                                                ,'12','dez/'
                                                ,'invalido')||TO_CHAR(resdata,'YYYY') MES_ANO,*/
                    SUM(nvl(res.val_cont, 0)) AS VLR_CONTABIL
                FROM
                    openrisow.resumo_fiscal res
                WHERE
                        1 = 1
                    AND (   (ind_es = 'S' AND origem = 'T')
                        OR  (ind_es = 'S' AND origem = 'M')
                        OR  (ind_es = 'E' AND origem = 'T')
                        OR  (ind_es = 'E' AND origem = 'M')
                        )
                    AND resdata         >= TO_DATE('%s', 'DD/MM/YYYY')              
                    AND resdata         <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1) 
                    AND FILI_COD_INSEST = '%s'
                    AND SUBSTR(CFOP_COD,1,1) IN ('3')
                    AND CFOP_COD        <> '0000'
                GROUP BY
                    --res.cfop_cod,
                    CFOP_COD,
                    TO_CHAR(resdata,'MM'),
                    DECODE(TO_CHAR(resdata,'MM'),'01','jan/'
                                                ,'02','fev/'
                                                ,'03','mar/'
                                                ,'04','abr/'
                                                ,'05','mai/'
                                                ,'06','jun/'
                                                ,'07','jul/'
                                                ,'08','ago/'
                                                ,'09','set/'
                                                ,'10','out/'
                                                ,'11','nov/'
                                                ,'12','dez/'
                                                ,'invalido')||TO_CHAR(resdata,'YYYY')
                )
                pivot 
                (
                   SUM(VLR_CONTABIL)
                   for MES in ('01','02','03','04','05','06','07','08','09','10','11','12')
                )
                order by CFOP_COD
    """%(vDataIni,vDataFim,vIE)

    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchall()
    return(result)


def Busca_RCF13(vIE,vDataIni,vDataFim):
    global ret
    query ="""select * from (
                SELECT 
                    CFOP_COD,
                    TO_CHAR(resdata,'MM') MES,/*
                    DECODE(TO_CHAR(resdata,'MM'),'01','jan/'
                                                ,'02','fev/'
                                                ,'03','mar/'
                                                ,'04','abr/'
                                                ,'05','mai/'
                                                ,'06','jun/'
                                                ,'07','jul/'
                                                ,'08','ago/'
                                                ,'09','set/'
                                                ,'10','out/'
                                                ,'11','nov/'
                                                ,'12','dez/'
                                                ,'invalido')||TO_CHAR(resdata,'YYYY') MES_ANO,*/
                    SUM(nvl(res.val_cont, 0)) AS VLR_CONTABIL
                FROM
                    openrisow.resumo_fiscal res
                WHERE
                        1 = 1
                    AND (   (ind_es = 'S' AND origem = 'T')
                        OR  (ind_es = 'S' AND origem = 'M')
                        OR  (ind_es = 'E' AND origem = 'T')
                        OR  (ind_es = 'E' AND origem = 'M')
                        )
                    AND resdata         >= TO_DATE('%s', 'DD/MM/YYYY')              
                    AND resdata         <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1) 
                    AND FILI_COD_INSEST = '%s'
                    AND SUBSTR(CFOP_COD,1,1) IN ('7')
                    AND CFOP_COD        <> '0000'
                GROUP BY
                    --res.cfop_cod,
                    CFOP_COD,
                    TO_CHAR(resdata,'MM'),
                    DECODE(TO_CHAR(resdata,'MM'),'01','jan/'
                                                ,'02','fev/'
                                                ,'03','mar/'
                                                ,'04','abr/'
                                                ,'05','mai/'
                                                ,'06','jun/'
                                                ,'07','jul/'
                                                ,'08','ago/'
                                                ,'09','set/'
                                                ,'10','out/'
                                                ,'11','nov/'
                                                ,'12','dez/'
                                                ,'invalido')||TO_CHAR(resdata,'YYYY')
                )
                pivot 
                (
                   SUM(VLR_CONTABIL)
                   for MES in ('01','02','03','04','05','06','07','08','09','10','11','12')
                )
                order by CFOP_COD
    """%(vDataIni,vDataFim,vIE)

    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchall()
    return(result)


def Busca_DD11(vIE,vDataIni,vDataFim):
    global ret
    query ="""SELECT 
                CASE WHEN SUBSTR(CFOP_COD,1,1) IN ('1','5') THEN '1.Estado'
                     WHEN SUBSTR(CFOP_COD,1,1) IN ('2','6') THEN '2.Outro Estado'
                     WHEN SUBSTR(CFOP_COD,1,1) IN ('3','7') THEN '3.Exterior'
                     ELSE 'Invalido'
                END Classificacao,
                --res.cfop_cod 			                AS CFOP,
                SUM(nvl(res.val_cont, 0))               AS VLR_CONTABIL
            FROM
                openrisow.resumo_fiscal res
            WHERE
                    1 = 1
                AND (   (ind_es = 'S' AND origem = 'T')
                    OR  (ind_es = 'S' AND origem = 'M')
                    OR  (ind_es = 'E' AND origem = 'T')
                    OR  (ind_es = 'E' AND origem = 'M')
                    )
                AND resdata         >= TO_DATE('%s', 'DD/MM/YYYY')               --<<PARÂMETRO: DATA_INI>>
                AND resdata         <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1)  --<<PARÂMETRO: DATA_FIM>>
                AND FILI_COD_INSEST = '%s'                                       --<<PARÂMETRO: INSCRIÇÃO ESTADUAL>>
                AND SUBSTR(CFOP_COD,1,1) IN ('1','2','3') 
                AND CFOP_COD        <> '0000'
            GROUP BY
                --res.cfop_cod,
                CASE WHEN SUBSTR(CFOP_COD,1,1) IN ('1','5') THEN '1.Estado'
                     WHEN SUBSTR(CFOP_COD,1,1) IN ('2','6') THEN '2.Outro Estado'
                     WHEN SUBSTR(CFOP_COD,1,1) IN ('3','7') THEN '3.Exterior'
                     ELSE 'Invalido'
                END
            ORDER BY
                Classificacao ASC
    """%(vDataIni,vDataFim,vIE)

    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchall()
    return(result)

def Busca_DD12(vIE,vDataIni,vDataFim):
    global ret
    query ="""SELECT 
                    CASE WHEN SUBSTR(CFOP_COD,1,1) IN ('1','5') THEN '1.Estado'
                         WHEN SUBSTR(CFOP_COD,1,1) IN ('2','6') THEN '2.Outro Estado'
                         WHEN SUBSTR(CFOP_COD,1,1) IN ('3','7') THEN '3.Exterior'
                         ELSE 'Invalido'
                    END Classificacao,
                    --res.cfop_cod 			            AS CFOP,
                    SUM(nvl(res.val_cont, 0))               AS VLR_CONTABIL
                FROM
                    openrisow.resumo_fiscal res
                WHERE
                        1 = 1
                    AND (   (ind_es = 'S' AND origem = 'T')
                        OR  (ind_es = 'S' AND origem = 'M')
                        OR  (ind_es = 'E' AND origem = 'T')
                        OR  (ind_es = 'E' AND origem = 'M')
                        )
                    AND resdata         >= TO_DATE('%s', 'DD/MM/YYYY')               --<<PARÂMETRO: DATA_INI>>
                    AND resdata         <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1) --<<PARÂMETRO: DATA_FIM>>
                    AND FILI_COD_INSEST = '%s'                                        --<<PARÂMETRO: INSCRIÇÃO ESTADUAL>>
                    AND SUBSTR(CFOP_COD,1,1) IN ('5','6','7') 
                    AND CFOP_COD        <> '0000'
                GROUP BY
                    --res.cfop_cod,
                    CASE WHEN SUBSTR(CFOP_COD,1,1) IN ('1','5') THEN '1.Estado'
                         WHEN SUBSTR(CFOP_COD,1,1) IN ('2','6') THEN '2.Outro Estado'
                         WHEN SUBSTR(CFOP_COD,1,1) IN ('3','7') THEN '3.Exterior'
                         ELSE 'Invalido'
                    END
                ORDER BY
                    Classificacao ASC
    """%(vDataIni,vDataFim,vIE)

    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchall()
    return(result)

def Busca_DD21(vIE,vDataIni,vDataFim):
    global ret
    query ="""SELECT 
                    CASE WHEN SUBSTR(CFOP_COD,1,1) IN ('1','5') THEN '1.Estado'
                         WHEN SUBSTR(CFOP_COD,1,1) IN ('2','6') THEN '2.Outro Estado'
                         WHEN SUBSTR(CFOP_COD,1,1) IN ('3','7') THEN '3.Exterior'
                         ELSE 'Invalido'
                    END Classificacao,
                    --res.cfop_cod 			            AS CFOP,
                    SUM(nvl(res.val_cont, 0))               AS VLR_CONTABIL
                FROM
                    openrisow.resumo_fiscal res
                WHERE
                        1 = 1
                    AND (   (ind_es = 'S' AND origem = 'T')
                        OR  (ind_es = 'S' AND origem = 'M')
                        OR  (ind_es = 'E' AND origem = 'T')
                        OR  (ind_es = 'E' AND origem = 'M')
                        )
                    AND resdata         >= TO_DATE('%s', 'DD/MM/YYYY')               --<<PARÂMETRO: DATA_INI>>
                    AND resdata         <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1) --<<PARÂMETRO: DATA_FIM>>
                    AND FILI_COD_INSEST = '%s'                                        --<<PARÂMETRO: INSCRIÇÃO ESTADUAL>>
                    AND SUBSTR(CFOP_COD,1,1) IN ('1','2','3')
                    AND SUBSTR(CFOP_COD,1,3) NOT IN ('130','230','330') 
                    AND CFOP_COD        <> '0000'
                GROUP BY
                    --res.cfop_cod,
                    CASE WHEN SUBSTR(CFOP_COD,1,1) IN ('1','5') THEN '1.Estado'
                         WHEN SUBSTR(CFOP_COD,1,1) IN ('2','6') THEN '2.Outro Estado'
                         WHEN SUBSTR(CFOP_COD,1,1) IN ('3','7') THEN '3.Exterior'
                         ELSE 'Invalido'
                    END
                ORDER BY
                    Classificacao ASC
    """%(vDataIni,vDataFim,vIE)

    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchall()
    return(result)


def Busca_DD22(vIE,vDataIni,vDataFim):
    global ret
    query ="""SELECT 
                    CASE WHEN SUBSTR(CFOP_COD,1,1) IN ('1','5') THEN '1.Estado'
                         WHEN SUBSTR(CFOP_COD,1,1) IN ('2','6') THEN '2.Outro Estado'
                         WHEN SUBSTR(CFOP_COD,1,1) IN ('3','7') THEN '3.Exterior'
                         ELSE 'Invalido'
                    END Classificacao,
                    --res.cfop_cod 			            AS CFOP,
                    SUM(nvl(res.val_cont, 0))               AS VLR_CONTABIL
                FROM
                    openrisow.resumo_fiscal res
                WHERE
                        1 = 1
                    AND (   (ind_es = 'S' AND origem = 'T')
                        OR  (ind_es = 'S' AND origem = 'M')
                        OR  (ind_es = 'E' AND origem = 'T')
                        OR  (ind_es = 'E' AND origem = 'M')
                        )
                    AND resdata         >= TO_DATE('%s', 'DD/MM/YYYY')               --<<PARÂMETRO: DATA_INI>>
                    AND resdata         <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1) --<<PARÂMETRO: DATA_FIM>>
                    AND FILI_COD_INSEST = '%s'                                        --<<PARÂMETRO: INSCRIÇÃO ESTADUAL>>
                    AND SUBSTR(CFOP_COD,1,1) IN ('5','6','7')
                    AND SUBSTR(CFOP_COD,1,3) NOT IN ('530','630','730') 
                    AND CFOP_COD        <> '0000'
                GROUP BY
                    --res.cfop_cod,
                    CASE WHEN SUBSTR(CFOP_COD,1,1) IN ('1','5') THEN '1.Estado'
                         WHEN SUBSTR(CFOP_COD,1,1) IN ('2','6') THEN '2.Outro Estado'
                         WHEN SUBSTR(CFOP_COD,1,1) IN ('3','7') THEN '3.Exterior'
                         ELSE 'Invalido'
                    END
                ORDER BY
                    Classificacao ASC                
    """%(vDataIni,vDataFim,vIE)

    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchall()
    return(result)



def Busca_DD31(vIE,vDataIni,vDataFim):
    global ret
    query ="""SELECT 
                    CASE WHEN CFOP_COD in ('1406', '1551', '1552', '1553', '1554', '1555', '2406',
                                           '2551', '2552', '2553', '2554', '2555', '3551', '3553') THEN '1. Operações Relativas ao Ativo Imobilizado'
                         WHEN CFOP_COD in ('1407', '2407', '1556', '2556', '3556', '1557', '2557') THEN '2. Operações Relativas ao Uso e Consumo'
                         WHEN CFOP_COD in ('1949', '1128', '1154', '1414', '1415', '1657', '1904',
                                           '1131', '1213', '1454', '1505', '1506', '1601', '1602',
                                           '1604', '1605', '1663', '1664', '1901', '1902', '1903',
                                           '1905', '1906', '1907', '1908', '1909', '1912', '1913',
                                           '1914', '1915', '1916', '1917', '1918', '1919', '1920',
                                           '1921', '1922', '1923', '1924', '1925', '1926', '1933',
                                           '1934', '2949', '2128', '2154', '2414', '2415', '2904',
                                           '2131', '2213', '2454', '2505', '2506', '2663', '2664',
                                           '2901', '2902', '2903', '2905', '2906', '2907', '2908',
                                           '2909', '2912', '2913', '2914', '2915', '2916', '2917', 
                                           '2918', '2919', '2920', '2921', '2922', '2923', '2924', 
                                           '2925', '2933', '2934', '3949', '3128', '3930')          THEN '3. Operações / Prestações que não são Fato Gerador do ICMS ou nao Utilizadas no VA'
                         WHEN CFOP_COD in ('1603', '2603') THEN '4. ICMS Retido por Substituição Tributária'
                         ELSE 'Invalido'
                    END Classificacao,
                    --res.cfop_cod 			            AS CFOP,
                    SUM(nvl(res.val_cont, 0))               AS VLR_CONTABIL
                FROM
                    openrisow.resumo_fiscal res
                WHERE
                        1 = 1
                    AND (   (ind_es = 'S' AND origem = 'T')
                        OR  (ind_es = 'S' AND origem = 'M')
                        OR  (ind_es = 'E' AND origem = 'T')
                        OR  (ind_es = 'E' AND origem = 'M')
                        )
                    AND resdata         >= TO_DATE('%s', 'DD/MM/YYYY')               --<<PARÂMETRO: DATA_INI>>
                    AND resdata         <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1) --<<PARÂMETRO: DATA_FIM>>
                    AND FILI_COD_INSEST = '%s'                                        --<<PARÂMETRO: INSCRIÇÃO ESTADUAL>>
                    AND CFOP_COD IN ('1406', '1551', '1552', '1553', '1554', '1555', '2406',
                                     '2551', '2552', '2553', '2554', '2555', '3551', '3553',
                                     '1407', '2407', '1556', '2556', '3556', '1557', '2557',
                                     '1949', '1128', '1154', '1414', '1415', '1657', '1904',
                                     '1131', '1213', '1454', '1505', '1506', '1601', '1602',
                                     '1604', '1605', '1663', '1664', '1901', '1902', '1903',
                                     '1905', '1906', '1907', '1908', '1909', '1912', '1913',
                                     '1914', '1915', '1916', '1917', '1918', '1919', '1920',
                                     '1921', '1922', '1923', '1924', '1925', '1926', '1933',
                                     '1934', '2949', '2128', '2154', '2414', '2415', '2904',
                                     '2131', '2213', '2454', '2505', '2506', '2663', '2664',
                                     '2901', '2902', '2903', '2905', '2906', '2907', '2908',
                                     '2909', '2912', '2913', '2914', '2915', '2916', '2917',
                                     '2918', '2919', '2920', '2921', '2922', '2923', '2924',
                                     '2925', '2933', '2934', '3949', '3128', '3930', '1603',
                                     '2603') 
                    AND CFOP_COD        <> '0000'
                GROUP BY
                    --res.cfop_cod,
                    CASE WHEN CFOP_COD in ('1406', '1551', '1552', '1553', '1554', '1555', '2406',
                                           '2551', '2552', '2553', '2554', '2555', '3551', '3553') THEN '1. Operações Relativas ao Ativo Imobilizado'
                         WHEN CFOP_COD in ('1407', '2407', '1556', '2556', '3556', '1557', '2557') THEN '2. Operações Relativas ao Uso e Consumo'
                         WHEN CFOP_COD in ('1949', '1128', '1154', '1414', '1415', '1657', '1904',
                                           '1131', '1213', '1454', '1505', '1506', '1601', '1602',
                                           '1604', '1605', '1663', '1664', '1901', '1902', '1903',
                                           '1905', '1906', '1907', '1908', '1909', '1912', '1913',
                                           '1914', '1915', '1916', '1917', '1918', '1919', '1920',
                                           '1921', '1922', '1923', '1924', '1925', '1926', '1933',
                                           '1934', '2949', '2128', '2154', '2414', '2415', '2904',
                                           '2131', '2213', '2454', '2505', '2506', '2663', '2664',
                                           '2901', '2902', '2903', '2905', '2906', '2907', '2908',
                                           '2909', '2912', '2913', '2914', '2915', '2916', '2917', 
                                           '2918', '2919', '2920', '2921', '2922', '2923', '2924', 
                                           '2925', '2933', '2934', '3949', '3128', '3930')          THEN '3. Operações / Prestações que não são Fato Gerador do ICMS ou nao Utilizadas no VA'
                         WHEN CFOP_COD in ('1603', '2603') THEN '4. ICMS Retido por Substituição Tributária'
                         ELSE 'Invalido'
                    END
                ORDER BY
                    Classificacao ASC
    """%(vDataIni,vDataFim,vIE)

    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchall()
    return(result)


def Busca_DD32(vIE,vDataIni,vDataFim):
    global ret
    query ="""SELECT 
                    CASE WHEN CFOP_COD in ('5412', '5551', '5552', '5553', '5554', '5555', '6555',
                                           '6412', '6551', '6552', '6553', '6554', '7551', '7553') THEN '1. Operações Relativas ao Ativo Imobilizado'
                         WHEN CFOP_COD in ('5413', '5556', '5557', '6413', '6556', '6557', '7556') THEN '2. Operações Relativas ao Uso e Consumo'
                         WHEN CFOP_COD in ('5949', '5210', '5414', '5415', '5657', '5904', '5131',
                                           '5213', '5454', '5504', '5505', '5601', '5602', '5605',
                                           '5606', '5663', '5664', '5665', '5666', '5901', '5902',
                                           '5903', '5905', '5906', '5907', '5908', '5909', '5912',
                                           '5913', '5914', '5915', '5916', '5917', '5918', '5919',
                                           '5920', '5921', '5922', '5923', '5924', '5925', '5926',
                                           '5929', '5933', '5934', '6949', '6210', '6414', '6415',
                                           '6657', '6904', '6131', '6213', '6454', '6504', '6505',
                                           '6663', '6664', '6665', '6666', '6901', '6902', '6903',
                                           '6905', '6906', '6907', '6908', '6909', '6912', '6913',
                                           '6914', '6915', '6916', '6917', '6918', '6919', '6920',
                                           '6921', '6922', '6923', '6924', '6925', '6929', '6933',
                                           '6934', '7949', '7210', '7930')                         THEN '3. Operações / Prestações que não são Fato Gerador do ICMS ou nao Utilizadas no VA'
                         WHEN CFOP_COD in ('5603', '6603') THEN '4. ICMS Retido por Substituição Tributária'
                         ELSE 'Invalido'
                    END Classificacao,
                    --res.cfop_cod 			            AS CFOP,
                    SUM(nvl(res.val_cont, 0))               AS VLR_CONTABIL
                FROM
                    openrisow.resumo_fiscal res
                WHERE
                        1 = 1
                    AND (   (ind_es = 'S' AND origem = 'T')
                        OR  (ind_es = 'S' AND origem = 'M')
                        OR  (ind_es = 'E' AND origem = 'T')
                        OR  (ind_es = 'E' AND origem = 'M')
                        )
                    AND resdata         >= TO_DATE('%s', 'DD/MM/YYYY')               --<<PARÂMETRO: DATA_INI>>
                    AND resdata         <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1) --<<PARÂMETRO: DATA_FIM>>
                    AND FILI_COD_INSEST = '%s'                                        --<<PARÂMETRO: INSCRIÇÃO ESTADUAL>>
                    AND CFOP_COD IN ('5412', '5551', '5552', '5553', '5554', '5555', '6555',	
                                     '6412', '6551', '6552', '6553', '6554', '7551', '7553',
                                     '5413', '5556', '5557', '6413', '6556', '6557', '7556',
                                     '5949', '5210', '5414', '5415', '5657', '5904', '5131',
                                     '5213', '5454', '5504', '5505', '5601', '5602', '5605',
                                     '5606', '5663', '5664', '5665', '5666', '5901', '5902',
                                     '5903', '5905', '5906', '5907', '5908', '5909', '5912',
                                     '5913', '5914', '5915', '5916', '5917', '5918', '5919',
                                     '5920', '5921', '5922', '5923', '5924', '5925', '5926',
                                     '5929', '5933', '5934', '6949', '6210', '6414', '6415',
                                     '6657', '6904', '6131', '6213', '6454', '6504', '6505',
                                     '6663', '6664', '6665', '6666', '6901', '6902', '6903',
                                     '6905', '6906', '6907', '6908', '6909', '6912', '6913',
                                     '6914', '6915', '6916', '6917', '6918', '6919', '6920',
                                     '6921', '6922', '6923', '6924', '6925', '6929', '6933',
                                     '6934', '7949', '7210', '7930', '5603', '6603') 
                    AND CFOP_COD        <> '0000'
                GROUP BY
                    --res.cfop_cod,
                    CASE WHEN CFOP_COD in ('5412', '5551', '5552', '5553', '5554', '5555', '6555',
                                           '6412', '6551', '6552', '6553', '6554', '7551', '7553') THEN '1. Operações Relativas ao Ativo Imobilizado'
                         WHEN CFOP_COD in ('5413', '5556', '5557', '6413', '6556', '6557', '7556') THEN '2. Operações Relativas ao Uso e Consumo'
                         WHEN CFOP_COD in ('5949', '5210', '5414', '5415', '5657', '5904', '5131',
                                           '5213', '5454', '5504', '5505', '5601', '5602', '5605',
                                           '5606', '5663', '5664', '5665', '5666', '5901', '5902',
                                           '5903', '5905', '5906', '5907', '5908', '5909', '5912',
                                           '5913', '5914', '5915', '5916', '5917', '5918', '5919',
                                           '5920', '5921', '5922', '5923', '5924', '5925', '5926',
                                           '5929', '5933', '5934', '6949', '6210', '6414', '6415',
                                           '6657', '6904', '6131', '6213', '6454', '6504', '6505',
                                           '6663', '6664', '6665', '6666', '6901', '6902', '6903',
                                           '6905', '6906', '6907', '6908', '6909', '6912', '6913',
                                           '6914', '6915', '6916', '6917', '6918', '6919', '6920',
                                           '6921', '6922', '6923', '6924', '6925', '6929', '6933',
                                           '6934', '7949', '7210', '7930')                         THEN '3. Operações / Prestações que não são Fato Gerador do ICMS ou nao Utilizadas no VA'
                         WHEN CFOP_COD in ('5603', '6603') THEN '4. ICMS Retido por Substituição Tributária'
                         ELSE 'Invalido'
                    END
                ORDER BY
                    Classificacao ASC
    """%(vDataIni,vDataFim,vIE)

    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchall()
    return(result)


def set_border( ws, cell_range):
        border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                cell.border = border


def set_border_edsi( ws, cell_range, edsi):
    b=[]
 
    n=Side(border_style= None    , color='00000000')
    f=Side(border_style='thin'   , color='00000000')
    m=Side(border_style='medium' , color='00000000')
    d=Side(border_style='double' , color='00000000')
    p=Side(border_style='dotted' , color='00000000')
    t=Side(border_style='dashed' , color='00000000')
    g=Side(border_style='thick'  , color='00000000')
    
    for x in (0,1,2,3):
        if (edsi[x]=="n"):
            b.append(n)
        elif (edsi[x]=='f'):
            b.append(f)
        elif (edsi[x]=='m'):
            b.append(m)
        elif (edsi[x]=='d'):
            b.append(d)
        elif (edsi[x]=='p'):
            b.append(p)
        elif (edsi[x]=='t'):
            b.append(t)
        elif (edsi[x]=='g'):
            b.append(g)
        else:
            return
    
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = Border(left=b[0],right=b[1],top=b[2],bottom=b[3])

def contorno_cli_clf( ws, ci,li,cf,lf, edsi ):

    n=Side(border_style= None    , color='00000000')
    f=Side(border_style='thin'   , color='00000000')
    m=Side(border_style='medium' , color='00000000')
    d=Side(border_style='double' , color='00000000')
    p=Side(border_style='dotted' , color='00000000')
    t=Side(border_style='dashed' , color='00000000')
    g=Side(border_style='thick'  , color='00000000')

    b=n

    if (edsi=="n"):
        b=n
    elif (edsi=='f'):
        b=f
    elif (edsi=='m'):
        b=m
    elif (edsi=='d'):
        b=d
    elif (edsi=='p'):
        b=p
    elif (edsi=='t'):
        b=t
    elif (edsi=='g'):
        b=g
    else:
        return

    for c in range(ci,cf+1):
        ws.cell(row=li-1,column=c).border = Border(n,n,n,b)
        ws.cell(row=lf+1,column=c).border = Border(n,n,b,n)
    for l in range(li,lf+1):
        ws.cell(row=l,column=ci-1).border = Border(n,b,n,n)
        ws.cell(row=l,column=cf+1).border = Border(b,n,n,n)

def cor_fundo_p_ci_li_cf_lf_c(planilha,colini,linini,colfim,linfim,cor):
    corfundo = PatternFill(start_color=cor,end_color=cor,fill_type='solid')
    for linha in range(linini,linfim+1):
        for coluna in range(colini,colfim+1):
            planilha.cell(linha, coluna).fill = corfundo


def adjust_column(ws, min_row, min_col, max_col):
    
    column_widths = []

    for i, col in \
        enumerate(
            ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
        ):

        for cell in col:
            value = cell.value
            if value is not None:

                if isinstance(value, str) is False:
                    value = str(value)

                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))

    for i, width in enumerate(column_widths):

        col_name = get_column_letter(min_col + i)
        value = column_widths[i] + 2
        value = int(round(value + value * 0.2,0))
        ws.column_dimensions[col_name].width = value

if __name__ == "__main__":
    global arquivo_destino
    arquivo_destino= ""
    log("-"*100)
    log(" - INICIO DO RELATÓRIO INSUMO CONSOLIDADO SPED FISCAL 21210708" ,sys.argv[0])
    variaveis = comum.carregaConfiguracoes(configuracoes)
#    variaveis = carregaConfiguracoes()
    ret = processar()
    if (ret > 0) :
        if(arquivo_destino):
            if os.path.isfile(arquivo_destino):
                os.remove(arquivo_destino)
    log("-" * 100)
    log(" - Código de execução = ", ret)
    log("-" * 100)
    log(" - FIM DO RELATÓRIO INSUMO CONSOLIDADO SPED FISCAL",sys.argv[0])
    sys.exit(ret)

