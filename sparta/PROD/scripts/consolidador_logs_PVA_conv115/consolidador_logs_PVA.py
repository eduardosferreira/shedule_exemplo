#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-
"""
----------------------------------------------------------------------------------------------
MODULO ...: TESHUVA
SCRIPT ...: loaderRelatorioPVA.py
CRIACAO ..: 05/07/2021
AUTOR ....: EDUARDO DA SILVA FERREIRA / KYROS TECNOLOGIA
            eduardof@kyros.com.br
DESCRICAO.: Geração de relatórios PVA 
----------------------------------------------------------------------------------------------
PARAMETROS: 
Parâmetros de entrada:
1)	MESANO: Mês e ano no formato MMAAAA - Obrigatório
2)	UF: UF do estado - Obrigatório
3)	IE: Inscrição estadual - Obrigatório
4)  DIR_BASE : Diretorio Base - Opcional  
----------------------------------------------------------------------------------------------
  HISTORICO : 
        Adequação para novo formato de script 
        SCRIPT ......: loader_sped_registro_O150.py
        AUTOR .......: Victor Santos
----------------------------------------------------------------------------------------------
"""

import os
import sys

global SD, dir_base
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)
import configuracoes

import datetime
import cx_Oracle
import re
import traceback
import string

# Lista de String
gv_lista_string = list(string.ascii_lowercase)

import comum
import sql
import layout
import util

log.gerar_log_em_arquivo = True
comum.carregaConfiguracoes(configuracoes)

from pathlib import Path
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook



# Parametros Globais
global ret # variavel para controle de retorno de erro 
           # (!= 0 -> ERRO, 0 - SUCESSO)
global txt # Variavel para tratamento de saida
# Entrada
global gv_mes_ano
global gv_mes
global gv_ano
global gv_uf
global gv_ie
global gv_diretorio_base

# Conexao
global gv_conexao
# Arquivo
global gv_usuario
global gv_senha
global gv_banco
global gv_caminho
global gv_relatorio
global gv_arquivos
global gv_arquivo_doc

"""
Retorna lista de alfabeto 
"""
def listAlphabet():
  return list(map(chr, range(97, 123)))

"""
Retorna a data formata 
"""
def dtf():
    return (datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S'))

"""
Retorna a lista de pasta
"""
def listar_Pasta(pasta):
    listas = list()
    if os.path.isdir(pasta):
        subpastas = list()
        items = os.listdir(pasta)
        for item in items:
            novo_item = os.path.join(pasta,item)
            if os.path.isdir(novo_item):
                subpastas.append(novo_item)
                listas.append(novo_item)

        if len(subpastas) > 0:            
            for subpasta in subpastas:
                lst = []
                lst = listar_Pasta(subpasta)
                if len(lst) > 0: 
                    listas.extend(lst)

    return listas

"""
Funcao para retornar UF 
"""
def retornar_UF(p_IE):
    l_UF = ""
    l_query="""
    select distinct f.unfe_sig  from openrisow.filial f where f.fili_cod_insest='%s' and rownum < 2
    """%(p_IE)

    l_cursor = sql.geraCnxBD(configuracoes)
    l_cursor.executa(l_query)
    l_result = l_cursor.fetchone()

    if(l_result != None): 
        for campo in l_result:
            l_UF = campo
    return(l_UF)

"""
Funcao para retornar DIRETORIO 
"""
def retornar_Diretorio(p_IE,p_MESANO):
    
    l_retorno = []
    l_query="""
    SELECT /*+ parallel(8) */
		'/portaloptrib/LEVCV115/'
		|| f.unfe_sig || '/'
		|| to_char(a.mes_ano,'YY/MM') ||'/' 
		|| f.emps_cod || '/' 
		|| a.fili_cod || '/SERIE/'
		|| a.id_serie_levantamento || '/PVA' 
		AS caminho
    FROM gfcarga.tsh_serie_levantamento a ,openrisow.filial f
    WHERE f.emps_cod = 'TBRA' 
    AND   f.fili_cod_insest = '%s'
    AND   a.emps_cod = f.emps_cod 
    AND   a.fili_cod = f.fili_cod
    AND   a.mes_ano = TO_DATE('01%s','DDMMYYYY')
	"""%(p_IE,p_MESANO)
    
    l_cursor = sql.geraCnxBD(configuracoes)
    l_cursor.executa(l_query)
    l_result = l_cursor.fetchone()

    if(l_result != None): 
        while l_result:
            for campo in l_result:
                l_retorno.append(campo)            
            l_result = l_cursor.fetchone() 
    return(l_retorno)
    
"""
Funcao para retornar arquivos existentes
"""
def ultimo_Arquivo_Diretorio(p_arq_mascara,p_diretorio,p_flag=0):
    l_qte = 0
    l_arquivo = ""
    l_diretorio = Path(p_diretorio)
    l_arq = l_diretorio.glob(p_arq_mascara)
    l_procura_arquivos = sorted(l_arq, reverse=False)
    v_numero =  int(0)
    v_numero_old =  int(0)
    v_arq = ""
    if l_procura_arquivos:        
        for arquivo in l_procura_arquivos:
            l_arquivo = str(arquivo)
            
            if l_arquivo.strip() == "":
                continue 
            
            l_qte = l_qte + 1
            
            try:
                v_numero =  int(l_arquivo.split(".")[0][-3:])
                if v_numero > v_numero_old:
                    v_numero_old = v_numero
                    v_arq = l_arquivo
            except:
                pass
            
            if p_flag == 1:
                break

        if l_qte == 0:
            if p_flag == 0:
                log('Arquivo ' + p_arq_mascara +' não está na pasta ' + p_diretorio)
            return ""
        else:
            if v_arq == "":
                v_arq = l_arquivo

            return(v_arq)

    else: 
        if p_flag == 0:
            log('Arquivo ' + p_arq_mascara +' não está na pasta ' + p_diretorio)
        return ""

"""
Retorna a definicao tipo de arquivo
"""
def tipoArquivo(path_arq) :
    try :
        fd = open(path_arq, 'r', encoding='iso-8859-1')
        fd.readline()
        fd.close()
    except :
        return 'utf-8'
    return 'iso-8859-1'

# Ponto de partida
if __name__ == "__main__" :
    
    ret = 0
    # Tratamento de excessao
    txt = ''
    # Conexao
    gv_conexao = None
    # Parametro Diretorio base 
    v_diretorio_base = ""
    
    try:
        log("-"*150)
        ### Cria os parametros do script 
        comum.addParametro( 'MESANO',  None, "MESANO (MMYYYY)", True, '122015' )
        comum.addParametro( 'UF', None, 'Unidade Federativa (UF)', True, 'SP')
        comum.addParametro( 'IE', None, 'Inscricao Estadual (IE)', True, '108383949112')
        comum.addParametro( 'DIR_BASE', None, 'Diretorio Base', False, '/arquivos/compartilhamento/PVA/2016/05')

        # Validacao dos parametros de entrada
        if not comum.validarParametros() :
            ret = 91
        else:
            gv_mes_ano = comum.getParametro('MESANO').upper().strip()
            gv_uf = comum.getParametro('UF').upper().strip()
            gv_ie = comum.getParametro('IE').upper().strip()
            # Opcional
            v_diretorio_base = str(comum.getParametro('DIR_BASE')).strip() 
            if (v_diretorio_base == "False"
                or v_diretorio_base is None
            ):
                v_diretorio_base = ""

            if (len(gv_mes_ano) != 6 
            ):
                log("PARAMETRO MES ANO: Invalido!") 
                ret = 91

            if not ret :
                gv_mes = gv_mes_ano[0:2]
                gv_ano = gv_mes_ano[2:6]

            if not ret :
                try:
                    if (int(gv_mes) < 1
                    or int(gv_mes) > 12 
                    ):
                        log("PARAMETRO MES : Invalido!") 
                        ret = 91
                except Exception as e:
                    log("PARAMETRO MES : Invalido!") 
                    ret = 91

            if not ret :
                try:
                    if (
                       int(gv_ano) > datetime.datetime.now().year
                    or int(gv_ano) < (datetime.datetime.now().year)-50
                    ):
                        log("PARAMETRO ANO : Invalido!") 
                        ret = 91
                except Exception as e:
                    log("PARAMETRO ANO : Invalido!") 
                    ret = 91
            
            if not ret :
                if len(gv_uf) != 2:
                    log("PARAMETRO UF: Invalido!") 
                    ret = 91

            if not ret :
                l_iei = re.sub('[^0-9]','',gv_ie)
                if ( (l_iei == "") or (l_iei == "''") or (l_iei == '""') or (int("0"+l_iei) == 0)):
                    log("PARAMETRO IE : Invalido!") 
                    ret = 91

            if not ret :
                if len(v_diretorio_base) > 0:
                    if not os.path.isdir(v_diretorio_base): 
                        log("PARAMETRO DIR_BASE: Invalido!"  + v_diretorio_base) 
                        ret = 91

        log("\n")

        # Verifica os parametros
        if not ret :
            try:
                gv_caminho = configuracoes.caminho.replace("<<MM>>",gv_mes).replace("<<AAAA>>",gv_ano).replace("<<UF>>",gv_uf).replace("<<IE>>",gv_ie).replace("<<MESANO>>",gv_mes_ano).strip()
                gv_relatorio = configuracoes.relatorio.replace("<<MM>>",gv_mes).replace("<<AAAA>>",gv_ano).replace("<<UF>>",gv_uf).replace("<<IE>>",gv_ie).replace("<<MESANO>>",gv_mes_ano).upper().strip()
                gv_arquivos = configuracoes.arquivos.replace("<<MM>>",gv_mes).replace("<<AAAA>>",gv_ano).replace("<<UF>>",gv_uf).replace("<<IE>>",gv_ie).replace("<<MESANO>>",gv_mes_ano).strip()
                gv_arquivo_doc = configuracoes.arquivo_documentacao.replace("<<MM>>",gv_mes).replace("<<AAAA>>",gv_ano).replace("<<UF>>",gv_uf).replace("<<IE>>",gv_ie).replace("<<MESANO>>",gv_mes_ano).strip()
                
                if not ret :
                    if (len(gv_relatorio) <= 5
                        or not gv_relatorio.endswith(".XLSX")
                        ):
                        log("PARAMETRO DO ARQUIVO  RELATORIO: INVALIDO! " + gv_relatorio) 
                        ret = 91
                    else:
                        log("Arquivo do relatorio : " + gv_relatorio)

                if not ret :
                    if (len(gv_arquivos) <= 5
                        or not gv_arquivos.upper().strip().endswith(".TXT")
                        ):
                        log("PARAMETRO DO ARQUIVOS: INVALIDO! " + gv_arquivos) 
                        ret = 91
                    else:
                        log("Arquivos : " + gv_arquivos)

                if not ret :
                    dir_path = os.path.dirname(os.path.realpath(__file__))
                    gv_arquivo_doc = os.path.join(dir_path, gv_arquivo_doc)
                    if not os.path.isfile(gv_arquivo_doc):
                        log("PARAMETRO DO ARQUIVO DOCUMENTACAO: INVALIDO! " + gv_arquivo_doc) 
                        ret = 91
                    else:
                        log("Arquivo DOCUMENTACAO : " + gv_arquivo_doc)
                        
                if not ret :
                    v_numero = "001"

                    if not os.path.isdir(gv_caminho):
                        log("Diretório não existente : " + gv_caminho)        
                        os.makedirs(gv_caminho)
                        log("Diretório criado : " + gv_caminho)
 
                    else:
                        v_mascara_arquivo = str(gv_relatorio).replace("_<<STATUS>>_","*")
                        v_mascara_arquivo = v_mascara_arquivo.replace("<<STATUS>>","*")
                        v_mascara_arquivo = v_mascara_arquivo.replace("_<<NNN>>_","*")
                        v_mascara_arquivo = v_mascara_arquivo.replace("<<NNN>>","*")
                        log(str(gv_caminho) + " >> Verificando ultimo relatorio : " + v_mascara_arquivo)
                        v_ultimo_arquivo = ultimo_Arquivo_Diretorio(v_mascara_arquivo,gv_caminho,2).strip()
                        
                        if len(v_ultimo_arquivo) > 0:
                            log("Ultimo relatorio gerado : " + v_ultimo_arquivo)
                            v_numero =  str(int(v_ultimo_arquivo.split(".")[0][-3:])+1).rjust(3,'0')  
                    
                    gv_relatorio = os.path.join(gv_caminho, gv_relatorio)
                    gv_relatorio = gv_relatorio.replace("_<<NNN>>_",v_numero)
                    gv_relatorio = gv_relatorio.replace("<<NNN>>",v_numero)                        
                    log("Novo relatorio a ser criado >: " + gv_relatorio) 
                
            except Exception as e:
                txt = traceback.format_exc()
                log("PARAMETRO DOS ARQUIVOS NAO ENCONTRADOS! " + str(e)) 
                ret = 91        

        v_diretorios = list()
        if not ret :
            try:
                v_diretorios_atuais = list()
                if not os.path.isdir(v_diretorio_base):
                    v_diretorios_atuais = retornar_Diretorio(p_IE=gv_ie,p_MESANO=gv_mes_ano)    
                else:
                    v_diretorios_atuais = listar_Pasta(v_diretorio_base)
                    v_diretorios_atuais.append(v_diretorio_base)
                if len(v_diretorios_atuais) == 0:
                    log("Diretórios atuais não existentes ! " + str(v_diretorios_atuais))
                    ret = 91
                else:
                    # log("Diretórios atuais : " + str(v_diretorios_atuais))
                    for diretorio in v_diretorios_atuais:
                        if len(str(diretorio).strip()) > 0:
                            v_ultimo_arquivo = ultimo_Arquivo_Diretorio(gv_arquivos,str(diretorio),1)
                            if len(str(v_ultimo_arquivo).strip()) > 0:
                                v_diretorios.append(str(diretorio).strip())   

                    if len(v_diretorios) == 0:
                        log("Diretórios não existentes com arquivos de erros ! " + str(v_diretorios))
                        ret = 91
            
            except Exception as e:
                txt = traceback.format_exc()
                log("ERRO VALIDAÇÃO DOS DIRETÓRIOS: " + str(e))
                ret = 93
                
        # Importar os dados para documentação
        v_dic_arq = dict()
        l_regrel=[[],[]]
        l_regrel[0]=["Código do Erro",
                    "Ocorrência",
                    "Arquivo",
                    "Campo",
                    "Posição Inicial",
                    "Posição Final",
                    "Descrição da Ocorrência"]
        l_nregrel = 0

        if not ret :
            try:
                log("Processando leitura do arquivo : " + gv_arquivo_doc)

                # inicializa a variaveis de controle
                l_flsair = 0 
                l_contador = 0
                l_tp_reg = ""

                # realiza a abertura do arquivo
                l_ent = open(gv_arquivo_doc,mode="r",encoding=tipoArquivo(gv_arquivo_doc))

                # realiza a leitura da linha do arquivo
                l_linhalida = l_ent.readline()

                # percorrre o arquivo
                while (l_linhalida and l_flsair == 0):
                    # inicio while
                    l_contador += 1

                    # Quebra em vetor        
                    l_dados_ent = l_linhalida.split("|")

                    l_tp_reg = ""
                    if len(l_dados_ent) >=7:
                        l_tp_reg = l_dados_ent[0].upper().strip()

                    if len(l_tp_reg) >=0:
                        # inicio if
                        l_nregrel += 1
                        l_regrel.append([])
                        l_regrel[l_nregrel].extend(l_dados_ent)
                        v_dic_arq[l_tp_reg] = l_dados_ent
                        # fim if

                    # proxima linha
                    l_linhalida = l_ent.readline()

                    # Fim while   

                # fechamento do arquivo
                l_ent.close()

                # Verifica se o arquivo teve algum processamentp                
                if l_nregrel == 0:
                    log("# " + str(len(l_regrel)) + " >> " + "Não processou nenhum dados do arquivo : " + gv_arquivo_doc)
                    ret = 91
                #else:
                #    log(str(v_dic_arq.items()) + " - " + str(len(v_dic_arq)))

            except Exception as e:
                txt = traceback.format_exc()
                log("ERRO IMPORTAÇÃO DO ARQUIVO DE DOCUMENTAÇÃO: " + str(e))
                ret = 93
                try:
                    l_ent.close()
                except:
                    pass    
        
        log("\n")
        l_regdados = [[],[]]
        l_regdados[0]=["Diretorio",
                        "Nome da Origem",
                        "Ocorrência",
                        "Tipo Arquivo",
                        "Numero Erro",
                        "Descrição",
                        "Quantidade"]
        l_nregdados = 0   
        v_STATUS = "OK"

        if not ret :

            for diretorio_local in v_diretorios:
            
                try:
                    
                    l_diretorio = Path(diretorio_local)
                    l_arq = l_diretorio.glob(gv_arquivos)
                    l_procura_arquivos = sorted(l_arq, reverse=False)
                
                    if l_procura_arquivos: 
            
                        for arquivo in l_procura_arquivos:
                        
                            l_arq_diretorio = os.path.join(diretorio_local, str(arquivo)) 
                                
                            try:
                        
                                log("Processando leitura do arquivo : " + l_arq_diretorio)
                    
                                # inicializa a variaveis de controle
                                l_flsair = 0 
                                l_contador = 0
                                l_tp_reg = ""
                    
                                # realiza a abertura do arquivo
                                l_ent = open(l_arq_diretorio,mode="r",encoding=tipoArquivo(l_arq_diretorio))
                    
                                # realiza a leitura da linha do arquivo
                                l_linhalida = l_ent.readline()
                    
                                # percorrre o arquivo
                                while (l_linhalida and l_flsair == 0):
                                    # inicio while
                                    l_contador += 1
                                    if (len(l_linhalida) > 98
                                        and l_linhalida[0:4].strip().upper() in ('DEST'
                                            , 'ITEM'
                                            , 'MEST')
                                        ):
                                        # inicio if
                                        l_nregdados += 1
                                        l_regdados.append([])
                                        # "Diretorio",
                                        l_regdados[l_nregdados].append(str(diretorio_local))
                                        # "Nome da Origem",
                                        l_regdados[l_nregdados].append(str(os.path.basename(arquivo)))
                                        # "Ocorrência",
                                        v_codigo_erro = l_linhalida[5:8].strip().upper()
                                        v_ocorrencia = ""
                                        if v_codigo_erro in v_dic_arq:
                                            v_ocorrencia = str(v_dic_arq[v_codigo_erro][1]).strip().upper()
                                            if (v_STATUS == "OK" and v_ocorrencia == "ERRO"):
                                                v_STATUS = "ERRO"    

                                        l_regdados[l_nregdados].append(v_ocorrencia)
                                        # "Tipo Arquivo",
                                        l_regdados[l_nregdados].append(str(l_linhalida[0:4].strip().upper()))
                                        # "Numero Erro",
                                        l_regdados[l_nregdados].append(str(v_codigo_erro))
                                        # "Descrição",
                                        l_regdados[l_nregdados].append(str(l_linhalida[9:-11].strip().upper()))
                                        # "Quantidade"
                                        l_regdados[l_nregdados].append(str(l_linhalida[-11:].strip().upper()))
                                        
                                        # l_regdados[l_nregdados].append(l_dados_ent)
                                        # fim if

                                    # proxima linha
                                    l_linhalida = l_ent.readline()
                                    # Fim while   
                    
                                # fechamento do arquivo
                                l_ent.close()
                    
                            except Exception as e:
                                txt = traceback.format_exc()
                                log(l_arq_diretorio + " >> ERRO LEITURA DO ARQUIVO: " + str(e))
                                ret = 93
                                try:
                                    l_ent.close()
                                except:
                                    pass    
                                break
                    
                    if ret != 0:
                        break
                    
                except Exception as e:
                    txt = traceback.format_exc()
                    log(str(diretorio_local) + " >> ERRO GERAÇÃO DADOS: " + str(e))
                    ret = 93
                    break

                if ret != 0:
                    break

            if l_nregdados == 0:
                log("# " + str(len(l_regdados)) + " >> " + "Não processou nenhum dados dos diretórios : " + str(v_diretorios))
                ret = 91
                
        log("\n")
        # Gerando Relatorio
        if not ret :
            try:
                
                log("Criando arquivo : " + gv_relatorio)
                #### Cria a planilha em memória....
                    
                arquivo_excel = Workbook()
                planilha1 = arquivo_excel.active
                planilha1.title = "Resultado PVA"
                planilha2 = arquivo_excel.create_sheet("Documentação", 1)
                    
                for linha in l_regdados:
                    planilha1.append(linha)

                for linha in l_regrel:
                    planilha2.append(linha)

                # Regulariza a dimensao das planilhas
                for nColP in gv_lista_string:
                    planilha1.column_dimensions[str(nColP).upper()].width = 30   
                    planilha2.column_dimensions[str(nColP).upper()].width = 30   

                for nColP in range(1,23):
                    planilha1.cell(1,nColP).font=Font(bold=True)                    
                    planilha2.cell(1,nColP).font=Font(bold=True)                    
    
                for nLinP in range(2,l_nregdados+2):
                    for nColP in (5,7):
                        planilha1.cell(nLinP,nColP).alignment = Alignment(horizontal='right')
                      
                # Grava a planilha Excel
                v_arquivo_excel = str(gv_relatorio).replace("<<STATUS>>",v_STATUS)                        
                arquivo_excel.save(v_arquivo_excel)

                log(" >> Processado arquivo : " + v_arquivo_excel)
            
            except Exception as e:
                txt = traceback.format_exc()
                log(str(gv_relatorio) + " >> ERRO GERAÇÃO RELATORIO: " + str(e))
                ret = 93

        if not ret :
            log("SUCESSO")
        else:
            log("ERRO")

        log("\n")
    
    except Exception as e:
        txt = traceback.format_exc()
        log("ERRO .: " + str(e))
        ret = 93
    
    sys.exit(ret if ret >= log.ret else log.ret )
