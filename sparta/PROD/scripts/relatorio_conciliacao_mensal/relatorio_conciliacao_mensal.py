#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-

"""
----------------------------------------------------------------------------------------------
  SISTEMA ..: 
  MODULO ...: 
  SCRIPT ...: relatorio_conciliacao_mensal.py
  CRIACAO ..: 16/10/2020
  AUTOR ....: Airton Borges da Silva Filho / KYROS Consultoria
  DESCRICAO : 
----------------------------------------------------------------------------------------------
  HISTORICO :
    * 16/10/2020 - Airton Borges da Silva Filho / KYROS Consultoria - Criacao do script.
    * 09/12/2020 - Airton - Incluido o import re
        corrigido se foi informado "" (aspas)  como um codigo de filial
    * 26/02/2021 - Airton Borges da Silva Filho / KYROS Consultoria
        substituida a query de selecão dos protocolados.
    
    * 23/03/2021 - Welber Pena - Kyros Consultoria
        ALT003 
            Documentação :
                - ALT003_AcertoSelectTabelaseFiliais.pdf
                - ALT003_AcertoSelectTabelaseFiliais_parte2.pdf
            Alterar a query que gera o relatorio do PROTOCOLADO.
    
    * 23/03/2021 - Welber Pena - Kyros Consultoria
        ALT004
            Documentação : 
                - ALT004_RemocaoColunaQtdNfs.pdf
            Remoção de coluna QTD_NFS nas abas "Detalhado REG, REG X PROT Dinamica e REG X PROT"
            do relatorio.

    * 23/03/2021 - Victor Santos - Kyros Consultoria
        ALT005
            Documentação : 
                Alteração Relatório Conciliação-Serie 4989 – Teshuvá
                Alteração na consulta que gera a primeira aba deste relatório

    Adequação para novo formato de script 
    SCRIPT ......: loader_sped_registro_O150.py
    AUTOR .......: Victor Santos
	
     * 13/01/2022 - Eduardo da Silva Ferreira - Kyros Tecnologia
        PTITES-1367 : Aonde que faz referencia na leitura das informações do protocolado, 
        verificar se a ultima entrega de acordo com a informação 
        "ORIGEM_PROTOCOLADO" de acordo com a tabela "TSH_SERIE_LEVANTAMENTO"
        https://jira.telefonica.com.br/browse/PTITES-1367
        https://wikicorp.telefonica.com.br/x/JKMPDQ
        Funcao alterada:
            - busca_regxprot() 
            - totaliza_coluna
   * 20/01/2022 - Victor Santos - Kyros Consultoria
        ALT006
            Documentação:
                1)Renomear aba "Detalhado REG" para "Detalhado ATUAL_TI"
                2)Renomear aba "REG X PROT Dinamica" para "ATUAL_TI x ULT_PROT Dinamica"
                3)Renomear aba "REG X PROT" para "ATUAL_TI x ULT_PROT"
                4)Atualizar aba "ATUAL_TI x ULT_PROT Dinamica" (antiga "REG X PROT Dinamica") alterar "Protocolado" para "ULTIMO_PROTOCOLADO"
                5)Atualizar "ATUAL_TI x ULT_PROT" (Antiga "REG X PROT")
                  5.1) Acrescentar coluna Original com 'S' para dados da tabela Original C2 e N para dados c6.
                  5.2) Ajustar para coluna Area possuir apenas 'ULTIMO_PROTOCOLADO' e 'ATUAL_TI'
    
    * 03/02/2021 - Welber Pena - Kyros Tecnologia
        ALT007
            - Alterado o diretorio final onde o script gera o relatorio.
----------------------------------------------------------------------------------------------
"""

import os
import sys

global SD, dir_base
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)

import configuracoes

import datetime
import os
import cx_Oracle
import re
from openpyxl import Workbook

import comum
import sql
import layout
import util

log.gerar_log_em_arquivo = True

ret = 0
nome_relatorio = "" 

SD = ('/' if os.name == 'posix' else '\\')

def semespacos(frase):
    retorno = "" 
    for l in frase:
        if (l != " "):
            retorno = retorno + l
    return(retorno)

#  inicio PTITES-1367
def validar_filiais(uf,fils,datai):
    query = """
        select distinct fili_cod
        from  gfcarga.TSH_SERIE_LEVANTAMENTO
        where uf_filial = '%s'
              AND EMPS_COD = 'TBRA'
              AND mes_ano  = TO_DATE('%s','dd/mm/yyyy')
              AND fili_cod IN %s
              """%(uf, datai,fils)
            
    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchone()

    if result == None:
        print("#### ATENÇÃO: Nenhum Resultado para lista de filiais")
        print("####     Query = ")
        print("####")
        print(query)
        print("####")
        ret=99
        return(result)
    else:
        filis = "('"
        while result:
            filis = filis + result[0] + "','"
            result = cursor.fetchone()
    filis = filis[:len(filis)-2]
    filis = filis + ")"

    return(filis)
#  fim PTITES-1367

def busca_filiais(uf,fils,datai):
    if fils : return(fils)
    query = """
        select distinct fili_cod
        from  gfcarga.TSH_SERIE_LEVANTAMENTO
        where uf_filial = '%s'
              AND EMPS_COD = 'TBRA'
              AND mes_ano  = TO_DATE('%s','dd/mm/yyyy')
              """%(uf, datai)
            
    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchone()

    if result == None:
        print("#### ATENÇÃO: Nenhum Resultado para lista de filiais")
        print("####     Query = ")
        print("####")
        print(query)
        print("####")
        ret=99
        return(result)
    else:
        filis = "('"
        while result:
            filis = filis + result[0] + "','"
            result = cursor.fetchone()
    filis = filis[:len(filis)-2]
    filis = filis + ")"

    return(filis)

def busca_detalhadoreg(datai,dataf,filis):
    query = """
            with controle as (
                select /*+ materialize */
                    EMPS_COD
                    , fili_cod
                    , CTR_APUR_DTINI
                    , CTR_SER_ORI                     
                    , CTR_NUM_NFINI                 
                    , CTR_NUM_NFFIN                 
                    , CTR_NF_NOMARQ                 
                    , CTR_ITEM_NOMARQ               
                    , CTR_CODH_ARQNF 
                    , CTR_CODH_ARQITEM 
                    , CTR_CODH_ARQCLI 
                    , CTR_CODH_REG 
                from (
                select
                    EMPS_COD
                    , fili_cod
                    , CTR_APUR_DTINI
                    , CTR_SER_ORI                     
                    , CTR_NUM_NFINI                 
                    , CTR_NUM_NFFIN                 
                    , CTR_NF_NOMARQ                 
                    , CTR_ITEM_NOMARQ               
                    , CTR_CODH_ARQNF 
                    , CTR_CODH_ARQITEM 
                    , CTR_CODH_ARQCLI 
                    , CTR_CODH_REG 
                    , RANK() OVER (PARTITION BY c.EMPS_COD, c.FILI_COD, c.CTR_SERIE, to_number(c.CTR_VOLUME) ORDER BY CTR_IND_RETIF desc) SEQ_RETIFICACAO
                from openrisow.CTR_IDENT_CNV115 c
                where CTR_APUR_DTINI  >= to_date('%s','dd/mm/yyyy')
                AND CTR_APUR_DTINI  <= to_date('%s','dd/mm/yyyy')
                AND EMPS_COD         = 'TBRA'
                AND FILI_COD        in %s

                )
                where SEQ_RETIFICACAO = 1
                )              
                SELECT /*+ parallel(inf,8), index(inf ITEM_NFTL_SERVI2) */ 
                    inf.EMPS_COD,
                    to_char(inf.infst_dtemiss,'MM/yy')       AS MES_ANO,
                    inf.INFST_SERIE                          AS serie,
                    inf.CFOP,
                    inf.estb_cod                             AS CST,
                    inf.infst_aliq_icms                      AS ALIQUOTA, 
                    inf.infst_tribicms                       AS IND_TRIB,
                    sum(inf.INFST_VAL_SERV)                  AS VALOR_TOTAL,
                    sum(inf.INFST_BASE_ICMS)                 AS BASE_ICMS,
                    sum(inf.INFST_VAL_ICMS)                  AS VALOR_ICMS,
                    sum(inf.INFST_ISENTA_ICMS)               AS ISENTAS_ICMS,
                    sum(inf.INFST_OUTRAS_ICMS)               AS OUTROS_VALORES,
                    sum(inf.INFST_VAL_DESC)                  AS DESCONTOS,
                    sum(CASE WHEN inf.infst_dtemiss < to_date('%s','dd/mm/yyyy') THEN NVL(inf.INFST_VAL_SERV,0) - NVL(inf.INFST_VAL_DESC,0)
                                ELSE NVL(inf.INFST_VAL_CONT,0)
                        END)                                 AS TOTAL_LIQUIDO,
                    SUM(inf.INFST_VAL_RED)                   AS VALOR_REDUCAO,
                    'ATUAL_TI'                               AS AREA,
                    ctr.CTR_CODH_ARQNF hash_nf,
                    ctr.CTR_CODH_ARQITEM hash_item,
                    ctr.CTR_CODH_ARQCLI hash_cli,
                    ctr.CTR_CODH_REG hash_reg      
                FROM OPENRISOW.ITEM_NFTL_SERV inf,
                    controle ctr
                WHERE 1=1
                and inf.infst_dtemiss  >= to_date('%s','dd/mm/yyyy')
                AND inf.infst_dtemiss  <= to_date('%s','dd/mm/yyyy')
                AND inf.emps_cod       = 'TBRA'
                AND inf.fili_cod       in %s  
                and ctr.emps_cod = inf.emps_cod
                and ctr.FILI_COD = inf.FILI_COD
                and CTR_APUR_DTINI = trunc(infst_dtemiss,'month')
                and to_number(inf.infst_num) between CTR_NUM_NFINI and CTR_NUM_NFFIN
                and ctr.CTR_SER_ORI = inf.infst_serie
                AND inf.INFST_IND_CANC  =  'N' 
                group by inf.EMPS_COD,
                        to_char(inf.infst_dtemiss,'MM/yy'),
                        inf.INFST_SERIE,
                        inf.CFOP,
                        inf.estb_cod,
                        inf.infst_aliq_icms,
                        inf.infst_tribicms,
                        ctr.CTR_CODH_ARQNF,
                        ctr.CTR_CODH_ARQITEM,
                        ctr.CTR_CODH_ARQCLI,
                        ctr.CTR_CODH_REG
    
    """%(datai,dataf,filis,datai,datai,dataf,filis)
    
    retorno = [[]]
    lin = 0 
    retorno[0]=[
        "EMPS_COD",
        "MES_ANO",
        "SERIE",
        "CFOP",
        "CST",
        "ALIQUOTA",
        "IND_TRIB",
        "VALOR_SERVICO",
        "BASE_ICMS",
        "VALOR_ICMS",
        "ISENTAS_ICMS",
        "OUTROS_VALORES",
        "DESCONTOS",
        "TOTAL_LIQUIDO",
        "VALOR_REDUCAO",
        "AREA",
        "HASH_NF",
        "HASH_ITEM",
        "HASH_CLI",
        "HASH_REG"
        ]
    
    cursor = sql.geraCnxBD(configuracoes)
    cursor.executa(query)
    result = cursor.fetchone()

    if result == None:
        print("#### ATENÇÃO: Nenhum Resultado para Detalhado ATUAL_TI")
        print("####     Query = ")
        print("####")
        print(query)
        print("####")
        ret=99
        return(retorno)
    else:
        while result:
            lin = lin + 1
            retorno.append([])
            for campo in result:
                retorno[lin].append(campo)
            result = cursor.fetchone()
    return(retorno)

def busca_regxprot(uf,datai,dataf,filis):
    
    c1 = ''
    if configuracoes.banco != 'GFCLONEDEV' :
        c1 = '@c1'

    # -- Inicio PTITES-1367
    c6 = '@c6'
    if configuracoes.banco:
        if 'C6' in str(configuracoes.banco).strip().upper():
            c6 = ''

    queryP = """ 
SELECT
  /*+parallel(15) */
  EMPS_COD,
  MES_ANO,
  SERIE,
  CFOP,
  SUM(VALOR_LIQUIDO)                       AS VALOR_LIQUIDO, 
  SUM(VALOR_SERVICO)                       AS VALOR_SERVICO,
  SUM(BASE_ICMS)                           AS BASE_ICMS,
  SUM(VALOR_ICMS)                          AS VALOR_ICMS,
  SUM(ISENTAS_ICMS)                        AS ISENTAS_ICMS,
  SUM(OUTROS_VALORES)                      AS OUTROS_VALORES,
  SUM(DESCONTO)                            AS DESCONTO,
  SUM(VALOR_REDUCAO)                       AS VALOR_REDUCAO,
  AREA,
  ORIGINAL                                  
FROM
  ( WITH TMP AS
  (SELECT '%s'                                         AS UF_FILIAL,
    s.MES_ANO                                          AS DATA_EMISSAO_INICIO,
    NVL(TO_DATE('%s', 'dd/mm/yyyy'),LAST_DAY(s.MES_ANO)) AS DATA_EMISSAO_FIM,
    s.ID_SERIE_LEVANTAMENTO ,
    s.SERIE ,
    s.MES_ANO,
    s.FILI_COD,
    NVL(s.ORIGEM_PROTOCOLADO,'ORIGINAL') AS ORIGEM_PROTOCOLADO
  FROM GFCARGA.TSH_SERIE_LEVANTAMENTO%s s
  WHERE s.MES_ANO = TO_DATE('%s', 'DD/MM/YYYY')
  AND s.FILI_COD IN %s
  )
SELECT
  'TBRA'                                          AS EMPS_COD,
  TO_CHAR(m.data_emissao,'mm/yy')                 AS MES_ANO,
  s.serie                                         AS SERIE,
  m.cfop                                          AS CFOP,
  (NVL(m.valor_total, 0) - NVL(m.desconto, 0))    AS VALOR_LIQUIDO, --Antes VALOR_CONTABIL
  (NVL(m.valor_total, 0))                         AS VALOR_SERVICO,
  (NVL(m.base_icms, 0))                           AS BASE_ICMS,
  (NVL(m.valor_icms, 0))                          AS VALOR_ICMS,
  (NVL(m.isentas_icms, 0))                        AS ISENTAS_ICMS,
  (NVL(m.outros_valores, 0))                      AS OUTROS_VALORES,
  (NVL(m.desconto, 0))                            AS DESCONTO,
  0                                               AS VALOR_REDUCAO,
  'ULTIMO_PROTOCOLADO'                            AS AREA,
  'S'                                             AS ORIGINAL  
FROM gfcarga.tsh_item_conv_115%s m
INNER JOIN TMP s
ON s.origem_protocolado     = 'ORIGINAL'
AND m.id_serie_levantamento = s.id_serie_levantamento
AND m.uf_filial             = s.uf_filial
AND m.data_emissao         >= s.DATA_EMISSAO_INICIO
AND m.data_emissao         <= s.DATA_EMISSAO_FIM
AND m.SIT_DOC               = 'N'
UNION ALL
SELECT
  'TBRA'                                          AS EMPS_COD,
  TO_CHAR(m.data_emissao,'mm/yy')                 AS MES_ANO,
  s.serie                                         AS SERIE,
  m.cfop                                          AS CFOP,
  (NVL(m.valor_total, 0) - NVL(m.desconto, 0))    AS VALOR_LIQUIDO, --Antes VALOR_CONTABIL
  (NVL(m.valor_total, 0))                         AS VALOR_SERVICO,
  (NVL(m.base_icms, 0))                           AS BASE_ICMS,
  (NVL(m.valor_icms, 0))                          AS VALOR_ICMS,
  (NVL(m.isentas_icms, 0))                        AS ISENTAS_ICMS,
  (NVL(m.outros_valores, 0))                      AS OUTROS_VALORES,
  (NVL(m.desconto, 0))                            AS DESCONTO,
  0                                               AS VALOR_REDUCAO,
  'ULTIMO_PROTOCOLADO'                            AS AREA,
  'N'                                             AS ORIGINAL 
FROM gfcarga.tsh_item_conv_115_ent%s m
INNER JOIN TMP s
ON s.origem_protocolado     = 'ULTIMO_ENTREGUE'
AND m.id_serie_levantamento = s.id_serie_levantamento
AND m.uf_filial             = s.uf_filial
AND m.data_emissao         >= s.DATA_EMISSAO_INICIO
AND m.data_emissao         <= s.DATA_EMISSAO_FIM
AND m.SIT_DOC               = 'N'
  )
GROUP BY EMPS_COD, MES_ANO,  SERIE,  CFOP, AREA, ORIGINAL
    """%(uf,dataf,c1,datai,filis,c1,c6) 
    
    # c2, c2, datai, uf, filis, datai, dataf )
    # -- Fim PTITES-1367
     
    queryR = """
        SELECT /*+ parallel(tmp,8), index(tmp ITEM_NFTL_SERVI2) */
            tmp.EMPS_COD,
            to_char(tmp.infst_dtemiss,'MM/yy')           as MES_ANO,
            tmp.INFST_SERIE                              as serie,
            tmp.CFOP,
            sum(CASE WHEN tmp.infst_dtemiss < to_date('01/01/2017','dd/mm/yyyy') 
                        THEN NVL(INFST_VAL_SERV,0) - NVL(INFST_VAL_DESC,0)
                        ELSE NVL(INFST_VAL_CONT,0)
                END)                                     as VALOR_LIQUIDO,--Antes VALOR_CONTABIL
            sum(INFST_VAL_SERV)                          as VALOR_SERVICO,--Antes VALOR_TOTAL
            sum(INFST_BASE_ICMS)                         as BASE_ICMS,
            sum(INFST_VAL_ICMS)                          as VALOR_ICMS,
            sum(INFST_ISENTA_ICMS)                       as ISENTAS_ICMS,
            sum(INFST_OUTRAS_ICMS)                       as OUTROS_VALORES,
            sum(INFST_VAL_DESC)                          as DESCONTO,
            sum(INFST_VAL_RED) VALOR_REDUCAO,
            'ATUAL_TI'                                   as AREA,
            ''                                           as ORIGINAL   
        FROM openrisow.ITEM_NFTL_SERV tmp
        WHERE 1 = 1
        and tmp.emps_cod         = 'TBRA'
        AND FILI_COD IN             %s                                -- ALTERAR A FILIAL PARA AS FILIAIS DO RELATÓRIO A SER REALIZADO
        and tmp.infst_dtemiss   >= to_date('%s','dd/mm/yyyy')         -- ALTERAR A DATA PARA O PERÍODO A SER REALIZADO O RELATÓRIO
        AND tmp.infst_dtemiss   <=  to_date('%s','dd/mm/yyyy')         -- ALTERAR A DATA PARA O PERÍODO A SER REALIZADO O RELATÓRIO
        and tmp.INFST_IND_CANC   = 'N'
        group by tmp.EMPS_COD,
                to_char(tmp.infst_dtemiss,'MM/yy'),
                tmp.INFST_SERIE,
                tmp.CFOP
        """%(filis,datai,dataf)

    retorno = [[]] 
    lin = 0
    retorno[0]=[
        "EMPS_COD",
        "MES_ANO",
        "SERIE",
        "CFOP",
        "VALOR_LIQUIDO",
        "VALOR_SERVICO",
        "BASE_ICMS",
        "VALOR_ICMS",
        "ISENTAS_ICMS",
        "OUTROS_VALORES",
        "DESCONTOS",
        "VALOR_REDUCAO",
        "AREA",
        "ORIGINAL"
        ]
    
    
    print("# ",dtf() , "        Buscando PROTOCOLADOS no BANCO DE DADOS....")
    cursorP = sql.geraCnxBD(configuracoes)
    cursorP.executa(queryP)
    result = cursorP.fetchone()
   
    if result == None:
        print("#### ATENÇÃO: Nenhum Resultado REG(ATUAL_TI) X PROT,  PROTOCOLADOS")
        print("####     Query para PROTOCOLADOS = ")
        print("####")
        print(queryP)
        print("####")
        ret=99
    else:
        while result:
            lin = lin + 1
            retorno.append([])
            for campo in result:
                retorno[lin].append(campo)
            result = cursorP.fetchone()

    print("# ",dtf() , "        Buscando REG(ATUAL_TI) no BANCO DE DADOS....")
    cursorR = sql.geraCnxBD(configuracoes)
    cursorR.executa(queryR)
    result = cursorR.fetchone()
    if result == None:
        print("#### ATENÇÃO: Nenhum Resultado REG(ATUAL_TI) X PROT, ATUAL_TI")
        print("####     Query para ATUAL_TI = ")
        print("####")
        print(queryR)
        print("####")
        ret=99
    else:
        while result:
            lin = lin + 1
            retorno.append([])
            for campo in result:
                retorno[lin].append(campo)
            result = cursorR.fetchone()
    return(retorno)

def busca_series(p):
    lseries = []    
    maxp = p.max_row
    for i in range(1,maxp + 1):
        serie = semespacos(p.cell(row=i,column=3).value)
        existe = False
        for eserie in lseries:
            if (eserie == serie):
                existe = True
        if (existe == False and serie):
            lseries.append(serie)
    return(lseries)

def totaliza_coluna(p,serie,col):
    tprot = 0.00
    treg  = 0.00
    maxl  = p.max_row
    for i in range(2,maxl + 1):
        if (semespacos(p.cell(row=i,column=3).value) == serie):
            if ('PROTOCOLADO' in p.cell(row=i,column=13).value):#PTITES-1367 
                tprot = tprot + (p.cell(row=i,column=col).value or 0.00)
            if ('ATUAL_TI' in p.cell(row=i,column=13).value):#PTITES-1367
                treg = treg + (p.cell(row=i,column=col).value or 0.00)
    return(tprot,treg)

def busca_regxprotdin(plan):

    retorno = [[]] 
    lin = 0
    retorno[0]=[
        "SERIE",
        "AREA",
        "VALOR_LIQUIDO",
        "VALOR_SERVICO",
        "BASE_ICMS",
        "VALOR_ICMS",
        "ISENTAS_ICMS",
        "OUTROS_VALORES",
        "DESCONTOS",
        # "VALOR_REDUCAO", #ALT006
        ]
    series=busca_series(plan)
    totalc = []
    for t in range(1, len(series)):
        lin = t * 2
        retorno.append([])
        retorno.append([])
        retorno[lin].append(semespacos(series[t]))            
        retorno[lin].append('ATUAL_TI')
        retorno[lin-1].append(semespacos(series[t]))
        retorno[lin-1].append('ULTIMO_PROTOCOLADO') #ALT006
        for col in range(5,5+(len(retorno[0])-2) ):
            totalc=[]
            totalc=totaliza_coluna(plan,semespacos(series[t]),col)
            retorno[lin].append(totalc[1])
            retorno[lin-1].append(totalc[0])
    return(retorno)

def processar():
    ufi = ""
    mesanoi = ""
    mesi = ""
    anoi = "" 
    filiaisi = ""
    ret = 0

    if (len(sys.argv) >= 3 ): 
        ufi = sys.argv[1].upper()
    if (len(sys.argv) >= 3 
        and util.validauf(ufi)
        and len(sys.argv[2])==6  
        and int(sys.argv[2][:2])>0 
        and int(sys.argv[2][:2])<13
        and int(sys.argv[2][2:])<=datetime.datetime.now().year
        and int(sys.argv[2][2:])>(datetime.datetime.now().year)-50
        ):
     
        mesanoi = sys.argv[2].upper()
        mesi = sys.argv[2][:2].upper()
        anoi = sys.argv[2][2:].upper()
        datai = "01/"+mesi+"/"+anoi
        dataf =  str(ultimodia(int(anoi),int(mesi)))+"/"+str(mesi)+"/"+str(anoi)

        if len(sys.argv) > 3:
            for i in range(3,len(sys.argv)):
                f = sys.argv[i].upper()
                f = re.sub('[^0-9]','',f)
                if ( (f == "") or (f == "''") or (f == '""') ): 
                    continue
                else:
                    if (len(filiaisi) < 1): 
                        filiaisi = "("
                    for x in range(len(sys.argv[i].split(','))) :
                        filiaisi = filiaisi + "'" + sys.argv[i].split(',')[x].strip() + "', "
            if (len(filiaisi) > 3):
                filiaisi = filiaisi[:len(filiaisi)-2] + ")"
                #  inicio PTITES-1367
                fl_validar_filiais = validar_filiais(ufi,filiaisi,datai)
                if not fl_validar_filiais:
                    print("-" * 100)
                    print("#### ")
                    print('#### ERRO - Erro nos parametros do script.')
                    print("FILIAIS INVÁLIDAS! ou inexistente para o periodo! ")
                    print('#### ')
                    print("-" * 100)
                    print("")
                    print("Retorno = 99") 
                    ret = 99
                    return(99)
                #  fim PTITES-1367
        
    else :
        print("-" * 100)
        print("#### ")
        print('#### ERRO - Erro nos parametros do script.')
        print("#### ")
        print('#### Exemplo de como deve ser :')
        print('####      %s <UF> <MMAAAA> [FILIAL] [FILIAL] ... '%(sys.argv[0]))
        print("#### ")
        print('#### Onde')
        print('####      <UF> = estado. Ex: SP')
        print('####      <MMAAAA> = mês e ano. Ex: Para junho de 2020 informe 062020')
        print('####      [FILIAL] = código da filial. É opcional, pode ou não ser informado.')
        print('####                 Se informado mais de um, devem ser separados por espaço Ex: 0001 9201 9144') 
        print('####                 caso não informado, será processado para todas filiais do estado <UF> informado.')
        print("#### ")
        print('#### Portanto, se o estado = SP, o mes = 06 e o ano = 2020, e deseja todas as filiais,  o comando correto deve ser :')  
        print('####      %s SP 062020'%(sys.argv[0]))  
        print("#### ")
        print('#### Se desejar processar apenas para a FILIAL 0001, o comando deve ser:')  
        print('####      %s SP 062020 0001'%(sys.argv[0]))  
        print("#### ")
        print('#### Se desejar processar para as FILIAIS 0001, 9144 e 9201, o comando deve ser:')  
        print('####      %s SP 062020 0001 9144 9201'%(sys.argv[0]))  
        print("#### ")
        print('#### Neste exemplo, a planilha resultado será:') 
        print('####      %sarquivos%sRELATORIOS%sCONV115%sSP%s2020%s06%sRelatorio_Conciliacao_SP_062020.xlsx'%(SD,SD,SD,SD,SD,SD,SD)) 
        print('#### ')
        print("-" * 100)
        print("")
        print("Retorno = 99") 
        ret = 99
        return(99)

#### ALT007 - Inicio
    ##  - Alterado o caminho base de geracao dos relatorios
    ## dir_base = configuracoes.dir_base
    #### Monta caminho e nome do relatório
    dir_base = os.path.join(configuracoes.dir_geracao_arquivos.split('relatorio_conciliacao')[0], 'Insumos', 'SPED_FISCAL')
    dir_relatorio = os.path.join(dir_base, ufi, anoi, mesi)
#### ALT007 - Fim
   
#### Se a pasta do relatório não existir, cria
    if not os.path.isdir(dir_relatorio) :
        os.makedirs(dir_relatorio)
   
#### Cria a planilha excel em memória....
    arquivo_excel = Workbook()
    planilha0 = arquivo_excel.active
    planilha0.title = "Detalhado ATUAL_TI"
    planilha1 = arquivo_excel.create_sheet("ATUAL_TI x ULT_PROT Dinamica", 1)
    planilha2 = arquivo_excel.create_sheet("ATUAL_TI x ULT_PROT", 2)

### Monta o nome da planilha = relatório
    nome_relatorio = os.path.join(dir_relatorio, "Relatorio_Conciliacao_"+ufi+"_"+mesanoi+".xlsx")
    print("# ",dtf() , " - Início do processamento da ATUAL_TI x ULT_PROT")
    print("# ",dtf() , "        Buscando lista de filiais do estado....")
    filiais = busca_filiais(ufi,filiaisi,datai)
    print("# ",dtf() , "            FIM da lista de filiais.")
    print("# ",dtf() , "        Buscando totalizações de PROTOCOLADO e ATUAL_TI....")
    dadosregxprot = busca_regxprot(ufi,datai,dataf,filiais)
    print("# ",dtf() , "            FIM da busca de totalizações PROTOCOLADO e ATUAL_TI")
    print("# ",dtf() , "        Buscando totalizações de Detalhado ATUAL_TI....")
    dadosdetalhadoreg = busca_detalhadoreg(datai,dataf,filiais)
    print("# ",dtf() , "            FIM da busca de totalizações Detalhado ATUAL_TI")

    for linha in dadosregxprot:
        planilha2.append(linha)
    arquivo_excel.save(nome_relatorio)
  
    for linha in dadosdetalhadoreg:
        planilha0.append(linha)
    arquivo_excel.save(nome_relatorio)
  
    print("# ",dtf() , "        Gerando a aba ATUAL_TI x ULT_PROT Dinamica....")
    dadosregxprotdin = busca_regxprotdin(planilha2)
    print("# ",dtf() , "            FIM da geração da aba ATUAL_TI x ULT_PROT Dinamica")
  
    for linha in dadosregxprotdin:
        planilha1.append(linha)
    arquivo_excel.save(nome_relatorio)
    
    print("#"*100)
    print("# ")
    print("####  ARQUIVO DE SAIDA = " )
    print("####     ",nome_relatorio )
    print("# ")
    print("#"*100)
    
    return(0)

# def validauf(uf):
#     return(True if (uf.upper() in ('AC','AL','AM','AP','BA','CE','DF','ES','GO','MA','MG','MS','MT','PA','PB','PE','PI','PR','RJ','RN','RO','RR','RS','SC','SE','SP','TO')) else False)

def dtf():
    return (datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S'))

def ultimodia(ano,mes):
   return(31 if mes == 12 else datetime.date.fromordinal((datetime.date(ano,mes+1,1)).toordinal()-1).day)

if __name__ == "__main__":
    cod_saida = 0
    print('-'*100)
    print("# ")  
    print("# ",dtf() , " - Início do processamento. conciliacao_serie")
    print("# ")
    comum.carregaConfiguracoes(configuracoes)
    cod_saida = processar()
    print('-'*100)
    print("Código de saida = ",cod_saida)
    sys.exit(cod_saida)
