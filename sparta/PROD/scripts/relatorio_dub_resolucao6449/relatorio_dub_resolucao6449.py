#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-
"""
----------------------------------------------------------------------------------------------
MODULO ...: TESHUVA
SCRIPT ...: loaderRelatorio_DUB_Resolucao6449.py
CRIACAO ..: 13/07/2021
AUTOR ....: EDUARDO DA SILVA FERREIRA / KYROS TECNOLOGIA
            eduardof@kyros.com.br
DESCRICAO.: Geração de relatórios PVA 
----------------------------------------------------------------------------------------------
PARAMETROS: 
Parâmetros de entrada:
1)	IE: Inscrição estadual - Obrigatório
2)	ANO: Ano no formato AAAA - Obrigatório
3)	TRIMESTRE: Periodo Trimestre - Obrigatório

----------------------------------------------------------------------------------------------
  HISTORICO : 
   01/09/2021 - Marcelo Gremonesi -  Adquações para novo painel
----------------------------------------------------------------------------------------------
"""
import sys
import os
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)
import configuracoes

import datetime
import cx_Oracle
import re
import shutil
import traceback
import string
import openpyxl
from openpyxl.utils import get_column_letter

from openpyxl.cell import cell

# Nome do script
nome_script = os.path.basename( sys.argv[0] ).replace('.py', '')

# Lista de String
gv_lista_string = list(string.ascii_lowercase)

import comum
import sql
variaveis = {'teste': '001'}
comum.variaveis = variaveis
sql.variaveis = variaveis
from layout import *
from pathlib import Path
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook

log.gerar_log_em_arquivo = True


"""
Funcao para retornar UF 
"""
def retornar_UF(p_IE):

    l_UF = ""

    l_query="""
    select distinct f.unfe_sig  from openrisow.filial f where f.fili_cod_insest='%s' and rownum < 2
    """%(p_IE)

    l_cursor = sql.geraCnxBD(configuracoes)
    l_cursor.executa(l_query)
    l_result = l_cursor.fetchone()

    if(l_result != None): 
        for campo in l_result:
            l_UF = campo

    return(l_UF)
"""
DEFINE O TAMANHO DAS COLUNAS
"""
def adjust_column(ws, min_row, min_col, max_col):
    column_widths = []
    for i, col in \
        enumerate(
            ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
        ):
        for cell in col:
            value = cell.value
            if value is not None:
                if isinstance(value, str) is False:
                    value = str(value)
                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))
    for i, width in enumerate(column_widths):
        col_name = get_column_letter(min_col + i)
        value = column_widths[i] + 2
        value = int(round(value + value * 0.2,0))
        ws.column_dimensions[col_name].width = value
"""
CONTORNO
"""
def contorno_cli_clf( ws, ci,li,cf,lf, edsi ):
    n=Side(border_style= None    , color='00000000')
    f=Side(border_style='thin'   , color='00000000')
    m=Side(border_style='medium' , color='00000000')
    d=Side(border_style='double' , color='00000000')
    p=Side(border_style='dotted' , color='00000000')
    t=Side(border_style='dashed' , color='00000000')
    g=Side(border_style='thick'  , color='00000000')

    b=n

    if (edsi=="n"):
        b=n
    elif (edsi=='f'):
        b=f
    elif (edsi=='m'):
        b=m
    elif (edsi=='d'):
        b=d
    elif (edsi=='p'):
        b=p
    elif (edsi=='t'):
        b=t
    elif (edsi=='g'):
        b=g
    else:
        return

    for c in range(ci,cf+1):
        ws.cell(row=li-1,column=c).border = Border(n,n,n,b)
        ws.cell(row=lf+1,column=c).border = Border(n,n,b,n)
    for l in range(li,lf+1):
        ws.cell(row=l,column=ci-1).border = Border(n,b,n,n)
        ws.cell(row=l,column=cf+1).border = Border(b,n,n,n)

"""
Retorna a validação de entrada e dos arquivos de configuração
"""
def validacaoEntrada():
    try:
        
        global gv_ie 
        global gv_ano
        global gv_trimestre
        global gv_usuario
        global gv_senha
        global gv_banco
        global gv_caminho
        global gv_relatorio
        

        l_ret = 0
        
        # Carrega os parametros do arquivo .cfg 

        log("-"*150)

        ### Cria os parametros do script 
        comum.addParametro( 'IE', None, 'Inscricao Estadual (IE)', True, '77452443')
        comum.addParametro( 'ANO',  None, "ANO (YYYY)", True, '2015' )
        comum.addParametro( 'TRIMESTRE', None, 'Trimestre - Periodo (1,2,3,4)', True, '1')
        
        # Validacao dos parametros de entrada
        if not comum.validarParametros() :
            l_ret = 91
        
        else:
            # INICIO ELSE
            gv_ie = comum.getParametro('IE').upper().strip()
            gv_ano = comum.getParametro('ANO').upper().strip()
            gv_trimestre = comum.getParametro('TRIMESTRE').upper().strip()

            if not l_ret :
                try:
                    l_iei = re.sub('[^0-9]','',gv_ie)
                    if ( (l_iei == "") or (l_iei == "''") or (l_iei == '""') or (int("0"+l_iei) == 0)):
                        log("PARAMETRO IE : Invalido! " + gv_ie) 
                        l_ret = 91
                except Exception as e:
                    log("PARAMETRO IE : Invalido! " + gv_ie) 
                    l_ret = 91

            if not l_ret :
                try:
                    if (len(gv_ano) != 4):
                        log("PARAMETRO ANO: Invalido! " + gv_ano) 
                        l_ret = 91           
                    else:
                        if (
                        int(gv_ano) > datetime.datetime.now().year
                        or int(gv_ano) < (datetime.datetime.now().year)-50
                        ):
                            log("PARAMETRO ANO : Invalido! " + gv_ano) 
                            l_ret = 91
                except Exception as e:
                    log("PARAMETRO ANO : Invalido! " + gv_ano) 
                    l_ret = 91
            
            if not l_ret :
                try:
                    if (int(gv_trimestre) > 4 or int(gv_trimestre) < 1 ):
                            log("PARAMETRO TRIMESTRE : Invalido! " + gv_trimestre) 
                            l_ret = 91
                except Exception as e:
                    log("PARAMETRO TRIMESTRE : Invalido! " + gv_trimestre) 
                    l_ret = 91
            # FIM ELSE
        
        # Verifica os parametros
        if not l_ret :
            try:

                
                gv_caminho = configuracoes.dir_salvar_arq  #config['caminho'].strip()
                gv_relatorio = configuracoes.nome_relatorio.replace("<<ANO>>",gv_ano).replace("<<IE>>",gv_ie).replace("<<TRIMESTRE>>",gv_trimestre).replace("<<DATA_HORA>>",datetime.datetime.now().strftime('%Y%m%Y%H%M%S')).upper().strip()  #config['relatorio'].replace("<<ANO>>",gv_ano).replace("<<IE>>",gv_ie).replace("<<TRIMESTRE>>",gv_trimestre).upper().strip()
                log('Nome do arquivo a ser gerado........:',gv_relatorio )
                log("Diretorio onde sera salvo o arquivo.:",gv_caminho)

                if not l_ret :
                    if (len(gv_relatorio) <= 5 or not gv_relatorio.endswith(".XLSX")   
                        ):
                        log("PARAMETRO DO ARQUIVO  RELATORIO: INVALIDO! " + gv_relatorio) 
                        l_ret = 91
                    else:
                        log("Arquivo do relatorio : " + gv_relatorio)
                        
                if not l_ret :
                    try:
                        if not os.path.isdir(gv_caminho):
                            log("Diretório não existente : " + gv_caminho)        
                            os.makedirs(gv_caminho)
                            log("Diretório criado : " + gv_caminho)
 
                        gv_relatorio = os.path.join(gv_caminho, gv_relatorio)
                    # victor
                        # if os.path.isfile(gv_relatorio):
                        #     log("Arquivo já existente !!!  " + gv_relatorio)        
                        #     l_ret = 91

                    except Exception as e:
                        txt = traceback.format_exc()
                        log(gv_caminho + " >> PARAMETRO DO RELATÓRIO " + gv_relatorio +  " INVÁLIDO! " + str(e)) 
                        l_ret = 91              
            
            except Exception as e:
                txt = traceback.format_exc()
                log("PARAMETRO DOS ARQUIVOS NAO ENCONTRADOS! " + str(e)) 
                l_ret = 91        

        return l_ret
    except Exception as e:
        txt = traceback.format_exc()
        log("ERRO VALIDAÇÃO DOS PARAMETROS DE ENTRADA: " + str(e))
        l_ret = 93
        return l_ret

def set_border_edsi( ws, cell_range, edsi):
    b=[]
 
    n=Side(border_style= None    , color='00000000')
    f=Side(border_style='thin'   , color='00000000')
    m=Side(border_style='medium' , color='00000000')
    d=Side(border_style='double' , color='00000000')
    p=Side(border_style='dotted' , color='00000000')
    t=Side(border_style='dashed' , color='00000000')
    g=Side(border_style='thick'  , color='00000000')
    
    for x in (0,1,2,3):
        if (edsi[x]=="n"):
            b.append(n)
        elif (edsi[x]=='f'):
            b.append(f)
        elif (edsi[x]=='m'):
            b.append(m)
        elif (edsi[x]=='d'):
            b.append(d)
        elif (edsi[x]=='p'):
            b.append(p)
        elif (edsi[x]=='t'):
            b.append(t)
        elif (edsi[x]=='g'):
            b.append(g)
        else:
            return
    
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = Border(left=b[0],right=b[1],top=b[2],bottom=b[3])


"""
Retorna dados
"""
def retorna_Dados(p_ie,
                    p_data_inicial,
                    p_data_final,
                    p_tipo=0):

    l_retorno=[]
 
    try:

        if str(p_tipo).strip() == "0":
            l_query = """ 
WITH consulado AS
  ( SELECT DISTINCT razao_social, cnpj FROM gfcadastro.tb_clientes_consulados
  )
SELECT
  /*+ parallel(12) */
  cc.razao_social                     AS CONSULADO,
  it.cgc_cpf                          AS CNPJ,
  TO_CHAR(it.infst_dtemiss,'mm/yyyy') AS MES,
  SUM(it.infst_isenta_icms)           AS ICMS_ISENCAO
FROM openrisow.ITEM_NFTL_SERV it,
  consulado cc
WHERE it.EMPS_COD = 'TBRA'
AND it.FILI_COD  IN
  (SELECT fili_cod
  FROM openrisow.filial f
  WHERE f.emps_cod     = 'TBRA'
  AND f.fili_cod_insest= '%s'
  ) 
AND it.infst_serie                     IS NOT NULL
AND it.infst_ind_canc                  <> 'S'
AND it.infst_dtemiss >= TRUNC(TO_DATE('%s', 'DD/MM/YYYY'),'MM')
AND it.infst_dtemiss <= LAST_DAY(TO_DATE('%s', 'DD/MM/YYYY'))
AND it.cgc_cpf                          = cc.cnpj
GROUP BY cc.razao_social,
  it.cgc_cpf,
  TO_CHAR(it.infst_dtemiss,'mm/yyyy')
ORDER BY cc.razao_social,
  it.cgc_cpf,
  TO_CHAR(it.infst_dtemiss,'mm/yyyy')
    """%(p_ie,p_data_inicial,p_data_final)

        else:
            l_query = """ 
WITH consulado AS (SELECT DISTINCT RAZAO_SOCIAL, CNPJ FROM gfcadastro.tb_clientes_consulados )   
SELECT /*+ parallel(12) */
       to_char(it.INFST_DTEMISS,'mm/yyyy') AS Periodo,
       it.CADG_COD,
       cc.razao_social,
       it.CGC_CPF,
       to_char(it.INFST_DTEMISS,'dd/mm/yyyy') AS DTEmissao,
       it.INFST_serie,
       it.INFST_NUM,
       it.SERV_COD,
       it.INFST_DSC_COMPL,
       it.INFST_VAL_CONT,
       it.INFST_BASE_ICMS,
       it.INFST_VAL_ICMS,
       it.INFST_ISENTA_ICMS,
       it.INFST_OUTRAS_ICMS
  FROM openrisow.ITEM_NFTL_SERV it,
       consulado cc
 WHERE it.EMPS_COD       = 'TBRA'
   AND it.FILI_COD       IN (select fili_cod from openrisow.filial f 
                            where f.emps_cod = 'TBRA' 
                            AND f.fili_cod_insest= '%s')
   AND it.INFST_SERIE   IS NOT NULL 
   AND it.INFST_IND_CANC <> 'S'     
   AND it.INFST_dtemiss >= TRUNC(TO_DATE('%s', 'DD/MM/YYYY'),'MM')
   AND it.INFST_dtemiss <= LAST_DAY(TO_DATE('%s', 'DD/MM/YYYY'))
   AND it.CGC_CPF = cc.cnpj
order by it.INFST_DTEMISS, CGC_CPF,  INFST_NUM, INFST_DSC_COMPL
    """%(p_ie,p_data_inicial,p_data_final)
  
        l_cursor = sql.geraCnxBD(configuracoes)
        l_cursor.executa(l_query)
        l_result = l_cursor.fetchone()
        l_lin = 0

        if l_result is not None:
            while l_result:
                #Cria uma linha para informar as colunas
                l_retorno.append([])
                
                # busca as colunas
                for field in l_result:
                    l_retorno[l_lin].append(field)
                
                # + 1
                l_lin = l_lin + 1
                # prox. registro
                l_result = l_cursor.fetchone()
        
       
    
    except Exception as e:
        txt = traceback.format_exc()
           
        try:
            log(" >> FALHA NA GERAÇÃO DOS DADOS ! " + str(txt) + " >> " +  str(e)) 
        except:
            pass
        try:
            l_retorno = None
        except:
            pass

    return l_retorno   

"""
Processamento de relatório
"""
def processarRelatorio(p_ie, 
                       p_ano,
                       p_trimestre,
                       p_relatorio):
    l_ret = 0
    try:
        
        l_dadosSintetico = []
        l_dadosAnalitico = []

        l_mes_01 = "10"
        l_mes_02 = "11"
        l_mes_03 = "12"       

        if int(p_trimestre) == 1:
            l_mes_01 = "01"
            l_mes_02 = "02"
            l_mes_03 = "03"
        elif int(p_trimestre) == 2:
            l_mes_01 = "04"
            l_mes_02 = "05"
            l_mes_03 = "06"        
        elif int(p_trimestre) == 3:
            l_mes_01 = "07"
            l_mes_02 = "08"
            l_mes_03 = "09"  

        l_data_01 = l_mes_01 + "/" + str(p_ano)
        l_data_02 = l_mes_02 + "/" + str(p_ano)
        l_data_03 = l_mes_03 + "/" + str(p_ano)

        if not l_ret:
            l_dadosSintetico = retorna_Dados(p_ie=p_ie,
                                             p_data_inicial="01/" + l_data_01,
                                             p_data_final="01/" + l_data_03)
            if l_dadosSintetico is None:
                log("Não foi encontrado nenhum dados sintetico! (1) ")
                l_ret = 1
            elif len(l_dadosSintetico) <= 1:
                log("Não foi encontrado nenhum dados sintetico! (2) ")
                l_ret = 1
            else:
                log("Quantidade de dados sintetico: " + str(len(l_dadosSintetico)))

        if not l_ret:
            l_dadosAnalitico = retorna_Dados(p_ie=p_ie,
                                            p_data_inicial="01/" + l_data_01,
                                            p_data_final="01/" + l_data_03,
                                            p_tipo=1)
            if l_dadosAnalitico is None:
                log("Não foi encontrado nenhum dados analitico! (1) ")
                l_ret = 1
            elif len(l_dadosAnalitico) <= 1:
                log("Não foi encontrado nenhum dados analitico! (2) ")
                l_ret = 1
            else:
                log("Quantidade de dados analitico: " + str(len(l_dadosAnalitico)))

        if not l_ret:
            log("\n")
            log("Iniciando a criacao do relatório " + str(p_relatorio)) 

            l_nome_planilha0 = "Dados - Trimestral"
            l_nome_padrao = "Dados Mensal <<MES>>_<<ANO>>"
            l_nome_planilha1 = l_nome_padrao.replace("<<MES>>",l_mes_01).replace("<<ANO>>",str(p_ano))
            l_nome_planilha2 = l_nome_padrao.replace("<<MES>>",l_mes_02).replace("<<ANO>>",str(p_ano))
            l_nome_planilha3 = l_nome_padrao.replace("<<MES>>",l_mes_03).replace("<<ANO>>",str(p_ano))
            
            # Criando arquivo principal
            arquivo_excel = Workbook()
            
            # Criando os worksheet
            log("Criando os worksheet : " + l_nome_planilha0
                                        + " >> " + l_nome_planilha1 
                                        + " >> " + l_nome_planilha2
                                        + " >> " + l_nome_planilha3
                                        )  
            planilha0 = arquivo_excel.active
            planilha0.title = l_nome_planilha0
            planilha1 = arquivo_excel.create_sheet(l_nome_planilha1 , 1)
            planilha2 = arquivo_excel.create_sheet(l_nome_planilha2 , 2)
            planilha3 = arquivo_excel.create_sheet(l_nome_planilha3 , 3)

            l_linha  = int(1)
            l_coluna = int(2)
            
            l_linha += 1

            log("Criando cabeçalho ... " + str(l_linha) + " >> " + str(l_coluna))
            planilha0.cell(l_linha,l_coluna,"Resolução 6449 -  (Trimestral)")
            planilha0.cell(l_linha,l_coluna).font=Font(bold=True)
            planilha0.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
            planilha0.cell(l_linha,l_coluna).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
            planilha0.merge_cells('B'+ str(l_linha) + ':E' + str(l_linha))

            planilha1.cell(l_linha,l_coluna,"Resolução 6449 - (Mensal)")
            planilha1.cell(l_linha,l_coluna).font=Font(bold=True)
            planilha1.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
            planilha1.cell(l_linha,l_coluna).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
            planilha1.merge_cells('B'+ str(l_linha) + ':N' + str(l_linha))

            planilha2.cell(l_linha,l_coluna,"Resolução 6449 - (Mensal)")
            planilha2.cell(l_linha,l_coluna).font=Font(bold=True)
            planilha2.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
            planilha2.cell(l_linha,l_coluna).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
            planilha2.merge_cells('B'+ str(l_linha) + ':N' + str(l_linha))

            planilha3.cell(l_linha,l_coluna,"Resolução 6449 - (Mensal)")
            planilha3.cell(l_linha,l_coluna).font=Font(bold=True)
            planilha3.cell(l_linha,l_coluna).alignment = Alignment(horizontal='center')
            planilha3.cell(l_linha,l_coluna).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
            planilha3.merge_cells('B'+ str(l_linha) + ':N' + str(l_linha))

            l_linha += 1

            planilha0.cell(l_linha,l_coluna,"Inscrição Estadual: "+str(p_ie))
            planilha0.cell(l_linha,l_coluna).font=Font(bold=True)
            planilha0.cell(l_linha,l_coluna).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
            planilha0.merge_cells('B'+ str(l_linha) + ':E' + str(l_linha))

            planilha1.cell(l_linha,l_coluna,"Inscrição Estadual: "+str(p_ie))
            planilha1.cell(l_linha,l_coluna).font=Font(bold=True)
            planilha1.cell(l_linha,l_coluna).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
            planilha1.merge_cells('B'+ str(l_linha) + ':N' + str(l_linha))

            planilha2.cell(l_linha,l_coluna,"Inscrição Estadual: "+str(p_ie))
            planilha2.cell(l_linha,l_coluna).font=Font(bold=True)
            planilha2.cell(l_linha,l_coluna).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
            planilha2.merge_cells('B'+ str(l_linha) + ':N' + str(l_linha))

            planilha3.cell(l_linha,l_coluna,"Inscrição Estadual: "+str(p_ie))
            planilha3.cell(l_linha,l_coluna).font=Font(bold=True)
            planilha3.cell(l_linha,l_coluna).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
            planilha3.merge_cells('B'+ str(l_linha) + ':N' + str(l_linha))

            l_linha += 1

            planilha0.cell(l_linha,l_coluna,"Trimestre: "+str(p_trimestre))
            planilha0.cell(l_linha,l_coluna).font=Font(bold=True)
            planilha0.cell(l_linha,l_coluna).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
            planilha0.merge_cells('B'+ str(l_linha) + ':C' + str(l_linha))
# victor
            planilha0.cell(l_linha,l_coluna+2,"Ano: "+str(p_ano))
            planilha0.cell(l_linha,l_coluna+2).font=Font(bold=True)
            planilha0.cell(l_linha,l_coluna+2).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
            planilha0.merge_cells('D'+ str(l_linha) + ':E' + str(l_linha))

            planilha1.cell(l_linha,l_coluna,"Mês/Ano: "+str(l_data_01))
            planilha1.cell(l_linha,l_coluna).font=Font(bold=True)
            planilha1.cell(l_linha,l_coluna).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
            planilha1.merge_cells('B'+ str(l_linha) + ':N' + str(l_linha))

            planilha2.cell(l_linha,l_coluna,"Mês/Ano: "+str(l_data_02))
            planilha2.cell(l_linha,l_coluna).font=Font(bold=True)
            planilha2.cell(l_linha,l_coluna).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
            planilha2.merge_cells('B'+ str(l_linha) + ':N' + str(l_linha))

            planilha3.cell(l_linha,l_coluna,"Mês/Ano: "+str(l_data_03))
            planilha3.cell(l_linha,l_coluna).font=Font(bold=True)
            planilha3.cell(l_linha,l_coluna).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
            planilha3.merge_cells('B'+ str(l_linha) + ':N' + str(l_linha))

            l_linha += 1

            planilha0.freeze_panes = 'A6'
            planilha1.freeze_panes = 'A6'
            planilha2.freeze_panes = 'A6'
            planilha3.freeze_panes = 'A6'

            l_col = l_coluna
            log("Criando as colunas ... " + str(l_linha) + " >> " + str(l_coluna))
            for nColuna in ("Consulado","CNPJ","Mês","ICMS - Isenção"):
                planilha0.cell(l_linha,l_col,nColuna)
                planilha0.cell(l_linha,l_col).font=Font(bold=True)
                planilha0.cell(l_linha,l_col).alignment = Alignment(horizontal='center')
                l_col = l_col + 1

            l_col = l_coluna
            for nColuna in ("Código do Cliente",	
                            "Razão Social",	
                            "CNPJ",	
                            "Data Emissão",	
                            "Série",	
                            "Número NF",
                            "Código Serviço",
                            "Descrição Serviço",	
                            "Valor Total",
                            "Base ICMS",
                            "Valor ICMS",	
                            "Valor Isentas",	
                            "Valor Outras"):
                planilha1.cell(l_linha,l_col,nColuna)
                planilha1.cell(l_linha,l_col).font=Font(bold=True)
                planilha1.cell(l_linha,l_col).alignment = Alignment(horizontal='center')
 
                planilha2.cell(l_linha,l_col,nColuna)
                planilha2.cell(l_linha,l_col).font=Font(bold=True)
                planilha2.cell(l_linha,l_col).alignment = Alignment(horizontal='center')

                planilha3.cell(l_linha,l_col,nColuna)
                planilha3.cell(l_linha,l_col).font=Font(bold=True)
                planilha3.cell(l_linha,l_col).alignment = Alignment(horizontal='center')
                
                l_col = l_col + 1
 
            # Guarda os valores iniciais
            l_col = l_coluna
            l_lin0 = l_linha 
            l_lin1 = l_linha
            l_lin2 = l_linha
            l_lin3 = l_linha
            soma1  = 0.00
            soma2  = 0.00
            soma3  = 0.00

            log("Executando os dados ... " + str(l_linha) + " >> " + str(l_coluna))
            #Dados - 0    
            for linha in l_dadosSintetico:
                l_lin0 += 1
                planilha0.cell(l_lin0,l_col+0,linha[0])
                planilha0.cell(l_lin0,l_col+1,linha[1])
                planilha0.cell(l_lin0,l_col+1).alignment = Alignment(horizontal='center')
                planilha0.cell(l_lin0,l_col+2,linha[2])
                planilha0.cell(l_lin0,l_col+2).alignment = Alignment(horizontal='center')
                planilha0.cell(l_lin0,l_col+3,linha[3])
                planilha0.cell(l_lin0,l_col+3).number_format = "#,##0.00"

                if linha[2] == str(l_data_01):
                    soma1 += linha[3]
                if linha[2] == str(l_data_02):
                    soma2 += linha[3]
                if linha[2] == str(l_data_03):
                    soma3 += linha[3]
         
            for linha in l_dadosAnalitico:
                if linha[0] == l_data_01:
                    l_lin1 += 1    
                    planilha1.cell(l_lin1,l_col+0,linha[1])
                    planilha1.cell(l_lin1,l_col+1,linha[2])
                    planilha1.cell(l_lin1,l_col+2,linha[3])
                    planilha1.cell(l_lin1,l_col+3,linha[4])
                    planilha1.cell(l_lin1,l_col+4,linha[5])
                    planilha1.cell(l_lin1,l_col+5,linha[6])
                    planilha1.cell(l_lin1,l_col+6,linha[7])
                    planilha1.cell(l_lin1,l_col+7,linha[8])
                    planilha1.cell(l_lin1,l_col+8,linha[9])
                    planilha1.cell(l_lin1,l_col+9,linha[10])
                    planilha1.cell(l_lin1,l_col+10,linha[11])
                    planilha1.cell(l_lin1,l_col+11,linha[12])
                    planilha1.cell(l_lin1,l_col+12,linha[13])
                    planilha1.cell(l_lin1,l_col+8).number_format = "#,##0.00"
                    planilha1.cell(l_lin1,l_col+9).number_format = "#,##0.00"
                    planilha1.cell(l_lin1,l_col+10).number_format = "#,##0.00"
                    planilha1.cell(l_lin1,l_col+11).number_format = "#,##0.00"
                    planilha1.cell(l_lin1,l_col+12).number_format = "#,##0.00"
                    
                elif linha[0] == l_data_02:
                    l_lin2 += 1    
                    planilha2.cell(l_lin2,l_col+0,linha[1])
                    planilha2.cell(l_lin2,l_col+1,linha[2])
                    planilha2.cell(l_lin2,l_col+2,linha[3])
                    planilha2.cell(l_lin2,l_col+3,linha[4])
                    planilha2.cell(l_lin2,l_col+4,linha[5])
                    planilha2.cell(l_lin2,l_col+5,linha[6])
                    planilha2.cell(l_lin2,l_col+6,linha[7])
                    planilha2.cell(l_lin2,l_col+7,linha[8])
                    planilha2.cell(l_lin2,l_col+8,linha[9])
                    planilha2.cell(l_lin2,l_col+9,linha[10])
                    planilha2.cell(l_lin2,l_col+10,linha[11])
                    planilha2.cell(l_lin2,l_col+11,linha[12])
                    planilha2.cell(l_lin2,l_col+12,linha[13])
                    planilha2.cell(l_lin2,l_col+8).number_format = "#,##0.00"
                    planilha2.cell(l_lin2,l_col+9).number_format = "#,##0.00"
                    planilha2.cell(l_lin2,l_col+10).number_format = "#,##0.00"
                    planilha2.cell(l_lin2,l_col+11).number_format = "#,##0.00"
                    planilha2.cell(l_lin2,l_col+12).number_format = "#,##0.00"

                elif linha[0] == l_data_03:
                    l_lin3 += 1    
                    planilha3.cell(l_lin3,l_col+0,linha[1])
                    planilha3.cell(l_lin3,l_col+1,linha[2])
                    planilha3.cell(l_lin3,l_col+2,linha[3])
                    planilha3.cell(l_lin3,l_col+3,linha[4])
                    planilha3.cell(l_lin3,l_col+4,linha[5])
                    planilha3.cell(l_lin3,l_col+5,linha[6])
                    planilha3.cell(l_lin3,l_col+6,linha[7])
                    planilha3.cell(l_lin3,l_col+7,linha[8])
                    planilha3.cell(l_lin3,l_col+8,linha[9])
                    planilha3.cell(l_lin3,l_col+9,linha[10])
                    planilha3.cell(l_lin3,l_col+10,linha[11])
                    planilha3.cell(l_lin3,l_col+11,linha[12])
                    planilha3.cell(l_lin3,l_col+12,linha[13])
                    planilha3.cell(l_lin3,l_col+8).number_format = "#,##0.00"
                    planilha3.cell(l_lin3,l_col+9).number_format = "#,##0.00"
                    planilha3.cell(l_lin3,l_col+10).number_format = "#,##0.00"
                    planilha3.cell(l_lin3,l_col+11).number_format = "#,##0.00"
                    planilha3.cell(l_lin3,l_col+12).number_format = "#,##0.00"
                    
            log("Executando os totais ... " + str(l_linha) + " >> " + str(l_coluna))    
            #Totais - 0

            if l_lin0 > l_linha:
                
                l_lin0 += 2

                planilha0.cell(l_lin0,l_col+0,"TOTAL TRIMESTRE:")
                planilha0.cell(l_lin0,l_col+0).font=Font(bold=True)
                planilha0.merge_cells('B'+ str(l_lin0) + ':D' + str(l_lin0))
                planilha0.cell(l_lin0,l_col+0).alignment = Alignment(horizontal='center')
                planilha0.cell(l_lin0,l_col+0).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

                planilha0.cell(l_lin0,l_col+3,"=SUM(E"+str(l_linha+1)+":E"+str(l_lin0-1)+")")
                planilha0.cell(l_lin0,l_col+3).font=Font(bold=True)
                planilha0.cell(l_lin0,l_col+3).number_format = "#,##0.00"
                planilha0.cell(l_lin0,l_col+3).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

                l_lin0 += 2
                
                planilha0.cell(l_lin0,l_col+0,"TOTAIS POR MÊS:")
                planilha0.cell(l_lin0,l_col+0).font=Font(bold=True)
                planilha0.merge_cells('B'+ str(l_lin0) + ':C' + str(l_lin0+2))
                planilha0.cell(l_lin0,l_col+0).alignment = Alignment(horizontal='center', vertical='center')
                planilha0.cell(l_lin0,l_col+0).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

                planilha0.cell(l_lin0,l_col+2,l_data_01)
                planilha0.cell(l_lin0,l_col+2).font=Font(bold=True)
                planilha0.cell(l_lin0,l_col+2).alignment = Alignment(horizontal='center', vertical='center')
                planilha0.cell(l_lin0,l_col+2).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

                planilha0.cell(l_lin0+1,l_col+2,l_data_02)
                planilha0.cell(l_lin0+1,l_col+2).font=Font(bold=True)
                planilha0.cell(l_lin0+1,l_col+2).alignment = Alignment(horizontal='center', vertical='center')
                planilha0.cell(l_lin0+1,l_col+2).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

                planilha0.cell(l_lin0+2,l_col+2,l_data_03)
                planilha0.cell(l_lin0+2,l_col+2).font=Font(bold=True)
                planilha0.cell(l_lin0+2,l_col+2).alignment = Alignment(horizontal='center', vertical='center')
                planilha0.cell(l_lin0+2,l_col+2).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

                planilha0.cell(l_lin0,l_col+3,soma1)
                planilha0.cell(l_lin0,l_col+3).font=Font(bold=True)
                planilha0.cell(l_lin0,l_col+3).number_format = "#,##0.00"
                planilha0.cell(l_lin0,l_col+3).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

                planilha0.cell(l_lin0+1,l_col+3,soma2)
                planilha0.cell(l_lin0+1,l_col+3).font=Font(bold=True)
                planilha0.cell(l_lin0+1,l_col+3).number_format = "#,##0.00"
                planilha0.cell(l_lin0+1,l_col+3).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

                planilha0.cell(l_lin0+2,l_col+3,soma3)
                planilha0.cell(l_lin0+2,l_col+3).font=Font(bold=True)
                planilha0.cell(l_lin0+2,l_col+3).number_format = "#,##0.00"
                planilha0.cell(l_lin0+2,l_col+3).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

            if l_lin1 > l_linha:
                l_lin1 += 2
                planilha1.cell(l_lin1,l_col+0,"TOTAIS:")
                planilha1.cell(l_lin1,l_col+0).font=Font(bold=True)
                planilha1.merge_cells('B'+ str(l_lin1) + ':I' + str(l_lin1))
                planilha1.cell(l_lin1,l_col+0).alignment = Alignment(horizontal='center')
                planilha1.cell(l_lin1,l_col+0).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

                planilha1.cell(l_lin1,l_col+8,"=SUM(J"+str(l_linha+1)+":J"+str(l_lin1-1)+")")
                planilha1.cell(l_lin1,l_col+8).font=Font(bold=True)
                planilha1.cell(l_lin1,l_col+8).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
                planilha1.cell(l_lin1,l_col+8).number_format = "#,##0.00"

                planilha1.cell(l_lin1,l_col+9,"=SUM(K"+str(l_linha+1)+":K"+str(l_lin1-1)+")")
                planilha1.cell(l_lin1,l_col+9).font=Font(bold=True)
                planilha1.cell(l_lin1,l_col+9).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
                planilha1.cell(l_lin1,l_col+9).number_format = "#,##0.00"

                planilha1.cell(l_lin1,l_col+10,"=SUM(L"+str(l_linha+1)+":L"+str(l_lin1-1)+")")
                planilha1.cell(l_lin1,l_col+10).font=Font(bold=True)
                planilha1.cell(l_lin1,l_col+10).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
                planilha1.cell(l_lin1,l_col+10).number_format = "#,##0.00"

                planilha1.cell(l_lin1,l_col+11,"=SUM(M"+str(l_linha+1)+":M"+str(l_lin1-1)+")")
                planilha1.cell(l_lin1,l_col+11).font=Font(bold=True)
                planilha1.cell(l_lin1,l_col+11).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
                planilha1.cell(l_lin1,l_col+11).number_format = "#,##0.00"

                planilha1.cell(l_lin1,l_col+12,"=SUM(N"+str(l_linha+1)+":N"+str(l_lin1-1)+")")
                planilha1.cell(l_lin1,l_col+12).font=Font(bold=True)
                planilha1.cell(l_lin1,l_col+12).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
                planilha1.cell(l_lin1,l_col+12).number_format = "#,##0.00"

            if l_lin2 > l_linha:
                l_lin2 += 2
                planilha2.cell(l_lin2,l_col+0,"TOTAIS:")
                planilha2.cell(l_lin2,l_col+0).font=Font(bold=True)
                planilha2.merge_cells('B'+ str(l_lin2) + ':I' + str(l_lin2))
                planilha2.cell(l_lin2,l_col+0).alignment = Alignment(horizontal='center')
                planilha2.cell(l_lin2,l_col+0).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

                planilha2.cell(l_lin2,l_col+8,"=SUM(J"+str(l_linha+1)+":J"+str(l_lin2-1)+")")
                planilha2.cell(l_lin2,l_col+8).font=Font(bold=True)
                planilha2.cell(l_lin2,l_col+8).number_format = "#,##0.00"

                planilha2.cell(l_lin2,l_col+9,"=SUM(K"+str(l_linha+1)+":K"+str(l_lin2-1)+")")
                planilha2.cell(l_lin2,l_col+9).font=Font(bold=True)
                planilha2.cell(l_lin2,l_col+9).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
                planilha2.cell(l_lin2,l_col+9).number_format = "#,##0.00"

                planilha2.cell(l_lin2,l_col+10,"=SUM(L"+str(l_linha+1)+":L"+str(l_lin2-1)+")")
                planilha2.cell(l_lin2,l_col+10).font=Font(bold=True)
                planilha2.cell(l_lin2,l_col+10).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
                planilha2.cell(l_lin2,l_col+10).number_format = "#,##0.00"

                planilha2.cell(l_lin2,l_col+11,"=SUM(M"+str(l_linha+1)+":M"+str(l_lin2-1)+")")
                planilha2.cell(l_lin2,l_col+11).font=Font(bold=True)
                planilha2.cell(l_lin2,l_col+11).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
                planilha2.cell(l_lin2,l_col+11).number_format = "#,##0.00"

                planilha2.cell(l_lin2,l_col+12,"=SUM(N"+str(l_linha+1)+":N"+str(l_lin2-1)+")")
                planilha2.cell(l_lin2,l_col+12).font=Font(bold=True)
                planilha2.cell(l_lin2,l_col+12).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
                planilha2.cell(l_lin2,l_col+12).number_format = "#,##0.00"

            if l_lin3 > l_linha:
                l_lin3 += 2
                planilha3.cell(l_lin3,l_col+0,"TOTAIS:")
                planilha3.cell(l_lin3,l_col+0).font=Font(bold=True)
                planilha3.merge_cells('B'+ str(l_lin3) + ':I' + str(l_lin3))
                planilha3.cell(l_lin3,l_col+0).alignment = Alignment(horizontal='center')
                planilha3.cell(l_lin3,l_col+0).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")

                planilha3.cell(l_lin3,l_col+8,"=SUM(J"+str(l_linha+1)+":J"+str(l_lin3-1)+")")
                planilha3.cell(l_lin3,l_col+8).font=Font(bold=True)
                planilha3.cell(l_lin3,l_col+8).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
                planilha3.cell(l_lin3,l_col+8).number_format = "#,##0.00"

                planilha3.cell(l_lin3,l_col+9,"=SUM(K"+str(l_linha+1)+":K"+str(l_lin3-1)+")")
                planilha3.cell(l_lin3,l_col+9).font=Font(bold=True)
                planilha3.cell(l_lin3,l_col+9).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
                planilha3.cell(l_lin3,l_col+9).number_format = "#,##0.00"

                planilha3.cell(l_lin3,l_col+10,"=SUM(L"+str(l_linha+1)+":L"+str(l_lin3-1)+")")
                planilha3.cell(l_lin3,l_col+10).font=Font(bold=True)
                planilha3.cell(l_lin3,l_col+10).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
                planilha3.cell(l_lin3,l_col+10).number_format = "#,##0.00"

                planilha3.cell(l_lin3,l_col+11,"=SUM(M"+str(l_linha+1)+":M"+str(l_lin3-1)+")")
                planilha3.cell(l_lin3,l_col+11).font=Font(bold=True)
                planilha3.cell(l_lin3,l_col+11).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
                planilha3.cell(l_lin3,l_col+11).number_format = "#,##0.00"

                planilha3.cell(l_lin3,l_col+12,"=SUM(N"+str(l_linha+1)+":N"+str(l_lin3-1)+")")
                planilha3.cell(l_lin3,l_col+12).font=Font(bold=True)
                planilha3.cell(l_lin3,l_col+12).fill = PatternFill(start_color="e1ecf4", end_color="e1ecf4", fill_type = "solid")
                planilha3.cell(l_lin3,l_col+12).number_format = "#,##0.00"

#BORDAS    
#BORDAS    
#BORDAS    
            set_border_edsi(planilha0, 'B6:E'+str(planilha0.max_row-6), 'gggg')

            set_border_edsi(planilha1, 'B6:N'+str(planilha1.max_row), 'ffff')
            set_border_edsi(planilha1, 'B'+str(planilha1.max_row)+':N'+str(planilha1.max_row), 'gggg')

            set_border_edsi(planilha2, 'B6:N'+str(planilha2.max_row), 'ffff')
            set_border_edsi(planilha2, 'B'+str(planilha2.max_row)+':N'+str(planilha2.max_row), 'gggg')

            set_border_edsi(planilha3, 'B6:N'+str(planilha3.max_row), 'ffff')
            set_border_edsi(planilha3, 'B'+str(planilha3.max_row)+':N'+str(planilha3.max_row), 'gggg')

            log("Aumentando as colunas ... ")

            adjust_column(planilha0,6,2,5)
            adjust_column(planilha1,6,2,14)
            adjust_column(planilha2,6,2,14)
            adjust_column(planilha3,6,2,14)
            

#MERGE CELULAS PLANILHA 0                            
#MERGE CELULAS PLANILHA 0                            
#MERGE CELULAS PLANILHA 0                            
            
            qtd      = 0
            anterior = planilha0.cell(5,3).value
            for linha in range(5,planilha0.max_row):
 
                if planilha0.cell(linha,3).value == "":
                    continue
                if anterior == planilha0.cell(linha,3).value:
                    qtd += 1
                else:
                    if qtd > 1: 
                        m1 = (linha) - qtd      
                        planilha0.merge_cells('B'+ str(m1) + ':B' + str(linha - 1))
                        
                        set_border_edsi(planilha0, 'C'+ str(m1) +':C'+ str(linha - 1), 'ffgg')

                        # UMA LINHA
                        if linha - 1 == m1: 
                            set_border_edsi(planilha0, 'D'+ str(m1+1) +':D'+ str(m1+1), 'ffgg')
                            set_border_edsi(planilha0, 'E'+ str(m1) +':E'+ str(m1), 'fggg')

                        # DUAS LINHA
                        if (linha - 1) - m1  == 1:
                            set_border_edsi(planilha0, 'D'+ str(m1) +':D'+ str(m1), 'ffgf')
                            set_border_edsi(planilha0, 'D'+ str(m1+1) +':D'+ str(m1+1), 'fffg')

                            set_border_edsi(planilha0, 'E'+ str(m1) +':E'+ str(m1), 'fggf')
                            set_border_edsi(planilha0, 'E'+ str(m1+1) +':E'+ str(m1+1), 'fgfg')
                        

                        # MAIS Q DUAS LINHA
                        if (linha - 1) - m1  > 1:
                            set_border_edsi(planilha0, 'D'+ str(m1) +':D'+ str(m1), 'ffgf')
                            set_border_edsi(planilha0, 'E'+ str(m1) +':E'+ str(m1), 'fggf')


                            for x in range(m1+1, (linha - 1)):
                                set_border_edsi(planilha0, 'D'+ str(x) +':D'+ str(x), 'ffff')
                                set_border_edsi(planilha0, 'E'+ str(x) +':E'+ str(x), 'fgff')


                            set_border_edsi(planilha0, 'D'+ str(linha - 1) +':D'+ str(linha - 1), 'fffg')
                            set_border_edsi(planilha0, 'E'+ str(linha - 1) +':E'+ str(linha - 1), 'fgfg')
                        
                        
                        planilha0.cell(m1, 2).alignment = Alignment(vertical='center')
                        
                        planilha0.merge_cells('C'+ str(m1) + ':C' + str(linha - 1))
                        
                        planilha0.cell(m1, 3).alignment = Alignment(vertical='center')

                        qtd = 1

                anterior = planilha0.cell(linha,3).value
            
            contorno_cli_clf(planilha0,2,2,5,4,'g')
            contorno_cli_clf(planilha1,2,2,14,4,'g')
            contorno_cli_clf(planilha2,2,2,14,4,'g')
            contorno_cli_clf(planilha3,2,2,14,4,'g')
            
            set_border_edsi(planilha0, 'B5:E5', 'gggg')

            contorno_cli_clf(planilha1,2,6,14,planilha1.max_row - 2,'g')
            contorno_cli_clf(planilha2,2,6,14,planilha2.max_row - 2,'g')
            contorno_cli_clf(planilha3,2,6,14,planilha3.max_row - 2,'g')

            set_border_edsi(planilha1, 'B5:N5', 'gggg')
            set_border_edsi(planilha2, 'B5:N5', 'gggg')
            set_border_edsi(planilha3, 'B5:N5', 'gggg')

            set_border_edsi(planilha0, 'B'+str(planilha0.max_row - 4)+':E'+str(planilha0.max_row - 4), 'gggg')
            set_border_edsi(planilha0, 'B'+str(planilha0.max_row - 2)+':E'+str(planilha0.max_row), 'gggg')

            
            # Grava a planilha Excel
            log("Salvando a planilha : " + str(p_relatorio))
            arquivo_excel.save(p_relatorio)
     

    except Exception as e:
        l_ret = 1    
        txt = traceback.format_exc()
        log("ERRO NA GERAÇÃO DO RELATÓRIO.: " + txt +  " >> " +  str(e))

    return l_ret

# Ponto de partida
if __name__ == "__main__" :
    
    # Codigo de Retorno
    ret = 0

    # Tratamento de excessao
    txt = ''

    # Tratamento de variaveis globais

    gv_ie = ""
    gv_ano = ""
    gv_trimestre = ""

    gv_caminho=""
    gv_relatorio=""

    try:

        log("\n")
        log("Carregando as variaveis de configuracao")
        variaveis = comum.carregaConfiguracoes(configuracoes)
        
        
        # Validacao dos parametros de entrada
        if not ret :
            ret = validacaoEntrada()
        
        log("\n")

        
        # Validar parametros que envolve o banco
        if not ret :
            l_str_uf = retornar_UF(p_IE=gv_ie) 
            if l_str_uf is None:
                log("PARAMETRO IE não encontrado! <1> " + str(l_str_uf)) 
                ret = 91
            elif (len(str(l_str_uf)) == 0 
                or str(l_str_uf).strip().upper() == "NONE"):
                log("PARAMETRO IE não encontrado ! <2> " + str(l_str_uf) ) 
                ret = 91

        log("\n")

        # Processar o relatório
        if not ret :
            ret = processarRelatorio(p_ie=gv_ie, 
                                     p_ano=gv_ano,
                                     p_trimestre=gv_trimestre,
                                     p_relatorio=gv_relatorio)        

        log("\n")

       
        # Finalizacao
        log("\n")            
        
        if not ret :
            log("***SUCESSO***")
        else:
            log("<<<ERRO>>>")

        log("\n")
    
    except Exception as e:
        txt = traceback.format_exc()
        log("ERRO .: " + str(e))
        ret = 93
    
    sys.exit(ret if ret >= log.ret else log.ret )
