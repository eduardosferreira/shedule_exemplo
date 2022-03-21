#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-

"""
----------------------------------------------------------------------------------------------
  SISTEMA ..: 
  MODULO ...: 
  SCRIPT ...: Relatório Insumo_SPED_Fiscal
  CRIACAO ..: 16/07/2021
  AUTOR ....: Airton Borges da Silva Filho - Kyros Consultoria 
  DESCRICAO : Relatório de análises dos insumos SPED
                

----------------------------------------------------------------------------------------------
    Exemplo de comando: ./insumo_SPED.py SP 102020 108383949112
    Diretório saída ..: /arquivos/RELATORIOS/SPED/<UF>/<AAAA>/<MM>/
    Arquivo saída ....: Insumo_Consolidado_SPED_Fiscal_<MESANO>_<UF>_<IE>_V<NNN>.xlsx
    Documentação......: CLONE1 /arquivos/TESHUVA/melhorias/XX - Fase 2 - Relatorio SPED Mercadoria/Teshuva_RMSV0_Insumo_SPED_FISCAL_V2.docx
----------------------------------------------------------------------------------------------


  HISTORICO : 

    27/09/2021 - Airton Borges - Kyros - Conversão para o novo padrão do Painel de execusões

  Detalhes úteis:

https://www.letscode.com.br/blog/aprenda-a-integrar-python-e-excel
-cd /arquivos/TESHUVA/scripts_rpa/insumo_SPED/
-echo "" > insumo_SPED.py & nano insumo_SPED.py
-./insumo_SPED.py rj 122016 77452443 
----------------------------------------------------------------------------------------------
"""


#### PATRONIZACAO PARA O PAINEL DE EXECUCOES....
import sys
import os
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)
import configuracoes
import util
import comum
import sql
comum.log.gerar_log_em_arquivo = False
comum.carregaConfiguracoes(configuracoes)
banco=sql.geraCnxBD(configuracoes)
#### PATRONIZACAO PARA O PAINEL DE EXECUCOES....


import datetime
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN, BORDER_MEDIUM, BORDER_DOUBLE
from openpyxl.reader.excel import load_workbook 
from openpyxl.utils import get_column_letter

sys.path.append( os.path.join( os.path.realpath('..'), "modulosPython" ) )
nome_script = os.path.basename( sys.argv[0] ).replace('.py', '')


global variaveis
global db
global listadeabas
global qtdabas


DEBUG = False
#DEBUG = True
listadeabas = ('01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21')
#listadeabas = ('09')


qtdabas = 21

fontMasterPreta  = Font(color='00000000', bold=True, size=12)
fontNegrito      = Font(color='00000000', bold=True)

#config = carregaConfiguracoes() 
#variaveis['config'] = config
#config = variaveis['config']

#config = carregaConfiguracoes()
#variaveis['config'] = config
#config = variaveis['config']

#uid = "gfcadastro"  
#pwd = "vivo2019" 
#db = config['banco']
 
ret = 0

def encodingDoArquivo(path_arq) :
    global ret
    
    try :
        fd = open(path_arq, 'r', encoding='iso-8859-1')
        fd.read()
        fd.close()
    except :
        return 'utf-8'

    return 'iso-8859-1'

def proximo_arquivo(mascara,diretorio):
    qdade = 0
    nomearq = "" 
    directory = Path(diretorio)
    files = directory.glob(mascara)
    sorted_files = sorted(files, reverse=False)    
   
    nomearq = mascara.replace("*", "V000")
    proximo = "001"
    
    if sorted_files:
        for f in sorted_files:
            qdade = qdade + 1
            nomearq = f
        proximo = '{:03d}'.format(int((str(nomearq).split(".")[0]).split("_")[-1][1:]) + 1)
    nomearq = mascara.replace("*", "V"+proximo)
    return(nomearq)

def nome_arquivo(mascara,diretorio):
    qdade = 0
    nomearq = "" 
    directory = Path(diretorio)
    files = directory.glob(mascara)
    sorted_files = sorted(files, reverse=False)
    if sorted_files:
        for f in sorted_files:
            qdade = qdade + 1
            nomearq = f
#    else: 
#        log('#### Erro encontrado : Arquivo %s não está na pasta %s'%(mascara,diretorio))
    return(nomearq)

def semespacos(frase):
    retorno = "" 
    for l in frase:
        if (l != " "):
            retorno = retorno + l
    return(retorno)


def processar():
    ret = 0
    ufi = ""
    mesanoi = ""
    mesi = ""
    anoi = "" 
    global listadeabas
    try :
        if (len(sys.argv) == 4 and len(sys.argv[1])==2 and len(sys.argv[2])==6 ) :
            iei = "*"
            ufi = sys.argv[1].upper()
            mesanoi = sys.argv[2].upper()
            mesi = sys.argv[2][:2].upper()
            anoi = sys.argv[2][2:].upper()
            iei = sys.argv[3].upper()
            iei = re.sub('[^0-9]','',iei)
        else :
            ret = 99
    except Exception as e :
        log('#### Erro encontrado :', e)
        ret = 99
    if ret == 99:
        log("-" * 100)
        log("#### ")
        log('#### ERRO - Erro nos parametros do script.')
        log("#### ")
        log('#### Exemplo de como deve ser :')
        log('####      %s <UF> <MMAAAA> <IE>'%(sys.argv[0]))
        log("#### ")
        log('#### Onde')
        log('####      <UF> = estado. Ex: SP')
        log('####      <MMAAAA> = mês e ano. Ex: Para junho de 2020 informe 062020')
        log('####      <IE> = inscrição desejada. Ex: 108383949112')
        log("#### ")
        log('#### Portanto, se o estado = SP, o mes = 06 e o ano = 2020, e IE = 108383949112, o comando correto deve ser :')  
        log('####      %s SP 062020 108383949112'%(sys.argv[0]))  
        log("#### ")
        log('#### Neste exemplo, a planilha resultado será:') 
        log('####      %sarquivos%sRELATORIOS%sMENSAIS%s2020%s06%sAnalise_SPED_Convenio115_GIA_062020_SP_108383949112.xlsx'%(SD,SD,SD,SD,SD,SD)) 
        log('#### ')
        log("-" * 100)
        log ("")
        log ("Retorno = 99")
        return(ret)
    
    #SETA O ARQUIVO DE SAIDA INSUMOS (PLANILHA EXCEL)
    dir_rel_INSUMOS     = os.path.join(SD,'arquivos','RELATORIOS','SPED',ufi,anoi,mesi)
    mask_INSUMOS        = "Insumo_Consolidado_SPED_Fiscal_"+mesanoi+"_"+ufi+"_"+iei+"_*.xlsx"  
    rel_INSUMOS         = proximo_arquivo(mask_INSUMOS,dir_rel_INSUMOS)
    ARQ_INSUMOS         = os.path.join(dir_rel_INSUMOS,rel_INSUMOS)

   #SETA O ARQUIVO DE MERCADORIA ENTRADA (PLANILHA EXCEL)
    dir_rel_ME  = os.path.join(SD,'arquivos','RELATORIOS','SPED',ufi,anoi,mesi)
    mask_ME     = "Mercadoria_Entrada_Mestre_Item_"+mesanoi+"_"+ufi+"_"+iei+"_*.xlsx"  
    rel_ME      = proximo_arquivo(mask_ME,dir_rel_ME)
    ARQ_ME      = os.path.join(dir_rel_ME,rel_ME)

   #SETA O ARQUIVO DE MERCADORIA SAIDA (PLANILHA EXCEL)
    dir_rel_MS  = os.path.join(SD,'arquivos','RELATORIOS','SPED',ufi,anoi,mesi)
    mask_MS     = "Mercadoria_Saida_Mestre_Item_"+mesanoi+"_"+ufi+"_"+iei+"_*.xlsx"  
    rel_MS      = proximo_arquivo(mask_MS,dir_rel_MS)
    ARQ_MS      = os.path.join(dir_rel_MS,rel_MS)

   #SELECIONA O ARQUIVO PROTOCOLADO ARQ_PRO
    dir_rel_SPED_PRO    = os.path.join(SD,'arquivos','SPED_FISCAL','PROTOCOLADOS',ufi,anoi, mesi)
    mask_SPED_PRO       = "SPED_"+mesanoi+"_"+ufi+"_"+iei+"_PROT_"+"*.txt"   
    rel_PRO             = nome_arquivo(mask_SPED_PRO,dir_rel_SPED_PRO)
    if (rel_PRO == ""):
        log('#### ERRO - Arquivo PROTOCOLADO ', mask_SPED_PRO, ' não encontrado na pasta: ', dir_rel_SPED_PRO )
        return(99)
    ARQ_PRO             = os.path.join(dir_rel_SPED_PRO,rel_PRO)
    
    #SELECIONA O ARQUIVO ENXERTADO ARQ_ENX    
    dir_rel_SPED_ENX    = os.path.join(SD,'arquivos','SPED_FISCAL','ENXERTADOS',ufi,anoi, mesi)
    mask_SPED_ENX       = "SPED_"+mesanoi+"_"+ufi+"_"+iei+"_ENX_"+"*.txt"   
    rel_ENX             = nome_arquivo(mask_SPED_ENX,dir_rel_SPED_ENX)
    if (rel_ENX == ""):
        log('#### ERRO - Arquivo ENXERTADO ', mask_SPED_ENX, ' não encontrado na pasta: ', dir_rel_SPED_ENX )
        return(99)
    ARQ_ENX             = os.path.join(dir_rel_SPED_ENX,rel_ENX)    
    
      
    log("-" * 100)    
    log(" ====> INSUMO = ",ARQ_INSUMOS)  
    log(" ====> ME     = ",ARQ_ME)  
    log(" ====> MS     = ",ARQ_MS)  
    log("-" * 100)    
  
    #### Se a pasta do relatório INSUMO não existir, cria
    if not os.path.isdir(dir_rel_INSUMOS) :
        os.makedirs(dir_rel_INSUMOS)
   
    #### Se a pasta do relatório MERCADORIA ENTRADA não existir, cria
    if not os.path.isdir(dir_rel_ME) :
        os.makedirs(dir_rel_ME)

    #### Se a pasta do relatório MERCADORIA ENTRADA não existir, cria
    if not os.path.isdir(dir_rel_MS) :
        os.makedirs(dir_rel_MS)

    #### Cria a planilha excel INSUMO em memória....
    arquivo_excel = Workbook()
    ABA_EP       = arquivo_excel.active
    ABA_EP.title =                            "ENTRADAS - PROTOCOLADO"            # 0)
    if ('02' in listadeabas):
        ABA_ER       = arquivo_excel.create_sheet("ENTRADAS - REGERADO"               , 1)
    if ('03' in listadeabas):
        ABA_SP       = arquivo_excel.create_sheet("SAIDAS - PROTOCOLADO"              , 2)
    if ('04' in listadeabas):
        ABA_SR       = arquivo_excel.create_sheet("SAIDAS - REGERADO"                 , 3)
    if ('05' in listadeabas):
        ABA_TP       = arquivo_excel.create_sheet("TELECOM - PROTOCOLADO"             , 4)
    if ('06' in listadeabas):
        ABA_TR       = arquivo_excel.create_sheet("TELECOM - REGERADO"                , 5)
    if ('07' in listadeabas):
        ABA_TRS      = arquivo_excel.create_sheet("TELECOM - POR SERIE - REGERADO"    , 6)
    if ('08' in listadeabas):
        ABA_TPS      = arquivo_excel.create_sheet("TELECOM-POR SERIE-PROTOCOLADO"     , 7)
    if ('09' in listadeabas):
        ABA_TRT      = arquivo_excel.create_sheet("TELECOM - REGERADO - TRIB ICMS"    , 8)
    if ('10' in listadeabas):
        ABA_RCFOP    = arquivo_excel.create_sheet("RESUMO CFOP"                       , 9)
    if ('11' in listadeabas):
        ABA_AICMS    = arquivo_excel.create_sheet("AJUSTES ICMS"                      ,10)
    if ('12' in listadeabas):
        ABA_SN       = arquivo_excel.create_sheet("SALTO_NOTAS"                       ,11)
    if ('13' in listadeabas):
        ABA_R1600    = arquivo_excel.create_sheet("REGISTRO_1600"                     ,12)
    if ('14' in listadeabas):
        ABA_HME      = arquivo_excel.create_sheet("HISTORICO - MESTRE - ENTRADA"      ,13)
    if ('15' in listadeabas):
        ABA_HIE      = arquivo_excel.create_sheet("HISTORICO - ITENS - ENTRADA"       ,14)
    if ('16' in listadeabas):
        ABA_HMS      = arquivo_excel.create_sheet("HISTORICO - MESTRE - SAIDA"        ,15)
    if ('17' in listadeabas):
        ABA_HIS      = arquivo_excel.create_sheet("HISTORICO - ITENS - SAIDA"         ,16)


    ##########  ABA 1-   ABA_EP = ENTRADAS PROTOCOLADO 0
    ##########  ABA 1-   ABA_EP = ENTRADAS PROTOCOLADO 0
    ##########  ABA 1-   ABA_EP = ENTRADAS PROTOCOLADO 0 
    ##########  ABA 1-   ABA_EP = ENTRADAS PROTOCOLADO 0
    ##########  ABA 1-   ABA_EP = ENTRADAS PROTOCOLADO 0 
    ##########  ABA 1-   ABA_EP = ENTRADAS PROTOCOLADO 0
    if ('01' in listadeabas):
        log("ABA 01 / ", qtdabas , " -> ENTRADAS - PROTOCOLADO")
        
        encP = encodingDoArquivo(ARQ_PRO)
        ARQPRO = open(ARQ_PRO, 'r', encoding=encP, errors='ignore')
        linhaABA_EP = {}
        chaves = []
        
        for linhaP in ARQPRO:

            #log("##### DEBUG ##### => linhaP = ",linhaP)

            camposlinhaP = linhaP.split("|")
            if (len(camposlinhaP) > 3 ):
              
                if ( (camposlinhaP[1].upper() == "C190" and camposlinhaP[3][0] in ("1", "2", "3")) or (camposlinhaP[1].upper() in ("C590", "D190", "D590"))):

                    if ( (camposlinhaP[1].upper() == "D190")):
                        camposlinhaP.append("")
                        camposlinhaP[11] == ""

                    
                    #log("##### DEBUG ##### => agrupamento = ",camposlinhaP[3].upper()+"x"+camposlinhaP[2].upper()+"x"+camposlinhaP[4].upper())
                    
                    agrupamento = camposlinhaP[3].upper()+"x"+camposlinhaP[2].upper()+"x"+camposlinhaP[4].upper()


                    
                    if (agrupamento in chaves):
                        #log(" = ",)
                        #log("Entrou no if = ",)
                        #log("linhaABA_EP = ", linhaABA_EP)
                        #log("Agrupamento = ", agrupamento)
                        #log("chaves      = ", chaves)
                        #input("Continua? ")
                        linhaABA_EP[agrupamento] = (linhaABA_EP[agrupamento][0],
                                                    linhaABA_EP[agrupamento][1],
                                                    linhaABA_EP[agrupamento][2],
                                                    linhaABA_EP[agrupamento][3]  + (0.00 if (camposlinhaP[5]  == "") else float(camposlinhaP[5].replace("," , "."))),
                                                    linhaABA_EP[agrupamento][4]  + (0.00 if (camposlinhaP[6]  == "") else float(camposlinhaP[6].replace("," , "."))),
                                                    linhaABA_EP[agrupamento][5]  + (0.00 if (camposlinhaP[7]  == "") else float(camposlinhaP[7].replace("," , "."))),
                                                    linhaABA_EP[agrupamento][6]  + (0.00 if (camposlinhaP[8]  == "") else float(camposlinhaP[8].replace("," , "."))),
                                                    linhaABA_EP[agrupamento][7]  + (0.00 if (camposlinhaP[9]  == "") else float(camposlinhaP[9].replace("," , "."))),
                                                    linhaABA_EP[agrupamento][8]  + (0.00 if (camposlinhaP[11]  == "") else float(camposlinhaP[11].replace("," , "."))))
                        #log("saiu do if",)
                        #log(" = ",)
                                                    
                    else:
                        #log(" = ",)
                        #log("entrou no else = ",)

                        #if (agrupamento == "2354x000x0"):
                            #log("##### DEBUG ##### => ARQPRO = ",ARQPRO)
                            #log("##### DEBUG ##### => ",camposlinhaP[1]," # ",camposlinhaP[2]," # ",camposlinhaP[3]," # ",camposlinhaP[4]," # ",camposlinhaP[5]," # ",camposlinhaP[6]," # ",camposlinhaP[7]," # ",camposlinhaP[8]," # ",)
                            #log("##### DEBUG ##### => QUANTIDADE DE REGISTROS = ",len(camposlinhaP))
                            #input("Vai dar o erro...")
                            #log("##### DEBUG ##### => ",camposlinhaP[1]," # ",camposlinhaP[2]," # ",camposlinhaP[3]," # ",camposlinhaP[4]," # ",camposlinhaP[5]," # ",camposlinhaP[6]," # ",camposlinhaP[7]," # ",camposlinhaP[8]," # ",camposlinhaP[9]," # ",camposlinhaP[11]," # ",)
                        linhaABA_EP[agrupamento] = (camposlinhaP[2],
                                                    camposlinhaP[3],
                                                    camposlinhaP[4],
                                                    (0.00 if (camposlinhaP[5]  == "") else float(camposlinhaP[5].replace("," , "."))),
                                                    (0.00 if (camposlinhaP[6]  == "") else float(camposlinhaP[6].replace("," , "."))),
                                                    (0.00 if (camposlinhaP[7]  == "") else float(camposlinhaP[7].replace("," , "."))),
                                                    (0.00 if (camposlinhaP[8]  == "") else float(camposlinhaP[8].replace("," , "."))),
                                                    (0.00 if (camposlinhaP[9]  == "") else float(camposlinhaP[9].replace("," , "."))),
                                                    (0.00 if (camposlinhaP[11] == "") else float(camposlinhaP[11].replace("," , "."))))
                        chaves.append(agrupamento)
                        #log("saiu else = ",)
                        #log(" = ",)
        ARQPRO.close()    
            
            
        l=1
        ABA_EP.cell(row=l, column=1,  value="Resumo por CST / CFOP - Entradas (PROTOCOLADO)")
        l=l+1
        ABA_EP.cell(row=l, column=1,  value="CST/CFOP")
        ABA_EP.cell(row=l, column=2,  value="CFOP")
        ABA_EP.cell(row=l, column=3,  value="ALÍQUOTA ICMS")
        ABA_EP.cell(row=l, column=4,  value="VLR_OPER")
        ABA_EP.cell(row=l, column=5,  value="VLR_BASE_ICMS")
        ABA_EP.cell(row=l, column=6,  value="VLR_ICMS")
        ABA_EP.cell(row=l, column=7,  value="VLR_BASE_ICMS_ST")
        ABA_EP.cell(row=l, column=8,  value="VLR_ICMS_ST")
        ABA_EP.cell(row=l, column=9,  value="VLR_IPI")
        
        l=l+1
        
        chaves.sort()
    
        for chave in chaves:        
    
            ABA_EP.cell(row=l, column=1,  value=linhaABA_EP[chave][0])
            ABA_EP.cell(row=l, column=2,  value=linhaABA_EP[chave][1])
            ABA_EP.cell(row=l, column=3,  value=linhaABA_EP[chave][2])
            ABA_EP.cell(row=l, column=4,  value=linhaABA_EP[chave][3])
            ABA_EP.cell(row=l, column=5,  value=linhaABA_EP[chave][4])
            ABA_EP.cell(row=l, column=6,  value=linhaABA_EP[chave][5])
            ABA_EP.cell(row=l, column=7,  value=linhaABA_EP[chave][6])
            ABA_EP.cell(row=l, column=8,  value=linhaABA_EP[chave][7])
            ABA_EP.cell(row=l, column=9,  value=linhaABA_EP[chave][8])
            
            l=l+1
    
    
    
        somad="=SUM(D3:D" + str(l-1) + ")"
        somae="=SUM(E3:E" + str(l-1) + ")"
        somaf="=SUM(F3:F" + str(l-1) + ")"
        somag="=SUM(G3:G" + str(l-1) + ")"
        somah="=SUM(H3:H" + str(l-1) + ")"
        somai="=SUM(I3:I" + str(l-1) + ")"
            
        ABA_EP.cell(row=l, column=1,  value="TOTAIS")
        ABA_EP.cell(row=l, column=4,  value=somad)
        ABA_EP.cell(row=l, column=5,  value=somae)
        ABA_EP.cell(row=l, column=6,  value=somaf)
        ABA_EP.cell(row=l, column=7,  value=somag)
        ABA_EP.cell(row=l, column=8,  value=somah)
        ABA_EP.cell(row=l, column=9,  value=somai)
    
        formata_EP(ABA_EP)
    
    
        arquivo_excel.save(ARQ_INSUMOS)
            

    ##########  ABA 2-   ABA_ER = ENTRADAS REGERADO 1
    ##########  ABA 2-   ABA_ER = ENTRADAS REGERADO 1
    ##########  ABA 2-   ABA_ER = ENTRADAS REGERADO 1
    ##########  ABA 2-   ABA_ER = ENTRADAS REGERADO 1
    ##########  ABA 2-   ABA_ER = ENTRADAS REGERADO 1
    ##########  ABA 2-   ABA_ER = ENTRADAS REGERADO 1
    if ('02' in listadeabas):
        log("ABA 02 / ", qtdabas , " -> ENTRADAS - REGERADO")
                
    
        encE = encodingDoArquivo(ARQ_ENX)
        ARQENX = open(ARQ_ENX, 'r', encoding=encE, errors='ignore')
        linhaABA_ER = {}
        chaves = []
        
        for linhaE in ARQENX:
            camposlinhaE = linhaE.split("|")
            if (len(camposlinhaE) > 3 ):
              
                if ( (camposlinhaE[1].upper() == "C190" and camposlinhaE[3][0] in ("1", "2", "3")) or (camposlinhaE[1].upper() in ("C590", "D190", "D590"))):


                    if ( (camposlinhaE[1].upper() == "D190")):
                        camposlinhaE.append("")
                        camposlinhaE[11] == ""




                    agrupamento = camposlinhaE[3].upper()+"x"+camposlinhaE[2].upper()+"x"+camposlinhaE[4].upper()
                    
                    if (agrupamento in chaves):
                        linhaABA_ER[agrupamento] = (linhaABA_ER[agrupamento][0],
                                                    linhaABA_ER[agrupamento][1],
                                                    linhaABA_ER[agrupamento][2],
                                                    linhaABA_ER[agrupamento][3]  + (0.00 if (camposlinhaE[5]   == "") else float(camposlinhaE[5].replace("," , "."))),
                                                    linhaABA_ER[agrupamento][4]  + (0.00 if (camposlinhaE[6]   == "") else float(camposlinhaE[6].replace("," , "."))),
                                                    linhaABA_ER[agrupamento][5]  + (0.00 if (camposlinhaE[7]   == "") else float(camposlinhaE[7].replace("," , "."))),
                                                    linhaABA_ER[agrupamento][6]  + (0.00 if (camposlinhaE[8]   == "") else float(camposlinhaE[8].replace("," , "."))),
                                                    linhaABA_ER[agrupamento][7]  + (0.00 if (camposlinhaE[9]   == "") else float(camposlinhaE[9].replace("," , "."))),
                                                    linhaABA_ER[agrupamento][8]  + (0.00 if (camposlinhaE[11]  == "") else float(camposlinhaE[11].replace("," , "."))))
                    else:
                        linhaABA_ER[agrupamento] = (camposlinhaE[2],
                                                    camposlinhaE[3],
                                                    camposlinhaE[4],
                                                    (0.00 if (camposlinhaE[5]  == "") else float(camposlinhaE[5].replace("," , "."))),
                                                    (0.00 if (camposlinhaE[6]  == "") else float(camposlinhaE[6].replace("," , "."))),
                                                    (0.00 if (camposlinhaE[7]  == "") else float(camposlinhaE[7].replace("," , "."))),
                                                    (0.00 if (camposlinhaE[8]  == "") else float(camposlinhaE[8].replace("," , "."))),
                                                    (0.00 if (camposlinhaE[9]  == "") else float(camposlinhaE[9].replace("," , "."))),
                                                    (0.00 if (camposlinhaE[11] == "") else float(camposlinhaE[11].replace("," , "."))))
                        chaves.append(agrupamento)
            
        ARQENX.close()
        l=1
        ABA_ER.cell(row=l, column=1,  value="Resumo por CST / CFOP - Entradas (REGERADO)")
        l=l+1
        ABA_ER.cell(row=l, column=1,  value="CST/CFOP")
        ABA_ER.cell(row=l, column=2,  value="CFOP")
        ABA_ER.cell(row=l, column=3,  value="ALÍQUOTA ICMS")
        ABA_ER.cell(row=l, column=4,  value="VLR_OPER")
        ABA_ER.cell(row=l, column=5,  value="VLR_BASE_ICMS")
        ABA_ER.cell(row=l, column=6,  value="VLR_ICMS")
        ABA_ER.cell(row=l, column=7,  value="VLR_BASE_ICMS_ST")
        ABA_ER.cell(row=l, column=8,  value="VLR_ICMS_ST")
        ABA_ER.cell(row=l, column=9,  value="VLR_IPI")
        
        l=l+1
        
        chaves.sort()
    
        for chave in chaves:        
    
            ABA_ER.cell(row=l, column=1,  value=linhaABA_ER[chave][0])
            ABA_ER.cell(row=l, column=2,  value=linhaABA_ER[chave][1])
            ABA_ER.cell(row=l, column=3,  value=linhaABA_ER[chave][2])
            ABA_ER.cell(row=l, column=4,  value=linhaABA_ER[chave][3])
            ABA_ER.cell(row=l, column=5,  value=linhaABA_ER[chave][4])
            ABA_ER.cell(row=l, column=6,  value=linhaABA_ER[chave][5])
            ABA_ER.cell(row=l, column=7,  value=linhaABA_ER[chave][6])
            ABA_ER.cell(row=l, column=8,  value=linhaABA_ER[chave][7])
            ABA_ER.cell(row=l, column=9,  value=linhaABA_ER[chave][8])
            
            l=l+1
    
    
    
        somad="=SUM(D3:D" + str(l-1) + ")"
        somae="=SUM(E3:E" + str(l-1) + ")"
        somaf="=SUM(F3:F" + str(l-1) + ")"
        somag="=SUM(G3:G" + str(l-1) + ")"
        somah="=SUM(H3:H" + str(l-1) + ")"
        somai="=SUM(I3:I" + str(l-1) + ")"
            
        ABA_ER.cell(row=l, column=1,  value="TOTAIS")
        ABA_ER.cell(row=l, column=4,  value=somad)
        ABA_ER.cell(row=l, column=5,  value=somae)
        ABA_ER.cell(row=l, column=6,  value=somaf)
        ABA_ER.cell(row=l, column=7,  value=somag)
        ABA_ER.cell(row=l, column=8,  value=somah)
        ABA_ER.cell(row=l, column=9,  value=somai)
    
        formata_EP(ABA_ER)
    
        arquivo_excel.save(ARQ_INSUMOS)
        
    ##########  ABA 3-   ABA_SP = SAÍDA PROTOCOLADO 2
    ##########  ABA 3-   ABA_SP = SAÍDA PROTOCOLADO 2
    ##########  ABA 3-   ABA_SP = SAÍDA PROTOCOLADO 2
    ##########  ABA 3-   ABA_SP = SAÍDA PROTOCOLADO 2
    ##########  ABA 3-   ABA_SP = SAÍDA PROTOCOLADO 2
    ##########  ABA 3-   ABA_SP = SAÍDA PROTOCOLADO 2
    if ('03' in listadeabas):
        log( "ABA 03 / ", qtdabas , " -> SAIDAS - PROTOCOLADO")

        encP = encodingDoArquivo(ARQ_PRO)
        ARQPRO = open(ARQ_PRO, 'r', encoding=encP, errors='ignore')
        linhaABA_SP = {}
        chaves = []
        
        for linhaP in ARQPRO:
            camposlinhaP = linhaP.split("|")
            if (len(camposlinhaP) > 3 ):
              
                if ( (camposlinhaP[1].upper() == "C190" and camposlinhaP[3][0] in ("5", "6", "7")) or (camposlinhaP[1].upper() in ("D696"))):
                    
                    agrupamento = camposlinhaP[3].upper()+"x"+camposlinhaP[2].upper()+"x"+camposlinhaP[4].upper()
                    
                    if (agrupamento in chaves):
                        #log(" = ",)
                        #log("Entrou no if = ",)
                        #log("linhaABA_SP = ", linhaABA_SP)
                        #log("Agrupamento = ", agrupamento)
                        #log("chaves      = ", chaves)
                        #input("Continua? ")
                        linhaABA_SP[agrupamento] = (linhaABA_SP[agrupamento][0],
                                                    linhaABA_SP[agrupamento][1],
                                                    linhaABA_SP[agrupamento][2],
                                                    linhaABA_SP[agrupamento][3]  + (0.00 if (camposlinhaP[5]  == "") else float(camposlinhaP[5].replace("," , "."))),
                                                    linhaABA_SP[agrupamento][4]  + (0.00 if (camposlinhaP[6]  == "") else float(camposlinhaP[6].replace("," , "."))),
                                                    linhaABA_SP[agrupamento][5]  + (0.00 if (camposlinhaP[7]  == "") else float(camposlinhaP[7].replace("," , "."))),
                                                    linhaABA_SP[agrupamento][6]  + (0.00 if (camposlinhaP[8]  == "") else float(camposlinhaP[8].replace("," , "."))),
                                                    linhaABA_SP[agrupamento][7]  + (0.00 if (camposlinhaP[9]  == "") else float(camposlinhaP[9].replace("," , "."))),
                                                    linhaABA_SP[agrupamento][8]  + (0.00 if (camposlinhaP[11] == "") else float(camposlinhaP[11].replace("," , "."))))
                        #log("saiu do if",)
                        #log(" = ",)
                                                    
                    else:
                        #log(" = ",)
                        #log("entrou no else = ",)
                        
                        #log("camposlinhaP[11] = ",camposlinhaP[11])
                        linhaABA_SP[agrupamento] = (camposlinhaP[2],
                                                    camposlinhaP[3],
                                                    camposlinhaP[4],
                                                    (0.00 if (camposlinhaP[5]  == "") else float(camposlinhaP[5].replace("," , "."))),
                                                    (0.00 if (camposlinhaP[6]  == "") else float(camposlinhaP[6].replace("," , "."))),
                                                    (0.00 if (camposlinhaP[7]  == "") else float(camposlinhaP[7].replace("," , "."))),
                                                    (0.00 if (camposlinhaP[8]  == "") else float(camposlinhaP[8].replace("," , "."))),
                                                    (0.00 if (camposlinhaP[9]  == "") else float(camposlinhaP[9].replace("," , "."))),
                                                    (0.00 if (camposlinhaP[11] == "") else float(camposlinhaP[11].replace("," , "."))))
                        chaves.append(agrupamento)
                        #log("saiu else = ",)
                        #log(" = ",)
            
            
            
        ARQPRO.close()   
        l=1
        ABA_SP.cell(row=l, column=1,  value="Resumo por CST / CFOP - Saidas (Mercadorias e Telecom) - PROTOCOLADO")
        l=l+1
        ABA_SP.cell(row=l, column=1,  value="CST/CFOP")
        ABA_SP.cell(row=l, column=2,  value="CFOP")
        ABA_SP.cell(row=l, column=3,  value="ALÍQUOTA ICMS")
        ABA_SP.cell(row=l, column=4,  value="VLR_OPER")
        ABA_SP.cell(row=l, column=5,  value="VLR_BASE_ICMS")
        ABA_SP.cell(row=l, column=6,  value="VLR_ICMS")
        ABA_SP.cell(row=l, column=7,  value="VLR_BASE_ICMS_ST")
        ABA_SP.cell(row=l, column=8,  value="VLR_ICMS_ST")
        ABA_SP.cell(row=l, column=9,  value="VLR_IPI")
        
        l=l+1
        
        chaves.sort()
    
        for chave in chaves:        
    
            ABA_SP.cell(row=l, column=1,  value=linhaABA_SP[chave][0])
            ABA_SP.cell(row=l, column=2,  value=linhaABA_SP[chave][1])
            ABA_SP.cell(row=l, column=3,  value=linhaABA_SP[chave][2])
            ABA_SP.cell(row=l, column=4,  value=linhaABA_SP[chave][3])
            ABA_SP.cell(row=l, column=5,  value=linhaABA_SP[chave][4])
            ABA_SP.cell(row=l, column=6,  value=linhaABA_SP[chave][5])
            ABA_SP.cell(row=l, column=7,  value=linhaABA_SP[chave][6])
            ABA_SP.cell(row=l, column=8,  value=linhaABA_SP[chave][7])
            ABA_SP.cell(row=l, column=9,  value=linhaABA_SP[chave][8])
            
            l=l+1
    
    
    
        somad="=SUM(D3:D" + str(l-1) + ")"
        somae="=SUM(E3:E" + str(l-1) + ")"
        somaf="=SUM(F3:F" + str(l-1) + ")"
        somag="=SUM(G3:G" + str(l-1) + ")"
        somah="=SUM(H3:H" + str(l-1) + ")"
        somai="=SUM(I3:I" + str(l-1) + ")"
            
        ABA_SP.cell(row=l, column=1,  value="TOTAIS")
        ABA_SP.cell(row=l, column=4,  value=somad)
        ABA_SP.cell(row=l, column=5,  value=somae)
        ABA_SP.cell(row=l, column=6,  value=somaf)
        ABA_SP.cell(row=l, column=7,  value=somag)
        ABA_SP.cell(row=l, column=8,  value=somah)
        ABA_SP.cell(row=l, column=9,  value=somai)
    
        formata_EP(ABA_SP)
    
    
        arquivo_excel.save(ARQ_INSUMOS)
            
    
    ##########  ABA 4-   ABA_SR = SAÍDA REGERADO 3
    ##########  ABA 4-   ABA_SR = SAÍDA REGERADO 3
    ##########  ABA 4-   ABA_SR = SAÍDA REGERADO 3
    ##########  ABA 4-   ABA_SR = SAÍDA REGERADO 3 
    ##########  ABA 4-   ABA_SR = SAÍDA REGERADO 3
    ##########  ABA 4-   ABA_SR = SAÍDA REGERADO 3
    if ('04' in listadeabas):
        log( "ABA 04 / ", qtdabas , " -> SAIDAS - REGERADO")
    
    
        encE = encodingDoArquivo(ARQ_ENX)
        ARQENX = open(ARQ_ENX, 'r', encoding=encE, errors='ignore')
        linhaABA_SR = {}
        chaves = []
        
        for linhaE in ARQENX:
            camposlinhaE = linhaE.split("|")
            if (len(camposlinhaE) > 3 ):
              
                if ( (camposlinhaE[1].upper() == "C190" and camposlinhaE[3][0] in ("5", "6", "7")) or (camposlinhaE[1].upper() in ("D696"))):
                    
                    agrupamento = camposlinhaE[3].upper()+"x"+camposlinhaE[2].upper()+"x"+camposlinhaE[4].upper()
                    
                    if (agrupamento in chaves):
                        linhaABA_SR[agrupamento] = (linhaABA_SR[agrupamento][0],
                                                    linhaABA_SR[agrupamento][1],
                                                    linhaABA_SR[agrupamento][2],
                                                    linhaABA_SR[agrupamento][3]  + (0.00 if (camposlinhaE[5]   == "") else float(camposlinhaE[5].replace("," , "."))),
                                                    linhaABA_SR[agrupamento][4]  + (0.00 if (camposlinhaE[6]   == "") else float(camposlinhaE[6].replace("," , "."))),
                                                    linhaABA_SR[agrupamento][5]  + (0.00 if (camposlinhaE[7]   == "") else float(camposlinhaE[7].replace("," , "."))),
                                                    linhaABA_SR[agrupamento][6]  + (0.00 if (camposlinhaE[8]   == "") else float(camposlinhaE[8].replace("," , "."))),
                                                    linhaABA_SR[agrupamento][7]  + (0.00 if (camposlinhaE[9]   == "") else float(camposlinhaE[9].replace("," , "."))),
                                                    linhaABA_SR[agrupamento][8]  + (0.00 if (camposlinhaE[11]  == "") else float(camposlinhaE[11].replace("," , "."))))
                    else:
                        linhaABA_SR[agrupamento] = (camposlinhaE[2],
                                                    camposlinhaE[3],
                                                    camposlinhaE[4],
                                                    (0.00 if (camposlinhaE[5]  == "") else float(camposlinhaE[5].replace("," , "."))),
                                                    (0.00 if (camposlinhaE[6]  == "") else float(camposlinhaE[6].replace("," , "."))),
                                                    (0.00 if (camposlinhaE[7]  == "") else float(camposlinhaE[7].replace("," , "."))),
                                                    (0.00 if (camposlinhaE[8]  == "") else float(camposlinhaE[8].replace("," , "."))),
                                                    (0.00 if (camposlinhaE[9]  == "") else float(camposlinhaE[9].replace("," , "."))),
                                                    (0.00 if (camposlinhaE[11] == "") else float(camposlinhaE[11].replace("," , "."))))
                        chaves.append(agrupamento)
            
        ARQENX.close()    
        l=1
        ABA_SR.cell(row=l, column=1,  value="Resumo por CST / CFOP - Saidas (Mercadoria e Telecom) - REGERADO")
        l=l+1
        ABA_SR.cell(row=l, column=1,  value="CST/CFOP")
        ABA_SR.cell(row=l, column=2,  value="CFOP")
        ABA_SR.cell(row=l, column=3,  value="ALÍQUOTA ICMS")
        ABA_SR.cell(row=l, column=4,  value="VLR_OPER")
        ABA_SR.cell(row=l, column=5,  value="VLR_BASE_ICMS")
        ABA_SR.cell(row=l, column=6,  value="VLR_ICMS")
        ABA_SR.cell(row=l, column=7,  value="VLR_BASE_ICMS_ST")
        ABA_SR.cell(row=l, column=8,  value="VLR_ICMS_ST")
        ABA_SR.cell(row=l, column=9,  value="VLR_IPI")
        
        l=l+1
        
        chaves.sort()
    
        for chave in chaves:        
    
            ABA_SR.cell(row=l, column=1,  value=linhaABA_SR[chave][0])
            ABA_SR.cell(row=l, column=2,  value=linhaABA_SR[chave][1])
            ABA_SR.cell(row=l, column=3,  value=linhaABA_SR[chave][2])
            ABA_SR.cell(row=l, column=4,  value=linhaABA_SR[chave][3])
            ABA_SR.cell(row=l, column=5,  value=linhaABA_SR[chave][4])
            ABA_SR.cell(row=l, column=6,  value=linhaABA_SR[chave][5])
            ABA_SR.cell(row=l, column=7,  value=linhaABA_SR[chave][6])
            ABA_SR.cell(row=l, column=8,  value=linhaABA_SR[chave][7])
            ABA_SR.cell(row=l, column=9,  value=linhaABA_SR[chave][8])
            
            l=l+1
    
    
    
        somad="=SUM(D3:D" + str(l-1) + ")"
        somae="=SUM(E3:E" + str(l-1) + ")"
        somaf="=SUM(F3:F" + str(l-1) + ")"
        somag="=SUM(G3:G" + str(l-1) + ")"
        somah="=SUM(H3:H" + str(l-1) + ")"
        somai="=SUM(I3:I" + str(l-1) + ")"
            
        ABA_SR.cell(row=l, column=1,  value="TOTAIS")
        ABA_SR.cell(row=l, column=4,  value=somad)
        ABA_SR.cell(row=l, column=5,  value=somae)
        ABA_SR.cell(row=l, column=6,  value=somaf)
        ABA_SR.cell(row=l, column=7,  value=somag)
        ABA_SR.cell(row=l, column=8,  value=somah)
        ABA_SR.cell(row=l, column=9,  value=somai)
    
        formata_EP(ABA_SR)
    
    
        arquivo_excel.save(ARQ_INSUMOS)
            
      
    ##########  ABA 5-   ABA_TP = TELECOM PROTOCOLADO 4 RF01.05
    ##########  ABA 5-   ABA_TP = TELECOM PROTOCOLADO 4 RF01.05
    ##########  ABA 5-   ABA_TP = TELECOM PROTOCOLADO 4 RF01.05
    ##########  ABA 5-   ABA_TP = TELECOM PROTOCOLADO 4 RF01.05
    ##########  ABA 5-   ABA_TP = TELECOM PROTOCOLADO 4 RF01.05
    ##########  ABA 5-   ABA_TP = TELECOM PROTOCOLADO 4 RF01.05
    if ('05' in listadeabas):
        log( "ABA 05 / ", qtdabas , " -> TELECOM - PROTOCOLADO")

        encP = encodingDoArquivo(ARQ_PRO)
        ARQPRO = open(ARQ_PRO, 'r', encoding=encP, errors='ignore')
        linhaABA_TP = {}
        chaves = []
        
        for linhaP in ARQPRO:
            camposlinhaP = linhaP.split("|")
            if (len(camposlinhaP) > 3 ):
              
                if ( camposlinhaP[1].upper() in ("D696")):
                    
                    agrupamento = camposlinhaP[3].upper()+"x"+camposlinhaP[2].upper()+"x"+camposlinhaP[4].upper()
                    
                    if (agrupamento in chaves):
                        linhaABA_TP[agrupamento] = (linhaABA_TP[agrupamento][0],
                                                    linhaABA_TP[agrupamento][1],
                                                    linhaABA_TP[agrupamento][2],
                                                    linhaABA_TP[agrupamento][3]  + (0.00 if (camposlinhaP[5]  == "") else float(camposlinhaP[5].replace("," , "."))),
                                                    linhaABA_TP[agrupamento][4]  + (0.00 if (camposlinhaP[6]  == "") else float(camposlinhaP[6].replace("," , "."))),
                                                    linhaABA_TP[agrupamento][5]  + (0.00 if (camposlinhaP[7]  == "") else float(camposlinhaP[7].replace("," , "."))),
                                                    linhaABA_TP[agrupamento][6]  + (0.00 if (camposlinhaP[8]  == "") else float(camposlinhaP[8].replace("," , "."))),
                                                    linhaABA_TP[agrupamento][7]  + (0.00 if (camposlinhaP[9]  == "") else float(camposlinhaP[9].replace("," , "."))))
                    else:
                        linhaABA_TP[agrupamento] = (camposlinhaP[2],
                                                    camposlinhaP[3],
                                                    camposlinhaP[4],
                                                    (0.00 if (camposlinhaP[5]  == "") else float(camposlinhaP[5].replace("," , "."))),
                                                    (0.00 if (camposlinhaP[6]  == "") else float(camposlinhaP[6].replace("," , "."))),
                                                    (0.00 if (camposlinhaP[7]  == "") else float(camposlinhaP[7].replace("," , "."))),
                                                    (0.00 if (camposlinhaP[8]  == "") else float(camposlinhaP[8].replace("," , "."))),
                                                    (0.00 if (camposlinhaP[9]  == "") else float(camposlinhaP[9].replace("," , "."))))
                        chaves.append(agrupamento)
        ARQPRO.close()                        
        l=1
        ABA_TP.cell(row=l, column=1,  value="Resumo por CST / CFOP - Saídas Telecom - PROTOCOLADO")
        l=l+1
        ABA_TP.cell(row=l, column=1,  value="CST/CFOP")
        ABA_TP.cell(row=l, column=2,  value="CFOP")
        ABA_TP.cell(row=l, column=3,  value="ALÍQUOTA ICMS")
        ABA_TP.cell(row=l, column=4,  value="VLR_OPER")
        ABA_TP.cell(row=l, column=5,  value="VLR_BASE_ICMS")
        ABA_TP.cell(row=l, column=6,  value="VLR_ICMS")
        ABA_TP.cell(row=l, column=7,  value="VLR_BASE_ICMS_ST")
        ABA_TP.cell(row=l, column=8,  value="VLR_ICMS_ST")

        
        l=l+1
        
        chaves.sort()
    
        for chave in chaves:        
    
            ABA_TP.cell(row=l, column=1,  value=linhaABA_TP[chave][0])
            ABA_TP.cell(row=l, column=2,  value=linhaABA_TP[chave][1])
            ABA_TP.cell(row=l, column=3,  value=linhaABA_TP[chave][2])
            ABA_TP.cell(row=l, column=4,  value=linhaABA_TP[chave][3])
            ABA_TP.cell(row=l, column=5,  value=linhaABA_TP[chave][4])
            ABA_TP.cell(row=l, column=6,  value=linhaABA_TP[chave][5])
            ABA_TP.cell(row=l, column=7,  value=linhaABA_TP[chave][6])
            ABA_TP.cell(row=l, column=8,  value=linhaABA_TP[chave][7])

            
            l=l+1
    
    
    
        somad="=SUM(D3:D" + str(l-1) + ")"
        somae="=SUM(E3:E" + str(l-1) + ")"
        somaf="=SUM(F3:F" + str(l-1) + ")"
        somag="=SUM(G3:G" + str(l-1) + ")"
        somah="=SUM(H3:H" + str(l-1) + ")"
        
            
        ABA_TP.cell(row=l, column=1,  value="TOTAIS")
        ABA_TP.cell(row=l, column=4,  value=somad)
        ABA_TP.cell(row=l, column=5,  value=somae)
        ABA_TP.cell(row=l, column=6,  value=somaf)
        ABA_TP.cell(row=l, column=7,  value=somag)
        ABA_TP.cell(row=l, column=8,  value=somah)

    
        formata_TP(ABA_TP)
    
    
        arquivo_excel.save(ARQ_INSUMOS)
            
    
    
    #ABA_TR       = arquivo_excel.create_sheet("Telecom REGERADO"               , 5)
    ##########  ABA 6-   ABA_TR = TELECOM REGERADO 5 RF01.06
    ##########  ABA 6-   ABA_TR = TELECOM REGERADO 5 RF01.06
    ##########  ABA 6-   ABA_TR = TELECOM REGERADO 5 RF01.06
    ##########  ABA 6-   ABA_TR = TELECOM REGERADO 5 RF01.06
    ##########  ABA 6-   ABA_TR = TELECOM REGERADO 5 RF01.06
    ##########  ABA 6-   ABA_TR = TELECOM REGERADO 5 RF01.06
    if ('06' in listadeabas):
        log( "ABA 06 / ", qtdabas , " -> TELECOM - REGERADO")
    
        encE = encodingDoArquivo(ARQ_ENX)
        ARQENX = open(ARQ_ENX, 'r', encoding=encE, errors='ignore')
        linhaABA_TR = {}
        chaves = []
        
        for linhaE in ARQENX:
            camposlinhaE = linhaE.split("|")
    
            if (len(camposlinhaE) > 3 ):
              
                if ( (camposlinhaE[1].upper() == "D696")) :
                   
                    agrupamento = camposlinhaE[3].upper()+"x"+camposlinhaE[2].upper()+"x"+camposlinhaE[4].upper()
                    
                    if (agrupamento in chaves):
                        linhaABA_TR[agrupamento] = (linhaABA_TR[agrupamento][0],
                                                    linhaABA_TR[agrupamento][1],
                                                    linhaABA_TR[agrupamento][2],
                                                    linhaABA_TR[agrupamento][3]  + (0.00 if (camposlinhaE[5]   == "") else float(camposlinhaE[5].replace("," , "."))),
                                                    linhaABA_TR[agrupamento][4]  + (0.00 if (camposlinhaE[6]   == "") else float(camposlinhaE[6].replace("," , "."))),
                                                    linhaABA_TR[agrupamento][5]  + (0.00 if (camposlinhaE[7]   == "") else float(camposlinhaE[7].replace("," , "."))),
                                                    linhaABA_TR[agrupamento][6]  + (0.00 if (camposlinhaE[8]   == "") else float(camposlinhaE[8].replace("," , "."))),
                                                    linhaABA_TR[agrupamento][7]  + (0.00 if (camposlinhaE[9]   == "") else float(camposlinhaE[9].replace("," , "."))))
                    else:
                        linhaABA_TR[agrupamento] = (camposlinhaE[2],
                                                    camposlinhaE[3],
                                                    camposlinhaE[4],
                                                    (0.00 if (camposlinhaE[5]  == "") else float(camposlinhaE[5].replace("," , "."))),
                                                    (0.00 if (camposlinhaE[6]  == "") else float(camposlinhaE[6].replace("," , "."))),
                                                    (0.00 if (camposlinhaE[7]  == "") else float(camposlinhaE[7].replace("," , "."))),
                                                    (0.00 if (camposlinhaE[8]  == "") else float(camposlinhaE[8].replace("," , "."))),
                                                    (0.00 if (camposlinhaE[9]  == "") else float(camposlinhaE[9].replace("," , "."))))

                        chaves.append(agrupamento)
            
        ARQENX.close()    
        l=1
        ABA_TR.cell(row=l, column=1,  value="Resumo por CST / CFOP - Saidas Telecom - REGERADO")
        l=l+1
        ABA_TR.cell(row=l, column=1,  value="CST/CFOP")
        ABA_TR.cell(row=l, column=2,  value="CFOP")
        ABA_TR.cell(row=l, column=3,  value="ALÍQUOTA ICMS")
        ABA_TR.cell(row=l, column=4,  value="VLR_OPER")
        ABA_TR.cell(row=l, column=5,  value="VLR_BASE_ICMS")
        ABA_TR.cell(row=l, column=6,  value="VLR_ICMS")
        ABA_TR.cell(row=l, column=7,  value="VLR_BASE_ICMS_ST")
        ABA_TR.cell(row=l, column=8,  value="VLR_ICMS_ST")
      
        l=l+1
        
        chaves.sort()
    
        for chave in chaves:        
    
            ABA_TR.cell(row=l, column=1,  value=linhaABA_TR[chave][0])
            ABA_TR.cell(row=l, column=2,  value=linhaABA_TR[chave][1])
            ABA_TR.cell(row=l, column=3,  value=linhaABA_TR[chave][2])
            ABA_TR.cell(row=l, column=4,  value=linhaABA_TR[chave][3])
            ABA_TR.cell(row=l, column=5,  value=linhaABA_TR[chave][4])
            ABA_TR.cell(row=l, column=6,  value=linhaABA_TR[chave][5])
            ABA_TR.cell(row=l, column=7,  value=linhaABA_TR[chave][6])
            ABA_TR.cell(row=l, column=8,  value=linhaABA_TR[chave][7])

            
            l=l+1
    
    
    
        somad="=SUM(D3:D" + str(l-1) + ")"
        somae="=SUM(E3:E" + str(l-1) + ")"
        somaf="=SUM(F3:F" + str(l-1) + ")"
        somag="=SUM(G3:G" + str(l-1) + ")"
        somah="=SUM(H3:H" + str(l-1) + ")"

            
        ABA_TR.cell(row=l, column=1,  value="TOTAIS")
        ABA_TR.cell(row=l, column=4,  value=somad)
        ABA_TR.cell(row=l, column=5,  value=somae)
        ABA_TR.cell(row=l, column=6,  value=somaf)
        ABA_TR.cell(row=l, column=7,  value=somag)
        ABA_TR.cell(row=l, column=8,  value=somah)
    
        formata_TP(ABA_TR)
    
    
        arquivo_excel.save(ARQ_INSUMOS)
    
    
    #ABA_TRS      = arquivo_excel.create_sheet("Telecom REGERADO - POR SERIE"   , 6)
    ##########  ABA 7-   ABA_TRS = TELECOM REGERADO - POR SERIE 6 RF01.07
    ##########  ABA 7-   ABA_TRS = TELECOM REGERADO - POR SERIE 6 RF01.07
    ##########  ABA 7-   ABA_TRS = TELECOM REGERADO - POR SERIE 6 RF01.07
    ##########  ABA 7-   ABA_TRS = TELECOM REGERADO - POR SERIE 6 RF01.07
    ##########  ABA 7-   ABA_TRS = TELECOM REGERADO - POR SERIE 6 RF01.07
    ##########  ABA 7-   ABA_TRS = TELECOM REGERADO - POR SERIE 6 RF01.07
    if ('07' in listadeabas):    
        log( "ABA 07 / ", qtdabas , " -> TELECOM - POR SERIE - REGERADO")
        
        linhaABA_TRS = {}
        chaves = []
        dadosABA_TRS = Busca_TRS(iei,"01/"+mesi+"/"+anoi)
        
        for camposlinhaE in dadosABA_TRS:
            agrupamento = camposlinhaE[0].upper()+"x"+camposlinhaE[1][3:5]+"x"+camposlinhaE[1][0:2]+"x"+camposlinhaE[2].upper()+"x"+camposlinhaE[3].upper()
            if (agrupamento in chaves):
                linhaABA_TRS[agrupamento] =(linhaABA_TRS[agrupamento][0],
                                            linhaABA_TRS[agrupamento][1],
                                            linhaABA_TRS[agrupamento][2],
                                            linhaABA_TRS[agrupamento][3],
                                            linhaABA_TRS[agrupamento][4]  + (0.00 if (camposlinhaE[4]    == None) else camposlinhaE[4]),
                                            linhaABA_TRS[agrupamento][5]  + (0.00 if (camposlinhaE[5]    == None) else camposlinhaE[5]),
                                            linhaABA_TRS[agrupamento][6]  + (0.00 if (camposlinhaE[6]    == None) else camposlinhaE[6]),
                                            linhaABA_TRS[agrupamento][7]  + (0.00 if (camposlinhaE[7]    == None) else camposlinhaE[7]),
                                            linhaABA_TRS[agrupamento][8]  + (0.00 if (camposlinhaE[8]    == None) else camposlinhaE[8]),
                                            linhaABA_TRS[agrupamento][9]  + (0.00 if (camposlinhaE[9]    == None) else camposlinhaE[9]),
                                            linhaABA_TRS[agrupamento][10] + (0.00 if (camposlinhaE[10]   == None) else camposlinhaE[10]),
                                            linhaABA_TRS[agrupamento][11] + (0.00 if (camposlinhaE[11]   == None) else camposlinhaE[11]))
            else:
                linhaABA_TRS[agrupamento] = (camposlinhaE[0],
                                            camposlinhaE[1],
                                            camposlinhaE[2],
                                            camposlinhaE[3],
                                            (0.00 if (camposlinhaE[4]  == None) else camposlinhaE[4]),
                                            (0.00 if (camposlinhaE[5]  == None) else camposlinhaE[5]),
                                            (0.00 if (camposlinhaE[6]  == None) else camposlinhaE[6]),
                                            (0.00 if (camposlinhaE[7]  == None) else camposlinhaE[7]),
                                            (0.00 if (camposlinhaE[8]  == None) else camposlinhaE[8]),
                                            (0.00 if (camposlinhaE[9]  == None) else camposlinhaE[9]),
                                            (0.00 if (camposlinhaE[10] == None) else camposlinhaE[10]),
                                            (0.00 if (camposlinhaE[11] == None) else camposlinhaE[11]))
            chaves.append(agrupamento)
            
            
        l=1
        ABA_TRS.cell(row=l, column=1,  value="Resumo por SERIE / CFOP - Saidas  Telecom - REGERADO - Origem Banco de Dados")
        l=l+1
        ABA_TRS.cell(row=l, column=1,  value="EMPS_COD")
        ABA_TRS.cell(row=l, column=2,  value="MES_ANO")
        ABA_TRS.cell(row=l, column=3,  value="SERIE")
        ABA_TRS.cell(row=l, column=4,  value="CFOP")
        ABA_TRS.cell(row=l, column=5,  value="VLR_LIQUIDO")
        ABA_TRS.cell(row=l, column=6,  value="VLR_SERVICO")
        ABA_TRS.cell(row=l, column=7,  value="VLR_BASE_ICMS")
        ABA_TRS.cell(row=l, column=8,  value="VLR_ICMS")
        ABA_TRS.cell(row=l, column=9,  value="VLR_ISENTAS")
        ABA_TRS.cell(row=l, column=10, value="VLR_OUTRAS")
        ABA_TRS.cell(row=l, column=11, value="VLR_DESCONTO")
        ABA_TRS.cell(row=l, column=12, value="VLR_REDUCAO")
        
        l=l+1
        
        chaves.sort()
    
        for chave in chaves:        
    
            ABA_TRS.cell(row=l, column=1,  value=linhaABA_TRS[chave][0])
            ABA_TRS.cell(row=l, column=2,  value=linhaABA_TRS[chave][1])
            ABA_TRS.cell(row=l, column=3,  value=linhaABA_TRS[chave][2])
            ABA_TRS.cell(row=l, column=4,  value=linhaABA_TRS[chave][3])
            ABA_TRS.cell(row=l, column=5,  value=linhaABA_TRS[chave][4])
            ABA_TRS.cell(row=l, column=6,  value=linhaABA_TRS[chave][5])
            ABA_TRS.cell(row=l, column=7,  value=linhaABA_TRS[chave][6])
            ABA_TRS.cell(row=l, column=8,  value=linhaABA_TRS[chave][7])
            ABA_TRS.cell(row=l, column=9,  value=linhaABA_TRS[chave][8])
            ABA_TRS.cell(row=l, column=10, value=linhaABA_TRS[chave][9])
            ABA_TRS.cell(row=l, column=11, value=linhaABA_TRS[chave][10])
            ABA_TRS.cell(row=l, column=12, value=linhaABA_TRS[chave][11])
            
            l=l+1
    
    
    
        somae="=SUM(E3:E" + str(l-1) + ")"
        somaf="=SUM(F3:F" + str(l-1) + ")"
        somag="=SUM(G3:G" + str(l-1) + ")"
        somah="=SUM(H3:H" + str(l-1) + ")"
        somai="=SUM(I3:I" + str(l-1) + ")"
        somaj="=SUM(J3:J" + str(l-1) + ")"
        somak="=SUM(K3:K" + str(l-1) + ")"
        somal="=SUM(L3:L" + str(l-1) + ")"
            
        ABA_TRS.cell(row=l, column=1,  value="TOTAIS")
        ABA_TRS.cell(row=l, column=5,  value=somae)
        ABA_TRS.cell(row=l, column=6,  value=somaf)
        ABA_TRS.cell(row=l, column=7,  value=somag)
        ABA_TRS.cell(row=l, column=8,  value=somah)
        ABA_TRS.cell(row=l, column=9,  value=somai)
        ABA_TRS.cell(row=l, column=10, value=somaj)
        ABA_TRS.cell(row=l, column=11, value=somak)
        ABA_TRS.cell(row=l, column=12, value=somal)
    
        formata_TRS(ABA_TRS)
    
    
        arquivo_excel.save(ARQ_INSUMOS)
    
    
    #ABA_TPS      = arquivo_excel.create_sheet("Telecom PROTOCOLADO - POR SERIE", 7)
    ##########  ABA 8-   ABA_TRS = Telecom PROTOCOLADO - POR SERIE", 7 RF01.08
    ##########  ABA 8-   ABA_TRS = Telecom PROTOCOLADO - POR SERIE", 7 RF01.08
    ##########  ABA 8-   ABA_TRS = Telecom PROTOCOLADO - POR SERIE", 7 RF01.08
    ##########  ABA 8-   ABA_TRS = Telecom PROTOCOLADO - POR SERIE", 7 RF01.08
    ##########  ABA 8-   ABA_TRS = Telecom PROTOCOLADO - POR SERIE", 7 RF01.08
    ##########  ABA 8-   ABA_TRS = Telecom PROTOCOLADO - POR SERIE", 7 RF01.08
    ##########  ABA 8-   ABA_TRS = Telecom PROTOCOLADO - POR SERIE", 7 RF01.08
    if ('08' in listadeabas):    
        log( "ABA 08 / ", qtdabas , " -> TELECOM-POR-SERIE-PROTOCOLADO")
       
        
        linhaABA_TPS = {}
        chaves = []
        dadosABA_TPS = Busca_TPS(iei,"01/"+mesi+"/"+anoi)
        
        for camposlinhaE in dadosABA_TPS:
            agrupamento = camposlinhaE[0].upper()+"x"+camposlinhaE[1][3:5]+"x"+camposlinhaE[1][0:2]+"x"+camposlinhaE[2].upper()+"x"+camposlinhaE[3].upper()
            if (agrupamento in chaves):
                linhaABA_TPS[agrupamento] =(linhaABA_TPS[agrupamento][0],
                                            linhaABA_TPS[agrupamento][1],
                                            linhaABA_TPS[agrupamento][2],
                                            linhaABA_TPS[agrupamento][3],
                                            linhaABA_TPS[agrupamento][4]  + (0.00 if (camposlinhaE[4]    == None) else camposlinhaE[4]),
                                            linhaABA_TPS[agrupamento][5]  + (0.00 if (camposlinhaE[5]    == None) else camposlinhaE[5]),
                                            linhaABA_TPS[agrupamento][6]  + (0.00 if (camposlinhaE[6]    == None) else camposlinhaE[6]),
                                            linhaABA_TPS[agrupamento][7]  + (0.00 if (camposlinhaE[7]    == None) else camposlinhaE[7]),
                                            linhaABA_TPS[agrupamento][8]  + (0.00 if (camposlinhaE[8]    == None) else camposlinhaE[8]),
                                            linhaABA_TPS[agrupamento][9]  + (0.00 if (camposlinhaE[9]    == None) else camposlinhaE[9]),
                                            linhaABA_TPS[agrupamento][10] + (0.00 if (camposlinhaE[10]   == None) else camposlinhaE[10]),
                                            linhaABA_TPS[agrupamento][11] + (0.00 if (camposlinhaE[11]   == None) else camposlinhaE[11]))
            else:
                linhaABA_TPS[agrupamento] = (camposlinhaE[0],
                                            camposlinhaE[1],
                                            camposlinhaE[2],
                                            camposlinhaE[3],
                                            (0.00 if (camposlinhaE[4]  == None) else camposlinhaE[4]),
                                            (0.00 if (camposlinhaE[5]  == None) else camposlinhaE[5]),
                                            (0.00 if (camposlinhaE[6]  == None) else camposlinhaE[6]),
                                            (0.00 if (camposlinhaE[7]  == None) else camposlinhaE[7]),
                                            (0.00 if (camposlinhaE[8]  == None) else camposlinhaE[8]),
                                            (0.00 if (camposlinhaE[9]  == None) else camposlinhaE[9]),
                                            (0.00 if (camposlinhaE[10] == None) else camposlinhaE[10]),
                                            (0.00 if (camposlinhaE[11] == None) else camposlinhaE[11]))
            chaves.append(agrupamento)
            
            
        l=1
        ABA_TPS.cell(row=l, column=1,  value="Resumo por SERIE / CFOP - Saidas  Telecom - PROTOCOLADO - Origem Arquivo Conv115 Protocolad")
        l=l+1
        ABA_TPS.cell(row=l, column=1,  value="EMPS_COD")
        ABA_TPS.cell(row=l, column=2,  value="MES_ANO")
        ABA_TPS.cell(row=l, column=3,  value="SERIE")
        ABA_TPS.cell(row=l, column=4,  value="CFOP")
        ABA_TPS.cell(row=l, column=5,  value="VLR_LIQUIDO")
        ABA_TPS.cell(row=l, column=6,  value="VLR_SERVICO")
        ABA_TPS.cell(row=l, column=7,  value="VLR_BASE_ICMS")
        ABA_TPS.cell(row=l, column=8,  value="VLR_ICMS")
        ABA_TPS.cell(row=l, column=9,  value="VLR_ISENTAS")
        ABA_TPS.cell(row=l, column=10, value="VLR_OUTRAS")
        ABA_TPS.cell(row=l, column=11, value="VLR_DESCONTO")
        ABA_TPS.cell(row=l, column=12, value="VLR_REDUCAO")
        
        l=l+1
        
        chaves.sort()
    
        for chave in chaves:        
    
            ABA_TPS.cell(row=l, column=1,  value=linhaABA_TPS[chave][0])
            ABA_TPS.cell(row=l, column=2,  value=linhaABA_TPS[chave][1])
            ABA_TPS.cell(row=l, column=3,  value=linhaABA_TPS[chave][2])
            ABA_TPS.cell(row=l, column=4,  value=linhaABA_TPS[chave][3])
            ABA_TPS.cell(row=l, column=5,  value=linhaABA_TPS[chave][4])
            ABA_TPS.cell(row=l, column=6,  value=linhaABA_TPS[chave][5])
            ABA_TPS.cell(row=l, column=7,  value=linhaABA_TPS[chave][6])
            ABA_TPS.cell(row=l, column=8,  value=linhaABA_TPS[chave][7])
            ABA_TPS.cell(row=l, column=9,  value=linhaABA_TPS[chave][8])
            ABA_TPS.cell(row=l, column=10, value=linhaABA_TPS[chave][9])
            ABA_TPS.cell(row=l, column=11, value=linhaABA_TPS[chave][10])
            ABA_TPS.cell(row=l, column=12, value=linhaABA_TPS[chave][11])
            
            l=l+1
    
    
    
        somae="=SUM(E3:E" + str(l-1) + ")"
        somaf="=SUM(F3:F" + str(l-1) + ")"
        somag="=SUM(G3:G" + str(l-1) + ")"
        somah="=SUM(H3:H" + str(l-1) + ")"
        somai="=SUM(I3:I" + str(l-1) + ")"
        somaj="=SUM(J3:J" + str(l-1) + ")"
        somak="=SUM(K3:K" + str(l-1) + ")"
        somal="=SUM(L3:L" + str(l-1) + ")"
            
        ABA_TPS.cell(row=l, column=1,  value="TOTAIS")
        ABA_TPS.cell(row=l, column=5,  value=somae)
        ABA_TPS.cell(row=l, column=6,  value=somaf)
        ABA_TPS.cell(row=l, column=7,  value=somag)
        ABA_TPS.cell(row=l, column=8,  value=somah)
        ABA_TPS.cell(row=l, column=9,  value=somai)
        ABA_TPS.cell(row=l, column=10, value=somaj)
        ABA_TPS.cell(row=l, column=11, value=somak)
        ABA_TPS.cell(row=l, column=12, value=somal)
    
        formata_TRS(ABA_TPS)
    
    
        arquivo_excel.save(ARQ_INSUMOS)
    
    
    
    #ABA_TRT      = arquivo_excel.create_sheet("Telecom REGERADO - TRIB. ICMS"  , 8)
    ##########  ABA 9-   ABA_TRT = Telecom REGERADO - Tributo ICMS", 8 RF01.09
    ##########  ABA 9-   ABA_TRT = Telecom REGERADO - Tributo ICMS", 8 RF01.09
    ##########  ABA 9-   ABA_TRT = Telecom REGERADO - Tributo ICMS", 8 RF01.09
    ##########  ABA 9-   ABA_TRT = Telecom REGERADO - Tributo ICMS", 8 RF01.09
    ##########  ABA 9-   ABA_TRT = Telecom REGERADO - Tributo ICMS", 8 RF01.09
    ##########  ABA 9-   ABA_TRT = Telecom REGERADO - Tributo ICMS", 8 RF01.09
    ##########  ABA 9-   ABA_TRT = Telecom REGERADO - Tributo ICMS", 8 RF01.09
    if ('09' in listadeabas):        
        log( "ABA 09 / ", qtdabas , " -> TELECOM - REGERADO - TRIB ICMS")
        linhaABA_TRT = {}
        chaves = []
        dadosABA_TRT = Busca_TRT(iei,"01/"+mesi+"/"+anoi)
        l=1
        ABA_TRT.cell(row=l, column=1,  value="TELECOM SAIDAS - Base ICMS igual a zero x Valor ICMS maior que zero e vice-versa - REGERADO")
        l=l+1
        ABA_TRT.cell(row=l, column=1,  value="EMPS_COD")
        ABA_TRT.cell(row=l, column=2,  value="FILI_COD")
        ABA_TRT.cell(row=l, column=3,  value="SERIE")
        ABA_TRT.cell(row=l, column=4,  value="NUMERO_NF")
        ABA_TRT.cell(row=l, column=5,  value="EMISSAO")
        ABA_TRT.cell(row=l, column=6,  value="CST")
        ABA_TRT.cell(row=l, column=7,  value="DESC_COMPL")
        ABA_TRT.cell(row=l, column=8,  value="VLR_LIQUIDO")
        ABA_TRT.cell(row=l, column=9,  value="ALIQ_ICMS")
        ABA_TRT.cell(row=l, column=10, value="VLR_BASE_ICMS")
        ABA_TRT.cell(row=l, column=11, value="VLR_ICMS")
        ABA_TRT.cell(row=l, column=12, value="VLR_ISENTAS")
        ABA_TRT.cell(row=l, column=13, value="VLR_OUTRAS")
        l=l+1
        for linha in dadosABA_TRT:
            for i in range(0,len(linha)):
                ABA_TRT.cell(row=l, column=i+1,  value=linha[i])
            l=l+1
        somae="=SUM(E3:E" + str(l-1) + ")"
        somaf="=SUM(F3:F" + str(l-1) + ")"
        somag="=SUM(G3:G" + str(l-1) + ")"
        somah="=SUM(H3:H" + str(l-1) + ")"
        somaj="=SUM(J3:J" + str(l-1) + ")"
        somak="=SUM(K3:K" + str(l-1) + ")"
        somal="=SUM(L3:L" + str(l-1) + ")"
        somam="=SUM(M3:M" + str(l-1) + ")"
        ABA_TRT.cell(row=l, column=1,  value="TOTAIS")
        ABA_TRT.cell(row=l, column=5,  value=somae)
        ABA_TRT.cell(row=l, column=6,  value=somaf)
        ABA_TRT.cell(row=l, column=7,  value=somag)
        ABA_TRT.cell(row=l, column=8,  value=somah)
        ABA_TRT.cell(row=l, column=10, value=somaj)
        ABA_TRT.cell(row=l, column=11, value=somak)
        ABA_TRT.cell(row=l, column=12, value=somal)
        ABA_TRT.cell(row=l, column=13, value=somam)
        formata_TRT(ABA_TRT)
        arquivo_excel.save(ARQ_INSUMOS)

    
    
    
    
    #ABA_RCFOP    = arquivo_excel.create_sheet("Resumo CFOP"                    , 9)
    ##########  ABA 10-   ABA_RCFOP = Resumo CFOP"  
    ##########  ABA 10-   ABA_RCFOP = Resumo CFOP"  
    ##########  ABA 10-   ABA_RCFOP = Resumo CFOP"  
    ##########  ABA 10-   ABA_RCFOP = Resumo CFOP"  
    ##########  ABA 10-   ABA_RCFOP = Resumo CFOP"  
    ##########  ABA 10-   ABA_RCFOP = Resumo CFOP"  

    if ('10' in listadeabas):        
        log( "ABA 10 / ", qtdabas , " -> RESUMO CFOP")
        dadosABA_RCFOP = Busca_RCFOP(iei,"01/"+mesi+"/"+anoi)
        l=1
        ABA_RCFOP.cell(row=l, column=1,  value="APURAÇÃO ICMS P9 - Resumo Fiscal (REGERADO)")
        l=l+1
        ABA_RCFOP.cell(row=l, column=1,  value="CFOP")
        ABA_RCFOP.cell(row=l, column=2,  value="VLR_CONTABIL")
        ABA_RCFOP.cell(row=l, column=3,  value="VLR_BASE_ICMS")
        ABA_RCFOP.cell(row=l, column=4,  value="VLR_ICMS")
        ABA_RCFOP.cell(row=l, column=5,  value="VLR_ISENTAS")
        ABA_RCFOP.cell(row=l, column=6,  value="VLR_OUTRAS")
        ABA_RCFOP.cell(row=l, column=7,  value="DIFERENCA")
        l=l+1
        somabe=0.00
        somace=0.00
        somade=0.00
        somaee=0.00
        somafe=0.00
        somage=0.00
        somabs=0.00
        somacs=0.00
        somads=0.00
        somaes=0.00
        somafs=0.00
        somags=0.00
        for linha in dadosABA_RCFOP:
            ABA_RCFOP.cell(row=l, column=1,  value=linha[0])
            ABA_RCFOP.cell(row=l, column=2,  value=linha[1])
            ABA_RCFOP.cell(row=l, column=3,  value=linha[2])
            ABA_RCFOP.cell(row=l, column=4,  value=linha[3])
            ABA_RCFOP.cell(row=l, column=5,  value=linha[4])
            ABA_RCFOP.cell(row=l, column=6,  value=linha[5])
            ABA_RCFOP.cell(row=l, column=7,  value=linha[6])
            if (linha[0][0] in ('1','2','3')):
                somabe=somabe+float(linha[1])
                somace=somace+float(linha[2])
                somade=somade+float(linha[3])
                somaee=somaee+float(linha[4])
                somafe=somafe+float(linha[5])
                somage=somage+float(linha[6])
            elif (linha[0][0] in ('5','6','7')):
                somabs=somabs+float(linha[1])
                somacs=somacs+float(linha[2])
                somads=somads+float(linha[3])
                somaes=somaes+float(linha[4])
                somafs=somafs+float(linha[5])
                somags=somags+float(linha[6])
            l=l+1
        l=l+1
        ABA_RCFOP.cell(row=l, column=1,  value="ENTRADA")
        ABA_RCFOP.cell(row=l, column=2,  value=somabe)
        ABA_RCFOP.cell(row=l, column=3,  value=somace)
        ABA_RCFOP.cell(row=l, column=4,  value=somade)
        ABA_RCFOP.cell(row=l, column=5,  value=somaee)
        ABA_RCFOP.cell(row=l, column=6,  value=somafe)
        ABA_RCFOP.cell(row=l, column=7,  value=somage)
        l=l+2
        ABA_RCFOP.cell(row=l, column=1,  value="SAIDA")
        ABA_RCFOP.cell(row=l, column=2,  value=somabs)
        ABA_RCFOP.cell(row=l, column=3,  value=somacs)
        ABA_RCFOP.cell(row=l, column=4,  value=somads)
        ABA_RCFOP.cell(row=l, column=5,  value=somaes)
        ABA_RCFOP.cell(row=l, column=6,  value=somafs)
        ABA_RCFOP.cell(row=l, column=7,  value=somags)
       
        formata_RCFOP(ABA_RCFOP)
        arquivo_excel.save(ARQ_INSUMOS)
       
        
       
        
       
        
        ###### ABA RESUMO E/S REGERADO
        ###### ABA RESUMO E/S REGERADO
        ###### ABA RESUMO E/S REGERADO
        ###### ABA RESUMO E/S REGERADO
        ########## PVA SPED FISCAL - REGERADO - ENTRADA   
        somale=0.00
        somame=0.00
        somane=0.00
        somals=0.00
        somams=0.00
        somans=0.00

        encE = encodingDoArquivo(ARQ_ENX)
        ARQENX = open(ARQ_ENX, 'r', encoding=encE, errors='ignore')
        ABA_RCFOP_R = {}
        chavesR = []
        for linha in ARQENX:
            camposlinha = linha.split("|")
            if (len(camposlinha) > 3 ):
                if ( (camposlinha[1].upper() == "C190" and camposlinha[3][0] in ("1", "2", "3")) or (camposlinha[1].upper() in ("C590", "D190", "D590"))):
                    agrupamento = camposlinha[3].upper()
                    if (agrupamento in chavesR):
                        ABA_RCFOP_R[agrupamento] = (ABA_RCFOP_R[agrupamento][0],
                                                    ABA_RCFOP_R[agrupamento][1]  + (0.00 if (camposlinha[5]   == "") else float(camposlinha[5].replace("," , "."))),
                                                    ABA_RCFOP_R[agrupamento][2]  + (0.00 if (camposlinha[6]   == "") else float(camposlinha[6].replace("," , "."))),
                                                    ABA_RCFOP_R[agrupamento][3]  + (0.00 if (camposlinha[7]   == "") else float(camposlinha[7].replace("," , "."))))
                    else:
                        ABA_RCFOP_R[agrupamento] = (camposlinha[3],
                                                    (0.00 if (camposlinha[5]  == "") else float(camposlinha[5].replace("," , "."))),
                                                    (0.00 if (camposlinha[6]  == "") else float(camposlinha[6].replace("," , "."))),
                                                    (0.00 if (camposlinha[7]  == "") else float(camposlinha[7].replace("," , "."))))
                        chavesR.append(agrupamento)

                    somale=somale+(0.00 if (camposlinha[5]   == "") else float(camposlinha[5].replace("," , ".")))
                    somame=somame+(0.00 if (camposlinha[6]   == "") else float(camposlinha[6].replace("," , ".")))
                    somane=somane+(0.00 if (camposlinha[7]   == "") else float(camposlinha[7].replace("," , ".")))

                        
                elif ( (camposlinha[1].upper() == "C190" and camposlinha[3][0] in ("5", "6", "7")) or (camposlinha[1].upper() in ("D696"))):
                    agrupamento = camposlinha[3].upper()
                    if (agrupamento in chavesR):
                        ABA_RCFOP_R[agrupamento] = (ABA_RCFOP_R[agrupamento][0],
                                                    ABA_RCFOP_R[agrupamento][1]  + (0.00 if (camposlinha[5]   == "") else float(camposlinha[5].replace("," , "."))),
                                                    ABA_RCFOP_R[agrupamento][2]  + (0.00 if (camposlinha[6]   == "") else float(camposlinha[6].replace("," , "."))),
                                                    ABA_RCFOP_R[agrupamento][3]  + (0.00 if (camposlinha[7]   == "") else float(camposlinha[7].replace("," , "."))))
                    else:
                        ABA_RCFOP_R[agrupamento] = (camposlinha[3],

                                                    (0.00 if (camposlinha[5]  == "") else float(camposlinha[5].replace("," , "."))),
                                                    (0.00 if (camposlinha[6]  == "") else float(camposlinha[6].replace("," , "."))),
                                                    (0.00 if (camposlinha[7]  == "") else float(camposlinha[7].replace("," , "."))))
                        chavesR.append(agrupamento)

                    somals=somals+(0.00 if (camposlinha[5]   == "") else float(camposlinha[5].replace("," , ".")))
                    somams=somams+(0.00 if (camposlinha[6]   == "") else float(camposlinha[6].replace("," , ".")))
                    somans=somans+(0.00 if (camposlinha[7]   == "") else float(camposlinha[7].replace("," , ".")))
        ARQENX.close()
        chavesR.sort()
 

        ###### ABA RESUMO E/S PROTOCOLADO
        ###### ABA RESUMO E/S PROTOCOLADO
        ###### ABA RESUMO E/S PROTOCOLADO
        ###### ABA RESUMO E/S PROTOCOLADO
        ########## PVA SPED FISCAL - PROTOCOLADO - ENTRADA   
        somaqe=0.00
        somare=0.00
        somase=0.00
        somaqs=0.00
        somars=0.00
        somass=0.00
        encP = encodingDoArquivo(ARQ_PRO)
        ARQPRO = open(ARQ_PRO, 'r', encoding=encP, errors='ignore')
        ABA_RCFOP_P = {}
        chavesP = []
        for linha in ARQPRO:
            camposlinha = linha.split("|")
            if (len(camposlinha) > 3 ):
              
                if ( (camposlinha[1].upper() == "C190" and camposlinha[3][0] in ("1", "2", "3")) or (camposlinha[1].upper() in ("C590", "D190", "D590"))):
                    agrupamento = camposlinha[3].upper()
                    if (agrupamento in chavesP):
                        ABA_RCFOP_P[agrupamento] = (ABA_RCFOP_P[agrupamento][0],
                                                    ABA_RCFOP_P[agrupamento][1]  + (0.00 if (camposlinha[5]   == "") else float(camposlinha[5].replace("," , "."))),
                                                    ABA_RCFOP_P[agrupamento][2]  + (0.00 if (camposlinha[6]   == "") else float(camposlinha[6].replace("," , "."))),
                                                    ABA_RCFOP_P[agrupamento][3]  + (0.00 if (camposlinha[7]   == "") else float(camposlinha[7].replace("," , "."))))
                    else:
                        ABA_RCFOP_P[agrupamento] = (camposlinha[3],
                                                    (0.00 if (camposlinha[5]  == "") else float(camposlinha[5].replace("," , "."))),
                                                    (0.00 if (camposlinha[6]  == "") else float(camposlinha[6].replace("," , "."))),
                                                    (0.00 if (camposlinha[7]  == "") else float(camposlinha[7].replace("," , "."))))
                        chavesP.append(agrupamento)
                    somaqe=somaqe+(0.00 if (camposlinha[5]   == "") else float(camposlinha[5].replace("," , ".")))
                    somare=somare+(0.00 if (camposlinha[6]   == "") else float(camposlinha[6].replace("," , ".")))
                    somase=somase+(0.00 if (camposlinha[7]   == "") else float(camposlinha[7].replace("," , ".")))

                elif ( (camposlinha[1].upper() == "C190" and camposlinha[3][0] in ("5", "6", "7")) or (camposlinha[1].upper() in ("D696"))):
                    agrupamento = camposlinha[3].upper()
                    if (agrupamento in chavesP):
                        ABA_RCFOP_P[agrupamento] = (ABA_RCFOP_P[agrupamento][0],
                                                    ABA_RCFOP_P[agrupamento][1]  + (0.00 if (camposlinha[5]   == "") else float(camposlinha[5].replace("," , "."))),
                                                    ABA_RCFOP_P[agrupamento][2]  + (0.00 if (camposlinha[6]   == "") else float(camposlinha[6].replace("," , "."))),
                                                    ABA_RCFOP_P[agrupamento][3]  + (0.00 if (camposlinha[7]   == "") else float(camposlinha[7].replace("," , "."))))
                    else:
                        ABA_RCFOP_P[agrupamento] = (camposlinha[3],
                                                    (0.00 if (camposlinha[5]  == "") else float(camposlinha[5].replace("," , "."))),
                                                    (0.00 if (camposlinha[6]  == "") else float(camposlinha[6].replace("," , "."))),
                                                    (0.00 if (camposlinha[7]  == "") else float(camposlinha[7].replace("," , "."))))
                        chavesP.append(agrupamento)
                    somaqs=somaqs+(0.00 if (camposlinha[5]   == "") else float(camposlinha[5].replace("," , ".")))
                    somars=somars+(0.00 if (camposlinha[6]   == "") else float(camposlinha[6].replace("," , ".")))
                    somass=somass+(0.00 if (camposlinha[7]   == "") else float(camposlinha[7].replace("," , ".")))
        ARQPRO.close()                
        chavesP.sort()
        
        ###### Montagem das planilhas:
        ###### Montagem das planilhas:
        ###### Montagem das planilhas:
        ###### Montagem das planilhas:
        ###### Montagem das planilhas:
        
        allkeys=[]
        for c in chavesR:
            if c not in allkeys:
                allkeys.append(c)
        for c in chavesP:
            if c not in allkeys:
                allkeys.append(c)
        allkeys.sort()

      
        l=1
        ABA_RCFOP.cell(row=l, column=11,  value="PVA SPED FISCAL - REGERADO")
        ABA_RCFOP.cell(row=l, column=16,  value="PVA SPED FISCAL - PROTOCOLADO")
        ABA_RCFOP.cell(row=l, column=21,  value="DIFERENÇA SPED REGERADO X SPED PROTOCOLADO")
        l=l+1
        ABA_RCFOP.cell(row=l, column=11,  value="CFOP")
        ABA_RCFOP.cell(row=l, column=12,  value="VLR_OPER")
        ABA_RCFOP.cell(row=l, column=13,  value="VLR_BASE_ICMS")
        ABA_RCFOP.cell(row=l, column=14,  value="VLR_ICMS")
        ABA_RCFOP.cell(row=l, column=16,  value="CFOP")
        ABA_RCFOP.cell(row=l, column=17,  value="VLR_OPER")
        ABA_RCFOP.cell(row=l, column=18,  value="VLR_BASE_ICMS")
        ABA_RCFOP.cell(row=l, column=19,  value="VLR_ICMS")
        ABA_RCFOP.cell(row=l, column=21,  value="VLR_OPER")
        ABA_RCFOP.cell(row=l, column=22,  value="VLR_BASE_ICMS")
        ABA_RCFOP.cell(row=l, column=23,  value="VLR_ICMS")
        l=l+1

       
        for chave in allkeys:
            if chave in chavesR:
                v12 = ABA_RCFOP_R[chave][1]
                v13 = ABA_RCFOP_R[chave][2]
                v14 = ABA_RCFOP_R[chave][3]
            else:
                v12 = 0 
                v13 = 0
                v14 = 0
            if chave in chavesP:
                v17 = ABA_RCFOP_P[chave][1]
                v18 = ABA_RCFOP_P[chave][2]
                v19 = ABA_RCFOP_P[chave][3]
            else:
                v17 = 0 
                v18 = 0
                v19 = 0
        
            ABA_RCFOP.cell(row=l, column=11,  value=chave)
            ABA_RCFOP.cell(row=l, column=12,  value=v12)
            ABA_RCFOP.cell(row=l, column=13,  value=v13)
            ABA_RCFOP.cell(row=l, column=14,  value=v14)
            ABA_RCFOP.cell(row=l, column=16,  value=chave)
            ABA_RCFOP.cell(row=l, column=17,  value=v17)
            ABA_RCFOP.cell(row=l, column=18,  value=v18)
            ABA_RCFOP.cell(row=l, column=19,  value=v19)
            ABA_RCFOP.cell(row=l, column=21,  value=v12-v17)
            ABA_RCFOP.cell(row=l, column=22,  value=v13-v18)
            ABA_RCFOP.cell(row=l, column=23,  value=v14-v19)
            l=l+1


        ########## TOTAIS    
            
        l=l+1
        ABA_RCFOP.cell(row=l, column=11,  value="ENTRADA")
        ABA_RCFOP.cell(row=l, column=12,  value=somale)
        ABA_RCFOP.cell(row=l, column=13,  value=somame)
        ABA_RCFOP.cell(row=l, column=14,  value=somane)
        ABA_RCFOP.cell(row=l, column=16,  value="ENTRADA")
        ABA_RCFOP.cell(row=l, column=17,  value=somaqe)
        ABA_RCFOP.cell(row=l, column=18,  value=somare)
        ABA_RCFOP.cell(row=l, column=19,  value=somase)
 
        ABA_RCFOP.cell(row=l, column=21,  value=somale-somaqe)
        ABA_RCFOP.cell(row=l, column=22,  value=somame-somare)
        ABA_RCFOP.cell(row=l, column=23,  value=somane-somase)
        l=l+2
        ABA_RCFOP.cell(row=l, column=11,  value="SAIDA")
        ABA_RCFOP.cell(row=l, column=12,  value=somals)
        ABA_RCFOP.cell(row=l, column=13,  value=somams)
        ABA_RCFOP.cell(row=l, column=14,  value=somans)
        ABA_RCFOP.cell(row=l, column=16,  value="SAIDA")
        ABA_RCFOP.cell(row=l, column=17,  value=somaqs)
        ABA_RCFOP.cell(row=l, column=18,  value=somars)
        ABA_RCFOP.cell(row=l, column=19,  value=somass)

        ABA_RCFOP.cell(row=l, column=21,  value=somals-somaqs)
        ABA_RCFOP.cell(row=l, column=22,  value=somams-somars)
        ABA_RCFOP.cell(row=l, column=23,  value=somans-somass)

        formata_RCFOP_DIF(ABA_RCFOP)
        arquivo_excel.save(ARQ_INSUMOS)
            
    
    #ABA_AICMS    = "Ajustes ICMS"                   ,10)    
    ##########  ABA 11-   ABA_AICMS    = "Ajustes ICMS - RF01.11"  
    ##########  ABA 11-   ABA_AICMS    = "Ajustes ICMS - RF01.11"  
    ##########  ABA 11-   ABA_AICMS    = "Ajustes ICMS - RF01.11"  
    ##########  ABA 11-   ABA_AICMS    = "Ajustes ICMS - RF01.11"  
    ##########  ABA 11-   ABA_AICMS    = "Ajustes ICMS - RF01.11"  
    ##########  ABA 11-   ABA_AICMS    = "Ajustes ICMS - RF01.11"  

    if ('11' in listadeabas):        
        log( "ABA 11 / ", qtdabas , " -> AJUSTES_ICMS")
    
        chaves = []
        dadosABA_AICMS = Busca_AICMS(iei,"01/"+mesi+"/"+anoi)
              
        l=1
        ABA_AICMS.cell(row=l, column=1,  value="Outros Ajustes Apuração ICMS - P9 (REGERADO)")
        l=l+1
     
        ABA_AICMS.cell(row=l, column=1,  value="EMPS_COD")
        ABA_AICMS.cell(row=l, column=2,  value="FILI_COD_INSEST")
        ABA_AICMS.cell(row=l, column=3,  value="CODIGO")
        ABA_AICMS.cell(row=l, column=4,  value="ITEM")
        ABA_AICMS.cell(row=l, column=5,  value="SEQ")
        ABA_AICMS.cell(row=l, column=6,  value="DATA")
        ABA_AICMS.cell(row=l, column=7,  value="GIA2_OCOR")
        ABA_AICMS.cell(row=l, column=8,  value="GIA2_VALOR")
        
        l=l+1
     
        for linha in dadosABA_AICMS:
            
            ABA_AICMS.cell(row=l, column=1,  value=linha[0])
            ABA_AICMS.cell(row=l, column=2,  value=linha[1])
            ABA_AICMS.cell(row=l, column=3,  value=linha[2])
            ABA_AICMS.cell(row=l, column=4,  value=linha[3])
            ABA_AICMS.cell(row=l, column=5,  value=linha[4])
            ABA_AICMS.cell(row=l, column=6,  value=linha[5])
            ABA_AICMS.cell(row=l, column=7,  value=linha[6])
            ABA_AICMS.cell(row=l, column=8,  value=linha[7])
            l=l+1
       
        formata_AICMS(ABA_AICMS)
    
    
        arquivo_excel.save(ARQ_INSUMOS)


    #ABA_SN    = "Salto de Notas"                            ,11)    
    ##########  ABA 12-   ABA_SN       = "Salto de Notas - RF01.12"  
    ##########  ABA 12-   ABA_SN       = "Salto de Notas - RF01.12"  
    ##########  ABA 12-   ABA_SN       = "Salto de Notas - RF01.12"  
    ##########  ABA 12-   ABA_SN       = "Salto de Notas - RF01.12"  
    ##########  ABA 12-   ABA_SN       = "Salto de Notas - RF01.12"  
    ##########  ABA 12-   ABA_SN       = "Salto de Notas - RF01.12"  

    if ('12' in listadeabas):        
        log( "ABA 12 / ", qtdabas , " -> SALTO_NOTAS")
    
        chaves = []
        dadosABA_SN = Busca_SN(iei,"01/"+mesi+"/"+anoi)
              
        l=1
        ABA_SN.cell(row=l, column=1,  value="TELECOM -> SALTO DE NOTAS")
        l=l+1
     
        ABA_SN.cell(row=l, column=1,  value="SERIE")
        ABA_SN.cell(row=l, column=2,  value="INICIO_SALTO")
        ABA_SN.cell(row=l, column=3,  value="FIM_SALTO")
        ABA_SN.cell(row=l, column=4,  value="PERIODO")
        ABA_SN.cell(row=l, column=5,  value="PERIODO_INICIAL")
        ABA_SN.cell(row=l, column=6,  value="PERIODO_FINAL")
        
        l=l+1
     
        for linha in dadosABA_SN:
            
            ABA_SN.cell(row=l, column=1,  value=linha[0])
            ABA_SN.cell(row=l, column=2,  value=linha[1])
            ABA_SN.cell(row=l, column=3,  value=linha[2])
            ABA_SN.cell(row=l, column=4,  value=linha[3])
            ABA_SN.cell(row=l, column=5,  value=linha[4])
            ABA_SN.cell(row=l, column=6,  value=linha[5])

            l=l+1
       
        formata_SN(ABA_SN)
    
    
        arquivo_excel.save(ARQ_INSUMOS)

  
    #ABA_R1600    = "Registro_1600"                            ,12)    
    ##########  ABA 13-   ABA_R1600    = "Registro_1600  - RF01.13"  
    ##########  ABA 13-   ABA_R1600    = "Registro_1600  - RF01.13"  
    ##########  ABA 13-   ABA_R1600    = "Registro_1600  - RF01.13"  
    ##########  ABA 13-   ABA_R1600    = "Registro_1600  - RF01.13"  
    ##########  ABA 13-   ABA_R1600    = "Registro_1600  - RF01.13"  
    ##########  ABA 13-   ABA_R1600    = "Registro_1600  - RF01.13"  

    if ('13' in listadeabas):       
        log( "ABA 13 / ", qtdabas , " -> REGISTRO_1600")
        anomesi = anoi+mesi    
        if( anomesi >= '201708' ):
            ARQ = ARQ_PRO
        else:
            ARQ = ARQ_ENX

        enc = encodingDoArquivo(ARQ)
        ARQ = open(ARQ, 'r', encoding=enc, errors='ignore')
        linhaABA_R1600 = {}
        chaves = []
        
        for linha in ARQ:
            camposlinha = linha.split("|")
            
            if linha.startswith('|1600|') :
                camposlinha = linha.split('|')
                agrupamento = camposlinha[2].upper()
                
                if (agrupamento in chaves):
                    linhaABA_R1600[agrupamento] = (linhaABA_R1600[agrupamento][0],
                                                linhaABA_R1600[agrupamento][1]  + (0.00 if (camposlinha[3]  == "") else float(camposlinha[3].replace("," , "."))),
                                                linhaABA_R1600[agrupamento][2]  + (0.00 if (camposlinha[4]  == "") else float(camposlinha[4].replace("," , "."))))
                else:
                    linhaABA_R1600[agrupamento] = ( camposlinha[2],
                                                    (0.00 if (camposlinha[3]  == "") else float(camposlinha[3].replace("," , "."))),
                                                    (0.00 if (camposlinha[4]  == "") else float(camposlinha[4].replace("," , "."))))
                    chaves.append(agrupamento)
        ARQ.close()
        l=1
        ABA_R1600.cell(row=l, column=1,  value="REGISTRO_1600")
        l=l+1
        ABA_R1600.cell(row=l, column=1,  value="COD_PART")
        ABA_R1600.cell(row=l, column=2,  value="TOT_CREDITO")
        ABA_R1600.cell(row=l, column=3,  value="TOT_DEBITO")
        l=l+1
        chaves.sort()
    
        for chave in chaves:        
            ABA_R1600.cell(row=l, column=1,  value=linhaABA_R1600[chave][0])
            ABA_R1600.cell(row=l, column=2,  value=linhaABA_R1600[chave][1])
            ABA_R1600.cell(row=l, column=3,  value=linhaABA_R1600[chave][2])
            l=l+1

        somab="=SUM(B3:B" + str(l-1) + ")"
        somac="=SUM(C3:C" + str(l-1) + ")"
        ABA_R1600.cell(row=l, column=1,  value="TOTAIS")
        ABA_R1600.cell(row=l, column=2,  value=somab)
        ABA_R1600.cell(row=l, column=3,  value=somac)
    
        formata_R1600(ABA_R1600)
        arquivo_excel.save(ARQ_INSUMOS)





    ##########  ABA 14-   ABA_HME       = "Historico Entradas - Mestre"    ,13) RF01.14
    ##########  ABA 14-   ABA_HME       = "Historico Entradas - Mestre"    ,13) RF01.14
    ##########  ABA 14-   ABA_HME       = "Historico Entradas - Mestre"    ,13) RF01.14
    ##########  ABA 14-   ABA_HME       = "Historico Entradas - Mestre"    ,13) RF01.14
    ##########  ABA 14-   ABA_HME       = "Historico Entradas - Mestre"    ,13) RF01.14
    ##########  ABA 14-   ABA_HME       = "Historico Entradas - Mestre"    ,13) RF01.14

    if ('14' in listadeabas):        
        log( "ABA 14 / ", qtdabas , " -> HISTORICO - MESTRE - ENTRADAS")
    
        chaves = []
        dadosABA_HME = Busca_HME(iei,"01/"+mesi+"/"+anoi)
              
        l=1
        ABA_HME.cell(row=l, column=1,  value="HISTORICO - MESTRE - ENTRADAS")
        l=l+1

        ABA_HME.cell(row=l, column=1,   value="NEGOGIO")
        ABA_HME.cell(row=l, column=2,   value="GAP")
        ABA_HME.cell(row=l, column=3,   value="ACAO")
        ABA_HME.cell(row=l, column=4,   value="TABELA")
        ABA_HME.cell(row=l, column=5,   value="CONTROLE")
        ABA_HME.cell(row=l, column=6,   value="DATA_CONTROLE")
        ABA_HME.cell(row=l, column=7,   value="TIPO_CONTROLE")
        ABA_HME.cell(row=l, column=8,   value="ROW_ID")
        ABA_HME.cell(row=l, column=9,   value="EMPS_COD")
        ABA_HME.cell(row=l, column=10,   value="FILI_COD")
        ABA_HME.cell(row=l, column=11,   value="TDOC_COD")
        ABA_HME.cell(row=l, column=12,   value="MNFEM_SERIE")
        ABA_HME.cell(row=l, column=13,  value="MNFEM_NUM")
        ABA_HME.cell(row=l, column=14,  value="MNFEM_DTEMIS")
        ABA_HME.cell(row=l, column=15,  value="MNFEM_IND_CONT")
        ABA_HME.cell(row=l, column=16,  value="CATG_COD")
        ABA_HME.cell(row=l, column=17,  value="CADG_COD")
        ABA_HME.cell(row=l, column=18,  value="MDOC_COD")
        ABA_HME.cell(row=l, column=19,  value="MNFEM_DTENTR")
        ABA_HME.cell(row=l, column=20,  value="MNFEM_VAL_TOT")
        ABA_HME.cell(row=l, column=21,  value="MNFEM_VAL_NF")
        ABA_HME.cell(row=l, column=22,  value="MNFEM_VAL_DESC")
        ABA_HME.cell(row=l, column=23,  value="MNFEM_NUM_NFREF")
        ABA_HME.cell(row=l, column=24,  value="MNFEM_SERIE_NFREF")
        ABA_HME.cell(row=l, column=25,  value="MNFEM_NUM_DECL")
        ABA_HME.cell(row=l, column=26,  value="MNFEM_VAL_REDIPI")
        ABA_HME.cell(row=l, column=27,  value="MNFEM_VAL_TOTIPI")
        ABA_HME.cell(row=l, column=28,  value="MNFEM_INSEST_SUBST")
        ABA_HME.cell(row=l, column=29,  value="MNFEM_OBSIPI")
        ABA_HME.cell(row=l, column=30,  value="MNFEM_INDCONTR")
        ABA_HME.cell(row=l, column=31,  value="MNFEM_IND_CANC")
        ABA_HME.cell(row=l, column=32,  value="MNFEM_AVISTA")
        ABA_HME.cell(row=l, column=33,  value="MNFEM_NF_PROPRIA")
        ABA_HME.cell(row=l, column=34,  value="NUM01")
        ABA_HME.cell(row=l, column=35,  value="NUM02")
        ABA_HME.cell(row=l, column=36,  value="NUM03")
        ABA_HME.cell(row=l, column=37,  value="VAR01")
        ABA_HME.cell(row=l, column=38,  value="VAR02")
        ABA_HME.cell(row=l, column=39,  value="VAR03")
        ABA_HME.cell(row=l, column=40,  value="VAR04")
        ABA_HME.cell(row=l, column=41,  value="VAR05")
        ABA_HME.cell(row=l, column=42,  value="MNFEM_REG")
        ABA_HME.cell(row=l, column=43,  value="MNFEM_CHV_NFE")
        ABA_HME.cell(row=l, column=44,  value="MNFEM_VL_ABAT_NT")
        ABA_HME.cell(row=l, column=45,  value="MFRT_COD")
        ABA_HME.cell(row=l, column=46,  value="MNFEM_VL_FRT")
        ABA_HME.cell(row=l, column=47,  value="MNFEM_VL_SEG")
        ABA_HME.cell(row=l, column=48,  value="MNFEM_VL_OUT_DA")
        ABA_HME.cell(row=l, column=49,  value="MNFEM_VL_BC_ICMS")
        ABA_HME.cell(row=l, column=50,  value="MNFEM_VL_ICMS")
        ABA_HME.cell(row=l, column=51,  value="MNFEM_BC_ICMS_ST")
        ABA_HME.cell(row=l, column=52,  value="MNFEM_VL_ICMS_ST")
        ABA_HME.cell(row=l, column=53,  value="MNFEM_VL_PIS")
        ABA_HME.cell(row=l, column=54,  value="MNFEM_VL_COFINS")
        ABA_HME.cell(row=l, column=55,  value="MNFEM_VL_PIS_ST")
        ABA_HME.cell(row=l, column=56,  value="MNFEM_COFINS_ST")
        ABA_HME.cell(row=l, column=57,  value="MNFEM_IND_COMPL")
        ABA_HME.cell(row=l, column=58,  value="MNFEM_DEN_IN")
        ABA_HME.cell(row=l, column=59,  value="TRANS_PES_BRUTO")
        ABA_HME.cell(row=l, column=60,  value="TRANS_PES_LIQ")
        ABA_HME.cell(row=l, column=61,  value="TRANS_QTD_VOL")
        ABA_HME.cell(row=l, column=62,  value="MNFEM_HR_EMIS")
        ABA_HME.cell(row=l, column=63,  value="MNFEM_TIP_ASSI")
        ABA_HME.cell(row=l, column=64,  value="MNFEM_TIP_UTIL")
        ABA_HME.cell(row=l, column=65,  value="MNFEM_GRP_TENS")
        ABA_HME.cell(row=l, column=66,  value="MNFEM_IND_EXTEMP")
        ABA_HME.cell(row=l, column=67,  value="MNFEM_DAT_EXTEMP")
        ABA_HME.cell(row=l, column=68,  value="MNFEM_TP_CTE")
        ABA_HME.cell(row=l, column=69,  value="MNFEM_NUM_DRWBCK")
        ABA_HME.cell(row=l, column=70,  value="MNFEM_VL_FCP")
        ABA_HME.cell(row=l, column=71,  value="MNFEM_IND_FO")
        ABA_HME.cell(row=l, column=72,  value="MNFEM_SIT_EDOC")
        ABA_HME.cell(row=l, column=73,  value="MNFEM_TIP_AQUI")
        ABA_HME.cell(row=l, column=74,  value="MNFEM_VLD_INSS")
        ABA_HME.cell(row=l, column=75,  value="MNFEM_VLD_GILRAT")
        ABA_HME.cell(row=l, column=76,  value="MNFEM_NUM_JUD")
        ABA_HME.cell(row=l, column=77,  value="MNFEM_VLD_SENAR")
        ABA_HME.cell(row=l, column=78,  value="MNFEM_DAT_LAN_PC")
        ABA_HME.cell(row=l, column=79,  value="MNFEM_FCP_DEST")
        ABA_HME.cell(row=l, column=80,  value="MNFEM_ICMS_DEST")
        ABA_HME.cell(row=l, column=81,  value="MNFEM_ICMS_ORIG")
        ABA_HME.cell(row=l, column=82,  value="MNFEM_VL_CPDAPPR")
        ABA_HME.cell(row=l, column=83,  value="MNFEM_VL_CFBCRAT")
        ABA_HME.cell(row=l, column=84,  value="MNFEM_SENAR_APPR")
        ABA_HME.cell(row=l, column=85,  value="MNFEM_DAT_AUT_NFE")
        ABA_HME.cell(row=l, column=86,  value="MNFEM_DAT_LT_ANT")
        ABA_HME.cell(row=l, column=87,  value="MNFEM_DAT_LT_ATU")
        ABA_HME.cell(row=l, column=88,  value="MNFEM_NUM_FAT")
        ABA_HME.cell(row=l, column=89,  value="MNFEM_VL_TOT_FAT")
        ABA_HME.cell(row=l, column=90,  value="MNFEN_BAS_RET_INSS")
        ABA_HME.cell(row=l, column=91,  value="ALIQ_INSS")
        ABA_HME.cell(row=l, column=92,  value="MNFEN_VL_RET_INSS")
        ABA_HME.cell(row=l, column=93,  value="MNFEN_TOT_DED")
        ABA_HME.cell(row=l, column=94,  value="MNFEN_RET_ADIC")
        ABA_HME.cell(row=l, column=95,  value="MNFEN_TOT_RET_ADIC")
        ABA_HME.cell(row=l, column=96,  value="MNFEN_TOT_INSS")
        ABA_HME.cell(row=l, column=97,  value="MNFEN_TOT_ESP")
        ABA_HME.cell(row=l, column=98,  value="SIST_ORIGEM")
        ABA_HME.cell(row=l, column=99,  value="USUA_ORIGEM")
        ABA_HME.cell(row=l, column=100,  value="DATA_CRIACAO")
        ABA_HME.cell(row=l, column=101,  value="ID_ORIGEM")
        ABA_HME.cell(row=l, column=102,  value="MNFEM_DT_ATUA")
        ABA_HME.cell(row=l, column=103, value="MNFEN_IND_OB_CIVIL")
        ABA_HME.cell(row=l, column=104, value="MNFEN_CEI")
        ABA_HME.cell(row=l, column=105, value="MIBGE_COD_ORIGEM")
        ABA_HME.cell(row=l, column=106, value="MIBGE_COD_DEST")
        ABA_HME.cell(row=l, column=107, value="MNFEN_FIN_DOCE")
        ABA_HME.cell(row=l, column=108, value="MNFEN_IND_DEST")



        l=l+1
     
        for linha in dadosABA_HME:
            #for i in range(0,len(linha)-1):
            for i in range(0,len(linha)):
                ABA_HME.cell(row=l, column=i+1,  value=linha[i])
            l=l+1
       
        formata_M(ABA_HME)
    
    
        arquivo_excel.save(ARQ_INSUMOS)



    #(15    "Historico Entradas - Item"      ,14) RF01.15
    ##########  ABA 15-   ABA_HEI       = "Historico Entradas - Item"      ,14) RF01.15
    ##########  ABA 15-   ABA_HEI       = "Historico Entradas - Item"      ,14) RF01.15
    ##########  ABA 15-   ABA_HEI       = "Historico Entradas - Item"      ,14) RF01.15
    ##########  ABA 15-   ABA_HEI       = "Historico Entradas - Item"      ,14) RF01.15
    ##########  ABA 15-   ABA_HEI       = "Historico Entradas - Item"      ,14) RF01.15
    ##########  ABA 15-   ABA_HEI       = "Historico Entradas - Item"      ,14) RF01.15

    if ('15' in listadeabas):        
        log( "ABA 15 / ", qtdabas , " -> HISTORICO - ITENS - ENTRADA")
    
        chaves = []
        dadosABA_HIE = Busca_HIE(iei,"01/"+mesi+"/"+anoi)
              
        l=1
        ABA_HIE.cell(row=l, column=1,  value="HISTORICO - ITENS - ENTRADA")
        l=l+1

        listadecabecalho = (
            'NEGOGIO'               ,'GAP'                   ,'ACAO'                  ,'TABELA'                ,'CONTROLE'              ,'DATA_CONTROLE'         ,'TIPO_CONTROLE'         ,'ROW_ID'                ,'EMPS_COD'              ,'FILI_COD'              ,'CGC_CPF'               ,'IE'                    ,'UF'                    ,'TP_LOC'                ,'LOCALIDADE'            ,'MOD_DOC'               ,'TDOC_COD'              ,'IND_CANC'              ,'IND_SIT'               ,'IND_TRIB_SUBSTRIB'     ,'VAL_ISEN_SUBSTRIB'     ,'VAL_OUTR_SUBSTRIB'     ,'INFEM_SERIE'           ,'INFEM_NUM'             ,'INFEM_DTEMIS'          ,'INFEM_DTENTR'          ,
            'CATG_COD'              ,'CADG_COD'              ,'INFEM_NUM_ITEM'        ,'MATE_COD'              ,'UNID_COD_VENDA'        ,'INFEM_DSC'             ,'CCUS_COD'              ,'CFOP_COD'              ,'CNBM_COD'              ,'NOPE_COD'              ,'UNID_COD_PADRAO'       ,'INFEM_QTD'             ,'INFEM_PES_LIQ'         ,'INFEM_VAL_PRECOUNIT'   ,'INFEM_VAL_PRECOTOT'    ,'INFEM_VAL_DESC'        ,'INFEM_NUM_ROMA'        ,'INFEM_DAT_ROMA'        ,'INFEM_VAL_FRETE'       ,'INFEM_VAL_SEGUR'       ,'INFEM_VAL_DESP'        ,'FEDE_COD'              ,'INFEM_TRIBIPI'         ,'INFEM_ALIQ_IPI'        ,'INFEM_BAS_IPI'         ,'INFEM_VAL_IPI'         ,
            'ESTA_COD'              ,'ESTB_COD'              ,'INFEM_TRIBICM'         ,'INFEM_ALIQ_ICMS'       ,'INFEM_BAS_ICMS'        ,'INFEM_VAL_ICMS'        ,'INFEM_BASSUBST_ICMS'   ,'INFEM_VALSUBST_ICMS'   ,'INFEM_VAL_REDICMS'     ,'INFEM_ALIQ_DIFICMS'    ,'INFEM_ISENTA_IPI'      ,'INFEM_ISENTA_ICMS'     ,'INFEM_OUTRA_IPI'       ,'INFEM_OUTRA_ICMS'      ,'INFEM_VAL_CONT'        ,'INFEM_COD_CONT'        ,'INFEM_RURAL'           ,'INFEM_PETROLEO'        ,'INFEM_CHASSI'          ,'INFEM_IND_MOV'         ,'INFEM_VAL_REDIPI'      ,'NUM01'                 ,'NUM02'                 ,'NUM03'                 ,'VAR01'                 ,'VAR02'                 ,
            'VAR03'                 ,'VAR04'                 ,'VAR05'                 ,'UNIN_COD'              ,'TIPI_COD'              ,'LIPI_COD'              ,'INFEM_ALIQ_ST'         ,'INFEM_IND_NAT'         ,'INFEM_SIMPLES_NAC'     ,'INFEM_NUM_FCI'         ,'INFEM_VLR_ADN'         ,'INFEM_FCP_DEST'        ,'INFEM_ICMS_DEST'       ,'INFEM_ICMS_ORIG'       ,'INFEM_COD_RRST'        ,'INFEM_COD_MRES'        ,'INFEM_MOD_DARR'        ,'INFEM_DOC_ARRE'        ,'INFEM_BC_ICMS_EMIT'    ,'INFEM_AL_ICMS_EMIT'    ,'INFEM_NUM_RET'         ,'INFEM_SER_RET'         ,'INFEM_DAT_RET'         ,'INFEM_ITEM_RET'        ,'CATG_COD_RET'          ,'CADG_COD_RET'          ,
            'INFEM_CHAVE_SUBT'      ,'INFEM_NUM_CONTR'       ,'INFEM_TIP_ISENCAO'     ,'INFEM_TAR_APLIC'       ,'INFEM_IND_DESC'        ,'INFEM_QTD_FAT'         ,'INFEM_UF_PREST'        ,'CPRB_COD'              ,'TSCP_COD'              ,'INFEN_CALC_RET_INSS'   ,'INFEN_ALIQ_INSS'       ,'INFEN_VL_RET_INSS'     ,'INFEN_VL_DED_MAT'      ,'INFEN_DED_ALIMENT'     ,'INFEN_VL_DED_TRANSP'   ,'INFEN_VL_DED_RET'      ,'INFEN_VL_RET_ADIC'     ,'INFEN_VL_RET_ADIC_N'   ,'INFEN_VL_INSS_N_RET'   ,'INFEN_VL_COND_ESP_15'  ,'INFEN_VL_COND_ESP_20'  ,'INFEN_VL_COND_ESP_25'  ,'SIST_ORIGEM'           ,'USUA_ORIGEM'           ,'DATA_CRIACAO'          ,'ID_ORIGEM'             ,
            'INFEM_FCP_PRO'         ,'INFEM_FCP_RET'         ,'INFEM_FCP_ST'          ,'INFEM_ABA_NT'          ,'CAIA_COD'              ,'INFEM_QTD_CONV'        ,'INFEM_UNID'            ,'INFEM_VLR_CONV'        ,'INFEM_ICMS_CONV'       ,'INFEM_BC_ICMS_ST_CONV' ,'INFEM_ICMS_ST_CONV'    ,'INFEM_FCP_ST_CONV'     ,'INFEM_VAL_IPI_NESC')
#EM          A                        B                        C                        D                        E                        F                        G                        H                        I                        J                        K                        L                        M                        N                        O                        P                        Q                        R                        S                        T                        U                        V                        W                        X                        Y                        Z                     

        for c in range(1, len(listadecabecalho)+1):
            ABA_HIE.cell(row=l, column=c,   value=listadecabecalho[c-1])    

        l=l+1
     
        for linha in dadosABA_HIE:
            #for i in range(0,len(linha)-1):
            for i in range(0,len(linha)):
                ABA_HIE.cell(row=l, column=i+1,  value=linha[i])
            l=l+1
        #formata_HIE(ABA_HIE,ARQ_INSUMOS,arquivo_excel )
        formata_M(ABA_HIE)
        arquivo_excel.save(ARQ_INSUMOS)



    ########## "ABA 16 / ", qtdabas , " - HISTORICO - MESTRE - SAIDA"
    ########## "ABA 16 / ", qtdabas , " - HISTORICO - MESTRE - SAIDA"
    ########## "ABA 16 / ", qtdabas , " - HISTORICO - MESTRE - SAIDA"
    ########## "ABA 16 / ", qtdabas , " - HISTORICO - MESTRE - SAIDA"
    ########## "ABA 16 / ", qtdabas , " - HISTORICO - MESTRE - SAIDA"
    ########## "ABA 16 / ", qtdabas , " - HISTORICO - MESTRE - SAIDA"

    if ('16' in listadeabas):        
        log( "ABA 16 / ", qtdabas , " -> HISTORICO - MESTRE - SAIDA")
    
        chaves = []
        dadosABA_HMS = Busca_HMS(iei,"01/"+mesi+"/"+anoi)
              
        l=1
        ABA_HMS.cell(row=l, column=1,  value="HISTORICO - MESTRE - SAIDA")
        l=l+1


        listadecabecalho = (
            'NEGOGIO'               ,'GAP'                   ,'ACAO'                  ,'TABELA'                ,'CONTROLE'              ,'DATA_CONTROLE'         ,'TIPO_CONTROLE'         ,'ROW_ID'                ,'EMPS_COD'              ,'FILI_COD'              ,'TDOC_COD'              ,'MNFSM_SERIE'           ,'MNFSM_NUM'             ,'MNFSM_DTEMISS'         ,'CATG_COD'              ,'CADG_COD'              ,'MNFSM_IND_CONT'        ,'MDOC_COD'              ,'MNFSM_DTSAIDA'         ,'MNFSM_VAL_TOTPROD'     ,'MNFSM_VAL_TOTNF'       ,'MNFSM_VAL_DESC'        ,'MNFSM_NUMNFREF'        ,'MNFSM_SERIENFREF'      ,'MNFSM_VAL_REDIPI'      ,'MNFSM_VAL_TOTIPI'      ,
            'MNFSM_INSEST_SUBST'    ,'MNFSM_OBSIPI'          ,'MNFSM_IND_CONTR'       ,'MNFSM_IND_CANC'        ,'MNFSM_AVISTA'          ,'CADG_COD_TRANSP'       ,'TRANS_VAL_FRETE'       ,'TRANS_VAL_SEGUR'       ,'MFRT_COD'              ,'TRANS_PES_BRUTO'       ,'TRANS_PES_LIQ'         ,'VTRP_COD'              ,'TRANS_QTD_VOL'         ,'EVOL_COD'              ,'TRANS_IDENT'           ,'NUM01'                 ,'NUM02'                 ,'NUM03'                 ,'VAR01'                 ,'VAR02'                 ,'VAR03'                 ,'VAR04'                 ,'VAR05'                 ,'MNFSM_REG'             ,'MNFSM_DEN_IN'          ,'MNFSM_CHV_NFE'         ,
            'MNFSM_VL_ABAT_NT'      ,'MNFSM_VL_OUT_DA'       ,'MNFSM_VL_BC_ICMS'      ,'MNFSM_VL_ICMS'         ,'MNFSM_BC_ICMS_ST'      ,'MNFSM_VL_ICMS_ST'      ,'MNFSM_VL_IPI'          ,'MNFSM_VL_PIS'          ,'MNFSM_VL_COFINS'       ,'MNFSM_VL_PIS_ST'       ,'MNFSM_COFINS_ST'       ,'MNFSM_IND_COMPL'       ,'UNFE_SIG_VEICULO'      ,'MNFSM_HR_EMIS'         ,'MNFSM_IND_EXTEMP'      ,'MNFSM_DAT_EXTEMP'      ,'MNFSM_VL_FCP'          ,'MNFSM_IND_FO'          ,'MNFSM_SIT_EDOC'        ,'MNFSM_DAT_COS'         ,'MNFSM_TIP_COM'         ,'MNFSM_VLD_INSS'        ,'MNFSM_VLD_GILRAT'      ,'MNFSM_VLD_SENAR'       ,'MNFSM_FCP_DEST'        ,'MNFSM_ICMS_DEST'       ,
            'MNFSM_ICMS_ORIG'       ,'SIST_ORIGEM'           ,'USUA_ORIGEM'           ,'DATA_CRIACAO'          ,'ID_ORIGEM')
#CE          A                        B                        C                        D                        E                        F                        G                        H                        I                        J                        K                        L                        M                        N                        O                        P                        Q                        R                        S                        T                        U                        V                        W                        X                        Y                        Z                     

        for c in range(1, len(listadecabecalho)+1):
            ABA_HMS.cell(row=l, column=c,   value=listadecabecalho[c-1])    

        l=l+1
     
        for linha in dadosABA_HMS:
            #for i in range(0,len(linha)-1):
            for i in range(0,len(linha)):
                ABA_HMS.cell(row=l, column=i+1,  value=linha[i])
            l=l+1
       
        formata_M(ABA_HMS)
    
    
        arquivo_excel.save(ARQ_INSUMOS)

    ##########"ABA 17 / ", qtdabas , " - HISTORICO - ITENS - SAIDA"
    ##########"ABA 17 / ", qtdabas , " - HISTORICO - ITENS - SAIDA"
    ##########"ABA 17 / ", qtdabas , " - HISTORICO - ITENS - SAIDA"
    ##########"ABA 17 / ", qtdabas , " - HISTORICO - ITENS - SAIDA"
    ##########"ABA 17 / ", qtdabas , " - HISTORICO - ITENS - SAIDA"
    ##########"ABA 17 / ", qtdabas , " - HISTORICO - ITENS - SAIDA"

    if ('17' in listadeabas):        
        log( "ABA 17 / ", qtdabas , " -> HISTORICO - ITENS - SAIDA")
    
        chaves = []
        dadosABA_HIS = Busca_HIS(iei,"01/"+mesi+"/"+anoi)
              
        l=1
        ABA_HIS.cell(row=l, column=1,  value="HISTORICO - ITENS - SAIDA")
        l=l+1

        listadecabecalho = (
            'NEGOGIO'               ,'GAP'                   ,'ACAO'                  ,'TABELA'                ,'CONTROLE'              ,'DATA_CONTROLE'         ,'TIPO_CONTROLE'         ,'ROW_ID'                ,'EMPS_COD'              ,'FILI_COD'              ,'CGC_CPF'               ,'IE'                    ,'UF'                    ,'TP_LOC'                ,'LOCALIDADE'            ,'MOD_DOC'               ,'TDOC_COD'              ,'IND_CANC'              ,'IND_SIT'               ,'IND_TRIB_SUBSTRIB'     ,'VAL_ISEN_SUBSTRIB'     ,'VAL_OUTR_SUBSTRIB'     ,'INFSM_SERIE'           ,'INFSM_NUM'             ,'INFSM_DTEMISS'         ,'CATG_COD'              ,
            'CADG_COD'              ,'INFSM_NUM_ITEM'        ,'MATE_COD'              ,'UNID_COD_VENDA'        ,'INFSM_DSC'             ,'CCUS_COD'              ,'CFOP_COD'              ,'NOPE_COD'              ,'CNBM_COD'              ,'UNID_COD'              ,'INFSM_QTD'             ,'INFSM_PES_LIQ'         ,'INFSM_VAL_PRECOUNIT'   ,'INFSM_VAL_PRECOTOT'    ,'INFSM_VAL_DESC'        ,'INFSM_NUM_ROMA'        ,'INFSM_DAT_ROMA'        ,'INFSM_VAL_FRETE'       ,'INFSM_VAL_SEGUR'       ,'INFSM_VAL_DESP'        ,'FEDE_COD'              ,'INFSM_TRIBIPI'         ,'INFSM_ALIQ_IPI'        ,'INFSM_BAS_IPI'         ,'INFSM_VAL_IPI'         ,'ESTA_COD'              ,
            'ESTB_COD'              ,'INFSM_TRIBICM'         ,'INFSM_ALIQ_ICMS'       ,'INFSM_BAS_ICMS'        ,'INFSM_VAL_ICMS'        ,'INFSM_BASSUBST_ICMS'   ,'INFSM_VALSUBST_ICMS'   ,'INFSM_VAL_REDICMS'     ,'INFSM_ALIQ_DIFICMS'    ,'INFSM_ISENTA_IPI'      ,'INFSM_ISENTA_ICMS'     ,'INFSM_OUTRA_IPI'       ,'INFSM_OUTRA_ICMS'      ,'INFSM_VAL_CONT'        ,'INFSM_COD_CONT'        ,'INFSM_ICMS_FRETE'      ,'INFSM_CHASSI'          ,'INFSM_IND_MOV'         ,'INFSM_VAL_REDIPI'      ,'NUM01'                 ,'NUM02'                 ,'NUM03'                 ,'VAR01'                 ,'VAR02'                 ,'VAR03'                 ,'VAR04'                 ,
            'VAR05'                 ,'UNIN_COD'              ,'TIPI_COD'              ,'LIPI_COD'              ,'INFSM_ALIQ_ST'         ,'INFSM_NUM_FCI'         ,'INFSM_FCP_DEST'        ,'INFSM_ICMS_DEST'       ,'INFSM_ICMS_ORIG'       ,'INFSM_MOT_RESS'        ,'INFSM_FCP_PRO'         ,'INFSM_FCP_RET'         ,'INFSM_FCP_ST'          ,'INFSM_ABA_NT'          ,'CAIA_COD'              ,'SIST_ORIGEM'           ,'USUA_ORIGEM'           ,'DATA_CRIACAO'          ,'ID_ORIGEM'             ,'TMRC_COD'              ,'INFSM_QTD_CONV'        ,'INFSM_UNID'            ,'INFSM_VLR_CONV'        ,'INFSM_ICMS_CONV'       ,'INFSM_ICMS_OP_CONV'    ,'INFSM_BC_ICMS_ST_CONV' ,
            'INFSM_ICMS_ST_EST'     ,'INFSM_FCP_ST_EST'      ,'INFSM_ICMS_ST_REST'    ,'INFSM_FCP_ST_REST'     ,'INFSM_ICMS_ST_COMP'    ,'INFSM_FCP_ST_COMP')
#DF          A                        B                        C                        D                        E                        F                        G                        H                        I                        J                        K                        L                        M                        N                        O                        P                        Q                        R                        S                        T                        U                        V                        W                        X                        Y                        Z                     


        for c in range(1, len(listadecabecalho)+1):
            ABA_HIS.cell(row=l, column=c,   value=listadecabecalho[c-1])    

        l=l+1
     
        for linha in dadosABA_HIS:
            #for i in range(0,len(linha)-1):
            for i in range(0,len(linha)):
                ABA_HIS.cell(row=l, column=i+1,  value=linha[i])
            l=l+1
       
        formata_M(ABA_HIS)
    
    
        arquivo_excel.save(ARQ_INSUMOS)




 #### Cria a planilha excel MERCADORIA ENTRADA em memória....
    arquivo_excel_me = Workbook()
    ABA_EMM       = arquivo_excel_me.active
    ABA_EMM.title =                               "ENTRADA MERCADORIA - MESTRE"       # 0)
    ########## " -> ENTRADA MERCADORIA - MESTRE"  RF02.01
    ########## " -> ENTRADA MERCADORIA - MESTRE"  RF02.01
    ########## " -> ENTRADA MERCADORIA - MESTRE"  RF02.01
    ########## " -> ENTRADA MERCADORIA - MESTRE"  RF02.01
    ########## " -> ENTRADA MERCADORIA - MESTRE"  RF02.01
    ########## " -> ENTRADA MERCADORIA - MESTRE"  RF02.01
    if ('18' in listadeabas):
        log( "ABA 18 / ", qtdabas , " -> ENTRADA MERCADORIA - MESTRE")
        chaves = []
        dadosABA_EMM = Busca_EMM(iei,"01/"+mesi+"/"+anoi)
        l=1
        ABA_EMM.cell(row=l, column=1,  value="ENTRADA MERCADORIA - MESTRE")
        l=l+1
        listadecabecalho = (
            'GAP'                  ,'ACAO'                 ,'TABELA'               ,'ROW_ID'               ,'EMPS_COD'             ,'FILI_COD'             ,'TDOC_COD'             ,'MNFEM_SERIE'          ,'MNFEM_NUM'            ,'MNFEM_DTEMIS'         ,'MNFEM_IND_CONT'       ,'CATG_COD'             ,'CADG_COD'             ,'MDOC_COD'             ,'MNFEM_DTENTR'         ,'MNFEM_VAL_TOT'        ,'MNFEM_VAL_NF'         ,'MNFEM_VAL_DESC'       ,'MNFEM_NUM_NFREF'      ,'MNFEM_SERIE_NFREF'    ,'MNFEM_NUM_DECL'       ,'MNFEM_VAL_REDIPI'     ,'MNFEM_VAL_TOTIPI'     ,'MNFEM_INSEST_SUBST'   ,'MNFEM_OBSIPI'         ,'MNFEM_INDCONTR'       ,
            'MNFEM_IND_CANC'       ,'MNFEM_AVISTA'         ,'MNFEM_NF_PROPRIA'     ,'NUM01'                ,'NUM02'                ,'NUM03'                ,'VAR01'                ,'VAR02'                ,'VAR03'                ,'VAR04'                ,'VAR05'                ,'MNFEM_REG'            ,'MNFEM_CHV_NFE'        ,'MNFEM_VL_ABAT_NT'     ,'MFRT_COD'             ,'MNFEM_VL_FRT'         ,'MNFEM_VL_SEG'         ,'MNFEM_VL_OUT_DA'      ,'MNFEM_VL_BC_ICMS'     ,'MNFEM_VL_ICMS'        ,'MNFEM_BC_ICMS_ST'     ,'MNFEM_VL_ICMS_ST'     ,'MNFEM_VL_PIS'         ,'MNFEM_VL_COFINS'      ,'MNFEM_VL_PIS_ST'      ,'MNFEM_COFINS_ST'      ,
            'MNFEM_IND_COMPL'      ,'MNFEM_DEN_IN'         ,'TRANS_PES_BRUTO'      ,'TRANS_PES_LIQ'        ,'TRANS_QTD_VOL'        ,'MNFEM_HR_EMIS'        ,'MNFEM_TIP_ASSI'       ,'MNFEM_TIP_UTIL'       ,'MNFEM_GRP_TENS'       ,'MNFEM_IND_EXTEMP'     ,'MNFEM_DAT_EXTEMP'     ,'MNFEM_TP_CTE'         ,'MNFEM_NUM_DRWBCK'     ,'MNFEM_VL_FCP'         ,'MNFEM_IND_FO'         ,'MNFEM_SIT_EDOC'       ,'MNFEM_TIP_AQUI'       ,'MNFEM_VLD_INSS'       ,'MNFEM_VLD_GILRAT'     ,'MNFEM_VLD_SENAR'      ,'MNFEM_NUM_JUD'        ,'MNFEM_DAT_LAN_PC'     ,'MNFEM_FCP_DEST'       ,'MNFEM_ICMS_DEST'      ,'MNFEM_ICMS_ORIG'      ,'MNFEM_VL_CPDAPPR'     ,
            'MNFEM_VL_CFBCRAT'     ,'MNFEM_SENAR_APPR'     ,'MNFEM_DAT_AUT_NFE'    ,'MNFEM_DAT_LT_ANT'     ,'MNFEM_DAT_LT_ATU'     ,'MNFEM_NUM_FAT'        ,'MNFEM_VL_TOT_FAT'     ,'MNFEN_BAS_RET_INSS'   ,'ALIQ_INSS'            ,'MNFEN_VL_RET_INSS'    ,'MNFEN_TOT_DED'        ,'MNFEN_RET_ADIC'       ,'MNFEN_TOT_RET_ADIC'   ,'MNFEN_TOT_INSS'       ,'MNFEN_TOT_ESP'        ,'SIST_ORIGEM'          ,'USUA_ORIGEM'          ,'DATA_CRIACAO'         ,'ID_ORIGEM'            ,'MNFEM_DT_ATUA'        ,'MNFEN_IND_OB_CIVIL'   ,'MNFEN_CEI'            ,'MIBGE_COD_ORIGEM'     ,'MIBGE_COD_DEST'       ,'MNFEN_FIN_DOCE'       ,'MNFEN_IND_DEST'       ,
            'EMPS_COD_1'           ,'FILI_COD_1'           ,'FILI_COD_CGC'         ,'UNFE_SIG'             ,'FILI_COD_INSEST'      ,'FILI_COD_INSMUN'      ,'FILI_COD_ATIVECON'    ,'FILI_NOM'             ,'FILI_NOM_FANTASIA'    ,'FILI_END'             ,'FILI_END_NUM'         ,'FILI_END_COMP'        ,'FILI_END_BAIRRO'      ,'FILI_END_MUNIC'       ,'FILI_END_CEP'         ,'FILI_COD_LOCAL'       ,'FILI_FAX'             ,'FILI_TEL'             ,'TP_LOC'               ,'FILI_SETOR'           ,'FILI_AUTOR'           ,'FILI_SANITARIA'       ,'FILI_LICEN'           ,'FILI_ESPEC'           ,'FILI_IND_RET_ISS'     ,'FILI_COD_INSCEN'      ,
            'FILI_EMAIL'           ,'FILI_NAT_JUR'         ,'NUM01_1'              ,'NUM02_1'              ,'NUM03_1'              ,'VAR01_1'              ,'VAR02_1'              ,'VAR03_1'              ,'VAR04_1'              ,'VAR05_1'              ,'FILI_CEI'             ,'FILI_NIT'             ,'FILI_SUFRAMA'         ,'FILI_MUN_IBGE'        ,'TPLD_COD'             ,'FILI_NIRE'            ,'FILI_MATRIZ'          ,'FILI_TIP_ATI_ECO'     ,'FILI_INFO_COMPL'      ,'FILI_CLAS_TRIB'       ,'FILI_IND_OB_CIVIL'    ,'FILI_COD_CNPJ_EFR'    ,'SIST_ORIGEM_1'        ,'USUA_ORIGEM_1'        ,'DATA_CRIACAO_1'       ,'ID_ORIGEM_1'          ,
            'FILI_CLAS_IND_EQ')
#EW          A                       B                       C                       D                       E                       F                       G                       H                       I                       J                       K                       L                       M                       N                       O                       P                       Q                       R                       S                       T                       U                       V                       W                       X                       Y                       Z
        for c in range(1, len(listadecabecalho)+1):
            ABA_EMM.cell(row=l, column=c,   value=listadecabecalho[c-1])    
        l=l+1
        for linha in dadosABA_EMM:
#            for i in range(0,len(linha)-1):
            for i in range(0,len(linha)):
                ABA_EMM.cell(row=l, column=i+1,  value=linha[i])
            l=l+1
        formata_M(ABA_EMM)
        arquivo_excel_me.save(ARQ_ME)    
    ########## " -> ENTRADA MERCADORIA - ITEM"  RF02.02
    ########## " -> ENTRADA MERCADORIA - ITEM"  RF02.02
    ########## " -> ENTRADA MERCADORIA - ITEM"  RF02.02
    ########## " -> ENTRADA MERCADORIA - ITEM"  RF02.02
    ########## " -> ENTRADA MERCADORIA - ITEM"  RF02.02
    ########## " -> ENTRADA MERCADORIA - ITEM"  RF02.02
    if ('19' in listadeabas):
        ABA_EMI       = arquivo_excel_me.create_sheet("ENTRADA MERCADORIA - ITEM",1)
        log( "ABA 19 / ", qtdabas , " -> ENTRADA MERCADORIA - ITEM")
        chaves = []
        dadosABA_EMI = Busca_EMI(iei,"01/"+mesi+"/"+anoi)
        l=1
        ABA_EMI.cell(row=l, column=1,  value="ENTRADA MERCADORIA - ITEM")
        l=l+1
        listadecabecalho = ('GAP'                  ,'ACAO'                 ,'TABELA'               ,'ROW_ID'               ,'EMPS_COD'             ,'FILI_COD'             ,'CGC_CPF'              ,'IE'                   ,'UF'                   ,'TP_LOC'               ,'LOCALIDADE'           ,'MOD_DOC'              ,'TDOC_COD'             ,'IND_CANC'             ,'IND_SIT'              ,'IND_TRIB_SUBSTRIB'    ,'VAL_ISEN_SUBSTRIB'    ,'VAL_OUTR_SUBSTRIB'    ,'INFEM_SERIE'          ,'INFEM_NUM'            ,'INFEM_DTEMIS'         ,'INFEM_DTENTR'         ,'CATG_COD'             ,'CADG_COD'             ,'INFEM_NUM_ITEM'       ,'MATE_COD'             ,
                            'UNID_COD_VENDA'       ,'INFEM_DSC'            ,'CCUS_COD'             ,'CFOP_COD'             ,'CNBM_COD'             ,'NOPE_COD'             ,'UNID_COD_PADRAO'      ,'INFEM_QTD'            ,'INFEM_PES_LIQ'        ,'INFEM_VAL_PRECOUNIT'  ,'INFEM_VAL_PRECOTOT'   ,'INFEM_VAL_DESC'       ,'INFEM_NUM_ROMA'       ,'INFEM_DAT_ROMA'       ,'INFEM_VAL_FRETE'      ,'INFEM_VAL_SEGUR'      ,'INFEM_VAL_DESP'       ,'FEDE_COD'             ,'INFEM_TRIBIPI'        ,'INFEM_ALIQ_IPI'       ,'INFEM_BAS_IPI'        ,'INFEM_VAL_IPI'        ,'ESTA_COD'             ,'ESTB_COD'             ,'INFEM_TRIBICM'        ,'INFEM_ALIQ_ICMS'      ,
                            'INFEM_BAS_ICMS'       ,'INFEM_VAL_ICMS'       ,'INFEM_BASSUBST_ICMS'  ,'INFEM_VALSUBST_ICMS'  ,'INFEM_VAL_REDICMS'    ,'INFEM_ALIQ_DIFICMS'   ,'INFEM_ISENTA_IPI'     ,'INFEM_ISENTA_ICMS'    ,'INFEM_OUTRA_IPI'      ,'INFEM_OUTRA_ICMS'     ,'INFEM_VAL_CONT'       ,'INFEM_COD_CONT'       ,'INFEM_RURAL'          ,'INFEM_PETROLEO'       ,'INFEM_CHASSI'         ,'INFEM_IND_MOV'        ,'INFEM_VAL_REDIPI'     ,'NUM01'                ,'NUM02'                ,'NUM03'                ,'VAR01'                ,'VAR02'                ,'VAR03'                ,'VAR04'                ,'VAR05'                ,'UNIN_COD'             ,
                            'TIPI_COD'             ,'LIPI_COD'             ,'INFEM_ALIQ_ST'        ,'INFEM_IND_NAT'        ,'INFEM_SIMPLES_NAC'    ,'INFEM_NUM_FCI'        ,'INFEM_VLR_ADN'        ,'INFEM_FCP_DEST'       ,'INFEM_ICMS_DEST'      ,'INFEM_ICMS_ORIG'      ,'INFEM_COD_RRST'       ,'INFEM_COD_MRES'       ,'INFEM_MOD_DARR'       ,'INFEM_DOC_ARRE'       ,'INFEM_BC_ICMS_EMIT'   ,'INFEM_AL_ICMS_EMIT'   ,'INFEM_NUM_RET'        ,'INFEM_SER_RET'        ,'INFEM_DAT_RET'        ,'INFEM_ITEM_RET'       ,'CATG_COD_RET'         ,'CADG_COD_RET'         ,'INFEM_CHAVE_SUBT'     ,'INFEM_NUM_CONTR'      ,'INFEM_TIP_ISENCAO'    ,'INFEM_TAR_APLIC'      ,
                            'INFEM_IND_DESC'       ,'INFEM_QTD_FAT'        ,'INFEM_UF_PREST'       ,'CPRB_COD'             ,'TSCP_COD'             ,'INFEN_CALC_RET_INSS'  ,'INFEN_ALIQ_INSS'      ,'INFEN_VL_RET_INSS'    ,'INFEN_VL_DED_MAT'     ,'INFEN_DED_ALIMENT'    ,'INFEN_VL_DED_TRANSP'  ,'INFEN_VL_DED_RET'     ,'INFEN_VL_RET_ADIC'    ,'INFEN_VL_RET_ADIC_N'  ,'INFEN_VL_INSS_N_RET'  ,'INFEN_VL_COND_ESP_15' ,'INFEN_VL_COND_ESP_20' ,'INFEN_VL_COND_ESP_25' ,'SIST_ORIGEM'          ,'USUA_ORIGEM'          ,'DATA_CRIACAO'         ,'ID_ORIGEM'            ,'INFEM_FCP_PRO'        ,'INFEM_FCP_RET'        ,'INFEM_FCP_ST'         ,'INFEM_ABA_NT'         ,
                            'CAIA_COD'             ,'INFEM_QTD_CONV'       ,'INFEM_UNID'           ,'INFEM_VLR_CONV'       ,'INFEM_ICMS_CONV'      ,'INFEM_BC_ICMS_ST_CONV','INFEM_ICMS_ST_CONV'   ,'INFEM_FCP_ST_CONV'    ,'INFEM_VAL_IPI_NESC'   ,'EMPS_COD_1'           ,'FILI_COD_1'           ,'FILI_COD_CGC'         ,'UNFE_SIG'             ,'FILI_COD_INSEST'      ,'FILI_COD_INSMUN'      ,'FILI_COD_ATIVECON'    ,'FILI_NOM'             ,'FILI_NOM_FANTASIA'    ,'FILI_END'             ,'FILI_END_NUM'         ,'FILI_END_COMP'        ,'FILI_END_BAIRRO'      ,'FILI_END_MUNIC'       ,'FILI_END_CEP'         ,'FILI_COD_LOCAL'       ,'FILI_FAX'             ,
                            'FILI_TEL'             ,'TP_LOC_1'             ,'FILI_SETOR'           ,'FILI_AUTOR'           ,'FILI_SANITARIA'       ,'FILI_LICEN'           ,'FILI_ESPEC'           ,'FILI_IND_RET_ISS'     ,'FILI_COD_INSCEN'      ,'FILI_EMAIL'           ,'FILI_NAT_JUR'         ,'NUM01_1'              ,'NUM02_1'              ,'NUM03_1'              ,'VAR01_1'              ,'VAR02_1'              ,'VAR03_1'              ,'VAR04_1'              ,'VAR05_1'              ,'FILI_CEI'             ,'FILI_NIT'             ,'FILI_SUFRAMA'         ,'FILI_MUN_IBGE'        ,'TPLD_COD'             ,'FILI_NIRE'            ,'FILI_MATRIZ'          ,
                            'FILI_TIP_ATI_ECO'     ,'FILI_INFO_COMPL'      ,'FILI_CLAS_TRIB'       ,'FILI_IND_OB_CIVIL'    ,'FILI_COD_CNPJ_EFR'    ,'SIST_ORIGEM_1'        ,'USUA_ORIGEM_1'        ,'DATA_CRIACAO_1'       ,'ID_ORIGEM_1'          ,'FILI_CLAS_IND_EQ')
                #GJ          A                       B                       C                       D                       E                       F                       G                       H                       I                       J                       K                       L                       M                       N                       O                       P                       Q                       R                       S                       T                       U                       V                       W                       X                       Y                       Z                     
        for c in range(1, len(listadecabecalho)+1):
            ABA_EMI.cell(row=l, column=c,   value=listadecabecalho[c-1])    
        l=l+1
        for linha in dadosABA_EMI:
#            for i in range(0,len(linha)-1):
            for i in range(0,len(linha)):
                ABA_EMI.cell(row=l, column=i+1,  value=linha[i])
            l=l+1
        formata_M(ABA_EMI)
        arquivo_excel_me.save(ARQ_ME)


    #### Cria a planilha excel MERCADORIA SAIDA em memória....
    arquivo_excel_ms = Workbook()
    ABA_SMM       = arquivo_excel_ms.active
    ABA_SMM.title =                               "SAIDA MERCADORIA - MESTRE"       # 0) 
    ########## " -> SAIDA MERCADORIA - MESTRE"  RF03.01
    ########## " -> SAIDA MERCADORIA - MESTRE"  RF03.01
    ########## " -> SAIDA MERCADORIA - MESTRE"  RF03.01
    ########## " -> SAIDA MERCADORIA - MESTRE"  RF03.01
    ########## " -> SAIDA MERCADORIA - MESTRE"  RF03.01
    ########## " -> SAIDA MERCADORIA - MESTRE"  RF03.01
    if ('20' in listadeabas):
        log( "ABA 20 / ", qtdabas , " -> SAIDA MERCADORIA - MESTRE")
        chaves = []
        dadosABA_SMM = Busca_SMM(iei,"01/"+mesi+"/"+anoi)
        l=1
        ABA_SMM.cell(row=l, column=1,  value="SAIDA MERCADORIA - MESTRE")
        l=l+1
        listadecabecalho = ( 'GAP'                  ,'ACAO'                 ,'TABELA'               ,'ROW_ID'               ,'EMPS_COD'             ,'FILI_COD'             ,'TDOC_COD'             ,'MNFSM_SERIE'          ,'MNFSM_NUM'            ,'MNFSM_DTEMISS'        ,'CATG_COD'             ,'CADG_COD'             ,'MNFSM_IND_CONT'       ,'MDOC_COD'             ,'MNFSM_DTSAIDA'        ,'MNFSM_VAL_TOTPROD'    ,'MNFSM_VAL_TOTNF'      ,'MNFSM_VAL_DESC'       ,'MNFSM_NUMNFREF'       ,'MNFSM_SERIENFREF'     ,'MNFSM_VAL_REDIPI'     ,'MNFSM_VAL_TOTIPI'     ,'MNFSM_INSEST_SUBST'   ,'MNFSM_OBSIPI'         ,'MNFSM_IND_CONTR'      ,'MNFSM_IND_CANC'       
                            ,'MNFSM_AVISTA'         ,'CADG_COD_TRANSP'      ,'TRANS_VAL_FRETE'      ,'TRANS_VAL_SEGUR'      ,'MFRT_COD'             ,'TRANS_PES_BRUTO'      ,'TRANS_PES_LIQ'        ,'VTRP_COD'             ,'TRANS_QTD_VOL'        ,'EVOL_COD'             ,'TRANS_IDENT'          ,'NUM01'                ,'NUM02'                ,'NUM03'                ,'VAR01'                ,'VAR02'                ,'VAR03'                ,'VAR04'                ,'VAR05'                ,'MNFSM_REG'            ,'MNFSM_DEN_IN'         ,'MNFSM_CHV_NFE'        ,'MNFSM_VL_ABAT_NT'     ,'MNFSM_VL_OUT_DA'      ,'MNFSM_VL_BC_ICMS'     ,'MNFSM_VL_ICMS'        
                            ,'MNFSM_BC_ICMS_ST'     ,'MNFSM_VL_ICMS_ST'     ,'MNFSM_VL_IPI'         ,'MNFSM_VL_PIS'         ,'MNFSM_VL_COFINS'      ,'MNFSM_VL_PIS_ST'      ,'MNFSM_COFINS_ST'      ,'MNFSM_IND_COMPL'      ,'UNFE_SIG_VEICULO'     ,'MNFSM_HR_EMIS'        ,'MNFSM_IND_EXTEMP'     ,'MNFSM_DAT_EXTEMP'     ,'MNFSM_VL_FCP'         ,'MNFSM_IND_FO'         ,'MNFSM_SIT_EDOC'       ,'MNFSM_DAT_COS'        ,'MNFSM_TIP_COM'        ,'MNFSM_VLD_INSS'       ,'MNFSM_VLD_GILRAT'     ,'MNFSM_VLD_SENAR'      ,'MNFSM_FCP_DEST'       ,'MNFSM_ICMS_DEST'      ,'MNFSM_ICMS_ORIG'      ,'SIST_ORIGEM'          ,'USUA_ORIGEM'          ,'DATA_CRIACAO'         
                            ,'ID_ORIGEM'            ,'EMPS_COD_1'           ,'FILI_COD_1'           ,'FILI_COD_CGC'        ,'UNFE_SIG'             ,'FILI_COD_INSEST'      ,'FILI_COD_INSMUN'      ,'FILI_COD_ATIVECON'    ,'FILI_NOM'             ,'FILI_NOM_FANTASIA'    ,'FILI_END'             ,'FILI_END_NUM'         ,'FILI_END_COMP'        ,'FILI_END_BAIRRO'      ,'FILI_END_MUNIC'       ,'FILI_END_CEP'         ,'FILI_COD_LOCAL'       ,'FILI_FAX'             ,'FILI_TEL'             ,'TP_LOC'               ,'FILI_SETOR'           ,'FILI_AUTOR'           ,'FILI_SANITARIA'       ,'FILI_LICEN'           ,'FILI_ESPEC'           ,'FILI_IND_RET_ISS'     
                            ,'FILI_COD_INSCEN'      ,'FILI_EMAIL'           ,'FILI_NAT_JUR'         ,'NUM01_1'             ,'NUM02_1'              ,'NUM03_1'              ,'VAR01_1'              ,'VAR02_1'              ,'VAR03_1'              ,'VAR04_1'              ,'VAR05_1'              ,'FILI_CEI'             ,'FILI_NIT'             ,'FILI_SUFRAMA'         ,'FILI_MUN_IBGE'        ,'TPLD_COD'             ,'FILI_NIRE'            ,'FILI_MATRIZ'          ,'FILI_TIP_ATI_ECO'     ,'FILI_INFO_COMPL'      ,'FILI_CLAS_TRIB'       ,'FILI_IND_OB_CIVIL'    ,'FILI_COD_CNPJ_EFR'    ,'SIST_ORIGEM_1'        ,'USUA_ORIGEM_1'        ,'DATA_CRIACAO_1'       
                            ,'ID_ORIGEM_1'          ,'FILI_CLAS_IND_EQ')
        #EB -                 A                       B                       C                       D                       E                       F                       G                       H                       I                       J                       K                       L                       M                       N                       O                       P                       Q                       R                       S                       T                       U                       V                       W                       X                       Y                       Z
        for c in range(1, len(listadecabecalho)+1):
            ABA_SMM.cell(row=l, column=c,   value=listadecabecalho[c-1])    
        l=l+1
        for linha in dadosABA_SMM:
#            for i in range(0,len(linha)-1):
            for i in range(0,len(linha)):
                ABA_SMM.cell(row=l, column=i+1,  value=linha[i])
            l=l+1
        formata_M(ABA_SMM)
        arquivo_excel_ms.save(ARQ_MS)    

    ########## " -> SAIDA MERCADORIA - ITEM"  RF03.02
    ########## " -> SAIDA MERCADORIA - ITEM"  RF03.02
    ########## " -> SAIDA MERCADORIA - ITEM"  RF03.02
    ########## " -> SAIDA MERCADORIA - ITEM"  RF03.02
    ########## " -> SAIDA MERCADORIA - ITEM"  RF03.02
    ########## " -> SAIDA MERCADORIA - ITEM"  RF03.02
    if ('21' in listadeabas):
        ABA_SMI       = arquivo_excel_ms.create_sheet("SAIDA MERCADORIA - ITEM",1)
        log( "ABA 21 / ", qtdabas , " -> SAIDA MERCADORIA - ITEM")
        chaves = []
        dadosABA_SMI = Busca_SMI(iei,"01/"+mesi+"/"+anoi)
        l=1
        ABA_SMI.cell(row=l, column=1,  value="SAIDA MERCADORIA - ITEM")
        l=l+1
        listadecabecalho = ('GAP'                  ,'ACAO'                 ,'TABELA'               ,'ROW_ID'               ,'EMPS_COD'             ,'FILI_COD'             ,'CGC_CPF'              ,'IE'                   ,'UF'                   ,'TP_LOC'               ,'LOCALIDADE'           ,'MOD_DOC'              ,'TDOC_COD'             ,'IND_CANC'             ,'IND_SIT'              ,'IND_TRIB_SUBSTRIB'    ,'VAL_ISEN_SUBSTRIB'    ,'VAL_OUTR_SUBSTRIB'    ,'INFSM_SERIE'          ,'INFSM_NUM'            ,'INFSM_DTEMISS'        ,'CATG_COD'             ,'CADG_COD'             ,'INFSM_NUM_ITEM'       ,'MATE_COD'             ,'UNID_COD_VENDA'       ,
                            'INFSM_DSC'            ,'CCUS_COD'             ,'CFOP_COD'             ,'NOPE_COD'             ,'CNBM_COD'             ,'UNID_COD'             ,'INFSM_QTD'            ,'INFSM_PES_LIQ'        ,'INFSM_VAL_PRECOUNIT'  ,'INFSM_VAL_PRECOTOT'   ,'INFSM_VAL_DESC'       ,'INFSM_NUM_ROMA'       ,'INFSM_DAT_ROMA'       ,'INFSM_VAL_FRETE'      ,'INFSM_VAL_SEGUR'      ,'INFSM_VAL_DESP'       ,'FEDE_COD'             ,'INFSM_TRIBIPI'        ,'INFSM_ALIQ_IPI'       ,'INFSM_BAS_IPI'        ,'INFSM_VAL_IPI'        ,'ESTA_COD'             ,'ESTB_COD'             ,'INFSM_TRIBICM'        ,'INFSM_ALIQ_ICMS'      ,'INFSM_BAS_ICMS'       ,
                            'INFSM_VAL_ICMS'       ,'INFSM_BASSUBST_ICMS'  ,'INFSM_VALSUBST_ICMS'  ,'INFSM_VAL_REDICMS'    ,'INFSM_ALIQ_DIFICMS'   ,'INFSM_ISENTA_IPI'     ,'INFSM_ISENTA_ICMS'    ,'INFSM_OUTRA_IPI'      ,'INFSM_OUTRA_ICMS'     ,'INFSM_VAL_CONT'       ,'INFSM_COD_CONT'       ,'INFSM_ICMS_FRETE'     ,'INFSM_CHASSI'         ,'INFSM_IND_MOV'        ,'INFSM_VAL_REDIPI'     ,'NUM01'                ,'NUM02'                ,'NUM03'                ,'VAR01'                ,'VAR02'                ,'VAR03'                ,'VAR04'                ,'VAR05'                ,'UNIN_COD'             ,'TIPI_COD'             ,'LIPI_COD'             ,
                            'INFSM_ALIQ_ST'        ,'INFSM_NUM_FCI'        ,'INFSM_FCP_DEST'       ,'INFSM_ICMS_DEST'      ,'INFSM_ICMS_ORIG'      ,'INFSM_MOT_RESS'       ,'INFSM_FCP_PRO'        ,'INFSM_FCP_RET'        ,'INFSM_FCP_ST'         ,'INFSM_ABA_NT'         ,'CAIA_COD'             ,'SIST_ORIGEM'          ,'USUA_ORIGEM'          ,'DATA_CRIACAO'         ,'ID_ORIGEM'            ,'TMRC_COD'             ,'INFSM_QTD_CONV'       ,'INFSM_UNID'           ,'INFSM_VLR_CONV'       ,'INFSM_ICMS_CONV'      ,'INFSM_ICMS_OP_CONV'   ,'INFSM_BC_ICMS_ST_CONV','INFSM_ICMS_ST_EST'    ,'INFSM_FCP_ST_EST'     ,'INFSM_ICMS_ST_REST'   ,'INFSM_FCP_ST_REST'    ,
                            'INFSM_ICMS_ST_COMP'   ,'INFSM_FCP_ST_COMP'    ,'EMPS_COD_1'           ,'FILI_COD_1'           ,'FILI_COD_CGC'         ,'UNFE_SIG'             ,'FILI_COD_INSEST'      ,'FILI_COD_INSMUN'      ,'FILI_COD_ATIVECON'    ,'FILI_NOM'             ,'FILI_NOM_FANTASIA'    ,'FILI_END'             ,'FILI_END_NUM'         ,'FILI_END_COMP'        ,'FILI_END_BAIRRO'      ,'FILI_END_MUNIC'       ,'FILI_END_CEP'         ,'FILI_COD_LOCAL'       ,'FILI_FAX'             ,'FILI_TEL'             ,'TP_LOC_1'             ,'FILI_SETOR'           ,'FILI_AUTOR'           ,'FILI_SANITARIA'       ,'FILI_LICEN'           ,'FILI_ESPEC'           ,
                            'FILI_IND_RET_ISS'     ,'FILI_COD_INSCEN'      ,'FILI_EMAIL'           ,'FILI_NAT_JUR'         ,'NUM01_1'              ,'NUM02_1'              ,'NUM03_1'              ,'VAR01_1'              ,'VAR02_1'              ,'VAR03_1'              ,'VAR04_1'              ,'VAR05_1'              ,'FILI_CEI'             ,'FILI_NIT'             ,'FILI_SUFRAMA'         ,'FILI_MUN_IBGE'        ,'TPLD_COD'             ,'FILI_NIRE'            ,'FILI_MATRIZ'          ,'FILI_TIP_ATI_ECO'     ,'FILI_INFO_COMPL'      ,'FILI_CLAS_TRIB'       ,'FILI_IND_OB_CIVIL'    ,'FILI_COD_CNPJ_EFR'    ,'SIST_ORIGEM_1'        ,'USUA_ORIGEM_1'        ,
                            'DATA_CRIACAO_1'       ,'ID_ORIGEM_1'          ,'FILI_CLAS_IND_EQ')                    
                    #FC -    A                       B                       C                       D                       E                       F                       G                       H                       I                       J                       K                       L                       M                       N                       O                       P                       Q                       R                       S                       T                       U                       V                       W                       X                       Y                       Z
        for c in range(1, len(listadecabecalho)+1):
            ABA_SMI.cell(row=l, column=c,   value=listadecabecalho[c-1])    
        l=l+1
        for linha in dadosABA_SMI:
#            for i in range(0,len(linha)-1):
            for i in range(0,len(linha)):
                ABA_SMI.cell(row=l, column=i+1,  value=linha[i])
            l=l+1
        formata_M(ABA_SMI)
        arquivo_excel_ms.save(ARQ_MS)    



    return(0)



########## FORMATACOES ###########
########## FORMATACOES ###########
########## FORMATACOES ###########
########## FORMATACOES ###########
########## FORMATACOES ###########
########## FORMATACOES ###########


def formata_M(planilha):
    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
    # TAMANHO DAS COLUNAS:
    adjust_column(planilha, 1,1,planilha.max_column) 
    # LINHA 1 = TITULO
    planilha.cell(1,  1).font = fontMasterPreta
    planilha.cell(1,  1).alignment = Alignment(horizontal='center')
    planilha.merge_cells('A1:P1')
    planilha.freeze_panes = 'A3'
    # LINHA 2 = CABEÇALHO
    for c in range(1,planilha.max_column+1):
        planilha.cell(2,  c).font = fontMasterPreta
        planilha.cell(2,  c).alignment = Alignment(horizontal='center')              
    return


def formata_R1600(planilha):
    planilha.freeze_panes = 'A3'
    # GERAL
    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
    planilha.column_dimensions['A'].width = 20  
    planilha.column_dimensions['B'].width = 20   
    planilha.column_dimensions['C'].width = 20  



    # LINHA 1 = TITULO
    planilha.merge_cells('A1:C1')
    planilha.cell(1,  1).font = fontMasterPreta
    planilha.cell(1,  1).alignment = Alignment(horizontal='center')
    
    # LINHA 2 = CABEÇALHO
    planilha.cell(2,  1).font = fontMasterPreta
    planilha.cell(2,  2).font = fontMasterPreta
    planilha.cell(2,  3).font = fontMasterPreta

    planilha.cell(2,  1).alignment = Alignment(horizontal='center')
    planilha.cell(2,  2).alignment = Alignment(horizontal='center')
    planilha.cell(2,  3).alignment = Alignment(horizontal='center')
 
    
    # LINHA 3 A PENÚLTIMA = DADOS
    lini = 3
    lfin = planilha.max_row
    
    for linha in range(lini,lfin):
        planilha.cell(linha,  2).number_format = "#,##0.00"
        planilha.cell(linha,  3).number_format = "#,##0.00"
        
    # LINHA FINAL = TOTAIS
    
    linha = planilha.max_row 
    planilha.cell(linha,  1, "TOTAIS")
    planilha.cell(linha,  1).alignment = Alignment(horizontal='center')
    planilha.cell(linha,  1).font = fontMasterPreta
    planilha.cell(linha,  2).font = fontMasterPreta
    planilha.cell(linha,  3).font = fontMasterPreta
    planilha.cell(linha,  2).number_format = "#,##0.00"
    planilha.cell(linha,  3).number_format = "#,##0.00"
    

# BORDAS    

    s=Side(border_style=BORDER_THIN, color='00000000')
    S=Side(border_style=BORDER_MEDIUM, color='00000000')
    d=Side(border_style=BORDER_DOUBLE, color='00000000')
    n=Side(border_style=None, color='00000000') 
  
    for a in range(2,linha):
         planilha.cell(row=a, column=1).border = Border(d,s,n,n)
         planilha.cell(row=a, column=2).border = Border(s,s,n,n)
         planilha.cell(row=a, column=3).border = Border(s,d,n,n)
     
    planilha.cell(row=1, column=1).border = Border(d,n,S,d)
    planilha.cell(row=1, column=2).border = Border(n,n,S,d)
    planilha.cell(row=1, column=3).border = Border(n,S,S,d)

    planilha.cell(row=2, column=1).border = Border(d,s,d,s)
    planilha.cell(row=2, column=2).border = Border(s,s,d,s)
    planilha.cell(row=2, column=3).border = Border(s,d,d,s)
    
    planilha.cell(row=linha, column=1).border = Border(d,n,S,d)
    planilha.cell(row=linha, column=2).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=3).border = Border(n,d,S,d)

    

def formata_SN(planilha):
    planilha.freeze_panes = 'A3'
    # GERAL
    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
    planilha.column_dimensions['A'].width = 20  
    planilha.column_dimensions['B'].width = 20   
    planilha.column_dimensions['C'].width = 20
    planilha.column_dimensions['D'].width = 20
    planilha.column_dimensions['E'].width = 20
    planilha.column_dimensions['F'].width = 20


    # LINHA 1 = TITULO
    planilha.merge_cells('A1:F1')
    planilha.cell(1,  1).font = fontMasterPreta
    planilha.cell(1,  1).alignment = Alignment(horizontal='center')
    
    # LINHA 2 = CABEÇALHO
    planilha.cell(2,  1).font = fontMasterPreta
    planilha.cell(2,  2).font = fontMasterPreta
    planilha.cell(2,  3).font = fontMasterPreta
    planilha.cell(2,  4).font = fontMasterPreta
    planilha.cell(2,  5).font = fontMasterPreta
    planilha.cell(2,  6).font = fontMasterPreta

    planilha.cell(2,  1).alignment = Alignment(horizontal='center')
    planilha.cell(2,  2).alignment = Alignment(horizontal='center')
    planilha.cell(2,  3).alignment = Alignment(horizontal='center')
    planilha.cell(2,  4).alignment = Alignment(horizontal='center')
    planilha.cell(2,  5).alignment = Alignment(horizontal='center')
    planilha.cell(2,  6).alignment = Alignment(horizontal='center')
 
    
    # LINHA 3 A PENÚLTIMA = DADOS
    lfin = planilha.max_row
    
#    for linha in range(lini,lfin+1):
#        planilha.cell(linha, 8).number_format = "#,##0.00"
        
# BORDAS    
    linha = lfin
    linha = linha + 1
    
    s=Side(border_style=BORDER_THIN, color='00000000')
    S=Side(border_style=BORDER_MEDIUM, color='00000000')
    d=Side(border_style=BORDER_DOUBLE, color='00000000')
    n=Side(border_style=None, color='00000000') 
  
    for a in range(2,linha):
         planilha.cell(row=a, column=1).border = Border(d,s,n,n)
         planilha.cell(row=a, column=2).border = Border(s,s,n,n)
         planilha.cell(row=a, column=3).border = Border(s,s,n,n)
         planilha.cell(row=a, column=4).border = Border(s,s,n,n)
         planilha.cell(row=a, column=5).border = Border(s,s,n,n)
         planilha.cell(row=a, column=6).border = Border(s,d,n,n)
     
    planilha.cell(row=1, column=1).border = Border(d,n,S,d)
    planilha.cell(row=1, column=2).border = Border(n,n,S,d)
    planilha.cell(row=1, column=3).border = Border(n,n,S,d)
    planilha.cell(row=1, column=4).border = Border(n,n,S,d)
    planilha.cell(row=1, column=5).border = Border(n,n,S,d)
    planilha.cell(row=1, column=6).border = Border(n,S,S,d)

    planilha.cell(row=2, column=1).border = Border(d,s,d,s)
    planilha.cell(row=2, column=2).border = Border(s,s,d,s)
    planilha.cell(row=2, column=3).border = Border(s,s,d,s)
    planilha.cell(row=2, column=4).border = Border(s,s,d,s)
    planilha.cell(row=2, column=5).border = Border(s,s,d,s)
    planilha.cell(row=2, column=6).border = Border(s,d,d,s)
    
    planilha.cell(row=linha-1, column=1).border = Border(d,s,n,d)
    planilha.cell(row=linha-1, column=2).border = Border(s,s,n,d)
    planilha.cell(row=linha-1, column=3).border = Border(s,s,n,d)
    planilha.cell(row=linha-1, column=4).border = Border(s,s,n,d)
    planilha.cell(row=linha-1, column=5).border = Border(s,s,n,d)
    planilha.cell(row=linha-1, column=6).border = Border(s,d,n,d)




       






def formata_AICMS(planilha):
    planilha.freeze_panes = 'A3'
    # GERAL
    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
    planilha.column_dimensions['A'].width = 11  
    planilha.column_dimensions['B'].width = 18   
    planilha.column_dimensions['C'].width = 8 
    planilha.column_dimensions['D'].width = 6
    planilha.column_dimensions['E'].width = 5 
    planilha.column_dimensions['F'].width = 12 
    planilha.column_dimensions['G'].width = 130  
    planilha.column_dimensions['H'].width = 20  


    # LINHA 1 = TITULO
    planilha.merge_cells('A1:H1')
    planilha.cell(1,  1).font = fontMasterPreta
    planilha.cell(1,  1).alignment = Alignment(horizontal='center')
    
    # LINHA 2 = CABEÇALHO
    planilha.cell(2,  1).font = fontMasterPreta
    planilha.cell(2,  2).font = fontMasterPreta
    planilha.cell(2,  3).font = fontMasterPreta
    planilha.cell(2,  4).font = fontMasterPreta
    planilha.cell(2,  5).font = fontMasterPreta
    planilha.cell(2,  6).font = fontMasterPreta
    planilha.cell(2,  7).font = fontMasterPreta
    planilha.cell(2,  8).font = fontMasterPreta

    planilha.cell(2,  1).alignment = Alignment(horizontal='center')
    planilha.cell(2,  2).alignment = Alignment(horizontal='center')
    planilha.cell(2,  3).alignment = Alignment(horizontal='center')
    planilha.cell(2,  4).alignment = Alignment(horizontal='center')
    planilha.cell(2,  5).alignment = Alignment(horizontal='center')
    planilha.cell(2,  6).alignment = Alignment(horizontal='center')
    planilha.cell(2,  7).alignment = Alignment(horizontal='center')
    planilha.cell(2,  8).alignment = Alignment(horizontal='center')
 
    
    # LINHA 3 A PENÚLTIMA = DADOS
    lini = 3
    lfin = planilha.max_row
    
    for linha in range(lini,lfin+1):
        planilha.cell(linha, 8).number_format = "#,##0.00"
        
# BORDAS    
    linha = linha + 1
    
    s=Side(border_style=BORDER_THIN, color='00000000')
    S=Side(border_style=BORDER_MEDIUM, color='00000000')
    d=Side(border_style=BORDER_DOUBLE, color='00000000')
    n=Side(border_style=None, color='00000000') 
  
    for a in range(2,linha):
         planilha.cell(row=a, column=1).border = Border(d,s,n,n)
         planilha.cell(row=a, column=2).border = Border(s,s,n,n)
         planilha.cell(row=a, column=3).border = Border(s,s,n,n)
         planilha.cell(row=a, column=4).border = Border(s,s,n,n)
         planilha.cell(row=a, column=5).border = Border(s,s,n,n)
         planilha.cell(row=a, column=6).border = Border(s,s,n,n)
         planilha.cell(row=a, column=7).border = Border(s,s,n,n)
         planilha.cell(row=a, column=8).border = Border(s,d,n,n)
     
    planilha.cell(row=1, column=1).border = Border(d,n,S,d)
    planilha.cell(row=1, column=2).border = Border(n,n,S,d)
    planilha.cell(row=1, column=3).border = Border(n,n,S,d)
    planilha.cell(row=1, column=4).border = Border(n,n,S,d)
    planilha.cell(row=1, column=5).border = Border(n,n,S,d)
    planilha.cell(row=1, column=6).border = Border(n,n,S,d)
    planilha.cell(row=1, column=7).border = Border(n,n,S,d)
    planilha.cell(row=1, column=8).border = Border(n,S,S,d)

    planilha.cell(row=2, column=1).border = Border(d,s,d,s)
    planilha.cell(row=2, column=2).border = Border(s,s,d,s)
    planilha.cell(row=2, column=3).border = Border(s,s,d,s)
    planilha.cell(row=2, column=4).border = Border(s,s,d,s)
    planilha.cell(row=2, column=5).border = Border(s,s,d,s)
    planilha.cell(row=2, column=6).border = Border(s,s,d,s)
    planilha.cell(row=2, column=7).border = Border(s,s,d,s)
    planilha.cell(row=2, column=8).border = Border(s,d,d,s)
    
    planilha.cell(row=linha-1, column=1).border = Border(d,s,n,d)
    planilha.cell(row=linha-1, column=2).border = Border(s,s,n,d)
    planilha.cell(row=linha-1, column=3).border = Border(s,s,n,d)
    planilha.cell(row=linha-1, column=4).border = Border(s,s,n,d)
    planilha.cell(row=linha-1, column=5).border = Border(s,s,n,d)
    planilha.cell(row=linha-1, column=6).border = Border(s,s,n,d)
    planilha.cell(row=linha-1, column=7).border = Border(s,s,n,d)
    planilha.cell(row=linha-1, column=8).border = Border(s,d,n,d)




       
    return



def formata_RCFOP(planilha):
    planilha.freeze_panes = 'A3'
    # GERAL
    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
    planilha.column_dimensions['A'].width = 10  
    planilha.column_dimensions['B'].width = 20   
    planilha.column_dimensions['C'].width = 20 
    planilha.column_dimensions['D'].width = 20
    planilha.column_dimensions['E'].width = 20 
    planilha.column_dimensions['F'].width = 20 
    planilha.column_dimensions['G'].width = 20  


    # LINHA 1 = TITULO
    planilha.merge_cells('A1:G1')
    planilha.cell(1,  1).font = fontMasterPreta
    planilha.cell(1,  1).alignment = Alignment(horizontal='center')
    
    # LINHA 2 = CABEÇALHO
    planilha.cell(2,  1).font = fontMasterPreta
    planilha.cell(2,  2).font = fontMasterPreta
    planilha.cell(2,  3).font = fontMasterPreta
    planilha.cell(2,  4).font = fontMasterPreta
    planilha.cell(2,  5).font = fontMasterPreta
    planilha.cell(2,  6).font = fontMasterPreta
    planilha.cell(2,  7).font = fontMasterPreta

    planilha.cell(2,  1).alignment = Alignment(horizontal='center')
    planilha.cell(2,  2).alignment = Alignment(horizontal='center')
    planilha.cell(2,  3).alignment = Alignment(horizontal='center')
    planilha.cell(2,  4).alignment = Alignment(horizontal='center')
    planilha.cell(2,  5).alignment = Alignment(horizontal='center')
    planilha.cell(2,  6).alignment = Alignment(horizontal='center')
    planilha.cell(2,  7).alignment = Alignment(horizontal='center')
 
    
    # LINHA 3 A PENÚLTIMA = DADOS
    lini = 3
    lfin = planilha.max_row
    
    for linha in range(lini,lfin):
        planilha.cell(linha, 2).number_format = "#,##0.00"
        planilha.cell(linha, 3).number_format = "#,##0.00"
        planilha.cell(linha, 4).number_format = "#,##0.00"
        planilha.cell(linha, 5).number_format = "#,##0.00"
        planilha.cell(linha, 6).number_format = "#,##0.00"
        planilha.cell(linha, 7).number_format = "#,##0.00"
        
    # LINHA FINAL = TOTAIS
    
    linha = planilha.max_row
#    planilha.cell(linha,  1, "ENTRADAS")
    planilha.cell(linha,  1).alignment = Alignment(horizontal='center')
    planilha.cell(linha,  1).font = fontMasterPreta
    planilha.cell(linha,  2).font = fontMasterPreta
    planilha.cell(linha,  3).font = fontMasterPreta
    planilha.cell(linha,  4).font = fontMasterPreta
    planilha.cell(linha,  5).font = fontMasterPreta
    planilha.cell(linha,  6).font = fontMasterPreta
    planilha.cell(linha,  7).font = fontMasterPreta

    planilha.cell(linha, 2).number_format = "#,##0.00"
    planilha.cell(linha, 3).number_format = "#,##0.00"
    planilha.cell(linha, 4).number_format = "#,##0.00"
    planilha.cell(linha, 5).number_format = "#,##0.00"
    planilha.cell(linha, 6).number_format = "#,##0.00"
    planilha.cell(linha, 7).number_format = "#,##0.00"
    
    linha = planilha.max_row-2
#    planilha.cell(linha,  1, "SAIDAS")
    planilha.cell(linha,  1).alignment = Alignment(horizontal='center')
    planilha.cell(linha,  1).font = fontMasterPreta
    planilha.cell(linha,  2).font = fontMasterPreta
    planilha.cell(linha,  3).font = fontMasterPreta
    planilha.cell(linha,  4).font = fontMasterPreta
    planilha.cell(linha,  5).font = fontMasterPreta
    planilha.cell(linha,  6).font = fontMasterPreta
    planilha.cell(linha,  7).font = fontMasterPreta

    planilha.cell(linha, 2).number_format = "#,##0.00"
    planilha.cell(linha, 3).number_format = "#,##0.00"
    planilha.cell(linha, 4).number_format = "#,##0.00"
    planilha.cell(linha, 5).number_format = "#,##0.00"
    planilha.cell(linha, 6).number_format = "#,##0.00"
    planilha.cell(linha, 7).number_format = "#,##0.00"
    
    #BORDAS    

    set_border_edsi(planilha, 'A1:G'+str(planilha.max_row-4), 'ffff')
    set_border_edsi(planilha, 'A'+str(planilha.max_row)+':G'+str(planilha.max_row) ,'dddd')      
    set_border_edsi(planilha, 'A'+str(planilha.max_row-2)+':G'+str(planilha.max_row-2) ,'dddd')      

    return


def formata_RCFOP_DIF(planilha):
    planilha.freeze_panes = 'A3'
    # GERAL
    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
    planilha.column_dimensions['H'].width = 2  
    planilha.column_dimensions['I'].width = 2  
    planilha.column_dimensions['J'].width = 2  
    
    planilha.column_dimensions['K'].width = 10  
    planilha.column_dimensions['L'].width = 20   
    planilha.column_dimensions['M'].width = 20 
    planilha.column_dimensions['N'].width = 20
    
    planilha.column_dimensions['O'].width = 2  
    
    planilha.column_dimensions['P'].width = 10   
    planilha.column_dimensions['Q'].width = 20 
    planilha.column_dimensions['R'].width = 20
    planilha.column_dimensions['S'].width = 20  
    
    planilha.column_dimensions['T'].width = 2   
    
    planilha.column_dimensions['U'].width = 20 
    planilha.column_dimensions['V'].width = 20
    planilha.column_dimensions['W'].width = 20  


    # LINHA 1 = TITULO
    planilha.merge_cells('K1:N1')
    planilha.merge_cells('P1:S1')
    planilha.merge_cells('U1:W1')
    
    planilha.cell(1,  11).font = fontMasterPreta
    planilha.cell(1,  11).alignment = Alignment(horizontal='center')
    
    planilha.cell(1,  16).font = fontMasterPreta
    planilha.cell(1,  16).alignment = Alignment(horizontal='center')
    
    planilha.cell(1,  21).font = fontMasterPreta
    planilha.cell(1,  21).alignment = Alignment(horizontal='center')
    
    # LINHA 2 = CABEÇALHO
    planilha.cell(2,  11).font = fontMasterPreta
    planilha.cell(2,  12).font = fontMasterPreta
    planilha.cell(2,  13).font = fontMasterPreta
    planilha.cell(2,  14).font = fontMasterPreta

    planilha.cell(2,  161).font = fontMasterPreta
    planilha.cell(2,  17).font = fontMasterPreta
    planilha.cell(2,  18).font = fontMasterPreta
    planilha.cell(2,  19).font = fontMasterPreta

    planilha.cell(2,  21).font = fontMasterPreta
    planilha.cell(2,  22).font = fontMasterPreta
    planilha.cell(2,  23).font = fontMasterPreta

    planilha.cell(2,  11).alignment = Alignment(horizontal='center')
    planilha.cell(2,  12).alignment = Alignment(horizontal='center')
    planilha.cell(2,  13).alignment = Alignment(horizontal='center')
    planilha.cell(2,  14).alignment = Alignment(horizontal='center')
 
    planilha.cell(2,  16).alignment = Alignment(horizontal='center')
    planilha.cell(2,  17).alignment = Alignment(horizontal='center')
    planilha.cell(2,  18).alignment = Alignment(horizontal='center')
    planilha.cell(2,  19).alignment = Alignment(horizontal='center')

    planilha.cell(2,  21).alignment = Alignment(horizontal='center')
    planilha.cell(2,  22).alignment = Alignment(horizontal='center')
    planilha.cell(2,  23).alignment = Alignment(horizontal='center')
    
    # LINHA 3 A PENÚLTIMA = DADOS
    lini = 3
    lfin = planilha.max_row
    
    for linha in range(lini,lfin):
        planilha.cell(linha, 12).number_format = "#,##0.00"
        planilha.cell(linha, 13).number_format = "#,##0.00"
        planilha.cell(linha, 14).number_format = "#,##0.00"

        planilha.cell(linha, 17).number_format = "#,##0.00"
        planilha.cell(linha, 18).number_format = "#,##0.00"
        planilha.cell(linha, 19).number_format = "#,##0.00"

        planilha.cell(linha, 21).number_format = "#,##0.00"
        planilha.cell(linha, 22).number_format = "#,##0.00"
        planilha.cell(linha, 23).number_format = "#,##0.00"
        
    # LINHA FINAL = TOTAIS
    
    linha = planilha.max_row
#    planilha.cell(linha,  1, "ENTRADAS")
    planilha.cell(linha,  11).alignment = Alignment(horizontal='center')
    planilha.cell(linha,  11).font = fontMasterPreta
    planilha.cell(linha,  12).font = fontMasterPreta
    planilha.cell(linha,  13).font = fontMasterPreta
    planilha.cell(linha,  14).font = fontMasterPreta

    planilha.cell(linha,  16).alignment = Alignment(horizontal='center')
    planilha.cell(linha,  16).font = fontMasterPreta
    planilha.cell(linha,  17).font = fontMasterPreta
    planilha.cell(linha,  18).font = fontMasterPreta
    planilha.cell(linha,  19).font = fontMasterPreta

    planilha.cell(linha,  21).font = fontMasterPreta
    planilha.cell(linha,  22).font = fontMasterPreta
    planilha.cell(linha,  23).font = fontMasterPreta

    planilha.cell(linha, 12).number_format = "#,##0.00"
    planilha.cell(linha, 13).number_format = "#,##0.00"
    planilha.cell(linha, 14).number_format = "#,##0.00"
    
    planilha.cell(linha, 17).number_format = "#,##0.00"
    planilha.cell(linha, 18).number_format = "#,##0.00"
    planilha.cell(linha, 19).number_format = "#,##0.00"
    
    planilha.cell(linha, 21).number_format = "#,##0.00"
    planilha.cell(linha, 22).number_format = "#,##0.00"
    planilha.cell(linha, 23).number_format = "#,##0.00"
    
    linha = planilha.max_row-2
#    planilha.cell(linha,  1, "SAIDAS")
    planilha.cell(linha,  11).alignment = Alignment(horizontal='center')
    planilha.cell(linha,  11).font = fontMasterPreta
    planilha.cell(linha,  12).font = fontMasterPreta
    planilha.cell(linha,  13).font = fontMasterPreta
    planilha.cell(linha,  14).font = fontMasterPreta

    planilha.cell(linha,  16).alignment = Alignment(horizontal='center')
    planilha.cell(linha,  16).font = fontMasterPreta
    planilha.cell(linha,  17).font = fontMasterPreta
    planilha.cell(linha,  18).font = fontMasterPreta
    planilha.cell(linha,  19).font = fontMasterPreta

    planilha.cell(linha,  21).font = fontMasterPreta
    planilha.cell(linha,  22).font = fontMasterPreta
    planilha.cell(linha,  23).font = fontMasterPreta

    planilha.cell(linha, 12).number_format = "#,##0.00"
    planilha.cell(linha, 13).number_format = "#,##0.00"
    planilha.cell(linha, 14).number_format = "#,##0.00"
    
    planilha.cell(linha, 17).number_format = "#,##0.00"
    planilha.cell(linha, 18).number_format = "#,##0.00"
    planilha.cell(linha, 19).number_format = "#,##0.00"
    
    planilha.cell(linha, 21).number_format = "#,##0.00"
    planilha.cell(linha, 22).number_format = "#,##0.00"
    planilha.cell(linha, 23).number_format = "#,##0.00"
    

    #BORDAS    

    set_border_edsi(planilha, 'K1:N'+str(planilha.max_row-4), 'ffff')
    set_border_edsi(planilha, 'K'+str(planilha.max_row)+':N'+str(planilha.max_row) ,'dddd')      
    set_border_edsi(planilha, 'K'+str(planilha.max_row-2)+':N'+str(planilha.max_row-2) ,'dddd')      

    set_border_edsi(planilha, 'P1:S'+str(planilha.max_row-4), 'ffff')
    set_border_edsi(planilha, 'P'+str(planilha.max_row)+':S'+str(planilha.max_row) ,'dddd')      
    set_border_edsi(planilha, 'P'+str(planilha.max_row-2)+':S'+str(planilha.max_row-2) ,'dddd')      

    set_border_edsi(planilha, 'U1:W'+str(planilha.max_row-4), 'ffff')
    set_border_edsi(planilha, 'U'+str(planilha.max_row)+':W'+str(planilha.max_row) ,'dddd')      
    set_border_edsi(planilha, 'U'+str(planilha.max_row-2)+':W'+str(planilha.max_row-2) ,'dddd')      


    return




def formata_TRT(planilha):
    planilha.freeze_panes = 'A3'
    # GERAL
    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
#    fontMasterAzul     = Font(color='000000FF', bold=True, size=12)
#    fontMasterVermelha = Font(color='00FF0000', bold=True, size=12)

    # TAMANHO DAS COLUNAS:
    adjust_column(planilha, 1,2,planilha.max_column)

    # LINHA 1 = TITULO
    planilha.merge_cells('A1:M1')
    planilha.cell(1,  1).font = fontMasterPreta
    planilha.cell(1,  1).alignment = Alignment(horizontal='center')
    
    # LINHA 2 = CABEÇALHO
    planilha.cell(2,  1).font = fontMasterPreta
    planilha.cell(2,  2).font = fontMasterPreta
    planilha.cell(2,  3).font = fontMasterPreta
    planilha.cell(2,  4).font = fontMasterPreta
    planilha.cell(2,  5).font = fontMasterPreta
    planilha.cell(2,  6).font = fontMasterPreta
    planilha.cell(2,  7).font = fontMasterPreta
    planilha.cell(2,  8).font = fontMasterPreta
    planilha.cell(2,  9).font = fontMasterPreta
    planilha.cell(2, 10).font = fontMasterPreta
    planilha.cell(2, 11).font = fontMasterPreta
    planilha.cell(2, 12).font = fontMasterPreta
    planilha.cell(2, 13).font = fontMasterPreta
    planilha.cell(2,  1).alignment = Alignment(horizontal='center')
    planilha.cell(2,  2).alignment = Alignment(horizontal='center')
    planilha.cell(2,  3).alignment = Alignment(horizontal='center')
    planilha.cell(2,  4).alignment = Alignment(horizontal='center')
    planilha.cell(2,  5).alignment = Alignment(horizontal='center')
    planilha.cell(2,  6).alignment = Alignment(horizontal='center')
    planilha.cell(2,  7).alignment = Alignment(horizontal='center')
    planilha.cell(2,  8).alignment = Alignment(horizontal='center')
    planilha.cell(2,  9).alignment = Alignment(horizontal='center')
    planilha.cell(2, 10).alignment = Alignment(horizontal='center')
    planilha.cell(2, 11).alignment = Alignment(horizontal='center')
    planilha.cell(2, 12).alignment = Alignment(horizontal='center')
    planilha.cell(2, 13).alignment = Alignment(horizontal='center')
    
    # LINHA 3 A PENÚLTIMA = DADOS
    lini = 3
    lfin = planilha.max_row
    
    for linha in range(lini,lfin):
        planilha.cell(linha,  8).number_format = "#,##0.00"
        planilha.cell(linha, 10).number_format = "#,##0.00"
        planilha.cell(linha, 11).number_format = "#,##0.00"
        planilha.cell(linha, 12).number_format = "#,##0.00"
        planilha.cell(linha, 13).number_format = "#,##0.00"
        
    # LINHA FINAL = TOTAIS
    
    linha = planilha.max_row 
    merg = 'A'+str(linha)+':G'+str(linha)
    planilha.merge_cells(merg)
    planilha.cell(linha,  1, "TOTAIS")
    planilha.cell(linha,  1).alignment = Alignment(horizontal='center')
    planilha.cell(linha,  1).font = fontMasterPreta
    planilha.cell(linha,  5).font = fontMasterPreta
    planilha.cell(linha,  6).font = fontMasterPreta
    planilha.cell(linha,  7).font = fontMasterPreta
    planilha.cell(linha,  8).font = fontMasterPreta
    planilha.cell(linha,  9).font = fontMasterPreta
    planilha.cell(linha,  9).font = fontMasterPreta
    planilha.cell(linha, 10).font = fontMasterPreta
    planilha.cell(linha, 11).font = fontMasterPreta
    planilha.cell(linha, 12).font = fontMasterPreta
    planilha.cell(linha, 13).font = fontMasterPreta

    planilha.cell(linha,  8).number_format = "#,##0.00"
    planilha.cell(linha, 10).number_format = "#,##0.00"
    planilha.cell(linha, 11).number_format = "#,##0.00"
    planilha.cell(linha, 12).number_format = "#,##0.00"
    planilha.cell(linha, 13).number_format = "#,##0.00"

    # BORDAS    

    s=Side(border_style=BORDER_THIN, color='00000000')
    S=Side(border_style=BORDER_MEDIUM, color='00000000')
    d=Side(border_style=BORDER_DOUBLE, color='00000000')
    n=Side(border_style=None, color='00000000') 
  
    for a in range(2,linha):
         planilha.cell(row=a, column=1).border = Border(d,s,n,n)
         planilha.cell(row=a, column=2).border = Border(s,s,n,n)
         planilha.cell(row=a, column=3).border = Border(s,s,n,n)
         planilha.cell(row=a, column=4).border = Border(s,s,n,n)
         planilha.cell(row=a, column=5).border = Border(s,s,n,n)
         planilha.cell(row=a, column=6).border = Border(s,s,n,n)
         planilha.cell(row=a, column=7).border = Border(s,s,n,n)
         planilha.cell(row=a, column=8).border = Border(s,s,n,n)
         planilha.cell(row=a, column=9).border = Border(s,s,n,n)
         planilha.cell(row=a, column=10).border = Border(s,s,n,n)
         planilha.cell(row=a, column=11).border = Border(s,s,n,n)
         planilha.cell(row=a, column=12).border = Border(s,s,n,n)
         planilha.cell(row=a, column=13).border = Border(s,d,n,n)
     
    planilha.cell(row=1, column=1).border = Border(d,n,S,d)
    planilha.cell(row=1, column=2).border = Border(n,n,S,d)
    planilha.cell(row=1, column=3).border = Border(n,n,S,d)
    planilha.cell(row=1, column=4).border = Border(n,n,S,d)
    planilha.cell(row=1, column=5).border = Border(n,n,S,d)
    planilha.cell(row=1, column=6).border = Border(n,n,S,d)
    planilha.cell(row=1, column=7).border = Border(n,n,S,d)
    planilha.cell(row=1, column=8).border = Border(n,n,S,d)
    planilha.cell(row=1, column=9).border = Border(n,n,S,d)
    planilha.cell(row=1, column=10).border = Border(n,n,S,d)
    planilha.cell(row=1, column=11).border = Border(n,n,S,d)
    planilha.cell(row=1, column=12).border = Border(n,n,S,d)
    planilha.cell(row=1, column=13).border = Border(n,S,S,d)

    planilha.cell(row=2, column=1).border = Border(d,s,d,s)
    planilha.cell(row=2, column=2).border = Border(s,s,d,s)
    planilha.cell(row=2, column=3).border = Border(s,s,d,s)
    planilha.cell(row=2, column=4).border = Border(s,s,d,s)
    planilha.cell(row=2, column=5).border = Border(s,s,d,s)
    planilha.cell(row=2, column=6).border = Border(s,s,d,s)
    planilha.cell(row=2, column=7).border = Border(s,s,d,s)
    planilha.cell(row=2, column=8).border = Border(s,s,d,s)
    planilha.cell(row=2, column=9).border = Border(s,s,d,s)
    planilha.cell(row=2, column=10).border = Border(s,s,d,s)
    planilha.cell(row=2, column=11).border = Border(s,s,d,s)
    planilha.cell(row=2, column=12).border = Border(s,s,d,s)
    planilha.cell(row=2, column=13).border = Border(s,d,d,s)
    
    planilha.cell(row=linha, column=1).border = Border(d,n,S,d)
    planilha.cell(row=linha, column=2).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=3).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=4).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=5).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=6).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=7).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=8).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=9).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=10).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=11).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=12).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=13).border = Border(n,d,S,d)

       
    return




def formata_TRS(planilha):
    planilha.freeze_panes = 'A3'
    # GERAL
    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
#    fontMasterAzul     = Font(color='000000FF', bold=True, size=12)
#    fontMasterVermelha = Font(color='00FF0000', bold=True, size=12)
    planilha.column_dimensions['A'].width = 12  
    planilha.column_dimensions['B'].width = 12   
    planilha.column_dimensions['C'].width = 6  
    planilha.column_dimensions['D'].width = 6 
    planilha.column_dimensions['E'].width = 20 
    planilha.column_dimensions['F'].width = 20  
    planilha.column_dimensions['G'].width = 20  
    planilha.column_dimensions['H'].width = 20  
    planilha.column_dimensions['I'].width = 20  
    planilha.column_dimensions['J'].width = 20  
    planilha.column_dimensions['K'].width = 20  
    planilha.column_dimensions['L'].width = 20  



    # LINHA 1 = TITULO
    planilha.merge_cells('A1:L1')
    planilha.cell(1,  1).font = fontMasterPreta
    planilha.cell(1,  1).alignment = Alignment(horizontal='center')
    
    # LINHA 2 = CABEÇALHO
    planilha.cell(2,  1).font = fontMasterPreta
    planilha.cell(2,  2).font = fontMasterPreta
    planilha.cell(2,  3).font = fontMasterPreta
    planilha.cell(2,  4).font = fontMasterPreta
    planilha.cell(2,  5).font = fontMasterPreta
    planilha.cell(2,  6).font = fontMasterPreta
    planilha.cell(2,  7).font = fontMasterPreta
    planilha.cell(2,  8).font = fontMasterPreta
    planilha.cell(2,  9).font = fontMasterPreta
    planilha.cell(2, 10).font = fontMasterPreta
    planilha.cell(2, 11).font = fontMasterPreta
    planilha.cell(2, 12).font = fontMasterPreta
    planilha.cell(2,  1).alignment = Alignment(horizontal='center')
    planilha.cell(2,  2).alignment = Alignment(horizontal='center')
    planilha.cell(2,  3).alignment = Alignment(horizontal='center')
    planilha.cell(2,  4).alignment = Alignment(horizontal='center')
    planilha.cell(2,  5).alignment = Alignment(horizontal='center')
    planilha.cell(2,  6).alignment = Alignment(horizontal='center')
    planilha.cell(2,  7).alignment = Alignment(horizontal='center')
    planilha.cell(2,  8).alignment = Alignment(horizontal='center')
    planilha.cell(2,  9).alignment = Alignment(horizontal='center')
    planilha.cell(2, 10).alignment = Alignment(horizontal='center')
    planilha.cell(2, 11).alignment = Alignment(horizontal='center')
    planilha.cell(2, 12).alignment = Alignment(horizontal='center')
    
    # LINHA 3 A PENÚLTIMA = DADOS
    lini = 3
    lfin = planilha.max_row
    
    for linha in range(lini,lfin):
        planilha.cell(linha,  5).number_format = "#,##0.00"
        planilha.cell(linha,  6).number_format = "#,##0.00"
        planilha.cell(linha,  7).number_format = "#,##0.00"
        planilha.cell(linha,  8).number_format = "#,##0.00"
        planilha.cell(linha,  9).number_format = "#,##0.00"
        planilha.cell(linha, 10).number_format = "#,##0.00"
        planilha.cell(linha, 11).number_format = "#,##0.00"
        planilha.cell(linha, 12).number_format = "#,##0.00"
        
    # LINHA FINAL = TOTAIS
    
    linha = planilha.max_row 
    merg = 'A'+str(linha)+':C'+str(linha)
    planilha.merge_cells(merg)
    planilha.cell(linha,  1, "TOTAIS")
    planilha.cell(linha,  1).alignment = Alignment(horizontal='center')
    planilha.cell(linha,  1).font = fontMasterPreta
    planilha.cell(linha,  5).font = fontMasterPreta
    planilha.cell(linha,  6).font = fontMasterPreta
    planilha.cell(linha,  7).font = fontMasterPreta
    planilha.cell(linha,  8).font = fontMasterPreta
    planilha.cell(linha,  9).font = fontMasterPreta
    planilha.cell(linha,  9).font = fontMasterPreta
    planilha.cell(linha, 10).font = fontMasterPreta
    planilha.cell(linha, 11).font = fontMasterPreta
    planilha.cell(linha, 12).font = fontMasterPreta

    planilha.cell(linha,  4).number_format = "#,##0.00"
    planilha.cell(linha,  5).number_format = "#,##0.00"
    planilha.cell(linha,  6).number_format = "#,##0.00"
    planilha.cell(linha,  7).number_format = "#,##0.00"
    planilha.cell(linha,  8).number_format = "#,##0.00"
    planilha.cell(linha,  9).number_format = "#,##0.00"
    planilha.cell(linha, 10).number_format = "#,##0.00"
    planilha.cell(linha, 11).number_format = "#,##0.00"
    planilha.cell(linha, 12).number_format = "#,##0.00"

# BORDAS    

    s=Side(border_style=BORDER_THIN, color='00000000')
    S=Side(border_style=BORDER_MEDIUM, color='00000000')
    d=Side(border_style=BORDER_DOUBLE, color='00000000')
    n=Side(border_style=None, color='00000000') 
  
    for a in range(2,linha):
         planilha.cell(row=a, column=1).border = Border(d,s,n,n)
         planilha.cell(row=a, column=2).border = Border(s,s,n,n)
         planilha.cell(row=a, column=3).border = Border(s,s,n,n)
         planilha.cell(row=a, column=4).border = Border(s,s,n,n)
         planilha.cell(row=a, column=5).border = Border(s,s,n,n)
         planilha.cell(row=a, column=6).border = Border(s,s,n,n)
         planilha.cell(row=a, column=7).border = Border(s,s,n,n)
         planilha.cell(row=a, column=8).border = Border(s,s,n,n)
         planilha.cell(row=a, column=9).border = Border(s,s,n,n)
         planilha.cell(row=a, column=10).border = Border(s,s,n,n)
         planilha.cell(row=a, column=11).border = Border(s,s,n,n)
         planilha.cell(row=a, column=12).border = Border(s,d,n,n)
     
    planilha.cell(row=1, column=1).border = Border(d,n,S,d)
    planilha.cell(row=1, column=2).border = Border(n,n,S,d)
    planilha.cell(row=1, column=3).border = Border(n,n,S,d)
    planilha.cell(row=1, column=4).border = Border(n,n,S,d)
    planilha.cell(row=1, column=5).border = Border(n,n,S,d)
    planilha.cell(row=1, column=6).border = Border(n,n,S,d)
    planilha.cell(row=1, column=7).border = Border(n,n,S,d)
    planilha.cell(row=1, column=8).border = Border(n,n,S,d)
    planilha.cell(row=1, column=9).border = Border(n,n,S,d)
    planilha.cell(row=1, column=10).border = Border(n,n,S,d)
    planilha.cell(row=1, column=11).border = Border(n,n,S,d)
    planilha.cell(row=1, column=12).border = Border(n,S,S,d)

    planilha.cell(row=2, column=1).border = Border(d,s,d,s)
    planilha.cell(row=2, column=2).border = Border(s,s,d,s)
    planilha.cell(row=2, column=3).border = Border(s,s,d,s)
    planilha.cell(row=2, column=4).border = Border(s,s,d,s)
    planilha.cell(row=2, column=5).border = Border(s,s,d,s)
    planilha.cell(row=2, column=6).border = Border(s,s,d,s)
    planilha.cell(row=2, column=7).border = Border(s,s,d,s)
    planilha.cell(row=2, column=8).border = Border(s,s,d,s)
    planilha.cell(row=2, column=9).border = Border(s,s,d,s)
    planilha.cell(row=2, column=10).border = Border(s,s,d,s)
    planilha.cell(row=2, column=11).border = Border(s,s,d,s)
    planilha.cell(row=2, column=12).border = Border(s,d,d,s)
    
    planilha.cell(row=linha, column=1).border = Border(d,n,S,d)
    planilha.cell(row=linha, column=2).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=3).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=4).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=5).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=6).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=7).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=8).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=9).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=10).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=11).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=12).border = Border(n,d,S,d)

       
    return





def formata_TP(planilha):
    
    #rows = planilha.nrows
    #columns = planilha.ncols
    
    planilha.freeze_panes = 'A3'

    # GERAL
    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
#    fontMasterAzul     = Font(color='000000FF', bold=True, size=12)
#    fontMasterVermelha = Font(color='00FF0000', bold=True, size=12)
    planilha.column_dimensions['A'].width = 11  
    planilha.column_dimensions['B'].width =7   
    planilha.column_dimensions['C'].width = 20  
    planilha.column_dimensions['D'].width = 20  
    planilha.column_dimensions['E'].width = 20 
    planilha.column_dimensions['F'].width = 20  
    planilha.column_dimensions['G'].width = 20  
    planilha.column_dimensions['H'].width = 20  




    # LINHA 1 = TITULO
    planilha.merge_cells('A1:H1')
    planilha.cell(1,  1).font = fontMasterPreta
    planilha.cell(1,  1).alignment = Alignment(horizontal='center')
    
    # LINHA 2 = CABEÇALHO
    planilha.cell(2,  1).font = fontMasterPreta
    planilha.cell(2,  2).font = fontMasterPreta
    planilha.cell(2,  3).font = fontMasterPreta
    planilha.cell(2,  4).font = fontMasterPreta
    planilha.cell(2,  5).font = fontMasterPreta
    planilha.cell(2,  6).font = fontMasterPreta
    planilha.cell(2,  7).font = fontMasterPreta
    planilha.cell(2,  8).font = fontMasterPreta

    planilha.cell(2,  1).alignment = Alignment(horizontal='center')
    planilha.cell(2,  2).alignment = Alignment(horizontal='center')
    planilha.cell(2,  3).alignment = Alignment(horizontal='center')
    planilha.cell(2,  4).alignment = Alignment(horizontal='center')
    planilha.cell(2,  5).alignment = Alignment(horizontal='center')
    planilha.cell(2,  6).alignment = Alignment(horizontal='center')
    planilha.cell(2,  7).alignment = Alignment(horizontal='center')
    planilha.cell(2,  8).alignment = Alignment(horizontal='center')

    # LINHA 3 A PENÚLTIMA = DADOS
    lini = 3
    lfin = planilha.max_row
    
    for linha in range(lini,lfin):
        planilha.cell(linha,  4).number_format = "#,##0.00"
        planilha.cell(linha,  5).number_format = "#,##0.00"
        planilha.cell(linha,  6).number_format = "#,##0.00"
        planilha.cell(linha,  7).number_format = "#,##0.00"
        planilha.cell(linha,  8).number_format = "#,##0.00"

        
    # LINHA FINAL = TOTAIS
    
    linha = planilha.max_row 
    merg = 'A'+str(linha)+':C'+str(linha)
    planilha.merge_cells(merg)
    planilha.cell(linha,  1, "TOTAIS")
    planilha.cell(linha,  1).alignment = Alignment(horizontal='center')
    planilha.cell(linha,  1).font = fontMasterPreta
    planilha.cell(linha,  4).font = fontMasterPreta
    planilha.cell(linha,  5).font = fontMasterPreta
    planilha.cell(linha,  6).font = fontMasterPreta
    planilha.cell(linha,  7).font = fontMasterPreta
    planilha.cell(linha,  8).font = fontMasterPreta
    planilha.cell(linha,  9).font = fontMasterPreta
    planilha.cell(linha,  4).number_format = "#,##0.00"
    planilha.cell(linha,  5).number_format = "#,##0.00"
    planilha.cell(linha,  6).number_format = "#,##0.00"
    planilha.cell(linha,  7).number_format = "#,##0.00"
    planilha.cell(linha,  8).number_format = "#,##0.00"
    

# BORDAS    

    s=Side(border_style=BORDER_THIN, color='00000000')
    S=Side(border_style=BORDER_MEDIUM, color='00000000')
    d=Side(border_style=BORDER_DOUBLE, color='00000000')
    n=Side(border_style=None, color='00000000') 
  
    for a in range(2,linha):
         planilha.cell(row=a, column=1).border = Border(d,s,n,n)
         planilha.cell(row=a, column=2).border = Border(s,s,n,n)
         planilha.cell(row=a, column=3).border = Border(s,s,n,n)
         planilha.cell(row=a, column=4).border = Border(s,s,n,n)
         planilha.cell(row=a, column=5).border = Border(s,s,n,n)
         planilha.cell(row=a, column=6).border = Border(s,s,n,n)
         planilha.cell(row=a, column=7).border = Border(s,s,n,n)
         planilha.cell(row=a, column=8).border = Border(s,d,n,n)
     
    planilha.cell(row=1, column=1).border = Border(d,n,S,d)
    planilha.cell(row=1, column=2).border = Border(n,n,S,d)
    planilha.cell(row=1, column=3).border = Border(n,n,S,d)
    planilha.cell(row=1, column=4).border = Border(n,n,S,d)
    planilha.cell(row=1, column=5).border = Border(n,n,S,d)
    planilha.cell(row=1, column=6).border = Border(n,n,S,d)
    planilha.cell(row=1, column=7).border = Border(n,n,S,d)
    planilha.cell(row=1, column=8).border = Border(n,S,S,d)

    planilha.cell(row=2, column=1).border = Border(d,s,d,s)
    planilha.cell(row=2, column=2).border = Border(s,s,d,s)
    planilha.cell(row=2, column=3).border = Border(s,s,d,s)
    planilha.cell(row=2, column=4).border = Border(s,s,d,s)
    planilha.cell(row=2, column=5).border = Border(s,s,d,s)
    planilha.cell(row=2, column=6).border = Border(s,s,d,s)
    planilha.cell(row=2, column=7).border = Border(s,s,d,s)
    planilha.cell(row=2, column=8).border = Border(s,d,d,s)
    
    planilha.cell(row=linha, column=1).border = Border(d,n,S,d)
    planilha.cell(row=linha, column=2).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=3).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=4).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=5).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=6).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=7).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=8).border = Border(n,d,S,d)

       
    return

def formata_EP(planilha):
    
    planilha.freeze_panes = 'A3'

    # GERAL
    fontMasterPreta    = Font(color='00000000', bold=True, size=12)
#    fontMasterAzul     = Font(color='000000FF', bold=True, size=12)
#    fontMasterVermelha = Font(color='00FF0000', bold=True, size=12)
    planilha.column_dimensions['A'].width = 11  
    planilha.column_dimensions['B'].width =7   
    planilha.column_dimensions['C'].width = 20  
    planilha.column_dimensions['D'].width = 20  
    planilha.column_dimensions['E'].width = 20 
    planilha.column_dimensions['F'].width = 20  
    planilha.column_dimensions['G'].width = 20  
    planilha.column_dimensions['H'].width = 20  
    planilha.column_dimensions['I'].width = 20  



    # LINHA 1 = TITULO
    planilha.merge_cells('A1:I1')
    planilha.cell(1,  1).font = fontMasterPreta
    planilha.cell(1,  1).alignment = Alignment(horizontal='center')
    
    # LINHA 2 = CABEÇALHO
    planilha.cell(2,  1).font = fontMasterPreta
    planilha.cell(2,  2).font = fontMasterPreta
    planilha.cell(2,  3).font = fontMasterPreta
    planilha.cell(2,  4).font = fontMasterPreta
    planilha.cell(2,  5).font = fontMasterPreta
    planilha.cell(2,  6).font = fontMasterPreta
    planilha.cell(2,  7).font = fontMasterPreta
    planilha.cell(2,  8).font = fontMasterPreta
    planilha.cell(2,  9).font = fontMasterPreta
    planilha.cell(2,  1).alignment = Alignment(horizontal='center')
    planilha.cell(2,  2).alignment = Alignment(horizontal='center')
    planilha.cell(2,  3).alignment = Alignment(horizontal='center')
    planilha.cell(2,  4).alignment = Alignment(horizontal='center')
    planilha.cell(2,  5).alignment = Alignment(horizontal='center')
    planilha.cell(2,  6).alignment = Alignment(horizontal='center')
    planilha.cell(2,  7).alignment = Alignment(horizontal='center')
    planilha.cell(2,  8).alignment = Alignment(horizontal='center')
    planilha.cell(2,  9).alignment = Alignment(horizontal='center')
    
    # LINHA 3 A PENÚLTIMA = DADOS
    lini = 3
    lfin = planilha.max_row
    
    for linha in range(lini,lfin):
        planilha.cell(linha,  4).number_format = "#,##0.00"
        planilha.cell(linha,  5).number_format = "#,##0.00"
        planilha.cell(linha,  6).number_format = "#,##0.00"
        planilha.cell(linha,  7).number_format = "#,##0.00"
        planilha.cell(linha,  8).number_format = "#,##0.00"
        planilha.cell(linha,  9).number_format = "#,##0.00"
        
    # LINHA FINAL = TOTAIS
    
    linha = planilha.max_row 
    merg = 'A'+str(linha)+':C'+str(linha)
    planilha.merge_cells(merg)
    planilha.cell(linha,  1, "TOTAIS")
    planilha.cell(linha,  1).alignment = Alignment(horizontal='center')
    planilha.cell(linha,  1).font = fontMasterPreta
    planilha.cell(linha,  4).font = fontMasterPreta
    planilha.cell(linha,  5).font = fontMasterPreta
    planilha.cell(linha,  6).font = fontMasterPreta
    planilha.cell(linha,  7).font = fontMasterPreta
    planilha.cell(linha,  8).font = fontMasterPreta
    planilha.cell(linha,  9).font = fontMasterPreta
    planilha.cell(linha,  4).number_format = "#,##0.00"
    planilha.cell(linha,  5).number_format = "#,##0.00"
    planilha.cell(linha,  6).number_format = "#,##0.00"
    planilha.cell(linha,  7).number_format = "#,##0.00"
    planilha.cell(linha,  8).number_format = "#,##0.00"
    planilha.cell(linha,  9).number_format = "#,##0.00"
    

# BORDAS    

    s=Side(border_style=BORDER_THIN, color='00000000')
    S=Side(border_style=BORDER_MEDIUM, color='00000000')
    d=Side(border_style=BORDER_DOUBLE, color='00000000')
    n=Side(border_style=None, color='00000000') 
  
    for a in range(2,linha):
         planilha.cell(row=a, column=1).border = Border(d,s,n,n)
         planilha.cell(row=a, column=2).border = Border(s,s,n,n)
         planilha.cell(row=a, column=3).border = Border(s,s,n,n)
         planilha.cell(row=a, column=4).border = Border(s,s,n,n)
         planilha.cell(row=a, column=5).border = Border(s,s,n,n)
         planilha.cell(row=a, column=6).border = Border(s,s,n,n)
         planilha.cell(row=a, column=7).border = Border(s,s,n,n)
         planilha.cell(row=a, column=8).border = Border(s,s,n,n)
         planilha.cell(row=a, column=9).border = Border(s,d,n,n)
     
    planilha.cell(row=1, column=1).border = Border(d,n,S,d)
    planilha.cell(row=1, column=2).border = Border(n,n,S,d)
    planilha.cell(row=1, column=3).border = Border(n,n,S,d)
    planilha.cell(row=1, column=4).border = Border(n,n,S,d)
    planilha.cell(row=1, column=5).border = Border(n,n,S,d)
    planilha.cell(row=1, column=6).border = Border(n,n,S,d)
    planilha.cell(row=1, column=7).border = Border(n,n,S,d)
    planilha.cell(row=1, column=8).border = Border(n,n,S,d)
    planilha.cell(row=1, column=9).border = Border(n,S,S,d)

    planilha.cell(row=2, column=1).border = Border(d,s,d,s)
    planilha.cell(row=2, column=2).border = Border(s,s,d,s)
    planilha.cell(row=2, column=3).border = Border(s,s,d,s)
    planilha.cell(row=2, column=4).border = Border(s,s,d,s)
    planilha.cell(row=2, column=5).border = Border(s,s,d,s)
    planilha.cell(row=2, column=6).border = Border(s,s,d,s)
    planilha.cell(row=2, column=7).border = Border(s,s,d,s)
    planilha.cell(row=2, column=8).border = Border(s,s,d,s)
    planilha.cell(row=2, column=9).border = Border(s,d,d,s)
    
    planilha.cell(row=linha, column=1).border = Border(d,n,S,d)
    planilha.cell(row=linha, column=2).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=3).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=4).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=5).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=6).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=7).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=8).border = Border(n,n,S,d)
    planilha.cell(row=linha, column=9).border = Border(n,d,S,d)

       
    return


def Busca_TPS(vIE,vDataIni):
    global ret

    query ="""SELECT /*+PARALLEL(M,15) */
                'TBRA'                            				AS EMPS_COD,
                TO_CHAR(M.DATA_EMISSAO,'MM/YY')   				AS MES_ANO,
                REPLACE(S.SERIE,' ','')                    		AS SERIE,
                M.CFOP                            				AS CFOP,
                SUM(NVL(M.VALOR_TOTAL, 0) - NVL(M.DESCONTO, 0)) AS VLR_LIQUIDO,
                SUM(NVL(M.VALOR_TOTAL, 0))      				AS VLR_SERVICO,
                SUM(NVL(M.BASE_ICMS, 0))        				AS VLR_BASE_ICMS,
                SUM(NVL(M.VALOR_ICMS, 0))       				AS VLR_ICMS,
                SUM(NVL(M.ISENTAS_ICMS, 0))     				AS VLR_ISENTAS,
                SUM(NVL(M.OUTROS_VALORES, 0))   				AS VLR_OUTRAS,
                SUM(NVL(M.DESCONTO, 0))         				AS VLR_DESCONTO,
                0                               				AS VLR_REDUCAO
            FROM GFCARGA.TSH_ITEM_CONV_115@C2 M 
                INNER JOIN GFCARGA.TSH_SERIE_LEVANTAMENTO@C2 S 
                ON S.ID_SERIE_LEVANTAMENTO = M.ID_SERIE_LEVANTAMENTO AND S.MES_ANO = TO_DATE('%s', 'DD/MM/YYYY')
            WHERE M.UF_FILIAL = 'RJ'
                AND S.EMPS_COD = 'TBRA'
                AND S.FILI_COD IN 
                    (select f1.fili_cod from openrisow.filial f1 
                    where f1.emps_cod = 'TBRA' 
                    AND f1.FILI_COD_INSEST = '%s'
                    ) 
                AND M.DATA_EMISSAO >= TO_DATE('%s', 'DD/MM/YYYY')
                AND M.DATA_EMISSAO <= LAST_DAY(TO_DATE('%s','DD/MM/YYYY'))
                AND M.SIT_DOC = 'N'
            GROUP BY TO_CHAR(M.DATA_EMISSAO,'MM/YY'),
                REPLACE(S.SERIE,' ',''),
                M.CFOP
            ORDER BY 3,4
    """%(vDataIni,vIE,vDataIni,vDataIni)

    banco.executa(query)
    result = banco.fetchall()
    return(result)


def Busca_TRT(vIE,vDataIni):
    global ret

    query = """SELECT  EMPS_COD,
            		FILI_COD,
            		SERIE,
            		NUMERO_NF,
                    DATA_EMISSAO,
            		CST,
            		DSC_COMPL,
            		SUM(VALOR_LIQUIDO),
            		SUM(ALIQUOTA_ICMS),
            		SUM(VLR_BASE_ICMS),
            		SUM(VLR_ICMS),
            		SUM(VLR_ISENTAS),
            		SUM(VLR_OUTRAS)
            	FROM 
            		(SELECT /*+ parallel(8) */ 
                		i.EMPS_COD                              AS EMPS_COD
                		,i.FILI_COD                             AS FILI_COD
                		,REPLACE(i.INFST_SERIE,' ','')          AS SERIE
                		,i.INFST_NUM                            AS NUMERO_NF
                		,to_char(i.INFST_DTEMISS,'YYYY/MM/DD')  AS DATA_EMISSAO
                		,ESTB_COD                               AS CST
                		,INFST_DSC_COMPL                        AS DSC_COMPL
                		,(CASE 	WHEN I.INFST_DTEMISS < TO_DATE('01/01/2017','dd/mm/yyyy') THEN NVL(INFST_VAL_SERV,0) - NVL(INFST_VAL_DESC,0)
            	            ELSE NVL(INFST_VAL_CONT,0)
                	        END) 	                            AS VALOR_LIQUIDO--Antes VALOR_CONTABIL
                		,INFST_ALIQ_ICMS                        AS ALIQUOTA_ICMS
                		,INFST_BASE_ICMS                        AS VLR_BASE_ICMS
                		,INFST_VAL_ICMS                         AS VLR_ICMS
                		,INFST_ISENTA_ICMS                      AS VLR_ISENTAS
                		,INFST_OUTRAS_ICMS                      AS VLR_OUTRAS
            		FROM openrisow.ITEM_NFTL_SERV    i,
                 		openrisow.filial f
             		WHERE 1 =1 
                		AND f.EMPS_COD = i.EMPS_COD
                		AND f.FILI_COD = i.FILI_COD
                		AND NVL(INFST_IND_CANC,'N') = 'N'	
                		AND f.FILI_COD_INSEST = '%s' --<<PARAMETRO: INSCRICAO ESTADUAL>>
                		AND i.INFST_dtemiss >= TO_DATE('%s', 'DD/MM/YYYY') --<<PARAMETRO: DATA_INI>>
            		    AND i.INFST_dtemiss <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1) --<<PARAMETRO: DATA_INI>>
                		AND f.EMPS_COD = 'TBRA'
                		AND (
                     		(nvl(INFST_BASE_ICMS,0) <> 0 and nvl(INFST_VAL_ICMS,0) = 0)
                  			or (nvl(INFST_BASE_ICMS,0) = 0  and nvl(INFST_VAL_ICMS,0) <> 0)
                  		))
                  GROUP BY 
                  	(EMPS_COD,
            		FILI_COD,
            		SERIE,
            		NUMERO_NF,
                    DATA_EMISSAO,
            		CST,
            		DSC_COMPL)
        """%(vIE,vDataIni,vDataIni)

    banco.executa(query)
    result = banco.fetchall()
    return(result)

def Busca_TRS(vIE,vDataIni):
    global ret

    query ="""SELECT /*+ PARALLEL(15) */
                 INF.EMPS_COD                                   EMPS_COD
                ,TO_CHAR(INFST_DTEMISS,'MM/YY')                 MES_ANO
                ,REPLACE(INFST_SERIE,' ','')                    SERIE
                ,CFOP                                           CFOP
                ,SUM    (CASE WHEN INF.INFST_DTEMISS < TO_DATE('01/01/2017','dd/mm/yyyy') THEN NVL(INFST_VAL_SERV,0) - NVL(INFST_VAL_DESC,0)
                        ELSE NVL(INFST_VAL_CONT,0)
                        END)                                    VALOR_LIQUIDO
                ,SUM(INFST_VAL_SERV)                            VLR_SERVICO
                ,SUM(INFST_BASE_ICMS)                           VLR_BASE_ICMS
                ,SUM(INFST_VAL_ICMS)                            VLR_ICMS
                ,SUM(INFST_ISENTA_ICMS)                         VLR_ISENTAS
                ,SUM(INFST_OUTRAS_ICMS)                         VLR_OUTRAS
                ,SUM(INFST_VAL_DESC)                            VLR_DESCONTO
                ,SUM(INFST_VAL_RED)                             VLR_REDUCAO
                
            FROM
                OPENRISOW.ITEM_NFTL_SERV INF,
                OPENRISOW.FILIAL F 
            WHERE 1=1
                AND INF.EMPS_COD        = 'TBRA'
                AND INF.EMPS_COD        = F.EMPS_COD 
                AND INF.FILI_COD        = F.FILI_COD
                AND INFST_IND_CANC      = 'N'
                AND F.FILI_COD_INSEST   = '%s' 
                AND INFST_DTEMISS       >= TO_DATE('%s','DD/MM/YYYY') 
                AND INFST_DTEMISS       <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1)
            GROUP BY
                INF.EMPS_COD,
                TO_CHAR(INFST_DTEMISS,'MM/YY'),
                REPLACE(INFST_SERIE,' ',''),
                CFOP
            ORDER BY 3,4
    """%(vIE,vDataIni,vDataIni)
    
    banco.executa(query)
    result = banco.fetchall()
    return(result)




def Busca_AICMS(vIE,vDataIni):
    global ret

    query =  """SELECT
                    emps_cod,
                    fili_cod_insest,
                    codigo,
                    item,
                    seq,
                    TO_CHAR(data,'DD/MM/YYYY'),
                    gia2_ocor,
                    gia2_valor
                FROM
                    openrisow.gia2
                WHERE 1=1
                    AND emps_cod = 'TBRA'
                    AND fili_cod_insest = '%s'
                    AND data >= TO_DATE('%s','DD/MM/YYYY')
                    AND data <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1)
                    AND codigo in('002', '003', '006')
             """%(vIE,vDataIni,vDataIni)
 
    banco.executa(query)
    result = banco.fetchall()
    return(result)


def Busca_SN(vIE,vDataIni):
    global ret

    query =  """WITH TMP_NF_MENOR AS
                  (SELECT nf.*
                  FROM
                    (
                    -- MONTH
                    SELECT nf.emps_cod,
                      nf.fili_cod,
                      nf.mdoc_cod,
                      nf.mnfst_serie,
                      LPAD('0',LENGTH(nf.mnfst_num) ,'0') mnfst_num ,
                      nf.mnfst_dtemiss ,
                      TO_CHAR(nf.mnfst_dtemiss, 'YYYY-MM') PERIODO ,
                      TO_NUMBER(0) NUM_NOTA ,
                      ROW_NUMBER() OVER(PARTITION BY nf.emps_cod,nf.fili_cod,nf.mdoc_cod,nf.mnfst_serie, TRUNC(nf.mnfst_dtemiss,'MM') ORDER BY TO_NUMBER(nf.mnfst_num)) rnk
                    FROM OPENRISOW.MESTRE_NFTL_SERV nf
                    INNER JOIN OPENRISOW.filial f
                    ON (nf.emps_cod = f.emps_cod
                    AND nf.fili_cod = f.fili_cod)
                      -- FILTRO
                    WHERE f.emps_cod                                    = 'TBRA'   
                    AND f.FILI_COD_INSEST                               = '%s' --<<PARAMETRO> INSCRICAO ESTADUAL>>
                    --AND nf.mnfst_serie = 'U' AND nf.mdoc_cod = '22'
                    AND UPPER(TRANSLATE(nf.mnfst_serie,'x ','x')) NOT  IN ('AS1', 'AS2', 'AS3', 'T1') --FIXO, NAO ALTERAR
                    AND (UPPER(TRANSLATE(nf.mnfst_serie,'x ','x')) NOT IN ('ASS') OR nf.mnfst_dtemiss >= TO_DATE('01/04/2017','DD/MM/YYYY')) --DATA FIXA. NÃO ALTERAR--
                    AND nf.MNFST_DTEMISS                                < TRUNC(TO_DATE('01/08/2018','DD/MM/YYYY'),'MM') --DATA FIXA. NÃO ALTERAR--
                      -- DATA INICIO : 01/03/2016
                    AND nf.MNFST_DTEMISS >= TRUNC(TO_DATE('%s','DD/MM/YYYY'),'MM') --<<PARAMETRO:DATA_INI>>
                      -- DATA FIM  : 31/03/2016
                    AND nf.MNFST_DTEMISS < ADD_MONTHS(TRUNC(TO_DATE('%s','DD/MM/YYYY'),'MM'), 1) --<<PARAMETRO:DATA_INI>>
                    ) nf
                  WHERE nf.rnk = 1 -- ROWNUM = 1
                  ) ,
                  TMP_NF_ANT AS --PARTE DO CURSOR c_sanea - NIVEL 1 CONSULTA MESTRE_NFTL_SERV -  PEGA ULTIMA SERIE DO MÊS ANTERIOR
                  (SELECT nf.*
                  FROM
                    (
                    -- MONTH
                    SELECT nf.emps_cod,
                      nf.fili_cod,
                      nf.mdoc_cod,
                      nf.mnfst_serie,
                      nf.mnfst_num,
                      nf.mnfst_dtemiss ,
                      'SEQUENCIAL' PERIODO ,
                      TO_NUMBER(nf.mnfst_num) NUM_NOTA ,
                      ROW_NUMBER() OVER(PARTITION BY nf.emps_cod,nf.fili_cod,nf.mdoc_cod,nf.mnfst_serie, nf.mnfst_dtemiss ORDER BY TO_NUMBER(nf.mnfst_num) DESC) rnk
                    FROM OPENRISOW.MESTRE_NFTL_SERV nf
                    INNER JOIN OPENRISOW.filial f
                    ON (nf.emps_cod = f.emps_cod
                    AND nf.fili_cod = f.fili_cod)
                      -- FILTRO
                    WHERE f.emps_cod                                    = 'TBRA'
                    AND f.FILI_COD_INSEST                               = '%s' --<<PARAMETRO:INSCRIÇÃO ESTADUAL>>
                    --AND nf.mnfst_serie = 'U' 
                    --AND nf.mdoc_cod = '22'
                    AND UPPER(TRANSLATE(nf.mnfst_serie,'x ','x')) NOT  IN ('AS1', 'AS2', 'AS3', 'T1') --FIXO, NÃO ALTERAR
                    AND (UPPER(TRANSLATE(nf.mnfst_serie,'x ','x')) NOT IN ('ASS') OR nf.mnfst_dtemiss >= TO_DATE('01/04/2017','DD/MM/YYYY')) --DATA FIXA. NÃO ALTERAR--
                    AND nf.mnfst_dtemiss                                =
                      (SELECT MAX(MNFST_DTEMISS)
                      FROM OPENRISOW.MESTRE_NFTL_SERV nf1
                      WHERE nf1.emps_cod    = nf.emps_cod
                      AND nf1.fili_cod      = nf.fili_cod
                      AND nf1.mnfst_serie   = nf.mnfst_serie
                      AND nf1.mdoc_cod      = nf.mdoc_cod
                      AND nf1.mnfst_dtemiss < (
                        CASE
                          WHEN TRUNC(TO_DATE('%s','DD/MM/YYYY'),'MM') < TO_DATE('01/08/2018','DD/MM/YYYY') --<<DATA DE 01/08/2018 É FIXA a DATA DE 01/01/2020 É PARAMETRO:DATA_INI>>
                          AND LAST_DAY(TO_DATE('%s','DD/MM/YYYY'))    > TO_DATE('01/08/2018','DD/MM/YYYY') --<<DATA DE 01/08/2018 É FIXA a DATA DE 01/01/2020 É PARAMETRO:DATA_INI>>
                          THEN TO_DATE('01/08/2018','DD/MM/YYYY') -- FIXO
                          WHEN TRUNC(TO_DATE('%s','DD/MM/YYYY'),'MM') >= TO_DATE('01/08/2018','DD/MM/YYYY') --<<DATA DE 01/08/2018 É FIXA a DATA DE 01/01/2020 É PARAMETRO:DATA_INI>>
                          THEN TRUNC(TO_DATE('%s','DD/MM/YYYY'),'MM') --<<PARAMETRO:DATA_INI>>
                          ELSE NULL
                        END)
                      )
                    ) nf
                  WHERE nf.rnk = 1 -- ROWNUM = 1
                  ) ,
                  TMP_NF AS --PARTE DO CURSOR c_sanea - NIVEL 1 CONSULTA MESTRE_NFTL_SERV - UNE COM A SERIE DO MES ATUAL
                  (SELECT nf.emps_cod,
                    nf.fili_cod,
                    nf.mdoc_cod,
                    nf.mnfst_serie,
                    nf.mnfst_num,
                    nf.mnfst_dtemiss ,
                    (
                    CASE
                      WHEN nf.mnfst_dtemiss >= TO_DATE('01/08/2018','DD/MM/YYYY') --FIXO
                      THEN 'SEQUENCIAL'
                      ELSE TO_CHAR(nf.mnfst_dtemiss, 'YYYY-MM')
                    END) PERIODO ,
                    TO_NUMBER(nf.mnfst_num) NUM_NOTA ,
                    1 AS rnk
                  FROM openrisow.mestre_nftl_serv nf
                  INNER JOIN OPENRISOW.filial f
                  ON (nf.emps_cod = f.emps_cod
                  AND nf.fili_cod = f.fili_cod)
                    -- FILTRO
                  WHERE f.emps_cod                                    = 'TBRA'
                  AND f.FILI_COD_INSEST                               = '%s' --<<PARAMETRO:INSCRICAO ESTADUAL>>
                  AND UPPER(TRANSLATE(nf.mnfst_serie,'x ','x')) NOT  IN ('AS1', 'AS2', 'AS3', 'T1') --FIXO
                  AND (UPPER(TRANSLATE(nf.mnfst_serie,'x ','x')) NOT IN ('ASS') OR nf.mnfst_dtemiss >= TO_DATE('01/04/2017','DD/MM/YYYY')) --FIXO
                    -- DATA INICIO : 01/01/2020
                  AND nf.MNFST_DTEMISS >= TRUNC(TO_DATE('%s','DD/MM/YYYY'),'MM') --<<PARAMETRO:DATA_INI>>
                    -- DATA FIM  : 01/01/2020
                  AND nf.MNFST_DTEMISS < ADD_MONTHS(TRUNC(TO_DATE('%s','DD/MM/YYYY'),'MM'), 1) --<<PARAMETRO:DATA_INI>>
                  ) ,
                  TMP_REL_NF AS
                  (SELECT LEAD(NUM_NOTA) OVER(PARTITION BY emps_cod, fili_cod, mdoc_cod , mnfst_serie, PERIODO ORDER BY NUM_NOTA) PROX_NUM_NOTA ,
                    TMP.*
                  FROM
                    ( SELECT * FROM TMP_NF_MENOR
                    UNION ALL
                      ( SELECT * FROM TMP_NF_ANT
                      UNION ALL
                      SELECT * FROM TMP_NF
                      )
                    ) TMP
                  ) ,
                  TMP_REL AS
                  (
                  --PARTE DO CURSOR c_sanea - NIVEL 3 CONSULTA TMP_REL_NF - LEVANTA O QUE ESTA ERRADO BASEADO NA SEQUENCIA LEVANTADA A CIMA
                  SELECT nf.*,
                    (
                    CASE
                      WHEN nf.PROX_NUM_NOTA = nf.NUM_NOTA
                      THEN 'D' -- NF duplicada
                      WHEN nf.PROX_NUM_NOTA - nf.NUM_NOTA > 1
                      THEN 'S' -- Salto de sequencia
                      ELSE NULL
                    END) TIPO_ERRO,
                    (nf.NUM_NOTA      + 1) INICIO_SALTO,
                    (nf.PROX_NUM_NOTA - 1) FIM_SALTO
                  FROM TMP_REL_NF nf
                  WHERE ( (nf.PROX_NUM_NOTA - nf.NUM_NOTA > 1)
                  OR (nf.PROX_NUM_NOTA                    = nf.NUM_NOTA) )
                  )
                SELECT  /*+ parallel(15) */ data_nf.MNFST_SERIE,INICIO_SALTO,FIM_SALTO,PERIODO,
                  (
                  CASE
                    WHEN data_nf.PERIODO = 'SEQUENCIAL'
                    THEN
                      CASE
                        WHEN TRUNC(TO_DATE('%s','DD/MM/YYYY'),'MM') > TO_DATE('01/08/2018','DD/MM/YYYY') --<<DATA DE 01/08/2018 É FIXA a DATA DE 01/01/2020 É PARAMETRO:DATA_INI>>
                        THEN TRUNC(TO_DATE('%s','DD/MM/YYYY'),'MM') --<<PARAMETRO:DATA_INI>>
                        ELSE TO_DATE('01/08/2018','DD/MM/YYYY') --FIXO
                      END
                    ELSE TRUNC(data_nf.DAT_NOTA,'MM')
                  END) PERIDO_INICIAL,
                  (
                  CASE
                    WHEN data_nf.PERIODO = 'SEQUENCIAL'
                    THEN LAST_DAY(TRUNC(TO_DATE('%s','DD/MM/YYYY'),'MM')) --<<PARAMETRO:DATA_INI>>
                    ELSE LAST_DAY(TRUNC(data_nf.DAT_NOTA,'MM'))
                  END) PERIDO_FINAL
                FROM
                  (SELECT
                    nf.mnfst_serie,
                    nf.mnfst_dtemiss AS DAT_NOTA,
                    tmp.INICIO_SALTO,
                    tmp.FIM_SALTO,
                    tmp.PERIODO
                  FROM openrisow.mestre_nftl_serv nf,
                    TMP_REL tmp
                  WHERE tmp.PERIODO           = 'SEQUENCIAL'
                  AND nf.emps_cod             = tmp.emps_cod
                  AND nf.fili_cod             = tmp.fili_cod
                  AND nf.mdoc_cod             = tmp.mdoc_cod
                  AND nf.mnfst_serie          = tmp.mnfst_serie
                  AND TO_NUMBER(nf.mnfst_num) = tmp.PROX_NUM_NOTA
                  AND nf.mnfst_dtemiss       >= (
                    CASE
                      WHEN TRUNC(TO_DATE('%s','DD/MM/YYYY'),'MM') > TO_DATE('01/08/2018','DD/MM/YYYY') --<<DATA DE 01/08/2018 É FIXA a DATA DE 01/01/2020 É PARAMETRO:DATA_INI>>
                      THEN TRUNC(TO_DATE('%s','DD/MM/YYYY'),'MM') --<<PARAMETRO:DATA_INI>>
                      ELSE TO_DATE('01/08/2018','DD/MM/YYYY') --FIXO
                    END)
                  AND nf.mnfst_dtemiss <= LAST_DAY(TRUNC(TO_DATE('%s','DD/MM/YYYY'),'MM')) --<<PARAMETRO:DATA_INI>>
                  UNION ALL
                  SELECT
                    REPLACE(nf.mnfst_serie,' ',''),
                    nf.mnfst_dtemiss AS DAT_NOTA,
                    tmp.INICIO_SALTO,
                    tmp.FIM_SALTO,
                    tmp.PERIODO
                  FROM openrisow.mestre_nftl_serv nf,
                    TMP_REL tmp
                  WHERE tmp.PERIODO          != 'SEQUENCIAL'
                  AND nf.emps_cod             = tmp.emps_cod
                  AND nf.fili_cod             = tmp.fili_cod
                  AND nf.mdoc_cod             = tmp.mdoc_cod
                  AND nf.mnfst_serie          = tmp.mnfst_serie
                  AND TO_NUMBER(nf.mnfst_num) = tmp.PROX_NUM_NOTA
                  AND nf.mnfst_dtemiss       >= TRUNC(tmp.mnfst_dtemiss,'MM')
                  AND nf.mnfst_dtemiss       <= LAST_DAY(TRUNC(tmp.mnfst_dtemiss,'MM'))
                  ) data_nf
        """%(vIE,vDataIni,vDataIni,vIE,vDataIni,vDataIni,vDataIni,vDataIni,vIE,vDataIni,vDataIni,vDataIni,vDataIni,vDataIni,vDataIni,vDataIni,vDataIni)
 
    banco.executa(query)
    result = banco.fetchall()
    return(result)


def Busca_HIE(vIE,vDataIni):
    global ret

    query =  """SELECT 'openrisow.item_nfem_merc' NEGOGIO ,
                    AUX.*
                FROM GFCADASTRO.GAP_ITEM_NFEM AUX
                WHERE NVL(UPPER(TRIM(AUX.CONTROLE)),'ABERTO') != 'ABERTO'
                    AND INFEM_DTENTR >= TO_DATE('%s','DD/MM/YYYY')
                    AND INFEM_DTENTR <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1)
                    AND NVL(UPPER(TRIM(AUX.TIPO_CONTROLE)),'ORIGEM') = 'ORIGEM'
              """%(vDataIni,vDataIni)
 
    banco.executa(query)
    result = banco.fetchall()
    return(result)


def Busca_HME(vIE,vDataIni):
    global ret

    query =  """SELECT 'openrisow.mestre_nfen_merc' NEGOGIO ,
                        AUX.*
                    FROM GFCADASTRO.GAP_MESTRE_NFEN AUX
                    WHERE NVL(UPPER(TRIM(AUX.CONTROLE)),'ABERTO') != 'ABERTO'
                        AND MNFEM_DTENTR >= TO_DATE('%s','DD/MM/YYYY')
                        AND MNFEM_DTENTR <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1)
                        AND NVL(UPPER(TRIM(AUX.TIPO_CONTROLE)),'ORIGEM') = 'ORIGEM'
            """%(vDataIni,vDataIni)
 
    banco.executa(query)
    result = banco.fetchall()
    return(result)


def Busca_HMS(vIE,vDataIni):
    global ret

    query =  """SELECT 'openrisow.mestre_nfsd_merc' NEGOGIO ,
                        AUX.*
                    FROM GFCADASTRO.GAP_MESTRE_NFSD AUX
                    WHERE NVL(UPPER(TRIM(AUX.CONTROLE)),'ABERTO') != 'ABERTO'
                        AND MNFSM_DTEMISS >= TO_DATE('%s','DD/MM/YYYY')
                        AND MNFSM_DTEMISS <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1)
                        AND NVL(UPPER(TRIM(AUX.TIPO_CONTROLE)),'ORIGEM') = 'ORIGEM'
            """%(vDataIni,vDataIni)
 
    banco.executa(query)
    result = banco.fetchall()
    return(result)


def Busca_HIS(vIE,vDataIni):
    global ret

    query =  """SELECT 'openrisow.item_nfsd_merc' NEGOGIO ,
                        AUX.*
                    FROM GFCADASTRO.GAP_ITEM_NFSD AUX
                    WHERE NVL(UPPER(TRIM(AUX.CONTROLE)),'ABERTO') != 'ABERTO'
                        AND INFSM_DTEMISS >= TO_DATE('%s','DD/MM/YYYY')
                        AND INFSM_DTEMISS <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1)
                        AND NVL(UPPER(TRIM(AUX.TIPO_CONTROLE)),'ORIGEM') = 'ORIGEM'
            """%(vDataIni,vDataIni)
 
    banco.executa(query)
    result = banco.fetchall()
    return(result)


def Busca_RCFOP(vIE,vDataIni):
    global ret

    query ="""SELECT DISTINCT
                    ( res.cfop_cod )			            AS CFOP,
                    SUM(nvl(res.val_cont, 0))               AS VLR_CONTABIL,
                    SUM(nvl(res.val_bas_icms, 0))           AS VLR_BASE_ICMS,
                    SUM(nvl(res.val_icms, 0))               AS VLR_ICMS,
                    SUM(nvl(res.val_ise_icms, 0))           AS VLR_ISENTAS,
                    SUM(nvl(res.val_out_icms, 0))           AS VLR_OUTRAS,
                    SUM(nvl(res.val_cont, 0) - nvl(res.val_bas_icms, 0) - nvl(res.val_ise_icms, 0) - nvl(res.val_out_icms, 0)) AS DIFERENCA
                FROM
                    openrisow.resumo_fiscal res
                WHERE
                    1 = 1
                    AND (   (ind_es = 'S' AND origem = 'T')
                         OR  (ind_es = 'S' AND origem = 'M')
                         OR  (ind_es = 'E' AND origem = 'T')
                         OR  (ind_es = 'E' AND origem = 'M')
                         )
                    AND resdata         >= TO_DATE('%s', 'DD/MM/YYYY')              
                    AND resdata         <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1) 
                    AND FILI_COD_INSEST = '%s'                                     
                    AND CFOP_COD        <> '0000'
                GROUP BY
                    res.cfop_cod
                ORDER BY
                    res.cfop_cod ASC
            """%(vDataIni,vDataIni,vIE)
 
    banco.executa(query)
    result = banco.fetchall()
    return(result)

def Busca_EMM(vIE,vDataIni):
    global ret

    query =  """SELECT /*+ PARALLEL(8) */ 
                    ''   								AS GAP,
                    'ALTERAR/MANTER/EXCLUIR/INCLUIR'    AS ACAO,
                    'mestre_nfen_merc'            		AS TABELA,
                    I.ROWID                    			AS ROW_ID,
                    I.*
                FROM  OPENRISOW.MESTRE_NFEN_MERC I,
                      OPENRISOW.FILIAL F
                WHERE  F.EMPS_COD = I.EMPS_COD
                  AND F.FILI_COD = I.FILI_COD
                  AND I.EMPS_COD = 'TBRA'
                  AND F.FILI_COD_INSEST = '%s' --<<PARAMETRO: INSCRICAO ESTADUAL>>
                  AND I.MNFEM_DTENTR   >= TO_DATE('%s','DD/MM/YYYY') --<<PARAMETRO: DATA_INI>>
                  AND I.MNFEM_DTENTR   <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1) --<<PARAMETRO: DATA_INI>>
             """%(vIE,vDataIni,vDataIni)
 
    banco.executa(query)
    result = banco.fetchall()
    return(result)

def Busca_EMI(vIE,vDataIni):
    global ret

    query =  """    SELECT /*+ PARALLEL(8) */ 
                        ''					   				AS GAP,
                        'ALTERAR/MANTER/EXCLUIR/INCLUIR'  	AS ACAO,
                        'item_nfem_merc'    	        	AS TABELA,
                        I.ROWID                 	   		AS ROW_ID,
                        I.*
                    FROM
                        openrisow.item_nfem_merc    i,
                        openrisow.filial            f
                    WHERE
                            f.emps_cod = i.emps_cod
                        AND f.fili_cod = i.fili_cod
                        AND i.emps_cod = 'TBRA'
                        AND f.fili_cod_insest = '%s'
                        AND i.infem_dtentr >= TO_DATE('%s', 'DD/MM/YYYY')
                        AND i.infem_dtentr  < ADD_MONTHS(TO_DATE('%s', 'DD/MM/YYYY'),1)
    """%(vIE,vDataIni,vDataIni)
 
    banco.executa(query)
    result = banco.fetchall()
    return(result)

def Busca_SMM(vIE,vDataIni):
    global ret
    query =  """SELECT /*+ PARALLEL(8) */ 
                    ''   								AS GAP,
                    'ALTERAR/MANTER/EXCLUIR/INCLUIR'    AS ACAO,
                    'mestre_nfsd_merc'            		AS TABELA,
                    I.ROWID                    			AS ROW_ID,
                    I.*
                FROM  OPENRISOW.MESTRE_NFSD_MERC I,
                      OPENRISOW.FILIAL F
                WHERE  F.EMPS_COD = I.EMPS_COD
                  AND F.FILI_COD = I.FILI_COD
                  AND I.EMPS_COD = 'TBRA'
                  AND F.FILI_COD_INSEST = '%s'
                  AND I.MNFSM_DTEMISS   >= TO_DATE('%s','DD/MM/YYYY')
                  AND I.MNFSM_DTEMISS   <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1)
    """%(vIE,vDataIni,vDataIni)

    banco.executa(query)
    result = banco.fetchall()
    return(result)

def Busca_SMI(vIE,vDataIni):
    global ret
    query = """SELECT /*+ PARALLEL(8) */ 
                    ''   								AS GAP,
                    'ALTERAR/MANTER/EXCLUIR/INCLUIR'    AS ACAO,
                    'item_nfsd_merc'            		AS TABELA,
                    I.ROWID                    			AS ROW_ID,
                    I.*
                FROM  OPENRISOW.ITEM_NFSD_MERC I,
                      OPENRISOW.FILIAL F
                WHERE  F.EMPS_COD = I.EMPS_COD
                  AND F.FILI_COD = I.FILI_COD
                  AND I.EMPS_COD = 'TBRA'
                  AND F.FILI_COD_INSEST = '%s'
                  AND I.INFSM_DTEMISS   >= TO_DATE('%s','DD/MM/YYYY')
                  AND I.INFSM_DTEMISS   <  ADD_MONTHS(TO_DATE('%s','DD/MM/YYYY'),1)
    """%(vIE,vDataIni,vDataIni)
  
    banco.executa(query)
    result = banco.fetchall()
    return(result)


def set_border( ws, cell_range):
        border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                cell.border = border


def set_border_edsi( ws, cell_range, edsi):
    b=[]
 
    n=Side(border_style= None    , color='00000000')
    f=Side(border_style='thin'   , color='00000000')
    m=Side(border_style='medium' , color='00000000')
    d=Side(border_style='double' , color='00000000')
    p=Side(border_style='dotted' , color='00000000')
    t=Side(border_style='dashed' , color='00000000')
    g=Side(border_style='thick'  , color='00000000')
    
    for x in (0,1,2,3):
        if (edsi[x]=="n"):
            b.append(n)
        elif (edsi[x]=='f'):
            b.append(f)
        elif (edsi[x]=='m'):
            b.append(m)
        elif (edsi[x]=='d'):
            b.append(d)
        elif (edsi[x]=='p'):
            b.append(p)
        elif (edsi[x]=='t'):
            b.append(t)
        elif (edsi[x]=='g'):
            b.append(g)
        else:
            return
    
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = Border(left=b[0],right=b[1],top=b[2],bottom=b[3])

def adjust_column(ws, min_row, min_col, max_col):
    
    column_widths = []

    for i, col in \
        enumerate(
            ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
        ):

        for cell in col:
            value = cell.value
            if value is not None:

                if isinstance(value, str) is False:
                    value = str(value)

                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))

    for i, width in enumerate(column_widths):

        col_name = get_column_letter(min_col + i)
        value = column_widths[i] + 2
        value = int(round(value + value * 0.2,0))
        ws.column_dimensions[col_name].width = value

if __name__ == "__main__":
    global arquivo_destino
    arquivo_destino= ""
    log("-"*100)
    log(" - INICIO DO RELATÓRIO INSUMO CONSOLIDADO SPED FISCAL 21210708" ,sys.argv[0])
#    variaveis = carregaConfiguracoes()
    ret = processar()
    if (ret > 0) :
        if(arquivo_destino):
            if os.path.isfile(arquivo_destino):
                os.remove(arquivo_destino)
    log("-" * 100)
    log(" - Código de execução = ", ret)
    log("-" * 100)
    log(" - FIM DO RELATÓRIO INSUMO CONSOLIDADO SPED FISCAL",sys.argv[0])
    sys.exit(ret)




















