#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-
"""
----------------------------------------------------------------------------------------------
MODULO ...: TESHUVA
SCRIPT ...: relatorio_unificado.py
CRIACAO ..: 05/01/2022
AUTOR ....: Victor Santos / Welber Pena - KYROS TECNOLOGIA
DESCRICAO.: 
----------------------------------------------------------------------------------------------
  HISTORICO : 
    * 01/02/2021 - ALT001 - Welber Pena 
        - Alterado diretorio de destino do relatorio conforme reuniao com Flavio Teixeira.

----------------------------------------------------------------------------------------------
"""
import os
import sys

global SD, dir_base
SD = ('/' if os.name == 'posix' else '\\')
dir_base = os.path.join( os.path.realpath('.').split(SD+'PROD'+SD)[0], 'PROD') if os.path.realpath('.').__contains__(SD+'PROD'+SD) else os.path.join( os.path.realpath('.').split(SD+'DEV'+SD)[0], 'DEV')
sys.path.append(dir_base)

import configuracoes
import comum
import sql
import util
from openpyxl import Workbook, load_workbook
import openpyxl
import datetime
import shutil
from pathlib import Path
from openpyxl.utils import get_column_letter
import converte_query_to_json
import converte_gia_to_json
import converte_sped_to_json
from copy import copy

NoneType = type(None)
StrType = type("string")
IntType = type(88)
FloatType = type(88.00)
DateType = type(datetime.datetime.now())
msize = 1.1
SD = ('/' if os.name == 'posix' else '\\')

from openpyxl import load_workbook 

global g_relatorio_saida
g_ix_linha_aba_1  = 1
g_temp_aba_2      = []
g_temp_aba_3      = []
lista_tupla_aba_2 = []
lista_tupla_aba_3 = []

def preencher_aba_1(p_registros, p_obrigacao, p_versao):

    obrigacao = p_obrigacao
    versao    = p_versao

    global g_relatorio_saida
    global g_ix_linha_aba_1

    v_aba = g_relatorio_saida["1 – Detalhamento Fiscal"]

    g_layout_aba_1 = {
          'Empresa'                 : 1
        , 'UF Filial'               : 2
        , 'Filial'                  : 3
        , 'Mês / Ano'               : 4
        , 'Série'                   : 5
        , 'Volume'                  : 6
        , 'Alíquota'                : 7
        , 'CST'                     : 8
        , 'CFOP'                    : 9
        , 'Chave CFOP'              : 10
        , 'Chave CFOP 0000'         : 11
        , 'Chave CFOP Telecom'      : 12
        , 'Chave Obrigação'         : 13
        , 'Chave por Série'         : 14
        , 'Chave Aba 2'             : 15
        , 'Obrigação'               : 16
        , 'Versão'                  : 17
        , 'Valor Líquido'           : 18
        , 'Valor Base'              : 19
        , 'Valor de ICMS'           : 20
        , 'Valor de Isentas'        : 21
        , 'Valor de Outras'         : 22
        , 'Valor Total'             : 23
        , 'Desconto / Redutores'    : 24
        , 'Substituo'               : 25
        , 'Substituído'             : 26
        , 'Outros Impostos'         : 27
        , 'E110'                    : 28
        , 'E116'                    : 29
        , 'Valor GIA'               : 30
        ,'Nome do Arquivo Mestre'   : 31
        ,'Código de Autenticação'   : 32
        ,'Indicador Retificação'    : 33
    }

    for registro in p_registros:

        v_indicador_cfop_0000     = 'Não'
        v_indicador_telecom_saida = 'N'
        
        if registro['CFOP'] == '0000':
            v_indicador_cfop_0000 = 'Sim'
        
        if registro['CFOP'] in ['0000','5301','5302','5303','5304','5305','5306','5307','6301','6302','6303','6304','6305','6306','6307','7301']:
            v_indicador_telecom_saida = 'S' 
        else:
            v_indicador_telecom_saida = 'N'
        
        registro['Chave CFOP']         = obrigacao + versao + registro['CFOP']
        registro['Chave CFOP 0000']    = obrigacao + versao + v_indicador_telecom_saida + v_indicador_cfop_0000
        registro['Chave CFOP Telecom'] = obrigacao + versao + v_indicador_telecom_saida
        registro['Chave Obrigação']    = obrigacao + ' - ' + versao
        registro['Chave por Série']    = registro['Série'] + obrigacao + versao
        registro['Obrigação']          = obrigacao 
        registro['Versão']             = versao 
        registro ['Chave Aba 2']       = registro['Empresa']+registro['UF Filial']+str(registro['Mês / Ano'])+str(registro.get('Série','N/a'))+str(registro.get('Volume','N/a'))+str(registro.get('Alíquota','N/a')).replace('.0','')+str(registro.get('CST','N/a'))+str(registro['CFOP'])+p_obrigacao+p_versao
        g_ix_linha_aba_1 += 1
        
        for coluna in g_layout_aba_1.keys():

            col = g_layout_aba_1[coluna]

            if registro.get(coluna, None) != None:
                celula = v_aba.cell(g_ix_linha_aba_1,col,registro[coluna])                    
            else:   
                celula = v_aba.cell(g_ix_linha_aba_1,col,'N/a')

            if g_ix_linha_aba_1 > 2:
                column_letter = get_column_letter(col)
                cell = v_aba[column_letter + str(2)]

                celula._style = copy(cell._style)               

        if g_ix_linha_aba_1 > 2: # 9 ate 44 

            for i in range(34, 36):
                
                letra = get_column_letter(i)

                celula = v_aba[letra + str(g_ix_linha_aba_1)]
                
                cell   = v_aba[letra + str(2)]                
                celula._style = copy(cell._style)
                celula.value  = copy(cell.value).replace('2', str(g_ix_linha_aba_1))               
    return True

def carrega_temporaria_aba_2(p_registros):

    global g_temp_aba_2
    global lista_tupla_aba_2

    ix_temp = 0
    v_chave_temp = ''
    p_registros  = sorted(p_registros, key=lambda row:(row['Empresa'],row['UF Filial'],row['Mês / Ano'],row['Série'],row['Volume'],row.get('Alíquota','N/a'),row.get('CST','N/a'),row['CFOP']),reverse=False)
    aliquota = 0.0
    for v_linha in p_registros:

        if v_linha.get('Alíquota', 'N/a') != 'N/a':
            aliquota = float(v_linha['Alíquota'])

        v_chave_linha =  str(v_linha['Empresa']).strip()                 +'|'+\
                         str(v_linha['UF Filial']).strip()               +'|'+\
                         str(v_linha['Mês / Ano']).strip()               +'|'+\
                         str(v_linha.get('Série','N/a'))                 +'|'+\
                         str(v_linha.get('Volume','N/a')).strip()        +'|'+\
                         str(aliquota).strip()                           +'|'+\
                         str(v_linha.get('CST', 'N/a')).strip()          +'|'+\
                         str(v_linha.get('CFOP','N/a')).strip()          +'|'

        if v_chave_linha not in lista_tupla_aba_2:

            lista_tupla_aba_2.append(v_chave_linha)

            while v_chave_temp < v_chave_linha:

                if ix_temp > (len(g_temp_aba_2) - 1):
                    v_chave_temp = 'ZZZZ'
                else:   
                    v_chave_temp =  g_temp_aba_2[ix_temp]['Empresa']                +'|'+\
                                    g_temp_aba_2[ix_temp]['UF Filial']              +'|'+\
                                    g_temp_aba_2[ix_temp]['Mês / Ano']              +'|'+\
                                    g_temp_aba_2[ix_temp].get('Série', 'N/a')       +'|'+\
                                    str(g_temp_aba_2[ix_temp].get('Volume', 'N/a')) +'|'+\
                                    str(g_temp_aba_2[ix_temp]['Alíquota'])          +'|'+\
                                    str(g_temp_aba_2[ix_temp]['CST'])               +'|'+\
                                    str(g_temp_aba_2[ix_temp]['CFOP'])              +'|'
                    ix_temp = ix_temp + 1

            if v_chave_temp > v_chave_linha:
                g_temp_aba_2.append(v_linha)
        
    g_temp_aba_2 = sorted(g_temp_aba_2, key=lambda row:(str(row['Empresa']),str(row['UF Filial']),str(row['Mês / Ano']),str(row['Série']),str(row['Volume']),str(row.get('Alíquota')),str(row.get('CST','N/a')),str(row['CFOP'])),reverse=False)

    return g_temp_aba_2

def preencher_aba_2():

    global g_relatorio_saida
    global g_temp_aba_2
    global NoneType

    g_layout_aba_2 = {'Empresa'   : 1
                     ,'UF Filial' : 2
                     ,'Mês / Ano' : 3
                     ,'Série'     : 4
                     ,'Volume'    : 5
                     ,'Alíquota'  : 6
                     ,'CST'       : 7
                     ,'CFOP'      : 8}

    v_aba = g_relatorio_saida["2 – Comparativo SPED"]   
    ix_linha = 2
    
    for v_linha in g_temp_aba_2:

        ix_linha = ix_linha + 1
        
        for coluna in g_layout_aba_2.keys():
            
            col = g_layout_aba_2[coluna]

            if v_linha.get(coluna, None) != None:
                celula = v_aba.cell(ix_linha,col,v_linha[coluna])
            else:   
                celula = v_aba.cell(ix_linha,col,'N/a')
            
            if ix_linha > 2:

                column_letter = get_column_letter(col)
                cell = v_aba[column_letter + str(3)]
                celula._style = copy(cell._style)

        if ix_linha > 3: # 9 ate 44 
            for i in range(9, 33):
                letra = get_column_letter(i)
                celula = v_aba[letra + str(ix_linha)]
                cell   = v_aba[letra + str(3)]                
                celula._style = copy(cell._style)

                if not cell.value == None:
                    celula.value = copy(cell.value).replace('3', str(ix_linha))
    
    log('Quantidade de linhas: ', len(g_temp_aba_2))
    return True

def carrega_temporaria_aba_3(p_registros):
    
    global g_temp_aba_3
    global lista_tupla_aba_3

    ix_temp_3 = 0
    v_chave_temp_3 = ''
    p_registros  = sorted(p_registros, key=lambda row:(row['Empresa'],row['UF Filial'],row['Mês / Ano'],row['Série'],row['Volume'],row['CFOP']),reverse=False)

    for v_linha_3 in p_registros:

        lista_cfop_telecom = ['5301','5302','5303','5304','5305','5306','5307','6301','6302','6303','6304','6305','6306','6307','7301','0000']
        
        if v_linha_3['CFOP'] in lista_cfop_telecom:
            v_chave_linha_3 =  str(v_linha_3['Empresa']).strip()                 +'|'+\
                               str(v_linha_3['UF Filial']).strip()               +'|'+\
                               str(v_linha_3['Mês / Ano']).strip()               +'|'+\
                               str(v_linha_3.get('Série','N/a'))                 +'|'+\
                               str(v_linha_3.get('Volume','N/a')).strip()        +'|'+\
                               str(v_linha_3.get('CFOP','N/a')).strip()          +'|'

            if v_chave_linha_3 not in lista_tupla_aba_3:

                lista_tupla_aba_3.append(v_chave_linha_3)

                while v_chave_temp_3 < v_chave_linha_3:

                    if ix_temp_3 > (len(g_temp_aba_3) - 1):
                        v_chave_temp_3 = 'ZZZZ'
                    else:   
                        v_chave_temp_3 =  g_temp_aba_3[ix_temp_3]['Empresa']                +'|'+\
                                          g_temp_aba_3[ix_temp_3]['UF Filial']              +'|'+\
                                          g_temp_aba_3[ix_temp_3]['Mês / Ano']              +'|'+\
                                          g_temp_aba_3[ix_temp_3].get('Série', 'N/a')       +'|'+\
                                          str(g_temp_aba_3[ix_temp_3].get('Volume', 'N/a')) +'|'+\
                                          str(g_temp_aba_3[ix_temp_3]['CFOP'])              +'|'
                        ix_temp_3 += 1

                if v_chave_temp_3 > v_chave_linha_3:
                    g_temp_aba_3.append(v_linha_3)
            
    g_temp_aba_3 = sorted(g_temp_aba_3, key=lambda row:(str(row['Empresa']),str(row['UF Filial']),str(row['Mês / Ano']),str(row['Série']),str(row['Volume']),str(row['CFOP'])),reverse=False)

    return g_temp_aba_3

def preencher_aba_3():
    
    global g_relatorio_saida
    global g_temp_aba_3

    g_layout_aba_3 = {'Empresa'   : 1
                     ,'UF Filial' : 2
                     ,'Mês / Ano' : 3
                     ,'Série'     : 4
                     ,'Volume'    : 5
                     ,'CFOP'      : 6}

    v_aba = g_relatorio_saida["3 – Comparativo SPED x C115"]   
    ix_linha = 2
    
    for v_linha in g_temp_aba_3:

        ix_linha = ix_linha + 1
        
        for coluna in g_layout_aba_3.keys():
            
            col = g_layout_aba_3[coluna]

            if v_linha.get(coluna, None) != None:
                celula = v_aba.cell(ix_linha,col,v_linha[coluna])
            else:   
                celula = v_aba.cell(ix_linha,col,'N/a')
            
            if ix_linha > 2:

                column_letter = get_column_letter(col)
                cell = v_aba[column_letter + str(3)]
                celula._style = copy(cell._style)

        if ix_linha > 3: # 9 ate 44 
            for i in range(7, 50):
                letra = get_column_letter(i)
                celula = v_aba[letra + str(ix_linha)]
                cell   = v_aba[letra + str(3)]                
                celula._style = copy(cell._style)

                if not cell.value == None:
                    celula.value = copy(cell.value).replace('3', str(ix_linha))
    
    log('Quantidade de linhas: ', len(g_temp_aba_3))
    return True

def ultimodia(ano,mes):
       return(31 if mes == 12 else datetime.date.fromordinal((datetime.date(ano,mes+1,1)).toordinal()-1).day)

def processar() :

    global g_relatorio_saida

    v_obrigacao_cv115 = 'Conv115'

    v_obrigacao_gia = 'GIA'

    v_obrigacao_sped = 'SPED'

    v_versao_atual_ti = 'Atual Ti'

    v_versao_ultimo_protocolado = 'Ultimo Protocolado'

    c6 = ''
    if configuracoes.banco != 'GFPRODC6' :
        c6 = '@c6'

    c1 = ''
    if configuracoes.banco != 'GFCLONEDEV' :
        c1 = '@c1'

    uf      = comum.getParametro('UF')
    mes_ano = comum.getParametro('MES_ANO')
    mesi    = mes_ano[:2]
    anoi    = mes_ano[2:]
   
    if util.validauf(uf) and len(mes_ano) == 6 and int(mesi) > 0 and int(mesi) < 13 and int(anoi) <= datetime.datetime.now().year and int(anoi) > (datetime.datetime.now().year)-50:
        v_ano = mes_ano[4:]
        datai = "01/" + mesi + "/" + anoi
        dataf =  str(ultimodia(int(anoi),int(mesi)))+"/"+str(mesi)+"/"+str(anoi)
    else:
        comum.imprimeHelp()
        ret = 91
        return ret

    data_atua = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    dic_ret = converte_query_to_json.le_arquivo_sql('query_conv115_atual_ti.sql')
    
    if dic_ret['status'] == 'Ok' :
        query_conv115_atual_ti = dic_ret['sql']        

        query_conv115_atual_ti = query_conv115_atual_ti.replace( '[uf_filial]', uf )
        query_conv115_atual_ti = query_conv115_atual_ti.replace( '[mes_ano]'  , mes_ano )
        query_conv115_atual_ti = query_conv115_atual_ti.replace( '[datai]'    , datai )
    
        dic_ret = converte_query_to_json.converter(query_conv115_atual_ti)

    if dic_ret['status'] == 'Ok' :
        v_regs_conv115_atual_ti = dic_ret['dados']

        dic_ret = converte_query_to_json.le_arquivo_sql('query_conv115_ultimo_protocolado_ent.sql')    

    if dic_ret['status'] == 'Ok' :
        query_conv115_ultimo_protocolado_e = dic_ret['sql']
        
        query_conv115_ultimo_protocolado_e = query_conv115_ultimo_protocolado_e.replace( '[uf_filial]', uf )
        query_conv115_ultimo_protocolado_e = query_conv115_ultimo_protocolado_e.replace( '[mes_ano]'  , mes_ano )
        query_conv115_ultimo_protocolado_e = query_conv115_ultimo_protocolado_e.replace( '[datai]'    , datai )
        query_conv115_ultimo_protocolado_e = query_conv115_ultimo_protocolado_e.replace( '[C6]'       , c6 )

        dic_ret = converte_query_to_json.converter(query_conv115_ultimo_protocolado_e)

    if dic_ret['status'] == 'Ok' : 
        v_regs_conv115_ultimo_protocolado_ent = dic_ret['dados']

        dic_ret = converte_query_to_json.le_arquivo_sql('query_conv115_ultimo_protocolado_ori.sql')

    if dic_ret['status'] == 'Ok' :
        query_conv115_ultimo_protocolado_o = dic_ret['sql']
        
        query_conv115_ultimo_protocolado_o = query_conv115_ultimo_protocolado_o.replace( '[uf_filial]', uf )
        query_conv115_ultimo_protocolado_o = query_conv115_ultimo_protocolado_o.replace( '[mes_ano]'  , mes_ano )
        query_conv115_ultimo_protocolado_o = query_conv115_ultimo_protocolado_o.replace( '[datai]'    , datai )
        query_conv115_ultimo_protocolado_o = query_conv115_ultimo_protocolado_o.replace( '[C1]'       , c1 )

        dic_ret = converte_query_to_json.converter(query_conv115_ultimo_protocolado_o)
    
    if dic_ret['status'] == 'Ok' : 
        v_regs_conv115_ultimo_protocolado_ori = dic_ret['dados']        

        dic_ret = converte_gia_to_json.converter(uf, mes_ano, v_versao_atual_ti )
        
    if dic_ret['status'] == 'Ok' :        
        v_regs_gia_atual_ti = dic_ret['dados'] 

    dic_ret = converte_gia_to_json.converter(uf, mes_ano, v_versao_ultimo_protocolado )
        
    if dic_ret['status'] == 'Ok' :        
        v_regs_gia_ultimo_protocolado = dic_ret['dados']     

        dic_ret = converte_sped_to_json.converter_valores(uf, mes_ano, v_versao_atual_ti )
        
    if dic_ret['status'] == 'Ok' :        
        v_regs_sped_atual_ti = dic_ret['dados']     

        dic_ret = converte_sped_to_json.converter_valores(uf, mes_ano, v_versao_ultimo_protocolado )
        
    if dic_ret['status'] == 'Ok' :        
        v_regs_sped_ultimo_protocolado = dic_ret['dados']     

    if dic_ret['status'] == 'Ok' :

        g_relatorio_saida = load_workbook(filename='Template_detalhamento_obrigacoes_fiscais.xlsm',keep_vba=True) 
        
        log(' Escrevendo dados na planilha - ABA 1 - CONV115 ATUAL TI '.center(100,'='))
        log('Quantidade de linhas:', len(v_regs_conv115_atual_ti))
        preencher_aba_1(v_regs_conv115_atual_ti,v_obrigacao_cv115,v_versao_atual_ti)

        log(' Escrevendo dados na planilha - ABA 1 - CONV115 ULTIMO PROTOCOLADO ( ULTIMO ENTREGUE ) '.center(100,'='))
        log('Quantidade de linhas:', len(v_regs_conv115_ultimo_protocolado_ent))
        preencher_aba_1(v_regs_conv115_ultimo_protocolado_ent,v_obrigacao_cv115,v_versao_ultimo_protocolado)

        log(' Escrevendo dados na planilha - ABA 1 - CONV115 ULTIMO PROTOCOLADO ( ORIGINAL ) '.center(100,'='))
        log('Quantidade de linhas:', len(v_regs_conv115_ultimo_protocolado_ori))
        preencher_aba_1(v_regs_conv115_ultimo_protocolado_ori,v_obrigacao_cv115,v_versao_ultimo_protocolado)
        
        log(' Escrevendo dados na planilha - ABA 1 - GIA ATUAL TI '.center(100,'='))
        log('Quantidade de linhas:', len(v_regs_gia_atual_ti))
        preencher_aba_1(v_regs_gia_atual_ti,v_obrigacao_gia,v_versao_atual_ti)

        log(' Escrevendo dados na planilha - ABA 1 - GIA ULTIMO PROTOCOLADO '.center(100,'='))
        log('Quantidade de linhas:', len(v_regs_gia_ultimo_protocolado))
        preencher_aba_1(v_regs_gia_ultimo_protocolado,v_obrigacao_gia,v_versao_ultimo_protocolado)

        log(' Escrevendo dados na planilha - ABA 1 - SPED ATUAL TI '.center(100,'='))
        log('Quantidade de linhas:', len(v_regs_sped_atual_ti))
        preencher_aba_1(v_regs_sped_atual_ti,v_obrigacao_sped,v_versao_atual_ti)

        log(' Escrevendo dados na planilha - ABA 1 - SPED ULTIMO PROTOCOLADO '.center(100,'='))
        log('Quantidade de linhas:', len(v_regs_sped_ultimo_protocolado))
        preencher_aba_1(v_regs_sped_ultimo_protocolado,v_obrigacao_sped,v_versao_ultimo_protocolado)

        log(' Carregando temporária ABA 2 - CONV 115 Atual Ti '.center(100,'='))
        log('Quantidade de linhas:', len(v_regs_conv115_atual_ti))
        carrega_temporaria_aba_2(v_regs_conv115_atual_ti)

        log(' Carregando temporária ABA 2 - SPED Atual Ti '.center(100,'='))
        log('Quantidade de linhas:', len(v_regs_sped_atual_ti))
        carrega_temporaria_aba_2(v_regs_sped_atual_ti)

        log(' Carregando temporária ABA 2 - SPED ULTIMO PROTOCOLADO '.center(100,'='))
        log('Quantidade de linhas:', len(v_regs_sped_ultimo_protocolado))
        carrega_temporaria_aba_2(v_regs_sped_ultimo_protocolado)

        log(' Escrevendo dados na planilha - ABA 2 '.center(100,'='))
        preencher_aba_2()

        log(' Carregando temporária ABA 3 - CONV 115 Atual Ti '.center(100,'='))
        log('Quantidade de linhas:', len(v_regs_conv115_atual_ti))
        carrega_temporaria_aba_3(v_regs_conv115_atual_ti)  

        log(' Carregando temporária ABA 3 -  SPED Atual Ti  '.center(100,'='))
        log('Quantidade de linhas:', len(v_regs_sped_atual_ti))
        carrega_temporaria_aba_3(v_regs_sped_atual_ti)       
        
        log(' Carregando temporária ABA 3 -  SPED ULTIMO PROTOCOLADO '.center(100,'='))
        log('Quantidade de linhas:', len(v_regs_sped_ultimo_protocolado))
        carrega_temporaria_aba_3(v_regs_sped_ultimo_protocolado)   

        log(' Escrevendo dados na planilha - ABA 3 '.center(100,'='))
        preencher_aba_3() 

        ### ALT001 - Incio
        v_diretorio = os.path.join( configuracoes.dir_geracao_arquivos.split('detalhamento_obrigacoes')[0] , 'Insumos', 'SPED_FISCAL', uf, anoi, mesi )
        ### Diretorio alterado conforme solicitação Flavio Teixeira
        # v_diretorio = configuracoes.v_caminho_relatorio
        #v_diretorio = v_diretorio.replace('<<UF>>', uf)
        #v_diretorio = v_diretorio.replace('<<ANO>>', anoi)
        #v_diretorio = v_diretorio.replace('<<MES>>', mesi)
        ### ALT001 - Fim
        
        v_nome_rel = 'Detalhamento_Obrigações_Fiscais_'+ uf + '_' + mes_ano + '_' + data_atua + '.xlsm'
        
        if not os.path.isdir(v_diretorio) :
            log(' Criando diretótrio'.center(100,'='))
            log(v_diretorio)
            os.makedirs(v_diretorio)

        v_nome_relatorio = os.path.join(v_diretorio, v_nome_rel)

        log('<  Salvando PLANILHA  >'.center(100,'='))
        g_relatorio_saida.save(v_nome_relatorio)
        log('- Caminho :', v_nome_relatorio)
        log('='*100)

    else:
        log('ERRO')
        return False

    return True


if __name__ == "__main__" :
    
    ret = 0
    print( )
    
    comum.addParametro( 'UF'      , None, "UNIDADE FEDERATIVA" , True , 'SP' )
    comum.addParametro( 'MES_ANO' , None, "MES E ANO"          , True , '012015')

    if not comum.validarParametros() :
        log('### ERRO AO VALIDAR OS PARÂMETROS')
        ret = 91
    else:
        configuracoes.uf       = comum.getParametro('UF').upper()
        configuracoes.mes_ano  = comum.getParametro('MES_ANO').upper()

        if not processar() :
            log('ERRO no processamento do relatorio !')
            ret = 92

    sys.exit(ret)
